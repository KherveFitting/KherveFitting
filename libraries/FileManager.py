import wx
import wx.grid
import os
import re
import numpy as np
import json
from matplotlib.ticker import ScalarFormatter
from copy import deepcopy
import shutil
import sys
from libraries.Save import save_state


class FileManagerWindow(wx.Frame):
    def __init__(self, parent, *args, **kwargs):
        # Check for maximum row index before initializing the window
        if hasattr(parent, 'Data') and 'Core levels' in parent.Data:
            max_index = 0
            for sheet_name in parent.Data['Core levels'].keys():
                match = re.match(r'[A-Za-z0-9]+?(\d*)$', sheet_name)
                if match:
                    index_str = match.group(1)
                    index = int(index_str) if index_str else 0
                    max_index = max(max_index, index)

            if max_index > 500:
                wx.MessageBox(
                    f"Cannot open Sample Manager: Sheet number {max_index} exceeds the maximum limit of 500.\n"
                    f"Please rename your sheets to use lower numbers.",
                    "Row Limit Exceeded", wx.OK | wx.ICON_ERROR)
                # Tell the parent that file_manager is None to enable reopening later
                parent.file_manager = None
                # Prevent the window from being created at all
                return

        # Only initialize the window if we didn't exceed the limit
        super().__init__(parent, title="Sample/Experiment Manager", size=(580, 350),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP, *args, **kwargs)

        # Add this line to set a minimum window size
        self.SetMinSize((630, 50))  # Ensure toolbar icons remain visible

        self.offset_multiplier = 1
        self.last_offset_sheets = []
        self.last_keypress_time = 0
        self.rapid_press_threshold = 1.0  # seconds

        self.parent = parent
        self.sample_names = {}  # Dictionary to store sample names by row index

        # Load sample names first
        self.load_sample_names()

        # Set up UI elements including the grid
        self.panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Create right panel with content
        self.right_panel = wx.Panel(self.panel)
        right_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create toolbar
        self.toolbar = wx.ToolBar(self.right_panel, style=wx.TB_HORIZONTAL | wx.TB_FLAT | wx.TB_NODIVIDER)
        self.toolbar.SetToolBitmapSize(wx.Size(25, 25))
        self.create_toolbar()
        self.toolbar.Realize()
        right_sizer.Add(self.toolbar, 0, wx.EXPAND)

        # Create grid
        self.grid = wx.grid.Grid(self.right_panel)
        self.core_levels = self.get_unique_core_levels()
        self.init_grid()
        right_sizer.Add(self.grid, 1, wx.EXPAND | wx.ALL, 5)

        self.right_panel.SetSizer(right_sizer)

        # Add panel to main sizer
        main_sizer.Add(self.right_panel, 1, wx.EXPAND)
        self.panel.SetSizer(main_sizer)

        # NOW load BE corrections after the grid is created
        self.load_be_corrections()

        # # Position window relative to main window
        # main_pos = parent.GetPosition()
        # main_size = parent.GetSize()
        # file_manager_size = self.GetSize()
        # pos_x = main_pos.x + (main_size.width - file_manager_size.width) // 2
        # pos_y = main_pos.y + (main_size.height - file_manager_size.height) // 2
        # self.SetPosition((pos_x, pos_y))

        # Position window using saved position or relative to main window
        if hasattr(parent, 'file_manager_position') and parent.file_manager_position:
            self.SetPosition(parent.file_manager_position)
        else:
            # Default positioning relative to main window
            main_pos = parent.GetPosition()
            main_size = parent.GetSize()
            file_manager_size = self.GetSize()
            pos_x = main_pos.x + (main_size.width - file_manager_size.width) // 2
            pos_y = main_pos.y + (main_size.height - file_manager_size.height) // 2
            self.SetPosition((pos_x, pos_y))

        # Initialize normalization values
        self.norm_min = 0
        self.norm_max = 1
        self.norm_vlines = [None, None]  # For the normalization cursors
        self.is_dragging_cursor = False

        # Bind events
        # self.norm_check.Bind(wx.EVT_CHECKBOX, self.on_norm_changed)
        # self.auto_check.Bind(wx.EVT_CHECKBOX, self.on_auto_changed)
        # self.parent.canvas.mpl_connect('key_press_event', self.on_key_press)
        # self.parent.canvas.mpl_connect('key_release_event', self.on_key_release)
        self.norm_check.Bind(wx.EVT_CHECKBOX, self.on_norm_changed)
        self.norm_type.Bind(wx.EVT_COMBOBOX, self.on_norm_type_changed)

        self.parent.canvas.mpl_connect('key_press_event', self.on_key_press)
        self.parent.canvas.mpl_connect('key_release_event', self.on_key_release)

        # Populate the grid with core levels
        self.populate_grid()

        # Bind grid events
        self.grid.Bind(wx.grid.EVT_GRID_CELL_CHANGING, self.on_cell_changing)
        self.grid.Bind(wx.EVT_KEY_DOWN, self.on_key_down)
        self.Bind(wx.EVT_KEY_DOWN, self.on_key_down)
        self.grid.Bind(wx.grid.EVT_GRID_CELL_CHANGED, self.on_cell_changed) # now redundant I believe
        self.grid.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.on_cell_focus_changed)


        self.Bind(wx.EVT_CLOSE, self.on_close)

        # Apply consistent fonts from parent
        from libraries.ConfigFile import set_consistent_fonts
        set_consistent_fonts(self)

        # Highlight the current sheet if any
        current_sheet = self.parent.sheet_combobox.GetValue()
        if current_sheet:
            self.highlight_current_sheet(current_sheet)

    def on_cell_changed(self, event):
        """Handle completed cell edit event"""
        row = event.GetRow()
        col = event.GetCol()

        event.Skip()

    def create_vertical_toolbar(self):
        """Create vertical toolbar with buttons for core level management"""
        # Get icon path
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Icons")
        v_toolbar_sizer = self.v_toolbar_panel.GetSizer()

        # Copy button
        copy_icon = os.path.join(icon_path, "copy-3.png")
        copy_btn = wx.BitmapButton(self.v_toolbar_panel, wx.ID_ANY,
                                   wx.Bitmap(copy_icon) if os.path.exists(copy_icon) else
                                   wx.ArtProvider.GetBitmap(wx.ART_COPY, wx.ART_BUTTON))
        copy_btn.SetToolTip("Copy Core Level")
        v_toolbar_sizer.Add(copy_btn, 0, wx.ALL, 2)
        self.Bind(wx.EVT_BUTTON, self.on_copy, copy_btn)

        # Paste button
        paste_btn = wx.BitmapButton(self.v_toolbar_panel, wx.ID_ANY,
                                    wx.ArtProvider.GetBitmap(wx.ART_PASTE, wx.ART_BUTTON))
        paste_btn.SetToolTip("Paste Core Level")
        v_toolbar_sizer.Add(paste_btn, 0, wx.ALL, 2)
        self.Bind(wx.EVT_BUTTON, self.on_paste, paste_btn)

        # Rename button
        rename_icon = os.path.join(icon_path, "rename-3.png")
        rename_btn = wx.BitmapButton(self.v_toolbar_panel, wx.ID_ANY,
                                     wx.Bitmap(rename_icon) if os.path.exists(rename_icon) else
                                     wx.ArtProvider.GetBitmap(wx.ART_INFORMATION, wx.ART_BUTTON))
        rename_btn.SetToolTip("Rename Core Level")
        v_toolbar_sizer.Add(rename_btn, 0, wx.ALL, 2)
        self.Bind(wx.EVT_BUTTON, self.on_rename, rename_btn)

        # Delete button
        delete_icon = os.path.join(icon_path, "delete-3.png")
        delete_btn = wx.BitmapButton(self.v_toolbar_panel, wx.ID_ANY,
                                     wx.Bitmap(delete_icon) if os.path.exists(delete_icon) else
                                     wx.ArtProvider.GetBitmap(wx.ART_DELETE, wx.ART_BUTTON))
        delete_btn.SetToolTip("Delete Core Level")
        v_toolbar_sizer.Add(delete_btn, 0, wx.ALL, 2)
        self.Bind(wx.EVT_BUTTON, self.on_delete, delete_btn)

        # Sum button
        sum_icon = os.path.join(icon_path, "SUM-25.png")
        sum_btn = wx.BitmapButton(self.v_toolbar_panel, wx.ID_ANY,
                                  wx.Bitmap(sum_icon) if os.path.exists(sum_icon) else
                                  wx.ArtProvider.GetBitmap(wx.ART_NEW, wx.ART_BUTTON))
        sum_btn.SetToolTip("Sum Selected")
        v_toolbar_sizer.Add(sum_btn, 0, wx.ALL, 2)
        self.Bind(wx.EVT_BUTTON, self.on_sum_selected, sum_btn)


    def create_toolbar(self):
        """Create toolbar with buttons for core level management"""
        # Get icon path
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Icons")

        add_rows_icon = os.path.join(icon_path, "add-rows-25.png")
        if os.path.exists(add_rows_icon):
            add_rows_bmp = wx.Bitmap(add_rows_icon)
        else:
            add_rows_bmp = wx.ArtProvider.GetBitmap(wx.ART_PLUS, wx.ART_TOOLBAR)
        add_rows_tool = self.toolbar.AddTool(wx.ID_ANY, "Add 1 Row", add_rows_bmp, "Add 1 row to the grid")
        self.Bind(wx.EVT_TOOL, lambda evt: self.add_single_row(), add_rows_tool)

        del_rows_icon = os.path.join(icon_path, "delete-rows-25.png")
        if os.path.exists(del_rows_icon):
            del_rows_bmp = wx.Bitmap(del_rows_icon)
        else:
            del_rows_bmp = wx.ArtProvider.GetBitmap(wx.ART_MINUS, wx.ART_TOOLBAR)
        del_rows_tool = self.toolbar.AddTool(wx.ID_ANY, "Delete Last Row", del_rows_bmp,
                                             "Delete the last 2 rows from the grid")
        self.Bind(wx.EVT_TOOL, lambda evt: self.delete_single_row(), del_rows_tool)

        # Copy button
        copy_icon = os.path.join(icon_path, "copy-3.png")
        if os.path.exists(copy_icon):
            copy_bmp = wx.Bitmap(copy_icon)
        else:
            copy_bmp = wx.ArtProvider.GetBitmap(wx.ART_COPY, wx.ART_TOOLBAR)
        copy_tool = self.toolbar.AddTool(wx.ID_ANY, "Copy Core Level", copy_bmp, "Copy selected core level")
        self.Bind(wx.EVT_TOOL, self.on_copy, copy_tool)

        # Paste button
        paste_icon = os.path.join(icon_path, "paste-3.png")
        if os.path.exists(paste_icon):
            paste_bmp = wx.Bitmap(paste_icon)
        else:
            paste_bmp = wx.ArtProvider.GetBitmap(wx.ART_PASTE, wx.ART_TOOLBAR)
        paste_tool = self.toolbar.AddTool(wx.ID_ANY, "Paste Core Level", paste_bmp, "Paste core level")
        self.Bind(wx.EVT_TOOL, self.on_paste, paste_tool)

        # Rename button
        rename_icon = os.path.join(icon_path, "rename-3.png")
        if os.path.exists(rename_icon):
            rename_bmp = wx.Bitmap(rename_icon)
        else:
            rename_bmp = wx.ArtProvider.GetBitmap(wx.ART_INFORMATION, wx.ART_TOOLBAR)
        rename_tool = self.toolbar.AddTool(wx.ID_ANY, "Rename Core Level", rename_bmp, "Rename selected core level")
        self.Bind(wx.EVT_TOOL, self.on_rename, rename_tool)

        # Delete button
        delete_icon = os.path.join(icon_path, "delete-3.png")
        if os.path.exists(delete_icon):
            delete_bmp = wx.Bitmap(delete_icon)
        else:
            delete_bmp = wx.ArtProvider.GetBitmap(wx.ART_DELETE, wx.ART_TOOLBAR)
        delete_tool = self.toolbar.AddTool(wx.ID_ANY, "Delete Core Level", delete_bmp, "Delete selected core level")
        self.Bind(wx.EVT_TOOL, self.on_delete, delete_tool)

        # Sum button
        sum_icon = os.path.join(icon_path, "SUM-25.png")
        sum_bmp = wx.Bitmap(sum_icon)
        sum_tool = self.toolbar.AddTool(wx.ID_ANY, "Sum Selected", sum_bmp, "Sum selected core levels")
        self.Bind(wx.EVT_TOOL, self.on_sum_selected, sum_tool)
        #
        # sort_icon = os.path.join(icon_path, "Sort-25.png")
        # if os.path.exists(sort_icon):
        #     sort_bmp = wx.Bitmap(sort_icon)
        # else:
        #     sort_bmp = wx.ArtProvider.GetBitmap(wx.ART_SORT_ASC, wx.ART_TOOLBAR)
        # sort_tool = self.toolbar.AddTool(wx.ID_ANY, "Sort Sheets", sort_bmp, "Sort sheets by sample groups")
        # self.Bind(wx.EVT_TOOL, self.sort_excel_sheets, sort_tool)

        # Plot button
        plot_icon = os.path.join(icon_path, "Plot2-25.png")
        if os.path.exists(plot_icon):
            plot_bmp = wx.Bitmap(plot_icon)
        else:
            plot_bmp = wx.ArtProvider.GetBitmap(wx.ART_FIND, wx.ART_TOOLBAR)
        plot_tool = self.toolbar.AddTool(wx.ID_ANY, "Plot Selected", plot_bmp, "Plot selected core level(s)"
                                        "\n Press F2 or Ctrl+2 to plot selected core level")
        self.Bind(wx.EVT_TOOL, self.on_plot_selected, plot_tool)

        # Offset plot button
        offset_plot_icon = os.path.join(icon_path, "Plot3-25.png")  # Using the same icon for now
        if os.path.exists(offset_plot_icon):
            offset_plot_bmp = wx.Bitmap(offset_plot_icon)
        else:
            offset_plot_bmp = wx.ArtProvider.GetBitmap(wx.ART_FIND, wx.ART_TOOLBAR)
        offset_plot_tool = self.toolbar.AddTool(wx.ID_ANY, "Plot with Offset", offset_plot_bmp,
                                                "Plot selected core level(s) with offset\n"
                                                "Press F3 or Ctrl+3 to plot with offset")
        self.Bind(wx.EVT_TOOL, self.on_plot_selected_with_offset, offset_plot_tool)

        self.norm_check = wx.CheckBox(self.toolbar, label="Norm.")
        self.norm_check.SetToolTip("Normalise multiple plot data")
        self.norm_type = wx.ComboBox(self.toolbar, choices=["Auto", "Norm. @ BE", "Norm. to A"], style=wx.CB_READONLY)
        self.norm_type.SetSelection(0)  # Default to "Auto"
        self.norm_type.SetToolTip("Choose normalization method. \n Press the shift key to activate the vLine in Norm. @ BE mode.")
        self.norm_check.SetValue(True)  # Auto is checked by default

        self.toolbar.AddControl(self.norm_check)
        self.toolbar.AddControl(self.norm_type)

        # self.toolbar.AddSeparator()



        # Toggle size button - using a different art ID
        self.toolbar.AddStretchableSpace()

        # Add experimental description info button
        exp_info_icon = os.path.join(icon_path, "info-25.png")
        if os.path.exists(exp_info_icon):
            exp_info_bmp = wx.Bitmap(exp_info_icon)
        else:
            exp_info_bmp = wx.ArtProvider.GetBitmap(wx.ART_INFORMATION, wx.ART_TOOLBAR)
        exp_info_tool = self.toolbar.AddTool(wx.ID_ANY, "Experimental Info", exp_info_bmp,
                                             "View experimental description information."
                                             "\nCurrently only upport .vms and .kal")
        self.Bind(wx.EVT_TOOL, self.on_view_exp_info, exp_info_tool)

        # # Backup button
        # backup_icon = os.path.join(icon_path, "backup-25.png")
        # if os.path.exists(backup_icon):
        #     backup_bmp = wx.Bitmap(backup_icon)
        # else:
        #     backup_bmp = wx.ArtProvider.GetBitmap(wx.ART_FILE_SAVE_AS, wx.ART_TOOLBAR)
        # backup_tool = self.toolbar.AddTool(wx.ID_ANY, "Backup", backup_bmp, "Create a backup of current files")
        # self.Bind(wx.EVT_TOOL, self.on_backup, backup_tool)

        # Add F2/Ctrl+2 info button
        f2_icon = os.path.join(icon_path, "Find-25.png")  # Use existing plot icon or another appropriate one
        if os.path.exists(f2_icon):
            f2_bmp = wx.Bitmap(f2_icon)
        else:
            f2_bmp = wx.ArtProvider.GetBitmap(wx.ART_QUESTION, wx.ART_TOOLBAR)
        f2_info_tool = self.toolbar.AddTool(wx.ID_ANY, "Plot Shortcuts", f2_bmp, "-- Press F2 or Ctrl+2 to plot peak "
                    "models. \n-- Press F3 or Ctrl+3 to plot multiple plots with an offset\n"
                    "-- In Norm. @BE mode, activate the plot & press the shift key to activate the vLine."
                    "")

        # pref_icon = os.path.join(icon_path, "settings-25.png")
        # pref_bmp = wx.Bitmap(pref_icon)
        # pref_tool = self.toolbar.AddTool(wx.ID_ANY, "Preferences", pref_bmp, "Normalization Settings")
        # self.Bind(wx.EVT_TOOL, self.on_preferences, pref_tool)

        size_icon = os.path.join(icon_path, "Minimize-25.png")
        size_bmp = wx.Bitmap(size_icon)
        self.size_tool = self.toolbar.AddTool(wx.ID_ANY, "Toggle Size", size_bmp, "Toggle window height")
        self.Bind(wx.EVT_TOOL, self.on_toggle_size, self.size_tool)

        # Store the original/expanded size
        self.is_collapsed = False
        self.expanded_height = 300  # Default expanded height
        self.collapsed_height = 70  # Collapsed height

    def on_toggle_size(self, event):
        """Toggle between normal and small window size"""
        current_size = self.GetSize()

        if self.is_collapsed:
            # Expand
            self.SetSize((current_size.width, self.expanded_height))
            self.is_collapsed = False
        else:
            # Collapse
            self.expanded_height = current_size.height  # Save current height
            self.SetSize((current_size.width, self.collapsed_height))
            self.is_collapsed = True

    def init_grid(self):
        """Initialize the grid with rows and columns"""
        # Get unique core level names from parent
        num_levels = len(self.core_levels)

        # Determine number of rows based on existing data
        max_row_index = self.get_max_core_level_row_index()

        # Check if max_row_index exceeds the limit
        if max_row_index > 500:
            wx.MessageBox(
                f"Cannot open Sample Manager: Sheet number {max_row_index} exceeds the maximum limit of 500.\n"
                f"Please rename your sheets to use lower numbers.",
                "Row Limit Exceeded", wx.OK | wx.ICON_ERROR)

        num_rows = max(10, max_row_index + 1)  # Ensure at least 10 rows

        # Create grid
        self.grid.CreateGrid(num_rows, num_levels)

        # Add "Sample Name" column at index 0
        self.grid.InsertCols(0, 1)
        self.grid.SetColLabelValue(0, "Experiment")
        self.grid.SetColSize(0, 70)  # Wider column for sample names

        # Add BE Correction and Normalization columns at the end
        self.grid.AppendCols(3)  # Add 3 columns instead of 2
        self.grid.SetColLabelValue(num_levels + 1, "Xshift")
        self.grid.SetColLabelValue(num_levels + 2, "Norm. @ BE")
        self.grid.SetColLabelValue(num_levels + 3, "Norm. to A")


        # Set column width for new columns
        self.grid.SetColSize(num_levels + 1, 60)
        self.grid.SetColSize(num_levels + 2, 70)  # Wider for the new name
        self.grid.SetColSize(num_levels + 3, 70)  # New column

        # # Enable cell editing for renaming
        # self.grid.EnableEditing(False)

        # Keep editing enabled but control which cells are editable
        self.grid.EnableEditing(True)

        # Make all core level columns read-only (columns 1 to len(self.core_levels))
        for row in range(self.grid.GetNumberRows()):
            for col in range(1, len(self.core_levels) + 1):  # Core level columns only
                self.grid.SetReadOnly(row, col, True)

        # Set column labels (core level names)
        self.grid.SetRowLabelSize(30)
        for i, level in enumerate(self.core_levels):
            self.grid.SetColLabelValue(i + 1, level)  # Add +1 to skip Sample Name column

        # Set row labels (0, 1, 2, etc.)
        for i in range(num_rows):
            self.grid.SetRowLabelValue(i, str(i))

        # Set column width and row height
        default_col_width = 50
        default_row_height = 20

        # Set column sizes and row heights
        for i in range(num_levels):
            self.grid.SetColSize(i+1, default_col_width)
        for i in range(num_rows):
            self.grid.SetRowSize(i, default_row_height)

        # Set cell alignment
        self.grid.SetColSize(num_levels + 1, 40)
        self.grid.SetColSize(num_levels + 2, 40)

        # After setting up grid and columns, make BE Correction column read-only
        be_col_index = len(self.core_levels) + 1
        for row in range(self.grid.GetNumberRows()):
            self.grid.SetReadOnly(row, be_col_index, True)

        # Set cell alignment
        for row in range(num_rows):
            for col in range(num_levels):
                self.grid.SetCellAlignment(row, col, wx.ALIGN_CENTER, wx.ALIGN_CENTER)


            # Apply consistent fonts to grid
            if 'wxMac' in wx.PlatformInfo:
                default_font = 'Helvetica'
                font_size = 11
            elif 'wxGTK' in wx.PlatformInfo:
                default_font = 'DejaVu Sans'
                font_size = 9
            else:
                default_font = 'Calibri'
                font_size = 9

            self.grid.SetDefaultCellFont(
                wx.Font(font_size, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, faceName=default_font))
            self.grid.SetLabelFont(
                wx.Font(font_size, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, faceName=default_font))

            self.grid.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.on_cell_click)
            self.grid.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.on_cursor_changed)
            self.grid.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.on_grid_right_click)
            self.grid.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.on_cell_double_click)

    def on_cell_double_click(self, event):
        """Handle double-click on grid cells"""
        row = event.GetRow()
        col = event.GetCol()

        # Allow double-click editing for experiment column (0) and normalization columns
        editable_columns = [0,  # Experiment/Sample name column
                            len(self.core_levels) + 2,  # Norm. @ BE column
                            len(self.core_levels) + 3]  # Norm. to A column

        if col in editable_columns:
            # For editable columns, let the default double-click behavior work
            event.Skip()
        else:
            # Call the plot function directly
            self.on_plot_selected(None)
            return  # Don't skip the event

    def get_unique_core_levels(self):
        """Get list of unique core level names from parent data"""
        unique_levels = set()

        if hasattr(self.parent, 'Data') and 'Core levels' in self.parent.Data:
            for sheet_name in self.parent.Data['Core levels'].keys():
                base_name = self.extract_base_name(sheet_name)
                if base_name:
                    unique_levels.add(base_name)

        return sorted(list(unique_levels))

    def extract_base_name(self, sheet_name):
        """Extract base core level name without any trailing numbers"""
        # Add support for Raman files with underscores
        if "Raman_" in sheet_name or "Ra_" in sheet_name:
            # Special handling for Raman files with underscores
            base_parts = sheet_name.split('_')
            if len(base_parts) > 1:
                return '_'.join(base_parts[:-1]) if base_parts[-1].isdigit() else sheet_name

        # Original pattern for typical core levels
        match = re.match(r'([A-Za-z0-9]+?)(\d*)$', sheet_name)
        if match:
            return match.group(1)
        return sheet_name

    def get_max_core_level_row_index(self):
        """Get the maximum row index based on core level naming"""
        max_index = 0

        if hasattr(self.parent, 'Data') and 'Core levels' in self.parent.Data:
            for sheet_name in self.parent.Data['Core levels'].keys():
                match = re.match(r'[A-Za-z0-9]+?(\d*)$', sheet_name)
                if match:
                    index_str = match.group(1)
                    index = int(index_str) if index_str else 0
                    max_index = max(max_index, index)

        return max_index

    def populate_grid(self):
        """Populate the grid with core levels from the data"""
        if not hasattr(self.parent, 'Data') or 'Core levels' not in self.parent.Data:
            return

        # Find maximum index across all sheets
        max_index = 0
        for sheet_name in self.parent.Data['Core levels'].keys():
            match = re.match(r'[A-Za-z0-9]+?(\d*)$', sheet_name)
            if match:
                index_str = match.group(1)
                index = int(index_str) if index_str else 0
                max_index = max(max_index, index)

        # Check if max_index exceeds the limit
        if max_index > 500:
            wx.MessageBox(f"Cannot open Sample Manager: Sheet number {max_index} exceeds the maximum limit of 500.\n"
                          f"Please rename your sheets to use lower numbers.",
                          "Row Limit Exceeded", wx.OK | wx.ICON_ERROR)
            self.Close()
            return

        # Clear grid first
        for row in range(self.grid.GetNumberRows()):
            for col in range(self.grid.GetNumberCols()):
                self.grid.SetCellValue(row, col, "")
                self.grid.SetCellBackgroundColour(row, col, wx.WHITE)

        # Initialize sample names from parent data if available
        if 'SampleNames' in self.parent.Data:
            self.sample_names = self.parent.Data['SampleNames']
        else:
            self.sample_names = {}

        # Map of all core levels by type and index
        core_level_map = {}

        self.grid.SetColSize(0, 70)  # Reset wider width for sample name column

        # First, categorize all core levels
        # First, categorize all core levels
        for sheet_name in self.parent.Data['Core levels'].keys():
            if "Raman_" in sheet_name or "Ra_" in sheet_name:
                # Handle Raman files with underscore
                base_parts = sheet_name.split('_')
                base_name = base_parts[0] + "_" + base_parts[1]  # Keep format as "Raman_bSiO4"
                index_str = ""
                if len(base_parts) > 2 and base_parts[2].isdigit():
                    index_str = base_parts[2]
                index = int(index_str) if index_str else 0
            else:
                # Original pattern for typical core levels
                match = re.match(r'([A-Za-z0-9]+?)(\d*)$', sheet_name)
                if match:
                    base_name = match.group(1)
                    index_str = match.group(2)
                    index = int(index_str) if index_str else 0
                else:
                    continue  # Skip if no match found

            if base_name not in core_level_map:
                core_level_map[base_name] = {}
            core_level_map[base_name][index] = sheet_name

        # Make sure grid has enough rows
        max_index = 0
        for base_name in core_level_map:
            if core_level_map[base_name]:
                max_index = max(max_index, max(core_level_map[base_name].keys()))

        if max_index >= self.grid.GetNumberRows():
            self.grid.AppendRows(max_index - self.grid.GetNumberRows() + 1)
            # Set row label for new rows
            for row in range(self.grid.GetNumberRows()):
                if not self.grid.GetRowLabelValue(row):
                    self.grid.SetRowLabelValue(row, str(row))


        # Set column width and row height
        default_col_width = 50
        # Set column sizes and row heights
        for i in range(len(self.core_levels)):
            if i + 1 < self.grid.GetNumberCols():
                self.grid.SetColSize(i + 1, default_col_width)

        # Make sure grid has enough columns (core levels + sample name column)
        if len(self.core_levels) + 1 > self.grid.GetNumberCols():
            self.grid.AppendCols(len(self.core_levels) + 1 - self.grid.GetNumberCols())
            # Update column labels
            self.grid.SetColLabelValue(0, "Experiment")
            for i, level in enumerate(self.core_levels):
                self.grid.SetColLabelValue(i + 1, level)

        # Make sure we have the BE correction and Normalization columns
        be_col_index = len(self.core_levels) + 1
        norm_col_index = len(self.core_levels) + 2
        norm_area_col_index = len(self.core_levels) + 3

        # Add these columns if they don't exist
        if self.grid.GetNumberCols() <= be_col_index:
            self.grid.AppendCols(norm_col_index + 1 - self.grid.GetNumberCols())

        # Set column labels
        self.grid.SetColLabelValue(be_col_index, "Xshift")
        self.grid.SetColLabelValue(norm_col_index, "Norm. @ BE")
        self.grid.SetColLabelValue(norm_area_col_index, "Norm. to A")

        if norm_col_index < self.grid.GetNumberCols():
            # Leave normalization column empty for now
            self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(180, 235, 208))

        if norm_area_col_index < self.grid.GetNumberCols():
            # Leave area normalization column empty for now
            self.grid.SetCellBackgroundColour(row, norm_area_col_index, wx.Colour(180, 235, 208))

        # Set column sizes
        self.grid.SetColSize(be_col_index, 40)
        self.grid.SetColSize(norm_col_index, 40)

        # Now populate the grid
        for col, base_name in enumerate(self.core_levels):
            if base_name in core_level_map:
                for index, sheet_name in core_level_map[base_name].items():
                    # Ensure we have enough rows
                    if index >= self.grid.GetNumberRows():
                        self.grid.AppendRows(index - self.grid.GetNumberRows() + 1)
                        # Set row label for new rows
                        for row in range(self.grid.GetNumberRows()):
                            if not self.grid.GetRowLabelValue(row):
                                self.grid.SetRowLabelValue(row, str(row))

                    # Set the cell value and color
                    self.grid.SetCellValue(index, col+1, sheet_name)
                    self.grid.SetCellBackgroundColour(index, col+1, wx.Colour(200, 245, 228))

        # Add BE correction values for each row
        for row in range(self.grid.GetNumberRows()):
            # Verify column index is valid before setting value
            be_col_index = len(self.core_levels) + 1
            if be_col_index < self.grid.GetNumberCols():
                # Set BE correction value if available
                be_correction = self.parent.Data.get('BEcorrections', {}).get(str(row), "0.0")
                self.grid.SetCellValue(row, be_col_index, str(be_correction))
                self.grid.SetCellBackgroundColour(row, be_col_index, wx.Colour(230, 230, 230))

            # Check if normalization column exists
            norm_col_index = len(self.core_levels) + 2
            if norm_col_index < self.grid.GetNumberCols():
                # Leave normalization column empty for now
                self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(230, 230, 230))

        be_col_index = len(self.core_levels) + 1
        if be_col_index < self.grid.GetNumberCols():
            for row in range(self.grid.GetNumberRows()):
                # Set BE correction value if available
                be_correction = self.parent.Data.get('BEcorrections', {}).get(str(row), "0.0")
                self.grid.SetCellValue(row, be_col_index, str(be_correction))
                self.grid.SetCellBackgroundColour(row, be_col_index, wx.Colour(230, 230, 230))
                self.grid.SetCellTextColour(row, be_col_index, wx.Colour(128, 128, 128))  # Set text color to gray

        # Add sample names to first column
        for row in range(self.grid.GetNumberRows()):
            sample_name = self.sample_names.get(str(row), "")
            self.grid.SetCellValue(row, 0, sample_name)
            self.grid.SetCellBackgroundColour(row, 0, wx.Colour(230, 230, 230))

        # Set background color for normalization columns for all rows
        norm_col_index = len(self.core_levels) + 2
        norm_area_col_index = len(self.core_levels) + 3
        for row in range(self.grid.GetNumberRows()):
            if norm_col_index < self.grid.GetNumberCols():
                # self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(180, 235, 208))
                self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(230, 230, 230))
            if norm_area_col_index < self.grid.GetNumberCols():
                self.grid.SetCellBackgroundColour(row, norm_area_col_index, wx.Colour(230, 230, 230))

        # Make sure grid has enough columns (core levels + sample name column + BE + Norm columns)
        required_cols = len(self.core_levels) + 4  # +1 for sample name, +3 for BE and two norm columns
        current_cols = self.grid.GetNumberCols()

        if current_cols < required_cols:
            self.grid.AppendCols(required_cols - current_cols)

        num_levels = len(self.core_levels)
        self.grid.SetColSize(num_levels + 2, 70)  # Wider for the new name
        self.grid.SetColSize(num_levels + 3, 70)  # New column

        # Calculate total width needed based on column sizes
        total_width = 0
        for col in range(self.grid.GetNumberCols()):
            total_width += self.grid.GetColSize(col)

        # Add some padding for grid borders, scrollbars, etc.
        total_width += 40

        # Add width for row labels
        total_width += self.grid.GetRowLabelSize()

        # Ensure minimum width for toolbar visibility
        minimum_width = 630
        total_width = max(total_width, minimum_width)

        # Calculate height based on current window size
        current_size = self.GetSize()

        # Set the new window size with calculated width and current height
        self.SetSize(total_width, current_size.GetHeight())

        # Force refresh
        self.grid.ForceRefresh()

    def load_be_corrections(self):
        """Load BE correction values from parent Data or JSON file"""

        # First, check if we have BE corrections in parent.Data
        if 'BEcorrections' in self.parent.Data:
            be_corrections = self.parent.Data['BEcorrections']
            # Apply these corrections to the grid
            be_col = len(self.core_levels) + 1
            for row, correction in be_corrections.items():
                try:
                    row_idx = int(row)
                    if 0 <= row_idx < self.grid.GetNumberRows():
                        self.grid.SetCellValue(row_idx, be_col, str(correction))
                except (ValueError, IndexError):
                    continue

        # As fallback, load from JSON file
        else:
            import json
            file_path = self.parent.Data.get('FilePath', '')
            if file_path:
                json_path = os.path.splitext(file_path)[0] + '.json'
                try:
                    if os.path.exists(json_path):
                        with open(json_path, 'r') as f:
                            json_data = json.load(f)

                        if 'BEcorrections' in json_data:
                            be_corrections = json_data['BEcorrections']
                            self.parent.Data['BEcorrections'] = be_corrections

                            # Update grid with BE corrections
                            be_col = len(self.core_levels) + 1
                            for row, correction in be_corrections.items():
                                try:
                                    row_idx = int(row)
                                    if 0 <= row_idx < self.grid.GetNumberRows():
                                        self.grid.SetCellValue(row_idx, be_col, str(correction))
                                except (ValueError, IndexError):
                                    continue
                except Exception as e:
                    print(f"Error loading BE corrections: {e}")


    def save_be_corrections(self):
        """Save BE correction values from grid to parent data only"""
        be_corrections = {}

        # Get BE correction values from grid
        be_col_index = len(self.core_levels) + 1

        # Check if BE column exists
        if be_col_index < self.grid.GetNumberCols():
            for row in range(self.grid.GetNumberRows()):
                value = self.grid.GetCellValue(row, be_col_index)
                if value.strip():
                    try:
                        be_corrections[str(row)] = float(value)
                    except ValueError:
                        be_corrections[str(row)] = 0.0

        # Save to parent.Data
        self.parent.Data['BEcorrections'] = be_corrections

        # Update current BE correction based on selected sheet
        current_sheet = self.parent.sheet_combobox.GetValue()
        sheet_found = False
        for row in range(self.grid.GetNumberRows()):
            for col in range(1, len(self.core_levels) + 1):
                if self.grid.GetCellValue(row, col) == current_sheet:
                    sheet_found = True
                    correction = be_corrections.get(str(row), 0.0)
                    self.parent.be_correction = correction
                    self.parent.Data['BEcorrection'] = correction  # For backward compatibility
                    self.parent.be_correction_spinbox.SetValue(correction)
                    break
            if sheet_found:
                break

    def load_sample_names(self):
        import json
        file_path = self.parent.Data.get('FilePath', '')
        if file_path:
            json_path = os.path.splitext(file_path)[0] + '.json'
            try:
                if os.path.exists(json_path):
                    with open(json_path, 'r') as f:
                        json_data = json.load(f)

                    if 'SampleNames' in json_data:
                        self.sample_names = json_data['SampleNames']
                        self.parent.Data['SampleNames'] = self.sample_names
            except Exception as e:
                print(f"Error loading sample names: {e}")

    # Add a method to save sample names:

    def save_sample_names(self):
        # Update sample_names from grid
        for row in range(self.grid.GetNumberRows()):
            name = self.grid.GetCellValue(row, 0)
            if name:
                self.sample_names[str(row)] = name
            elif str(row) in self.sample_names:
                del self.sample_names[str(row)]

        # Save to parent.Data
        self.parent.Data['SampleNames'] = self.sample_names

        # Save to JSON file
        file_path = self.parent.Data.get('FilePath', '')
        if file_path:
            json_path = os.path.splitext(file_path)[0] + '.json'
            try:
                # Load existing JSON data
                json_data = {}
                if os.path.exists(json_path):
                    with open(json_path, 'r') as f:
                        json_data = json.load(f)

                # Update sample names in JSON
                json_data['SampleNames'] = self.sample_names

                # Save back to JSON file
                with open(json_path, 'w') as f:
                    json.dump(json_data, f, indent=2)

            except Exception as e:
                print(f"Error saving sample names to JSON: {e}")

    def on_key_down(self, event):
        """Handle key press events"""
        key_code = event.GetKeyCode()

        if key_code == wx.WXK_F2:
            # Call the plot function directly
            self.on_plot_selected(None)
            return  # Don't skip the event
        elif key_code == wx.WXK_F3:
            # Call the offset plot function
            self.on_plot_selected_with_offset(None)
            return  # Don't skip the event
        elif event.ControlDown() and key_code == wx.WXK_F2:
            # CTRL+F2: Standard multiple plot
            self.on_plot_selected(None)
            return
        elif event.ControlDown() and key_code == wx.WXK_F3:
            # CTRL+F3: Offset multiple plot
            self.on_plot_selected_with_offset(None)
            return
        elif event.ControlDown() and key_code == ord('2'):
            # CTRL+2: Standard multiple plot
            self.on_plot_selected(None)
            return
        elif event.ControlDown() and key_code == ord('3'):
            # CTRL+3: Offset multiple plot
            self.on_plot_selected_with_offset(None)
            return
        else:
            event.Skip()
    def on_plot_selected(self, event):
        """Plot the currently selected core level(s)"""
        sheet_names = self.get_selected_sheet_names()

        if sheet_names:
            if len(sheet_names) == 1:
                # Single sheet - update combobox and plot
                self.parent.sheet_combobox.SetValue(sheet_names[0])
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, sheet_names[0])
            else:
                # Multiple sheets - overlay plot
                self.plot_multiple_sheets(sheet_names)
                # Update combobox with first sheet name
                self.parent.sheet_combobox.SetValue(sheet_names[0])

            # Highlight the selected cell(s)
            self.highlight_current_sheet(sheet_names[0])
            self.Raise()  # Bring the file manager window to the front

    def on_sum_selected(self, event):
        """Sum/average the Y values of selected cells in the same column"""
        save_state(self.parent)
        # Get selected cells grouped by column
        selected_by_column = {}

        # Check for selected blocks first
        blocks = self.grid.GetSelectedBlocks()
        for block in blocks:  # Directly iterate over blocks instead of using GetCount()
            top = block.GetTopRow()
            bottom = block.GetBottomRow()
            left = block.GetLeftCol()
            right = block.GetRightCol()

            for col in range(left, right + 1):
                if col not in selected_by_column:
                    selected_by_column[col] = []

                for row in range(top, bottom + 1):
                    cell_value = self.grid.GetCellValue(row, col)
                    if cell_value and cell_value in self.parent.Data['Core levels']:
                        selected_by_column[col].append(cell_value)

        # Add individually selected cells
        selected_cells = self.grid.GetSelectedCells()
        for cell in selected_cells:
            row, col = cell.GetRow(), cell.GetCol()

            if col not in selected_by_column:
                selected_by_column[col] = []

            cell_value = self.grid.GetCellValue(row, col)
            if cell_value and cell_value in self.parent.Data['Core levels']:
                selected_by_column[col].append(cell_value)

        # Process each column separately
        for col, sheet_names in selected_by_column.items():
            if len(sheet_names) > 1:
                self.create_summed_spectrum(sheet_names, self.grid.GetColLabelValue(col))

    def create_summed_spectrum(self, sheet_names, base_name):
        """Create a new spectrum that is the average of the selected spectra"""
        if not sheet_names:
            return

        # Find the earliest available row
        used_rows = []
        for sheet in self.parent.Data['Core levels'].keys():
            match = re.match(r'([A-Za-z0-9]+?)(\d*)$', sheet)
            if match:
                sheet_base = match.group(1)
                row_str = match.group(2)
                if sheet_base == base_name:
                    row_num = int(row_str) if row_str else 0
                    used_rows.append(row_num)

        # Find the first unused row number
        row_num = 0
        while row_num in used_rows:
            row_num += 1

        # Create the new sheet name with the available row
        new_sheet_name = f"{base_name}{row_num}" if row_num > 0 else base_name

        # Get X values from the first sheet (assuming they're similar)
        first_sheet = sheet_names[0]
        x_values = self.parent.Data['Core levels'][first_sheet]['B.E.']

        # Sum Y values from all sheets
        summed_y = np.zeros_like(x_values, dtype=float)
        for sheet_name in sheet_names:
            if sheet_name in self.parent.Data['Core levels']:
                current_y = self.parent.Data['Core levels'][sheet_name]['Raw Data']
                summed_y += np.array(current_y)

        # Average the Y values
        avg_y = summed_y / len(sheet_names)

        # Create a new entry in the parent data
        if 'Core levels' not in self.parent.Data:
            self.parent.Data['Core levels'] = {}
            self.parent.Data['Number of Core levels'] = 0

        self.parent.Data['Core levels'][new_sheet_name] = {
            'Name': new_sheet_name,
            'B.E.': x_values,
            'Raw Data': avg_y.tolist(),
            'Background': {
                'Bkg Type': 'Linear',
                'Bkg Low': min(x_values),
                'Bkg High': max(x_values),
                'Bkg Offset Low': 0,
                'Bkg Offset High': 0,
                'Bkg Y': avg_y.tolist()  # Initially use raw data as background
            }
        }
        self.parent.Data['Number of Core levels'] += 1

        # Update the Excel file
        import pandas as pd
        df = pd.DataFrame({
            'BE': x_values,
            'Raw Data': avg_y.tolist(),
            'Background': avg_y.tolist(),
            'Transmission': [1.0] * len(x_values)
        })

        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        # Update combobox in parent
        self.parent.sheet_combobox.Append(new_sheet_name)

        # Plot the new sheet
        self.parent.sheet_combobox.SetValue(new_sheet_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, new_sheet_name)

        # Update the grid
        self.populate_grid()

        # Notify the user
        # wx.MessageBox(f"Created summed spectrum: {new_sheet_name}", "Sum Complete", wx.OK | wx.ICON_INFORMATION)
        self.parent.show_popup_message2("Sum Complete", f"Created summed spectrum: {new_sheet_name}")

    def get_selected_sheet_names(self):
        """Get names of all currently selected sheets in the grid"""
        sheet_names = []

        # Check for selected blocks first
        blocks = self.grid.GetSelectedBlocks()
        for block in blocks:
            top = block.GetTopRow()
            bottom = block.GetBottomRow()
            left = block.GetLeftCol()
            right = block.GetRightCol()

            for row in range(top, bottom + 1):
                for col in range(left, right + 1):
                    cell_value = self.grid.GetCellValue(row, col)
                    if cell_value and cell_value in self.parent.Data['Core levels']:
                        sheet_names.append(cell_value)

        # Add individually selected cells
        selected_cells = self.grid.GetSelectedCells()
        for cell in selected_cells:
            row, col = cell.GetRow(), cell.GetCol()
            cell_value = self.grid.GetCellValue(row, col)
            if cell_value and cell_value in self.parent.Data['Core levels']:
                sheet_names.append(cell_value)

        # If no selection, use the current cell
        if not sheet_names:
            row = self.grid.GetGridCursorRow()
            col = self.grid.GetGridCursorCol()
            if row >= 0 and col >= 0:
                cell_value = self.grid.GetCellValue(row, col)
                if cell_value and cell_value in self.parent.Data['Core levels']:
                    sheet_names.append(cell_value)

        # Deduplicate the list before returning
        return list(dict.fromkeys(sheet_names))  # Preserves order unlike set conversion

    def plot_selected_cell(self):
        # Get the current cell
        row = self.grid.GetGridCursorRow()
        col = self.grid.GetGridCursorCol()

        # Only plot if the cell contains a valid sheet name
        cell_value = self.grid.GetCellValue(row, col)
        if cell_value and cell_value in self.parent.Data['Core levels']:
            self.on_plot_selected(None)

    def quick_plot_sheet(self, sheet_name):
        """Plot just the raw data quickly without background or peaks"""
        if sheet_name not in self.parent.Data['Core levels']:
            return

        # Update parent's combobox
        self.parent.sheet_combobox.SetValue(sheet_name)

        # Highlight this cell
        self.highlight_current_sheet(sheet_name)

        # Store the original residuals state
        original_residuals_state = self.parent.plot_manager.residuals_state

        # Clear the plot
        self.parent.ax.clear()

        # Get data
        x_values = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y_values = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        # Update parent's data arrays
        self.parent.x_values = np.array(x_values)
        self.parent.y_values = np.array(y_values)
        self.parent.background = np.array(y_values)  # Initialize background with raw data

        # Apply scientific format to Y-axis
        self.parent.ax.yaxis.set_major_formatter(ScalarFormatter(useMathText=True))
        self.parent.ax.ticklabel_format(style='sci', axis='y', scilimits=(0, 0))

        # Check if it's a Raman file
        is_raman = "Raman_" in sheet_name or "Ra_" in sheet_name

        # Plot simple black line
        if self.parent.energy_scale == 'KE':
            self.parent.ax.plot(self.parent.photons - x_values, y_values, 'k-', linewidth=1)
        else:
            self.parent.ax.plot(x_values, y_values, 'k-', linewidth=1)

        # Add core level text - skip for Raman files
        if not is_raman:
            base_name = self.extract_base_name(sheet_name)
            formatted_name = self.parent.plot_manager.format_sheet_name(base_name)
            self.parent.ax.text(0.98, 0.98, formatted_name, transform=self.parent.ax.transAxes,
                                fontsize=self.parent.core_level_text_size, fontweight='bold',
                                va='top', ha='right')



        # Set x-axis direction based on data type
        if is_raman:
            self.parent.ax.set_xlabel("Wavenumber (cm$^{-1}$)")
            self.parent.ax.set_ylabel("Intensity (CPS)")
            self.parent.ax.set_xlim(min(x_values), max(x_values))  # Normal direction for Raman
        else:
            # Set axes
            self.parent.ax.set_xlabel("Binding Energy (eV)")
            self.parent.ax.set_ylabel("Intensity (CPS)")
            self.parent.ax.set_xlim(max(x_values), min(x_values))  # Reversed for XPS

        # Remove any residual subplot temporarily
        if hasattr(self.parent.plot_manager, 'residuals_subplot') and self.parent.plot_manager.residuals_subplot:
            self.parent.figure.delaxes(self.parent.plot_manager.residuals_subplot)
            self.parent.plot_manager.residuals_subplot = None
            self.parent.ax.set_position([0.1, 0.125, 0.85, 0.85])
            self.parent.ax.get_xaxis().set_visible(True)

        # Apply text settings from preference window
        self.parent.plot_manager.apply_text_settings(self.parent)

        # Draw
        self.parent.canvas.draw_idle()

        # Restore the original residuals state in the manager
        self.parent.plot_manager.residuals_state = original_residuals_state



    def plot_multiple_sheets(self, sheet_names):
        """Plot multiple core levels together on the same graph"""
        if not sheet_names:
            return

        # Store the original residuals state
        original_residuals_state = self.parent.plot_manager.residuals_state

        # Set the first sheet as the active one in the parent window
        self.parent.sheet_combobox.SetValue(sheet_names[0])
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, sheet_names[0])

        # Clear the plot
        self.parent.ax.clear()

        # Remove any residual subplot temporarily
        if hasattr(self.parent.plot_manager, 'residuals_subplot') and self.parent.plot_manager.residuals_subplot:
            self.parent.figure.delaxes(self.parent.plot_manager.residuals_subplot)
            self.parent.plot_manager.residuals_subplot = None
            self.parent.ax.set_position([0.1, 0.125, 0.85, 0.85])
            self.parent.ax.get_xaxis().set_visible(True)

        # Track min/max x values
        x_min = float('inf')
        x_max = float('-inf')

        # Determine if normalization is needed
        normalize = self.norm_check.GetValue()
        norm_method = self.norm_type.GetValue()

        # For auto normalization, we need to calculate global min/max
        global_min = float('inf')
        global_max = float('-inf')

        # Check if all sheets are from the same column (core level)
        base_names = set(self.extract_base_name(name) for name in sheet_names)
        same_column = len(base_names) == 1
        column_name = list(base_names)[0] if same_column else None

        if normalize and norm_method == "Auto":
            # Get global min/max across all selected datasets
            for sheet_name in sheet_names:
                if sheet_name in self.parent.Data['Core levels']:
                    y_values = self.parent.Data['Core levels'][sheet_name]['Raw Data']
                    global_min = min(global_min, min(y_values))
                    global_max = max(global_max, max(y_values))

        # Clear the plot
        self.parent.ax.clear()

        # Plot each selected sheet
        for i, sheet_name in enumerate(sheet_names):
            if sheet_name in self.parent.Data['Core levels']:
                core_level = self.parent.Data['Core levels'][sheet_name]
                x_values = core_level['B.E.']
                y_values = np.array(core_level['Raw Data'])

                # Update min/max x values
                x_min = min(x_min, min(x_values))
                x_max = max(x_max, max(x_values))

                # Apply normalization if enabled
                if normalize:
                    if norm_method == "Auto":
                        # Original auto normalization code
                        norm_min = min(y_values)
                        norm_max = max(y_values)
                        # Avoid division by zero
                        if norm_max != norm_min:
                            y_values = (y_values - norm_min) / (norm_max - norm_min) * 1000
                    elif norm_method == "Norm. @ BE":
                        # Original auto normalization code
                        norm_min = min(y_values)
                        norm_max = max(y_values)


                        # Get the BE value for normalization
                        row_found = -1
                        col_found = -1
                        for row in range(self.grid.GetNumberRows()):
                            for col in range(1, len(self.core_levels) + 1):
                                if self.grid.GetCellValue(row, col) == sheet_name:
                                    row_found = row
                                    col_found = col
                                    break
                            if row_found >= 0:
                                break

                        if row_found >= 0:
                            norm_be_str = self.grid.GetCellValue(row_found, len(self.core_levels) + 2)
                            norm_area_str = self.grid.GetCellValue(row_found, len(self.core_levels) + 3)

                            try:
                                # Normalize at a specific binding energy
                                norm_be = float(norm_be_str) if norm_be_str else None
                                if norm_be is not None:
                                    # Find closest x value
                                    closest_idx = np.argmin(np.abs(np.array(x_values) - norm_be))
                                    norm_value = y_values[closest_idx] - norm_min
                                    # Avoid division by zero
                                    if norm_value != 0:
                                        y_values = (y_values - norm_min) / norm_value * 1000

                            except ValueError:
                                pass
                    elif norm_method == "Norm. to A":
                        # Get the normalization factor directly from the "Norm. to A" column
                        row_found = -1
                        col_found = -1
                        for row in range(self.grid.GetNumberRows()):
                            for col in range(1, len(self.core_levels) + 1):
                                if self.grid.GetCellValue(row, col) == sheet_name:
                                    row_found = row
                                    col_found = col
                                    break
                            if row_found >= 0:
                                break

                        if row_found >= 0:
                            norm_area_str = self.grid.GetCellValue(row_found, len(self.core_levels) + 3)
                            try:
                                # Simply multiply the y values by the factor from the column
                                norm_factor = float(norm_area_str) if norm_area_str else None
                                if norm_factor is not None:
                                    norm_min = min(y_values)
                                    y_values = (y_values - norm_min) / norm_factor * 1000
                            except ValueError:
                                pass

                # Use a different color for each plot
                color = self.parent.peak_colors[i % len(self.parent.peak_colors)]

                # Plot the data
                if self.parent.energy_scale == 'KE':
                    self.parent.ax.plot(self.parent.photons - x_values, y_values, label=sheet_name, color=color,
                                        linewidth=self.parent.line_width)
                else:
                    self.parent.ax.plot(x_values, y_values, label=sheet_name, color=color,
                                        linewidth=self.parent.line_width)

        # Set labels and formatting
        self.parent.ax.set_xlabel("Binding Energy (eV)")
        self.parent.ax.set_ylabel("Intensity (CPS)")

        # Apply scientific format to Y-axis
        self.parent.ax.yaxis.set_major_formatter(ScalarFormatter(useMathText=True))
        self.parent.ax.ticklabel_format(style='sci', axis='y', scilimits=(0, 0))

        # Set legend on the left
        self.parent.ax.legend(loc='upper left')

        # Set labels and formatting
        self.parent.ax.set_xlabel("Binding Energy (eV)")
        if normalize:
            self.parent.ax.set_ylabel("Normalized Intensity")
        else:
            self.parent.ax.set_ylabel("Intensity (CPS)")

        # Set x-axis limits to min/max values from all datasets
        self.parent.ax.set_xlim(x_max, x_min)  # Reversed for XPS

        # If all sheets are from the same column, add core level text in top right
        if same_column:
            formatted_name = self.parent.plot_manager.format_sheet_name(column_name)
            sheet_name_text = self.parent.ax.text(
                0.98, 0.98,  # Position (top-right corner)
                formatted_name,
                transform=self.parent.ax.transAxes,
                fontsize=self.parent.core_level_text_size,
                fontfamily=[self.parent.plot_font],
                fontweight='bold',
                verticalalignment='top',
                horizontalalignment='right',
                bbox=dict(facecolor='none', edgecolor='none', alpha=1),
            )
            sheet_name_text.sheet_name_text = True  # Mark this text object

        # Apply text settings from preferences
        self.parent.ax.tick_params(axis='both', labelsize=self.parent.axis_number_size)
        self.parent.ax.xaxis.label.set_size(self.parent.axis_title_size)
        self.parent.ax.yaxis.label.set_size(self.parent.axis_title_size)

        # Update the plot
        self.parent.canvas.draw_idle()

        # Restore the original residuals state
        self.parent.plot_manager.residuals_state = original_residuals_state

    def replot_with_normalization(self):
        """Replot the current selection with updated normalization settings"""
        sheet_names = self.get_selected_sheet_names()
        if sheet_names:
            self.plot_multiple_sheets(sheet_names)

    def on_cell_focus_changed(self, event):
        """Save sample names when cell focus changes"""
        if event.GetCol() == 0:  # Sample name column
            self.save_sample_names()
        event.Skip()

    def on_cell_changing(self, event):
        """Handle cell edit event for renaming and repositioning core levels"""
        row = event.GetRow()
        col = event.GetCol()
        old_value = self.grid.GetCellValue(row, col)
        new_value = event.GetString()

        # Prevent editing BE correction column
        if col == len(self.core_levels) + 1:  # BE correction column
            event.Veto()
            return

        # Handle sample name column separately
        if col == 0:
            # Just update the sample name
            self.sample_names[str(row)] = new_value
            self.save_sample_names()
            event.Skip()
            return

        # Handle BE correction column separately
        if col == len(self.core_levels) + 1:  # BE correction column
            try:
                new_correction = float(new_value)

                # Update parent.Data['BEcorrections']
                if 'BEcorrections' not in self.parent.Data:
                    self.parent.Data['BEcorrections'] = {}
                self.parent.Data['BEcorrections'][str(row)] = new_correction

                # Check if current sheet in main window belongs to this row
                current_sheet = self.parent.sheet_combobox.GetValue()
                sheet_found = False
                for cell_col in range(1, len(self.core_levels) + 1):
                    if self.grid.GetCellValue(row, cell_col) == current_sheet:
                        sheet_found = True
                        # Update main window spinbox and apply correction
                        self.parent.be_correction = new_correction
                        self.parent.be_correction_spinbox.SetValue(new_correction)
                        self.parent.apply_be_correction(new_correction)
                        break

                # Even if this row doesn't contain the current sheet, save corrections
                if not sheet_found:
                    self.save_be_corrections()

            except ValueError:
                # wx.MessageBox("BE correction must be a number", "Invalid Value", wx.OK | wx.ICON_ERROR)
                self.parent.show_popup_message2( "Invalid Value", "BE correction must be a number")
                event.Veto()
                return

        # Only process if there's a real change and the cell isn't empty
        if old_value and old_value != new_value and new_value.strip():
            # Check if old name exists in parent data
            if old_value in self.parent.Data['Core levels']:
                # Check if new name already exists
                if new_value in self.parent.Data['Core levels']:
                    # wx.MessageBox(f"A core level named '{new_value}' already exists.",
                    #               "Duplicate Name", wx.OK | wx.ICON_ERROR)
                    self.parent.show_popup_message2("Duplicate Name", f"A core level named '{new_value}' already "
                                                                      f"exists.")
                    event.Veto()
                    return

                # Set the sheet in parent before renaming
                self.parent.sheet_combobox.SetValue(old_value)
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, old_value)

                # Perform the rename
                from libraries.Utilities import rename_sheet
                rename_sheet(self.parent, new_value)

                # Always refresh the grid to properly position items
                wx.CallAfter(self.populate_grid)

    def on_cell_click(self, event):
        row = event.GetRow()
        col = event.GetCol()
        cell_value = self.grid.GetCellValue(row, col)

        # Process the event to select the cell
        event.Skip()

        # Check if shift or ctrl is being held down
        if not wx.GetKeyState(wx.WXK_SHIFT) and not wx.GetKeyState(wx.WXK_CONTROL):
            if cell_value and cell_value in self.parent.Data['Core levels']:
                wx.CallAfter(self.quick_plot_sheet, cell_value)

    def on_cursor_changed(self, event):
        # Process the event first to change the cursor
        event.Skip()

        # Check if shift or ctrl is being held down
        if wx.GetKeyState(wx.WXK_SHIFT) or wx.GetKeyState(wx.WXK_CONTROL):
            return

        row = event.GetRow()  # Use event row instead of cursor position
        col = event.GetCol()  # Use event column instead of cursor position
        cell_value = self.grid.GetCellValue(row, col)
        if cell_value and cell_value in self.parent.Data['Core levels']:
            wx.CallAfter(self.quick_plot_sheet, cell_value)


            # THIS MAY TO BE DELETED AS IT IS TOO SLOW
            # Get BE correction for the current row
            be_col_index = len(self.core_levels) + 1
            be_correction = self.grid.GetCellValue(row, be_col_index)

            # Update parent's BE correction if valid
            if be_correction.strip():
                try:
                    correction_value = float(be_correction)
                    self.parent.be_correction = correction_value
                    self.parent.be_correction_spinbox.SetValue(correction_value)
                    # Apply the correction to the current view
                    # self.parent.apply_be_correction(correction_value)
                except ValueError:
                    pass  # Ignore invalid correction values

    def on_copy_OLD(self, event):
        """Copy the selected core levels with columns C and D preserved exactly"""
        import os
        import json
        import tempfile
        from copy import deepcopy

        sheet_names = self.get_selected_sheet_names()
        if not sheet_names:
            return

        clipboard_data = {}

        for sheet_name in sheet_names:
            if sheet_name in self.parent.Data['Core levels']:
                clipboard_data[sheet_name] = deepcopy(self.parent.Data['Core levels'][sheet_name])

                file_path = self.parent.Data.get('FilePath', '')
                if file_path and os.path.exists(file_path):
                    try:
                        # Read all data including columns C and D
                        import pandas as pd
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # Store column names
                        column_names = df.columns.tolist()
                        clipboard_data[sheet_name]['column_names'] = column_names

                        # Store exact data from columns C and D (indices 2 and 3)
                        if df.shape[1] > 3:
                            clipboard_data[sheet_name]['column_C_data'] = df.iloc[:, 2].tolist()
                            clipboard_data[sheet_name]['column_D_data'] = df.iloc[:, 3].tolist()

                    except Exception as e:
                        print(f"Error reading Excel data for {sheet_name}: {e}")

        # Show preview dialog
        preview_dialog = CoreLevelPreviewDialog(self, "Copy Core Levels", clipboard_data, "copy")
        if preview_dialog.ShowModal() != wx.ID_OK:
            preview_dialog.Destroy()
            return
        preview_dialog.Destroy()

        # Save to clipboard file
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')
        with open(clipboard_file, 'w') as f:
            json.dump(clipboard_data, f)

        # self.parent.show_popup_message2("Copy Successful",
        #                                 f"{len(clipboard_data)} core level(s) copied with exact C and D columns")

    def on_copy(self, event):
        """Copy the selected core levels with original uncorrected BE values"""
        import os
        import json
        import tempfile
        from copy import deepcopy

        sheet_names = self.get_selected_sheet_names()
        if not sheet_names:
            return

        clipboard_data = {}

        for sheet_name in sheet_names:
            if sheet_name in self.parent.Data['Core levels']:
                clipboard_data[sheet_name] = deepcopy(self.parent.Data['Core levels'][sheet_name])

                # Get BE correction for this sheet's row
                source_row = 0
                if "Raman_" in sheet_name or "Ra_" in sheet_name:
                    base_parts = sheet_name.split('_')
                    if len(base_parts) > 2 and base_parts[2].isdigit():
                        source_row = int(base_parts[2])
                else:
                    match_row = re.search(r'(\d+)$', sheet_name)
                    if match_row:
                        source_row = int(match_row.group(1))

                be_correction = self.parent.Data.get('BEcorrections', {}).get(str(source_row), 0.0)

                # Store original BE values (subtract the applied correction)
                original_be_values = [be - be_correction for be in clipboard_data[sheet_name]['B.E.']]
                clipboard_data[sheet_name]['B.E.'] = original_be_values

                # Also adjust background limits if present
                if 'Background' in clipboard_data[sheet_name]:
                    if 'Bkg Low' in clipboard_data[sheet_name]['Background'] and \
                            clipboard_data[sheet_name]['Background']['Bkg Low'] != '':
                        try:
                            clipboard_data[sheet_name]['Background']['Bkg Low'] -= be_correction
                        except (TypeError, ValueError):
                            pass

                    if 'Bkg High' in clipboard_data[sheet_name]['Background'] and \
                            clipboard_data[sheet_name]['Background']['Bkg High'] != '':
                        try:
                            clipboard_data[sheet_name]['Background']['Bkg High'] -= be_correction
                        except (TypeError, ValueError):
                            pass

                # Adjust peak positions if present
                if 'Fitting' in clipboard_data[sheet_name] and 'Peaks' in clipboard_data[sheet_name]['Fitting']:
                    for peak in clipboard_data[sheet_name]['Fitting']['Peaks'].values():
                        if 'Position' in peak:
                            peak['Position'] -= be_correction
                        if 'Constraints' in peak:
                            pos_constraint = peak['Constraints'].get('Position', '')
                            if pos_constraint and ',' in pos_constraint and not any(
                                    c in pos_constraint for c in 'ABCDEFGHIJKLMNOP'):
                                try:
                                    min_val, max_val = map(float, pos_constraint.split(','))
                                    peak['Constraints'][
                                        'Position'] = f"{min_val - be_correction:.2f},{max_val - be_correction:.2f}"
                                except ValueError:
                                    pass

                file_path = self.parent.Data.get('FilePath', '')
                if file_path and os.path.exists(file_path):
                    try:
                        # Read all data including columns C and D
                        import pandas as pd
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # Store column names
                        column_names = df.columns.tolist()
                        clipboard_data[sheet_name]['column_names'] = column_names

                        # Store exact data from columns C and D (indices 2 and 3)
                        if df.shape[1] > 3:
                            clipboard_data[sheet_name]['column_C_data'] = df.iloc[:, 2].tolist()
                            clipboard_data[sheet_name]['column_D_data'] = df.iloc[:, 3].tolist()

                    except Exception as e:
                        print(f"Error reading Excel data for {sheet_name}: {e}")

        # Show preview dialog
        preview_dialog = CoreLevelPreviewDialog(self, "Copy Core Levels", clipboard_data, "copy")
        if preview_dialog.ShowModal() != wx.ID_OK:
            preview_dialog.Destroy()
            return
        preview_dialog.Destroy()

        # Save to clipboard file
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')
        with open(clipboard_file, 'w') as f:
            json.dump(clipboard_data, f)

    def on_paste(self, event):
        """Paste the core levels with original column names and experimental description"""
        import json
        import tempfile
        import os
        import re
        import numpy as np
        import pandas as pd
        from copy import deepcopy

        # Get the clipboard file
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')

        if not os.path.exists(clipboard_file):
            wx.MessageBox("No core levels in clipboard", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        # Load the clipboard data
        try:
            with open(clipboard_file, 'r') as f:
                clipboard_data = json.load(f)
        except json.JSONDecodeError:
            wx.MessageBox("Invalid clipboard data", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        if not clipboard_data:
            wx.MessageBox("Clipboard is empty", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        # Show preview dialog
        preview_dialog = CoreLevelPreviewDialog(self, "Paste Core Levels", clipboard_data, "paste")
        if preview_dialog.ShowModal() != wx.ID_OK:
            preview_dialog.Destroy()
            return
        preview_dialog.Destroy()

        # Perform backup before pasting
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(self.parent)

        # Get the clipboard file
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')

        if not os.path.exists(clipboard_file):
            wx.MessageBox("No core levels in clipboard", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        # Load the clipboard data
        try:
            with open(clipboard_file, 'r') as f:
                clipboard_data = json.load(f)
        except json.JSONDecodeError:
            wx.MessageBox("Invalid clipboard data", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        if not clipboard_data:
            wx.MessageBox("Clipboard is empty", "Paste Failed", wx.OK | wx.ICON_ERROR)
            return

        # Determine target row based on cursor position
        target_row = self.grid.GetGridCursorRow()

        # Get BE corrections for target row
        target_correction = self.parent.Data.get('BEcorrections', {}).get(str(target_row), 0.0)

        # Group core levels by base name (e.g., "C1s" without the number)
        core_level_groups = {}
        for sheet_name in clipboard_data.keys():
            # Extract the true base name (e.g., "C1s" from "C1s2")
            match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
            if match:
                base_name = match.group(1)
                if base_name not in core_level_groups:
                    core_level_groups[base_name] = []
                core_level_groups[base_name].append(sheet_name)

        # For each base name, determine the starting row
        base_name_row_map = {}
        current_row = target_row

        for base_name, sheets in core_level_groups.items():
            base_name_row_map[base_name] = current_row

            # If there are multiple sheets with the same base name,
            # increment the current row for the next base name
            if len(sheets) > 1:
                current_row += len(sheets)
            # If it's just one sheet, keep the same row for the next base name

        # Process each core level
        for sheet_name, core_level_data in clipboard_data.items():
            # Extract source row from original sheet name
            match_row = re.search(r'(\d+)$', sheet_name)
            source_row = match_row.group(1) if match_row else "0"

            # Try to get source correction from clipboard data, or default to 0.0
            # This allows pasting between different instances
            source_correction = 0.0
            if "BEcorrection" in core_level_data:
                source_correction = core_level_data["BEcorrection"]

            # Calculate the BE adjustment needed
            be_adjustment = source_correction - target_correction

            # Extract base name and determine which group it belongs to
            match_base = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
            if match_base:
                base_name = match_base.group(1)

                # Find position in the group to determine row offset
                group = core_level_groups[base_name]
                position = group.index(sheet_name)

                # Calculate the actual row for this sheet
                actual_row = base_name_row_map[base_name] + position

                # Create new sheet name
                new_sheet_name = f"{base_name}{actual_row}"

                # Check if new name already exists
                counter = 0
                while new_sheet_name in self.parent.Data['Core levels']:
                    counter += 1
                    new_sheet_name = f"{base_name}{actual_row}_{counter}"

                # Ensure the required fields are present in core_level_data
                if 'B.E.' not in core_level_data or 'Raw Data' not in core_level_data:
                    wx.MessageBox(f"Invalid data structure for {sheet_name}", "Paste Failed", wx.OK | wx.ICON_ERROR)
                    continue

                # Adjust BE values to remove previous correction and apply new row's correction
                adjusted_be_values = [be - be_adjustment for be in core_level_data['B.E.']]

                # Create new core level data with adjusted BE values
                new_core_level_data = deepcopy(core_level_data)
                new_core_level_data['B.E.'] = adjusted_be_values
                new_core_level_data['Name'] = new_sheet_name

                # Ensure background structure exists
                if 'Background' not in new_core_level_data:
                    new_core_level_data['Background'] = {
                        'Bkg Y': new_core_level_data['Raw Data'],
                        'Bkg Type': '',
                        'Bkg Low': min(adjusted_be_values),
                        'Bkg High': max(adjusted_be_values),
                        'Bkg Offset Low': 0,
                        'Bkg Offset High': 0
                    }

                # Adjust background limits if present
                if 'Background' in new_core_level_data:
                    if 'Bkg Low' in new_core_level_data['Background'] and new_core_level_data['Background'][
                        'Bkg Low'] != '':
                        try:
                            new_core_level_data['Background']['Bkg Low'] -= be_adjustment
                        except (TypeError, ValueError):
                            # Handle case where Bkg Low is not a number
                            new_core_level_data['Background']['Bkg Low'] = min(adjusted_be_values)

                    if 'Bkg High' in new_core_level_data['Background'] and new_core_level_data['Background'][
                        'Bkg High'] != '':
                        try:
                            new_core_level_data['Background']['Bkg High'] -= be_adjustment
                        except (TypeError, ValueError):
                            # Handle case where Bkg High is not a number
                            new_core_level_data['Background']['Bkg High'] = max(adjusted_be_values)

                    # Ensure Bkg Y exists
                    if 'Bkg Y' not in new_core_level_data['Background'] or not new_core_level_data['Background'][
                        'Bkg Y']:
                        new_core_level_data['Background']['Bkg Y'] = new_core_level_data['Raw Data']

                # Adjust peak positions if present
                if 'Fitting' in new_core_level_data and 'Peaks' in new_core_level_data['Fitting']:
                    for peak in new_core_level_data['Fitting']['Peaks'].values():
                        if 'Position' in peak:
                            peak['Position'] -= be_adjustment
                        if 'Constraints' in peak:
                            pos_constraint = peak['Constraints'].get('Position', '')
                            if pos_constraint and ',' in pos_constraint and not any(
                                    c in pos_constraint for c in 'ABCDEFGHIJKLMNOP'):
                                try:
                                    min_val, max_val = map(float, pos_constraint.split(','))
                                    peak['Constraints'][
                                        'Position'] = f"{min_val - be_adjustment:.2f},{max_val - be_adjustment:.2f}"
                                except ValueError:
                                    pass

                # Add to parent data
                if 'Core levels' not in self.parent.Data:
                    self.parent.Data['Core levels'] = {}
                if 'Number of Core levels' not in self.parent.Data:
                    self.parent.Data['Number of Core levels'] = 0

                self.parent.Data['Core levels'][new_sheet_name] = new_core_level_data
                self.parent.Data['Number of Core levels'] += 1

                # Update Excel file
                try:
                    # Get data length for this core level
                    data_length = len(adjusted_be_values)

                    # Ensure arrays have the correct length
                    raw_data = core_level_data['Raw Data']
                    if len(raw_data) > data_length:
                        raw_data = raw_data[:data_length]
                    elif len(raw_data) < data_length:
                        # Pad with zeros if too short
                        raw_data = raw_data + [0] * (data_length - len(raw_data))

                    # Get background data, defaulting to raw data if not available
                    bkg_data = core_level_data.get('Background', {}).get('Bkg Y', raw_data)
                    if len(bkg_data) > data_length:
                        bkg_data = bkg_data[:data_length]
                    elif len(bkg_data) < data_length:
                        # Pad with zeros if too short
                        bkg_data = bkg_data + [0] * (data_length - len(bkg_data))

                    # Check if it's a Raman file
                    is_raman = "Raman_" in new_sheet_name or "Ra_" in new_sheet_name

                    # Use original column names for Raman, standardized names for others
                    if is_raman:
                        column_names = core_level_data.get('column_names',
                                                           ['BE', 'Raw Data', 'Background', 'Transmission'])
                    else:
                        # Force standard column names for all XPS core levels and surveys
                        column_names = ['Binding Energy (eV)', 'Corrected Data', 'Raw Data', 'Transmission']

                    # Make sure we have enough column names
                    while len(column_names) < 4:
                        if is_raman:
                            column_names.append(f"Column{len(column_names) + 1}")
                        else:
                            default_names = ['Binding Energy (eV)', 'Corrected Data', 'Raw Data', 'Transmission']
                            if len(column_names) < len(default_names):
                                column_names.append(default_names[len(column_names)])
                            else:
                                column_names.append(f"Column{len(column_names) + 1}")

                    # Calculate original BE values by adding back the source correction
                    # original_be_values = [be - source_correction for be in core_level_data['B.E.']]

                    # Create DataFrame with original column names
                    data_dict = {
                        column_names[0]: core_level_data['B.E.'],
                        column_names[1]: raw_data
                    }

                    # Use the exact C and D data from clipboard if available
                    if 'column_C_data' in core_level_data and len(column_names) > 2:
                        col_c_data = core_level_data['column_C_data']
                        # Adjust length if needed
                        if len(col_c_data) != data_length:
                            if len(col_c_data) > data_length:
                                col_c_data = col_c_data[:data_length]
                            else:
                                col_c_data = col_c_data + [None] * (data_length - len(col_c_data))
                        data_dict[column_names[2]] = col_c_data
                    elif len(column_names) > 2:
                        data_dict[column_names[2]] = bkg_data

                    if 'column_D_data' in core_level_data and len(column_names) > 3:
                        col_d_data = core_level_data['column_D_data']
                        # Adjust length if needed
                        if len(col_d_data) != data_length:
                            if len(col_d_data) > data_length:
                                col_d_data = col_d_data[:data_length]
                            else:
                                col_d_data = col_d_data + [None] * (data_length - len(col_d_data))
                        data_dict[column_names[3]] = col_d_data
                    elif len(column_names) > 3:
                        data_dict[column_names[3]] = [1.0] * data_length

                    df = pd.DataFrame(data_dict)

                    # Write to Excel with original column names
                    with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a',
                                        if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=new_sheet_name, index=False)

                        # Remove borders from header row
                        workbook = writer.book
                        worksheet = workbook[new_sheet_name]
                        from openpyxl.styles import Border
                        no_border = Border()

                        # Apply no border to header row (row 1)
                        for col in range(1, len(column_names) + 1):
                            worksheet.cell(row=1, column=col).border = no_border

                        # Add experimental description data if available
                        if 'experimental_description' in core_level_data and core_level_data[
                            'experimental_description']:
                            # Calculate column 'AX' index (typically 49)
                            exp_col = 49

                            # Get the workbook and sheet
                            workbook = writer.book
                            worksheet = workbook[new_sheet_name]

                            # Add experimental description header
                            worksheet.cell(row=1, column=exp_col + 1, value="Experimental Description")
                            worksheet.cell(row=1, column=exp_col + 2, value="Value")

                            # Add all experimental description data
                            for i, item in enumerate(core_level_data['experimental_description']):
                                if len(item) >= 2:
                                    worksheet.cell(row=i + 2, column=exp_col + 1, value=item[0])
                                    worksheet.cell(row=i + 2, column=exp_col + 2, value=item[1])

                except Exception as e:
                    wx.MessageBox(f"Error writing to Excel: {str(e)}", "Warning", wx.OK | wx.ICON_WARNING)
                    # Continue even if Excel write fails

                # Update combobox in parent
                if hasattr(self.parent, 'sheet_combobox'):
                    self.parent.sheet_combobox.Append(new_sheet_name)

        # Refresh the grid
        self.populate_grid()

        # Update JSON file
        try:
            json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

            self.save_be_corrections()
            self.save_sample_names()
        except Exception as e:
            print(f"Error in final update steps: {e}")

        # Refresh sheets after pasting
        try:
            from libraries.Sheet_Operations import on_sheet_selected
            from libraries.Save import refresh_sheets
            refresh_sheets(window, on_sheet_selected, update_console)
        except Exception as refresh_err:
            print(f"Error refreshing sheets: {refresh_err}")

        # Close and reopen the file manager to refresh all columns
        self.parent.file_manager = None  # Clear the reference
        self.Destroy()  # Close current file manager
        wx.CallAfter(self.parent.on_open_file_manager, None)  # Reopen file manager

        self.parent.show_popup_message2("Paste Successful",
                                        f"{len(clipboard_data)} core level(s) pasted."
                                        f"\nCreated backup of the original data in Backup folder"
                                        f"\nBeware that core level retain their BE correction values")

    def on_rename(self, event):
        """Rename the selected core level"""
        save_state(self.parent)
        sheet_names = self.get_selected_sheet_names()

        if sheet_names:
            sheet_name = sheet_names[0]  # Use the first selected sheet

            # Set the sheet in parent before renaming
            self.parent.sheet_combobox.SetValue(sheet_name)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(self.parent, sheet_name)

            dlg = wx.TextEntryDialog(self, f"Enter new name for {sheet_name}:", "Rename Core Level", sheet_name)
            if dlg.ShowModal() == wx.ID_OK:
                new_name = dlg.GetValue()
                if new_name and new_name != sheet_name:
                    from libraries.Utilities import rename_sheet
                    rename_sheet(self.parent, new_name)

                    # Reload the grid after renaming
                    self.populate_grid()
            dlg.Destroy()

    def on_delete_OLD(self, event):
        """Delete selected core level(s)."""
        # Gather all sheet names to delete
        sheet_names = []

        # Check if cells are selected
        selected_cells = []
        for row in range(self.grid.GetNumberRows()):
            for col in range(1, len(self.core_levels) + 1):  # Skip sample name column
                if self.grid.IsInSelection(row, col):
                    cell_value = self.grid.GetCellValue(row, col)
                    if cell_value and cell_value in self.parent.Data['Core levels']:
                        sheet_names.append(cell_value)

        # If no cells selected, try current cursor position
        if not sheet_names:
            row = self.grid.GetGridCursorRow()
            col = self.grid.GetGridCursorCol()

            if col > 0 and col <= len(self.core_levels):
                cell_value = self.grid.GetCellValue(row, col)
                if cell_value and cell_value in self.parent.Data['Core levels']:
                    sheet_names.append(cell_value)

        # Remove duplicates
        sheet_names = list(set(sheet_names))

        if not sheet_names:
            self.parent.show_popup_message2("Information", "No core levels selected.")
            return

        # Confirm deletion
        if wx.MessageBox(f"Are you sure you want to delete {len(sheet_names)} core level(s)?",
                         "Confirm Delete", wx.YES_NO | wx.ICON_QUESTION) != wx.YES:
            return

        # Backup before deletion
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(self.parent)

        # Delete the sheets
        for sheet_name in sheet_names:
            # Delete the sheet from parent Data
            if sheet_name in self.parent.Data['Core levels']:
                del self.parent.Data['Core levels'][sheet_name]
                self.parent.Data['Number of Core levels'] -= 1

                # Also remove from Excel file if possible
                try:
                    import pandas as pd
                    from openpyxl import load_workbook

                    excel_path = self.parent.Data.get('FilePath', '')
                    if excel_path and os.path.exists(excel_path):
                        book = load_workbook(excel_path)
                        if sheet_name in book.sheetnames:
                            del book[sheet_name]
                            book.save(excel_path)
                except Exception as e:
                    print(f"Error removing sheet from Excel: {e}")

        # Save JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        from libraries.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(self.parent.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Update the parent's combobox
        if hasattr(self.parent, 'sheet_combobox'):
            current_sheet = self.parent.sheet_combobox.GetValue()
            self.parent.sheet_combobox.Clear()
            for sheet in self.parent.Data['Core levels'].keys():
                self.parent.sheet_combobox.Append(sheet)

            # Select an available sheet
            if current_sheet in self.parent.Data['Core levels']:
                self.parent.sheet_combobox.SetValue(current_sheet)
            elif self.parent.sheet_combobox.GetCount() > 0:
                self.parent.sheet_combobox.SetSelection(0)
                new_sheet = self.parent.sheet_combobox.GetValue()
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, new_sheet)

        # Close and reopen the file manager to refresh all columns
        self.parent.file_manager = None  # Clear the reference
        self.Destroy()  # Close current file manager
        wx.CallAfter(self.parent.on_open_file_manager, None)  # Reopen file manager

        self.parent.show_popup_message2("Success", f"Deleted {len(sheet_names)} core level(s).")

    def on_delete(self, event):
        """Delete selected core level(s)."""
        # Gather all sheet names to delete
        sheet_names = []

        # Check if cells are selected
        selected_cells = []
        for row in range(self.grid.GetNumberRows()):
            for col in range(1, len(self.core_levels) + 1):  # Skip sample name column
                if self.grid.IsInSelection(row, col):
                    cell_value = self.grid.GetCellValue(row, col)
                    if cell_value and cell_value in self.parent.Data['Core levels']:
                        sheet_names.append(cell_value)

        # If no cells selected, try current cursor position
        if not sheet_names:
            row = self.grid.GetGridCursorRow()
            col = self.grid.GetGridCursorCol()

            if col > 0 and col <= len(self.core_levels):
                cell_value = self.grid.GetCellValue(row, col)
                if cell_value and cell_value in self.parent.Data['Core levels']:
                    sheet_names.append(cell_value)

        # Remove duplicates
        sheet_names = list(set(sheet_names))

        if not sheet_names:
            self.parent.show_popup_message2("Information", "No core levels selected.")
            return

        # Show preview dialog
        preview_dialog = CoreLevelPreviewDialog(self, "Delete Core Levels",
                                                {name: self.parent.Data['Core levels'][name] for name in sheet_names},
                                                "delete")
        if preview_dialog.ShowModal() != wx.ID_OK:
            preview_dialog.Destroy()
            return
        preview_dialog.Destroy()

        # Backup before deletion
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(self.parent)

        # Delete the sheets
        for sheet_name in sheet_names:
            # Delete the sheet from parent Data
            if sheet_name in self.parent.Data['Core levels']:
                del self.parent.Data['Core levels'][sheet_name]
                self.parent.Data['Number of Core levels'] -= 1

                # Also remove from Excel file if possible
                try:
                    import pandas as pd
                    from openpyxl import load_workbook

                    excel_path = self.parent.Data.get('FilePath', '')
                    if excel_path and os.path.exists(excel_path):
                        book = load_workbook(excel_path)
                        if sheet_name in book.sheetnames:
                            del book[sheet_name]
                            book.save(excel_path)
                except Exception as e:
                    print(f"Error removing sheet from Excel: {e}")

        # Save JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        from libraries.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(self.parent.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Update the parent's combobox
        if hasattr(self.parent, 'sheet_combobox'):
            current_sheet = self.parent.sheet_combobox.GetValue()
            self.parent.sheet_combobox.Clear()
            for sheet in self.parent.Data['Core levels'].keys():
                self.parent.sheet_combobox.Append(sheet)

            # Select an available sheet
            if current_sheet in self.parent.Data['Core levels']:
                self.parent.sheet_combobox.SetValue(current_sheet)
            elif self.parent.sheet_combobox.GetCount() > 0:
                self.parent.sheet_combobox.SetSelection(0)
                new_sheet = self.parent.sheet_combobox.GetValue()
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, new_sheet)

        # Close and reopen the file manager to refresh all columns
        self.parent.file_manager = None  # Clear the reference
        self.Destroy()  # Close current file manager
        wx.CallAfter(self.parent.on_open_file_manager, None)  # Reopen file manager

        self.parent.show_popup_message2("Success", f"Deleted {len(sheet_names)} core level(s).")

    def on_preferences(self, event):
        dlg = wx.Dialog(self, title="Normalization Settings", size=(300, 200))
        panel = wx.Panel(dlg)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Min value input
        min_sizer = wx.BoxSizer(wx.HORIZONTAL)
        min_sizer.Add(wx.StaticText(panel, label="Min Value:"), 0, wx.ALL, 5)
        min_ctrl = wx.SpinCtrlDouble(panel, value=str(self.norm_min), min=0, max=1000000, inc=0.1)
        min_sizer.Add(min_ctrl, 1, wx.ALL, 5)

        # Max value input
        max_sizer = wx.BoxSizer(wx.HORIZONTAL)
        max_sizer.Add(wx.StaticText(panel, label="Max Value:"), 0, wx.ALL, 5)
        max_ctrl = wx.SpinCtrlDouble(panel, value=str(self.norm_max), min=0, max=1000000, inc=0.1)
        max_sizer.Add(max_ctrl, 1, wx.ALL, 5)

        # Buttons
        btn_sizer = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(panel, wx.ID_OK)
        cancel_btn = wx.Button(panel, wx.ID_CANCEL)
        btn_sizer.AddButton(ok_btn)
        btn_sizer.AddButton(cancel_btn)
        btn_sizer.Realize()

        # Add to main sizer
        sizer.Add(min_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(max_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 10)

        panel.SetSizer(sizer)

        if dlg.ShowModal() == wx.ID_OK:
            self.norm_min = min_ctrl.GetValue()
            self.norm_max = max_ctrl.GetValue()

        dlg.Destroy()

    def on_norm_changed(self, event):
        """Handle normalization checkbox toggle."""
        if self.norm_check.GetValue():
            # self.replot_with_normalization()
            if self.norm_type.GetValue() != "Auto":
                self.show_norm_cursors()
        else:
            self.hide_norm_cursors()
            self.replot_with_normalization()

    def on_norm_type_changed(self, event):
        """Handle normalization type selection change."""
        if self.norm_check.GetValue():
            norm_method = self.norm_type.GetValue()
            # Reapply normalization with new method setting
            self.replot_with_normalization()

            # Show or hide normalization cursors based on selected method
            if norm_method == "Auto" and not self.is_dragging_cursor:
                self.hide_norm_cursors()
            elif norm_method == "Norm @ BE":
                self.hide_norm_cursors()  # No cursors needed for BE normalization

    def on_auto_changed(self, event):
        if self.norm_check.GetValue():
            # Re-apply normalization with new Auto setting
            self.replot_with_normalization()
            # Show or hide normalization cursors based on auto setting
            if self.auto_check.GetValue():
                self.hide_norm_cursors()
            else:
                self.show_norm_cursors()


    def on_key_press(self, event):
        """Handle key press events"""
        try:
            # Only show the line if shift is pressed and Norm. @ BE is selected
            if event.key == 'shift' and self.norm_type.GetValue() == "Norm. @ BE":
                # Create vertical line at current mouse position
                if hasattr(self, 'be_norm_line') and self.be_norm_line is not None:
                    self.be_norm_line.remove()

                # Place line at mouse position if available, otherwise center of plot
                if event.xdata is not None:
                    x_pos = event.xdata
                else:
                    x_pos = (self.parent.ax.get_xlim()[0] + self.parent.ax.get_xlim()[1]) / 2

                self.be_norm_line = self.parent.ax.axvline(x_pos, color='red', linestyle='-',
                                                           linewidth=2, alpha=0.8)

                # Connect mouse events for dragging
                self.motion_id = self.parent.canvas.mpl_connect('motion_notify_event',
                                                                self.on_norm_line_motion)
                self.press_id = self.parent.canvas.mpl_connect('button_press_event',
                                                               self.on_norm_line_press)
                self.release_id = self.parent.canvas.mpl_connect('button_release_event',
                                                                 self.on_norm_line_release)

                # Initialize dragging state
                self.is_dragging = False

                # Draw the line
                self.parent.canvas.draw_idle()
        except RuntimeError:
            # Handle the case where the control has been deleted
            pass

    def on_key_release(self, event):
        """Handle key release events"""
        if event.key == 'shift':
            self.remove_norm_line()

    def on_norm_line_motion(self, event):
        """Handle mouse movement for the normalization line"""
        # Only process if shift is still pressed
        if not wx.GetKeyState(wx.WXK_SHIFT):
            self.remove_norm_line()
            return

        # Update line position if mouse is in plot area
        if event.inaxes and hasattr(self, 'be_norm_line') and self.be_norm_line is not None:
            # Get the current energy value from parent
            if hasattr(self.parent, 'current_energy_value'):
                x_value = self.parent.current_energy_value

                self.be_norm_line.set_xdata([x_value, x_value])
                self.parent.canvas.draw_idle()

                # If dragging, update normalization values in grid
                if self.is_dragging:
                    self.update_norm_be_values(x_value)

    def on_norm_line_press(self, event):
        """Handle mouse button press on the normalization line"""
        if not wx.GetKeyState(wx.WXK_SHIFT) or not event.inaxes or event.button != 1:
            return

        # Start dragging and updating values
        self.is_dragging = True

        # Update values immediately on press using parent's stored energy value
        if hasattr(self.parent, 'current_energy_value'):
            self.update_norm_be_values(self.parent.current_energy_value)

    def on_norm_line_release(self, event):
        """Handle mouse button release"""
        if self.is_dragging:
            self.is_dragging = False

            # Final update of values using parent's stored energy value
            if hasattr(self.parent, 'current_energy_value'):
                self.update_norm_be_values(self.parent.current_energy_value)

            # Remove the line
            self.remove_norm_line()

    def remove_norm_line(self):
        """Remove the normalization line and disconnect events"""
        # Remove the line
        if hasattr(self, 'be_norm_line') and self.be_norm_line is not None:
            self.be_norm_line.remove()
            self.be_norm_line = None

        # Disconnect events
        if hasattr(self, 'motion_id'):
            self.parent.canvas.mpl_disconnect(self.motion_id)
        if hasattr(self, 'press_id'):
            self.parent.canvas.mpl_disconnect(self.press_id)
        if hasattr(self, 'release_id'):
            self.parent.canvas.mpl_disconnect(self.release_id)

        # Reset dragging state
        self.is_dragging = False

        # Update the canvas
        self.parent.canvas.draw_idle()

    def update_norm_be_values(self, x_value):
        """Update the Norm. @ BE values in the grid with the given x value"""
        # Get selected cells to determine which rows to update
        selected_cells = self.get_selected_sheet_names()
        rows_to_update = []

        # Find grid rows for each selected sheet
        for sheet_name in selected_cells:
            for row in range(self.grid.GetNumberRows()):
                for col in range(1, len(self.core_levels) + 1):
                    if self.grid.GetCellValue(row, col) == sheet_name:
                        rows_to_update.append(row)
                        break

        # If no specific rows selected, update all rows
        if not rows_to_update:
            rows_to_update = list(range(self.grid.GetNumberRows()))

        # Update the BE normalization column for each row
        norm_be_col = len(self.core_levels) + 2
        for row in rows_to_update:
            self.grid.SetCellValue(row, norm_be_col, f"{x_value:.2f}")

        # Refresh grid and replot if normalization is active
        self.grid.ForceRefresh()
        if self.norm_check.GetValue():
            self.replot_with_normalization()



    def on_close(self, event):
        self.save_sample_names()
        self.save_be_corrections()  # Add this line
        self.parent.file_manager_position = self.GetPosition()
        event.Skip()

    def sort_excel_sheets(self, event):
        """Sort Excel sheets by sample group and element name"""
        save_state(self.parent)
        if not hasattr(self.parent, 'Data') or 'Core levels' not in self.parent.Data:
            wx.MessageBox("No data available to sort.", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get all sheet names from Data
        sheet_names = list(self.parent.Data['Core levels'].keys())

        # Group sheets by sample number
        grouped_sheets = {}

        for sheet_name in sheet_names:
            # Handle wide/survey scans specially
            if "wide" in sheet_name.lower() or "survey" in sheet_name.lower():
                match = re.match(r'(wide|survey)(\d*)$', sheet_name.lower(), re.IGNORECASE)
                if match:
                    base_name = match.group(1).capitalize()
                    sample_num = match.group(2)
                else:
                    base_name = sheet_name
                    sample_num = ""
            else:
                # Regular core level
                match = re.match(r'([A-Za-z]+\d*[spdfg]*)(\d*)$', sheet_name)
                if match:
                    base_name, sample_num = match.groups()
                else:
                    base_name = sheet_name
                    sample_num = ""

            sample_num = int(sample_num) if sample_num else 0

            if sample_num not in grouped_sheets:
                grouped_sheets[sample_num] = []

            grouped_sheets[sample_num].append((base_name, sheet_name))

        # Sort each group alphabetically by base name
        for sample_num in grouped_sheets:
            # Put "Wide" or "Survey" at the end of each group
            def sort_key(item):
                base = item[0].lower()
                if "wide" in base or "survey" in base:
                    return "zzz"  # This ensures these come last alphabetically
                return base

            grouped_sheets[sample_num].sort(key=sort_key)

        # Create final sorted list of sheet names
        sorted_sheet_names = []
        for sample_num in sorted(grouped_sheets.keys()):
            for _, sheet_name in grouped_sheets[sample_num]:
                sorted_sheet_names.append(sheet_name)

        # Check if already sorted
        if sheet_names == sorted_sheet_names:
            # wx.MessageBox("Sheets are already sorted.", "Information", wx.OK | wx.ICON_INFORMATION)
            self.parent.show_popup_message2("Information", "Sheets are already sorted.")
            return


        # Sort the Excel file
        try:
            import pandas as pd

            file_path = self.parent.Data['FilePath']

            # Check if file is accessible
            try:
                with open(file_path, 'rb') as f:
                    pass
            except PermissionError:
                wx.MessageBox("Cannot sort sheets: Excel file is open in another program.",
                              "File Locked", wx.OK | wx.ICON_ERROR)
                return

            # Read all data into memory
            data_frames = {}
            for sheet_name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                data_frames[sheet_name] = df

            # Write sheets in sorted order
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name in sorted_sheet_names:
                    data_frames[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

            # Update Data structure
            sorted_core_levels = {}
            for sheet_name in sorted_sheet_names:
                sorted_core_levels[sheet_name] = self.parent.Data['Core levels'][sheet_name]
            self.parent.Data['Core levels'] = sorted_core_levels

            # Update UI
            current_sheet = self.parent.sheet_combobox.GetValue()
            self.parent.sheet_combobox.Clear()
            for sheet_name in sorted_sheet_names:
                self.parent.sheet_combobox.Append(sheet_name)

            if current_sheet in sorted_sheet_names:
                self.parent.sheet_combobox.SetValue(current_sheet)
            elif sorted_sheet_names:
                self.parent.sheet_combobox.SetValue(sorted_sheet_names[0])
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, sorted_sheet_names[0])

            # Refresh grid
            self.populate_grid()

            # wx.MessageBox("Sheets sorted successfully.", "Success", wx.OK | wx.ICON_INFORMATION)
            self.parent.show_popup_message2("Success", "Sheets sorted successfully.")


        except Exception as e:
            wx.MessageBox(f"Error sorting sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def on_backup(self, event):
        """Create a backup of the current Excel and JSON files"""
        if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
            wx.MessageBox("No file currently open to backup.", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get current file paths
        excel_file = self.parent.Data['FilePath']
        json_file = os.path.splitext(excel_file)[0] + '.json'

        # Check if files exist
        if not os.path.exists(excel_file):
            wx.MessageBox(f"Excel file not found: {excel_file}", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Create backup folder in the executable directory
        executable_dir = os.path.dirname(os.path.abspath(sys.executable))
        # For development environment, fall back to current script directory
        if not "KherveFitting" in executable_dir:
            executable_dir = os.path.dirname(os.path.abspath(__file__))
            # Go up one level if in libraries folder
            if os.path.basename(executable_dir) == "libraries":
                executable_dir = os.path.dirname(executable_dir)

        backup_folder = os.path.join(executable_dir, "Backup")
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        # Generate timestamp in format: YYcDD_HHMMSS
        import datetime
        now = datetime.datetime.now()
        month_letter = chr(ord('a') + now.month - 1)  # a=Jan, b=Feb, c=Mar, etc.
        timestamp = f"{now.year % 100}{month_letter}{now.day:02d}_{now.hour:02d}{now.minute:02d}{now.second:02d}"

        # Create backup filenames
        excel_filename = os.path.basename(excel_file)
        excel_backup = os.path.join(backup_folder, f"{excel_filename}_{timestamp}")

        # Copy the Excel file
        try:
            shutil.copy2(excel_file, excel_backup)
            files_backed_up = [excel_backup]

            # Copy the JSON file if it exists
            if os.path.exists(json_file):
                json_filename = os.path.basename(json_file)
                json_backup = os.path.join(backup_folder, f"{json_filename}_{timestamp}")
                shutil.copy2(json_file, json_backup)
                files_backed_up.append(json_backup)

            # Show success message
            # wx.MessageBox(f"Backup created successfully:\n" + "\n".join(files_backed_up),
            #               "Backup Complete", wx.OK | wx.ICON_INFORMATION)
            self.parent.show_popup_message2("Backup Complete", f"Backup created successfully:\n" + "\n".join(
                files_backed_up))

        except Exception as e:
            # wx.MessageBox(f"Error creating backup: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
            self.parent.show_popup_message2("Error", f"Error creating backup: {str(e)}")

    def rename_from_context_menu(self, sheet_name):
        """Rename a core level from the right-click context menu"""
        save_state(self.parent)

        # Set the sheet in parent before renaming
        self.parent.sheet_combobox.SetValue(sheet_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, sheet_name)

        dlg = wx.TextEntryDialog(self, f"Enter new name for {sheet_name}:", "Rename Core Level", sheet_name)
        if dlg.ShowModal() == wx.ID_OK:
            new_name = dlg.GetValue()
            if new_name and new_name != sheet_name:
                from libraries.Utilities import rename_sheet
                rename_sheet(self.parent, new_name)
                # Reload the grid after renaming
                self.populate_grid()
        dlg.Destroy()

    def on_grid_right_click(self, event):
        """Handle right-click on grid to show context menu"""
        row = event.GetRow()
        col = event.GetCol()

        # Create context menu
        menu = wx.Menu()

        # Add standard menu items
        copy_item = menu.Append(wx.ID_ANY, "Copy Core Level(s)")
        paste_item = menu.Append(wx.ID_ANY, "Paste Core Level(s)")

        # Add rename option for core level cells only
        if col > 0 and col <= len(self.core_levels):  # Only for core level columns
            cell_value = self.grid.GetCellValue(row, col)
            if cell_value and cell_value in self.parent.Data['Core levels']:
                menu.AppendSeparator()
                rename_item = menu.Append(wx.ID_ANY, f"Rename '{cell_value}'")
                self.Bind(wx.EVT_MENU, lambda evt, sheet=cell_value: self.rename_from_context_menu(sheet), rename_item)

        # Add insert row option
        menu.AppendSeparator()
        insert_row_item = menu.Append(wx.ID_ANY, f"Insert Row")
        self.Bind(wx.EVT_MENU, lambda evt, r=row: self.on_insert_row(r), insert_row_item)

        delete_row_item = menu.Append(wx.ID_ANY, f"Delete Row {row}")
        self.Bind(wx.EVT_MENU, lambda evt, r=row: self.on_delete_row(r), delete_row_item)

        # Check if paste should be enabled (clipboard has data)
        import os
        import tempfile
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')
        paste_item.Enable(os.path.exists(clipboard_file))

        # Bind events
        self.Bind(wx.EVT_MENU, self.on_copy, copy_item)
        self.Bind(wx.EVT_MENU, self.on_paste, paste_item)

        # Add normalization propagation options for BE and Area normalization columns
        norm_be_col = len(self.core_levels) + 2
        norm_area_col = len(self.core_levels) + 3

        if col == norm_be_col or col == norm_area_col:
            cell_value = self.grid.GetCellValue(row, col)
            if cell_value:
                menu.AppendSeparator()
                propagate_item = menu.Append(wx.ID_ANY, f"Propagate {cell_value} to all rows")
                self.Bind(wx.EVT_MENU, lambda evt, r=row, c=col, v=cell_value: self.propagate_norm_value(r, c, v),
                          propagate_item)

        # Show the menu
        self.grid.PopupMenu(menu)
        menu.Destroy()

    def propagate_norm_value(self, source_row, column, value):
        """Propagate a normalization value to all rows in the same column"""
        # Ask for confirmation
        col_label = self.grid.GetColLabelValue(column)
        if wx.MessageBox(f"Propagate value '{value}' to all rows in column '{col_label}'?",
                         "Confirm Propagation", wx.YES_NO | wx.ICON_QUESTION) != wx.YES:
            return

        # Apply the value to all rows
        for row in range(self.grid.GetNumberRows()):
            self.grid.SetCellValue(row, column, value)

        # Refresh grid
        self.grid.ForceRefresh()

        # If currently plotting multiple sheets, update the plot with new normalization
        self.replot_with_normalization()

        # wx.MessageBox(f"Value '{value}' propagated to all rows.", "Success", wx.OK | wx.ICON_INFORMATION)
        self.parent.show_popup_message2("Success", f"Value '{value}' propagated to all rows.")

    def highlight_current_sheet(self, sheet_name):
        """Highlight the cell containing the current sheet name"""
        # Clear existing highlights
        for row in range(self.grid.GetNumberRows()):
            for col in range(self.grid.GetNumberCols()):
                if self.grid.GetCellBackgroundColour(row, col) == wx.YELLOW:
                    base_color = wx.Colour(200, 245, 228) if col > 0 else wx.Colour(180, 235, 208)
                    self.grid.SetCellBackgroundColour(row, col, base_color)

        # Find and highlight the cell with sheet_name
        for row in range(self.grid.GetNumberRows()):
            for col in range(1, len(self.core_levels) + 1):  # Skip sample name column
                if self.grid.GetCellValue(row, col) == sheet_name:
                    self.grid.SetCellBackgroundColour(row, col, wx.YELLOW)
                    self.grid.MakeCellVisible(row, col)
                    break

        self.grid.ForceRefresh()

    def add_more_rows(self):
        """Add 10 more rows to the grid."""
        current_rows = self.grid.GetNumberRows()
        self.grid.AppendRows(10)

        # Set row labels for new rows
        for row in range(current_rows, current_rows + 10):
            self.grid.SetRowLabelValue(row, str(row))

        # Set appropriate background colors for new rows
        for row in range(current_rows, current_rows + 10):
            # Sample name column (Experiment)
            self.grid.SetCellBackgroundColour(row, 0, wx.Colour(230, 230, 230))

            # BE correction column (Xshift)
            be_col_index = len(self.core_levels) + 1
            if be_col_index < self.grid.GetNumberCols():
                self.grid.SetCellBackgroundColour(row, be_col_index, wx.Colour(230, 230, 230))
                self.grid.SetCellTextColour(row, be_col_index, wx.Colour(128, 128, 128))

            # Norm @ BE column
            norm_col_index = len(self.core_levels) + 2
            if norm_col_index < self.grid.GetNumberCols():
                self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(230, 230, 230))

            # Norm to A column
            norm_area_col_index = len(self.core_levels) + 3
            if norm_area_col_index < self.grid.GetNumberCols():
                self.grid.SetCellBackgroundColour(row, norm_area_col_index, wx.Colour(230, 230, 230))

        self.grid.ForceRefresh()

    def delete_last_rows(self):
        """Delete the last 2 rows from the grid."""
        current_rows = self.grid.GetNumberRows()
        if current_rows >= 2:
            self.grid.DeleteRows(current_rows - 2, 2)
            self.grid.ForceRefresh()
        else:
            wx.MessageBox("Not enough rows to delete.", "Warning", wx.OK | wx.ICON_WARNING)

    def add_single_row(self):
        """Add 1 row to the grid."""
        current_rows = self.grid.GetNumberRows()
        self.grid.AppendRows(1)

        # Set row label for new row
        self.grid.SetRowLabelValue(current_rows, str(current_rows))

        # Set appropriate background colors for new row
        row = current_rows
        # Sample name column (Experiment)
        self.grid.SetCellBackgroundColour(row, 0, wx.Colour(230, 230, 230))

        # BE correction column (Xshift)
        be_col_index = len(self.core_levels) + 1
        if be_col_index < self.grid.GetNumberCols():
            self.grid.SetCellValue(row, be_col_index, "0.0")
            self.grid.SetCellBackgroundColour(row, be_col_index, wx.Colour(230, 230, 230))
            self.grid.SetCellTextColour(row, be_col_index, wx.Colour(128, 128, 128))

        # Norm @ BE column
        norm_col_index = len(self.core_levels) + 2
        if norm_col_index < self.grid.GetNumberCols():
            self.grid.SetCellBackgroundColour(row, norm_col_index, wx.Colour(230, 230, 230))

        # Norm to A column
        norm_area_col_index = len(self.core_levels) + 3
        if norm_area_col_index < self.grid.GetNumberCols():
            self.grid.SetCellBackgroundColour(row, norm_area_col_index, wx.Colour(230, 230, 230))

        self.grid.ForceRefresh()



    def delete_single_row(self):
        """Delete the last row from the grid."""
        current_rows = self.grid.GetNumberRows()
        if current_rows >= 1:
            self.grid.DeleteRows(current_rows - 1, 1)
            self.grid.ForceRefresh()
        else:
            wx.MessageBox("No rows to delete.", "Warning", wx.OK | wx.ICON_WARNING)


    def plot_multiple_sheets_with_offset(self, sheet_names):
        """Plot multiple core levels with vertical offset between plots"""
        if not sheet_names:
            return

        # Use current time to determine if this is a rapid keypress
        import time
        current_time = time.time()

        # Check if this is the same set of sheets as last time
        if self.last_offset_sheets == sheet_names:
            # Check if the keypress was rapid (within threshold)
            if current_time - self.last_keypress_time < self.rapid_press_threshold:
                # Increment the offset multiplier for rapid presses
                self.offset_multiplier += 1
            else:
                # Reset multiplier if too much time has passed
                self.offset_multiplier = 1
        else:
            # Reset for new selection
            self.offset_multiplier = 1
            self.last_offset_sheets = sheet_names.copy()

        # Update last keypress time
        self.last_keypress_time = current_time

        # Store the original residuals state
        original_residuals_state = self.parent.plot_manager.residuals_state

        # Set the first sheet as the active one in the parent window
        self.parent.sheet_combobox.SetValue(sheet_names[0])
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, sheet_names[0])

        # Clear the plot
        self.parent.ax.clear()

        # Remove any residual subplot temporarily
        if hasattr(self.parent.plot_manager, 'residuals_subplot') and self.parent.plot_manager.residuals_subplot:
            self.parent.figure.delaxes(self.parent.plot_manager.residuals_subplot)
            self.parent.plot_manager.residuals_subplot = None
            self.parent.ax.set_position([0.1, 0.125, 0.85, 0.85])
            self.parent.ax.get_xaxis().set_visible(True)

        # Track min/max x values
        x_min = float('inf')
        x_max = float('-inf')

        # Determine if normalization is needed
        normalize = self.norm_check.GetValue()
        norm_method = self.norm_type.GetValue()

        # For auto normalization, we need to calculate global min/max
        global_min = float('inf')
        global_max = float('-inf')

        # Check if all sheets are from the same column (core level)
        base_names = set(self.extract_base_name(name) for name in sheet_names)
        same_column = len(base_names) == 1
        column_name = list(base_names)[0] if same_column else None

        if normalize and norm_method == "Auto":
            # Get global min/max across all selected datasets
            for sheet_name in sheet_names:
                if sheet_name in self.parent.Data['Core levels']:
                    y_values = self.parent.Data['Core levels'][sheet_name]['Raw Data']
                    global_min = min(global_min, min(y_values))
                    global_max = max(global_max, max(y_values))

        # Plot each selected sheet with progressive offset
        for i, sheet_name in enumerate(sheet_names):
            if sheet_name in self.parent.Data['Core levels']:
                core_level = self.parent.Data['Core levels'][sheet_name]
                x_values = core_level['B.E.']
                y_values = np.array(core_level['Raw Data'])                     

                # Update min/max x values
                x_min = min(x_min, min(x_values))
                x_max = max(x_max, max(x_values))

                # Apply normalization if enabled
                if normalize:
                    # print("Normalise_before")
                    if norm_method == "Auto":
                        # Auto normalization
                        norm_min = min(y_values)
                        norm_max = max(y_values)
                        # print("Normalise")
                        # Avoid division by zero
                        if norm_max != norm_min:
                            y_values = (y_values - norm_min) / (norm_max - norm_min) * 1000

                            # Add offset based on position and multiplier
                            offset = i * (0.1 * self.offset_multiplier) * 1000
                            y_values += offset

                    elif norm_method == "Norm. @ BE":
                        # Get the normalization point
                        norm_min = min(y_values)
                        norm_max = max(y_values)

                        row_found = -1
                        for row in range(self.grid.GetNumberRows()):
                            for col in range(1, len(self.core_levels) + 1):
                                if self.grid.GetCellValue(row, col) == sheet_name:
                                    row_found = row
                                    break
                            if row_found >= 0:
                                break

                        if row_found >= 0:
                            norm_be_str = self.grid.GetCellValue(row_found, len(self.core_levels) + 2)
                            try:
                                norm_be = float(norm_be_str) if norm_be_str else None
                                if norm_be is not None:
                                    closest_idx = np.argmin(np.abs(np.array(x_values) - norm_be))
                                    norm_value = y_values[closest_idx] - norm_min
                                    if norm_value != 0:
                                        y_values = (y_values - norm_min) / norm_value * 1000

                                        # Add offset based on position and multiplier
                                        offset = i * (0.1 * self.offset_multiplier) * 1000
                                        y_values += offset
                            except ValueError:
                                pass
                    elif norm_method == "Norm. to A":
                        # Use area normalization factor
                        row_found = -1
                        for row in range(self.grid.GetNumberRows()):
                            for col in range(1, len(self.core_levels) + 1):
                                if self.grid.GetCellValue(row, col) == sheet_name:
                                    row_found = row
                                    break
                            if row_found >= 0:
                                break

                        if row_found >= 0:
                            norm_area_str = self.grid.GetCellValue(row_found, len(self.core_levels) + 3)
                            try:
                                norm_factor = float(norm_area_str) if norm_area_str else None
                                if norm_factor is not None:
                                    norm_min = min(y_values)
                                    y_values = (y_values - norm_min) / norm_factor * 1000

                                    # Add offset based on position and multiplier
                                    offset = i * (0.1 * self.offset_multiplier) * 1000
                                    y_values += offset
                            except ValueError:
                                pass

                # Use a different color for each plot
                color = self.parent.peak_colors[i % len(self.parent.peak_colors)]

                # Plot the data
                if self.parent.energy_scale == 'KE':
                    self.parent.ax.plot(self.parent.photons - x_values, y_values, label=sheet_name, color=color,
                                        linewidth=self.parent.line_width)
                else:
                    self.parent.ax.plot(x_values, y_values, label=sheet_name, color=color,
                                        linewidth=self.parent.line_width)

        # Set labels and formatting
        self.parent.ax.set_xlabel("Binding Energy (eV)")
        if normalize:
            self.parent.ax.set_ylabel(f"Normalized Intensity (offset×{self.offset_multiplier / 10:.1f})")
        else:
            self.parent.ax.set_ylabel("Intensity (CPS)")

        # Apply scientific format to Y-axis
        self.parent.ax.yaxis.set_major_formatter(ScalarFormatter(useMathText=True))
        self.parent.ax.ticklabel_format(style='sci', axis='y', scilimits=(0, 0))

        # Set legend on the left
        self.parent.ax.legend(loc='upper left')

        # Set x-axis limits to min/max values from all datasets
        self.parent.ax.set_xlim(x_max, x_min)  # Reversed for XPS

        # If all sheets are from the same column, add core level text in top right
        if same_column:
            formatted_name = self.parent.plot_manager.format_sheet_name(column_name)
            sheet_name_text = self.parent.ax.text(
                0.98, 0.98,  # Position (top-right corner)
                formatted_name,
                transform=self.parent.ax.transAxes,
                fontsize=self.parent.core_level_text_size,
                fontfamily=[self.parent.plot_font],
                fontweight='bold',
                verticalalignment='top',
                horizontalalignment='right',
                bbox=dict(facecolor='none', edgecolor='none', alpha=1),
            )

        # Apply text settings from preferences
        self.parent.ax.tick_params(axis='both', labelsize=self.parent.axis_number_size)
        self.parent.ax.xaxis.label.set_size(self.parent.axis_title_size)
        self.parent.ax.yaxis.label.set_size(self.parent.axis_title_size)

        # Update the plot
        self.parent.canvas.draw_idle()

        # Restore the original residuals state
        self.parent.plot_manager.residuals_state = original_residuals_state

    def on_plot_selected_with_offset(self, event):
        """Plot the currently selected core level(s) with offset"""
        sheet_names = self.get_selected_sheet_names()

        if sheet_names:
            if len(sheet_names) == 1:
                # Single sheet - update combobox and plot
                self.parent.sheet_combobox.SetValue(sheet_names[0])
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, sheet_names[0])
            else:
                # Multiple sheets - offset plot
                self.plot_multiple_sheets_with_offset(sheet_names)
                # Update combobox with first sheet name
                self.parent.sheet_combobox.SetValue(sheet_names[0])

            # Highlight the selected cell(s)
            self.highlight_current_sheet(sheet_names[0])
            self.Raise()  # Bring the file manager window to the front

    def hide_norm_cursors(self):
        """Hide normalization cursors if they exist"""
        if hasattr(self, 'norm_vlines') and self.norm_vlines:
            for vline in self.norm_vlines:
                if vline is not None:
                    vline.remove()
            self.norm_vlines = [None, None]
        self.is_dragging_cursor = False
        self.parent.canvas.draw_idle()

    def show_norm_cursors(self):
        """Show normalization cursors for manual range selection"""
        # Implementation details would depend on how you want to display and interact with the cursors
        pass

    def on_view_exp_info(self, event):
        """Display experimental description information for the selected sheet"""
        sheet_names = self.get_selected_sheet_names()

        if not sheet_names:
            wx.MessageBox("No sheet selected.", "Information", wx.OK | wx.ICON_INFORMATION)
            return

        sheet_name = sheet_names[0]
        exp_window = ExperimentalDescriptionWindow(self, sheet_name)
        exp_window.Show()

    def on_insert_row(self, target_row):
        """Insert a new row above the target row, incrementing all higher row numbers"""
        if wx.MessageBox(f"Insert new row above row {target_row}?\nThis will increment all higher row numbers.",
                         "Confirm Insert Row", wx.YES_NO | wx.ICON_QUESTION) != wx.YES:
            return

        # Backup before operation
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(self.parent)

        # Get all core levels and group by row number
        core_levels_by_row = {}
        sheets_to_rename = []

        for sheet_name in list(self.parent.Data['Core levels'].keys()):
            # Handle Raman files with underscore
            if "Raman_" in sheet_name or "Ra_" in sheet_name:
                base_parts = sheet_name.split('_')
                base_name = base_parts[0] + "_" + base_parts[1]
                if len(base_parts) > 2 and base_parts[2].isdigit():
                    row_num = int(base_parts[2])
                else:
                    row_num = 0
            else:
                # Regular core level parsing
                match = re.match(r'([A-Za-z]+\d*[spdfg]*)(\d*)$', sheet_name)
                if match:
                    base_name = match.group(1)
                    row_str = match.group(2)
                    row_num = int(row_str) if row_str else 0
                else:
                    continue

            if row_num >= target_row:
                sheets_to_rename.append((sheet_name, base_name, row_num))

            if row_num not in core_levels_by_row:
                core_levels_by_row[row_num] = []
            core_levels_by_row[row_num].append(sheet_name)

        # Sort sheets to rename by row number (descending to avoid conflicts)
        sheets_to_rename.sort(key=lambda x: x[2], reverse=True)

        try:
            import pandas as pd
            from openpyxl import load_workbook

            file_path = self.parent.Data['FilePath']

            # Read Excel file
            wb = load_workbook(file_path)
            data_frames = {}

            # Store data for sheets that will be renamed
            for sheet_name in self.parent.Data['Core levels']:
                if sheet_name in wb.sheetnames:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    data_frames[sheet_name] = df

            # Rename sheets in reverse order (highest row numbers first)
            renamed_sheets = {}
            for old_name, base_name, old_row in sheets_to_rename:
                new_row = old_row + 1

                # Create new name
                if "Raman_" in old_name or "Ra_" in old_name:
                    new_name = f"{base_name}_{new_row}" if new_row > 0 else base_name
                else:
                    new_name = f"{base_name}{new_row}" if new_row > 0 else base_name

                # Update Data structure
                if old_name in self.parent.Data['Core levels']:
                    core_level_data = self.parent.Data['Core levels'][old_name]
                    core_level_data['Name'] = new_name
                    self.parent.Data['Core levels'][new_name] = core_level_data
                    del self.parent.Data['Core levels'][old_name]
                    renamed_sheets[old_name] = new_name

            # Update BE corrections - shift row numbers
            if 'BEcorrections' in self.parent.Data:
                new_be_corrections = {}
                for row_str, correction in self.parent.Data['BEcorrections'].items():
                    row_num = int(row_str)
                    if row_num >= target_row:
                        new_be_corrections[str(row_num + 1)] = correction
                    else:
                        new_be_corrections[row_str] = correction
                self.parent.Data['BEcorrections'] = new_be_corrections

            # Update sample names - shift row numbers
            new_sample_names = {}
            for row_str, name in self.sample_names.items():
                row_num = int(row_str)
                if row_num >= target_row:
                    new_sample_names[str(row_num + 1)] = name
                else:
                    new_sample_names[row_str] = name
            self.sample_names = new_sample_names
            self.parent.Data['SampleNames'] = self.sample_names

            # Write Excel file with new sheet names
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                # Write all sheets with updated names
                for old_name, df in data_frames.items():
                    if old_name in renamed_sheets:
                        new_name = renamed_sheets[old_name]
                        df.to_excel(writer, sheet_name=new_name, index=False)
                    else:
                        df.to_excel(writer, sheet_name=old_name, index=False)

            # Update parent combobox
            current_sheet = self.parent.sheet_combobox.GetValue()
            self.parent.sheet_combobox.Clear()
            for sheet_name in sorted(self.parent.Data['Core levels'].keys()):
                self.parent.sheet_combobox.Append(sheet_name)

            # Update current selection if it was renamed
            if current_sheet in renamed_sheets:
                new_current = renamed_sheets[current_sheet]
                self.parent.sheet_combobox.SetValue(new_current)
            elif current_sheet in self.parent.Data['Core levels']:
                self.parent.sheet_combobox.SetValue(current_sheet)
            elif self.parent.sheet_combobox.GetCount() > 0:
                self.parent.sheet_combobox.SetSelection(0)

            # Save JSON file
            json_file_path = os.path.splitext(file_path)[0] + '.json'
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

            # Refresh grid
            self.populate_grid()

            # self.parent.show_popup_message2("Success", f"Inserted row above row {target_row}. "
            #                                            f"{len(sheets_to_rename)} sheets renamed.")

        except Exception as e:
            self.parent.show_popup_message2("Error", f"Error inserting row: {str(e)}")

    def on_delete_row(self, target_row):
        """Delete all core levels in the target row and decrement higher row numbers"""
        # Find sheets in target row
        sheets_in_row = []
        for col in range(1, len(self.core_levels) + 1):
            cell_value = self.grid.GetCellValue(target_row, col)
            if cell_value and cell_value in self.parent.Data['Core levels']:
                sheets_in_row.append(cell_value)

        # Handle empty row case
        if not sheets_in_row:
            if wx.MessageBox(f"Delete empty row {target_row}?\n"
                             f"This will decrement all higher row numbers.",
                             "Confirm Delete Empty Row", wx.YES_NO | wx.ICON_QUESTION) != wx.YES:
                return
            sheets_to_delete = set()
        else:
            # Confirm deletion for non-empty row
            if wx.MessageBox(f"Delete all {len(sheets_in_row)} core level(s) in row {target_row}?\n"
                             f"This will decrement all higher row numbers.\n\n"
                             f"Sheets to delete: {', '.join(sheets_in_row)}",
                             "Confirm Delete Row", wx.YES_NO | wx.ICON_QUESTION) != wx.YES:
                return
            sheets_to_delete = set(sheets_in_row)

        # Backup before operation
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(self.parent)

        # Get all core levels and identify sheets to rename
        sheets_to_rename = []

        for sheet_name in list(self.parent.Data['Core levels'].keys()):
            # Handle Raman files with underscore
            if "Raman_" in sheet_name or "Ra_" in sheet_name:
                base_parts = sheet_name.split('_')
                base_name = base_parts[0] + "_" + base_parts[1]
                if len(base_parts) > 2 and base_parts[2].isdigit():
                    row_num = int(base_parts[2])
                else:
                    row_num = 0
            else:
                # Regular core level parsing
                match = re.match(r'([A-Za-z]+\d*[spdfg]*)(\d*)$', sheet_name)
                if match:
                    base_name = match.group(1)
                    row_str = match.group(2)
                    row_num = int(row_str) if row_str else 0
                else:
                    continue

            if row_num > target_row:
                sheets_to_rename.append((sheet_name, base_name, row_num))

        # Sort sheets to rename by row number (ascending to avoid conflicts)
        sheets_to_rename.sort(key=lambda x: x[2])

        try:
            import pandas as pd
            from openpyxl import load_workbook

            file_path = self.parent.Data['FilePath']

            # Read Excel file
            wb = load_workbook(file_path)
            data_frames = {}

            # Store data for sheets that will be kept
            for sheet_name in self.parent.Data['Core levels']:
                if sheet_name not in sheets_to_delete and sheet_name in wb.sheetnames:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    data_frames[sheet_name] = df

            # Delete sheets from Data structure (only if row wasn't empty)
            for sheet_name in sheets_to_delete:
                if sheet_name in self.parent.Data['Core levels']:
                    del self.parent.Data['Core levels'][sheet_name]
                    self.parent.Data['Number of Core levels'] -= 1

            # Rename sheets with higher row numbers (decrement by 1)
            renamed_sheets = {}
            for old_name, base_name, old_row in sheets_to_rename:
                new_row = old_row - 1

                # Create new name
                if "Raman_" in old_name or "Ra_" in old_name:
                    new_name = f"{base_name}_{new_row}" if new_row > 0 else base_name
                else:
                    new_name = f"{base_name}{new_row}" if new_row > 0 else base_name

                # Update Data structure
                if old_name in self.parent.Data['Core levels']:
                    core_level_data = self.parent.Data['Core levels'][old_name]
                    core_level_data['Name'] = new_name
                    self.parent.Data['Core levels'][new_name] = core_level_data
                    del self.parent.Data['Core levels'][old_name]
                    renamed_sheets[old_name] = new_name

                    # Update data_frames dict
                    if old_name in data_frames:
                        data_frames[new_name] = data_frames[old_name]
                        del data_frames[old_name]

            # Update BE corrections - remove target row and shift higher rows
            if 'BEcorrections' in self.parent.Data:
                new_be_corrections = {}
                for row_str, correction in self.parent.Data['BEcorrections'].items():
                    row_num = int(row_str)
                    if row_num == target_row:
                        continue  # Skip deleted row
                    elif row_num > target_row:
                        new_be_corrections[str(row_num - 1)] = correction
                    else:
                        new_be_corrections[row_str] = correction
                self.parent.Data['BEcorrections'] = new_be_corrections

            # Update sample names - remove target row and shift higher rows
            new_sample_names = {}
            for row_str, name in self.sample_names.items():
                row_num = int(row_str)
                if row_num == target_row:
                    continue  # Skip deleted row
                elif row_num > target_row:
                    new_sample_names[str(row_num - 1)] = name
                else:
                    new_sample_names[row_str] = name
            self.sample_names = new_sample_names
            self.parent.Data['SampleNames'] = self.sample_names

            # Write Excel file with updated sheet names (excluding deleted sheets)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df in data_frames.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Update parent combobox
            current_sheet = self.parent.sheet_combobox.GetValue()
            self.parent.sheet_combobox.Clear()
            for sheet_name in sorted(self.parent.Data['Core levels'].keys()):
                self.parent.sheet_combobox.Append(sheet_name)

            # Update current selection
            if current_sheet in sheets_to_delete:
                # Current sheet was deleted, select first available
                if self.parent.sheet_combobox.GetCount() > 0:
                    self.parent.sheet_combobox.SetSelection(0)
                    from libraries.Sheet_Operations import on_sheet_selected
                    on_sheet_selected(self.parent, self.parent.sheet_combobox.GetValue())
            elif current_sheet in renamed_sheets:
                # Current sheet was renamed
                new_current = renamed_sheets[current_sheet]
                self.parent.sheet_combobox.SetValue(new_current)
            elif current_sheet in self.parent.Data['Core levels']:
                self.parent.sheet_combobox.SetValue(current_sheet)

            # Save JSON file
            json_file_path = os.path.splitext(file_path)[0] + '.json'
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

            # Refresh grid
            self.populate_grid()

            # if sheets_to_delete:
            #     self.parent.show_popup_message2("Success",
            #                                     f"Deleted row {target_row} with {len(sheets_to_delete)} sheets. "
            #                                     f"{len(sheets_to_rename)} sheets renumbered.")
            # else:
            #     self.parent.show_popup_message2("Success", f"Deleted empty row {target_row}. "
            #                                                f"{len(sheets_to_rename)} sheets renumbered.")

        except Exception as e:
            self.parent.show_popup_message2("Error", f"Error deleting row: {str(e)}")

class CoreLevelPreviewDialog(wx.Dialog):
    def __init__(self, parent, title, core_levels_data, operation_type):
        super().__init__(parent, title=title, size=(290, 400))

        self.core_levels_data = core_levels_data
        self.operation_type = operation_type  # "copy", "paste", or "delete"

        panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Title label
        if operation_type == "copy":
            label_text = f"Core levels to copy ({len(core_levels_data)}):"
        elif operation_type == "paste":
            label_text = f"Core levels to paste ({len(core_levels_data)}):"
        else:  # delete
            label_text = f"Core levels to delete ({len(core_levels_data)}):"

        title_label = wx.StaticText(panel, label=label_text)
        main_sizer.Add(title_label, 0, wx.ALL, 10)

        # List control
        self.list_ctrl = wx.ListCtrl(panel, style=wx.LC_REPORT | wx.LC_SINGLE_SEL)
        self.list_ctrl.AppendColumn("Name", width=50)
        self.list_ctrl.AppendColumn("N# Pts", width=50)
        self.list_ctrl.AppendColumn("BE Range", width=80)
        self.list_ctrl.AppendColumn("Fitting?", width=70)

        # Populate list
        for i, (sheet_name, core_level) in enumerate(core_levels_data.items()):
            index = self.list_ctrl.InsertItem(i, sheet_name)

            # Data points count
            data_points = len(core_level.get('Raw Data', []))
            self.list_ctrl.SetItem(index, 1, str(data_points))

            # BE range
            be_values = core_level.get('B.E.', [])
            if be_values:
                be_range = f"{min(be_values):.1f} - {max(be_values):.1f}"
            else:
                be_range = "N/A"
            self.list_ctrl.SetItem(index, 2, be_range)

            # Has peaks
            has_peaks = "Yes" if 'Fitting' in core_level and 'Peaks' in core_level['Fitting'] else "No"
            self.list_ctrl.SetItem(index, 3, has_peaks)

        main_sizer.Add(self.list_ctrl, 1, wx.EXPAND | wx.ALL, 10)

        # Buttons
        btn_sizer = wx.StdDialogButtonSizer()
        if operation_type == "delete":
            ok_btn = wx.Button(panel, wx.ID_OK, "Delete")
            ok_btn.SetBackgroundColour(wx.Colour(255, 100, 100))  # Red background for delete
        else:
            ok_btn = wx.Button(panel, wx.ID_OK)
        cancel_btn = wx.Button(panel, wx.ID_CANCEL)
        btn_sizer.AddButton(ok_btn)
        btn_sizer.AddButton(cancel_btn)
        btn_sizer.Realize()

        main_sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 10)

        panel.SetSizer(main_sizer)
        self.CenterOnParent()


class ExperimentalDescriptionWindow(wx.Frame):
    def __init__(self, parent, sheet_name):
        super().__init__(parent, title="Experimental Description",
                         size=(600, 600), style=wx.DEFAULT_FRAME_STYLE)

        self.parent = parent
        self.sheet_name = sheet_name

        self.panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        self.grid = wx.grid.Grid(self.panel)
        self.grid.CreateGrid(30, 2)

        self.grid.SetColLabelValue(0, "Parameter")
        self.grid.SetColLabelValue(1, "Value")
        self.grid.SetColSize(0, 200)
        self.grid.SetColSize(1, 350)

        self.populate_grid()

        main_sizer.Add(self.grid, 1, wx.EXPAND | wx.ALL, 10)
        self.panel.SetSizer(main_sizer)
        self.CenterOnParent()

        from libraries.ConfigFile import set_consistent_fonts
        set_consistent_fonts(self)

    def populate_grid(self):
        """Populate the grid with experimental description data"""
        file_path = self.parent.parent.Data.get('FilePath', '')
        if not file_path or not os.path.exists(file_path):
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            if self.sheet_name not in wb.sheetnames:
                return

            sheet = wb[self.sheet_name]

            # Find experimental description column
            exp_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "Experimental Description":
                    exp_col = col
                    break

            if not exp_col:
                return

            # Count rows with data
            row_count = 0
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=exp_col).value:
                    row_count += 1
                else:
                    # Continue checking a few more rows before breaking
                    empty_count = 0
                    for r in range(row, min(row + 5, sheet.max_row + 1)):
                        if not sheet.cell(row=r, column=exp_col).value:
                            empty_count += 1
                    if empty_count >= 3:
                        break
                    row_count += 1

            # Resize grid if needed
            if row_count > self.grid.GetNumberRows():
                self.grid.AppendRows(row_count - self.grid.GetNumberRows())

            # Populate grid with data
            for i in range(row_count):
                row = i + 2  # Start from row 2 in Excel (after header)
                parameter = sheet.cell(row=row, column=exp_col).value
                value = sheet.cell(row=row, column=exp_col + 1).value

                if parameter:
                    self.grid.SetCellValue(i, 0, str(parameter))
                    self.grid.SetCellValue(i, 1, str(value) if value is not None else "")

                    self.grid.SetCellAlignment(i, 0, wx.ALIGN_LEFT, wx.ALIGN_CENTER)
                    self.grid.SetCellAlignment(i, 1, wx.ALIGN_LEFT, wx.ALIGN_CENTER)
                    self.grid.SetReadOnly(i, 0)
                    self.grid.SetReadOnly(i, 1)

            # Hide unused rows
            for i in range(row_count, self.grid.GetNumberRows()):
                self.grid.SetRowSize(i, 0)

        except Exception as e:
            wx.MessageBox(f"Error loading experimental description: {str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
