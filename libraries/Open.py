import wx
import os
import json
import openpyxl
import xlrd
import wx
import re
import sys
import pandas as pd
import struct
from pathlib import Path
import shutil
from vamas import Vamas
from openpyxl import Workbook
import numpy as np
from scipy.interpolate import interp1d
import pandas as pd
from openpyxl.styles import Alignment
from yadg.extractors.phi.spe import extract  # NOTE THIS LIBRARY HAS BEEN TRANSFORMED

from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
from libraries.Save import update_undo_redo_state, save_state
from libraries.Sheet_Operations import on_sheet_selected
from libraries.Grid_Operations import populate_results_grid


class ExcelDropTarget(wx.FileDropTarget):
    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        from libraries.Open import open_xlsx_file, open_vamas_file
        for file in filenames:
            if not any(file.lower().endswith(ext) for ext in ['.xlsx', '.xls', '.vms', '.kal', '.avg', '.spe',
                                                              '.mrs', '.1']):
                wx.MessageBox(f"Only .xlsx/.xls (Khervefitting or Avantage), .vms (Vamas), "
                              f".kal (Kratos), .avg (Thermo), .mrs, .1 (VG-Microtech) and .spe (Phi) files can be "
                              f"dropped.",
                              "Invalid File "
                                                                                                     "Type",
                              wx.OK | wx.ICON_ERROR)
                return False
            if file.lower().endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(file)
                    if "Titles" in wb.sheetnames:
                        wx.CallAfter(import_avantage_file_direct, self.window, file)
                    else:
                        wx.CallAfter(open_xlsx_file, self.window, file)
                    return True
                except Exception:
                    return False
            elif file.lower().endswith('.xls'):
                try:
                    wb = xlrd.open_workbook(file)
                    if "Titles" in wb.sheet_names():
                        wx.CallAfter(import_avantage_file_direct_xls, self.window, file)
                    else:
                        wx.CallAfter(open_xlsx_file, self.window, file)
                    return True
                except Exception:
                    print("Error opening Excel file:", sys.exc_info())
                    return False
            elif file.lower().endswith('.vms'):
                wx.CallAfter(open_vamas_file, self.window, file)
                return True
            elif file.lower().endswith('.kal'):
                wx.CallAfter(open_kal_file, self.window, file)
                return True
            elif file.lower().endswith('.avg'):
                wx.CallAfter(open_avg_file_direct, self.window, file)
                return True
            elif file.lower().endswith('.spe'):
                wx.CallAfter(open_spe_file, self.window, file)
                return True
            elif file.lower().endswith('.mrs'):
                wx.CallAfter(open_mrs_file, self.window, file)
                return True
            elif file.lower().endswith('.1'):
                wx.CallAfter(open_vg_microtech_file, self.window, file)
                return True
        return False


def load_library_data_WITHEXCEL():
   wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
   sheet = wb['Library']
   data = {}
   for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
       element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
       key = (element, orbital)
       if key not in data:
           data[key] = {}
       data[key][instrument] = {
           'position': position,
           'ds': ds,
           'rsf': rsf,
           'row': row_idx
       }
   return data


def load_library_data_JSON_ONLY():
    with open('KherveFitting_library.json', 'r') as f:
        json_data = json.load(f)

    # Convert string keys back to tuples
    data = {}
    for k, v in json_data.items():
        element, orbital = k.split('_')
        data[(element, orbital)] = v
    return data


def load_library_data():
    """Load library data from parquet or json file"""
    import os

    parquet_file = 'KherveFitting_library.parquet'
    json_file = 'KherveFitting_library.json'

    if os.path.exists(parquet_file):
        print("Convert Parquet Database")
        import pandas as pd
        df = pd.read_parquet(parquet_file)

        # Convert back to nested dict structure
        data = {}
        for _, row in df.iterrows():
            key = (row['element'], row['orbital'])
            if key not in data:
                data[key] = {}
            data[key][row['instrument']] = {
                'position': row['position'],
                'ds': row['ds'],
                'rsf': row['rsf'],
                'full_name': row['full_name'],
                'auger': row['auger'],
                'ke_be': row['ke_be']
            }
        return data

    elif os.path.exists(json_file):
        print("Convert Json Database")
        import json
        with open(json_file, 'r') as f:
            json_data = json.load(f)

        # Convert string keys back to tuples
        data = {}
        for k, v in json_data.items():
            element, orbital = k.split('_')
            data[(element, orbital)] = v
        return data

    else:
        raise FileNotFoundError("Neither parquet nor json library file found")




def load_library_data_NEWBUTNO():
    wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
    sheet = wb['Library']
    data = {}
    instruments = set()  # Create set for unique instruments

    for row in sheet.iter_rows(min_row=2, values_only=True):
        element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
        instruments.add(instrument)  # Add each instrument to set
        key = (element, orbital)
        if key not in data:
            data[key] = {}
        data[key][instrument] = {
            'position': position,
            'ds': ds,
            'rsf': rsf
        }
    return data, sorted(list(instruments))  # Return data and sorted instruments list

def load_recent_files_from_config(window):
    config = window.load_config()
    window.recent_files = config.get('recent_files', [])
    update_recent_files_menu(window)

def update_recent_files(window, file_path):
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    window.save_config()  # Call save_config directly on the window object


def open_spe_file_OLD_PHI(window, file_path):
    try:
        from yadg.extractors.phi.spe import extract
        import openpyxl

        # Extract data from SPE file
        data = extract(fn=file_path)

        # Create new Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Process each core level
        for core_level in data:
            # Create sheet with core level name
            ws = wb.create_sheet(title=core_level)

            # Get energy and intensity values
            energy = data[core_level].coords["E"].values
            intensity = data[core_level].data_vars["y"].values

            # Set column headers
            ws["A1"] = "BE"
            ws["B1"] = "Corrected Data"
            ws["C1"] = "Raw Data"
            ws["D1"] = "Transmission"

            # Fill data
            for i, (e, inten) in enumerate(zip(energy, intensity), start=2):
                ws[f"A{i}"] = e
                ws[f"B{i}"] = inten
                ws[f"C{i}"] = inten
                ws[f"D{i}"] = 1.0

        # Create Experimental Description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Read header information using binary mode
        with open(file_path, 'rb') as f:
            header_lines = []
            in_header = False
            for line in f:
                try:
                    decoded_line = line.decode('utf-8', errors='ignore').strip()
                    if decoded_line == 'SOFH':
                        in_header = True
                        continue
                    if decoded_line == 'EOFH':
                        break
                    if in_header and decoded_line:
                        header_lines.append(decoded_line)
                except UnicodeDecodeError:
                    continue

        # Write header info to exp sheet
        for i, line in enumerate(header_lines, start=1):
            if ':' in line:
                key, value = line.split(':', 1)
                exp_sheet[f"A{i}"] = key.strip()
                exp_sheet[f"B{i}"] = value.strip()

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the created Excel file using existing function
        open_xlsx_file(window, excel_path)

    except Exception as e:
        # wx.MessageBox(f"Error processing SPE file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")


def open_spe_file(window, file_path):
    """Process a PHI SPE file and convert it to Excel format"""
    import numpy as np
    import openpyxl
    import re
    import os
    import struct

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Read file content
        with open(file_path, 'rb') as f:
            content = f.read()

        # Extract header
        header_match = re.search(rb'SOFH(.*?)EOFH', content, re.DOTALL)
        if not header_match:
            raise ValueError("Cannot find header section (SOFH...EOFH)")

        header_text = header_match.group(1).decode('utf-8', errors='ignore')
        header_lines = [line.strip() for line in header_text.strip().split('\n')]

        # Extract intensity calibration coefficients
        a, b = 31.826, 0.229  # Default values
        for line in header_lines:
            if 'IntensityCalCoeff:' in line:
                _, coeffs = line.split(':', 1)
                a, b = map(float, coeffs.strip().split())
                print(f"Extracted transmission function coefficients: a={a}, b={b}")
                break

        # Find active regions
        regions = []
        for line in header_lines:
            if line.startswith('SpectralRegDef:'):
                parts = line.split()
                if len(parts) >= 9 and int(parts[2]) == 1:  # Active region
                    region = {
                        'number': int(parts[1]),
                        'name': parts[3],
                        'atomic_number': int(parts[4]),
                        'num_points': int(parts[5]),
                        'start_energy': float(parts[7]),
                        'end_energy': float(parts[8])
                    }

                    # Normalize region name
                    if region['name'] == 'Su1s':
                        region['name'] = 'Survey'

                    regions.append(region)

        # Extract metadata for experimental description
        metadata = {
            'Sample ID': os.path.basename(file_path),
            'Date': '1970/1/1',
            'Time': '0:0:0',
            'Technique': 'XPS',
            'Species & Transition': '',
            'Number of scans': '1',
            'Source Label': 'Al',
            'Source Energy': '1486.6',
            'Source width X': '100',
            'Source width Y': '100',
            'Pass Energy': '224',
            'Work Function': '4.339',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': 'Kinetic Energy',
            'X Units': 'eV',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '1',
            'Collection Time': '0.16',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Update metadata from header
        for line in header_lines:
            if 'FileDateTime:' in line:
                parts = line.split(':', 1)[1].strip().split()
                if len(parts) >= 2:
                    date_part = parts[0]
                    time_part = parts[1]
                    metadata['Date'] = date_part
                    metadata['Time'] = time_part
            elif 'PassEnergy:' in line:
                metadata['Pass Energy'] = line.split(':', 1)[1].strip()
            elif 'WorkFunction:' in line:
                metadata['Work Function'] = line.split(':', 1)[1].strip()
            elif 'Comments:' in line:
                metadata['Block Comment'] = line.split(':', 1)[1].strip()
                metadata['# Comment Lines'] = '1'

        # Get binary data section
        data_start = content.find(b'EOFH') + 4
        binary_data = content[data_start:]

        # Implementation of FindFirstValidOffset
        def find_first_valid_offset(binary_data, region):
            num_points = region['num_points']

            # Iterate through possible offsets
            for offset in range(0, min(5000, len(binary_data) - num_points * 4), 2):
                try:
                    # Check first value - must be significant
                    first_val = struct.unpack('<f', binary_data[offset:offset + 4])[0]
                    if first_val < 0.5:
                        continue

                    # Try reading all points
                    values = []
                    for i in range(num_points):
                        data_offset = offset + (i * 4)
                        if data_offset + 4 <= len(binary_data):
                            val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                            if 0 <= val < 1e6:  # Reasonable value check
                                values.append(val)
                            else:
                                values = []
                                break
                        else:
                            values = []
                            break

                    # Check if we have the correct number of values with meaningful variation
                    if len(values) == num_points:
                        max_val = max(values)
                        min_val = min(values)
                        if max_val - min_val > 50:  # Meaningful variation
                            return offset
                except Exception:
                    pass

            # Try looking for marker patterns
            for offset in range(0, 5000, 2):
                try:
                    # Look for a marker (value near zero)
                    marker_val = struct.unpack('<f', binary_data[offset:offset + 4])[0]

                    # If found possible marker, check data after it
                    if 0 <= marker_val < 1e-6:
                        data_offset = offset + 4

                        # Validate the first actual data value
                        first_data_val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]

                        # Data must start with significant value
                        if first_data_val > 100:
                            # Check all values can be read
                            values = []
                            valid = True
                            for i in range(num_points):
                                point_offset = data_offset + (i * 4)
                                if point_offset + 4 <= len(binary_data):
                                    val = struct.unpack('<f', binary_data[point_offset:point_offset + 4])[0]
                                    if 0 <= val < 1e6:
                                        values.append(val)
                                    else:
                                        valid = False
                                        break
                                else:
                                    valid = False
                                    break

                            if valid and len(values) == num_points:
                                return data_offset
                except Exception:
                    pass

            # Try common offset 114 as a last resort
            try:
                values = []
                valid = True
                for i in range(num_points):
                    data_offset = 114 + (i * 4)
                    if data_offset + 4 <= len(binary_data):
                        val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                        if 0 <= val < 1e6:
                            values.append(val)
                        else:
                            valid = False
                            break
                    else:
                        valid = False
                        break

                if valid and len(values) == num_points:
                    max_val = max(values)
                    min_val = min(values)
                    if max_val - min_val > 100:
                        return 114
            except Exception:
                pass

            # If all else fails
            return -1

        # Implementation of CalculateAllOffsets
        def calculate_all_offsets(binary_data, regions, first_offset):
            if first_offset < 0:
                return []

            offsets = [first_offset]
            current_offset = first_offset

            # Calculate offsets for remaining regions
            for i in range(len(regions) - 1):
                # Next offset = current + (current region points * 4)
                current_offset += regions[i]['num_points'] * 4
                offsets.append(current_offset)

                # Verify this offset works
                valid = True
                for j in range(regions[i + 1]['num_points']):
                    data_offset = current_offset + (j * 4)
                    if data_offset + 4 <= len(binary_data):
                        try:
                            val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                            if val < 0 or val >= 1e6:
                                valid = False
                                break
                        except:
                            valid = False
                            break
                    else:
                        valid = False
                        break

                # If not valid, scan for a better offset
                if not valid:
                    test_offset = current_offset - 20
                    while test_offset < current_offset + 100:
                        test_valid = True
                        for j in range(regions[i + 1]['num_points']):
                            data_offset = test_offset + (j * 4)
                            if data_offset + 4 <= len(binary_data):
                                try:
                                    val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                                    if val < 0 or val >= 1e6:
                                        test_valid = False
                                        break
                                except:
                                    test_valid = False
                                    break
                            else:
                                test_valid = False
                                break

                        if test_valid:
                            offsets[i + 1] = test_offset
                            current_offset = test_offset
                            break

                        test_offset += 4

            return offsets

        # Find first valid offset
        first_offset = find_first_valid_offset(binary_data, regions[0])
        print(f"Found first valid offset at: {first_offset}")

        if first_offset < 0:
            raise ValueError("Could not find valid data offset in file")

        # Calculate offsets for all regions
        region_offsets = calculate_all_offsets(binary_data, regions, first_offset)
        print(f"Calculated offsets for all regions: {region_offsets}")

        if not region_offsets or len(region_offsets) != len(regions):
            raise ValueError("Failed to calculate valid offsets for all regions")

        # Process each region
        for i, region in enumerate(regions):
            sheet_name = region['name']
            num_points = region['num_points']
            offset = region_offsets[i]

            # Create sheet
            ws = wb.create_sheet(title=sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy (eV)"
            ws["B1"] = "Corrected Data"
            ws["C1"] = "Raw Data"
            ws["D1"] = "Transmission"

            # Update region-specific metadata
            region_metadata = metadata.copy()
            region_metadata['Species & Transition'] = sheet_name
            region_metadata['X Start'] = str(region['start_energy'])
            region_metadata['X Step'] = str((region['end_energy'] - region['start_energy']) / (num_points - 1))
            region_metadata['Num Y Values'] = str(num_points)

            # Read intensity values
            intensity_values = []
            for j in range(num_points):
                data_offset = offset + (j * 4)
                if data_offset + 4 <= len(binary_data):
                    val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                    if 0 <= val < 1e6:  # Reasonable value check
                        intensity_values.append(val)
                    else:
                        intensity_values.append(0.0)  # Use 0 for invalid values
                else:
                    intensity_values.append(0.0)

            # Make sure we have the expected number of points
            if len(intensity_values) < num_points:
                intensity_values.extend([0.0] * (num_points - len(intensity_values)))

            # Calculate energy values and transmission
            energy_values = np.linspace(region['start_energy'], region['end_energy'], num_points)
            ke_values = 1486.6 - energy_values  # Al K-alpha energy
            transmission = a * np.power(ke_values, -b)
            corrected_intensity = np.array(intensity_values) / transmission

            # Write data
            for j, (e, ci, ri, t) in enumerate(zip(energy_values, corrected_intensity, intensity_values, transmission),
                                               start=2):
                ws[f"A{j}"] = float(e)
                ws[f"B{j}"] = float(ci)
                ws[f"C{j}"] = float(ri)
                ws[f"D{j}"] = float(t)

            # Add experimental description at column AX (50)
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            # Add metadata fields
            row = 2
            for key, value in region_metadata.items():
                ws.cell(row=row, column=exp_col, value=key)
                ws.cell(row=row, column=exp_col + 1, value=value)
                row += 1

            # Set column widths for experimental description
            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40

        # Create Experimental description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata to the experimental description sheet
        row = 1
        exp_sheet.cell(row=row, column=1, value="Experimental Description")
        row += 1

        for key, value in metadata.items():
            exp_sheet.cell(row=row, column=1, value=key)
            exp_sheet.cell(row=row, column=2, value=value)
            row += 1

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the Excel file
        from libraries.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")
        return False




def open_spe_file_dialog(window):
    with wx.FileDialog(window, "Open SPE file", wildcard="PHI files (*.spe)|*.spe",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_spe_file(window, file_path)



def validate_data(df):
    """Validate the processed data"""
    issues = []

    if df['Binding Energy'].isna().any():
        issues.append("Missing binding energy values detected")

    if df['Raw Data'].isna().any():
        issues.append("Missing intensity values detected")

    if not np.allclose(df['Raw Data'], df['Corrected Data']):
        issues.append("Mismatch between raw and corrected data")

    if not np.allclose(df['Transmission'], 1.0):
        issues.append("Transmission values not all set to 1.0")

    return issues


def get_core_level_from_filename(filename):
    """
    Extract core level name from MRS filename.
    Examples:
    - xxx_C.MRS → C1s
    - xxx_CL.MRS → Cl2p
    - xxx_SU.MRS → Survey
    - xxx_O.MRS → O1s
    """
    import re

    # Define mapping of filename suffixes to standard core level names
    core_level_map = {
        'C': 'C1s',
        'N': 'N1s',
        'O': 'O1s',
        'F': 'F1s',
        'S': 'S2p',
        'CL': 'Cl2p',
        'BR': 'Br3d',
        'I': 'I3d',
        'SI': 'Si2p',
        'P': 'P2p',
        'B': 'B1s',
        'LI': 'Li1s',
        'NA': 'Na1s',
        'K': 'K2p',
        'CA': 'Ca2p',
        'MG': 'Mg2p',
        'AL': 'Al2p',
        'TI': 'Ti2p',
        'V': 'V2p',
        'CR': 'Cr2p',
        'MN': 'Mn2p',
        'FE': 'Fe2p',
        'CO': 'Co2p',
        'NI': 'Ni2p',
        'CU': 'Cu2p',
        'ZN': 'Zn2p',
        'GA': 'Ga3d',
        'GE': 'Ge3d',
        'AS': 'As3d',
        'SE': 'Se3d',
        'RB': 'Rb3d',
        'SR': 'Sr3d',
        'Y': 'Y3d',
        'ZR': 'Zr3d',
        'NB': 'Nb3d',
        'MO': 'Mo3d',
        'RU': 'Ru3d',
        'RH': 'Rh3d',
        'PD': 'Pd3d',
        'AG': 'Ag3d',
        'CD': 'Cd3d',
        'IN': 'In3d',
        'SN': 'Sn3d',
        'SB': 'Sb3d',
        'TE': 'Te3d',
        'CS': 'Cs3d',
        'BA': 'Ba3d',
        'LA': 'La3d',
        'CE': 'Ce3d',
        'HF': 'Hf4f',
        'TA': 'Ta4f',
        'W': 'W4f',
        'RE': 'Re4f',
        'OS': 'Os4f',
        'IR': 'Ir4f',
        'PT': 'Pt4f',
        'AU': 'Au4f',
        'HG': 'Hg4f',
        'PB': 'Pb4f',
        'BI': 'Bi4f',
        'TH': 'Th4f',
        'U': 'U4f',
        'SU': 'Survey',
        'VB': 'VB'  # Valence Band
    }

    # Extract the suffix after the last underscore and before the extension
    base_name = os.path.basename(filename)
    match = re.search(r'_([A-Z]+)\.mrs$', base_name.upper())

    if match:
        suffix = match.group(1)
        return core_level_map.get(suffix, suffix)
    # If no underscore or no match in mapping, try to use the whole filename
    # This is a fallback for unusual naming conventions
    base_name_without_ext = os.path.splitext(base_name)[0]
    for key, value in core_level_map.items():
        if key in base_name_without_ext.upper():
            return value

    # Last resort: just return the base filename without extension
    return base_name_without_ext


def open_mrs_file(window, file_path):
    """
    Import a Physical Electronics MRS file (XPS data) and convert it to Excel format
    suitable for KherveFitting.
    """
    try:
        # Read the MRS file content
        with open(file_path, 'r', errors='ignore') as f:
            content = f.read()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Extract sample info
        sample_name = os.path.splitext(os.path.basename(file_path))[0]
        desc_match = re.search(r'desc=(.*?)[\r\n]', content)
        sample_description = desc_match.group(1).strip() if desc_match else "Unknown"

        # Determine sheet name from filename using the new function
        sheet_name = get_core_level_from_filename(file_path)

        # Create sheet
        ws = wb.create_sheet(title=sheet_name)

        # Find data section
        data_section_match = re.search(
            r'array_size=(\d+).*?data=Data Array.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
            content, re.DOTALL)

        # If not found, try with Auto Survey
        if not data_section_match:
            data_section_match = re.search(
                r'array_size=(\d+).*?data=Auto Survey.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                content, re.DOTALL)

        # If still not found, try a more general pattern
        if not data_section_match:
            data_section_match = re.search(r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                                           content, re.DOTALL)

        if data_section_match:
            array_size = int(data_section_match.group(1))
            be_start = float(data_section_match.group(2))
            be_end = float(data_section_match.group(3))
            data_text = data_section_match.group(4).strip()

            # Extract numeric values
            data_values = []
            for line in data_text.split('\n'):
                line = line.strip()
                if line and line.isdigit():
                    data_values.append(int(line))

            if data_values:
                # Calculate binding energy values
                num_points = len(data_values)
                step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                be_values = [be_end - i * step_size for i in range(num_points)]

                # Add headers
                ws["A1"] = "Binding Energy (eV)"
                ws["B1"] = "Corrected Data"
                ws["C1"] = "Raw Data"
                ws["D1"] = "Transmission"

                # Add data
                for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                    ws[f"A{i}"] = be
                    ws[f"B{i}"] = intensity
                    ws[f"C{i}"] = intensity
                    ws[f"D{i}"] = 1.0  # Default transmission value

                # Create experimental description sheet
                exp_sheet = wb.create_sheet("Experimental description")
                exp_sheet.column_dimensions['A'].width = 30
                exp_sheet.column_dimensions['B'].width = 50

                # Add basic info to experimental sheet
                exp_sheet["A1"] = "File"
                exp_sheet["B1"] = os.path.basename(file_path)
                exp_sheet["A2"] = "Energy Range"
                exp_sheet["B2"] = f"{be_start} - {be_end} eV"
                exp_sheet["A3"] = "Number of Points"
                exp_sheet["B3"] = str(num_points)
                exp_sheet["A4"] = "Description"
                exp_sheet["B4"] = sample_description
                exp_sheet["A5"] = "Core Level"
                exp_sheet["B5"] = sheet_name

                # Extract additional metadata if available
                metadata_matches = {
                    "Scan Count": re.search(r'scan_total=(\d+)', content),
                    "Resolution": re.search(r'res=(\d+)', content),
                    "Spot Size": re.search(r'spot=(\d+)', content),
                    "Date": re.search(r'time_stamp=(.*?)[\r\n]', content),
                    "Operator": re.search(r'oper=(.*?)[\r\n]', content)
                }

                row = 6
                for key, match in metadata_matches.items():
                    if match:
                        exp_sheet[f"A{row}"] = key
                        exp_sheet[f"B{row}"] = match.group(1).strip()
                        row += 1

                # Save Excel file
                excel_path = os.path.splitext(file_path)[0] + ".xlsx"
                wb.save(excel_path)

                # Open the created Excel file
                from libraries.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
                return True
            else:
                window.show_popup_message2("Error", "No valid data found in the MRS file.")
                return False
        else:
            window.show_popup_message2("Error", "Could not locate data section in MRS file.")
            return False

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS file: {str(e)}")
        return False


def import_multiple_mrs_files(window):
    """
    Import multiple Physical Electronics MRS files from a folder.
    """
    with wx.DirDialog(window, "Choose a directory containing MRS files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all MRS files in the directory
        mrs_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.mrs')]

        if not mrs_files:
            window.show_popup_message2("Information", "No MRS files found in the selected folder.")
            return

        # Ask if user wants individual Excel files or one combined file
        dlg = wx.MessageDialog(window,
                               "Do you want to create individual Excel files for each MRS file or combine them into one Excel file?",
                               "Import Options",
                               wx.YES_NO | wx.ICON_QUESTION)
        dlg.SetYesNoLabels("Individual Files", "One Combined File")

        result = dlg.ShowModal()
        individual_files = (result == wx.ID_YES)
        dlg.Destroy()

        if individual_files:
            # Process each MRS file individually
            processed_count = 0
            for mrs_file in mrs_files:
                mrs_path = os.path.join(folder_path, mrs_file)
                if open_mrs_file(window, mrs_path):
                    processed_count += 1

            window.show_popup_message2("Success", f"Processed {processed_count} of {len(mrs_files)} MRS files.")
        else:
            # Combine all MRS files into one Excel file
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            processed_count = 0
            for mrs_file in mrs_files:
                mrs_path = os.path.join(folder_path, mrs_file)

                try:
                    # Read the MRS file content
                    with open(mrs_path, 'r', errors='ignore') as f:
                        content = f.read()

                    # Determine sheet name
                    sheet_name = get_core_level_from_filename(mrs_file)

                    # Ensure unique sheet name
                    suffix = 1
                    original_name = sheet_name
                    while sheet_name in combined_wb.sheetnames:
                        sheet_name = f"{original_name}{suffix}"
                        suffix += 1

                    # Ensure unique sheet name
                    suffix = 1
                    original_name = sheet_name
                    while sheet_name in combined_wb.sheetnames:
                        sheet_name = f"{original_name}{suffix}"
                        suffix += 1

                    # Create sheet
                    ws = combined_wb.create_sheet(title=sheet_name)

                    # Find and extract data
                    data_section_match = re.search(
                        r'array_size=(\d+).*?(?:data=Data Array|data=Auto Survey).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                        content, re.DOTALL)

                    if not data_section_match:
                        data_section_match = re.search(
                            r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                            content, re.DOTALL)

                    if data_section_match:
                        array_size = int(data_section_match.group(1))
                        be_start = float(data_section_match.group(2))
                        be_end = float(data_section_match.group(3))
                        data_text = data_section_match.group(4).strip()

                        # Extract numeric values
                        data_values = []
                        for line in data_text.split('\n'):
                            line = line.strip()
                            if line and line.isdigit():
                                data_values.append(int(line))

                        if data_values:
                            # Calculate binding energy values
                            num_points = len(data_values)
                            step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                            be_values = [be_end - i * step_size for i in range(num_points)]

                            # Add headers
                            ws["A1"] = "Binding Energy (eV)"
                            ws["B1"] = "Corrected Data"
                            ws["C1"] = "Raw Data"
                            ws["D1"] = "Transmission"

                            # Add data
                            for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                                ws[f"A{i}"] = be
                                ws[f"B{i}"] = intensity
                                ws[f"C{i}"] = intensity
                                ws[f"D{i}"] = 1.0  # Default transmission value

                            processed_count += 1
                except Exception as e:
                    print(f"Error processing {mrs_file}: {str(e)}")

            # Create experimental description sheet
            exp_sheet = combined_wb.create_sheet("Experimental description")
            exp_sheet.column_dimensions['A'].width = 30
            exp_sheet.column_dimensions['B'].width = 50

            exp_sheet["A1"] = "Source Folder"
            exp_sheet["B1"] = folder_path
            exp_sheet["A2"] = "Processed Files"
            exp_sheet["B2"] = f"{processed_count} of {len(mrs_files)}"

            # Save combined Excel file
            folder_name = os.path.basename(folder_path)
            excel_path = os.path.join(folder_path, f"{folder_name}_combined.xlsx")
            combined_wb.save(excel_path)

            # Open the combined Excel file
            if processed_count > 0:
                from libraries.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
                window.show_popup_message2("Success",
                                           f"Created combined Excel file with {processed_count} sheets from MRS files.")
            else:
                window.show_popup_message2("Error", "No valid data found in any MRS file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS files: {str(e)}")


def import_mrs_file(window):
    """
    Import a Physical Electronics MRS file (XPS data) and convert it to Excel format
    suitable for KherveFitting.
    """
    import wx
    import os
    import re
    import numpy as np
    from openpyxl import Workbook

    with wx.FileDialog(window, "Open MRS file", wildcard="MRS files (*.mrs)|*.mrs",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Read the MRS file content
        with open(file_path, 'r', errors='ignore') as f:
            content = f.read()

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Extract sample info
        sample_name = os.path.splitext(os.path.basename(file_path))[0]
        desc_match = re.search(r'desc=(.*?)[\r\n]', content)
        if desc_match:
            sample_description = desc_match.group(1).strip()

        # Determine sheet name from filename
        # base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Look for core level identifier in filename
        sheet_name = get_core_level_from_filename(file_path)
        if "_" in base_name:
            core_level = base_name.split("_")[1]
            if core_level.lower() in ["c1s", "o1s", "n1s", "s2p", "su", "vb"]:
                if core_level.lower() == "su":
                    sheet_name = "Survey"
                else:
                    sheet_name = core_level.upper()

        # Create sheet
        ws = wb.create_sheet(title=sheet_name)

        # Find data section
        data_section_match = re.search(
            r'array_size=(\d+).*?data=Data Array.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
            content, re.DOTALL)

        # If not found, try with Auto Survey
        if not data_section_match:
            data_section_match = re.search(
                r'array_size=(\d+).*?data=Auto Survey.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                content, re.DOTALL)

        # If still not found, try a more general pattern
        if not data_section_match:
            data_section_match = re.search(r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                                           content, re.DOTALL)

        if data_section_match:
            array_size = int(data_section_match.group(1))
            be_start = float(data_section_match.group(2))
            be_end = float(data_section_match.group(3))
            data_text = data_section_match.group(4).strip()

            # Extract numeric values
            data_values = []
            for line in data_text.split('\n'):
                line = line.strip()
                if line and line.isdigit():
                    data_values.append(int(line))

            if data_values:
                # Calculate binding energy values
                num_points = len(data_values)
                step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                be_values = [be_end - i * step_size for i in range(num_points)]

                # Add headers
                ws["A1"] = "Binding Energy (eV)"
                ws["B1"] = "Raw Data"

                # Add data
                for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                    ws[f"A{i}"] = be
                    ws[f"B{i}"] = intensity

                # Create experimental description sheet
                exp_sheet = wb.create_sheet("Experimental description")
                exp_sheet.column_dimensions['A'].width = 30
                exp_sheet.column_dimensions['B'].width = 50

                # Add basic info to experimental sheet
                exp_sheet["A1"] = "File"
                exp_sheet["B1"] = os.path.basename(file_path)
                exp_sheet["A2"] = "Energy Range"
                exp_sheet["B2"] = f"{be_start} - {be_end} eV"
                exp_sheet["A3"] = "Number of Points"
                exp_sheet["B3"] = str(num_points)

                if desc_match:
                    exp_sheet["A4"] = "Description"
                    exp_sheet["B4"] = sample_description

                # Save and open Excel file
                excel_path = os.path.splitext(file_path)[0] + ".xlsx"
                wb.save(excel_path)

                # Open the created Excel file
                from libraries.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
            else:
                window.show_popup_message2("Error", "No valid data found in the MRS file.")
        else:
            window.show_popup_message2("Error", "Could not locate data section in MRS file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS file: {str(e)}")


def import_avantage_file_direct(window, file_path):
    import re
    import openpyxl

    wb = openpyxl.load_workbook(file_path)
    new_file_path = os.path.splitext(file_path)[0] + "_Kfitting.xlsx"
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    sheets_to_process = []
    for sheet_name in wb.sheetnames:
        if "Survey" in sheet_name or "Scan" in sheet_name:
            sheets_to_process.append(sheet_name)

    for sheet_name in sheets_to_process:
        sheet = wb[sheet_name]

        # Extract element name (e.g., C1s, O1s)
        if "Survey" in sheet_name or "survey" in sheet_name:
            base_name = "Survey"
        else:
            parts = sheet_name.split()
            base_name = parts[0]  # Get element (C1s, O1s, etc.)

        # Check if multi-sample (D8 cell not empty and > 1)
        multi_sample = False
        num_samples = 1
        if sheet.cell(row=8, column=4).value is not None:
            try:
                num_samples = int(sheet.cell(row=8, column=4).value)
                if num_samples > 1:
                    multi_sample = True
            except (ValueError, TypeError):
                pass

        # Find data start row (default to 19 for Thermo files)
        start_row = 19
        for row_idx in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == "eV":
                start_row = row_idx + 1
                break

        if multi_sample:
            # Process each sample into its own sheet
            for sample_idx in range(num_samples):
                sample_name = f"{base_name}{sample_idx if sample_idx > 0 else ''}"
                new_sheet = new_wb.create_sheet(sample_name)
                new_sheet['A1'] = "Binding Energy"
                new_sheet['B1'] = "Raw Data"

                # A column is binding energy (col 1), data is in col C+sample_idx (col 3+sample_idx)
                data_col = 3 + sample_idx  # C is 3, D is 4, etc.

                # Copy data
                row_new = 2
                for row in range(start_row, sheet.max_row + 1):
                    be_value = sheet.cell(row=row, column=1).value
                    intensity_value = sheet.cell(row=row, column=data_col).value

                    if be_value is None or intensity_value is None:
                        continue

                    new_sheet.cell(row=row_new, column=1, value=be_value)
                    new_sheet.cell(row=row_new, column=2, value=intensity_value)
                    row_new += 1
        else:
            # Single sample - process as before
            new_name = base_name
            # Extract number if present in format like "C1s Scan (2)"
            number_match = re.search(r'\((\d+)\)', sheet_name)
            if number_match:
                number = number_match.group(1)
                new_name = f"{base_name}{number}"

            new_sheet = new_wb.create_sheet(new_name)
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"

            for row in range(start_row, sheet.max_row + 1):
                be_value = sheet.cell(row=row, column=1).value
                intensity_value = sheet.cell(row=row, column=3).value  # Column C

                if be_value is None or intensity_value is None:
                    continue

                new_sheet['A{}'.format(row - start_row + 2)] = be_value
                new_sheet['B{}'.format(row - start_row + 2)] = intensity_value

    new_wb.save(new_file_path)
    open_xlsx_file(window, new_file_path)

def import_avantage_file_direct_xls(window, file_path):
    import xlrd
    wb_xls = xlrd.open_workbook(file_path)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    for sheet_name in wb_xls.sheet_names():
        sheet = wb_xls.sheet_by_name(sheet_name)
        if "Survey" in sheet_name or "Scan" in sheet_name:
            # Handle Survey sheets
            if "Survey" in sheet_name or "survey" in sheet_name:
                # Extract number if present in "Survey Scan (2)" format
                number_match = re.search(r'\((\d+)\)', sheet_name)
                if number_match:
                    number = number_match.group(1)
                    new_name = f"Survey{number}"
                else:
                    new_name = "Survey"
            else:
                # Extract element name and number for patterns like "C1s Scan (1)"
                parts = sheet_name.split()
                element = parts[0]
                # Check if there's a number in parentheses
                number_match = re.search(r'\((\d+)\)', sheet_name)
                if number_match:
                    number = number_match.group(1)
                    new_name = f"{element}{number}"
                else:
                    new_name = element

            wb_new.create_sheet(new_name)
            new_sheet = wb_new[new_name]
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"

            start_row = 17
            for row_idx in range(sheet.nrows):
                if sheet.cell_value(row_idx, 0) == "eV":
                    start_row = row_idx + 1
                    break

            for row_idx in range(start_row, sheet.nrows):
                row_values = sheet.row_values(row_idx)
                new_sheet.append([row_values[0]] + row_values[2:])

            for col in new_sheet.iter_cols(min_col=3, max_col=24):
                for cell in col:
                    cell.value = None

    new_file_path = os.path.splitext(file_path)[0] + "_Kfitting.xlsx"
    wb_new.save(new_file_path)
    open_xlsx_file(window, new_file_path)


def open_avg_file_direct(window, avg_file_path):

    # Get the basename without extension
    raw_sheet_name = os.path.basename(avg_file_path).split('.')[0]

    # Normalize the sheet name to follow KherveFitting conventions
    sheet_name = normalize_sheet_name(raw_sheet_name)

    photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)
    be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

    df = pd.DataFrame({
        'BE': be_values,
        'Intensity': y_values[:num_points]
    })

    output_path = avg_file_path.rsplit('.', 1)[0] + '.xlsx'

    # Extract metadata
    metadata = extract_metadata_from_avg(avg_file_path)

    # Fields in the order shown in the example
    field_order = [
        'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
        'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
        'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
        'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
        'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
        'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
        'Collection Time', 'Time Correction', 'Y Unit'
    ]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Use the normalized sheet name instead of the raw file name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add metadata to the main sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Add experimental description at column 50
        exp_col = 50
        worksheet.cell(row=1, column=exp_col, value="Experimental Description")

        for i, field in enumerate(field_order, start=2):
            worksheet.cell(row=i, column=exp_col, value=field)
            worksheet.cell(row=i, column=exp_col + 1, value=metadata.get(field, "N/A"))

        # Set column widths
        from openpyxl.utils import get_column_letter
        worksheet.column_dimensions[get_column_letter(exp_col)].width = 25
        worksheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        # Create separate experimental description sheet
        exp_sheet = workbook.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        exp_sheet['A1'] = "Experimental Description"

        for row, field in enumerate(field_order, start=2):
            exp_sheet[f'A{row}'] = field
            exp_sheet[f'B{row}'] = metadata.get(field, "N/A")

    return output_path

def import_avantage_file(window):
    # Opens a file dialog for user to select file
    with wx.FileDialog(window, "Open Avantage Excel file", wildcard="Excel files (*.xlsx;*.xls)|*.xlsx;*.xls",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

        # Call appropriate function based on extension
        if file_path.lower().endswith('.xlsx'):
            import_avantage_file_direct(window, file_path)
        elif file_path.lower().endswith('.xls'):
            import_avantage_file_direct_xls(window, file_path)


def parse_avg_file(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    photon_energy = float(re.search(r'DS_SOPROPID_ENERGY\s+:\s+VT_R4\s+=\s+(\d+\.\d+)', content).group(1))
    start_energy, width, num_points = map(float, re.search(r'\$SPACEAXES=1\s+0=\s+(\d+\.\d+),\s+(\d+\.\d+),\s+(\d+),',
                                                           content).groups())

    # Extract all intensity values using a different approach
    y_values = []
    in_list_section = False
    list_counter = 0

    for line in content.split('\n'):
        if line.strip().startswith('LIST@'):
            in_list_section = True
            # Get values from the first line
            values_part = line.split('=', 1)[1].strip()
            # Process values from this line
            for val in values_part.split(','):
                if val.strip():
                    try:
                        y_values.append(float(val.strip()))
                    except ValueError:
                        pass
        elif in_list_section:
            # Check if this line likely contains data values
            if ',' in line and not line.strip().startswith('$') and not line.strip().startswith('DS_'):
                # Process values from continuation lines
                for val in line.split(','):
                    if val.strip():
                        try:
                            y_values.append(float(val.strip()))
                        except ValueError:
                            pass
            else:
                # If we hit a line that doesn't look like data, we're probably out of the LIST section
                in_list_section = False

    # Ensure we have the correct number of points
    if len(y_values) > int(num_points):
        y_values = y_values[:int(num_points)]

    # If we still don't have enough points, print warning
    if len(y_values) < int(num_points):
        print(f"Warning: Expected {int(num_points)} points but found only {len(y_values)}.")
        # This is serious - don't silently pad with zeros as it would give misleading data
        # Instead, let's just return what we have and make it known there's an issue

    return photon_energy, start_energy, width, int(num_points), y_values


def create_excel_from_avg(avg_file_path):
    from openpyxl.utils import get_column_letter

    # Extract the raw sheet name from the file name
    raw_sheet_name = os.path.basename(avg_file_path).split()[0]

    # Normalize the sheet name to follow KherveFitting conventions
    sheet_name = normalize_sheet_name(raw_sheet_name)

    photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)

    be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

    df = pd.DataFrame({
        'BE': be_values,
        'Intensity': y_values[:num_points]
    })

    output_path = avg_file_path.rsplit('.', 1)[0] + '.xlsx'

    # Extract metadata
    metadata = extract_metadata_from_avg(avg_file_path)

    # Fields in the order shown in the example
    field_order = [
        'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
        'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
        'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
        'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
        'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
        'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
        'Collection Time', 'Time Correction', 'Y Unit'
    ]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Use normalized sheet name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add metadata to the main sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Add experimental description at column 50
        exp_col = 50
        worksheet.cell(row=1, column=exp_col, value="Experimental Description")

        for i, field in enumerate(field_order, start=2):
            worksheet.cell(row=i, column=exp_col, value=field)
            worksheet.cell(row=i, column=exp_col + 1, value=metadata.get(field, "N/A"))

        # Set column widths
        worksheet.column_dimensions[get_column_letter(exp_col)].width = 25
        worksheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        # Create separate experimental description sheet
        exp_sheet = workbook.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        exp_sheet['A1'] = "Experimental Description"

        for row, field in enumerate(field_order, start=2):
            exp_sheet[f'A{row}'] = field
            exp_sheet[f'B{row}'] = metadata.get(field, "N/A")

    return output_path


def extract_metadata_from_avg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    metadata = {
        'Sample ID': "Group",
        'Date': "N/A",
        'Time': "N/A",
        'Technique': "XPS",
        'Species & Transition': "C 1s",
        'Number of scans': "20",
        'Source Label': "KA1486",
        'Source Energy': "1486.68",
        'Source width X': "1E+37",
        'Source width Y': "1E+37",
        'Pass Energy': "20",
        'Work Function': "1E+37",
        'Analyzer Mode': "FAT",
        'Sputtering Energy': "N/A",
        'Take-off Polar Angle': "1E+37",
        'Take-off Azimuth': "1E+37",
        'Target Bias': "1E+37",
        'Analysis Width X': "1E+37",
        'Analysis Width Y': "1E+37",
        'X Label': "Kinetic Energy",
        'X Units': "eV",
        'X Start': "1188.6",
        'X Step': "0.1",
        'Num Y Values': "382",
        'Num Scans': "20",
        'Collection Time': "0.05",
        'Time Correction': "1E+37",
        'Y Unit': "d"
    }

    # Extract values from file where available
    created_match = re.search(r'DS_EXT_SUPROPID_CREATED\s+:\s+VT_DATE\s+=\s+(\d+)/(\d+)/(\d+)\s+(\d+):(\d+):(\d+)',
                              content)
    if created_match:
        day, month, year = created_match.group(1), created_match.group(2), created_match.group(3)
        hour, minute, second = created_match.group(4), created_match.group(5), created_match.group(6)
        metadata['Date'] = f"{year}/{month}/{day}"
        metadata['Time'] = f"{hour}:{minute}:{second}"

    # Update other metadata from file where patterns match
    # Patterns already exist in the parse_avg_file function

    return metadata


def open_avg_file(window):
    with wx.FileDialog(window, "Open AVG file", wildcard="AVG files (*.avg)|*.avg",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        avg_file_path = fileDialog.GetPath()

    try:
        # Use create_excel_from_avg which we need to modify to normalize sheet names
        excel_file_path = create_excel_from_avg(avg_file_path)

        from libraries.Open import open_xlsx_file
        open_xlsx_file(window, excel_file_path)
    except Exception as e:
        wx.MessageBox(f"Error processing AVG file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def import_multiple_avg_files(window):
    with wx.DirDialog(window, "Choose a directory containing AVG files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        root_folder_path = dirDialog.GetPath()

    try:
        # Function to process a single folder
        def process_folder(folder_path):
            folder_name = os.path.basename(folder_path)
            avg_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.avg')]

            if not avg_files:
                return None  # No AVG files in this folder

            # Create Excel file for this folder
            excel_file_path = os.path.join(folder_path, f"{folder_name}.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Process each AVG file in this folder
            for avg_file in avg_files:
                avg_file_path = os.path.join(folder_path, avg_file)

                # Extract raw sheet name from file name (without extension)
                raw_sheet_name = os.path.splitext(avg_file)[0]

                # Normalize the sheet name according to KherveFitting nomenclature
                sheet_name = normalize_sheet_name(raw_sheet_name)

                # Handle duplicate sheet names if needed
                suffix = 1
                original_sheet_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_sheet_name}{suffix}"
                    suffix += 1

                # Parse AVG file
                try:
                    photon_energy, start_energy, width, expected_points, y_values = parse_avg_file(avg_file_path)

                    # Use the actual number of points we retrieved, not the expected number
                    actual_points = len(y_values)

                    # If we got fewer points than expected, adjust the width of the energy range
                    if actual_points < expected_points:
                        print(f"Warning: {avg_file} - Expected {expected_points} points but found {actual_points}.")
                        # Generate BE values based on the actual number of points we have
                        be_values = [photon_energy - (start_energy + i * width) for i in range(actual_points)]
                    else:
                        be_values = [photon_energy - (start_energy + i * width) for i in range(actual_points)]

                    # Create new sheet
                    ws = wb.create_sheet(title=sheet_name)

                    # Add headers
                    ws.append(["BE", "Intensity"])

                    # Add data - only use the points we actually have
                    for be, intensity in zip(be_values, y_values):
                        ws.append([be, intensity])

                    # Extract metadata if available
                    metadata = extract_metadata_from_avg(avg_file_path)

                    # Add metadata starting at column 50
                    exp_col = 50
                    ws.cell(row=1, column=exp_col, value="Experimental Description")

                    for i, (key, value) in enumerate(metadata.items(), start=2):
                        ws.cell(row=i, column=exp_col, value=key)
                        ws.cell(row=i, column=exp_col + 1, value=value)

                except Exception as e:
                    print(f"Error processing {avg_file}: {str(e)}")
                    continue

            # Save the Excel file if any sheets were created
            if len(wb.sheetnames) > 0:
                wb.save(excel_file_path)
                return excel_file_path
            else:
                return None

        # Process the root folder and all subfolders
        excel_files = []

        # Walk through directory tree
        for dirpath, dirnames, filenames in os.walk(root_folder_path):
            # Process current folder
            excel_path = process_folder(dirpath)
            if excel_path:
                excel_files.append(excel_path)

        # Show results
        if excel_files:
            message = f"Created {len(excel_files)} Excel files:\n" + "\n".join(excel_files)
            wx.MessageBox(message, "Success", wx.OK | wx.ICON_INFORMATION)

            # Open the first Excel file
            from libraries.Open import open_xlsx_file
            open_xlsx_file(window, excel_files[0])
        else:
            wx.MessageBox("No AVG files found in the selected folder or subfolders.",
                          "Information", wx.OK | wx.ICON_INFORMATION)

    except Exception as e:
        wx.MessageBox(f"Error processing AVG files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def open_xlsx_file_OLD(window, file_path=None):
    """
    Opens an Excel file, loads its data, and updates the application's state accordingly.
    If a corresponding JSON file exists, it loads data from there instead.
    """
    print("Starting open_xlsx_file function")
    if file_path is None:
        with wx.FileDialog(window, "Open XLSX file", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
            else:
                return

    # Store reference to file manager if open
    file_manager_was_open = False
    file_manager_position = None
    try:
        if hasattr(window, 'file_manager') and window.file_manager is not None and window.file_manager.IsShown():
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
    except RuntimeError:
        # The file_manager window was deleted but the reference still exists
        window.file_manager = None
        file_manager_was_open = False
        file_manager_position = None

    window.SetStatusText(f"Selected File: {file_path}", 0)

    try:
        # Clear undo and redo history
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        # Clear the results grid
        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Look for corresponding .json file
        json_file = os.path.splitext(file_path)[0] + '.json'
        json_data_loaded = False
        if os.path.exists(json_file):
            print(f"Found corresponding .json file: {json_file}")
            with open(json_file, 'r') as f:
                loaded_data = json.load(f)

            # Convert data structure without changing types
            window.Data = convert_from_serializable(loaded_data)
            json_data_loaded = True
            print("Loaded data from .json file")

            # Populate the results grid
            populate_results_grid(window)
        else:
            print("No corresponding .json file found. Initializing new data.")
            # Initialize the measurement data
            window.Data = Init_Measurement_Data(window)

        # Read the Excel file
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        # Check for invalid sheet names before proceeding
        invalid_sheets = [name for name in sheet_names if name.startswith('Sheet')]
        if invalid_sheets:
            wx.MessageBox(
                f"File contains invalid sheet names: {', '.join(invalid_sheets)}\n\nAll sheets must be named after "
                f"their core level (e.g., C1s, O1s) using one word in proper format without spaces.",
                "Invalid Sheet Names", wx.OK | wx.ICON_WARNING)
            return

        # Check first row values in each sheet
        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            col1_value = str(df.iloc[0, 0]).strip().upper()
            col2_value = str(df.iloc[0, 1]).strip().upper()

            # Check if it's XPS data
            xps_valid = ('BE' in col1_value or 'BINDING' in col1_value) and \
                        ('RAW DATA' in col2_value or 'CORRECTED DATA' in col2_value or 'INTENSITY' in col2_value)

            # Check if it's Raman data
            raman_valid = ('WAVENUMBER' in col1_value or 'CM-1' in col1_value) and \
                          ('RAW DATA' in col2_value or 'INTENSITY' in col2_value)

            if not (xps_valid or raman_valid):
                wx.MessageBox(
                    f"Sheet '{sheet_name}' has invalid column labels in row 1.\n"
                    f"Column A should contain 'BE'/'Binding Energy' (for XPS) or 'Wavenumber'/'Wavenumber (cm-1)' (for Raman)\n"
                    f"Column B should contain 'Raw Data', 'Corrected Data' or 'Intensity'",
                    "Invalid Column Labels", wx.OK | wx.ICON_WARNING)
                return

        results_table_index = -1
        for i, name in enumerate(all_sheet_names):
            if name.lower() == "results table":
                results_table_index = i
                break

        if results_table_index != -1:
            sheet_names = sheet_names[:results_table_index]

        print(f"Number of sheets: {len(sheet_names)}")

        # Update file path
        window.Data['FilePath'] = file_path

        # Create progress dialog
        max_progress = len(sheet_names) + 4  # +4 for initialization, processing, BE correction, final setup
        progress_dlg = wx.ProgressDialog(
            "Loading Excel File",
            "Initializing...",
            maximum=max_progress,
            parent=window,
            style=wx.PD_APP_MODAL | wx.PD_AUTO_HIDE | wx.PD_CAN_ABORT
        )

        try:
            progress_count = 0

            # Update for initialization
            if not progress_dlg.Update(progress_count, "Initializing data structure..."):
                return
            wx.GetApp().Yield()
            progress_count += 1

            # If we didn't load from json, populate the data from Excel
            if not json_data_loaded and ('Core levels' not in window.Data or not window.Data['Core levels']):
                window.Data['Number of Core levels'] = 0
                for sheet_name in sheet_names:
                    # Check if user cancelled
                    if not progress_dlg.Update(progress_count, f"Loading sheet: {sheet_name}"):
                        return
                    wx.GetApp().Yield()

                    window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)
                    progress_count += 1
            else:
                # Skip sheet loading but update progress
                progress_count += len(sheet_names)
                if not progress_dlg.Update(progress_count, "Data loaded from JSON file..."):
                    return
                wx.GetApp().Yield()

            print(f"Final number of core levels: {window.Data['Number of Core levels']}")

            # Update for BE correction
            if not progress_dlg.Update(progress_count, "Loading BE corrections..."):
                return
            wx.GetApp().Yield()
            window.load_be_correction()
            progress_count += 1

            # Update for UI setup
            if not progress_dlg.Update(progress_count, "Setting up interface..."):
                return
            wx.GetApp().Yield()

            # Update sheet names in the combobox
            window.sheet_combobox.Clear()
            window.sheet_combobox.AppendItems(sheet_names)

            # Set the first sheet as the selected one
            first_sheet = sheet_names[0]
            window.sheet_combobox.SetValue(first_sheet)

            # Use on_sheet_selected to update peak parameter grid and plot
            event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
            event.SetString(first_sheet)
            window.plot_config.plot_limits.clear()
            on_sheet_selected(window, event)
            progress_count += 1

            # Final setup
            if not progress_dlg.Update(progress_count, "Finalizing..."):
                return
            wx.GetApp().Yield()

            # undo and redo
            save_state(window)

            # Update recent files list
            update_recent_files(window, file_path)

            if hasattr(window, 'setup_backup_timer'):
                window.setup_backup_timer()
                print("Auto backup timer checked/updated after file loaded")

            progress_dlg.Update(max_progress, "Complete!")
            wx.GetApp().Yield()

        finally:
            progress_dlg.Destroy()

        # Refresh any open FileManager windows
        for top_window in wx.GetTopLevelWindows():
            if hasattr(top_window, '__class__') and top_window.__class__.__name__ == 'FileManagerWindow':
                # Force a complete refresh by getting new core levels
                top_window.core_levels = top_window.get_unique_core_levels()
                max_row_index = top_window.get_max_core_level_row_index()
                num_rows = max(10, max_row_index + 1)

                # Clear the grid completely
                if top_window.grid.GetNumberRows() > 0:
                    top_window.grid.DeleteRows(0, top_window.grid.GetNumberRows())
                if top_window.grid.GetNumberCols() > 0:
                    top_window.grid.DeleteCols(0, top_window.grid.GetNumberCols())

                # Add new columns and rows
                num_levels = len(top_window.core_levels)
                top_window.grid.AppendCols(num_levels)
                top_window.grid.AppendRows(num_rows)

                # Set column labels
                for i, level in enumerate(top_window.core_levels):
                    top_window.grid.SetColLabelValue(i, level)

                # Set row labels
                for i in range(num_rows):
                    top_window.grid.SetRowLabelValue(i, str(i))

                # Set column width and row height
                default_col_width = 50
                default_row_height = 20

                for i in range(num_levels):
                    top_window.grid.SetColSize(i, default_col_width)
                for i in range(num_rows):
                    top_window.grid.SetRowSize(i, default_row_height)

                # Set cell alignment
                for row in range(num_rows):
                    for col in range(num_levels):
                        top_window.grid.SetCellAlignment(row, col, wx.ALIGN_CENTER, wx.ALIGN_CENTER)

                # Now populate the grid with new data
                top_window.populate_grid()

        # Create backup before opening file
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(window)

        # Refresh Sheet
        from libraries.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected)

        # Reopen file manager if it was open
        if file_manager_was_open:
            from libraries.FileManager import FileManagerWindow
            window.file_manager = FileManagerWindow(window)
            if file_manager_position:
                window.file_manager.SetPosition(file_manager_position)
            window.file_manager.Show()

        print("open_xlsx_file function completed successfully")
    except Exception as e:
        print(f"Error in open_xlsx_file: {str(e)}")
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error reading file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def open_xlsx_file(window, file_path=None):
    if file_path is None:
        with wx.FileDialog(window, "Open XLSX file", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
            else:
                return

    # Store file manager state
    file_manager_was_open = False
    file_manager_position = None
    try:
        if hasattr(window, 'file_manager') and window.file_manager is not None and window.file_manager.IsShown():
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
    except RuntimeError:
        window.file_manager = None

    try:
        # Create console window centered on parent
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Loading Excel File", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

        update_console("Initializing...")
        window.SetStatusText(f"Selected File: {file_path}", 0)

        # Clear history and results
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)
        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Check for JSON file
        json_file = os.path.splitext(file_path)[0] + '.json'
        json_data_loaded = False
        if os.path.exists(json_file):
            update_console("Found corresponding JSON file - loading saved data...")
            with open(json_file, 'r') as f:
                loaded_data = json.load(f)
            window.Data = convert_from_serializable(loaded_data)
            json_data_loaded = True
            populate_results_grid(window)
        else:
            update_console("Initializing new data structure...")
            window.Data = Init_Measurement_Data(window)

        # Read Excel file
        update_console("Reading Excel file structure...")
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        # Validation
        invalid_sheets = [name for name in sheet_names if name.startswith('Sheet')]
        if invalid_sheets:
            console_frame.Close()
            wx.MessageBox(f"File contains invalid sheet names: {', '.join(invalid_sheets)}", "Invalid Sheet Names",
                          wx.OK | wx.ICON_WARNING)
            return

        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            col1_value = str(df.iloc[0, 0]).strip().upper()
            col2_value = str(df.iloc[0, 1]).strip().upper()

            xps_valid = ('BE' in col1_value or 'BINDING' in col1_value) and \
                        ('RAW DATA' in col2_value or 'CORRECTED DATA' in col2_value or 'INTENSITY' in col2_value)
            raman_valid = ('WAVENUMBER' in col1_value or 'CM-1' in col1_value) and \
                          ('RAW DATA' in col2_value or 'INTENSITY' in col2_value)

            if not (xps_valid or raman_valid):
                console_frame.Close()
                wx.MessageBox(f"Sheet '{sheet_name}' has invalid column labels", "Invalid Column Labels",
                              wx.OK | wx.ICON_WARNING)
                return

        window.Data['FilePath'] = file_path
        update_console(f"Found {len(sheet_names)} sheets to process...")

        # Load core level data
        if not json_data_loaded and ('Core levels' not in window.Data or not window.Data['Core levels']):
            window.Data['Number of Core levels'] = 0
            for i, sheet_name in enumerate(sheet_names, 1):
                update_console(f"Loading sheet {i}/{len(sheet_names)}: {sheet_name}")
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)
        else:
            update_console("Data loaded from JSON file...")
            update_console("Available sheets:")
            for i, sheet_name in enumerate(sheet_names, 1):
                update_console(f"  {i}/{len(sheet_names)}: {sheet_name}")

        update_console("Loading BE corrections...")
        window.load_be_correction()

        update_console("Setting up interface...")
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        first_sheet = sheet_names[0]
        window.sheet_combobox.SetValue(first_sheet)

        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(first_sheet)
        window.plot_config.plot_limits.clear()
        on_sheet_selected(window, event)

        save_state(window)
        update_recent_files(window, file_path)

        if hasattr(window, 'setup_backup_timer'):
            window.setup_backup_timer()


        # Create backup before opening file
        from libraries.Utilities import perform_auto_backup
        update_console("Performing auto backup...")
        perform_auto_backup(window)

        # Refresh Sheet
        update_console("Refreshing sheets...")
        from libraries.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected, update_console)

        update_console("File loaded successfully!")
        wx.CallLater(500, console_frame.Close)

        # Restore file manager
        if file_manager_was_open:
            from libraries.FileManager import FileManagerWindow
            window.file_manager = FileManagerWindow(window)
            if file_manager_position:
                window.file_manager.SetPosition(file_manager_position)
            window.file_manager.Show()

    except Exception as e:
        wx.MessageBox(f"Error reading file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def convert_from_serializable(obj):
    """
    Recursively converts a serializable object (list or dict) back into its original structure.
    This is used to restore complex data structures that were serialized to JSON.
    """
    if isinstance(obj, list):
        return [convert_from_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_from_serializable(v) for k, v in obj.items()}
    else:
        return obj

def update_recent_files(window, file_path):
    """
    Updates the list of recently opened files, adding the new file to the top and removing duplicates.
    """
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    save_recent_files_to_config(window)

def update_recent_files_menu(window):
    """
    Refreshes the 'Recent Files' menu with the updated list of recently opened files.
    """
    if hasattr(window, 'recent_files_menu'):
        # Remove all existing menu items
        for item in window.recent_files_menu.GetMenuItems():
            window.recent_files_menu.DestroyItem(item)

        # Add new menu items for recent files
        for i, file_path in enumerate(window.recent_files):
            item = window.recent_files_menu.Append(wx.ID_ANY, os.path.basename(file_path))
            window.Bind(wx.EVT_MENU, lambda evt, fp=file_path: open_xlsx_file(window, fp), item)

def save_recent_files_to_config(window):
    """
    Saves the updated list of recently opened files to the application's configuration.
    """
    window.recent_files = window.recent_files[:window.max_recent_files]
    window.save_config()


def normalize_sheet_name(name):
    """Normalize core level name to standard format."""
    import re
    new_name = name

    lower_name = name.lower()
    if 'xps survey' in lower_name:
        new_name = 'Survey'
    elif 'survey' in lower_name:
        new_name = 'Survey'
    elif 'survey scan' in lower_name:
        new_name = 'Survey'
    elif 'wide' in lower_name:
        new_name = 'Wide'
    elif 'wide scan' in lower_name:
        new_name = 'Wide'
    elif 'su1s' in lower_name or '_su' in lower_name or name.lower().endswith('_su'):
        new_name = 'Survey'
    else:
        # Remove spaces between element and orbital (e.g., "C 1s" → "C1s")
        match = re.search(r'([A-Z][a-z]?)\s+(\d+[spdf])', name)
        if match:
            element, orbital = match.groups()
            new_name = f"{element}{orbital}"

        # Simplify names like "C1s Scan" to just "C1s"
        match = re.search(r'([A-Z][a-z]?\d+[spdf])', new_name)
        if match and len(new_name) > len(match.group(1)):
            new_name = match.group(1)

    # Preserve sample number suffix if it exists
    suffix_match = re.search(r'(\d+)$', name)
    if suffix_match and not re.search(r'\d+$', new_name):
        new_name = f"{new_name}{suffix_match.group(1)}"

    return new_name

def open_vamas_file_dialog(window):
    """
    Open a file dialog for selecting a VAMAS file and process it.

    This function displays a file dialog for the user to select a VAMAS file,
    then calls open_vamas_file to process the selected file.

    Args:
    window: The main application window object.
    """
    with wx.FileDialog(window, "Open VAMAS file", wildcard="VAMAS files (*.vms)|*.vms",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_vamas_file(window, file_path)



def open_vamas_file(window, file_path):
    """
    Open and process a VAMAS file, converting it to an Excel file format.
    This function reads a VAMAS file, extracts its data and metadata,
    and creates a new Excel file with multiple sheets for each data block
    and an additional sheet for experimental description.

    Args:
    window: The main application window object.
    file_path: The path to the VAMAS file to be opened.
    """
    try:
        # Clear undo and redo history
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")

        # Copy VAMAS file to current working directory
        vamas_filename = os.path.basename(file_path)
        destination_path = os.path.join(os.getcwd(), vamas_filename)
        shutil.copy2(file_path, destination_path)

        # Read VAMAS file
        vamas_data = Vamas(vamas_filename)

        # Check for Casa peak fitting information
        has_casa_fitting = check_for_casa_fitting(vamas_data)
        import_fitting = False

        if has_casa_fitting:
            dlg = wx.MessageDialog(window,
                                   "Peak fitting information detected in VAMAS file.\n\nDo you want to import the "
                                   "peak fitting data?\n"
                                   "Note that this feature is still in beta testing and may not work perfectly.\n",
                                   "Import Peak Fitting [Beta testing]",
                                   wx.YES_NO | wx.ICON_QUESTION)
            result = dlg.ShowModal()
            import_fitting = (result == wx.ID_YES)
            dlg.Destroy()

        # Create new Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        exp_data = []

        # Create console window centered on parent
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Processing VAMAS File", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

        num_blocks = len(vamas_data.blocks)
        update_console(f"Found {num_blocks} core levels to process...")
        update_console("Starting conversion to Excel format...")

        # Process each block
        for i, block in enumerate(vamas_data.blocks, start=1):
            if block.species_label.lower() == "wide" or block.transition_or_charge_state_label.lower() == "none":
                raw_sheet_name = block.species_label
            else:
                raw_sheet_name = f"{block.species_label}{block.transition_or_charge_state_label}"
            raw_sheet_name = raw_sheet_name.replace("/", "_")

            sheet_name = normalize_sheet_name(raw_sheet_name)

            if sheet_name in wb.sheetnames:
                count = 1
                while f"{sheet_name}{count}" in wb.sheetnames:
                    count += 1
                sheet_name = f"{sheet_name}{count}"

            update_console(f"Processing {i}/{num_blocks}: {sheet_name}")

            # Create new sheet with normalized name
            ws = wb.create_sheet(title=sheet_name)

            # Extract and process data
            num_points = block.num_y_values
            x_start = block.x_start
            x_step = block.x_step
            x_values = [x_start + i * x_step for i in range(num_points)]
            y_values = block.corresponding_variables[0].y_values
            y_unit = block.corresponding_variables[0].unit
            num_scans = block.num_scans_to_compile_block

            # # Convert counts to counts per second if necessary
            # if y_unit != "c/s":
            #     y_values = [y / num_scans for y in y_values]

            # Convert counts to counts per second if necessary
            collection_time = block.signal_collection_time
            print(f"Collection time: {collection_time}, Num scans: {num_scans}, Y unit: {y_unit}")
            print(f'Number of Scans: {num_scans}')
            num_scan=num_scans # To stop division by number of scan quickly if needed
            if y_unit != "c/s" and collection_time > 0:
                y_values = [y / (num_scan * collection_time) for y in y_values]
                print(f'Maximum Y value after conversion: {max(y_values)}')
            elif y_unit != "c/s":
                y_values = [y / num_scan for y in y_values]

            # Convert to Binding Energy if necessary
            if block.x_label.lower() in ["kinetic energy", "ke"]:
                x_values = [window.photons - x - window.workfunction for x in x_values]
                x_label = "Binding Energy"
            else:
                x_label = block.x_label

            # Write data to sheet
            ws.append([x_label, "Corrected Data", "Raw Data", "Transmission"])

            # Get transmission data if it exists
            transmission_data = None
            if hasattr(block, 'corresponding_variables') and len(block.corresponding_variables) > 1:
                transmission_data = block.corresponding_variables[1].y_values
            else:
                transmission_data = [1.0] * len(y_values)

            # Write data row by row
            for j, (x, y) in enumerate(zip(x_values, y_values)):
                trans = transmission_data[j] if j < len(transmission_data) else 1.0
                corrected_y = y / trans
                ws.append([x, corrected_y, y, trans])


            if import_fitting:
                if update_console:
                    update_console(f"Processing fitting data for {sheet_name}...")

                # Parse Casa fitting information
                casa_data = parse_casa_peak_fitting(block.block_comment, block.num_scans_to_compile_block,
                                                    window.photons, transmission_data)

                if import_fitting and casa_data:
                    print(f"DEBUG: Casa data keys: {list(casa_data.keys())}")
                    print(f"DEBUG: Background data: {casa_data.get('Background', 'No Background key')}")

                    num_peaks = len(casa_data['Peaks'])
                    if num_peaks > 0:
                        peak_names = list(casa_data['Peaks'].keys())
                        if update_console:
                            update_console(f"  Found {num_peaks} peaks: {', '.join(peak_names)}")

                    # Store fitting data in a way that will be transferred to window.Data
                    if not hasattr(wb, '_fitting_data'):
                        wb._fitting_data = {}

                    wb._fitting_data[sheet_name] = {
                        'Fitting': {
                            'Peaks': casa_data['Peaks']
                        }
                    }

                    if casa_data['Background']:
                        print(f"DEBUG: Background exists, storing it")
                        wb._fitting_data[sheet_name]['Background'] = casa_data['Background']
                        wb._fitting_data[sheet_name]['Background']['Bkg X'] = x_values

                        # Calculate corrected_y (same as done in Excel writing loop)
                        corrected_y_values = []
                        for j, y in enumerate(y_values):
                            trans = transmission_data[j] if j < len(transmission_data) else 1.0
                            corrected_y_values.append(y / trans)

                        wb._fitting_data[sheet_name]['Background']['Bkg Y'] = corrected_y_values
                        # print(f"DEBUG: Stored background data: {wb._fitting_data[sheet_name]['Background']}")
                        if update_console:
                            update_console(
                                f"  Background: {casa_data['Background'].get('Bkg Type')} from {casa_data['Background'].get('Bkg Low'):.1f} to {casa_data['Background'].get('Bkg High'):.1f} eV")
                    else:
                        print(f"DEBUG: No background data in casa_data or background is empty")
                        print(f"DEBUG: casa_data['Background'] = {casa_data.get('Background', 'KEY NOT FOUND')}")
                else:
                    if update_console:
                        update_console(f"  No Casa fitting data found for {sheet_name}")

            # Store experimental setup data
            block_exp_data = [
                f"Block {i}",
                block.sample_identifier,
                f"{block.year}/{block.month}/{block.day}",
                f"{block.hour}:{block.minute}:{block.second}",
                block.technique,
                f"{block.species_label} {block.transition_or_charge_state_label}",
                block.num_scans_to_compile_block,
                block.analysis_source_label,
                block.analysis_source_characteristic_energy,
                block.analysis_source_beam_width_x,
                block.analysis_source_beam_width_y,
                block.analyzer_pass_energy_or_retard_ratio_or_mass_res,
                block.analyzer_work_function_or_acceptance_energy,
                block.analyzer_mode,
                block.sputtering_source_energy if hasattr(block, 'sputtering_source_energy') else 'N/A',
                block.analyzer_axis_take_off_polar_angle,
                block.analyzer_axis_take_off_azimuth,
                block.target_bias,
                block.analysis_width_x,
                block.analysis_width_y,
                block.x_label,
                block.x_units,
                block.x_start,
                block.x_step,
                block.num_y_values,
                block.num_scans_to_compile_block,
                block.signal_collection_time,
                block.signal_time_correction,
                y_unit,
                block.num_lines_block_comment,
                block.block_comment
            ]
            exp_data.append(block_exp_data)

            # Add experimental description data to this sheet starting at column 50
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            exp_labels = [
                "Sample ID", "Date", "Time", "Technique", "Species & Transition", "Number of scans",
                "Source Label", "Source Energy", "Source width X", "Source width Y", "Pass Energy", "Work Function",
                "Analyzer Mode", "Sputtering Energy", "Take-off Polar Angle", "Take-off Azimuth", "Target Bias",
                "Analysis Width X", "Analysis Width Y", "X Label", "X Units", "X Start", "X Step", "Num Y Values",
                "Num Scans", "Collection Time", "Time Correction", "Y Unit", "# Comment Lines", "Block Comment"
            ]

            for j, (label, value) in enumerate(zip(exp_labels, block_exp_data[1:])):
                ws.cell(row=j + 2, column=exp_col, value=label)
                ws.cell(row=j + 2, column=exp_col + 1, value=value)

            # Set column width for experimental data
            ws.column_dimensions[chr(64 + exp_col)].width = 25
            ws.column_dimensions[chr(64 + exp_col + 1)].width = 40

        update_console("Creating experimental description sheet...")

        # Create "Experimental description" sheet (keep this for backward compatibility)
        exp_sheet = wb.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 50
        exp_sheet.column_dimensions['B'].width = 100
        left_aligned = Alignment(horizontal='left')

        # Add VAMAS header information
        exp_sheet.append(["VAMAS Header Information"])
        for item in [
            ("Format Identifier", vamas_data.header.format_identifier),
            ("Institution Identifier", vamas_data.header.institution_identifier),
            ("Instrument Model", vamas_data.header.instrument_model_identifier),
            ("Operator Identifier", vamas_data.header.operator_identifier),
            ("Experiment Identifier", vamas_data.header.experiment_identifier),
            ("Number of Comment Lines", vamas_data.header.num_lines_comment),
            ("Comment", vamas_data.header.comment),
            ("Experiment Mode", vamas_data.header.experiment_mode),
            ("Scan Mode", vamas_data.header.scan_mode),
            ("Number of Spectral Regions", vamas_data.header.num_spectral_regions),
            ("Number of Analysis Positions", vamas_data.header.num_analysis_positions),
            ("Number of Discrete X Coordinates", vamas_data.header.num_discrete_x_coords_in_full_map),
            ("Number of Discrete Y Coordinates", vamas_data.header.num_discrete_y_coords_in_full_map)
        ]:
            exp_sheet.append(item)

        exp_sheet.append([])  # Add a blank row for separation

        # Define the order of block information
        block_info_order = [
            "Sample ID", "Year/Month/Day", "Time HH,MM,SS", "Technique", "Species & Transition", "Number of scans",
            "Source Label", "Source Energy", "Source width X", "Source width Y", "Pass Energy", "Work Function",
            "Analyzer Mode", "Sputtering Energy", "Take-off Polar Angle", "Take-off Azimuth", "Target Bias",
            "Analysis Width X", "Analysis Width Y", "X Label", "X Units", "X Start", "X Step", "Num Y Values",
            "Num Scans", "Collection Time", "Time Correction", "Y Unit", "# Comment Lines", "Block Comment"
        ]

        # Add block information
        for i, block_data in enumerate(exp_data, start=1):
            exp_sheet.append([f"Block {i}", ""])
            for j, info in enumerate(block_info_order):
                exp_sheet.append([info, block_data[j + 1]])
            exp_sheet.append([])  # Add a blank row between blocks

        # Set alignment for all cells in column B
        for row in exp_sheet.iter_rows(min_row=1, max_row=exp_sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = left_aligned

        update_console("Saving Excel file...")
        excel_filename = os.path.splitext(vamas_filename)[0] + ".xlsx"
        excel_path = os.path.join(os.path.dirname(file_path), excel_filename)
        wb.save(excel_path)

        # Save fitting data to JSON file immediately after Excel file is saved
        if import_fitting and hasattr(wb, '_fitting_data'):
            print(f"VAMAS: Saving fitting data for {len(wb._fitting_data)} sheets")
            import json
            fitting_file = excel_path.replace('.xlsx', '_fitting.json')
            with open(fitting_file, 'w') as f:
                json.dump(wb._fitting_data, f)
            print(f"VAMAS: Fitting data saved to: {fitting_file}")

        update_console("Excel file created successfully!")
        update_console("Loading Excel file into KherveFitting...")

        os.remove(destination_path)
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        # Pass console to next function
        open_xlsx_file_vamas(window, excel_path, console_frame, update_console, import_fitting)

    except Exception as e:
        wx.MessageBox(f"Error processing VAMAS file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def parse_casa_peak_fitting(block_comment, num_scans=1, photon_energy=1486.67, transmission_data=None):
    """
    Parse Casa XPS peak fitting information from VAMAS block comment.

    Args:
        block_comment (str): The block comment containing Casa Info
        num_scans (int): Number of scans to divide area by
        photon_energy (float): X-ray photon energy for KE to BE conversion
        transmission_data (list): Transmission values for area correction

    Returns:
        dict: Dictionary containing background and peak fitting information
    """
    if "Casa Info Follows" not in block_comment:
        return None

    # Calculate average transmission if available
    avg_transmission = 1.0
    if transmission_data and len(transmission_data) > 0:
        avg_transmission = np.mean(transmission_data)

    lines = block_comment.split('\n')
    casa_start = -1

    # Find the start of Casa Info
    for i, line in enumerate(lines):
        if "Casa Info Follows" in line:
            casa_start = i + 1
            break

    if casa_start == -1:
        return None

    fitting_data = {
        'Background': {},
        'Peaks': {}
    }

    # Parse the lines after "Casa Info Follows"
    i = casa_start
    while i < len(lines):
        line = lines[i].strip()

        if line.startswith('CASA region'):
            # Parse background information
            parts = line.split()
            if len(parts) >= 5:
                # Extract background type - look for second (*...*) pattern
                bg_matches = re.findall(r'\(\*([^*]+)\*\)', line)
                bg_type = bg_matches[1] if len(bg_matches) > 1 else "Shirley"

                # Map Casa background types to KherveFitting types
                if bg_type.lower() == "shirley":
                    kf_bg_type = "Shirley"
                else:
                    kf_bg_type = bg_type

                # Extract energy range and convert from KE to BE
                try:
                    low_ke = float(parts[5])
                    high_ke = float(parts[6])
                    # Convert KE to BE
                    low_energy = photon_energy - high_ke
                    high_energy = photon_energy - low_ke
                    print(f"DEBUG: Parsed background - Type: {kf_bg_type}, Low: {low_energy}, High: {high_energy}")

                    fitting_data['Background'] = {
                        'Bkg Type': kf_bg_type,
                        'Bkg Low': low_energy,
                        'Bkg High': high_energy,
                        'Bkg Offset Low': '0',
                        'Bkg Offset High': '0',
                        'Bkg Y': []
                    }
                except (ValueError, IndexError):
                    pass

        elif line.startswith('CASA comp'):
            # Parse peak information
            # Extract peak name (first pattern)
            peak_matches = re.findall(r'\(\*([^*]+)\*\)', line)
            peak_name = peak_matches[0] if peak_matches else f"Peak_{len(fitting_data['Peaks']) + 1}"

            # Extract model (second pattern)
            model_str = peak_matches[1] if len(peak_matches) > 1 else "GL(30)"

            # Parse model and determine KherveFitting equivalent
            model = "GL (Area)"
            lg_value = 30
            sigma_value = 0.6
            gamma_value = 0.4

            if 'SGL(' in model_str:
                lg_match = re.search(r'SGL\((\d+)\)', model_str)
                lg_value = float(lg_match.group(1)) if lg_match else 30
                model = "SGL (Area)"
            elif 'GL(' in model_str:
                lg_match = re.search(r'GL\((\d+)\)', model_str)
                lg_value = float(lg_match.group(1)) if lg_match else 30
                model = "GL (Area)"
            elif 'LA(' in model_str:
                # Extract LA parameters: LA(sigma, gamma, w_g)
                la_params = re.search(r'LA\(([\d.]+),\s*([\d.]+),\s*([\d.]+)\)', model_str)
                if la_params:
                    sigma_value = float(la_params.group(1))
                    gamma_value = float(la_params.group(2))
                    w_g = float(la_params.group(3))

                    if w_g > 100:
                        model = "LA*G (Area, σ/γ, γ)"
                    else:
                        model = "LA (Area, σ, γ)"

                    # Calculate L/G ratio from sigma and gamma
                    lg_value = 100 * sigma_value / (sigma_value + gamma_value)
                else:
                    # Check for single parameter LA(number) format
                    single_param = re.search(r'LA\((\d+)\)', model_str)
                    if single_param:
                        ratio_value = int(single_param.group(1))

                        # Mapping for gamma values based on sigma/gamma ratio
                        gamma_mapping = {
                            20: 2.7, 30: 2.4, 40: 2.2, 50: 2.0, 60: 1.8,
                            70: 1.6, 80: 1.4, 90: 1.2, 100: 1.0
                        }

                        gamma_value = gamma_mapping.get(ratio_value, 2.0)  # Default to 2.0 if not found
                        sigma_value = gamma_value  # sigma = (sigma/gamma) * gamma
                        lg_value = ratio_value  # L/G ratio is the number itself
                        model = "LA (Area, σ/γ, γ)"
                    else:
                        model = "LA (Area, σ, γ)"
                        sigma_value = 0.6
                        gamma_value = 0.4
                        lg_value = 100 * sigma_value / (sigma_value + gamma_value)

            # Extract parameters with constraints
            area_match = re.search(r'Area\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)
            fwhm_match = re.search(r'MFWHM\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)
            pos_match = re.search(r'Position\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)

            # Calculate area (divide by average transmission only)
            area_value = float(area_match.group(1)) / avg_transmission if area_match else 1000
            area_value = round(area_value,2)
            # Handle area constraints
            if area_match:
                area_constrained_peak = int(area_match.group(4))
                area_constraint_value = float(area_match.group(5))

                if area_constrained_peak >= 0:
                    peak_letter = chr(65 + area_constrained_peak)  # A=0, B=1, etc.
                    area_constraint = f"{peak_letter}*{area_constraint_value}"
                else:
                    area_min = float(area_match.group(2)) if float(area_match.group(2)) > 1e-19 else 1
                    area_max = float(area_match.group(3)) / avg_transmission
                    area_constraint = f"{area_min}:{area_max}"
            else:
                area_constraint = "1:1e7"

            # Convert position from KE to BE
            pos_ke = float(pos_match.group(1)) if pos_match else 800
            position_be = photon_energy - pos_ke

            # Handle position constraints
            if pos_match:
                pos_constrained_peak = int(pos_match.group(4))
                pos_constraint_value = float(pos_match.group(5))

                if pos_constrained_peak >= 0:
                    peak_letter = chr(65 + pos_constrained_peak)
                    pos_constraint = f"{peak_letter}+0.0"  # Position typically uses +
                else:
                    pos_ke_min = float(pos_match.group(2))
                    pos_ke_max = float(pos_match.group(3))
                    pos_be_min = photon_energy - pos_ke_max  # Inverted
                    pos_be_max = photon_energy - pos_ke_min  # Inverted
                    pos_constraint = f"{pos_be_min:.2f}:{pos_be_max:.2f}"
            else:
                pos_constraint = "1:1000"

            # Handle FWHM constraints
            fwhm_value = float(fwhm_match.group(1)) if fwhm_match else 1.5
            fwhm_value = round(fwhm_value,2)
            if fwhm_match:
                fwhm_constrained_peak = int(fwhm_match.group(4))
                fwhm_constraint_value = float(fwhm_match.group(5))

                if fwhm_constrained_peak >= 0:
                    peak_letter = chr(65 + fwhm_constrained_peak)
                    fwhm_constraint = f"{peak_letter}*{fwhm_constraint_value}"
                else:
                    fwhm_constraint = f"{fwhm_match.group(2)}:{fwhm_match.group(3)}"
            else:
                fwhm_constraint = "0.3:3.5"

            # Calculate height from area and FWHM
            if 'LA' in model:
                height = area_value / (fwhm_value * 1.5)  # Approximation
            else:
                height = area_value / (fwhm_value * np.sqrt(np.pi / (4 * np.log(2))))

            peak_data = {
                'Position': position_be,
                'Height': round(height,2),
                'FWHM': round(fwhm_value,2),
                'L/G': lg_value,
                'Area': round(area_value,2),
                'Sigma': sigma_value,
                'Gamma': gamma_value,
                'Skew': 0.1,
                'Fitting Model': model,
                'Bkg Type': fitting_data['Background'].get('Bkg Type', ''),
                'Bkg Low': fitting_data['Background'].get('Bkg Low', '0'),
                'Bkg High': fitting_data['Background'].get('Bkg High', '0'),
                'Bkg Offset Low': '0',
                'Bkg Offset High': '0'
            }

            # Set constraints
            if 'LA' in model:
                lg_constraint = "Fixed"
            else:
                lg_constraint = "Fixed" if any(x in model_str for x in ['GL(', 'SGL(']) else "5:80"

            constraints = {
                'Position': pos_constraint,
                'Height': "1:1e7",
                'FWHM': fwhm_constraint,
                'L/G': lg_constraint,
                'Area': area_constraint,
                'Sigma': "0.01:10" if 'LA' in model else "0.3:3",
                'Gamma': "0.01:10" if 'LA' in model else "0.3:3",
                'Skew': "0.01:2"
            }

            peak_data['Constraints'] = constraints
            fitting_data['Peaks'][peak_name] = peak_data

        i += 1

    return fitting_data if (fitting_data['Background'] or fitting_data['Peaks']) else None


def check_for_casa_fitting(vamas_data):
    """
    Check if any blocks contain Casa peak fitting information.

    Args:
        vamas_data: VAMAS data object

    Returns:
        bool: True if Casa fitting info found
    """
    for block in vamas_data.blocks:
        if hasattr(block, 'block_comment') and "Casa Info Follows" in block.block_comment:
            return True
    return False

def open_xlsx_file_vamas_OLD(window, file_path, console_frame=None, update_console=None):
    """
    Open and process an Excel file created from a VAMAS file.

    This function initializes the data structure, reads the Excel file,
    populates the window.Data dictionary with core level information,
    updates the GUI elements, and plots the data for the first sheet.

    Args:
    window: The main application window object.
    file_path: The path to the Excel file to be opened.
    """
    try:
        if update_console:
            update_console("Reading Excel file structure...")

        window.SetStatusText(f"Selected File: {file_path}", 0)
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = file_path

        excel_file = pd.ExcelFile(file_path)
        sheet_names = [name for name in excel_file.sheet_names if name != "Experimental description"]
        window.Data['Number of Core levels'] = 0

        if update_console:
            update_console(f"Found {len(sheet_names)} sheets to load...")

        for i, sheet_name in enumerate(sheet_names, 1):
            if update_console:
                update_console(f"Loading sheet {i}/{len(sheet_names)}: {sheet_name}")
            window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        if update_console:
            update_console("Updating interface...")

        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        window.sheet_combobox.SetValue(sheet_names[0])
        window.plot_manager.plot_data(window)

        if update_console:
            update_console("Loading complete!")
            wx.CallLater(500, console_frame.Close)

    except Exception as e:
        wx.MessageBox(f"Error reading Excel file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def open_xlsx_file_vamas(window, file_path, console_frame=None, update_console=None, has_fitting=False):
    """
    Open and process an Excel file created from a VAMAS file.

    This function initializes the data structure, reads the Excel file,
    populates the window.Data dictionary with core level information,
    updates the GUI elements, and plots the data for the first sheet.

    Args:
    window: The main application window object.
    file_path: The path to the Excel file to be opened.
    console_frame: Console window for progress updates.
    update_console: Function to update console output.
    has_fitting: Whether fitting data should be loaded.
    """
    try:
        if update_console:
            update_console("Reading Excel file structure...")

        window.SetStatusText(f"Selected File: {file_path}", 0)
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = file_path

        # Load fitting data if available
        fitting_data = {}
        if has_fitting:
            fitting_file = file_path.replace('.xlsx', '_fitting.json')
            if os.path.exists(fitting_file):
                if update_console:
                    update_console("Loading peak fitting data...")

                with open(fitting_file, 'r') as f:
                    fitting_data = json.load(f)

                # Update console with peak information
                for sheet_name, data in fitting_data.items():
                    num_peaks = len(data.get('Fitting', {}).get('Peaks', {}))
                    if num_peaks > 0:
                        peak_names = list(data.get('Fitting', {}).get('Peaks', {}).keys())
                        if update_console:
                            update_console(f"  {sheet_name}: {num_peaks} peaks - {', '.join(peak_names)}")

                # Clean up temporary file
                os.remove(fitting_file)

        excel_file = pd.ExcelFile(file_path)
        sheet_names = [name for name in excel_file.sheet_names if name != "Experimental description"]
        window.Data['Number of Core levels'] = 0

        if update_console:
            update_console(f"Found {len(sheet_names)} sheets to load...")

        for i, sheet_name in enumerate(sheet_names, 1):
            if update_console:
                update_console(f"Loading sheet {i}/{len(sheet_names)}: {sheet_name}")

            window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

            # Add fitting data if available for this sheet
            if fitting_data and sheet_name in fitting_data:
                if 'Core levels' in window.Data and sheet_name in window.Data['Core levels']:
                    # Add the fitting data to the core level
                    window.Data['Core levels'][sheet_name].update(fitting_data[sheet_name])

        if update_console:
            update_console("Updating interface...")

        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        window.sheet_combobox.SetValue(sheet_names[0])

        # Refresh peak fitting grid to show imported peaks
        from libraries.Sheet_Operations import on_sheet_selected
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(sheet_names[0])
        on_sheet_selected(window, event)

        # Calculate backgrounds for sheets with fitting data
        for sheet_name in sheet_names:
            if (fitting_data and sheet_name in fitting_data and
                    'Background' in fitting_data[sheet_name]):
                window.sheet_combobox.SetValue(sheet_name)
                bg_data = fitting_data[sheet_name]['Background']
                window.bg_min_energy = bg_data.get('Bkg Low')  # Add print here
                window.bg_max_energy = bg_data.get('Bkg High')  # Add print here
                print(f"DEBUG: Setting bg range: {window.bg_min_energy} to {window.bg_max_energy}")
                window.background_method = bg_data.get('Bkg Type', 'Multi-Regions Smart')
                window.plot_manager.plot_background(window, use_smoothing=False)

        # Set back to first sheet
        window.sheet_combobox.SetValue(sheet_names[0])

        window.plot_manager.plot_data(window)

        window.plot_manager.clear_and_replot(window)

        if update_console:
            update_console("Loading complete!")
            wx.CallLater(500, console_frame.Close)

    except Exception as e:
        wx.MessageBox(f"Error reading Excel file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def extract_transmission_data(block):
    """
    Extract transmission function data from a Kratos .kal file block.

    Args:
        block (str): Text block containing transmission data

    Returns:
        tuple: (ke_trans, trans_values) arrays of kinetic energies and transmission values
    """
    if 'Transmission Function Object (ke,t)' in block:
        trans_section = block.split('Transmission Function Object (ke,t)')[1].split('Dataset filename')[0]

        ke_line = trans_section.split('Transmission Function Kinetic Energy =')[1].split('\n')[0]
        ke_str = ke_line.strip().strip('{}').strip()
        ke_trans = np.array([float(x.strip()) for x in ke_str.split(',')])

        val_line = trans_section.split('Transmission Function Value          =')[1].split('\n')[0]
        val_str = val_line.strip().strip('{}').strip()
        trans_values = np.array([float(x.strip()) for x in val_str.split(',')])

        return ke_trans, trans_values
    return None, None


def convert_kal_to_excel(file_path):
    """
    Convert Kratos .kal file to Excel format suitable for KherveFitting.

    Args:
        file_path (str): Path to the .kal file

    Returns:
        str: Path to the created Excel file
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    with open(file_path, 'r') as f:
        content = f.read()

    blocks = content.split('Dataset filename')
    spectra = {}
    PHOTON_ENERGY = 1486.67

    # Extract sample ID
    sample_id = "Unknown"
    for block in blocks:
        for line in block.split('\n'):
            if 'Stage Position Name' in line:
                sample_id = line.split('=')[1].strip()
                break
        if sample_id != "Unknown":
            break

    # Process each spectrum block
    for block in blocks:
        if 'Ordinate values' not in block or 'Object name' not in block:
            continue

        # Get spectrum name and number
        object_name_line = next((line for line in block.split('\n') if 'Object name' in line), None)
        if not object_name_line:
            continue

        name_parts = object_name_line.split('=')[1].strip().split('/')
        name = name_parts[0].strip()
        spectrum_id = name_parts[1].strip() if len(name_parts) > 1 else ""

        # Skip non-spectrum blocks
        if 'Sample Position' in name or 'Counter' in name:
            continue

        # Initialize metadata dictionary with default values
        metadata = {
            'Sample ID': sample_id,
            'Date': '',
            'Time': '',
            'Technique': 'XPS',
            'Species & Transition': '',
            'Number of scans': '',
            'Source Label': 'Al mono',
            'Source Energy': '1486.6',
            'Source width X': '1E+37',
            'Source width Y': '1E+37',
            'Pass Energy': '',
            'Work Function': '1E+37',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': '',
            'X Units': '',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '',
            'Collection Time': '',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Extract metadata from block
        for line in block.split('\n'):
            line = line.strip()

            # Date and Time
            if 'Date Acquired' in line:
                parts = line.split('=')[1].strip().split()
                if len(parts) >= 2:
                    metadata['Date'] = parts[0]
                    metadata['Time'] = parts[1]

            # Species & Transition
            elif 'Chemical symbol or formula' in line:
                metadata['species'] = line.split('=')[1].strip()
            elif 'Transition or charge state' in line:
                metadata['transition'] = line.split('=')[1].strip()

            # Number of scans
            elif '# Sweeps completed' in line:
                num_scans = line.split('=')[1].strip()
                metadata['Number of scans'] = num_scans
                metadata['Num Scans'] = num_scans

            # Pass Energy
            elif 'Pass energy' in line:
                metadata['Pass Energy'] = line.split('=')[1].strip()

            # X Label, Units, Start, Step
            elif 'Abscissa label' in line:
                metadata['X Label'] = line.split('=')[1].strip()
            elif 'Abscissa units' in line:
                metadata['X Units'] = line.split('=')[1].strip()
            elif 'Spectrum scan start' in line:
                metadata['X Start'] = line.split('=')[1].split('eV')[0].strip()
            elif 'Spectrum scan step size' in line:
                metadata['X Step'] = line.split('=')[1].split('eV')[0].strip()

            # Collection Time
            elif 'Dwell time' in line:
                metadata['Collection Time'] = line.split('=')[1].replace('seconds', '').strip()

        # Combine species and transition
        if 'species' in metadata and 'transition' in metadata:
            metadata['Species & Transition'] = f"{metadata['species']} {metadata['transition']}"
            metadata.pop('species')
            metadata.pop('transition')

        # Count Y values
        if 'Ordinate values' in block:
            try:
                values_str = block.split('Ordinate values')[1].split('=')[1].split('}')[0].strip().strip('{').strip()
                values = values_str.split(',')
                metadata['Num Y Values'] = str(len(values))
            except:
                pass

        # Extract block comment
        comment_lines = []
        in_comment = False
        for line in block.split('\n'):
            if 'Casa Info Follows' in line:
                in_comment = True
                comment_lines.append(line.strip())
            elif in_comment and line.strip():
                comment_lines.append(line.strip())

        if comment_lines:
            metadata['# Comment Lines'] = str(len(comment_lines))
            metadata['Block Comment'] = '"' + '\n'.join(comment_lines) + '"'

        # Process spectral data
        ke_trans, trans_values = extract_transmission_data(block)
        if ke_trans is not None and trans_values is not None:
            # Create interpolation function
            trans_func = interp1d(ke_trans, trans_values, kind='linear', bounds_error=False,
                                  fill_value='extrapolate')

            start_ke = float(metadata['X Start'])
            step = float(metadata['X Step'])
            raw_str = block.split('Ordinate values')[1].split('=')[1].split('}')[0].strip().strip('{').strip()
            raw_data = np.array([float(x.strip()) for x in raw_str.split(',')])

            num_points = len(raw_data)
            ke_values = np.linspace(start_ke, start_ke + (num_points - 1) * step, num_points)
            be_values = PHOTON_ENERGY - ke_values

            transmission = trans_func(ke_values)
            corrected_data = raw_data / transmission

            df = pd.DataFrame({
                'BE': be_values,
                'Corrected Data': corrected_data,
                'Raw Data': raw_data,
                'Transmission': transmission
            })

            spectra[name] = {
                'data': df,
                'metadata': metadata,
                'id': spectrum_id
            }

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets for each spectrum
    for name, info in spectra.items():
        sheet_name = name.replace(' ', '')
        df = info['data']
        metadata = info['metadata']

        # Create sheet and write data
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data values
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add experimental description (column 50)
        exp_col = 50
        ws.cell(row=1, column=exp_col, value="Experimental Description")

        # Metadata fields in specified order
        fields = [
            'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
            'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
            'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
            'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
            'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
            'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
            'Collection Time', 'Time Correction', 'Y Unit', '# Comment Lines',
            'Block Comment'
        ]

        # Write metadata in order
        for i, field in enumerate(fields, 2):
            ws.cell(row=i, column=exp_col, value=field)
            ws.cell(row=i, column=exp_col + 1, value=metadata.get(field, ''))

        # Set column widths
        ws.column_dimensions[get_column_letter(exp_col)].width = 25
        ws.column_dimensions[get_column_letter(exp_col + 1)].width = 50

    # Create Experimental description sheet
    exp_sheet = wb.create_sheet("Experimental description")
    exp_sheet.column_dimensions['A'].width = 30
    exp_sheet.column_dimensions['B'].width = 100
    left_aligned = Alignment(horizontal='left')

    # Write metadata to Experimental description sheet
    row = 1
    for name, info in spectra.items():
        metadata = info['metadata']
        spectrum_id = info['id']

        # Add spectrum header
        header = exp_sheet.cell(row=row, column=1, value=f"Spectrum: {name}/{spectrum_id}")
        header.alignment = left_aligned
        row += 1

        # Write metadata fields
        for field in fields:
            cell_a = exp_sheet.cell(row=row, column=1, value=field)
            cell_b = exp_sheet.cell(row=row, column=2, value=metadata.get(field, ''))

            cell_a.alignment = left_aligned
            cell_b.alignment = left_aligned

            row += 1

        # Add space between spectra
        row += 1

    # Save Excel file
    output_file = file_path.replace('.kal', '.xlsx')
    wb.save(output_file)

    return output_file

def open_kal_file_dialog(window):
    """Open a file dialog for selecting a Kratos .kal file"""
    with wx.FileDialog(window, "Open Kratos file", wildcard="Kratos files (*.kal)|*.kal",
                      style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_kal_file(window, file_path)

def open_kal_file(window, file_path):
    """Process a Kratos .kal file and convert it to Excel format"""

    try:
        output_excel = convert_kal_to_excel(file_path)
        open_xlsx_file(window, output_excel)
    except Exception as e:
        wx.MessageBox(f"Error processing Kratos file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_raman_txt_file(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.FileDialog(window, "Open Raman text file", wildcard="Text files (*.txt)|*.txt",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = f"Raman_{base_filename}"

        # Read data from txt file
        data = []
        with open(file_path, 'r') as f:
            for line in f:
                if line.strip() and not line.startswith('#'):
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        try:
                            wavenumber = float(parts[0])
                            intensity = float(parts[1])
                            data.append([wavenumber, intensity])
                        except ValueError:
                            continue

        if not data:
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "Wavenumber (cm-1)"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (wavenumber, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = wavenumber
            ws[f"B{i}"] = intensity

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")


def import_multiple_raman_files(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing Raman text files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all txt files in the directory
        txt_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.txt')]

        if not txt_files:
            window.show_popup_message2("Information", "No .txt files found in the selected folder.")
            return

        # Create an Excel file for each group of files
        excel_path = os.path.join(dir_path, "Raman_Data.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each txt file
        for txt_file in txt_files:
            file_path = os.path.join(dir_path, txt_file)
            base_filename = os.path.splitext(txt_file)[0]
            sheet_name = f"Raman_{base_filename}"

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from txt file
            data = []
            with open(file_path, 'r') as f:
                for line in f:
                    if line.strip() and not line.startswith('#'):
                        parts = line.strip().split()
                        if len(parts) >= 2:
                            try:
                                wavenumber = float(parts[0])
                                intensity = float(parts[1])
                                data.append([wavenumber, intensity])
                            except ValueError:
                                continue

            if not data:
                continue  # Skip files with no valid data

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "Wavenumber (cm-1)"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (wavenumber, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = wavenumber
                ws[f"B{i}"] = intensity

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} Raman data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the text files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")

def import_xps_asc_file(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.FileDialog(window, "Open XPS ASC file", wildcard="ASC files (*.asc)|*.asc",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = base_filename

        # Read data from asc file
        data = []
        with open(file_path, 'r') as f:
            for line in f:
                if line.strip() and not line.startswith('#'):
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        try:
                            binding_energy = float(parts[0])
                            intensity = float(parts[1])
                            data.append([binding_energy, intensity])
                        except ValueError:
                            continue

        if not data:
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "BE"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (binding_energy, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = binding_energy
            ws[f"B{i}"] = intensity

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")


def import_multiple_xps_asc_files(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing XPS ASC files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all asc files in the directory
        asc_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.asc')]

        if not asc_files:
            window.show_popup_message2("Information", "No .asc files found in the selected folder.")
            return

        # Create an Excel file for each group of files
        excel_path = os.path.join(dir_path, "XPS_Data.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each asc file
        for asc_file in asc_files:
            file_path = os.path.join(dir_path, asc_file)
            base_filename = os.path.splitext(asc_file)[0]
            sheet_name = base_filename

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from asc file
            data = []
            with open(file_path, 'r') as f:
                for line in f:
                    if line.strip() and not line.startswith('#'):
                        parts = line.strip().split()
                        if len(parts) >= 2:
                            try:
                                binding_energy = float(parts[0])
                                intensity = float(parts[1])
                                data.append([binding_energy, intensity])
                            except ValueError:
                                continue

            if not data:
                continue  # Skip files with no valid data

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "BE"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (binding_energy, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = binding_energy
                ws[f"B{i}"] = intensity

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} XPS data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the ASC files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")

def import_xps_csv_file(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.FileDialog(window, "Open XPS CSV file", wildcard="CSV files (*.csv)|*.csv",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = base_filename

        # Read data from csv file
        df = pd.read_csv(file_path, delimiter=',', header=None)
        if df.shape[1] < 2:
            # Try other common separators if comma doesn't work
            for sep in [';', '\t', ' ']:
                test_df = pd.read_csv(file_path, delimiter=sep, header=None)
                if test_df.shape[1] >= 2:
                    df = test_df
                    break

        if df.shape[1] < 2:
            window.show_popup_message2("Error", "CSV file must contain at least two columns (BE and Intensity).")
            return

        # Extract the first two columns (assuming BE and intensity)
        data = df.iloc[:, :2].values

        if len(data) == 0:  # Use len() instead of direct boolean check
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "Binding Energy (eV)"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (binding_energy, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = float(binding_energy)
            ws[f"B{i}"] = float(intensity)

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")

def import_multiple_xps_csv_files(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing XPS CSV files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all csv files in the directory
        csv_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.csv')]

        if not csv_files:
            window.show_popup_message2("Information", "No .csv files found in the selected folder.")
            return

        # Create an Excel file for all files
        excel_path = os.path.join(dir_path, "XPS_Data_CSV.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each csv file
        for csv_file in csv_files:
            file_path = os.path.join(dir_path, csv_file)
            base_filename = os.path.splitext(csv_file)[0]
            sheet_name = base_filename

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from csv file
            try:
                df = pd.read_csv(file_path, delimiter=',', header=None)
                if df.shape[1] < 2:
                    # Try other common separators if comma doesn't work
                    for sep in [';', '\t', ' ']:
                        test_df = pd.read_csv(file_path, delimiter=sep, header=None)
                        if test_df.shape[1] >= 2:
                            df = test_df
                            break

                if df.shape[1] < 2:
                    continue  # Skip files without at least two columns

                # Extract the first two columns (assuming BE and intensity)
                data = df.iloc[:, :2].values

                if len(data) == 0:  # Use len() instead of direct boolean check
                    continue  # Skip files with no valid data
            except Exception as e:
                print(f"Error reading {csv_file}: {e}")
                continue

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy (eV)"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (binding_energy, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = float(binding_energy)
                ws[f"B{i}"] = float(intensity)

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} XPS data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the CSV files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")

def open_file_location(window):
    if 'FilePath' in window.Data:
        file_path = window.Data['FilePath']
        folder_path = os.path.dirname(file_path)
        if os.path.exists(folder_path):
            if sys.platform == 'win32':
                os.startfile(folder_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{folder_path}"')
            else:
                os.system(f'xdg-open "{folder_path}"')


def open_vg_microtech_file(window, file_path):
    """
    Process a VG-Microtech .1 file and convert it to Excel format suitable for KherveFitting.

    Args:
        window: The main application window
        file_path: Path to the .1 file

    Returns:
        bool: Success status
    """
    import numpy as np
    import openpyxl
    import os

    try:
        # Read the file content
        with open(file_path, 'r') as f:
            lines = f.readlines()

        # Parse header information
        header = lines[1].strip().split()
        print(f'Header: {header}')
        be_start = float(header[0])
        be_end = float(header[1])
        energy_step = float(header[2])
        unknown_param = float(header[3])
        dwell_time = float(header[4])
        num_points = int(header[5])
        pass_energy = float(header[6])
        photon_energy = abs(float(header[7]))  # Absolute value for negative photon energies

        # Get measurement type
        measurement_type = lines[2].strip()

        # Extract intensity values
        intensity_values = []
        for i in range(3, len(lines)):
            if lines[i].strip():
                try:
                    intensity_values.append(float(lines[i].strip()))
                except ValueError:
                    continue

        # Ensure we have the right number of intensity values
        if len(intensity_values) < num_points:
            window.show_popup_message2("Warning",
                                       f"Expected {num_points} data points but found only {len(intensity_values)}.")
        elif len(intensity_values) > num_points:
            intensity_values = intensity_values[:num_points]

        # Calculate binding energy values
        be_values = np.linspace(be_start, be_end, num_points)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheet with normalized measurement type as name
        from libraries.Open import normalize_sheet_name
        sheet_name = normalize_sheet_name(measurement_type)
        ws = wb.create_sheet(title=sheet_name)

        # Add column headers
        ws["A1"] = "Binding Energy (eV)"
        ws["B1"] = "Corrected Data"
        ws["C1"] = "Raw Data"
        ws["D1"] = "Transmission"

        # Add data rows
        for i, (be, intensity) in enumerate(zip(be_values, intensity_values), start=2):
            ws[f"A{i}"] = be
            ws[f"B{i}"] = intensity
            ws[f"C{i}"] = intensity
            ws[f"D{i}"] = 1.0  # Default transmission value

        # Create experimental description sheet
        exp_sheet = wb.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata
        exp_sheet["A1"] = "Sample ID"
        exp_sheet["B1"] = os.path.basename(file_path)
        exp_sheet["A2"] = "BE Start"
        exp_sheet["B2"] = str(be_start)
        exp_sheet["A3"] = "BE End"
        exp_sheet["B3"] = str(be_end)
        exp_sheet["A4"] = "Energy Step"
        exp_sheet["B4"] = str(energy_step)
        exp_sheet["A5"] = "Dwell Time"
        exp_sheet["B5"] = str(dwell_time)
        exp_sheet["A6"] = "Number of Points"
        exp_sheet["B6"] = str(num_points)
        exp_sheet["A7"] = "Pass Energy"
        exp_sheet["B7"] = str(pass_energy)
        exp_sheet["A8"] = "Photon Energy"
        exp_sheet["B8"] = str(photon_energy)
        exp_sheet["A9"] = "Technique"
        exp_sheet["B9"] = "XPS"
        exp_sheet["A10"] = "Species & Transition"
        exp_sheet["B10"] = measurement_type

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the created Excel file
        from libraries.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing VG-Microtech file: {str(e)}")
        return False


def open_vg_microtech_file_dialog(window):
    """
    Open a file dialog for selecting a VG-Microtech .1 file and process it.
    """
    import wx

    with wx.FileDialog(window, "Open VG-Microtech file", wildcard="VG-Microtech files (*.1)|*.1",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_vg_microtech_file(window, file_path)


def import_multiple_vg_microtech_files(window):
    """
    Import multiple VG-Microtech .1 files from a folder.
    """
    import wx
    import os
    import numpy as np
    import openpyxl
    from libraries.Open import normalize_sheet_name, open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing VG-Microtech .1 files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all .1 files in the directory
        vg_files = [f for f in os.listdir(folder_path) if f.endswith('.1')]

        if not vg_files:
            window.show_popup_message2("Information", "No .1 files found in the selected folder.")
            return

        # Ask if user wants individual Excel files or one combined file
        dlg = wx.MessageDialog(window,
                               "Do you want to create individual Excel files for each .1 file or combine them into one Excel file?",
                               "Import Options",
                               wx.YES_NO | wx.ICON_QUESTION)
        dlg.SetYesNoLabels("Individual Files", "One Combined File")

        result = dlg.ShowModal()
        individual_files = (result == wx.ID_YES)
        dlg.Destroy()

        if individual_files:
            # Process each .1 file individually
            processed_count = 0
            for vg_file in vg_files:
                vg_path = os.path.join(folder_path, vg_file)
                if open_vg_microtech_file(window, vg_path):
                    processed_count += 1

            window.show_popup_message2("Success", f"Processed {processed_count} of {len(vg_files)} VG-Microtech files.")
        else:
            # Combine all .1 files into one Excel file
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            processed_count = 0
            exp_sheet = combined_wb.create_sheet(title="Experimental description")
            exp_sheet.column_dimensions['A'].width = 30
            exp_sheet.column_dimensions['B'].width = 50

            row = 1

            for vg_file in vg_files:
                vg_path = os.path.join(folder_path, vg_file)

                try:
                    # Read the file content
                    with open(vg_path, 'r') as f:
                        lines = f.readlines()

                    # Parse header information
                    header = lines[1].strip().split()
                    be_start = float(header[0])
                    be_end = float(header[1])
                    energy_step = float(header[2])
                    unknown_param = float(header[3])
                    dwell_time = float(header[4])
                    num_points = int(header[5])
                    pass_energy = float(header[6])
                    photon_energy = abs(float(header[7]))

                    # Get measurement type
                    measurement_type = lines[2].strip()

                    # Create a unique sheet name
                    sheet_name = normalize_sheet_name(measurement_type)
                    if sheet_name in combined_wb.sheetnames:
                        # If sheet name already exists, append a number
                        count = 1
                        while f"{sheet_name}{count}" in combined_wb.sheetnames:
                            count += 1
                        sheet_name = f"{sheet_name}{count}"

                    # Create sheet
                    ws = combined_wb.create_sheet(title=sheet_name)

                    # Add column headers
                    ws["A1"] = "Binding Energy (eV)"
                    ws["B1"] = "Corrected Data"
                    ws["C1"] = "Raw Data"
                    ws["D1"] = "Transmission"

                    # Extract intensity values
                    intensity_values = []
                    for i in range(3, len(lines)):
                        if lines[i].strip():
                            try:
                                intensity_values.append(float(lines[i].strip()))
                            except ValueError:
                                continue

                    # Ensure we have the right number of intensity values
                    if len(intensity_values) > num_points:
                        intensity_values = intensity_values[:num_points]

                    # Calculate binding energy values
                    be_values = np.linspace(be_start, be_end, num_points)

                    # Match be_values and intensity_values lengths
                    actual_points = min(len(be_values), len(intensity_values))
                    be_values = be_values[:actual_points]
                    intensity_values = intensity_values[:actual_points]

                    # Add data rows
                    for i, (be, intensity) in enumerate(zip(be_values, intensity_values), start=2):
                        ws[f"A{i}"] = be
                        ws[f"B{i}"] = intensity
                        ws[f"C{i}"] = intensity
                        ws[f"D{i}"] = 1.0  # Default transmission value

                    # Add metadata to experimental description sheet
                    exp_sheet[f"A{row}"] = f"File: {vg_file}"
                    exp_sheet[f"B{row}"] = sheet_name
                    row += 1

                    exp_sheet[f"A{row}"] = "BE Start"
                    exp_sheet[f"B{row}"] = str(be_start)
                    row += 1

                    exp_sheet[f"A{row}"] = "BE End"
                    exp_sheet[f"B{row}"] = str(be_end)
                    row += 1

                    exp_sheet[f"A{row}"] = "Energy Step"
                    exp_sheet[f"B{row}"] = str(energy_step)
                    row += 1

                    exp_sheet[f"A{row}"] = "Dwell Time"
                    exp_sheet[f"B{row}"] = str(dwell_time)
                    row += 1

                    exp_sheet[f"A{row}"] = "Number of Points"
                    exp_sheet[f"B{row}"] = str(num_points)
                    row += 1

                    exp_sheet[f"A{row}"] = "Pass Energy"
                    exp_sheet[f"B{row}"] = str(pass_energy)
                    row += 1

                    exp_sheet[f"A{row}"] = "Photon Energy"
                    exp_sheet[f"B{row}"] = str(photon_energy)
                    row += 1

                    # Add a blank row for separation
                    row += 1

                    processed_count += 1

                except Exception as e:
                    print(f"Error processing {vg_file}: {str(e)}")

            # Save combined Excel file
            if processed_count > 0:
                combined_excel_path = os.path.join(folder_path, "VG_Microtech_combined.xlsx")
                combined_wb.save(combined_excel_path)

                # Open the combined Excel file
                open_xlsx_file(window, combined_excel_path)

                window.show_popup_message2("Success",
                                           f"Created combined Excel file with {processed_count} sheets from VG-Microtech files.")
            else:
                window.show_popup_message2("Error", "No valid data found in any VG-Microtech file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing VG-Microtech files: {str(e)}")
