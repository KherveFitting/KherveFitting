# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial


# CONFIG FILE FOR XPS DATA ------------------------------------
# -------------------------------------------------------------
import pandas as pd
import wx
import matplotlib.pyplot as plt


def Init_Measurement_Data(window):
    Data = {
        'FilePath': '',
        'Number of Core levels': 0,
        'Core levels': {},
    }

    # Initialize 10 Results Tables
    for i in range(10):
        Data[f'Results Table{i}'] = {
            'Peak': {}
        }

    return Data


def add_core_level_Data(Data, window, file_path, sheet_name):
    """
    Add core level data from Excel sheet to Data structure, including experimental info
    """
    import openpyxl
    import pandas as pd
    import numpy as np

    try:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # ========== Handle EDX~Plot sheets ==========
        if sheet_name == 'EELS~Plot' or sheet_name.startswith('EELS~Plot'):
            be_values = []
            raw_data = []

            for index, row in df.iterrows():
                if index == 0:
                    continue

                be_val = row.iloc[0]
                raw_val = row.iloc[1]

                if pd.isna(be_val) or pd.isna(raw_val):
                    continue

                try:
                    be_values.append(float(be_val))
                    raw_data.append(float(raw_val))
                except (ValueError, TypeError):
                    continue

            # Create complete Background structure with all required keys
            core_level = {
                'Name': sheet_name,
                'B.E.': be_values,
                'Raw Data': raw_data,
                '_EELS_type': 'plot',
                'Background': {
                    'Bkg Type': '',
                    'Bkg Low': '',
                    'Bkg High': '',
                    'Bkg Offset Low': '',
                    'Bkg Offset High': '',
                    'Bkg X': be_values.copy() if be_values else [],
                    'Bkg Y': raw_data.copy() if raw_data else []
                }
            }

            Data['Core levels'][sheet_name] = core_level
            return Data

            # ========== Handle EELS~Map sheets ==========
        if sheet_name == 'EELS~Map':
            # For EELS~Map, just store metadata - actual map will be loaded from DM3
            energy_range = None

            for index, row in df.iterrows():
                if index == 0:
                    header_text = str(row.iloc[0])
                    if 'Range:' in header_text:
                        energy_range = header_text.split('Range:')[1].strip()
                    break  # Only need the header

            # Find DM3/DM4 file - try multiple naming patterns
            dm3_path = None
            base_dir = os.path.dirname(file_path)
            base_name = os.path.basename(file_path)

            # Pattern 1: filename_EELS.xlsx -> filename_EELS.dm3
            test_path = file_path.replace('.xlsx', '.dm3')
            if os.path.exists(test_path):
                dm3_path = test_path

            # Pattern 2: filename_EELS.xlsx -> filename_EELS.dm4
            if not dm3_path:
                test_path = file_path.replace('.xlsx', '.dm4')
                if os.path.exists(test_path):
                    dm3_path = test_path

            # Pattern 3: filename_EELS.xlsx -> filename.dm3 (original file without _EELS)
            if not dm3_path:
                test_path = file_path.replace('_EELS.xlsx', '.dm3')
                if os.path.exists(test_path):
                    dm3_path = test_path

            # Pattern 4: filename_EELS.xlsx -> filename.dm4 (original file without _EELS)
            if not dm3_path:
                test_path = file_path.replace('_EELS.xlsx', '.dm4')
                if os.path.exists(test_path):
                    dm3_path = test_path

            # Pattern 5: Look for any .dm3 or .dm4 file with similar base name in same directory
            if not dm3_path:
                # Get base name without _EELS suffix
                name_without_eels = base_name.replace('_EELS.xlsx', '')
                for ext in ['.dm3', '.dm4']:
                    # Try exact match
                    test_path = os.path.join(base_dir, name_without_eels + ext)
                    if os.path.exists(test_path):
                        dm3_path = test_path
                        break
                    # Try with any suffix pattern
                    import glob
                    pattern = os.path.join(base_dir, name_without_eels + '*' + ext)
                    matches = glob.glob(pattern)
                    if matches:
                        dm3_path = matches[0]
                        break

            print(f"EELS~Map: Looking for DM3/DM4 file")
            print(f"  Excel path: {file_path}")
            print(f"  Found DM3 path: {dm3_path}")

            core_level = {
                'Name': sheet_name,
                'Energy_Range': energy_range if energy_range else 'N/A',
                '_EELS_type': 'map',
                '_DM3_Path': dm3_path
            }

            Data['Core levels'][sheet_name] = core_level
            return Data

        # ========== Handle XPS~Map sheets ==========
        if sheet_name.startswith('XPS~Map'):
            be_values = []
            y_columns = {}
            num_sweeps = 0

            header_row = df.iloc[0] if len(df) > 0 else None
            if header_row is not None:
                for col_idx, val in enumerate(header_row):
                    if str(val).startswith('Y') and str(val)[1:].isdigit():
                        num_sweeps = max(num_sweeps, int(str(val)[1:]))

            for i in range(1, num_sweeps + 1):
                y_columns[f'Y{i}'] = []

            for index, row in df.iterrows():
                if index == 0:
                    continue
                be_val = row.iloc[0]
                if pd.isna(be_val):
                    continue
                try:
                    be_values.append(float(be_val))
                    for i in range(1, num_sweeps + 1):
                        col_idx = i
                        if col_idx < len(row) and not pd.isna(row.iloc[col_idx]):
                            y_columns[f'Y{i}'].append(float(row.iloc[col_idx]))
                        else:
                            y_columns[f'Y{i}'].append(0.0)
                except (ValueError, TypeError):
                    continue

            experimental_info = {}
            wb = openpyxl.load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                exp_col = None
                for col in range(num_sweeps + 5, min(num_sweeps + 30, ws.max_column + 1)):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and "Experimental Description" in str(cell_value):
                        exp_col = col
                        break
                if exp_col:
                    for row in range(2, ws.max_row + 1):
                        param_cell = ws.cell(row=row, column=exp_col)
                        value_cell = ws.cell(row=row, column=exp_col + 1)
                        if param_cell.value is not None and str(param_cell.value).strip():
                            experimental_info[str(param_cell.value).strip()] = str(value_cell.value).strip() if value_cell.value else ""
            wb.close()

            core_level = {
                'Name': sheet_name,
                'B.E.': [float(f"{val:.2f}") for val in be_values],
                '_Map_type': 'scienta',
                '_num_sweeps': num_sweeps,
                '_core_level': experimental_info.get('Core Level', ''),
            }
            for col_name, values in y_columns.items():
                core_level[col_name] = [float(f"{val:.2f}") for val in values]
            if experimental_info:
                core_level['ExperimentalInfo'] = experimental_info
            if be_values:
                core_level['Background'] = {
                    'Bkg Type': '',
                    'Bkg Low': float(f"{min(be_values):.2f}"),
                    'Bkg High': float(f"{max(be_values):.2f}"),
                    'Bkg Offset Low': 0,
                    'Bkg Offset High': 0
                }
            Data['Core levels'][sheet_name] = core_level
            return Data

        # Extract B.E. and Raw Data columns
        be_values = []
        raw_data = []
        corrected_data = []
        transmission = []

        # Skip the header row and extract data
        for index, row in df.iterrows():
            if index == 0:  # Skip header
                continue

            be_val = row.iloc[0]  # Column A (B.E.)
            raw_val = row.iloc[1]  # Column B (Raw Data)

            # Handle missing or invalid data
            if pd.isna(be_val) or pd.isna(raw_val):
                continue

            try:
                be_values.append(float(be_val))
                raw_data.append(float(raw_val))

                # Check if corrected data column exists
                if len(row) > 2 and not pd.isna(row.iloc[2]):
                    corrected_data.append(float(row.iloc[2]))
                else:
                    corrected_data.append(float(raw_val))

                # Check if transmission column exists
                if len(row) > 3 and not pd.isna(row.iloc[3]):
                    transmission.append(float(row.iloc[3]))
                else:
                    transmission.append(1.0)

            except (ValueError, TypeError):
                continue

        # Extract experimental description data from Excel file
        experimental_info = {}

        # Load workbook to access experimental description columns
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Search for experimental description column (typically around column 45-50)
            exp_col = None
            for col in range(40, min(61, ws.max_column + 1)):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and "Experimental Description" in str(cell_value):
                    exp_col = col
                    break

            if exp_col:
                # Read experimental description data
                for row in range(2, ws.max_row + 1):
                    param_cell = ws.cell(row=row, column=exp_col)
                    value_cell = ws.cell(row=row, column=exp_col + 1)

                    if param_cell.value is not None and str(param_cell.value).strip():
                        param_name = str(param_cell.value).strip()
                        param_value = str(value_cell.value).strip() if value_cell.value is not None else ""
                        experimental_info[param_name] = param_value

        # Create the core level data structure with .2f formatting
        core_level_data = {
            'B.E.': [float(f"{val:.2f}") for val in be_values],
            'Raw Data': [float(f"{val:.2f}") for val in raw_data],
            'Corrected Data': [float(f"{val:.2f}") for val in corrected_data],
            'Transmission': [float(f"{val:.2f}") for val in transmission],
            'Name': sheet_name
        }

        # Add experimental info to the core level data if found
        if experimental_info:
            core_level_data['ExperimentalInfo'] = experimental_info

        # Initialize background structure
        if be_values:
            core_level_data['Background'] = {
                'Bkg Y': core_level_data['Raw Data'],
                'Bkg Type': '',
                'Bkg Low': float(f"{min(be_values):.2f}"),
                'Bkg High': float(f"{max(be_values):.2f}"),
                'Bkg Offset Low': 0,
                'Bkg Offset High': 0
            }

        # Add to Data structure
        if 'Core levels' not in Data:
            Data['Core levels'] = {}

        Data['Core levels'][sheet_name] = core_level_data
        Data['Number of Core levels'] = len(Data['Core levels'])

        return Data

    except Exception as e:
        print(f"Error adding core level data for {sheet_name}: {e}")
        return Data


def add_peak_to_core_level_Data(data, core_name, peak_data):
    if core_name in data['Core levels']:
        fitting = data['Core levels'][core_name]['Fitting']
        fitting.update(peak_data)
    else:
        print(f"Core level {core_name} does not exist.")


def set_consistent_fonts(window):
    if 'wxMac' in wx.PlatformInfo:
        default_font = 'Helvetica'
        # default_font = 'Calibri'
        # default_font = 'Arial'

        STANDARD_FONT_SIZE = 11
        # STANDARD_FONT_SIZE = 5
        HEADER_FONT_SIZE = 12
        # HEADER_FONT_SIZE = 5
        LABEL_FONT_SIZE = 11
        # LABEL_FONT_SIZE = 5
    elif 'wxGTK' in wx.PlatformInfo:  # wxGTK is the Linux version of wxWidgets
        default_font = 'DejaVu Sans'  # A common Linux font
        STANDARD_FONT_SIZE = 9
        HEADER_FONT_SIZE = 10
        LABEL_FONT_SIZE = 9
    else:
        # default_font = 'Arial'
        # default_font = 'Helvetica'
        default_font = 'Calibri'

        STANDARD_FONT_SIZE = 9
        # STANDARD_FONT_SIZE = 5
        HEADER_FONT_SIZE = 10
        # HEADER_FONT_SIZE = 5
        LABEL_FONT_SIZE = 9
        # LABEL_FONT_SIZE = 5

    window.SetFont(wx.Font(STANDARD_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                           faceName=default_font))

    # Handle all wx widgets including buttons
    def set_font_recursive(widget):
        widget.SetFont(wx.Font(STANDARD_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                               faceName=default_font))
        if isinstance(widget, wx.Button):
            widget.SetFont(wx.Font(STANDARD_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                                   faceName=default_font))
        for child in widget.GetChildren():
            set_font_recursive(child)

    set_font_recursive(window)

    # Handle matplotlib
    if hasattr(window, 'ax'):
        plt.rcParams['font.family'] = default_font
        window.ax.tick_params(axis='both', labelsize=STANDARD_FONT_SIZE)
        window.ax.set_xlabel(window.ax.get_xlabel(), fontsize=STANDARD_FONT_SIZE, fontname=default_font)
        window.ax.set_ylabel(window.ax.get_ylabel(), fontsize=STANDARD_FONT_SIZE, fontname=default_font)
        if window.ax.get_title():
            window.ax.set_title(window.ax.get_title(), fontsize=HEADER_FONT_SIZE, fontname=default_font)

    # Handle grids
    if hasattr(window, 'peak_params_grid'):
        window.peak_params_grid.SetDefaultCellFont(
            wx.Font(STANDARD_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                    faceName=default_font))
        window.peak_params_grid.SetLabelFont(
            wx.Font(LABEL_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                    faceName=default_font))

    if hasattr(window, 'results_grid'):
        window.results_grid.SetDefaultCellFont(
            wx.Font(STANDARD_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                    faceName=default_font))
        window.results_grid.SetLabelFont(
            wx.Font(LABEL_FONT_SIZE, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,
                    faceName=default_font))

    if hasattr(window, 'fitting_window') and window.fitting_window:
        set_consistent_fonts(window.fitting_window)

    if hasattr(window, 'background_window') and window.background_window:
        set_consistent_fonts(window.background_window)

    if hasattr(window, 'canvas'):
        window.canvas.draw_idle()