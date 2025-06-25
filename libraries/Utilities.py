# libraries/utilities.py
import wx
import numpy as np
import openpyxl
from openpyxl import load_workbook
import os
import sys
import shutil
import pandas as pd
import libraries.Sheet_Operations
import json
from scipy.ndimage import gaussian_filter
from scipy.signal import savgol_filter
from scipy.integrate import cumtrapz

# from KherveFitting import FIRST_TIME_USE


def check_first_time_use(frame):
    config = frame.load_config()
    times_opened = config.get('times_opened', 0)


    # Check if registration is needed - use frame instead of window
    if not hasattr(frame, 'registered') or not frame.registered:
        from libraries.MarketResearch import check_registration_needed, show_registration_form
        if check_registration_needed():
            registered = show_registration_form()
            if registered:
                # Update frame's registered status
                frame.registered = True
                print("User registered successfully.")
                # Reload config
                config = frame.load_config()
                # Make sure config is also updated
                frame.save_config()

    # if times_opened == 0:
    #     # Show registration form on first run
    #     from libraries.MarketResearch import check_registration_needed, show_registration_form
    #     if check_registration_needed():
    #         show_registration_form()
    elif times_opened == 2:
        # Show manual dialog on second run
        dlg = wx.MessageDialog(frame,
                              "Would you like to open the manual to the Getting Started section?",
                              "Welcome to KherveFitting",
                              wx.YES_NO | wx.ICON_QUESTION)

        if dlg.ShowModal() == wx.ID_YES:
            import os
            import sys
            import platform
            import subprocess

            if getattr(sys, 'frozen', False):
                application_path = os.path.dirname(sys.executable)
            else:
                application_path = os.path.dirname(os.path.abspath(__file__))

            manual_path = os.path.join(application_path, "Manual.pdf")

            if platform.system() == 'Windows':
                os.startfile(manual_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', manual_path])
            else:  # Linux and other Unix-like systems
                subprocess.call(['xdg-open', manual_path])

        dlg.Destroy()

    # Update times_opened in the frame's attribute
    frame.times_opened = times_opened + 1

    # Save config
    frame.save_config()

def _clear_peak_params_grid(window):
    num_rows = window.peak_params_grid.GetNumberRows()
    if num_rows > 0:
        window.peak_params_grid.DeleteRows(0, num_rows)
    window.peak_count = 0  # Reset peak count when clearing the grid


def copy_cell(grid):
    print("Copy cell function called")
    print(f"Grid has focus: {grid.HasFocus()}")

    if grid.HasFocus():
        selected_blocks = grid.GetSelectedBlocks()
        if selected_blocks.GetCount() > 0:
            block = selected_blocks.GetBlock(0)
            row, col = block.GetTopRow(), block.GetLeftCol()
            print(f"Selected cell: ({row}, {col})")
            print("CTL C has focus")
            data = grid.GetCellValue(row, col)
            if wx.TheClipboard.Open():
                wx.TheClipboard.SetData(wx.TextDataObject(data))
                wx.TheClipboard.Close()
        else:
            print("No cell selected")
    else:
        print("Grid does not have focus")


def paste_cell(grid):
    if grid.HasFocus():
        selected_blocks = grid.GetSelectedBlocks()
        if selected_blocks.GetCount() > 0 and wx.TheClipboard.Open():
            if wx.TheClipboard.IsSupported(wx.DataFormat(wx.DF_TEXT)):
                data = wx.TextDataObject()
                wx.TheClipboard.GetData(data)
                block = selected_blocks.GetBlock(0)
                row, col = block.GetTopRow(), block.GetLeftCol()
                grid.SetCellValue(row, col, data.GetText())
            wx.TheClipboard.Close()

def load_rsf_data(file_path):
    rsf_dict = {}
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.split()
            if len(parts) == 2:
                core_level, rsf = parts
                rsf_dict[core_level] = float(rsf)
    return rsf_dict


class DraggableText:
    def __init__(self, text):
        self.text = text
        self.press = None
        self.menu = None
        self.connect()

    def connect(self):
        self.cidpress = self.text.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.cidrelease = self.text.figure.canvas.mpl_connect('button_release_event', self.on_release)
        self.cidmotion = self.text.figure.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.cidkeypress = self.text.figure.canvas.mpl_connect('key_press_event', self.on_key)
        self.cidrightclick = self.text.figure.canvas.mpl_connect('button_press_event', self.on_right_click)

    def on_right_click(self, event):
        if event.button != 3: return
        if event.inaxes != self.text.axes: return
        contains, _ = self.text.contains(event)
        if not contains: return

        window = event.canvas.GetParent().GetParent().GetParent()
        menu = wx.Menu()

        rotate_item = menu.Append(wx.ID_ANY, "Rotate")
        size_item = menu.Append(wx.ID_ANY, "Change Size")
        delete_item = menu.Append(wx.ID_ANY, "Delete")

        window.Bind(wx.EVT_MENU, self.on_rotate, rotate_item)
        window.Bind(wx.EVT_MENU, self.on_change_size, size_item)
        window.Bind(wx.EVT_MENU, self.on_delete, delete_item)

        window.PopupMenu(menu)
        menu.Destroy()

    def on_rotate(self, event):
        window = self.text.figure.canvas.GetParent().GetParent().GetParent()
        dlg = wx.TextEntryDialog(window, 'Enter rotation angle (degrees):', 'Rotate Text')
        if dlg.ShowModal() == wx.ID_OK:
            try:
                angle = float(dlg.GetValue())
                self.text.set_rotation(angle)
                self.text.figure.canvas.draw()
            except ValueError:
                wx.MessageBox('Please enter a valid number', 'Error')
        dlg.Destroy()

    def on_change_size(self, event):
        window = self.text.figure.canvas.GetParent().GetParent().GetParent()
        dlg = wx.TextEntryDialog(window, 'Enter font size:', 'Change Text Size')
        if dlg.ShowModal() == wx.ID_OK:
            try:
                size = float(dlg.GetValue())
                self.text.set_fontsize(size)
                self.text.figure.canvas.draw()
            except ValueError:
                wx.MessageBox('Please enter a valid number', 'Error')
        dlg.Destroy()

    def on_delete(self, event):
        self.text.remove()
        self.text.figure.canvas.draw()

    def on_press(self, event):
        if event.button != 1: return
        if event.inaxes != self.text.axes: return
        contains, _ = self.text.contains(event)
        if not contains: return
        self.press = self.text.get_position(), event.xdata, event.ydata

    def on_motion(self, event):
        if self.press is None: return
        if event.inaxes != self.text.axes: return
        pos, xpress, ypress = self.press
        dx = event.xdata - xpress
        dy = event.ydata - ypress
        self.text.set_position((pos[0] + dx, pos[1] + dy))
        self.text.figure.canvas.draw()

    def on_release(self, event):
        self.press = None

    def on_key(self, event):
        if event.key == 'delete':
            contains, _ = self.text.contains(event)
            if contains:
                self.text.remove()
                self.text.figure.canvas.draw()


def add_draggable_text(window):
    window.text_mode = not getattr(window, 'text_mode', False)
    if window.text_mode:
        window.canvas.Bind(wx.EVT_LEFT_DOWN, on_canvas_click)
    else:
        window.canvas.Unbind(wx.EVT_LEFT_DOWN)
        window.canvas.mpl_connect('button_press_event', lambda event: None)  # Reset Matplotlib listener


def on_canvas_click(event):
    parent = event.GetEventObject().GetParent()
    while not isinstance(parent, wx.Frame):
        parent = parent.GetParent()
    window = parent

    if not window.text_mode:
        event.Skip()
        return

    x, y = event.GetPosition()
    ax = window.ax
    display_point = ax.transData.inverted().transform((x, y))

    dlg = wx.TextEntryDialog(window, 'Enter text:', 'Add Text Annotation')
    if dlg.ShowModal() == wx.ID_OK:
        text = dlg.GetValue()
        annotation = ax.text(display_point[0], display_point[1], text,
                             picker=5,
                             bbox=dict(facecolor='white', edgecolor='none', alpha=0.7))

        sheet_name = window.sheet_combobox.GetValue()
        if 'Labels' not in window.Data['Core levels'][sheet_name]:
            window.Data['Core levels'][sheet_name]['Labels'] = []

        window.Data['Core levels'][sheet_name]['Labels'].append({
            'text': text,
            'x': display_point[0],
            'y': display_point[1]
        })

        draggable = DraggableText(annotation)
        window.canvas.draw()
    dlg.Destroy()

    window.text_mode = False
    window.canvas.Unbind(wx.EVT_LEFT_DOWN)


def on_delete_sheet(window, event):
    import wx
    sheet_name = window.sheet_combobox.GetValue()
    dlg = wx.MessageDialog(window, f"Are you sure you want to delete sheet {sheet_name}?",
                           "Confirm Delete", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION)

    if dlg.ShowModal() == wx.ID_YES:
        # Backup before deletion
        from libraries.Utilities import perform_auto_backup
        perform_auto_backup(window)

        # Remove from Excel file
        wb = openpyxl.load_workbook(window.Data['FilePath'])
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            wb.save(window.Data['FilePath'])

        # Remove from window.Data
        if sheet_name in window.Data['Core levels']:
            del window.Data['Core levels'][sheet_name]
            window.Data['Number of Core levels'] -= 1

        # Save JSON file
        json_file_path = os.path.splitext(window.Data['FilePath'])[0] + '.json'
        from libraries.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Update combobox
        window.sheet_combobox.Delete(window.sheet_combobox.FindString(sheet_name))
        if window.sheet_combobox.GetCount() > 0:
            window.sheet_combobox.SetSelection(0)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, window.sheet_combobox.GetString(0))

        # Clear plots if no sheets remain
        if window.sheet_combobox.GetCount() == 0:
            window.ax.clear()
            window.canvas.draw()

        # Refresh sheets after deletion
        try:
            from libraries.Save import refresh_sheets
            from libraries.Sheet_Operations import on_sheet_selected
            refresh_sheets(window, on_sheet_selected, update_console)
        except Exception as refresh_err:
            print(f"Error refreshing sheets: {refresh_err}")

        # Close and reopen the file manager if it exists
        if hasattr(window, 'file_manager') and window.file_manager is not None:
            try:
                # Close existing file manager
                window.file_manager.Close()
                window.file_manager.Destroy()
                window.file_manager = None

                # Reopen file manager
                import wx
                wx.CallAfter(window.on_open_file_manager, None)
            except Exception as e:
                print(f"Error refreshing file manager: {e}")
                pass

    dlg.Destroy()


def rename_sheet(window, new_sheet_name):
    if not new_sheet_name:
        return

    # Create backup before renaming
    from libraries.Utilities import perform_auto_backup
    perform_auto_backup(window)

    old_sheet_name = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    # Rename in Excel file
    wb = openpyxl.load_workbook(file_path)
    if old_sheet_name in wb.sheetnames:
        sheet = wb[old_sheet_name]
        sheet.title = new_sheet_name
        wb.save(file_path)

    # Update window.Data
    window.Data['Core levels'][new_sheet_name] = window.Data['Core levels'].pop(old_sheet_name)

    # Update combobox
    index = window.sheet_combobox.FindString(old_sheet_name)
    window.sheet_combobox.Delete(index)
    window.sheet_combobox.Insert(new_sheet_name, index)
    window.sheet_combobox.SetSelection(index)

    # Update plot
    from libraries.Sheet_Operations import on_sheet_selected
    on_sheet_selected(window, new_sheet_name)

    # Save JSON file
    json_file_path = os.path.splitext(file_path)[0] + '.json'
    from libraries.Save import convert_to_serializable_and_round
    json_data = convert_to_serializable_and_round(window.Data)
    with open(json_file_path, 'w') as json_file:
        json.dump(json_data, json_file, indent=2)

    # Refresh sheets after renaming
    from libraries.Sheet_Operations import on_sheet_selected
    from libraries.Save import refresh_sheets
    refresh_sheets(window, on_sheet_selected, None)

    # Close and reopen the file manager if it exists
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            # Close existing file manager
            window.file_manager.Close()
            window.file_manager.Destroy()
            window.file_manager = None

            # Reopen file manager
            import wx
            wx.CallAfter(window.on_open_file_manager, None)
        except Exception as e:
            print(f"Error refreshing file manager: {e}")
            pass


def get_unique_sheet_name(base_name, existing_sheets):
    """Get unique sheet name by incrementing number suffix if name exists."""
    if base_name not in existing_sheets:
        return base_name

    # Extract the base and any existing number
    parts = base_name.split('_')
    prefix = '_'.join(parts[:-1]) if len(parts) > 1 else base_name

    # Check for existing number at end
    number = 1
    if parts[-1].isdigit():
        number = int(parts[-1])
        prefix = '_'.join(parts[:-1])

    # Find next available number
    while f"{prefix}_{number}" in existing_sheets:
        number += 1

    return f"{prefix}_{number}"


def copy_sheet(window):
    sheet_name = window.sheet_combobox.GetValue()

    # Extract the base name and numeric suffix with more detailed pattern
    import re
    match = re.match(r'([A-Za-z]+\d+[spdfg]*)(\d*)$', sheet_name)
    if match:
        base_name = match.group(1)  # Ti2p
        number_suffix = match.group(2)  # 8 or "" if none
    else:
        base_name = sheet_name
        number_suffix = ""

    # Find existing sheets with the same base name
    existing_sheets = list(window.Data['Core levels'].keys())
    existing_base_sheets = [s for s in existing_sheets if re.match(r'^' + re.escape(base_name) + r'\d*$', s)]

    # Find the highest number suffix
    max_suffix = 0
    for s in existing_base_sheets:
        # Extract just the numeric suffix at the end
        suffix_match = re.search(r'^' + re.escape(base_name) + r'(\d+)$', s)
        if suffix_match:
            suffix_num = int(suffix_match.group(1))
            max_suffix = max(max_suffix, suffix_num)
        elif s == base_name:  # Handle base without suffix (equivalent to suffix 0)
            max_suffix = max(max_suffix, 0)

    # Create new sheet name with incremented suffix
    if base_name in existing_sheets or f"{base_name}0" in existing_sheets or max_suffix > 0:
        # Increment the highest suffix by 1
        new_sheet_name = f"{base_name}{max_suffix + 1}"
    else:
        # First instance, don't add a suffix
        new_sheet_name = base_name

    # Final check to ensure uniqueness (prevents collisions)
    while new_sheet_name in existing_sheets:
        suffix_match = re.search(r'^' + re.escape(base_name) + r'(\d+)$', new_sheet_name)
        if suffix_match:
            current_suffix = int(suffix_match.group(1))
            new_sheet_name = f"{base_name}{current_suffix + 1}"
        else:
            new_sheet_name = f"{base_name}1"

    # Copy data to new sheet
    window.Data['Core levels'][new_sheet_name] = window.Data['Core levels'][sheet_name].copy()
    window.Data['Number of Core levels'] += 1

    # Create Excel sheet
    import pandas as pd
    import openpyxl
    wb = openpyxl.load_workbook(window.Data['FilePath'])
    df = pd.DataFrame({
        'BE': window.Data['Core levels'][sheet_name]['B.E.'],
        'Raw Data': window.Data['Core levels'][sheet_name]['Raw Data'],
        'Background': window.Data['Core levels'][sheet_name]['Background']['Bkg Y'],
        'Transmission': [1.0] * len(window.Data['Core levels'][sheet_name]['B.E.'])
    })

    with pd.ExcelWriter(window.Data['FilePath'], engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=new_sheet_name, index=False)

    # Update JSON file
    json_file_path = os.path.splitext(window.Data['FilePath'])[0] + '.json'
    if os.path.exists(json_file_path):
        from libraries.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

    # Update sheet list and display
    window.sheet_combobox.Append(new_sheet_name)
    window.sheet_combobox.SetValue(new_sheet_name)
    from libraries.Sheet_Operations import on_sheet_selected
    on_sheet_selected(window, new_sheet_name)

    # Close and reopen the file manager if it exists
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            # Close existing file manager
            window.file_manager.Close()
            window.file_manager.Destroy()
            window.file_manager = None

            # Reopen file manager
            import wx
            wx.CallAfter(window.on_open_file_manager, None)
        except Exception as e:
            print(f"Error refreshing file manager: {e}")
            pass

    return new_sheet_name

def save_modified_data(self, x, y, sheet_name, operation_type):
    wb = openpyxl.load_workbook(self.parent.Data['FilePath'])
    sheet_name = get_unique_sheet_name(sheet_name, wb.sheetnames)

    # Get unique sheet name
    existing_sheets = wb.sheetnames
    sheet_name = get_unique_sheet_name(sheet_name, existing_sheets)


def propagate_fwhm_difference(window, row, col):
    """Propagate FWHM differences from reference peak to all other peaks"""
    # Save state
    from libraries.Save import save_state
    save_state(window)

    if col != 4 or row % 2 != 1:  # Only for FWHM constraint rows
        return

    param_row = row - 1
    ref_peak_letter = window.peak_params_grid.GetCellValue(param_row, 0)
    ref_peak_index = ord(ref_peak_letter) - 65
    ref_fwhm = float(window.peak_params_grid.GetCellValue(ref_peak_index * 2, 4))

    num_peaks = window.peak_params_grid.GetNumberRows() // 2

    # Get sheet and peaks data for saving
    sheet_name = window.sheet_combobox.GetValue()
    peaks = window.Data['Core levels'][sheet_name]['Fitting']['Peaks']

    for i in range(num_peaks):
        if i != ref_peak_index:  # Skip the reference peak itself
            param_row_i = i * 2
            constraint_row_i = i * 2 + 1
            current_fwhm = float(window.peak_params_grid.GetCellValue(param_row_i, 4))

            # Calculate difference
            difference = current_fwhm - ref_fwhm

            # Create constraint string: B+0.4#0.1 or B-0.2#0.1
            if difference >= 0:
                constraint_str = f"{ref_peak_letter}+{difference:.2f}#0.1"
            else:
                constraint_str = f"{ref_peak_letter}{difference:.2f}#0.1"  # difference is already negative

            window.peak_params_grid.SetCellValue(constraint_row_i, col, constraint_str)

            # ASave constraint to Data structure
            correct_peak_key = list(peaks.keys())[i]

            if 'Constraints' not in peaks[correct_peak_key]:
                peaks[correct_peak_key]['Constraints'] = {}

            peaks[correct_peak_key]['Constraints']['FWHM'] = constraint_str

    window.peak_params_grid.ForceRefresh()

def propagate_constraint(window, row, col):
    """Propagate the constraint in the selected cell to all other peaks in the same column"""
    if row % 2 != 1 or col not in [2, 3, 4, 5, 6, 7, 8, 9]:
        return  # Only work on constraint rows and specific columns

    from libraries.Save import save_state
    save_state(window)  # Save state for undo

    peak_index = row // 2
    peak_letter = chr(65 + peak_index)  # Convert peak index to letter (A, B, C...)

    # Get values from the source peak
    data_row = peak_index * 2
    source_value = float(window.peak_params_grid.GetCellValue(data_row, col))

    # Update all other rows based on column type
    num_peaks = window.peak_params_grid.GetNumberRows() // 2
    sheet_name = window.sheet_combobox.GetValue()
    peaks = window.Data['Core levels'][sheet_name]['Fitting']['Peaks']

    constraint_names = {
        2: 'Position',
        3: 'Height',
        4: 'FWHM',
        5: 'L/G',
        6: 'Area',
        7: 'Sigma',
        8: 'Gamma',
        9: 'Skew'
    }

    constraint_name = constraint_names.get(col)

    for i in range(num_peaks):
        if i == peak_index:  # Skip the source peak
            continue

        data_row_i = i * 2
        constraint_row_i = i * 2 + 1

        # Different handling based on column
        if col == 2:  # Position
            # Calculate split (difference in position)
            target_pos = float(window.peak_params_grid.GetCellValue(data_row_i, col))
            split = target_pos - source_value
            constraint_value = f"{peak_letter}+{split:.2f}#0.1"

        elif col == 6:  # Area
            # Calculate ratio
            target_area = float(window.peak_params_grid.GetCellValue(data_row_i, col))
            source_area = float(window.peak_params_grid.GetCellValue(data_row, col))
            ratio = (target_area / source_area) if source_area != 0 else 1
            constraint_value = f"{peak_letter}*{ratio:.2f}"

        elif col == 7:  # Sigma - also update Gamma
            constraint_value = f"{peak_letter}*1"
            # Update sigma value
            window.peak_params_grid.SetCellValue(data_row_i, col, f"{source_value:.2f}")

            # Also update gamma and its constraint
            source_gamma = float(window.peak_params_grid.GetCellValue(data_row, 8))
            window.peak_params_grid.SetCellValue(data_row_i, 8, f"{source_gamma:.2f}")
            window.peak_params_grid.SetCellValue(constraint_row_i, 8, f"{peak_letter}*1")

            # Update in Data structure
            if i < len(list(peaks.values())):
                peak_data = list(peaks.values())[i]
                peak_data['Sigma'] = source_value
                peak_data['Gamma'] = source_gamma
                if 'Constraints' not in peak_data:
                    peak_data['Constraints'] = {}
                peak_data['Constraints']['Sigma'] = constraint_value
                peak_data['Constraints']['Gamma'] = f"{peak_letter}*1"

        elif col == 8:  # Gamma - also update Sigma
            constraint_value = f"{peak_letter}*1"
            # Update gamma value
            window.peak_params_grid.SetCellValue(data_row_i, col, f"{source_value:.2f}")

            # Also update sigma and its constraint
            source_sigma = float(window.peak_params_grid.GetCellValue(data_row, 7))
            window.peak_params_grid.SetCellValue(data_row_i, 7, f"{source_sigma:.2f}")
            window.peak_params_grid.SetCellValue(constraint_row_i, 7, f"{peak_letter}*1")

            # Update in Data structure
            if i < len(list(peaks.values())):
                peak_data = list(peaks.values())[i]
                peak_data['Gamma'] = source_value
                peak_data['Sigma'] = source_sigma
                if 'Constraints' not in peak_data:
                    peak_data['Constraints'] = {}
                peak_data['Constraints']['Gamma'] = constraint_value
                peak_data['Constraints']['Sigma'] = f"{peak_letter}*1"

        elif col in [4, 5, 9]:  # FWHM, L/G, Skew
            # Set all to the same value for these parameters
            constraint_value = f"{peak_letter}*1"
            # Also update the actual value in the data row
            window.peak_params_grid.SetCellValue(data_row_i, col, f"{source_value:.2f}")

            # Update in Data structure
            if constraint_name and i < len(list(peaks.values())):
                peak_data = list(peaks.values())[i]
                peak_data[constraint_name] = source_value

        else:  # Other columns
            constraint_value = f"{peak_letter}*1"

        # Set the constraint in the grid (if not already set for special cases)
        if col not in [7, 8] or not window.peak_params_grid.GetCellValue(constraint_row_i, col):
            window.peak_params_grid.SetCellValue(constraint_row_i, col, constraint_value)

        # Update in Data structure (if not already handled in special cases)
        if constraint_name and col not in [7, 8] and i < len(list(peaks.values())):
            peak_data = list(peaks.values())[i]
            if 'Constraints' not in peak_data:
                peak_data['Constraints'] = {}
            peak_data['Constraints'][constraint_name] = constraint_value

    window.peak_fitting_grid.refresh_peak_params_grid()
    window.clear_and_replot()

class CropWindow(wx.Frame):
    def __init__(self, parent,*args, **kw):
        super().__init__(parent, *args, **kw, style=wx.DEFAULT_FRAME_STYLE & ~(
                    wx.RESIZE_BORDER | wx.MAXIMIZE_BOX | wx.MINIMIZE_BOX | wx.SYSTEM_MENU) | wx.STAY_ON_TOP)
        self.parent = parent
        self.SetTitle("Crop Data")
        self.SetSize(250, 220)

        self.panel = wx.Panel(self)


        sizer = wx.BoxSizer(wx.VERTICAL)

        main_pos = parent.GetPosition()
        main_size = parent.GetSize()
        crop_size = self.GetSize()

        x = main_pos.x + (main_size.width - crop_size.width) // 2
        y = main_pos.y + (main_size.height - crop_size.height) // 2

        self.SetPosition((x, y))

        # Range controls
        range_box = wx.StaticBox(self.panel, label="Crop Range")
        range_sizer = wx.StaticBoxSizer(range_box, wx.VERTICAL)

        min_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.min_ctrl = wx.SpinCtrlDouble(self.panel, min=0, max=2000, inc=0.1)
        min_sizer.Add(wx.StaticText(self.panel, label="Min BE:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        min_sizer.Add(self.min_ctrl, 1)

        max_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.max_ctrl = wx.SpinCtrlDouble(self.panel, min=0, max=2000, inc=0.1)
        max_sizer.Add(wx.StaticText(self.panel, label="Max BE:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        max_sizer.Add(self.max_ctrl, 1)

        range_sizer.Add(min_sizer, 0, wx.EXPAND | wx.ALL, 5)
        range_sizer.Add(max_sizer, 0, wx.EXPAND | wx.ALL, 5)

        # Sheet name control
        name_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.name_ctrl = wx.TextCtrl(self.panel)
        name_sizer.Add(wx.StaticText(self.panel, label="New Sheet:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        name_sizer.Add(self.name_ctrl, 1)

        # Crop button
        self.crop_btn = wx.Button(self.panel, label="Crop")
        self.crop_btn.SetMinSize((125, 40))
        self.crop_btn.Bind(wx.EVT_BUTTON, self.on_crop)

        sizer.Add(range_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(name_sizer, 0, wx.EXPAND | wx.ALL, 5)
        sizer.Add(self.crop_btn, 0, wx.EXPAND | wx.ALL, 5)

        self.panel.SetSizer(sizer)

        self.vline_min = None
        self.vline_max = None

        self.init_values()
        self.Bind(wx.EVT_CLOSE, self.on_close)
        self.min_ctrl.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_range_change)
        self.max_ctrl.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_range_change)

    def init_values(self):
        sheet_name = self.parent.sheet_combobox.GetValue()
        x_values = self.parent.Data['Core levels'][sheet_name]['B.E.']
        min_be = min(x_values) + 2
        max_be = max(x_values) - 2
        self.min_ctrl.SetValue(min_be)
        self.max_ctrl.SetValue(max_be)

        # Extract base name from sheet_name
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        if match:
            base_name = match.group(1)
        else:
            base_name = sheet_name

        # Find the earliest available row name
        new_name = self.get_earliest_row_name(base_name)
        self.name_ctrl.SetValue(new_name)
        self.show_vlines()

    def get_earliest_row_name(self, base_name):
        """
        Find the earliest available row for a core level.
        Returns the sheet name with appropriate row number.
        """
        import re

        # Check all sheets in parent data
        all_sheets = list(self.parent.Data['Core levels'].keys())

        # Check if base name without number exists (row 0)
        if base_name in all_sheets:
            # Base name exists, need to find next available
            rows_used = []
        else:
            # Base name doesn't exist, it's available
            return base_name

        # Pattern to match base name followed by optional number
        pattern = re.compile(f"^{re.escape(base_name)}(\\d+)$")

        # Collect all used row numbers
        for sheet in all_sheets:
            if sheet == base_name:  # This is row 0
                rows_used.append(0)
            else:
                match = pattern.match(sheet)
                if match:
                    rows_used.append(int(match.group(1)))

        # Find first unused row number
        for i in range(1000):  # Reasonable upper limit
            if i not in rows_used:
                if i == 0:
                    return base_name  # No suffix for row 0
                else:
                    return f"{base_name}{i}"

        # Fallback (unlikely to reach)
        return f"{base_name}{len(all_sheets)}"

    def show_vlines(self):
        if self.vline_min:
            self.vline_min.remove()
        if self.vline_max:
            self.vline_max.remove()

        self.vline_min = self.parent.ax.axvline(self.min_ctrl.GetValue(), color='green', linestyle='--', alpha=0.5)
        self.vline_max = self.parent.ax.axvline(self.max_ctrl.GetValue(), color='green', linestyle='--', alpha=0.5)
        self.parent.canvas.draw_idle()

    def on_range_change(self, event):
        self.show_vlines()

    def on_crop(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        suggested_name = self.name_ctrl.GetValue()
        min_be = self.min_ctrl.GetValue()
        max_be = self.max_ctrl.GetValue()

        # Extract base name and ensure we don't create a '0' suffix
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)(\d*)$', suggested_name)
        if match:
            base_name = match.group(1)  # Base name like C1s
            number_suffix = match.group(2)  # Numeric suffix or empty

            # Find existing sheets with this base name
            existing_sheets = list(self.parent.Data['Core levels'].keys())
            existing_base_sheets = [s for s in existing_sheets if re.match(r'^' + re.escape(base_name) + r'\d*$', s)]

            # Find the earliest available row
            used_suffixes = set()
            for s in existing_base_sheets:
                suffix_match = re.search(r'^' + re.escape(base_name) + r'(\d+)$', s)
                if suffix_match:
                    used_suffixes.add(int(suffix_match.group(1)))
                elif s == base_name:  # Base without suffix is like having suffix 0
                    used_suffixes.add(0)

            # Find first unused number starting from 0
            suffix = 0
            while suffix in used_suffixes:
                suffix += 1

            # Create new name (don't add 0 suffix if base name is unused)
            if suffix == 0 and base_name not in existing_sheets:
                new_name = base_name
            else:
                new_name = f"{base_name}{suffix}"
        else:
            new_name = suggested_name

        data = self.parent.Data['Core levels'][sheet_name]
        x_values = np.array(data['B.E.'])
        mask = (x_values >= min_be) & (x_values <= max_be)

        # Create new sheet data
        new_data = {
            'B.E.': x_values[mask].tolist(),
            'Raw Data': np.array(data['Raw Data'])[mask].tolist(),
            'Background': {'Bkg Y': np.array(data['Background']['Bkg Y'])[mask].tolist()},
            'Name': new_name
        }

        # Update window.Data
        self.parent.Data['Core levels'][new_name] = new_data
        self.parent.Data['Number of Core levels'] += 1

        # Update Excel file
        import pandas as pd
        df = pd.DataFrame({
            'BE': new_data['B.E.'],
            'Raw Data': new_data['Raw Data'],
            'Background': new_data['Background']['Bkg Y'],
            'Transmission': np.ones(sum(mask)).tolist()
        })

        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=new_name, index=False)

        # Update JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        if os.path.exists(json_file_path):
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

        # Update sheet list and display
        self.parent.sheet_combobox.Append(new_name)
        self.parent.sheet_combobox.SetValue(new_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, new_name)

        # Close and reopen the file manager if it exists
        if hasattr(self.parent, 'file_manager') and self.parent.file_manager is not None:
            try:
                # Close existing file manager
                self.parent.file_manager.Close()
                self.parent.file_manager.Destroy()
                self.parent.file_manager = None

                # Reopen file manager
                import wx
                wx.CallAfter(self.parent.on_open_file_manager, None)
            except Exception as e:
                print(f"Error refreshing file manager: {e}")

        self.Close()

    def on_close(self, event):
        if self.vline_min:
            self.vline_min.remove()
        if self.vline_max:
            self.vline_max.remove()
        self.parent.canvas.draw_idle()
        self.Destroy()


class PlotModWindow(wx.Frame):
    def __init__(self, parent, *args, **kw):
        super().__init__(parent, *args, **kw, style=wx.DEFAULT_FRAME_STYLE & ~(
                wx.RESIZE_BORDER | wx.MAXIMIZE_BOX | wx.MINIMIZE_BOX | wx.SYSTEM_MENU) | wx.STAY_ON_TOP)

        self.SetTitle("Plot Modifications")
        self.SetSize(340, 450)

        self.parent = parent
        panel = wx.Panel(self)

        main_pos = parent.GetPosition()
        main_size = parent.GetSize()
        mod_size = self.GetSize()

        x = main_pos.x + (main_size.width - mod_size.width) // 2
        y = main_pos.y + (main_size.height - mod_size.height) // 2

        self.SetPosition((x, y))

        grid_sizer = wx.GridBagSizer(5, 5)

        # First row - Smoothing
        smooth_box = wx.StaticBox(panel, label="Smoothing")
        smooth_sizer = wx.StaticBoxSizer(smooth_box, wx.VERTICAL)

        self.smooth_method = wx.ComboBox(panel, choices=["Gaussian", "Savitzky-Golay", "Moving Average"],
                                         style=wx.CB_READONLY)
        self.smooth_method.SetValue("Gaussian")
        self.smooth_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)

        smooth_sizer.Add(wx.StaticText(panel, label="Method:"), 0, wx.ALL, 5)
        smooth_sizer.Add(self.smooth_method, 0, wx.EXPAND | wx.ALL, 5)
        smooth_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        smooth_sizer.Add(self.smooth_width, 0, wx.EXPAND | wx.ALL, 5)

        smooth_btn = wx.Button(panel, label="Apply Smoothing")
        smooth_btn.SetMinSize((125, 40))
        smooth_btn.Bind(wx.EVT_BUTTON, self.on_smooth)
        smooth_sizer.Add(smooth_btn, 0, wx.EXPAND | wx.ALL, 5)

        # First row - Differentiation
        diff_box = wx.StaticBox(panel, label="Differentiation")
        diff_sizer = wx.StaticBoxSizer(diff_box, wx.VERTICAL)

        self.diff_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)
        diff_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        diff_sizer.Add(self.diff_width, 0, wx.EXPAND | wx.ALL, 5)

        diff_btn = wx.Button(panel, label="Apply Differentiation")
        diff_btn.SetMinSize((125, 40))
        diff_btn.Bind(wx.EVT_BUTTON, self.on_differentiate)
        diff_sizer.Add(diff_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Second row - Integration
        int_box = wx.StaticBox(panel, label="Integration")
        int_sizer = wx.StaticBoxSizer(int_box, wx.VERTICAL)

        self.int_width = wx.SpinCtrl(panel, min=1, max=100, initial=5)
        int_sizer.Add(wx.StaticText(panel, label="Width:"), 0, wx.ALL, 5)
        int_sizer.Add(self.int_width, 0, wx.EXPAND | wx.ALL, 5)

        int_btn = wx.Button(panel, label="Apply Integration")
        int_btn.SetMinSize((125, 40))
        int_btn.Bind(wx.EVT_BUTTON, self.on_integrate)
        int_sizer.Add(int_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Constant Operation
        const_box = wx.StaticBox(panel, label="Constant Operation")
        const_sizer = wx.StaticBoxSizer(const_box, wx.VERTICAL)

        self.const_op = wx.ComboBox(panel, choices=["Multiply", "Divide", "Add", "Subtract"], style=wx.CB_READONLY)
        self.const_op.SetValue("Multiply")
        const_sizer.Add(self.const_op, 0, wx.EXPAND | wx.ALL, 5)

        self.const_value = wx.SpinCtrlDouble(panel, value="1.0", min=0.001, max=10000000.0, inc=0.1)
        const_sizer.Add(wx.StaticText(panel, label="Value:"), 0, wx.ALL, 5)
        const_sizer.Add(self.const_value, 0, wx.EXPAND | wx.ALL, 5)

        const_btn = wx.Button(panel, label="Apply Operation")
        const_btn.SetMinSize((125, 40))
        const_btn.Bind(wx.EVT_BUTTON, self.on_apply_constant)
        const_sizer.Add(const_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Add to grid
        grid_sizer.Add(smooth_sizer, pos=(0, 0), flag=wx.EXPAND | wx.ALL, border=5)
        grid_sizer.Add(diff_sizer, pos=(0, 1), flag=wx.EXPAND | wx.ALL, border=5)
        grid_sizer.Add(int_sizer, pos=(1, 0), flag=wx.EXPAND | wx.ALL, border=5)
        grid_sizer.Add(const_sizer, pos=(1, 1), flag=wx.EXPAND | wx.ALL, border=5)

        from libraries.ConfigFile import set_consistent_fonts
        set_consistent_fonts(self)

        panel.SetSizer(grid_sizer)
        self.Centre()

    # Method for constant operation
    def on_apply_constant(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        constant = self.const_value.GetValue()
        operation = self.const_op.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        # Apply the operation
        if operation == "Multiply":
            modified_y = [val * constant for val in y]
        elif operation == "Divide":
            modified_y = [val / constant for val in y]
        elif operation == "Add":
            modified_y = [val + constant for val in y]
        elif operation == "Subtract":
            modified_y = [val - constant for val in y]

        # Get the base name
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        if match:
            base_name = match.group(1)
        else:
            base_name = sheet_name

        # Find the earliest available row
        new_sheet_name = self.get_earliest_row_name(base_name)

        # Save the data
        self.save_modified_data(x, modified_y, new_sheet_name, f"{operation}d")

    def get_earliest_row_name(self, base_name):
        """
        Find the earliest available row for a core level.
        Returns the sheet name with appropriate row number.
        """
        import re

        # Check all sheets in parent data
        all_sheets = list(self.parent.Data['Core levels'].keys())

        # Check if base name without number exists (row 0)
        if base_name in all_sheets:
            # Base name exists, need to find next available
            rows_used = []
        else:
            # Base name doesn't exist, it's available
            return base_name

        # Pattern to match base name followed by optional number
        pattern = re.compile(f"^{re.escape(base_name)}(\\d+)$")

        # Collect all used row numbers
        for sheet in all_sheets:
            if sheet == base_name:  # This is row 0
                rows_used.append(0)
            else:
                match = pattern.match(sheet)
                if match:
                    rows_used.append(int(match.group(1)))

        # Find first unused row number
        for i in range(1000):  # Reasonable upper limit
            if i not in rows_used:
                if i == 0:
                    return base_name  # No suffix for row 0
                else:
                    return f"{base_name}{i}"

        # Fallback (unlikely to reach)
        return f"{base_name}{len(all_sheets)}"

    def save_modified_data(self, x, y, sheet_name, operation_type):
        """Save modified data to a new sheet and refresh file manager."""
        import pandas as pd
        import openpyxl

        # Create DataFrame with required columns
        df = pd.DataFrame({
            'BE': x,
            'Raw Data': y,
            'Background': y,
            'Transmission': [1.0] * len(x)
        })

        # Load workbook
        wb = openpyxl.load_workbook(self.parent.Data['FilePath'])

        # Save to Excel
        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Update window.Data
        self.parent.Data['Core levels'][sheet_name] = {
            'B.E.': x if isinstance(x, list) else x.tolist(),
            'Raw Data': y if isinstance(y, list) else y.tolist(),
            'Background': {'Bkg Y': y if isinstance(y, list) else y.tolist()},
            'Name': sheet_name
        }
        self.parent.Data['Number of Core levels'] += 1

        # Update JSON file
        json_file_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'
        if os.path.exists(json_file_path):
            from libraries.Save import convert_to_serializable_and_round
            json_data = convert_to_serializable_and_round(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

        # Close and reopen file manager if it exists
        if hasattr(self.parent, 'file_manager') and self.parent.file_manager is not None:
            try:
                # Close existing file manager
                self.parent.file_manager.Close()
                self.parent.file_manager.Destroy()
                self.parent.file_manager = None

                # Reopen file manager
                import wx
                wx.CallAfter(self.parent.on_open_file_manager, None)
            except Exception as e:
                print(f"Error refreshing file manager: {e}")
                pass

        # Update sheet list
        self.parent.sheet_combobox.Append(sheet_name)
        self.parent.sheet_combobox.SetValue(sheet_name)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, sheet_name)

    def on_smooth(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        method = self.smooth_method.GetValue()
        width = self.smooth_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        if method == "Gaussian":
            smoothed = gaussian_filter(y, width)
        elif method == "Savitzky-Golay":
            if width % 2 == 0:  # Ensure odd window length for savgol_filter
                width += 1
            smoothed = savgol_filter(y, width, 3)
        else:  # Moving Average
            kernel = np.ones(width) / width
            smoothed = np.convolve(y, kernel, mode='same')

        # Get the base name
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        if match:
            base_name = match.group(1)
        else:
            base_name = sheet_name

        # Find the earliest available row
        new_sheet_name = self.get_earliest_row_name(base_name)

        # Save the data
        self.save_modified_data(x, smoothed, new_sheet_name, "Smoothed")

    def on_differentiate(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        width = self.diff_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        derivative = np.gradient(y, x)

        # Ensure odd window length for savgol_filter
        if width % 2 == 0:
            width += 1

        smoothed_deriv = savgol_filter(derivative, width, 3)

        # Get the base name
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        if match:
            base_name = match.group(1)
        else:
            base_name = sheet_name

        # Find the earliest available row
        new_sheet_name = self.get_earliest_row_name(base_name)

        # Save the data
        self.save_modified_data(x, smoothed_deriv, new_sheet_name, "Differentiated")

    def on_integrate(self, event):
        sheet_name = self.parent.sheet_combobox.GetValue()
        width = self.int_width.GetValue()

        x = self.parent.Data['Core levels'][sheet_name]['B.E.']
        y = self.parent.Data['Core levels'][sheet_name]['Raw Data']

        integrated = cumtrapz(y, x, initial=0)

        # Ensure odd window length for savgol_filter
        if width % 2 == 0:
            width += 1

        smoothed_int = savgol_filter(integrated, width, 3)

        # Get the base name
        import re
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        if match:
            base_name = match.group(1)
        else:
            base_name = sheet_name

        # Find the earliest available row
        new_sheet_name = self.get_earliest_row_name(base_name)

        # Save the data
        self.save_modified_data(x, smoothed_int, new_sheet_name, "Integrated")


class JoinSheetsWindow(wx.Frame):
    def __init__(self, parent):
        super().__init__(parent, title="Join Sheets", size=(300, 400))
        self.parent = parent
        panel = wx.Panel(self)

        sizer = wx.BoxSizer(wx.VERTICAL)

        self.sheet_list = wx.CheckListBox(panel, choices=self.parent.sheet_combobox.GetStrings())
        sizer.Add(self.sheet_list, 1, wx.EXPAND | wx.ALL, 5)

        join_btn = wx.Button(panel, label="Join Sheets")
        join_btn.Bind(wx.EVT_BUTTON, self.on_join)
        sizer.Add(join_btn, 0, wx.EXPAND | wx.ALL, 5)

        panel.SetSizer(sizer)

    def on_join(self, event):
        selected_sheets = self.sheet_list.GetCheckedStrings()
        if len(selected_sheets) < 2:
            wx.MessageBox("Select at least 2 sheets to join", "Error")
            return

        # Sort sheets by highest BE
        sheet_max_be = {sheet: max(self.parent.Data['Core levels'][sheet]['B.E.']) for sheet in selected_sheets}
        sorted_sheets = sorted(selected_sheets, key=lambda x: sheet_max_be[x], reverse=True)

        # Join data
        joined_be = []
        joined_data = []

        for sheet in sorted_sheets:
            be = self.parent.Data['Core levels'][sheet]['B.E.']
            data = self.parent.Data['Core levels'][sheet]['Raw Data']
            joined_be.extend(be)
            joined_data.extend(data)

        # Create new sheet
        new_sheet = "Joined_Scan"
        self.parent.Data['Core levels'][new_sheet] = {
            'B.E.': joined_be,
            'Raw Data': joined_data,
            'Background': {'Bkg Y': joined_data},
            'Transmission': [1.0] * len(joined_be)
        }

        # Update Excel file
        df = pd.DataFrame({
            'BE': joined_be,
            'Raw Data': joined_data,
            'Background': joined_data,
            'Transmission': [1.0] * len(joined_be)
        })

        with pd.ExcelWriter(self.parent.Data['FilePath'], engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=new_sheet, index=False)

        # Update UI
        self.parent.sheet_combobox.Append(new_sheet)
        self.parent.sheet_combobox.SetValue(new_sheet)
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, new_sheet)

        self.Close()

def perform_auto_backup(parent):
    """Create an automatic backup of the current data files"""
    if 'FilePath' not in parent.Data or not parent.Data['FilePath']:
        print("No file open, skipping auto backup")
        return

    # Get current file paths
    excel_file = parent.Data['FilePath']
    json_file = os.path.splitext(excel_file)[0] + '.json'

    # Check if files exist
    if not os.path.exists(excel_file):
        print(f"Excel file not found: {excel_file}, skipping auto backup")
        return

    # Create backup folder in the executable directory
    executable_dir = os.path.dirname(os.path.abspath(sys.executable))
    # For development environment, fall back to current script directory
    if not "KherveFitting" in executable_dir:
        executable_dir = os.path.dirname(os.path.abspath(__file__))
        # Go up one level if in libraries folder
        if os.path.basename(executable_dir) == "libraries":
            executable_dir = os.path.dirname(executable_dir)

    backup_folder = os.path.join(executable_dir, "Backup")
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)

    # Generate timestamp in format: YYcDD_HHMMSS
    import datetime
    now = datetime.datetime.now()
    month_letter = chr(ord('a') + now.month - 1)  # a=Jan, b=Feb, c=Mar, etc.
    timestamp = f"{now.year % 100}{month_letter}{now.day:02d}_{now.hour:02d}{now.minute:02d}{now.second:02d}"

    # Create backup filenames
    excel_filename = os.path.basename(excel_file)
    excel_name_no_ext = os.path.splitext(excel_filename)[0]
    excel_backup = os.path.join(backup_folder, f"{excel_name_no_ext}_{timestamp}.xlsx")

    # Copy the Excel file
    try:
        shutil.copy2(excel_file, excel_backup)

        # Copy the JSON file if it exists
        if os.path.exists(json_file):
            json_filename = os.path.basename(json_file)
            json_name_no_ext = os.path.splitext(json_filename)[0]
            json_backup = os.path.join(backup_folder, f"{json_name_no_ext}_{timestamp}.json")
            shutil.copy2(json_file, json_backup)

        print(f"Auto backup completed at {datetime.datetime.now().strftime('%H:%M:%S')}")
        return True
    except Exception as e:
        print(f"Auto backup error: {str(e)}")
        return False


def sort_excel_sheets(window):
    """Sort Excel sheets by sample group and element name without modifying data"""
    from libraries.Save import save_state
    save_state(window)

    if not hasattr(window, 'Data') or 'Core levels' not in window.Data:
        wx.MessageBox("No data available to sort.", "Error", wx.OK | wx.ICON_ERROR)
        return

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Sorting Sheets", size=(300, 350))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 300) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console("Starting sheet sorting process...")

        file_path = window.Data['FilePath']

        # Check if file is accessible
        update_console("Checking file accessibility...")
        try:
            with open(file_path, 'rb') as f:
                pass
        except PermissionError:
            update_console("Error: Excel file is open in another program.")
            wx.CallLater(2000, console_frame.Close)
            wx.MessageBox("Cannot sort sheets: Excel file is open in another program.",
                          "File Locked", wx.OK | wx.ICON_ERROR)
            return

        import openpyxl
        import re

        # Load workbook
        update_console("Loading Excel workbook...")
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames.copy()
        update_console(f"Found {len(sheet_names)} sheets total")

        # Separate special sheets from regular sheets
        update_console("Identifying special sheets...")
        regular_sheets = []
        results_table_sheet = None
        experimental_desc_sheet = None

        for sheet_name in sheet_names:
            if sheet_name == "Results Table":
                results_table_sheet = sheet_name
                update_console("Found Results Table sheet")
            elif sheet_name == "Experimental Description":
                experimental_desc_sheet = sheet_name
                update_console("Found Experimental Description sheet")
            else:
                regular_sheets.append(sheet_name)

        update_console(f"Processing {len(regular_sheets)} regular data sheets...")

        # Group regular sheets by sample number
        update_console("Analyzing sheet names and grouping by sample...")
        grouped_sheets = {}
        for sheet_name in regular_sheets:
            if "wide" in sheet_name.lower() or "survey" in sheet_name.lower():
                match = re.match(r'(wide|survey)(\d*)$', sheet_name.lower(), re.IGNORECASE)
                if match:
                    base_name = match.group(1).capitalize()
                    sample_num = match.group(2)
                else:
                    base_name = sheet_name
                    sample_num = ""
            else:
                match = re.match(r'([A-Za-z]+\d*[spdfg]*)(\d*)$', sheet_name)
                if match:
                    base_name, sample_num = match.groups()
                else:
                    base_name = sheet_name
                    sample_num = ""

            sample_num = int(sample_num) if sample_num else 0

            if sample_num not in grouped_sheets:
                grouped_sheets[sample_num] = []
            grouped_sheets[sample_num].append((base_name, sheet_name))

        update_console(f"Grouped sheets into {len(grouped_sheets)} sample groups")

        # Sort each group
        update_console("Sorting sheets within each group...")
        for sample_num in grouped_sheets:
            def sort_key(item):
                base = item[0].lower()
                if "wide" in base or "survey" in base:
                    return "zzz"
                return base

            grouped_sheets[sample_num].sort(key=sort_key)

        # Create sorted list - regular sheets first
        sorted_sheet_names = []
        for sample_num in sorted(grouped_sheets.keys()):
            for _, sheet_name in grouped_sheets[sample_num]:
                sorted_sheet_names.append(sheet_name)

        # Add special sheets at the end
        if results_table_sheet:
            sorted_sheet_names.append(results_table_sheet)
        if experimental_desc_sheet:
            sorted_sheet_names.append(experimental_desc_sheet)

        # Check if already sorted
        if sheet_names == sorted_sheet_names:
            update_console("Sheets are already sorted - no changes needed.")
            wx.CallLater(1000, console_frame.Close)
            return

        # Reorder sheets using openpyxl
        update_console("Reordering sheets in Excel file...")
        for i, sheet_name in enumerate(sorted_sheet_names):
            update_console(f"Moving sheet {i + 1}/{len(sorted_sheet_names)}: {sheet_name}")
            sheet = wb[sheet_name]
            wb.move_sheet(sheet, offset=i - wb.index(sheet))

        # Save workbook
        update_console("Saving Excel file...")
        wb.save(file_path)
        wb.close()

        # Update Data structure order to match (only for core levels)
        update_console("Updating data structure...")
        sorted_core_levels = {}
        for sheet_name in sorted_sheet_names:
            if sheet_name in window.Data['Core levels']:
                sorted_core_levels[sheet_name] = window.Data['Core levels'][sheet_name]
        window.Data['Core levels'] = sorted_core_levels

        # Update UI
        update_console("Updating interface...")
        current_sheet = window.sheet_combobox.GetValue()
        window.sheet_combobox.Clear()
        for sheet_name in sorted_sheet_names:
            if sheet_name in window.Data['Core levels']:  # Only add core level sheets to combobox
                window.sheet_combobox.Append(sheet_name)

        if current_sheet in window.Data['Core levels'] and current_sheet in [window.sheet_combobox.GetString(i) for i in
                                                                             range(window.sheet_combobox.GetCount())]:
            window.sheet_combobox.SetValue(current_sheet)
        elif window.sheet_combobox.GetCount() > 0:
            window.sheet_combobox.SetValue(window.sheet_combobox.GetString(0))
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, window.sheet_combobox.GetString(0))

        update_console("Sheet sorting completed successfully!")
        wx.CallLater(500, console_frame.Close)

    except Exception as e:
        update_console(f"Error during sorting: {str(e)}")
        wx.CallLater(2000, console_frame.Close)
        wx.MessageBox(f"Error sorting sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
