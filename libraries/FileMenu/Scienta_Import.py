# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
Scienta Analyzer File Import Module for KherveFitting

Supports two types of Scienta .txt files:
1. Map files: Multiple regions with heat map data (BE x Sweeps)
2. Single-sweep files: Multiple regions with single spectrum each

Features:
- Map preview with scrollable heatmaps (Greens colormap)
- Drop individual sweeps
- Bin sweeps into N groups
- Sum all sweeps
- Save maps as ~Map sheets
- Direct import for single-sweep files
"""

import os
import re
import numpy as np
import wx
import json
import openpyxl

try:
    import h5py
    HAS_H5PY = True
except ImportError:
    HAS_H5PY = False

import matplotlib
matplotlib.use('WXAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.backends.backend_wxagg import NavigationToolbar2WxAgg as NavigationToolbar
from matplotlib.figure import Figure


def clean_region_name(region_name):
    """
    Clean and normalize region name from Scienta file.
    Handles:
    - Standard core levels: C1s, O1s, Sr3d, Ce3d, etc.
    - Multi-word core levels: Ce 3d, Ce_3d, Ce 3d 1
    - Auger lines: C KLL, O KLL, Fe LMM, etc.
    - Special scans: Wide, Survey, Valence, Valence Band, Fermi
    - Suffixes to remove: _fast, _fast2, _slow, _quick, etc.

    Returns:
        tuple: (clean_name, is_special)
        - clean_name: The cleaned region name
        - is_special: True if this is a special scan type (Wide, Survey, Valence, Auger)
    """
    # First, remove common acquisition suffixes
    suffixes_to_remove = [
        '_fast2', '_fast', '_slow', '_quick', '_scan', '_Scan', '_SCAN',
        ' fast2', ' fast', ' slow', ' quick', ' scan', ' Scan', ' SCAN',
        '_region', '_Region', '_REGION', ' region', ' Region', ' REGION',
        '_spectrum', '_Spectrum', ' spectrum', ' Spectrum'
    ]

    name = region_name.strip()
    for suffix in suffixes_to_remove:
        if name.endswith(suffix):
            name = name[:-len(suffix)]
            break

    name = name.strip()
    lower_name = name.lower()

    # Check for special scan types
    if any(x in lower_name for x in ['survey', 'wide scan', 'wide']):
        if 'survey' in lower_name:
            return 'Survey', True
        else:
            return 'Wide', True

    if any(x in lower_name for x in ['valence band', 'valence_band', 'valenceband', 'valence']):
        return 'VB', True

    if 'fermi' in lower_name:
        return 'Fermi', True

    # Check for Auger lines
    auger_patterns = [
        r'([A-Z][a-z]?)\s*(KLL\d*)',
        r'([A-Z][a-z]?)\s*(LMM\d*)',
        r'([A-Z][a-z]?)\s*(MNN\d*)',
        r'([A-Z][a-z]?)\s*(MVV\d*)',
        r'([A-Z][a-z]?)\s*(MNV\d*)',
        r'([A-Z][a-z]?)\s*(NOO\d*)',
        r'([A-Z][a-z]?)\s*(KL\d+)',
        r'([A-Z][a-z]?)\s*(LM\d+)',
        r'([A-Z][a-z]?)\s*(MN\d+)',
    ]

    for pattern in auger_patterns:
        match = re.match(pattern, name, re.IGNORECASE)
        if match:
            element = match.group(1)
            auger_type = match.group(2).upper()
            return f"{element}{auger_type}", True

    # Core levels with space/underscore
    core_level_patterns = [
        r'^([A-Z][a-z]?)\s+(\d+[spdfgh])(\d*)$',
        r'^([A-Z][a-z]?)_(\d+[spdfgh])(\d*)$',
        r'^([A-Z][a-z]?)\s+(\d+[spdfgh])\s+(\d+)$',
        r'^([A-Z][a-z]?)_(\d+[spdfgh])_(\d+)$',
    ]

    for pattern in core_level_patterns:
        match = re.match(pattern, name)
        if match:
            element = match.group(1)
            orbital = match.group(2)
            return f"{element}{orbital}", False

    # Standard core level pattern
    standard_pattern = r'^([A-Z][a-z]?\d+[spdfgh])(\d*)$'
    match = re.match(standard_pattern, name)
    if match:
        core_level = match.group(1)
        return core_level, False

    # Fallback
    base_match = re.match(r'^(.+?)(\d*)$', name)
    if base_match and base_match.group(1):
        return base_match.group(1).strip(), False

    return name, False


def is_auger_line(name):
    """Check if a region name is an Auger line."""
    lower_name = name.lower()

    if any(lower_name.endswith(x) for x in ['kll', 'mnn', 'mvv', 'mnv', 'lmm', 'noo']):
        return True

    auger_patterns = [
        r'kll\d*', r'kl\d+', r'lmm\d*', r'lm\d+',
        r'mnn\d*', r'mn\d+', r'mvv\d*', r'mv\d+',
        r'mnv\d*', r'noo\d*', r'no\d+'
    ]

    for pattern in auger_patterns:
        if re.search(pattern, lower_name):
            return True

    return False


def parse_scienta_file(file_path):
    """
    Parse a Scienta .txt file and extract all regions with their data and metadata.
    """
    with open(file_path, 'r') as f:
        content = f.read()

    lines = content.split('\n')

    result = {
        'num_regions': 0,
        'version': '',
        'regions': [],
        'file_path': file_path,
        'is_map': False
    }

    # Parse [Info] section
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line == '[Info]':
            i += 1
            while i < len(lines) and not lines[i].strip().startswith('['):
                info_line = lines[i].strip()
                if info_line.startswith('Number of Regions='):
                    result['num_regions'] = int(info_line.split('=')[1])
                elif info_line.startswith('Version='):
                    result['version'] = info_line.split('=')[1]
                i += 1
            break
        i += 1

    # Parse each region
    for region_num in range(1, result['num_regions'] + 1):
        region_data = parse_region(lines, region_num)
        if region_data:
            result['regions'].append(region_data)
            if region_data.get('sweeps', 1) > 1 and region_data.get('data_2d') is not None:
                result['is_map'] = True

    return result


def parse_region(lines, region_num):
    """Parse a single region from the file."""
    region = {
        'name': '',
        'be_values': None,
        'sweeps': 1,
        'sweep_scale': None,
        'data_2d': None,
        'intensities': None,
        'metadata': {}
    }

    region_header = f'[Region {region_num}]'
    info_header = f'[Info {region_num}]'
    data_header = f'[Data {region_num}]'

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        if line == region_header:
            i += 1
            while i < len(lines) and not lines[i].strip().startswith('['):
                region_line = lines[i].strip()
                if region_line.startswith('Region Name='):
                    region['name'] = region_line.split('=')[1]
                elif region_line.startswith('Dimension 1 size='):
                    region['dim1_size'] = int(region_line.split('=')[1])
                elif region_line.startswith('Dimension 1 scale='):
                    scale_str = region_line.split('=')[1]
                    region['be_values'] = np.array([float(x) for x in scale_str.split()])
                elif region_line.startswith('Dimension 2 size='):
                    region['sweeps'] = int(region_line.split('=')[1])
                elif region_line.startswith('Dimension 2 scale='):
                    scale_str = region_line.split('=')[1]
                    region['sweep_scale'] = np.array([int(x) for x in scale_str.split()])
                i += 1

        elif line == info_header:
            i += 1
            while i < len(lines) and not lines[i].strip().startswith('['):
                info_line = lines[i].strip()
                if '=' in info_line:
                    key, value = info_line.split('=', 1)
                    region['metadata'][key] = value
                i += 1

        elif line == data_header:
            i += 1
            data_rows = []
            while i < len(lines) and not lines[i].strip().startswith('[') and lines[i].strip():
                data_line = lines[i].strip()
                if data_line:
                    values = [float(x) for x in data_line.split()]
                    data_rows.append(values)
                i += 1

            if data_rows:
                data_array = np.array(data_rows)
                if data_array.shape[1] == 2:
                    region['be_values'] = data_array[:, 0]
                    region['intensities'] = data_array[:, 1]
                    region['sweeps'] = 1
                else:
                    if region['be_values'] is None:
                        region['be_values'] = data_array[:, 0]
                    region['data_2d'] = data_array[:, 1:].T
        else:
            i += 1

    return region if (region['data_2d'] is not None or region['intensities'] is not None) else None


class ScientaMapPreviewWindow(wx.Frame):
    """
    Scrollable window to preview Scienta map data with heat maps for all regions.
    """

    def __init__(self, parent, parsed_data, callback_on_import):
        # Calculate window size
        num_regions = len(parsed_data['regions'])
        cols = 3
        rows = (num_regions + cols - 1) // cols

        # Fixed window size with scrolling
        width = min(2000, 360 * cols + 50)
        height = min(1000, 320 * 2 + 120)

        super().__init__(parent, title=f"Scienta Omicron Map Preview - {os.path.basename(parsed_data['file_path'])}",
                        size=(width, height), style=wx.DEFAULT_FRAME_STYLE)

        self.parent = parent
        self.parsed_data = parsed_data
        self.callback_on_import = callback_on_import
        self.regions = [r.copy() for r in parsed_data['regions']]

        # Deep copy the 2D data arrays
        for i, region in enumerate(self.regions):
            if parsed_data['regions'][i].get('data_2d') is not None:
                region['data_2d'] = np.copy(parsed_data['regions'][i]['data_2d'])
            region['dropped_sweeps'] = []

        self.current_sweep_index = 0
        self.max_sweeps = max(r.get('sweeps', 1) for r in self.regions) if self.regions else 1

        self.init_ui()
        self.update_heatmaps()

        self.Centre()
        self.Show()

    def init_ui(self):
        """Initialize the user interface with scrollable panel."""
        main_panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # --- Region name editor panel ---
        name_panel = wx.Panel(main_panel)
        name_sizer = wx.BoxSizer(wx.HORIZONTAL)

        name_label = wx.StaticText(name_panel, label="Names:")
        name_label.SetFont(wx.Font(8, wx.FONTFAMILY_DEFAULT,
                                   wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        name_sizer.Add(name_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT | wx.RIGHT, 6)

        self.name_fields = []
        for region in self.regions:
            proposed, _ = clean_region_name(region['name'])
            if not proposed:
                proposed = region['name']
            col_sizer = wx.BoxSizer(wx.VERTICAL)
            # Grey label showing original name
            orig_lbl = wx.StaticText(name_panel, label=region['name'],
                                     style=wx.ST_ELLIPSIZE_END)
            orig_lbl.SetForegroundColour(wx.Colour(130, 130, 130))
            orig_lbl.SetFont(wx.Font(7, wx.FONTFAMILY_DEFAULT,
                                     wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_NORMAL))
            # Editable proposed name
            tf = wx.TextCtrl(name_panel, value=proposed, size=(78, -1))
            tf.SetToolTip(f"Original name: {region['name']}\n"
                          f"Proposed: {proposed}\n"
                          f"Edit to change the sheet name used on import.")
            col_sizer.Add(orig_lbl, 0, wx.EXPAND | wx.BOTTOM, 1)
            col_sizer.Add(tf, 0, wx.EXPAND)
            name_sizer.Add(col_sizer, 0, wx.ALL, 3)
            self.name_fields.append(tf)

        name_panel.SetSizer(name_sizer)
        main_sizer.Add(name_panel, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, 4)

        # Thin separator line
        sep = wx.StaticLine(main_panel)
        main_sizer.Add(sep, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 4)

        # Control panel at top (fixed)
        control_panel = wx.Panel(main_panel)
        control_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Sweep selector
        sweep_label = wx.StaticText(control_panel, label="Sweep:")
        control_sizer.Add(sweep_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        self.sweep_spin = wx.SpinCtrl(control_panel, min=1, max=self.max_sweeps, initial=1, size=(60, -1))
        self.sweep_spin.Bind(wx.EVT_SPINCTRL, self.on_sweep_changed)
        control_sizer.Add(self.sweep_spin, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        sweep_info = wx.StaticText(control_panel, label=f"/ {self.max_sweeps}")
        control_sizer.Add(sweep_info, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)

        # Drop button
        self.drop_btn = wx.Button(control_panel, label="Drop", size=(50, -1))
        self.drop_btn.SetBackgroundColour(wx.Colour(255, 200, 200))
        self.drop_btn.Bind(wx.EVT_BUTTON, self.on_drop_sweep)
        control_sizer.Add(self.drop_btn, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        # Undo button
        self.undo_btn = wx.Button(control_panel, label="Undo", size=(50, -1))
        self.undo_btn.Bind(wx.EVT_BUTTON, self.on_undo_drop)
        self.undo_btn.Enable(False)
        control_sizer.Add(self.undo_btn, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        control_sizer.AddSpacer(15)

        # Binning controls
        bin_label = wx.StaticText(control_panel, label="Bins:")
        control_sizer.Add(bin_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        self.bin_spin = wx.SpinCtrl(control_panel, min=1, max=self.max_sweeps, initial=1, size=(60, -1))
        control_sizer.Add(self.bin_spin, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        self.bin_btn = wx.Button(control_panel, label="Bin && Import", size=(90, -1))
        self.bin_btn.SetBackgroundColour(wx.Colour(200, 200, 255))
        self.bin_btn.Bind(wx.EVT_BUTTON, self.on_bin_import)
        control_sizer.Add(self.bin_btn, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        control_sizer.AddSpacer(15)

        # Checkbox for saving maps
        self.save_maps_checkbox = wx.CheckBox(control_panel, label="Save Maps")
        self.save_maps_checkbox.SetValue(False)  # Default: do not save maps
        control_sizer.Add(self.save_maps_checkbox, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        control_sizer.AddStretchSpacer()

        # Status text
        self.status_text = wx.StaticText(control_panel, label="")
        control_sizer.Add(self.status_text, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 3)

        control_sizer.AddStretchSpacer()

        # SUM button
        self.sum_btn = wx.Button(control_panel, label="SUM && Import", size=(100, -1))
        self.sum_btn.SetBackgroundColour(wx.Colour(200, 255, 200))
        self.sum_btn.SetFont(wx.Font(10, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        self.sum_btn.Bind(wx.EVT_BUTTON, self.on_sum_sweeps)
        control_sizer.Add(self.sum_btn, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        control_panel.SetSizer(control_sizer)
        main_sizer.Add(control_panel, 0, wx.EXPAND | wx.ALL, 3)

        # Scrolled window for the plots
        self.scroll_panel = wx.ScrolledWindow(main_panel, style=wx.VSCROLL | wx.HSCROLL)
        self.scroll_panel.SetScrollRate(20, 20)

        scroll_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create figure for heatmaps
        num_regions = len(self.regions)
        cols = 3
        rows = (num_regions + cols - 1) // cols

        # Calculate figure size based on number of rows
        fig_height = max(2, rows * 3.2)
        fig_width = cols * 3.6

        self.figure = Figure(figsize=(fig_width, fig_height), dpi=100)
        self.canvas = FigureCanvas(self.scroll_panel, -1, self.figure)

        # Set canvas size for scrolling
        canvas_height = int(fig_height * 100)
        canvas_width = int(fig_width * 100)
        self.canvas.SetMinSize((canvas_width, canvas_height))

        scroll_sizer.Add(self.canvas, 0, wx.EXPAND | wx.ALL, 3)

        self.scroll_panel.SetSizer(scroll_sizer)
        self.scroll_panel.FitInside()

        main_sizer.Add(self.scroll_panel, 1, wx.EXPAND | wx.ALL, 3)

        main_panel.SetSizer(main_sizer)

        self.axes = []
        self.heatmaps = []
        self.sweep_lines = []

        self.update_status()

    def update_heatmaps(self):
        """Update all heatmaps with current data."""
        self.figure.clear()
        self.axes = []
        self.heatmaps = []
        self.sweep_lines = []

        num_regions = len(self.regions)
        cols = 3
        rows = (num_regions + cols - 1) // cols

        for idx, region in enumerate(self.regions):
            ax = self.figure.add_subplot(rows, cols, idx + 1)
            self.axes.append(ax)

            if region.get('data_2d') is None:
                ax.text(0.5, 0.5, 'Single\nSweep', ha='center', va='center', transform=ax.transAxes)
                ax.set_title(region['name'], fontsize=8)
                self.heatmaps.append(None)
                self.sweep_lines.append(None)
                continue

            # Get data - mask dropped sweeps with NaN
            data = region['data_2d'].copy()
            for dropped_idx in region.get('dropped_sweeps', []):
                data[dropped_idx, :] = np.nan

            be_values = region['be_values']
            num_sweeps = region['sweeps']
            be_min, be_max = be_values.min(), be_values.max()

            # Check if BE is in descending order (typical for XPS)
            be_descending = be_values[0] > be_values[-1] if len(be_values) > 1 else False

            # Plot heatmap with Greens colormap
            if be_descending:
                # BE already goes high to low, plot normally
                im = ax.imshow(data, aspect='auto', origin='lower',
                               extent=[be_values[0], be_values[-1], -0.5, num_sweeps - 0.5],
                               cmap='Greens')
            else:
                # BE goes low to high, flip data for display
                data_flipped = np.fliplr(data)
                im = ax.imshow(data_flipped, aspect='auto', origin='lower',
                               extent=[be_max, be_min, -0.5, num_sweeps - 0.5],
                               cmap='Greens')
            self.heatmaps.append(im)

            # Set Y-axis limits to start at 0
            ax.set_ylim(0, num_sweeps)

            # Add horizontal line at current sweep
            line = ax.axhline(y=self.current_sweep_index + 0.5, color='red',
                              linewidth=1.5, linestyle='--')
            self.sweep_lines.append(line)

            ax.set_title(region['name'], fontsize=8)
            ax.set_xlabel('BE (eV)', fontsize=7)
            ax.set_ylabel('Sweep', fontsize=7)
            ax.tick_params(labelsize=6)

        self.figure.tight_layout(pad=0.5)
        self.canvas.draw()

    def update_sweep_lines(self):
        """Update just the sweep indicator lines."""
        for line in self.sweep_lines:
            if line is not None:
                line.set_ydata([self.current_sweep_index + 0.5, self.current_sweep_index + 0.5])
        self.canvas.draw_idle()

    def update_status(self):
        """Update the status text."""
        total_dropped = sum(len(r.get('dropped_sweeps', [])) for r in self.regions)
        if total_dropped > 0:
            self.status_text.SetLabel(f"Dropped: {total_dropped}")
            self.undo_btn.Enable(True)
        else:
            self.status_text.SetLabel("")
            self.undo_btn.Enable(False)

    def on_sweep_changed(self, event):
        self.current_sweep_index = self.sweep_spin.GetValue() - 1
        self.update_sweep_lines()

    def on_drop_sweep(self, event):
        sweep_idx = self.current_sweep_index

        for region in self.regions:
            if region.get('data_2d') is not None:
                if sweep_idx < region['sweeps'] and sweep_idx not in region.get('dropped_sweeps', []):
                    region['dropped_sweeps'].append(sweep_idx)

        self.update_heatmaps()
        self.update_status()

        if self.current_sweep_index < self.max_sweeps - 1:
            self.current_sweep_index += 1
            self.sweep_spin.SetValue(self.current_sweep_index + 1)

    def on_undo_drop(self, event):
        last_dropped = None
        for region in self.regions:
            if region.get('dropped_sweeps'):
                if last_dropped is None:
                    last_dropped = region['dropped_sweeps'][-1]

        if last_dropped is not None:
            for region in self.regions:
                if last_dropped in region.get('dropped_sweeps', []):
                    region['dropped_sweeps'].remove(last_dropped)

            self.current_sweep_index = last_dropped
            self.sweep_spin.SetValue(self.current_sweep_index + 1)
            self.update_heatmaps()
            self.update_status()

    def on_bin_import(self, event):
        """Bin sweeps into N groups and import."""
        num_bins = self.bin_spin.GetValue()
        save_maps = self.save_maps_checkbox.GetValue()

        # Apply user-edited names
        for i, (region, tf) in enumerate(zip(self.regions, self.name_fields)):
            new_name = tf.GetValue().strip()
            if new_name:
                region['name'] = new_name

        binned_data = {
            'file_path': self.parsed_data['file_path'],
            'regions': [],
            'bin_mode': True,
            'num_bins': num_bins,
            'save_maps': save_maps
        }

        for region in self.regions:
            if region.get('data_2d') is None:
                binned_region = {
                    'name': region['name'],
                    'be_values': region['be_values'],
                    'binned_intensities': [region['intensities']],
                    'metadata': region['metadata'],
                    'bin_indices': [[0]],
                    'total_sweeps': 1,
                    'data_2d': None
                }
                binned_data['regions'].append(binned_region)
                continue

            dropped = set(region.get('dropped_sweeps', []))
            valid_indices = [i for i in range(region['sweeps']) if i not in dropped]

            if not valid_indices:
                wx.MessageBox(f"Region '{region['name']}' has all sweeps dropped!",
                             "Error", wx.OK | wx.ICON_ERROR)
                return

            n_valid = len(valid_indices)
            bin_size = n_valid // num_bins
            remainder = n_valid % num_bins

            binned_intensities = []
            bin_indices_list = []

            start = 0
            for b in range(num_bins):
                size = bin_size + (1 if b < remainder else 0)
                if size == 0:
                    continue
                end = start + size
                bin_indices = valid_indices[start:end]
                bin_indices_list.append(bin_indices)

                bin_data = region['data_2d'][bin_indices, :]
                summed = np.sum(bin_data, axis=0) / len(bin_indices)
                binned_intensities.append(summed)

                start = end

            binned_region = {
                'name': region['name'],
                'be_values': region['be_values'],
                'binned_intensities': binned_intensities,
                'metadata': region['metadata'],
                'bin_indices': bin_indices_list,
                'total_sweeps': region['sweeps'],
                'dropped_sweeps': list(dropped),
                'data_2d': region['data_2d']  # Include original map data
            }
            binned_data['regions'].append(binned_region)

        self.Close()
        self.callback_on_import(binned_data)

    def on_sum_sweeps(self, event):
        """Sum all remaining sweeps and trigger import."""
        save_maps = self.save_maps_checkbox.GetValue()

        # Apply user-edited names
        for i, (region, tf) in enumerate(zip(self.regions, self.name_fields)):
            new_name = tf.GetValue().strip()
            if new_name:
                region['name'] = new_name

        summed_data = {
            'file_path': self.parsed_data['file_path'],
            'regions': [],
            'bin_mode': False,
            'save_maps': save_maps
        }

        for region in self.regions:
            if region.get('data_2d') is None:
                summed_region = {
                    'name': region['name'],
                    'be_values': region['be_values'],
                    'intensities': region['intensities'],
                    'metadata': region['metadata'],
                    'num_sweeps_summed': 1,
                    'total_sweeps': 1,
                    'dropped_sweeps': [],
                    'data_2d': None
                }
                summed_data['regions'].append(summed_region)
                continue

            dropped = set(region.get('dropped_sweeps', []))
            valid_indices = [i for i in range(region['sweeps']) if i not in dropped]

            if not valid_indices:
                wx.MessageBox(f"Region '{region['name']}' has all sweeps dropped!",
                             "Error", wx.OK | wx.ICON_ERROR)
                return

            valid_data = region['data_2d'][valid_indices, :]
            summed_intensities = np.sum(valid_data, axis=0) / len(valid_indices)

            summed_region = {
                'name': region['name'],
                'be_values': region['be_values'],
                'intensities': summed_intensities,
                'metadata': region['metadata'],
                'num_sweeps_summed': len(valid_indices),
                'total_sweeps': region['sweeps'],
                'dropped_sweeps': list(dropped),
                'data_2d': region['data_2d']  # Include original map data
            }
            summed_data['regions'].append(summed_region)

        self.Close()
        self.callback_on_import(summed_data)


def import_scienta_map(window):
    """Import Scienta map file with preview window."""
    with wx.FileDialog(window, "Open Scienta Map File",
                       wildcard="Scienta files (*.txt)|*.txt|All files (*.*)|*.*",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dialog:

        if dialog.ShowModal() == wx.ID_CANCEL:
            return

        file_path = dialog.GetPath()

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Loading Scienta File", size=(400, 200))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 400) // 2,
        parent_pos.y + (parent_size.height - 200) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Parsing: {os.path.basename(file_path)}")

        parsed_data = parse_scienta_file(file_path)

        update_console(f"Found {parsed_data['num_regions']} region(s)")
        for region in parsed_data['regions']:
            sweeps = region.get('sweeps', 1)
            update_console(f"  - {region['name']}: {sweeps} sweep(s), {len(region['be_values'])} points")

        if not parsed_data['is_map']:
            update_console("\nSingle-sweep file detected - use 'Import Scienta File' instead")
            console_frame.Close()
            wx.MessageBox("This file contains single-sweep data.\nUse 'Import Scienta File' for direct import.",
                         "Info", wx.OK | wx.ICON_INFORMATION)
            return

        update_console("\nOpening preview window...")
        console_frame.Close()

        def on_import_complete(data):
            finalize_scienta_import(window, data)

        preview = ScientaMapPreviewWindow(window, parsed_data, on_import_complete)

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error parsing Scienta file:\n\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_scienta_file(window):
    """Import Scienta single-sweep file directly (no preview)."""
    with wx.FileDialog(window, "Open Scienta File",
                       wildcard="Scienta files (*.txt)|*.txt|All files (*.*)|*.*",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dialog:

        if dialog.ShowModal() == wx.ID_CANCEL:
            return

        file_path = dialog.GetPath()

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing Scienta File", size=(400, 200))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 400) // 2,
        parent_pos.y + (parent_size.height - 200) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Parsing: {os.path.basename(file_path)}")

        parsed_data = parse_scienta_file(file_path)

        update_console(f"Found {parsed_data['num_regions']} region(s)")

        if parsed_data['is_map']:
            update_console("\nMap file detected - use 'Import Scienta Map' instead")
            console_frame.Close()
            wx.MessageBox("This file contains map data (multiple sweeps).\nUse 'Import Scienta Map' for preview and import.",
                         "Info", wx.OK | wx.ICON_INFORMATION)
            return

        import_data = {
            'file_path': file_path,
            'regions': [],
            'bin_mode': False,
            'save_maps': False
        }

        for region in parsed_data['regions']:
            import_region = {
                'name': region['name'],
                'be_values': region['be_values'],
                'intensities': region['intensities'],
                'metadata': region['metadata'],
                'num_sweeps_summed': 1,
                'total_sweeps': 1,
                'dropped_sweeps': [],
                'data_2d': None
            }
            import_data['regions'].append(import_region)
            update_console(f"  - {region['name']}: {len(region['be_values'])} points")

        console_frame.Close()
        finalize_scienta_import(window, import_data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing Scienta file:\n\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def finalize_scienta_import(window, import_data):
    """
    Finalize the import after summing/binning sweeps.
    Creates Excel file, JSON, and populates window.Data.
    Also saves map data as ~Map sheets.
    """
    from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
    from libraries.FileMenu.Save import update_undo_redo_state, save_state, convert_to_serializable_and_round
    from libraries.Sheet_Operations import on_sheet_selected

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing Scienta Data", size=(400, 250))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 400) // 2,
        parent_pos.y + (parent_size.height - 250) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        input_file = import_data['file_path']
        base_name = os.path.splitext(input_file)[0]
        excel_path = base_name + '.xlsx'
        json_path = base_name + '.json'

        update_console(f"Creating Excel file: {os.path.basename(excel_path)}")

        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        sheet_names = []
        map_sheet_names = []
        first_sample_name = None

        is_binned = import_data.get('bin_mode', False)
        save_maps = import_data.get('save_maps', False)

        for region in import_data['regions']:
            region_name = region['name']

            # Clean up region name
            sheet_base, is_special = clean_region_name(region_name)

            if is_binned and 'binned_intensities' in region:
                # Create multiple sheets for binned data
                for bin_idx, intensities in enumerate(region['binned_intensities']):
                    if len(region['binned_intensities']) > 1:
                        sheet_name = f"{sheet_base}{bin_idx + 1}"
                    else:
                        sheet_name = sheet_base

                    base_sheet_name = sheet_name
                    counter = 1
                    while sheet_name in sheet_names:
                        sheet_name = f"{base_sheet_name}{counter}"
                        counter += 1

                    sheet_names.append(sheet_name)
                    update_console(f"  Creating sheet: {sheet_name}")

                    ws = wb.create_sheet(sheet_name)
                    write_sheet_data(ws, region, intensities, sheet_name,
                                   bin_idx if is_binned else None,
                                   region.get('bin_indices', [[]])[bin_idx] if is_binned else None)

                    if first_sample_name is None:
                        first_sample_name = region['metadata'].get('Sample',
                                          region['metadata'].get('Spectrum Name', os.path.basename(input_file)))
            else:
                # Single summed spectrum
                sheet_name = sheet_base
                base_sheet_name = sheet_name
                counter = 1
                while sheet_name in sheet_names:
                    sheet_name = f"{base_sheet_name}{counter}"
                    counter += 1

                sheet_names.append(sheet_name)
                update_console(f"  Creating sheet: {sheet_name}")

                ws = wb.create_sheet(sheet_name)
                write_sheet_data(ws, region, region['intensities'], sheet_name)

                if first_sample_name is None:
                    first_sample_name = region['metadata'].get('Sample',
                                      region['metadata'].get('Spectrum Name', os.path.basename(input_file)))

            # Save map data if available
            if save_maps and region.get('data_2d') is not None:
                # Use XPS~Map naming: XPS~Map, XPS~Map1, XPS~Map2, etc.
                if len(map_sheet_names) == 0:
                    map_sheet_name = "XPS~Map"
                else:
                    map_sheet_name = f"XPS~Map{len(map_sheet_names)}"

                # Ensure unique
                base_map_name = map_sheet_name
                counter = len(map_sheet_names) + 1
                while map_sheet_name in sheet_names or map_sheet_name in map_sheet_names:
                    map_sheet_name = f"XPS~Map{counter}"
                    counter += 1

                map_sheet_names.append(map_sheet_name)
                update_console(f"  Creating map sheet: {map_sheet_name}")

                ws_map = wb.create_sheet(map_sheet_name)
                write_map_sheet_data(ws_map, region, map_sheet_name, sheet_base)

        update_console(f"Saving Excel file...")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} spectrum sheet(s)")
        if map_sheet_names:
            update_console(f"  Created {len(map_sheet_names)} map sheet(s)")

        # Initialize window.Data
        update_console("Initializing data structures...")
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Add spectrum sheets
        for sheet_name in sheet_names:
            add_core_level_Data(window.Data, window, excel_path, sheet_name)
            update_console(f"  Imported: {sheet_name}")

        # Add map sheets to Data structure using the actual map_sheet_names created
        map_idx = 0
        for idx, region in enumerate(import_data['regions']):
            if save_maps and region.get('data_2d') is not None:
                if map_idx < len(map_sheet_names):
                    map_sheet_name = map_sheet_names[map_idx]
                    sheet_base, _ = clean_region_name(region['name'])
                    add_map_to_data(window.Data, region, map_sheet_name, sheet_base)
                    update_console(f"  Imported map: {map_sheet_name}")
                    map_idx += 1

        window.Data['SampleNames'] = {0: first_sample_name}

        update_console("Creating JSON file...")
        serializable_data = convert_to_serializable_and_round(window.Data)
        with open(json_path, 'w') as f:
            json.dump(serializable_data, f, indent=4)

        # Update UI - include both spectrum and map sheets
        all_sheets = sheet_names + map_sheet_names
        window.sheet_combobox.Clear()
        for sheet_name in all_sheets:
            window.sheet_combobox.Append(sheet_name)
        window.sheet_combobox.SetSelection(0)
        window.current_sheet = all_sheets[0]

        on_sheet_selected(window, None)
        save_state(window)

        update_console("Refreshing display...")
        from libraries.FileMenu.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected, update_console)

        update_console("\nImport complete!")
        update_console(f"  {len(sheet_names)} spectrum sheet(s)")
        if map_sheet_names:
            update_console(f"  {len(map_sheet_names)} map sheet(s)")

        wx.CallLater(1500, console_frame.Close)

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing Scienta data:\n\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def write_sheet_data(ws, region, intensities, sheet_name, bin_idx=None, bin_indices=None):
    """Write spectrum data to an Excel worksheet with experimental info."""
    ws.cell(row=1, column=1, value='BE')
    ws.cell(row=1, column=2, value='Raw Data')

    be_values = region['be_values']

    for i, (be, intensity) in enumerate(zip(be_values, intensities), start=2):
        ws.cell(row=i, column=1, value=round(float(be), 2))
        ws.cell(row=i, column=2, value=round(float(intensity), 2))

    # Add experimental info
    exp_col = 50
    ws.cell(row=1, column=exp_col, value="Experimental Description")

    metadata = region['metadata']

    exp_metadata = {
        'Sample ID': metadata.get('Sample', ''),
        'Spectrum Name': metadata.get('Spectrum Name', ''),
        'Region Name': region['name'],
        'Instrument': metadata.get('Instrument', ''),
        'Location': metadata.get('Location', ''),
        'User': metadata.get('User', ''),
        'Date': metadata.get('Date', ''),
        'Time': metadata.get('Time', ''),
        'Technique': 'XPS',
        'Species & Transition': sheet_name,
        'Excitation Energy': metadata.get('Excitation Energy', ''),
        'Pass Energy': metadata.get('Pass Energy', ''),
        'Energy Scale': metadata.get('Energy Scale', ''),
        'Lens Mode': metadata.get('Lens Mode', ''),
        'Acquisition Mode': metadata.get('Acquisition Mode', ''),
        'Energy Step': metadata.get('Energy Step', ''),
        'Step Time': metadata.get('Step Time', ''),
        'Number of Points': str(len(be_values)),
        'BE Start': f"{be_values[0]:.2f}",
        'BE End': f"{be_values[-1]:.2f}",
    }

    if 'total_sweeps' in region:
        exp_metadata['Total Sweeps in File'] = str(region.get('total_sweeps', 1))
    if 'num_sweeps_summed' in region:
        exp_metadata['Sweeps Summed'] = str(region.get('num_sweeps_summed', 1))
    if 'dropped_sweeps' in region and region['dropped_sweeps']:
        exp_metadata['Dropped Sweeps'] = ', '.join(map(str, region['dropped_sweeps']))

    if bin_idx is not None:
        exp_metadata['Bin Number'] = str(bin_idx + 1)
    if bin_indices is not None:
        exp_metadata['Sweeps in Bin'] = ', '.join(map(str, [i + 1 for i in bin_indices]))

    row = 2
    for key, value in exp_metadata.items():
        ws.cell(row=row, column=exp_col, value=key)
        ws.cell(row=row, column=exp_col + 1, value=str(value))
        row += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40


def write_map_sheet_data(ws, region, map_sheet_name, core_level_name=None):
    """Write map data to an Excel worksheet."""
    be_values = region['be_values']
    data_2d = region['data_2d']
    num_sweeps = data_2d.shape[0]

    # Header row: BE, Y1, Y2, Y3, ...
    ws.cell(row=1, column=1, value='BE')
    for sweep_idx in range(num_sweeps):
        ws.cell(row=1, column=sweep_idx + 2, value=f'Y{sweep_idx + 1}')

    # Data rows
    for i, be in enumerate(be_values):
        ws.cell(row=i + 2, column=1, value=round(float(be), 2))
        for sweep_idx in range(num_sweeps):
            ws.cell(row=i + 2, column=sweep_idx + 2, value=round(float(data_2d[sweep_idx, i]), 2))

    # Add experimental info
    exp_col = num_sweeps + 10
    ws.cell(row=1, column=exp_col, value="Experimental Description")

    metadata = region['metadata']

    exp_metadata = {
        'Sample ID': metadata.get('Sample', ''),
        'Spectrum Name': metadata.get('Spectrum Name', ''),
        'Region Name': region['name'],
        'Core Level': core_level_name or '',
        'Map Sheet Name': map_sheet_name,
        'Instrument': metadata.get('Instrument', ''),
        'Location': metadata.get('Location', ''),
        'User': metadata.get('User', ''),
        'Date': metadata.get('Date', ''),
        'Time': metadata.get('Time', ''),
        'Technique': 'XPS Map',
        'Excitation Energy': metadata.get('Excitation Energy', ''),
        'Pass Energy': metadata.get('Pass Energy', ''),
        'Energy Scale': metadata.get('Energy Scale', ''),
        'Lens Mode': metadata.get('Lens Mode', ''),
        'Number of Sweeps': str(num_sweeps),
        'Number of Points': str(len(be_values)),
        'BE Start': f"{be_values[0]:.2f}",
        'BE End': f"{be_values[-1]:.2f}",
    }

    row = 2
    for key, value in exp_metadata.items():
        ws.cell(row=row, column=exp_col, value=key)
        ws.cell(row=row, column=exp_col + 1, value=str(value))
        row += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40


def add_map_to_data(Data, region, map_sheet_name, core_level_name=None):
    """Add map data to window.Data structure."""
    be_values = region['be_values']
    data_2d = region['data_2d']
    num_sweeps = data_2d.shape[0]

    # Get the core level name from the region if not provided
    if core_level_name is None:
        core_level_name, _ = clean_region_name(region['name'])

    # Create the data structure similar to standard core levels
    map_data = {
        'Name': map_sheet_name,
        'B.E.': [round(float(be), 2) for be in be_values],
        '_Map_type': 'scienta',
        '_num_sweeps': num_sweeps,
        '_core_level': core_level_name,
    }

    # Add Y columns: Y1, Y2, Y3, ...
    for sweep_idx in range(num_sweeps):
        col_name = f'Y{sweep_idx + 1}'
        map_data[col_name] = [round(float(data_2d[sweep_idx, i]), 2) for i in range(len(be_values))]

    # Add experimental info
    metadata = region['metadata']
    map_data['ExperimentalInfo'] = {
        'Sample ID': metadata.get('Sample', ''),
        'Spectrum Name': metadata.get('Spectrum Name', ''),
        'Region Name': region['name'],
        'Core Level': core_level_name,
        'Map Sheet Name': map_sheet_name,
        'Instrument': metadata.get('Instrument', ''),
        'Location': metadata.get('Location', ''),
        'User': metadata.get('User', ''),
        'Date': metadata.get('Date', ''),
        'Time': metadata.get('Time', ''),
        'Technique': 'XPS Map',
        'Excitation Energy': metadata.get('Excitation Energy', ''),
        'Pass Energy': metadata.get('Pass Energy', ''),
        'Number of Sweeps': str(num_sweeps),
        'Number of Points': str(len(be_values)),
        'BE Start': f"{be_values[0]:.2f}",
        'BE End': f"{be_values[-1]:.2f}",
    }

    # Add background structure (empty for maps)
    map_data['Background'] = {
        'Bkg Type': '',
        'Bkg Low': round(float(min(be_values)), 2),
        'Bkg High': round(float(max(be_values)), 2),
        'Bkg Offset Low': 0,
        'Bkg Offset High': 0
    }

    Data['Core levels'][map_sheet_name] = map_data
    Data['Number of Core levels'] = len(Data['Core levels'])


def parse_h5_scienta_file(file_path):
    """
    Parse a Scienta Omicron HDF5 (.h5) file.

    Returns a dict compatible with the existing Scienta map/import pipeline:
        {
            'file_path': str,
            'name': str,               # spectrum name from file
            'element_set': str,        # element set / region label
            'source_energy': float,    # photon energy (eV)
            'pass_energy': float,
            'work_function': float,
            'lens_mode': str,
            'acq_mode': str,           # 'Image' etc.
            'energy_mode': str,        # 'Kinetic' or 'Binding'
            'dwell_time': float,
            'acq_time': float,
            'start_time': str,
            'stop_time': str,
            'instrument_model': str,
            'be_values': np.ndarray,   # shape (n_energy,) — always BE
            'ke_values': np.ndarray,   # shape (n_energy,)
            'y_axis': np.ndarray,      # shape (n_y,) spatial / angle / etch
            'y_label': str,
            'y_unit': str,
            'data_2d': np.ndarray,     # shape (n_y, n_energy) — intensities
            'data_1d': np.ndarray,     # shape (n_energy,)     — reduced sum
            'is_map': True,
        }
    """
    if not HAS_H5PY:
        raise ImportError("h5py is required to import HDF5 files.\n"
                          "Install it with:  pip install h5py")

    with h5py.File(file_path, 'r') as f:

        def _str(ds):
            v = ds[()]
            if isinstance(v, bytes):
                return v.decode('utf-8', errors='replace')
            return str(v)

        def _float(ds):
            return float(ds[()])

        def _int(ds):
            return int(ds[()])

        # --- spectrum_definition ---
        sd = f['acquisition/spectrum_definition']
        acq_mode    = _str(sd['acquisition_mode'])
        energy_mode = _str(sd['energy_mode'])
        lens_mode   = _str(sd['lens_mode_name'])
        pass_energy = _float(sd['pass_energy'])
        dwell_time  = _float(sd['dwell_time'])
        acq_time    = _float(sd['acquisition_time'])
        element_set = _str(sd['element_set_name'])

        # --- spectrum_log ---
        sl = f['acquisition/spectrum_log']
        start_time = _str(sl['start_time'])
        stop_time  = _str(sl['stop_time'])

        # --- spectrum ---
        sp = f['acquisition/spectrum']
        name = _str(sp['name'])

        # --- instrument ---
        src = f['instrument/analyser/excitation_source']
        source_energy = _float(src['energy'])

        analyser = f['instrument/analyser']
        work_function = _float(analyser['work_function'])
        instrument_model = _str(analyser['model'])

        # --- 2D data ---
        data_grp = f['acquisition/spectrum/data']
        data_2d_raw = data_grp['data'][()]           # shape (n_y, n_energy)
        x_axis_raw  = data_grp['x_axis'][()]         # Analyser Energy or KE, shape (n_energy,)
        y_axis      = data_grp['y_axis'][()]         # spatial / angle, shape (n_y,)
        y_label     = data_grp['y_axis'].attrs.get('label', b'Y').decode() \
                      if isinstance(data_grp['y_axis'].attrs.get('label', b'Y'), bytes) \
                      else str(data_grp['y_axis'].attrs.get('label', 'Y'))
        y_unit      = data_grp['y_axis'].attrs.get('units', b'').decode() \
                      if isinstance(data_grp['y_axis'].attrs.get('units', b''), bytes) \
                      else str(data_grp['y_axis'].attrs.get('units', ''))

        # --- 1D reduced ---
        red_grp = f['acquisition/spectrum/data_reduced_1d']
        data_1d_raw   = red_grp['data'][()]          # shape (n_energy,)
        # x_axis for 1d is always Kinetic Energy
        ke_1d         = red_grp['x_axis'][()]

    # Convert KE → BE: BE = hν − KE
    # x_axis_raw may be labelled 'Analyser Energy' which equals KE here
    ke_values = np.array(x_axis_raw, dtype=float)
    be_values = source_energy - ke_values            # descending if KE ascending

    # Flip so BE is in descending order (high → low, standard XPS convention)
    if be_values[0] < be_values[-1]:
        be_values = be_values[::-1]
        ke_values = ke_values[::-1]
        data_2d_raw = data_2d_raw[:, ::-1]
        data_1d_raw = data_1d_raw[::-1]

    return {
        'file_path':        file_path,
        'name':             name,
        'element_set':      element_set,
        'source_energy':    round(source_energy, 4),
        'pass_energy':      round(pass_energy, 2),
        'work_function':    round(work_function, 4),
        'lens_mode':        lens_mode,
        'acq_mode':         acq_mode,
        'energy_mode':      energy_mode,
        'dwell_time':       dwell_time,
        'acq_time':         acq_time,
        'start_time':       start_time,
        'stop_time':        stop_time,
        'instrument_model': instrument_model,
        'be_values':        be_values,
        'ke_values':        ke_values,
        'y_axis':           y_axis,
        'y_label':          y_label,
        'y_unit':           y_unit,
        'data_2d':          np.array(data_2d_raw, dtype=float),  # (n_y, n_energy)
        'data_1d':          np.array(data_1d_raw, dtype=float),  # (n_energy,)
        'is_map':           True,
    }


def import_h5_scienta_file(window, file_path=None):
    """
    Import a Scienta Omicron HDF5 (.h5) spatial map file.

    Parses the file, reshapes data into the ScientaMapPreviewWindow format,
    then opens the preview window so the user can drop sweeps, bin, or sum
    before finalising the import via finalize_scienta_import.

    The 'sweeps' axis corresponds to Y positions (spatial / angle / etch).
    """
    if not HAS_H5PY:
        wx.MessageBox(
            "The 'h5py' library is required to import HDF5 files.\n\n"
            "Install it with:\n  pip install h5py",
            "Missing Library", wx.OK | wx.ICON_ERROR)
        return

    if file_path is None:
        with wx.FileDialog(
                window, "Open Scienta HDF5 file",
                wildcard="HDF5 files (*.h5;*.hdf5)|*.h5;*.hdf5|All files (*.*)|*.*",
                style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

    parent_pos  = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Loading Scienta HDF5 File", size=(440, 220))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width  - 440) // 2,
        parent_pos.y + (parent_size.height - 220) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        update_console(f"Opening: {os.path.basename(file_path)}")
        h5data = parse_h5_scienta_file(file_path)

        n_y      = h5data['data_2d'].shape[0]
        n_energy = h5data['data_2d'].shape[1]

        update_console(f"  Instrument:    {h5data['instrument_model']}")
        update_console(f"  Element set:   {h5data['element_set']}")
        update_console(f"  Source energy: {h5data['source_energy']:.3f} eV")
        update_console(f"  BE range:      {h5data['be_values'][-1]:.3f} – {h5data['be_values'][0]:.3f} eV")
        update_console(f"  Map size:      {n_y} {h5data['y_label']} positions × {n_energy} energy points")

        # Derive a clean core-level / region name from the element set
        core_level, _ = clean_region_name(h5data['element_set'])
        if not core_level:
            core_level = base_name

        # Build metadata dict matching the Scienta .txt metadata keys
        source_label = (
            'He I (UPS)'  if abs(h5data['source_energy'] - 21.218) < 0.05 else
            'He II (UPS)' if abs(h5data['source_energy'] - 40.814) < 0.05 else
            'Al K-alpha'  if abs(h5data['source_energy'] - 1486.68) < 0.1  else
            'Mg K-alpha'  if abs(h5data['source_energy'] - 1253.6)  < 0.1  else
            f"Photon {h5data['source_energy']:.3f} eV"
        )

        metadata = {
            'Sample':           h5data['element_set'],
            'Spectrum Name':    h5data['name'],
            'Instrument':       h5data['instrument_model'],
            'Location':         '',
            'User':             '',
            'Date':             h5data['start_time'][:10] if h5data['start_time'] else '',
            'Time':             h5data['start_time'][11:] if len(h5data['start_time']) > 10 else '',
            'Technique':        'UPS' if h5data['source_energy'] < 100 else 'XPS',
            'Excitation Energy': f"{h5data['source_energy']:.3f}",
            'Source Label':     source_label,
            'Pass Energy':      f"{h5data['pass_energy']:.2f}",
            'Work Function':    f"{h5data['work_function']:.4f}",
            'Lens Mode':        h5data['lens_mode'],
            'Acquisition Mode': h5data['acq_mode'],
            'Energy Mode':      h5data['energy_mode'],
            'Energy Scale':     'Binding',
            'Energy Step':      f"{abs(float(h5data['be_values'][1]) - float(h5data['be_values'][0])):.4f}",
            'Step Time':        f"{h5data['dwell_time']:.4f}",
            'Start Time':       h5data['start_time'],
            'Stop Time':        h5data['stop_time'],
            'Y Axis Label':     h5data['y_label'],
            'Y Axis Unit':      h5data['y_unit'],
            'Y Axis Start':     f"{h5data['y_axis'][0]:.4f}",
            'Y Axis End':       f"{h5data['y_axis'][-1]:.4f}",
            'Y Axis Points':    str(n_y),
        }

        # Build sweep_scale array (Y-axis values, used as sweep labels)
        sweep_scale = np.arange(n_y)

        # Build the single region dict that ScientaMapPreviewWindow expects
        # data_2d shape must be (sweeps, n_be) = (n_y, n_energy) — already correct
        region = {
            'name':        core_level,
            'be_values':   h5data['be_values'],         # shape (n_energy,)
            'sweeps':      n_y,
            'sweep_scale': sweep_scale,
            'data_2d':     h5data['data_2d'],           # shape (n_y, n_energy)
            'intensities': None,
            'metadata':    metadata,
            'dropped_sweeps': [],
        }

        # parsed_data dict matching what ScientaMapPreviewWindow and
        # finalize_scienta_import expect
        parsed_data = {
            'num_regions': 1,
            'version':     'HDF5',
            'regions':     [region],
            'file_path':   file_path,
            'is_map':      True,
        }

        update_console("\nOpening preview window...")
        console_frame.Close()

        def on_import_complete(data):
            finalize_scienta_import(window, data)

        ScientaMapPreviewWindow(window, parsed_data, on_import_complete)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error loading HDF5 file:\n\n{exc}", "Error", wx.OK | wx.ICON_ERROR)
