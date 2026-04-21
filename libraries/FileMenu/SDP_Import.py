# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
SDP File Import Module for KherveFitting
XPS International / Spectral Data Processor Binary Format

SDP files store XPS spectra in a proprietary binary format used by SDP v8 software
from XPS International LLC (B. Vincent Crist). Each file can contain multiple spectra
(survey + narrow scans). Data is stored as Binding Energy with float64 intensity values.

Format structure:
  - 18-byte magic: "XI SDP BINARY FILE"
  - 2-byte version (uint16)
  - 2-byte declared spectrum count (uint16)
  - Global header (variable)
  - Per-spectrum blocks, each containing:
      * Axis descriptors with "Binding Energy" and "Counts" labels
      * Length-prefixed metadata strings (file path, description, date, format)
      * After "VAMAS/ISO\0" marker: parameter block with:
          +0:  BE start (float64)
          +8:  BE step  (float64)
          +16: BE end   (float64)
          +24: Y min    (float64)
          +32: Y max    (float64)
          +40: 2-byte pad
          +42: Source energy (float64)
          +50: (reserved, float64)
          +58: Work function (float64)
          +66: Pass energy (float64)
          +74: Sweeps/scans (float64)
          +82: (reserved, float64)
          +90: Flag (uint32 = 1)
          +94: Data count (uint32)
          +98: Intensity data (data_count * float64)
  - Optional footer block (partial descriptor without full data)
"""

import os
import struct
import numpy as np
import openpyxl
import wx
import json
import re


MAGIC = b'XI SDP BINARY FILE'
MAGIC_LEN = 18

# Marker used to locate the start of each spectrum block
BE_MARKER = b'Binding Energy'

# Fixed offset from a "Binding Energy" marker to the first metadata string.
# Structure between them: axis descriptor fields (unit, display params,
# Counts descriptor, y-range, fixed 26-byte block, padding).
BE_TO_META_STRINGS = 112   # bytes from start of "Binding Energy" to first uint16 len

# Number of length-prefixed metadata strings before the parameter block:
#   0: file path,  1: description,  2: acquisition params,
#   3: date,  4: format string (e.g. "VAMAS/ISO", "XI ASCII")
NUM_META_STRINGS = 5

# Field offsets within the parameter block (starts right after the 5th string)
OFF_BE_START = 0
OFF_BE_STEP = 8
OFF_BE_END = 16
OFF_Y_MIN = 24
OFF_Y_MAX = 32
OFF_SOURCE_E = 42      # 2-byte gap at +40
OFF_WORK_FN = 58
OFF_PASS_E = 66
OFF_SWEEPS = 74
OFF_FLAG = 90
OFF_DATA_COUNT = 94
OFF_DATA_START = 98


# ---------------------------------------------------------------------------
# Core-level identification
# ---------------------------------------------------------------------------

def identify_core_level_from_be(be_start, be_end):
    """
    Identify core level name from the Binding Energy range.

    Uses the primary XPS core level for each element (one entry per element).
    Only the strongest / most commonly measured line is listed — no sub-shells
    (e.g. Ni2p is listed but not Ni3p, Ni3s, Fe2p but not Fe3p, etc.).

    Returns 'VB' for valence-band scans (0–20 eV), 'Survey' for wide scans
    (> 200 eV range), or the closest matching core level name.
    """
    center = (be_start + be_end) / 2.0
    width = abs(be_start - be_end)

    # Wide scans are surveys
    if width > 200:
        return "Survey"

    # Valence band: scan centred in the 0–20 eV region
    if center <= 20 and width < 60:
        return "VB"

    # ---- Primary XPS core levels (one per element, sorted by BE) ----------
    #
    # Binding energies are approximate centres of the spin-orbit doublet
    # (or the single peak for s-levels) measured with Al K-alpha (1486.68 eV).
    #
    # Period 6 4f / Period 5 3d / Period 4 3d / Period 3 2p / Period 2 1s
    # plus lanthanides (3d) and actinides (4f).

    core_levels = [
        # BE (eV)  Name       # Element – line
        (14,  "Hf4f"),        # Hafnium 4f7/2
        (18,  "Ga3d"),        # Gallium 3d
        (22,  "Ta4f"),        # Tantalum 4f7/2
        (26,  "Ge3d"),        # Germanium 3d
        (31,  "W4f"),         # Tungsten 4f7/2
        (40,  "Re4f"),        # Rhenium 4f7/2
        (42,  "As3d"),        # Arsenic 3d5/2
        (49,  "Mg2p"),        # Magnesium 2p
        (51,  "Os4f"),        # Osmium 4f7/2
        (55,  "Se3d"),        # Selenium 3d  (Li1s also ~55)
        (61,  "Ir4f"),        # Iridium 4f7/2
        (69,  "Br3d"),        # Bromine 3d
        (71,  "Pt4f"),        # Platinum 4f7/2
        (74,  "Al2p"),        # Aluminium 2p
        (84,  "Au4f"),        # Gold 4f7/2
        (99,  "Si2p"),        # Silicon 2p
        (111, "Rb3d"),        # Rubidium 3d5/2
        (118, "Tl4f"),        # Thallium 4f7/2
        (130, "P2p"),         # Phosphorus 2p
        (134, "Sr3d"),        # Strontium 3d5/2
        (139, "Pb4f"),        # Lead 4f7/2
        (142, "Gd4d"),        # Gadolinium 4d
        (151, "Si2s"),        # Silicon 2s
        (157, "Y3d"),         # Yttrium 3d5/2  (Bi4f ~157)
        (162, "S2p"),         # Sulfur 2p
        (170, "Bi4f"),        # Bismuth 4f7/2
        (182, "Zr3d"),        # Zirconium 3d5/2
        (189, "B1s"),         # Boron 1s
        (199, "Cl2p"),        # Chlorine 2p
        (207, "Nb3d"),        # Niobium 3d5/2
        (228, "Mo3d"),        # Molybdenum 3d5/2
        (232, "S2s"),         # Sulfur 2s
        (242, "Ar2p"),        # Argon 2p  (implanted)
        (280, "C1s"),         # Ruthenium 3d5/2 overlaps — C1s predominates
        (285, "C1s"),         # Carbon 1s
        (293, "K2p"),         # Potassium 2p
        (307, "Rh3d"),        # Rhodium 3d5/2
        (334, "Pd3d"),        # Palladium 3d5/2  (Th4f ~333)
        (347, "Ca2p"),        # Calcium 2p3/2
        (368, "Ag3d"),        # Silver 3d5/2
        (382, "U4f"),         # Uranium 4f7/2
        (399, "N1s"),         # Nitrogen 1s
        (405, "Cd3d"),        # Cadmium 3d5/2
        (412, "Sc2p"),        # Scandium 2p
        (444, "In3d"),        # Indium 3d5/2
        (459, "Ti2p"),        # Titanium 2p3/2
        (487, "Sn3d"),        # Tin 3d5/2
        (517, "V2p"),         # Vanadium 2p3/2
        (529, "O1s"),         # Antimony 3d5/2 overlaps — O1s predominates
        (532, "O1s"),         # Oxygen 1s
        (573, "Te3d"),        # Tellurium 3d5/2
        (577, "Cr2p"),        # Chromium 2p3/2
        (619, "I3d"),         # Iodine 3d5/2
        (641, "Mn2p"),        # Manganese 2p3/2
        (686, "F1s"),         # Fluorine 1s
        (711, "Fe2p"),        # Iron 2p3/2
        (726, "Cs3d"),        # Caesium 3d5/2
        (780, "Co2p"),        # Cobalt 2p3/2  (Ba3d5/2 ~780)
        (796, "Ba3d"),        # Barium 3d5/2
        (836, "La3d"),        # Lanthanum 3d5/2
        (855, "Ni2p"),        # Nickel 2p3/2
        (883, "Ce3d"),        # Cerium 3d5/2
        (929, "Pr3d"),        # Praseodymium 3d5/2
        (933, "Cu2p"),        # Copper 2p3/2
        (980, "Nd3d"),        # Neodymium 3d5/2
        (1004, "Pm3d"),       # Promethium 3d (very rare)
        (1022, "Zn2p"),       # Zinc 2p3/2
        (1072, "Na1s"),       # Sodium 1s
        (1083, "Sm3d"),       # Samarium 3d5/2
        (1117, "Ga2p"),       # Gallium 2p3/2
        (1126, "Eu3d"),       # Europium 3d
        (1186, "Gd3d"),       # Gadolinium 3d
        (1220, "Tb3d"),       # Terbium 3d
        (1296, "Dy3d"),       # Dysprosium 3d
        (1351, "Ho3d"),       # Holmium 3d
        (1409, "Er3d"),       # Erbium 3d
        (1468, "Tm3d"),       # Thulium 3d  (at limit of Al K-alpha)
    ]

    best_name = "Unknown"
    best_dist = 999.0
    for be, name in core_levels:
        dist = abs(center - be)
        if dist < best_dist and dist < 30:
            best_dist = dist
            best_name = name

    return best_name


def identify_core_level_from_description(desc):
    """
    Extract core-level name from the SDP description string.

    Handles common notations:
        'C1s Scan'  -> 'C1s'
        'O 1s'      -> 'O1s'
        'Fe 2p'     -> 'Fe2p'
        'Survey'    -> 'Survey'
        'VB scan'   -> 'VB'
    """
    if not desc:
        return None

    # Core level pattern: Element (1-2 chars) + optional space + number + orbital
    match = re.match(r'^([A-Z][a-z]?\s*\d+[spdfgh]\d*)', desc)
    if match:
        return match.group(1).replace(' ', '')

    # Check for Survey / Wide / VB / Valence keywords
    low = desc.lower().strip()
    if low.startswith(('survey', 'wide')):
        return 'Survey'
    if low.startswith(('vb', 'valence')):
        return 'VB'

    return None


def _extract_core_level_from_filename(filename):
    """
    Extract core-level name from a filename (SDP or original source).

    The outer SDP filename or the stored original path may encode
    the core level when the file holds a single spectrum:
        'O1s.sdp'        -> 'O1s'
        'C1s_Scan.sdp'   -> 'C1s'
        'Fe 2p scan.sdp' -> 'Fe2p'
        'Survey.sdp'     -> 'Survey'
        'VB.sdp'         -> 'VB'
    """
    if not filename:
        return None

    base = os.path.splitext(os.path.basename(filename))[0]

    # Remove common suffixes before matching
    for suffix in ['_Scan', '_scan', '_SCAN', ' Scan', ' scan',
                   '_Region', '_region', ' Region', ' region',
                   '_spe', '_SPE', '_txt', '_TXT', '_vgd', '_VGD',
                   '_vms', '_VMS', '_iso', '_ISO', '_mrs', '_MRS']:
        if base.endswith(suffix):
            base = base[:-len(suffix)]
            break

    # Core level at the start of the filename
    match = re.match(r'^([A-Z][a-z]?\s*\d+[spdfgh]\d*)', base)
    if match:
        return match.group(1).replace(' ', '')

    low = base.lower().strip()
    if low.startswith(('survey', 'wide')):
        return 'Survey'
    if low.startswith(('vb', 'valence')):
        return 'VB'

    return None


# ---------------------------------------------------------------------------
# Filename helpers
# ---------------------------------------------------------------------------

def extract_sample_name(filename):
    """
    Extract clean sample name from SDP filename.
    Removes common suffixes indicating source format.
    """
    base_name = os.path.splitext(filename)[0]

    suffixes_to_remove = [
        '_spe', '_SPE', '_txt', '_TXT', '_vgd', '_VGD',
        '_vms', '_VMS', '_iso', '_ISO', '_mrs', '_MRS',
    ]
    result = base_name
    for suffix in suffixes_to_remove:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break

    return result.strip()


# ---------------------------------------------------------------------------
# Binary parser
# ---------------------------------------------------------------------------

def parse_sdp_file(file_path):
    """
    Parse an SDP binary file and return extracted spectra data.

    Locates spectrum blocks by finding "Binding Energy" markers, then walks
    through the fixed axis-descriptor structure and 5 length-prefixed metadata
    strings (path, description, acq params, date, format) to reach the
    parameter + data block that follows.

    This approach is format-marker agnostic — it works regardless of whether
    the source format string is "VAMAS/ISO", "XI ASCII", or any other variant.

    Returns:
        dict with keys:
            'spectra': list of spectrum dicts
            'metadata': dict with file-level info
    """
    with open(file_path, 'rb') as f:
        data = f.read()

    file_size = len(data)

    if len(data) < MAGIC_LEN or data[:MAGIC_LEN] != MAGIC:
        raise ValueError("Not a valid SDP file: missing 'XI SDP BINARY FILE' header")

    version = struct.unpack('<H', data[18:20])[0]
    num_spectra_declared = struct.unpack('<H', data[20:22])[0]

    # Locate spectrum blocks by finding all "Binding Energy" markers
    be_offsets = []
    pos = 0
    while True:
        pos = data.find(BE_MARKER, pos)
        if pos == -1:
            break
        be_offsets.append(pos)
        pos += 1

    if not be_offsets:
        raise ValueError("No spectrum data found in SDP file "
                         "(no 'Binding Energy' markers)")

    spectra = []

    for block_idx, be_off in enumerate(be_offsets):
        # ---- Walk through the 5 metadata strings ----------------------------
        meta_start = be_off + BE_TO_META_STRINGS
        if meta_start + 4 > file_size:
            continue

        off = meta_start
        meta_strings = []   # [path, description, acq_params, date, format]
        valid_meta = True
        for _ in range(NUM_META_STRINGS):
            if off + 2 > file_size:
                valid_meta = False
                break
            slen = struct.unpack('<H', data[off:off + 2])[0]
            if slen > 1000 or off + 2 + slen > file_size:
                valid_meta = False
                break
            raw = data[off + 2:off + 2 + slen]
            meta_strings.append(
                raw.split(b'\x00')[0].decode('ascii', errors='replace').strip())
            off += 2 + slen

        if not valid_meta or len(meta_strings) < NUM_META_STRINGS:
            continue

        # ---- Read parameter block -------------------------------------------
        params_base = off   # immediately after the 5th metadata string

        if params_base + OFF_DATA_START > file_size:
            continue

        try:
            be_start = struct.unpack('<d', data[params_base + OFF_BE_START:
                                                params_base + OFF_BE_START + 8])[0]
            be_step = struct.unpack('<d', data[params_base + OFF_BE_STEP:
                                               params_base + OFF_BE_STEP + 8])[0]
            be_end = struct.unpack('<d', data[params_base + OFF_BE_END:
                                              params_base + OFF_BE_END + 8])[0]
        except struct.error:
            continue

        # Sanity checks on BE range
        if abs(be_start) > 5000 or abs(be_end) > 5000:
            continue
        if abs(be_step) > 100 or be_step == 0:
            continue

        # Read acquisition parameters
        try:
            y_min = struct.unpack('<d', data[params_base + OFF_Y_MIN:
                                             params_base + OFF_Y_MIN + 8])[0]
            y_max = struct.unpack('<d', data[params_base + OFF_Y_MAX:
                                             params_base + OFF_Y_MAX + 8])[0]
            source_energy = struct.unpack('<d', data[params_base + OFF_SOURCE_E:
                                                     params_base + OFF_SOURCE_E + 8])[0]
            work_fn = struct.unpack('<d', data[params_base + OFF_WORK_FN:
                                               params_base + OFF_WORK_FN + 8])[0]
            pass_energy = struct.unpack('<d', data[params_base + OFF_PASS_E:
                                                   params_base + OFF_PASS_E + 8])[0]
            sweeps = struct.unpack('<d', data[params_base + OFF_SWEEPS:
                                              params_base + OFF_SWEEPS + 8])[0]
        except struct.error:
            source_energy = 1486.68
            work_fn = pass_energy = sweeps = 0.0
            y_min = y_max = 0.0

        # Read data count
        try:
            data_count = struct.unpack('<I', data[params_base + OFF_DATA_COUNT:
                                                  params_base + OFF_DATA_COUNT + 4])[0]
        except struct.error:
            continue

        data_start = params_base + OFF_DATA_START

        if data_count == 0 or data_count > 100000:
            continue
        if data_start + data_count * 8 > file_size:
            continue

        # Read intensity data
        intensities = []
        for i in range(data_count):
            val = struct.unpack('<d', data[data_start + i * 8:
                                           data_start + i * 8 + 8])[0]
            intensities.append(val)

        if not intensities or all(v == 0 for v in intensities):
            continue

        # Verify data looks like real XPS counts (not garbage bytes)
        valid_count = sum(1 for v in intensities if abs(v) < 1e12 and v == v)
        if valid_count < len(intensities) * 0.9:
            continue

        # Generate BE axis
        be_values = [be_start + i * be_step for i in range(data_count)]

        # ---- Extract metadata from the 5 strings ----------------------------
        file_path_str = meta_strings[0]   # original file path
        description = meta_strings[1]      # sample / scan description
        acq_params = meta_strings[2]       # acquisition parameters
        date_str = meta_strings[3]         # date
        format_str = meta_strings[4]       # source format ("VAMAS/ISO", "XI ASCII", …)

        # Identify core level — priority:
        #   1. Description field (e.g. "C1s Scan", "O 1s")
        #   2. Outer SDP filename (e.g. "O1s.sdp", "VB.sdp")
        #   3. BE range fallback
        core_level = identify_core_level_from_description(description)
        if core_level is None:
            core_level = _extract_core_level_from_filename(file_path)
        if core_level is None:
            core_level = identify_core_level_from_be(be_start, be_values[-1])

        # Sanitise source energy
        if not (100 < source_energy < 5000):
            source_energy = 1486.68

        spectra.append({
            'be_start': be_start,
            'be_step': be_step,
            'be_end': be_end,
            'be_values': be_values,
            'intensities': intensities,
            'num_points': data_count,
            'core_level': core_level,
            'file_path': file_path_str,
            'description': description,
            'acq_params': acq_params,
            'date': date_str,
            'format': format_str,
            'source_energy': source_energy,
            'pass_energy': pass_energy if 0 < pass_energy < 1000 else None,
            'work_function': work_fn if 2.0 < work_fn < 8.0 else None,
            'sweeps': int(sweeps) if 0 < sweeps < 10000 else None,
        })

    if not spectra:
        raise ValueError("No valid spectrum data found in SDP file")

    # Build file-level metadata
    sample_name = extract_sample_name(os.path.basename(file_path))

    # Use the first description as sample name only if it's meaningful
    # (not a scan label like "C1s Scan" and not a filename)
    first_desc = spectra[0].get('description', '')
    scan_label_pattern = re.compile(
        r'^[A-Z][a-z]?\d+[spdfgh]\d*\s+(Scan|Region|scan|region)', re.IGNORECASE)
    if first_desc:
        is_scan_label = bool(scan_label_pattern.match(first_desc))
        is_filename = any(c in first_desc for c in ['\\', '/']) or first_desc.endswith(('.txt', '.vms'))
        if not is_scan_label and not is_filename:
            sample_name = first_desc

    source_label = "Al K-alpha Monochromated" if abs(
        spectra[0]['source_energy'] - 1486.68) < 0.5 else (
        f"X-ray {spectra[0]['source_energy']:.2f} eV")

    return {
        'spectra': spectra,
        'metadata': {
            'file_path': file_path,
            'sample_name': sample_name,
            'version': version,
            'num_spectra_declared': num_spectra_declared,
            'num_spectra_actual': len(spectra),
            'source_energy': spectra[0].get('source_energy', 1486.68),
            'source_label': source_label,
            'source_file': spectra[0].get('file_path', ''),
        }
    }


# ---------------------------------------------------------------------------
# Excel sheet writing
# ---------------------------------------------------------------------------

def _write_sheet(ws, spec, meta, spec_idx, total_spectra):
    """Write data + metadata columns to a worksheet."""
    import openpyxl

    core_level = spec['core_level']

    # Data columns A-B
    ws.cell(row=1, column=1, value='BE')
    ws.cell(row=1, column=2, value='Raw Data')

    for i, (be, intensity) in enumerate(
            zip(spec['be_values'], spec['intensities']), start=2):
        ws.cell(row=i, column=1, value=round(be, 2))
        ws.cell(row=i, column=2, value=round(intensity, 2))

    # Experimental metadata in column AX (50)
    exp_col = 50
    ws.cell(row=1, column=exp_col, value="Experimental Description")

    source_energy = spec.get('source_energy', 1486.68)
    source_label = meta.get('source_label', '')

    exp_metadata = {
        'Sample ID': meta['sample_name'],
        'Source File': os.path.basename(meta.get('source_file', '')),
        'Original Path': spec.get('file_path', ''),
        'Description': spec.get('description', ''),
        'Date': spec.get('date', ''),
        'Technique': 'XPS',
        'Species & Transition': core_level,
        'Spectrum Index': str(spec_idx),
        'Total Spectra': str(total_spectra),
        'Source Label': source_label,
        'Source Energy': f"{source_energy:.2f}",
        'Pass Energy': (f"{spec['pass_energy']:.2f}"
                        if spec.get('pass_energy') else 'Unknown'),
        'Work Function': (f"{spec['work_function']:.2f}"
                          if spec.get('work_function') else 'Unknown'),
        'Sweeps': str(spec['sweeps']) if spec.get('sweeps') else 'Unknown',
        'Number of Points': str(spec['num_points']),
        'BE Start': f"{spec['be_start']:.2f}",
        'BE End': f"{spec['be_end']:.2f}",
        'BE Step': f"{abs(spec['be_step']):.4f}",
        'SDP Version': str(meta.get('version', '')),
    }

    row = 2
    for key, value in exp_metadata.items():
        ws.cell(row=row, column=exp_col, value=key)
        ws.cell(row=row, column=exp_col + 1, value=str(value))
        row += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40


# ---------------------------------------------------------------------------
# Shared KherveFitting loader
# ---------------------------------------------------------------------------

def _load_into_kherve(window, excel_path, sheet_names, sample_name, update_console):
    """
    Load a completed Excel file into the KherveFitting data model, save JSON,
    update the UI combobox, and refresh sheet display.
    """
    from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
    from libraries.FileMenu.Save import (update_undo_redo_state, save_state,
                                         convert_to_serializable_and_round)
    from libraries.Sheet_Operations import on_sheet_selected

    update_console("\nLoading into KherveFitting...")

    window.Data = Init_Measurement_Data(window)
    window.Data['FilePath'] = excel_path

    window.history = []
    window.redo_stack = []
    update_undo_redo_state(window)

    window.results_grid.ClearGrid()
    if window.results_grid.GetNumberRows() > 0:
        window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

    for sheet_name in sheet_names:
        add_core_level_Data(window.Data, window, excel_path, sheet_name)
        update_console(f"  Imported: {sheet_name}")

    window.Data['SampleNames'] = {0: sample_name}

    update_console("Creating JSON file...")
    json_path = excel_path.replace('.xlsx', '.json')
    serializable_data = convert_to_serializable_and_round(window.Data)
    with open(json_path, 'w') as fh:
        json.dump(serializable_data, fh, indent=4)

    window.sheet_combobox.Clear()
    for sheet_name in sheet_names:
        window.sheet_combobox.Append(sheet_name)
    window.sheet_combobox.SetSelection(0)
    window.current_sheet = sheet_names[0]

    on_sheet_selected(window, None)
    save_state(window)

    update_console("Refreshing sheets...")
    from libraries.FileMenu.Save import refresh_sheets
    refresh_sheets(window, on_sheet_selected, update_console)

    # Refresh FileManager if open
    file_manager_was_open = False
    file_manager_position = None
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            window.file_manager.GetSize()
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
        except RuntimeError:
            window.file_manager = None

    update_console(f"\nImport complete!  ({len(sheet_names)} sheet(s))")

    if file_manager_was_open:
        from libraries.ViewMenu.FileManager import FileManagerWindow
        window.file_manager = FileManagerWindow(window)
        if file_manager_position:
            window.file_manager.SetPosition(file_manager_position)
        window.file_manager.Show()


# ---------------------------------------------------------------------------
# Unique sheet name helper
# ---------------------------------------------------------------------------

def _unique_sheet_name(core_level, sheet_names):
    """Return a unique sheet name: C1s, C1s1, C1s2, ..."""
    if core_level not in sheet_names:
        return core_level
    counter = 1
    while f"{core_level}{counter}" in sheet_names:
        counter += 1
    return f"{core_level}{counter}"


# ---------------------------------------------------------------------------
# Single-file import
# ---------------------------------------------------------------------------

def import_sdp_file(window, file_path=None, show_message=False):
    """Import an SDP binary file (single file, may contain multiple spectra)."""
    if file_path is None:
        with wx.FileDialog(window, "Open SDP file",
                           wildcard="SDP files (*.sdp;*.SDP)|*.sdp;*.SDP",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing SDP File", size=(420, 380))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 420) // 2,
        parent_pos.y + (parent_size.height - 380) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Opening: {os.path.basename(file_path)}")
        parsed = parse_sdp_file(file_path)

        meta = parsed['metadata']
        spectra = parsed['spectra']
        update_console(f"  Spectra found: {len(spectra)}")
        update_console(f"  Source: {meta['source_label']}")
        if meta['source_file']:
            update_console(f"  Origin: {os.path.basename(meta['source_file'])}")

        for s in spectra:
            update_console(f"  {s['core_level']}: {s['be_start']:.1f}-"
                           f"{s['be_end']:.1f} eV ({s['num_points']} pts)")

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}.xlsx")

        update_console("\nCreating Excel file...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []
        for spec_idx, spec in enumerate(spectra):
            sheet_name = _unique_sheet_name(spec['core_level'], sheet_names)
            sheet_names.append(sheet_name)
            update_console(f"  Processing: {sheet_name}")

            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, spec, meta, spec_idx, len(spectra))

        update_console(f"Saving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        _load_into_kherve(window, excel_path, sheet_names,
                          meta['sample_name'], update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing SDP file:\n\n{str(e)}", "Error",
                      wx.OK | wx.ICON_ERROR)


# ---------------------------------------------------------------------------
# Multi-file import
# ---------------------------------------------------------------------------

def import_multiple_sdp_files(window, file_paths=None, show_message=False):
    """Import multiple SDP files into a single Excel workbook."""
    if file_paths is None:
        with wx.FileDialog(window, "Select SDP files",
                           wildcard="SDP files (*.sdp;*.SDP)|*.sdp;*.SDP",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE
                           ) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_paths = dlg.GetPaths()

    if not file_paths:
        return

    output_dir = os.path.dirname(file_paths[0])
    first_base = os.path.splitext(os.path.basename(file_paths[0]))[0]
    default_name = f"{first_base}_combined" if len(file_paths) > 1 else first_base

    with wx.TextEntryDialog(window, "Enter name for the output file:",
                            "Output File Name", default_name) as dlg:
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        output_name = dlg.GetValue().strip() or default_name

    excel_path = os.path.join(output_dir, f"{output_name}.xlsx")

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing SDP Files", size=(420, 400),
                             style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 420) // 2,
        parent_pos.y + (parent_size.height - 400) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Importing {len(file_paths)} SDP file(s)...")
        update_console(f"Output: {output_name}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []
        first_sample_name = None

        for file_idx, fp in enumerate(file_paths):
            try:
                update_console(f"\n({file_idx + 1}/{len(file_paths)}) "
                               f"{os.path.basename(fp)}")
                parsed = parse_sdp_file(fp)

                meta = parsed['metadata']
                spectra = parsed['spectra']
                update_console(f"  Spectra in file: {len(spectra)}")

                if first_sample_name is None:
                    first_sample_name = meta['sample_name']

                for spec_idx, spec in enumerate(spectra):
                    sheet_name = _unique_sheet_name(spec['core_level'], sheet_names)
                    sheet_names.append(sheet_name)
                    update_console(f"  Sheet: {sheet_name}")

                    ws = wb.create_sheet(sheet_name)
                    _write_sheet(ws, spec, meta, spec_idx, len(spectra))

            except Exception as e:
                update_console(f"  Error: {e}")
                continue

        if not sheet_names:
            update_console("No SDP files could be processed.")
            wx.CallLater(2000, console_frame.Close)
            wx.MessageBox("No SDP files could be processed.", "Error",
                          wx.OK | wx.ICON_ERROR)
            return

        update_console(f"\nSaving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        _load_into_kherve(window, excel_path, sheet_names,
                          first_sample_name or output_name, update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing SDP files:\n\n{exc}", "Error",
                      wx.OK | wx.ICON_ERROR)