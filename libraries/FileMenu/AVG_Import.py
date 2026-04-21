# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
AVG File Import Module for KherveFitting
Thermo Scientific / VG Scienta ASCII Text Format (.avg)

AVG files are text-based exports from VGD DataSpaces. They can contain:
  - Single spectra (e.g., a simple XPS scan)
  - Multiple spectra along a second axis (depth profiles, ARXPS angle series)

The data is stored as Kinetic Energy and is converted to Binding Energy:
    BE = Source_Energy - KE

Each spectrum from the second axis becomes its own sheet:
    Core level name: e.g. Ti2p, Ti2p1, Ti2p2, ...
"""

import os
import re
import json
import numpy as np
import openpyxl
import wx


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_core_level_name(filename):
    """
    Extract clean core level name from filename.

    Examples:
        'Ti2p Snap.avg' -> 'Ti2p'
        'O1s_Scan.avg'  -> 'O1s'
        'C1s Region.avg'-> 'C1s'
    """
    base_name = os.path.splitext(filename)[0]

    suffixes_to_remove = [
        '_Snap', '_snap', '_SNAP',
        ' Snap', ' snap', ' SNAP',
        '_Scan', '_scan', '_SCAN',
        ' Scan', ' scan', ' SCAN',
        '_Region', '_region', '_REGION',
        ' Region', ' region', ' REGION',
        '_core level', '_Core Level', '_Core level',
        ' core level', ' Core Level', ' Core level',
        '_spectrum', '_Spectrum', ' spectrum', ' Spectrum',
    ]

    result = base_name
    for suffix in suffixes_to_remove:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break

    # Regex fallback: grab element + orbital at start
    match = re.match(r'^([A-Z][a-z]?\d+[spdfgh]\d*)[\s_]', result + ' ')
    if match:
        result = match.group(1)

    return result.strip()


def _get_bstr(line):
    """Extract string value from a  VT_BSTR property line."""
    m = re.search(r"VT_BSTR\s*=\s*'([^']*)'", line)
    return m.group(1) if m else ''


def _get_r4(line):
    """Extract float value from a VT_R4 property line."""
    m = re.search(r'VT_R4\s*=\s*([\d.eE+\-]+)', line)
    return float(m.group(1)) if m else None


def _get_i4(line):
    """Extract int value from a VT_I4 property line."""
    m = re.search(r'VT_I4\s*=\s*(\d+)', line)
    return int(m.group(1)) if m else None


def _get_date(line):
    """Extract date/time string from a VT_DATE property line."""
    m = re.search(r'VT_DATE\s*=\s*(.+)', line)
    return m.group(1).strip() if m else ''


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_avg_file(file_path):
    """
    Parse an AVG file and return structured data.

    Returns a dict:
        metadata        : dict of header properties
        energy_axis     : dict  {ke_start, ke_step, num_points}
        second_axis     : dict  {label, unit, axis_type, values}  or None
        spectra         : list of lists  (one list of floats per spectrum)
        num_spectra     : int
    """
    with open(file_path, 'r', encoding='latin-1') as fh:
        lines = fh.readlines()

    metadata = {
        'title': '',
        'subject': '',
        'author': '',
        'created': '',
        'saved': '',
        'instrument': '',
        'source_energy': 1486.68,
        'pass_energy': None,
        'work_function': None,
        'acq_time': None,
        'periods': None,
        'txf_coeffs': [],
        'lens_mode': '',
    }

    energy_axis = {
        'ke_start': None,
        'ke_step': None,
        'num_points': None,
    }

    second_axis = None   # populated if a non-energy second axis exists
    spectra = []         # list of flat float lists, one per spectrum
    current_spectrum = []
    in_data_block = False
    num_spectra_expected = 1
    txf_coeff_list = []

    for line in lines:
        line_s = line.rstrip()

        # ---- Metadata ----
        if 'DS_EXT_SUPROPID_TITLE' in line_s:
            metadata['title'] = _get_bstr(line_s)
        elif 'DS_EXT_SUPROPID_SUBJECT' in line_s:
            metadata['subject'] = _get_bstr(line_s)
        elif 'DS_EXT_SUPROPID_AUTHOR' in line_s:
            metadata['author'] = _get_bstr(line_s)
        elif 'DS_EXT_SUPROPID_CREATED' in line_s:
            metadata['created'] = _get_date(line_s)
        elif 'DS_EXT_SUPROPID_SAVED' in line_s:
            metadata['saved'] = _get_date(line_s)
        elif 'DS_GEPROPID_INSTRUMENT' in line_s:
            metadata['instrument'] = _get_bstr(line_s)
        elif 'DS_SOPROPID_ENERGY' in line_s:
            v = _get_r4(line_s)
            if v is not None:
                metadata['source_energy'] = v
        elif 'DS_ANPROPID_PASS' in line_s:
            v = _get_r4(line_s)
            if v is not None:
                metadata['pass_energy'] = v
        elif 'DS_ANPROPID_WORK_FTN' in line_s:
            v = _get_r4(line_s)
            if v is not None:
                metadata['work_function'] = v
        elif 'DS_ACPROPID_ACQ_TIME' in line_s:
            v = _get_r4(line_s)
            if v is not None:
                metadata['acq_time'] = v
        elif 'DS_ACPROPID_PERIODS' in line_s:
            v = _get_i4(line_s)
            if v is not None:
                metadata['periods'] = v
        elif 'DS_ANPROPID_LENS_MODE_NAME' in line_s:
            metadata['lens_mode'] = _get_bstr(line_s)
        elif 'DS_ANPROPID_TXFN_COEFF[' in line_s:
            v = _get_r4(line_s)
            if v is not None:
                txf_coeff_list.append(v)

        # ---- DATAAXES: find how many spectra ----
        elif line_s.strip().startswith('$DATAAXES='):
            # $DATAAXES=2,#empty#  or $DATAAXES=1,#empty#
            pass  # we derive count from SPACEAXES

        # ---- SPACEAXES ----
        elif line_s.strip().startswith('$SPACEAXES='):
            pass  # handled by individual axis lines below

        # ---- Space axis definitions ----
        # Format:  N=   start,  width,  numPoints,  axisType,  linear,  symbol,  unit,  label
        elif re.match(r'\s+\d+=\s+[\d.\-]+,', line_s):
            m = re.match(
                r'\s+(\d+)=\s+([\d.\-]+),\s+([\d.\-]+),\s+(\d+),\s+'
                r'(\w+),\s+\w+,\s+\'([^\']*)\',\s+\'([^\']*)\',\s+\'([^\']*)\'',
                line_s)
            if m:
                axis_idx = int(m.group(1))
                ax_start = float(m.group(2))
                ax_width = float(m.group(3))
                ax_num = int(m.group(4))
                ax_type = m.group(5)
                ax_symbol = m.group(6)
                ax_unit = m.group(7)
                ax_label = m.group(8)

                if ax_type == 'ENERGY':
                    energy_axis['ke_start'] = ax_start
                    energy_axis['ke_step'] = ax_width
                    energy_axis['num_points'] = ax_num
                else:
                    # This is the second (non-energy) axis
                    num_spectra_expected = ax_num
                    second_axis = {
                        'label': ax_label,
                        'unit': ax_unit,
                        'axis_type': ax_type,
                        'values': [],
                    }

        # ---- AXISVALUE lines: capture second-axis values ----
        elif line_s.startswith('$AXISVALUE=') and second_axis is not None:
            # $AXISVALUE= DATAXIS=1 SPACEAXIS=1 LABEL='Etch Time' POINT=0 VALUE=0.000000;
            vm = re.search(r'VALUE=([\d.\-]+)', line_s)
            lm = re.search(r"LABEL='([^']+)'", line_s)
            sm = re.search(r'SPACEAXIS=(\d+)', line_s)
            if vm and sm and int(sm.group(1)) == 1:  # only first non-energy axis
                second_axis['values'].append(float(vm.group(1)))

        # ---- DATA blocks ----
        elif line_s.startswith('$DATA='):
            # Save previous spectrum if any
            if current_spectrum:
                spectra.append(current_spectrum)
            current_spectrum = []
            in_data_block = True

        elif in_data_block and line_s.startswith('LIST@'):
            # LIST@   0=  val1,  val2,  val3,  val4
            # Extract the values after the '='
            parts = line_s.split('=', 1)
            if len(parts) == 2:
                vals = re.findall(r'[\d.\-+eE]+', parts[1])
                for v in vals:
                    try:
                        current_spectrum.append(float(v))
                    except ValueError:
                        pass

        elif in_data_block and (line_s.startswith(';') or line_s.startswith('$')):
            # End of current data block on next section header
            if current_spectrum and not line_s.startswith('$AXISVALUE') and not line_s.startswith('$DATA'):
                # Do not flush here; flush happens at next $DATA or end of file
                pass

    # Flush last spectrum
    if current_spectrum:
        spectra.append(current_spectrum)

    metadata['txf_coeffs'] = txf_coeff_list

    # If no second axis was found, ensure we have exactly 1 spectrum
    if second_axis is None:
        num_spectra_expected = max(1, len(spectra))

    return {
        'metadata': metadata,
        'energy_axis': energy_axis,
        'second_axis': second_axis,
        'spectra': spectra,
        'num_spectra': len(spectra) if spectra else num_spectra_expected,
    }


# ---------------------------------------------------------------------------
# Data calculation
# ---------------------------------------------------------------------------

def calculate_avg_data(parsed, spectrum_index=0):
    """
    Convert KE axis to BE and compute corrected intensities.

    Returns dict: be_values, ke_values, raw_intensities,
                  corrected_data, transmission_values,
                  be_start, be_end, be_step
    """
    ea = parsed['energy_axis']
    meta = parsed['metadata']

    ke_start = ea['ke_start']
    ke_step = ea['ke_step']
    num_points = ea['num_points']
    source_energy = meta['source_energy']

    raw_intensities = parsed['spectra'][spectrum_index] if spectrum_index < len(parsed['spectra']) else []

    # Trim / pad to expected number of points
    raw_intensities = list(raw_intensities[:num_points])
    while len(raw_intensities) < num_points:
        raw_intensities.append(0.0)

    ke_values = [ke_start + i * ke_step for i in range(num_points)]
    be_values = [source_energy - ke for ke in ke_values]

    be_start = be_values[0]
    be_end = be_values[-1]
    be_step = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0.0

    acq_time = meta.get('acq_time')
    periods = meta.get('periods')

    if acq_time and periods and acq_time > 0 and periods > 0:
        factor = acq_time * periods
        corrected_data = [v / factor for v in raw_intensities]
        transmission_values = [factor] * num_points
        txf_valid = True
    else:
        corrected_data = list(raw_intensities)
        transmission_values = [1.0] * num_points
        txf_valid = False

    return {
        'be_values': be_values,
        'ke_values': ke_values,
        'raw_intensities': raw_intensities,
        'corrected_data': corrected_data,
        'transmission_values': transmission_values,
        'be_start': be_start,
        'be_end': be_end,
        'be_step': be_step,
        'txf_valid': txf_valid,
    }


# ---------------------------------------------------------------------------
# Excel / JSON writing helpers
# ---------------------------------------------------------------------------

def _write_sheet(ws, core_level, calc_data, parsed, spectrum_index, base_name, sheet_name):
    """Write data + metadata columns to a worksheet."""
    meta = parsed['metadata']
    ea = parsed['energy_axis']
    source_energy = meta['source_energy']
    source_label = (
        "Al K-alpha Monochromated"
        if abs(source_energy - 1486.68) < 0.1
        else f"X-ray {source_energy:.2f} eV"
    )

    # Data columns A-D
    ws.cell(row=1, column=1, value='BE')
    ws.cell(row=1, column=2, value='Corrected Data')
    ws.cell(row=1, column=3, value='Raw Data')
    ws.cell(row=1, column=4, value='Transmission')

    for i, (be, corr, raw, trans) in enumerate(zip(
            calc_data['be_values'],
            calc_data['corrected_data'],
            calc_data['raw_intensities'],
            calc_data['transmission_values']), start=2):
        ws.cell(row=i, column=1, value=round(be, 2))
        ws.cell(row=i, column=2, value=round(corr, 2))
        ws.cell(row=i, column=3, value=round(raw, 2))
        ws.cell(row=i, column=4, value=round(trans, 2))

    # Experimental metadata in column 50 (AX)
    exp_col = 50

    second_axis = parsed.get('second_axis')
    second_axis_label = ''
    second_axis_value = ''
    if second_axis and spectrum_index < len(second_axis['values']):
        second_axis_label = f"{second_axis['label']} ({second_axis['unit']})"
        second_axis_value = f"{second_axis['values'][spectrum_index]:.4f}"

    ws.cell(row=1, column=exp_col, value='Experimental Description')

    exp_metadata = {
        'Sample ID':          meta['subject'] if meta['subject'] else base_name,
        'Title':              meta['title'],
        'Author':             meta['author'],
        'Date Created':       meta['created'].split()[0] if meta['created'] else '',
        'Time Created':       (meta['created'].split()[1]
                               if meta['created'] and len(meta['created'].split()) > 1 else ''),
        'Date Saved':         meta['saved'].split()[0] if meta['saved'] else '',
        'Time Saved':         (meta['saved'].split()[1]
                               if meta['saved'] and len(meta['saved'].split()) > 1 else ''),
        'Instrument':         meta['instrument'],
        'Lens Mode':          meta['lens_mode'],
        'Technique':          'XPS',
        'Species & Transition': core_level,
        'Spectrum Index':     str(spectrum_index),
        'Total Spectra':      str(parsed['num_spectra']),
        'Source Label':       source_label,
        'Source Energy':      f"{source_energy:.2f}",
        'Pass Energy':        (f"{meta['pass_energy']:.2f}"
                               if meta['pass_energy'] is not None else 'Unknown'),
        'Work Function':      (f"{meta['work_function']:.2f}"
                               if meta['work_function'] is not None else 'Unknown'),
        'Acq Time (s)':       (f"{meta['acq_time']:.4f}"
                               if meta['acq_time'] is not None else 'Unknown'),
        'Periods':            str(meta['periods']) if meta['periods'] is not None else 'Unknown',
        'Number of Points':   str(ea['num_points']),
        'BE Start':           f"{calc_data['be_start']:.2f}",
        'BE End':             f"{calc_data['be_end']:.2f}",
        'BE Step':            f"{abs(calc_data['be_step']):.4f}",
        'KE Start':           (f"{ea['ke_start']:.2f}"
                               if ea['ke_start'] is not None else 'Unknown'),
        'KE Step':            (f"{ea['ke_step']:.4f}"
                               if ea['ke_step'] is not None else 'Unknown'),
        'TXF Applied':        'Yes' if calc_data['txf_valid'] else 'No',
        'TXF Coefficients':   (', '.join([f"{c:.6f}" for c in meta['txf_coeffs']])
                               if meta['txf_coeffs'] else 'N/A'),
    }

    if second_axis_label:
        exp_metadata[second_axis_label] = second_axis_value

    row = 2
    for key, value in exp_metadata.items():
        ws.cell(row=row, column=exp_col, value=key)
        ws.cell(row=row, column=exp_col + 1, value=str(value))
        row += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 28
    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40


# ---------------------------------------------------------------------------
# Main import functions
# ---------------------------------------------------------------------------

def open_avg_file(window, file_path=None, show_message=False):
    """
    Import a single Thermo AVG file.

    Args:
        window:      Main KherveFitting window instance
        file_path:   Path to .avg file (None = show dialog)
        show_message: Show success message box after import
    """
    if file_path is None:
        with wx.FileDialog(
                window, "Open AVG file",
                wildcard="AVG files (*.avg;*.AVG)|*.avg;*.AVG",
                style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing AVG File", size=(420, 360))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 420) // 2,
        parent_pos.y + (parent_size.height - 360) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        core_level = extract_core_level_name(os.path.basename(file_path))

        update_console(f"Opening: {os.path.basename(file_path)}")
        parsed = parse_avg_file(file_path)

        num_spectra = parsed['num_spectra']
        ea = parsed['energy_axis']
        meta = parsed['metadata']

        update_console(f"  Core level:       {core_level}")
        update_console(f"  Points/spectrum:  {ea['num_points']}")
        update_console(f"  Spectra found:    {num_spectra}")
        update_console(f"  Source energy:    {meta['source_energy']:.2f} eV")
        if meta['pass_energy'] is not None:
            update_console(f"  Pass energy:      {meta['pass_energy']:.2f} eV")
        if meta['acq_time'] is not None and meta['periods'] is not None:
            update_console(f"  Acq time:         {meta['acq_time']:.4f} s  x {meta['periods']} periods")

        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}.xlsx")

        update_console("\nCreating Excel file...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []

        for idx in range(num_spectra):
            calc_data = calculate_avg_data(parsed, idx)

            # Naming: Ti2p, Ti2p1, Ti2p2, ...
            sheet_name = core_level if idx == 0 else f"{core_level}{idx}"
            sheet_names.append(sheet_name)
            update_console(f"  Processing sheet: {sheet_name}")

            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, core_level, calc_data, parsed, idx, base_name, sheet_name)

        update_console(f"Saving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        _load_into_kherve(window, excel_path, sheet_names,
                          meta['subject'] or base_name,
                          update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing AVG file:\n\n{exc}", "Error", wx.OK | wx.ICON_ERROR)


def open_avg_file_direct(window, file_path, update_console=None):
    """
    Import an AVG file directly (used from FileManager drag-and-drop etc.).
    Returns True on success.
    """
    if update_console is None:
        def update_console(msg):
            print(msg)

    try:
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        core_level = extract_core_level_name(os.path.basename(file_path))

        update_console(f"Opening: {os.path.basename(file_path)}")
        parsed = parse_avg_file(file_path)

        num_spectra = parsed['num_spectra']
        meta = parsed['metadata']
        ea = parsed['energy_axis']

        update_console(f"  Core level: {core_level}, Spectra: {num_spectra}")

        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sheet_names = []

        for idx in range(num_spectra):
            calc_data = calculate_avg_data(parsed, idx)
            sheet_name = core_level if idx == 0 else f"{core_level}{idx}"
            sheet_names.append(sheet_name)
            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, core_level, calc_data, parsed, idx, base_name, sheet_name)

        wb.save(excel_path)
        update_console(f"  Saved: {os.path.basename(excel_path)} ({len(sheet_names)} sheets)")

        _load_into_kherve(window, excel_path, sheet_names,
                          meta['subject'] or base_name, update_console)
        return True

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"  Error: {exc}")
        return False


def import_multiple_avg_files(window, file_paths=None, show_message=False):
    """
    Import multiple AVG files and merge them into a single Excel workbook.

    Args:
        window:      Main KherveFitting window instance
        file_paths:  List of .avg file paths (None = show dialog)
        show_message: Show success message box after import
    """
    if file_paths is None:
        with wx.FileDialog(
                window, "Select AVG files",
                wildcard="AVG files (*.avg;*.AVG)|*.avg;*.AVG",
                style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_paths = dlg.GetPaths()

    if not file_paths:
        return

    output_dir = os.path.dirname(file_paths[0])
    first_base = os.path.splitext(os.path.basename(file_paths[0]))[0]
    default_name = f"{first_base}_combined" if len(file_paths) > 1 else first_base

    with wx.TextEntryDialog(
            window, "Enter name for the output file:",
            "Output File Name", default_name) as dlg:
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        output_name = dlg.GetValue().strip() or default_name

    excel_path = os.path.join(output_dir, f"{output_name}.xlsx")

    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing AVG Files", size=(420, 400),
                             style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 420) // 2,
        parent_pos.y + (parent_size.height - 400) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Importing {len(file_paths)} AVG file(s)...")
        update_console(f"Output: {output_name}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []
        first_sample_name = None

        for file_idx, fp in enumerate(file_paths):
            try:
                base_name = os.path.splitext(os.path.basename(fp))[0]
                core_level = extract_core_level_name(os.path.basename(fp))

                update_console(f"\n({file_idx + 1}/{len(file_paths)}) {os.path.basename(fp)}")
                parsed = parse_avg_file(fp)

                num_spectra = parsed['num_spectra']
                meta = parsed['metadata']
                update_console(f"  Core level: {core_level}, Spectra: {num_spectra}")

                if first_sample_name is None:
                    first_sample_name = meta['subject'] or base_name

                for idx in range(num_spectra):
                    calc_data = calculate_avg_data(parsed, idx)

                    # Unique sheet name
                    candidate = core_level if idx == 0 else f"{core_level}{idx}"
                    if candidate not in sheet_names:
                        sheet_name = candidate
                    else:
                        counter = 1
                        while f"{core_level}{counter}" in sheet_names:
                            counter += 1
                        sheet_name = f"{core_level}{counter}"

                    sheet_names.append(sheet_name)
                    update_console(f"  Sheet: {sheet_name}")

                    ws = wb.create_sheet(sheet_name)
                    _write_sheet(ws, core_level, calc_data, parsed, idx, base_name, sheet_name)

            except Exception as e:
                update_console(f"  Error in file: {e}")
                continue

        if not sheet_names:
            update_console("No files could be processed.")
            wx.CallLater(2000, console_frame.Close)
            wx.MessageBox("No AVG files could be processed.", "Error", wx.OK | wx.ICON_ERROR)
            return

        update_console(f"\nSaving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        _load_into_kherve(window, excel_path, sheet_names,
                          first_sample_name or output_name, update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing AVG files:\n\n{exc}", "Error", wx.OK | wx.ICON_ERROR)


# ---------------------------------------------------------------------------
# Shared KherveFitting loader
# ---------------------------------------------------------------------------

def _load_into_kherve(window, excel_path, sheet_names, sample_name, update_console):
    """
    Load a completed Excel file into the KherveFitting data model, save JSON,
    update the UI combobox, and refresh sheet display.
    """
    from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
    from libraries.FileMenu.Save import (update_undo_redo_state, save_state,
                                         convert_to_serializable_and_round)
    from libraries.Sheet_Operations import on_sheet_selected

    update_console("\nLoading into KherveFitting...")

    window.Data = Init_Measurement_Data(window)
    window.Data['FilePath'] = excel_path

    window.history = []
    window.redo_stack = []
    update_undo_redo_state(window)

    window.results_grid.ClearGrid()
    if window.results_grid.GetNumberRows() > 0:
        window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

    for sheet_name in sheet_names:
        add_core_level_Data(window.Data, window, excel_path, sheet_name)
        update_console(f"  Imported: {sheet_name}")

    window.Data['SampleNames'] = {0: sample_name}

    update_console("Creating JSON file...")
    json_path = excel_path.replace('.xlsx', '.json')
    serializable_data = convert_to_serializable_and_round(window.Data)
    with open(json_path, 'w') as fh:
        json.dump(serializable_data, fh, indent=4)

    window.sheet_combobox.Clear()
    for sheet_name in sheet_names:
        window.sheet_combobox.Append(sheet_name)
    window.sheet_combobox.SetSelection(0)
    window.current_sheet = sheet_names[0]

    on_sheet_selected(window, None)
    save_state(window)

    update_console("Refreshing sheets...")
    from libraries.FileMenu.Save import refresh_sheets
    refresh_sheets(window, on_sheet_selected, update_console)

    # Refresh FileManager if open
    file_manager_was_open = False
    file_manager_position = None
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            window.file_manager.GetSize()
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
        except RuntimeError:
            window.file_manager = None

    update_console(f"\nImport complete!  ({len(sheet_names)} sheet(s))")

    if file_manager_was_open:
        from libraries.ViewMenu.FileManager import FileManagerWindow
        window.file_manager = FileManagerWindow(window)
        if file_manager_position:
            window.file_manager.SetPosition(file_manager_position)
        window.file_manager.Show()