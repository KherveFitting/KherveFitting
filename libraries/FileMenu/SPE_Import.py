# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

import os
import re
import struct
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import wx


def open_spe_file(window, file_path):
    """
    Process a PHI SPE file and convert it to Excel format.

    Supports multiple PHI instrument formats:
    - PHI VersaProbe III: float32 (4 bytes), data after metadata header
    - PHI X-tool: float64 (8 bytes), data after metadata header
    - PHI Quantum 2000: float64 (8 bytes), data at END of binary section
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Read file content
        with open(file_path, 'rb') as f:
            content = f.read()

        # Extract header
        header_match = re.search(rb'SOFH(.*?)EOFH', content, re.DOTALL)
        if not header_match:
            raise ValueError("Cannot find header section (SOFH...EOFH)")

        header_text = header_match.group(1).decode('utf-8', errors='ignore')
        header_lines = [line.strip() for line in header_text.strip().split('\n')]

        # Extract intensity calibration coefficients
        a, b = 31.826, 0.229  # Default values
        for line in header_lines:
            if 'IntensityCalCoeff:' in line:
                try:
                    _, coeffs = line.split(':', 1)
                    parts = coeffs.strip().split()
                    if len(parts) >= 2:
                        a = float(parts[0])
                        b = float(parts[1])
                        print(f"Extracted transmission function coefficients: a={a}, b={b}")
                except:
                    pass
                break

        # Extract source energy
        source_energy = 1486.6  # Default Al K-alpha
        for line in header_lines:
            if 'XraySource:' in line:
                if 'Al' in line:
                    source_energy = 1486.6
                elif 'Mg' in line:
                    source_energy = 1253.6
                # Try to extract exact value if present
                try:
                    parts = line.split()
                    for part in parts:
                        try:
                            val = float(part)
                            if 1200 < val < 1600:
                                source_energy = val
                                break
                        except:
                            pass
                except:
                    pass
                break

        # Get instrument model
        instrument_model = "Unknown"
        for line in header_lines:
            if 'InstrumentModel:' in line:
                instrument_model = line.split(':', 1)[1].strip()
                print(f"Instrument: {instrument_model}")
                break

        # Find active regions - use SpectralRegDef: (not Full or 2:)
        regions = []
        for line in header_lines:
            if line.startswith('SpectralRegDef:') and 'Full' not in line and '2:' not in line:
                parts = line.split()
                if len(parts) >= 9:
                    region = {
                        'number': int(parts[1]),
                        'name': parts[3],
                        'atomic_number': int(parts[4]) if parts[4].lstrip('-').isdigit() else 0,
                        'num_points': int(parts[5]),
                        'step': float(parts[6]),
                        'start_energy': float(parts[7]),
                        'end_energy': float(parts[8])
                    }

                    # Normalize region name
                    if region['name'] == 'Su1s':
                        region['name'] = 'Survey'

                    regions.append(region)

        if not regions:
            raise ValueError("No spectral regions found in file")

        # Extract metadata for experimental description
        metadata = {
            'Sample ID': os.path.basename(file_path),
            'Date': '1970/1/1',
            'Time': '0:0:0',
            'Technique': 'XPS',
            'Instrument': instrument_model,
            'Species & Transition': '',
            'Number of scans': '1',
            'Source Label': 'Al' if source_energy > 1400 else 'Mg',
            'Source Energy': str(source_energy),
            'Source width X': '100',
            'Source width Y': '100',
            'Pass Energy': '224',
            'Work Function': '4.339',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': 'Kinetic Energy',
            'X Units': 'eV',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '1',
            'Collection Time': '0.16',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Update metadata from header
        for line in header_lines:
            if 'FileDateTime:' in line:
                parts = line.split(':', 1)[1].strip().split()
                if len(parts) >= 2:
                    metadata['Date'] = parts[0]
                    metadata['Time'] = parts[1]
            elif 'FileDate:' in line:
                metadata['Date'] = line.split(':', 1)[1].strip().replace(' ', '/')
            elif 'PassEnergy:' in line:
                try:
                    metadata['Pass Energy'] = line.split(':', 1)[1].strip().split()[0]
                except:
                    pass
            elif 'AnalyserWorkFcn:' in line or 'WorkFunction:' in line:
                try:
                    metadata['Work Function'] = line.split(':', 1)[1].strip().split()[0]
                except:
                    pass
            elif 'AnalyserMode:' in line:
                metadata['Analyzer Mode'] = line.split(':', 1)[1].strip()
            elif 'XrayBeamDiameter:' in line:
                try:
                    size = line.split(':', 1)[1].strip().split()[0]
                    metadata['Source width X'] = size
                    metadata['Source width Y'] = size
                except:
                    pass
            elif 'Comments:' in line:
                metadata['Block Comment'] = line.split(':', 1)[1].strip()
                metadata['# Comment Lines'] = '1'

        # Get binary data section
        data_start = content.find(b'EOFH') + 4
        binary_data = content[data_start:]

        # Calculate total points
        total_points = sum(r['num_points'] for r in regions)

        # =====================================================================
        # AUTO-DETECT DATA FORMAT AND OFFSET
        # =====================================================================

        def try_find_offset_standard(binary_data, first_region, dtype, bytes_per_val):
            """Try to find data offset for standard formats (VersaProbe, X-tool)."""
            fmt = '<f' if dtype == 'float32' else '<d'
            num_points = first_region['num_points']

            first_region_size = num_points * bytes_per_val
            max_offset = len(binary_data) - first_region_size
            if max_offset < 0:
                return None
            max_offset = min(2000, max_offset)

            for offset in range(0, max_offset + 1, 2):
                try:
                    values = []
                    valid = True
                    for i in range(min(20, num_points)):
                        byte_pos = offset + i * bytes_per_val
                        if byte_pos + bytes_per_val > len(binary_data):
                            valid = False
                            break
                        val = struct.unpack(fmt, binary_data[byte_pos:byte_pos + bytes_per_val])[0]
                        if not (1 < val < 1e8):
                            valid = False
                            break
                        values.append(val)

                    if valid and len(values) >= 10 and max(values) - min(values) > 1:
                        # Verify all points in first region
                        all_valid = True
                        for i in range(num_points):
                            byte_pos = offset + i * bytes_per_val
                            if byte_pos + bytes_per_val > len(binary_data):
                                all_valid = False
                                break
                            val = struct.unpack(fmt, binary_data[byte_pos:byte_pos + bytes_per_val])[0]
                            if not (0 <= val < 1e8):
                                all_valid = False
                                break

                        if all_valid:
                            return offset
                except:
                    pass

            return None

        def try_quantum2000_format(binary_data, first_region, total_points):
            """Try Quantum 2000 format: float64 data at END of binary section."""
            expected_size = total_points * 8
            if expected_size > len(binary_data):
                return None

            offset = len(binary_data) - expected_size
            num_points = first_region['num_points']

            try:
                values = []
                valid = True
                for i in range(min(20, num_points)):
                    byte_pos = offset + i * 8
                    val = struct.unpack('<d', binary_data[byte_pos:byte_pos + 8])[0]
                    if not (0 < val < 1e8):
                        valid = False
                        break
                    values.append(val)

                if valid and len(values) >= 10 and max(values) - min(values) > 1:
                    return offset
            except:
                pass

            return None

        # Try different formats
        data_type = None
        bytes_per_val = None
        first_offset = None

        # Method 1: Try float32 (VersaProbe format)
        offset = try_find_offset_standard(binary_data, regions[0], 'float32', 4)
        if offset is not None:
            data_type = 'float32'
            bytes_per_val = 4
            first_offset = offset
            print(f"Detected VersaProbe format (float32) at offset {offset}")

        # Method 2: Try float64 standard (X-tool format)
        if first_offset is None:
            offset = try_find_offset_standard(binary_data, regions[0], 'float64', 8)
            if offset is not None:
                data_type = 'float64'
                bytes_per_val = 8
                first_offset = offset
                print(f"Detected X-tool format (float64) at offset {offset}")

        # Method 3: Try Quantum 2000 format (float64 at end)
        if first_offset is None:
            offset = try_quantum2000_format(binary_data, regions[0], total_points)
            if offset is not None:
                data_type = 'float64'
                bytes_per_val = 8
                first_offset = offset
                print(f"Detected Quantum 2000 format (float64 at end) at offset {offset}")

        if first_offset is None:
            raise ValueError("Could not detect data format or find valid data offset")

        # Calculate offsets for all regions (contiguous data)
        region_offsets = []
        current_offset = first_offset
        for region in regions:
            region_offsets.append(current_offset)
            current_offset += region['num_points'] * bytes_per_val

        print(f"Data type: {data_type}, Bytes per value: {bytes_per_val}")
        print(f"Region offsets: {region_offsets}")

        # =====================================================================
        # READ DATA AND CREATE EXCEL
        # =====================================================================

        fmt = '<f' if data_type == 'float32' else '<d'

        # Process each region
        for i, region in enumerate(regions):
            sheet_name = region['name']
            num_points = region['num_points']
            offset = region_offsets[i]

            # Read intensity values
            intensity_values = []
            valid_count = 0
            for j in range(num_points):
                data_offset = offset + (j * bytes_per_val)
                if data_offset + bytes_per_val <= len(binary_data):
                    try:
                        val = struct.unpack(fmt, binary_data[data_offset:data_offset + bytes_per_val])[0]
                        if 0 <= val < 1e8:
                            intensity_values.append(val)
                            valid_count += 1
                        else:
                            intensity_values.append(0.0)
                    except:
                        intensity_values.append(0.0)
                else:
                    intensity_values.append(0.0)

            # Skip regions with mostly invalid data (< 50% valid)
            if valid_count < num_points * 0.5:
                print(f"Skipping {sheet_name}: only {valid_count}/{num_points} valid points")
                continue

            # Make sure we have the expected number of points
            if len(intensity_values) < num_points:
                intensity_values.extend([0.0] * (num_points - len(intensity_values)))

            # Create sheet (max 31 chars for sheet name)
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy"
            ws["B1"] = "Raw Data"

            # Update region-specific metadata
            region_metadata = metadata.copy()
            region_metadata['Species & Transition'] = region['name']
            region_metadata['X Start'] = f"{region['start_energy']:.2f}"
            region_metadata['X Step'] = f"{region['step']:.4f}"
            region_metadata['Num Y Values'] = str(num_points)

            # Calculate energy values and transmission
            energy_values = np.linspace(region['start_energy'], region['end_energy'], num_points)
            ke_values = source_energy - energy_values
            # Protect against negative or zero KE values
            ke_values = np.maximum(ke_values, 1.0)
            transmission = a * np.power(ke_values, -b)
            corrected_intensity = np.array(intensity_values) / transmission

            # Write data with .2f format
            for j, (e, ci) in enumerate(zip(energy_values, corrected_intensity), start=2):
                ws[f"A{j}"] = f"{float(e):.2f}"
                ws[f"B{j}"] = f"{float(ci):.2f}"

            # Set column widths
            for col in ['A', 'B']:
                ws.column_dimensions[col].width = 18

            # Add experimental description at column AX (50)
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            # Add metadata fields
            row = 2
            for key, value in region_metadata.items():
                ws.cell(row=row, column=exp_col, value=key)
                ws.cell(row=row, column=exp_col + 1, value=str(value))
                row += 1

            # Set column widths for experimental description
            ws.column_dimensions[get_column_letter(exp_col)].width = 25
            ws.column_dimensions[get_column_letter(exp_col + 1)].width = 40

            print(f"Exported {sheet_name}: {num_points} points, {valid_count} valid")

        # Check if any sheets were created
        if len(wb.sheetnames) == 0:
            raise ValueError("No valid data regions found in file")

        # Create Experimental description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata to the experimental description sheet
        row = 1
        exp_sheet.cell(row=row, column=1, value="Experimental Description")
        row += 1

        for key, value in metadata.items():
            exp_sheet.cell(row=row, column=1, value=key)
            exp_sheet.cell(row=row, column=2, value=str(value))
            row += 1

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        print(f"Successfully saved: {excel_path}")

        # Open the Excel file in KherveFitting
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")
        return False


def open_spe_file_dialog(window):
    """Open file dialog for SPE files."""
    with wx.FileDialog(window, "Open SPE file", wildcard="PHI files (*.spe)|*.spe",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_spe_file(window, file_path)




# OLD Data from Open.py

def open_spe_file_OPEN_PY(window, file_path):
    """
    Process a PHI SPE file and convert it to Excel format.

    Supports multiple PHI instrument formats:
    - PHI VersaProbe III: float32 (4 bytes), data after metadata header
    - PHI X-tool: float64 (8 bytes), data after metadata header
    - PHI Quantum 2000: float64 (8 bytes), data at END of binary section
    """
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter
    import re
    import os
    import struct

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Read file content
        with open(file_path, 'rb') as f:
            content = f.read()

        # Extract header
        header_match = re.search(rb'SOFH(.*?)EOFH', content, re.DOTALL)
        if not header_match:
            raise ValueError("Cannot find header section (SOFH...EOFH)")

        header_text = header_match.group(1).decode('utf-8', errors='ignore')
        header_lines = [line.strip() for line in header_text.strip().split('\n')]

        # Extract intensity calibration coefficients
        a, b = 31.826, 0.229  # Default values
        for line in header_lines:
            if 'IntensityCalCoeff:' in line:
                try:
                    _, coeffs = line.split(':', 1)
                    parts = coeffs.strip().split()
                    if len(parts) >= 2:
                        a = float(parts[0])
                        b = float(parts[1])
                        print(f"Extracted transmission function coefficients: a={a}, b={b}")
                except:
                    pass
                break

        # Extract source energy
        source_energy = 1486.6  # Default Al K-alpha
        for line in header_lines:
            if 'XraySource:' in line:
                if 'Al' in line:
                    source_energy = 1486.6
                elif 'Mg' in line:
                    source_energy = 1253.6
                # Try to extract exact value if present
                try:
                    parts = line.split()
                    for part in parts:
                        try:
                            val = float(part)
                            if 1200 < val < 1600:
                                source_energy = val
                                break
                        except:
                            pass
                except:
                    pass
                break

        # Get instrument model
        instrument_model = "Unknown"
        for line in header_lines:
            if 'InstrumentModel:' in line:
                instrument_model = line.split(':', 1)[1].strip()
                print(f"Instrument: {instrument_model}")
                break

        # Find active regions - use SpectralRegDef: (not Full or 2:)
        regions = []
        for line in header_lines:
            if line.startswith('SpectralRegDef:') and 'Full' not in line and '2:' not in line:
                parts = line.split()
                if len(parts) >= 9:
                    region = {
                        'number': int(parts[1]),
                        'name': parts[3],
                        'atomic_number': int(parts[4]) if parts[4].lstrip('-').isdigit() else 0,
                        'num_points': int(parts[5]),
                        'step': float(parts[6]),
                        'start_energy': float(parts[7]),
                        'end_energy': float(parts[8])
                    }

                    # Normalize region name
                    if region['name'] == 'Su1s':
                        region['name'] = 'Survey'

                    regions.append(region)

        if not regions:
            raise ValueError("No spectral regions found in file")

        # Extract metadata for experimental description
        metadata = {
            'Sample ID': os.path.basename(file_path),
            'Date': '1970/1/1',
            'Time': '0:0:0',
            'Technique': 'XPS',
            'Instrument': instrument_model,
            'Species & Transition': '',
            'Number of scans': '1',
            'Source Label': 'Al' if source_energy > 1400 else 'Mg',
            'Source Energy': str(source_energy),
            'Source width X': '100',
            'Source width Y': '100',
            'Pass Energy': '224',
            'Work Function': '4.339',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': 'Kinetic Energy',
            'X Units': 'eV',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '1',
            'Collection Time': '0.16',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Update metadata from header
        for line in header_lines:
            if 'FileDateTime:' in line:
                parts = line.split(':', 1)[1].strip().split()
                if len(parts) >= 2:
                    metadata['Date'] = parts[0]
                    metadata['Time'] = parts[1]
            elif 'FileDate:' in line:
                metadata['Date'] = line.split(':', 1)[1].strip().replace(' ', '/')
            elif 'PassEnergy:' in line:
                try:
                    metadata['Pass Energy'] = line.split(':', 1)[1].strip().split()[0]
                except:
                    pass
            elif 'AnalyserWorkFcn:' in line or 'WorkFunction:' in line:
                try:
                    metadata['Work Function'] = line.split(':', 1)[1].strip().split()[0]
                except:
                    pass
            elif 'AnalyserMode:' in line:
                metadata['Analyzer Mode'] = line.split(':', 1)[1].strip()
            elif 'XrayBeamDiameter:' in line:
                try:
                    size = line.split(':', 1)[1].strip().split()[0]
                    metadata['Source width X'] = size
                    metadata['Source width Y'] = size
                except:
                    pass
            elif 'Comments:' in line:
                metadata['Block Comment'] = line.split(':', 1)[1].strip()
                metadata['# Comment Lines'] = '1'

        # Get binary data section
        data_start = content.find(b'EOFH') + 4
        binary_data = content[data_start:]

        # Calculate total points
        total_points = sum(r['num_points'] for r in regions)

        # =====================================================================
        # AUTO-DETECT DATA FORMAT AND OFFSET
        # =====================================================================

        def try_find_offset_standard(binary_data, first_region, dtype, bytes_per_val):
            """Try to find data offset for standard formats (VersaProbe, X-tool)."""
            fmt = '<f' if dtype == 'float32' else '<d'
            num_points = first_region['num_points']

            first_region_size = num_points * bytes_per_val
            max_offset = len(binary_data) - first_region_size
            if max_offset < 0:
                return None
            max_offset = min(2000, max_offset)

            for offset in range(0, max_offset + 1, 2):
                try:
                    values = []
                    valid = True
                    for i in range(min(20, num_points)):
                        byte_pos = offset + i * bytes_per_val
                        if byte_pos + bytes_per_val > len(binary_data):
                            valid = False
                            break
                        val = struct.unpack(fmt, binary_data[byte_pos:byte_pos + bytes_per_val])[0]
                        if not (1 < val < 1e8):
                            valid = False
                            break
                        values.append(val)

                    if valid and len(values) >= 10 and max(values) - min(values) > 1:
                        # Verify all points in first region
                        all_valid = True
                        for i in range(num_points):
                            byte_pos = offset + i * bytes_per_val
                            if byte_pos + bytes_per_val > len(binary_data):
                                all_valid = False
                                break
                            val = struct.unpack(fmt, binary_data[byte_pos:byte_pos + bytes_per_val])[0]
                            if not (0 <= val < 1e8):
                                all_valid = False
                                break

                        if all_valid:
                            return offset
                except:
                    pass

            return None

        def try_quantum2000_format(binary_data, first_region, total_points):
            """Try Quantum 2000 format: float64 data at END of binary section."""
            expected_size = total_points * 8
            if expected_size > len(binary_data):
                return None

            offset = len(binary_data) - expected_size
            num_points = first_region['num_points']

            try:
                values = []
                valid = True
                for i in range(min(20, num_points)):
                    byte_pos = offset + i * 8
                    val = struct.unpack('<d', binary_data[byte_pos:byte_pos + 8])[0]
                    if not (0 < val < 1e8):
                        valid = False
                        break
                    values.append(val)

                if valid and len(values) >= 10 and max(values) - min(values) > 1:
                    return offset
            except:
                pass

            return None

        # Try different formats
        data_type = None
        bytes_per_val = None
        first_offset = None

        # Method 1: Try float32 (VersaProbe format)
        offset = try_find_offset_standard(binary_data, regions[0], 'float32', 4)
        if offset is not None:
            data_type = 'float32'
            bytes_per_val = 4
            first_offset = offset
            print(f"Detected VersaProbe format (float32) at offset {offset}")

        # Method 2: Try float64 standard (X-tool format)
        if first_offset is None:
            offset = try_find_offset_standard(binary_data, regions[0], 'float64', 8)
            if offset is not None:
                data_type = 'float64'
                bytes_per_val = 8
                first_offset = offset
                print(f"Detected X-tool format (float64) at offset {offset}")

        # Method 3: Try Quantum 2000 format (float64 at end)
        if first_offset is None:
            offset = try_quantum2000_format(binary_data, regions[0], total_points)
            if offset is not None:
                data_type = 'float64'
                bytes_per_val = 8
                first_offset = offset
                print(f"Detected Quantum 2000 format (float64 at end) at offset {offset}")

        if first_offset is None:
            raise ValueError("Could not detect data format or find valid data offset")

        # Calculate offsets for all regions (contiguous data)
        region_offsets = []
        current_offset = first_offset
        for region in regions:
            region_offsets.append(current_offset)
            current_offset += region['num_points'] * bytes_per_val

        print(f"Data type: {data_type}, Bytes per value: {bytes_per_val}")
        print(f"Region offsets: {region_offsets}")

        # =====================================================================
        # READ DATA AND CREATE EXCEL
        # =====================================================================

        fmt = '<f' if data_type == 'float32' else '<d'

        # Process each region
        for i, region in enumerate(regions):
            sheet_name = region['name']
            num_points = region['num_points']
            offset = region_offsets[i]

            # Read intensity values
            intensity_values = []
            valid_count = 0
            for j in range(num_points):
                data_offset = offset + (j * bytes_per_val)
                if data_offset + bytes_per_val <= len(binary_data):
                    try:
                        val = struct.unpack(fmt, binary_data[data_offset:data_offset + bytes_per_val])[0]
                        if 0 <= val < 1e8:
                            intensity_values.append(val)
                            valid_count += 1
                        else:
                            intensity_values.append(0.0)
                    except:
                        intensity_values.append(0.0)
                else:
                    intensity_values.append(0.0)

            # Skip regions with mostly invalid data (< 50% valid)
            if valid_count < num_points * 0.5:
                print(f"Skipping {sheet_name}: only {valid_count}/{num_points} valid points")
                continue

            # Make sure we have the expected number of points
            if len(intensity_values) < num_points:
                intensity_values.extend([0.0] * (num_points - len(intensity_values)))

            # Create sheet (max 31 chars for sheet name)
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy (eV)"
            ws["B1"] = "Corrected Data"
            ws["C1"] = "Raw Data"
            ws["D1"] = "Transmission"

            # Update region-specific metadata
            region_metadata = metadata.copy()
            region_metadata['Species & Transition'] = region['name']
            region_metadata['X Start'] = f"{region['start_energy']:.2f}"
            region_metadata['X Step'] = f"{region['step']:.4f}"
            region_metadata['Num Y Values'] = str(num_points)

            # Calculate energy values and transmission
            energy_values = np.linspace(region['start_energy'], region['end_energy'], num_points)
            ke_values = source_energy - energy_values
            # Protect against negative or zero KE values
            ke_values = np.maximum(ke_values, 1.0)
            transmission = a * np.power(ke_values, -b)
            corrected_intensity = np.array(intensity_values) / transmission

            # Write data with .2f format
            for j, (e, ci, ri, t) in enumerate(zip(energy_values, corrected_intensity,
                                                   intensity_values, transmission), start=2):
                ws[f"A{j}"] = round(float(e), 2)
                ws[f"B{j}"] = round(float(ci), 2)
                ws[f"C{j}"] = round(float(ri), 2)
                ws[f"D{j}"] = round(float(t), 2)

            # Set column widths
            for col in ['A', 'B', 'C', 'D']:
                ws.column_dimensions[col].width = 18

            # Add experimental description at column AX (50)
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            # Add metadata fields
            row = 2
            for key, value in region_metadata.items():
                ws.cell(row=row, column=exp_col, value=key)
                ws.cell(row=row, column=exp_col + 1, value=str(value))
                row += 1

            # Set column widths for experimental description
            ws.column_dimensions[get_column_letter(exp_col)].width = 25
            ws.column_dimensions[get_column_letter(exp_col + 1)].width = 40

            print(f"Exported {sheet_name}: {num_points} points, {valid_count} valid")

        # Check if any sheets were created
        if len(wb.sheetnames) == 0:
            raise ValueError("No valid data regions found in file")

        # Create Experimental description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata to the experimental description sheet
        row = 1
        exp_sheet.cell(row=row, column=1, value="Experimental Description")
        row += 1

        for key, value in metadata.items():
            exp_sheet.cell(row=row, column=1, value=key)
            exp_sheet.cell(row=row, column=2, value=str(value))
            row += 1

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        print(f"Successfully saved: {excel_path}")

        # Open the Excel file in KherveFitting
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")
        return False

def open_spe_file_dialog_OPEN_PY(window):
    with wx.FileDialog(window, "Open SPE file", wildcard="PHI files (*.spe)|*.spe",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_spe_file(window, file_path)