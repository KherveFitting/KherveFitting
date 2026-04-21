# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
XAS File Import Module for KherveFitting
Diamond Light Source B07 Beamline - NEXAFS / XAS Data

Supports two Diamond-B07 XAS file variants:

Format A  -  Standard 4-column tab-separated  (*.txt)
  Header row + tab-separated data:
    col 0  =  Photon Energy
    col 1  =  (unused)
    col 2  =  (unused)
    col 3  =  Raw Signal   <-- last column when 4 cols

Format B  -  3-column tab-separated  (*.dat, e.g. OCP.dat)
  Header row (e.g. "pgm_energy  ca18b  ca36b") + tab-separated data:
    col 0  =  Photon Energy
    col 1  =  (unused)
    col 2  =  Raw Signal   <-- last column when 3 cols

Both formats are auto-detected by the number of columns in the first data row.
In both cases:  X = first column,  Y = last column.

Intensity normalisation (same for both formats):
    intensity = (raw_value / first_raw_value) * 10000

This always yields values in a human-readable range regardless of the
absolute scale of the raw signal.

Sheet naming (KherveFitting XAS convention):
    Single file   ->  XAS~Co-L~
    Multi-file    ->  XAS~Co-L~0,  XAS~Co-L~1, ...

Column headers written to Excel:
    A1 = "Photon Energy"
    B1 = "Raw Data"

Public API (mirrors VGD_Import / AVG_Import conventions):
    import_xas_file(window, file_path=None)
    import_xas_file_direct(window, file_path, ask_edge=False, edge_name=None)
    import_multiple_xas_files(window, file_paths=None)
"""

import json
import os
import openpyxl
import wx

from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
from libraries.FileMenu.Save import (update_undo_redo_state, save_state,
                                     convert_to_serializable_and_round)
from libraries.Sheet_Operations import on_sheet_selected


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_xas_file(file_path):
    """
    Parse a Diamond-B07 XAS file and return (photon_energy, intensity) pairs.

    Rules:
      - Skip the first (header) line
      - Skip blank lines
      - X = first column,  Y = last column
      - Intensity is normalised: (raw / first_raw_value) * 10000

    Returns:
        list of [photon_energy: float, intensity: float]

    Raises:
        ValueError  if no valid data rows are found
    """
    raw_data = []

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as fh:
        lines = fh.readlines()

    # Skip header line (line 0), iterate from line 1 onward
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        parts = line.split('\t')
        if len(parts) < 2:
            # Try whitespace split as fallback
            parts = line.split()
        if len(parts) < 2:
            continue
        try:
            energy = float(parts[0])       # first column  = Photon Energy
            raw    = float(parts[-1])      # last  column  = Raw Signal
            raw_data.append([energy, raw])
        except ValueError:
            continue

    if not raw_data:
        raise ValueError("No valid numeric data rows found in the file.")

    # Normalise: divide every raw value by the first raw value, scale to 10000
    first_raw = raw_data[0][1]
    if first_raw == 0:
        raise ValueError("First raw intensity value is zero; cannot normalise.")

    data = [[e, (r / first_raw) * 10000] for e, r in raw_data]
    return data


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _write_xas_sheet(ws, data):
    """Write Photon Energy / Raw Data + XAS technique tag to an openpyxl sheet."""
    ws["A1"]  = "Photon Energy"
    ws["B1"]  = "Raw Data"
    ws["AX1"] = "Technique"
    ws["AX2"] = "XAS"
    for row_idx, (energy, intensity) in enumerate(data, start=2):
        ws[f"A{row_idx}"] = float(f"{energy:.2f}")
        ws[f"B{row_idx}"] = float(f"{intensity:.2f}")


def _make_sheet_name(edge_name, index=None):
    """
    Build the KherveFitting XAS sheet name.
        single file  ->  XAS~Co-L~
        multi-file   ->  XAS~Co-L~0
    """
    base = f"XAS~{edge_name}~" if edge_name else "XAS~"
    return base if index is None else f"{base}{index}"


def _strip_nexafs(name):
    """Remove _NEXAFS / -NEXAFS suffix from a sample name."""
    for suffix in ('_NEXAFS', '-NEXAFS', '_nexafs', '-nexafs'):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def _ask_edge(window, default="Co-L"):
    """Prompt the user for the absorption edge label."""
    dlg = wx.TextEntryDialog(
        window,
        "Enter the transition edge (e.g., Co-L, Fe-K, O-K):",
        "XAS Edge Information",
        default,
    )
    if dlg.ShowModal() == wx.ID_OK:
        edge = dlg.GetValue().strip()
        dlg.Destroy()
        return edge
    dlg.Destroy()
    return None


# ---------------------------------------------------------------------------
# KherveFitting loader  (mirrors _load_into_kherve in AVG_Import.py)
# ---------------------------------------------------------------------------

def _load_into_kherve(window, excel_path, sheet_names, sample_name, update_console):
    """Load a completed Excel file into the KherveFitting data model."""
    update_console("\nLoading into KherveFitting...")

    window.Data = Init_Measurement_Data(window)
    window.Data['FilePath'] = excel_path

    window.history = []
    window.redo_stack = []
    update_undo_redo_state(window)

    window.results_grid.ClearGrid()
    if window.results_grid.GetNumberRows() > 0:
        window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

    for sheet_name in sheet_names:
        add_core_level_Data(window.Data, window, excel_path, sheet_name)
        update_console(f"  Imported: {sheet_name}")

    window.Data['SampleNames'] = {0: sample_name}

    update_console("Creating JSON file...")
    json_path = excel_path.replace('.xlsx', '.json')
    serializable_data = convert_to_serializable_and_round(window.Data)
    with open(json_path, 'w') as fh:
        json.dump(serializable_data, fh, indent=4)

    window.sheet_combobox.Clear()
    for sheet_name in sheet_names:
        window.sheet_combobox.Append(sheet_name)
    window.sheet_combobox.SetSelection(0)
    window.current_sheet = sheet_names[0]

    on_sheet_selected(window, None)
    save_state(window)

    update_console("Refreshing sheets...")
    from libraries.FileMenu.Save import refresh_sheets
    refresh_sheets(window, on_sheet_selected, update_console)

    # Refresh FileManager if it is open
    file_manager_was_open = False
    file_manager_position = None
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            window.file_manager.GetSize()
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
        except RuntimeError:
            window.file_manager = None

    update_console(f"\nImport complete!  ({len(sheet_names)} sheet(s))")

    if file_manager_was_open:
        from libraries.ViewMenu.FileManager import FileManagerWindow
        window.file_manager = FileManagerWindow(window)
        if file_manager_position:
            window.file_manager.SetPosition(file_manager_position)
        window.file_manager.Show()


# ---------------------------------------------------------------------------
# Public API  --  single file
# ---------------------------------------------------------------------------

def import_xas_file(window, file_path=None):
    """
    Import a single Diamond-B07 XAS file (.txt or .dat).

    Shows a file-chooser dialog when file_path is None.
    Supports both the 4-column .txt and the 3-column .dat variants.
    """
    if file_path is None:
        with wx.FileDialog(
                window, "Open Diamond-B07 XAS file",
                wildcard="XAS data files (*.txt;*.dat)|*.txt;*.dat|All files (*.*)|*.*",
                style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

    import_xas_file_direct(window, file_path, ask_edge=True)


def import_xas_file_direct(window, file_path, ask_edge=False, edge_name=None):
    """
    Core single-file import — called by the menu handler and drag-and-drop.

    Args:
        window:    Main KherveFitting window.
        file_path: Absolute path to the XAS file.
        ask_edge:  Prompt the user for the edge when True and edge_name is None.
        edge_name: Pre-set edge label (skips the dialog).
    """
    parent_pos  = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing XAS File", size=(420, 300))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width  - 420) // 2,
        parent_pos.y + (parent_size.height - 300) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        base_filename = os.path.splitext(os.path.basename(file_path))[0]

        # Ask for edge if needed
        if ask_edge and edge_name is None:
            edge_name = _ask_edge(window)
            if edge_name is None:
                console_frame.Close()
                return

        sheet_name = _make_sheet_name(edge_name)
        update_console(f"Opening: {os.path.basename(file_path)}")
        update_console(f"  Edge: {edge_name}  ->  sheet: {sheet_name}")

        # Parse
        try:
            data = _parse_xas_file(file_path)
        except ValueError as exc:
            update_console(f"\nError: {exc}")
            wx.CallLater(3000, console_frame.Close)
            return

        update_console(f"  Points: {len(data)}")

        # Build Excel
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(sheet_name)
        _write_xas_sheet(ws, data)

        update_console(f"Saving: {os.path.basename(excel_path)}")
        wb.save(excel_path)

        sample_name = _strip_nexafs(base_filename)
        _load_into_kherve(window, excel_path, [sheet_name], sample_name, update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing XAS file:\n\n{exc}", "Error", wx.OK | wx.ICON_ERROR)


# ---------------------------------------------------------------------------
# Public API  --  multiple files
# ---------------------------------------------------------------------------

def import_multiple_xas_files(window, file_paths=None):
    """
    Import multiple Diamond-B07 XAS files into one Excel workbook.

    Shows a multi-select file dialog when file_paths is None.
    Both Format A (.txt) and Format B (.dat) files may be mixed in the same batch.
    """
    if file_paths is None:
        with wx.FileDialog(
                window, "Select Diamond-B07 XAS files",
                wildcard="XAS data files (*.txt;*.dat)|*.txt;*.dat|All files (*.*)|*.*",
                style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_paths = dlg.GetPaths()

    if not file_paths:
        return

    # Ask for edge once for the whole batch
    edge_name = _ask_edge(window)
    if edge_name is None:
        return

    # Ask for output filename
    output_dir  = os.path.dirname(file_paths[0])
    first_base  = os.path.splitext(os.path.basename(file_paths[0]))[0]
    default_name = f"{first_base}_combined" if len(file_paths) > 1 else first_base

    with wx.TextEntryDialog(
            window, "Enter name for the output file:",
            "Output File Name", default_name) as dlg:
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        output_name = dlg.GetValue().strip() or default_name

    excel_path = os.path.join(output_dir, f"{output_name}.xlsx")

    # Console window
    parent_pos  = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing XAS Files", size=(420, 360),
                             style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width  - 420) // 2,
        parent_pos.y + (parent_size.height - 360) // 2,
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(msg):
        console_text.AppendText(msg + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Importing {len(file_paths)} file(s)...")
        update_console(f"Edge: {edge_name}   Output: {output_name}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names      = []
        first_sample_name = None

        for idx, fp in enumerate(file_paths):
            fname = os.path.basename(fp)
            base  = os.path.splitext(fname)[0]
            sheet_name = _make_sheet_name(edge_name, idx)

            update_console(f"\n({idx + 1}/{len(file_paths)}) {fname}")
            update_console(f"  -> {sheet_name}")

            try:
                data = _parse_xas_file(fp)
            except ValueError as exc:
                update_console(f"  Skipped: {exc}")
                continue

            ws = wb.create_sheet(sheet_name)
            _write_xas_sheet(ws, data)

            sheet_names.append(sheet_name)
            if first_sample_name is None:
                first_sample_name = _strip_nexafs(base)

            update_console(f"  {len(data)} points")

        if not sheet_names:
            update_console("\nNo valid data found in any of the files.")
            wx.CallLater(2000, console_frame.Close)
            wx.MessageBox("No XAS files could be processed.", "Error", wx.OK | wx.ICON_ERROR)
            return

        update_console(f"\nSaving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        _load_into_kherve(window, excel_path, sheet_names,
                          first_sample_name or output_name, update_console)

        wx.CallLater(1500, console_frame.Close)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {exc}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing XAS files:\n\n{exc}", "Error", wx.OK | wx.ICON_ERROR)