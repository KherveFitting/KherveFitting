# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
Igor File Import Module for KherveFitting
Supports .dat and .itx file formats from Igor Pro
"""

import os
import re
import numpy as np
import openpyxl
import wx
import json

from libraries.ConfigFile import Init_Measurement_Data
from libraries.FileMenu.Save import update_undo_redo_state, save_state
from libraries.Sheet_Operations import on_sheet_selected


def import_igor_dat_file(window, file_path=None):
    """
    Import Igor .dat file containing XPS data.

    The .dat file is tab-separated with:
    - Header row containing [All Core_Level] SampleName patterns
    - Data rows with BE in first column, intensity values in subsequent columns
    - Fitted data columns (bck, fitresult, peak) are ignored

    Args:
        window: Main KherveFitting window instance
        file_path: Path to .dat file (if None, shows file dialog)
    """
    if file_path is None:
        with wx.FileDialog(window, "Open Igor .dat file",
                          wildcard="Igor DAT files (*.dat)|*.dat",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            file_path = fileDialog.GetPath()

    try:
        # Read file
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        if not lines:
            wx.MessageBox("File is empty", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Parse header to extract sample information
        # Structure varies:
        # - First sample: Column 0 has both header AND is used for BE
        # - Other samples: Empty column before data is the BE column
        header = lines[0].strip().split('\t')
        sample_pattern = r'\[All (.*?)\] (\w+)'
        samples_info = []
        seen_samples = set()

        # Parse data first to determine actual BE columns
        all_data = {}
        for line in lines[1:]:
            if not line.strip():
                continue
            values = line.strip().split('\t')
            try:
                for col_idx in range(len(values)):
                    if col_idx not in all_data:
                        all_data[col_idx] = []
                    val = float(values[col_idx]) if col_idx < len(values) and values[col_idx] else 0.0
                    all_data[col_idx].append(val)
            except (ValueError, IndexError):
                continue

        # Now identify samples and their BE/data columns
        for i, h in enumerate(header):
            if not h.strip():  # Empty header
                continue

            match = re.search(sample_pattern, h)
            if match:
                core_level_raw = match.group(1)
                sample_name = match.group(2)
                core_level = core_level_raw.replace(' ', '')

                # Only process raw data columns (first occurrence of sample, not bck/fitresult/peak)
                if sample_name not in seen_samples:
                    if 'bck' not in h.lower() and 'fitresult' not in h.lower() and 'peak' not in h.lower():
                        # Check if this column contains BE values (800-1000 range) or intensity
                        col_values = all_data.get(i, [])

                        if col_values and min(col_values) > 500:
                            # This column has BE values, intensity is in next empty column
                            be_col = i
                            data_col = i + 1  # Next column should be empty header with data

                            # Verify next column exists and has reasonable data
                            if data_col >= len(header) or data_col not in all_data:
                                continue
                        else:
                            # This column has intensity data
                            data_col = i

                            # Find BE column - look backwards for empty header or use col 0
                            be_col = None
                            for j in range(i-1, -1, -1):
                                if not header[j].strip():
                                    be_col = j
                                    break
                            if be_col is None:
                                be_col = 0  # Default to first column

                        samples_info.append({
                            'sample_name': sample_name,
                            'core_level': core_level,
                            'data_col': data_col,
                            'be_col': be_col
                        })
                        seen_samples.add(sample_name)

        if not samples_info:
            wx.MessageBox("No valid sample data found in file", "Error", wx.OK | wx.ICON_ERROR)
            return

        if not all_data:
            wx.MessageBox("No valid data found in file", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Check if normalization is needed (max < 100 means normalized to 1)
        def check_normalization_needed(data_values):
            """Check if data needs to be scaled up to avoid .2f precision loss"""
            max_val = max(data_values) if data_values else 0
            return max_val < 100

        # Create Excel file
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}_imported.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create a sheet for each sample with incrementing number
        for idx, sample_info in enumerate(samples_info):
            sheet_name = f"{sample_info['core_level']}{idx}"
            ws = wb.create_sheet(sheet_name)

            # Write headers
            ws['A1'] = 'BE'
            ws['B1'] = 'Raw Data'

            # Get BE and data
            be_col = sample_info['be_col']
            data_col = sample_info['data_col']

            if be_col not in all_data or data_col not in all_data:
                continue

            be_values = all_data[be_col]
            data_values = all_data[data_col]

            # Check if normalization is needed
            needs_normalization = check_normalization_needed(data_values)
            scale_factor = 10000 if needs_normalization else 1

            # Reverse data order - XPS convention is high BE to low BE
            be_values_reversed = list(reversed(be_values))
            data_values_reversed = list(reversed(data_values))

            # Write data with .2f formatting
            for i, (be, intensity) in enumerate(zip(be_values_reversed, data_values_reversed), start=2):
                ws[f'A{i}'] = f"{be:.2f}"
                ws[f'B{i}'] = f"{(intensity * scale_factor):.2f}"

        wb.save(excel_path)

        # Now open the file and update window.Data
        _open_imported_excel_file(window, excel_path, samples_info)

        window.show_popup_message2("Success",
                                   f"Imported {len(samples_info)} samples from Igor .dat file")

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error importing Igor .dat file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_igor_itx_file(window, file_path=None):
    """
    Import Igor .itx file containing XPS data.

    The .itx file format:
    - Starts with 'IGOR' header
    - Contains WAVES declarations: WAVES 'wavename'
    - Data blocks between BEGIN and END
    - X SetScale commands define BE range: X SetScale/P x start,step

    Args:
        window: Main KherveFitting window instance
        file_path: Path to .itx file (if None, shows file dialog)
    """
    if file_path is None:
        with wx.FileDialog(window, "Open Igor .itx file",
                          wildcard="Igor ITX files (*.itx)|*.itx",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            file_path = fileDialog.GetPath()

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        # Validate Igor file
        if not lines or lines[0].strip() != 'IGOR':
            wx.MessageBox("Not a valid Igor ITX file", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Parse waves
        waves_data = {}
        i = 0

        while i < len(lines):
            line = lines[i].strip()

            if line.startswith('WAVES'):
                # Extract wave name - handle WAVES, WAVES/D, WAVES/T, etc.
                wave_name_match = re.search(r"WAVES[/\w]*\s+'([^']+)'", line)
                if not wave_name_match:
                    i += 1
                    continue

                wave_name = wave_name_match.group(1)

                # Skip fitted data (bck, fitresult, peak)
                if any(term in wave_name.lower() for term in ['bck', 'fitresult', 'peak']):
                    i += 1
                    continue

                # Find BEGIN
                i += 1
                while i < len(lines) and lines[i].strip() != 'BEGIN':
                    i += 1

                if i >= len(lines):
                    break

                # Read data until END
                i += 1
                data_values = []
                while i < len(lines) and lines[i].strip() != 'END':
                    try:
                        value = float(lines[i].strip())
                        data_values.append(value)
                    except ValueError:
                        pass
                    i += 1

                if data_values:
                    waves_data[wave_name] = np.array(data_values)

                    # Look for X SetScale to get BE range
                    i += 1
                    if i < len(lines):
                        setscale_line = lines[i].strip()
                        if setscale_line.startswith('X SetScale'):
                            # Extract: X SetScale/P x 874.7,0.1
                            match = re.search(r'x\s+([\d.]+),([\d.]+)', setscale_line)
                            if match:
                                start_be = float(match.group(1))
                                step = float(match.group(2))

                                # Create BE array
                                be_array = np.arange(start_be,
                                                     start_be + step * len(data_values),
                                                     step)[:len(data_values)]
                                waves_data[f'{wave_name}_BE'] = be_array

            i += 1

        if not waves_data:
            wx.MessageBox("No valid wave data found in file", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Extract sample information
        sample_pattern = r'\[All (.*?)\] (\w+)'
        samples_info = []

        for wave_name in sorted(waves_data.keys()):
            if '_BE' in wave_name:
                continue

            match = re.search(sample_pattern, wave_name)
            if match:
                # Standard pattern: [All CoreLevel] SampleName
                core_level_raw = match.group(1)
                sample_name = match.group(2)
                core_level = core_level_raw.replace(' ', '')

                samples_info.append({
                    'sample_name': sample_name,
                    'core_level': core_level,
                    'wave_name': wave_name
                })
            else:
                # Alternative pattern: extract from wave name directly
                # Try to extract core level from wave name (e.g., O1s, C1s, etc.)
                core_level_match = re.search(r'([A-Z][a-z]?\d[spdf])', wave_name)
                if core_level_match:
                    core_level = core_level_match.group(1)
                # Check for valence band patterns
                elif '_V' in wave_name or 'VB' in wave_name.upper() or 'valence' in wave_name.lower():
                    # Extract valence band range if present
                    vb_range_match = re.search(r'_V\(([\d.]+)\s+to\s+([\d.]+)\)', wave_name)
                    if vb_range_match:
                        core_level = f"VB_{vb_range_match.group(1)}to{vb_range_match.group(2)}"
                    else:
                        core_level = 'VB'
                # Check for other energy ranges with "eV"
                elif 'eV' in wave_name:
                    # Try to extract energy information
                    ev_match = re.search(r'([\d.]+)\s*eV', wave_name)
                    if ev_match:
                        core_level = f"{ev_match.group(1)}eV"
                    else:
                        core_level = 'Spectrum'
                else:
                    # Use a generic name if no core level found
                    core_level = 'Unknown'

                # Use wave name as sample name, cleaned up
                sample_name = wave_name.strip("'").replace('.txt', '').replace('.itx', '')

                samples_info.append({
                    'sample_name': sample_name,
                    'core_level': core_level,
                    'wave_name': wave_name
                })

        if not samples_info:
            wx.MessageBox("No valid sample data found in file", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Create Excel file
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}_imported.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Check if normalization is needed
        def check_normalization_needed(data_values):
            """Check if data needs to be scaled up to avoid .2f precision loss"""
            max_val = np.max(data_values) if len(data_values) > 0 else 0
            return max_val < 100

        # Create a sheet for each sample
        for idx, sample_info in enumerate(samples_info):
            wave_name = sample_info['wave_name']
            sheet_name = f"{sample_info['core_level']}{idx}"

            ws = wb.create_sheet(sheet_name)

            # Write headers
            ws['A1'] = 'BE'
            ws['B1'] = 'Raw Data'

            # Get BE and data
            be_key = f'{wave_name}_BE'
            if be_key in waves_data:
                be_values = waves_data[be_key]
            else:
                # If no BE found, create index
                be_values = np.arange(len(waves_data[wave_name]))

            data_values = waves_data[wave_name]

            # Check if normalization is needed
            needs_normalization = check_normalization_needed(data_values)
            scale_factor = 10000 if needs_normalization else 1

            # Reverse data order - XPS convention is high BE to low BE
            be_values_reversed = be_values[::-1]
            data_values_reversed = data_values[::-1]

            # Write data with .2f formatting and normalization
            for i, (be, intensity) in enumerate(zip(be_values_reversed, data_values_reversed), start=2):
                ws[f'A{i}'] = f"{be:.2f}"
                ws[f'B{i}'] = f"{(intensity * scale_factor):.2f}"

        wb.save(excel_path)

        # Open the file and update window.Data
        _open_imported_excel_file(window, excel_path, samples_info)

        window.show_popup_message2("Success",
                                   f"Imported {len(samples_info)} samples from Igor .itx file")

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error importing Igor .itx file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_multiple_igor_files(window):
    """
    Import multiple Igor files (.dat and/or .itx) at once.

    Args:
        window: Main KherveFitting window instance
    """
    with wx.FileDialog(window, "Open Igor files",
                      wildcard="Igor files (*.dat;*.itx)|*.dat;*.itx",
                      style=wx.FD_OPEN | wx.FD_MULTIPLE | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        file_paths = fileDialog.GetPaths()

    if not file_paths:
        return

    processed_count = 0
    error_count = 0

    for file_path in file_paths:
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.dat':
                import_igor_dat_file(window, file_path)
                processed_count += 1
            elif ext == '.itx':
                import_igor_itx_file(window, file_path)
                processed_count += 1
            else:
                error_count += 1
        except Exception as e:
            print(f"Error processing {os.path.basename(file_path)}: {e}")
            error_count += 1

    if processed_count > 0:
        msg = f"Processed {processed_count} of {len(file_paths)} Igor files"
        if error_count > 0:
            msg += f"\n{error_count} files had errors"
        window.show_popup_message2("Import Complete", msg)


def _open_imported_excel_file(window, excel_path, samples_info):
    """
    Open the imported Excel file and properly populate window.Data structure.

    Args:
        window: Main KherveFitting window instance
        excel_path: Path to the Excel file to open
        samples_info: List of dicts containing sample metadata
    """
    import pandas as pd
    from libraries.ConfigFile import add_core_level_Data

    # Close any open windows
    if hasattr(window, 'file_manager') and window.file_manager is not None:
        try:
            window.file_manager.Close()
        except:
            pass
        window.file_manager = None

    # Initialize Data structure
    window.Data = Init_Measurement_Data(window)
    window.Data['FilePath'] = excel_path

    # Clear history
    window.history = []
    window.redo_stack = []
    update_undo_redo_state(window)

    # Clear results grid
    window.results_grid.ClearGrid()
    if window.results_grid.GetNumberRows() > 0:
        window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

    # Read Excel file and add core level data
    excel_file = pd.ExcelFile(excel_path)
    sheet_names = excel_file.sheet_names

    for sheet_name in sheet_names:
        window.Data = add_core_level_Data(window.Data, window, excel_path, sheet_name)

    # Add sample names to window.Data
    if 'SampleNames' not in window.Data:
        window.Data['SampleNames'] = {}

    for idx, sample_info in enumerate(samples_info):
        window.Data['SampleNames'][idx] = sample_info['sample_name']

    # Create and save JSON file
    json_path = os.path.splitext(excel_path)[0] + '.json'
    from libraries.FileMenu.Save import convert_to_serializable_and_round
    serializable_data = convert_to_serializable_and_round(window.Data)

    with open(json_path, 'w') as json_file:
        json.dump(serializable_data, json_file, indent=2)

    # Update UI
    window.sheet_combobox.Clear()
    window.sheet_combobox.AppendItems(sheet_names)

    if sheet_names:
        first_sheet = sheet_names[0]
        window.sheet_combobox.SetValue(first_sheet)

        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(first_sheet)
        window.plot_config.plot_limits.clear()
        on_sheet_selected(window, event)

    # Save state
    save_state(window)

    # Update recent files
    from libraries.FileMenu.Open import update_recent_files
    update_recent_files(window, excel_path)

    # Refresh sheets
    from libraries.FileMenu.Save import refresh_sheets
    refresh_sheets(window, on_sheet_selected)

    window.SetStatusText(f"Selected File: {excel_path}", 0)