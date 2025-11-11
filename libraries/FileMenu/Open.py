import os
import json
import openpyxl
import xlrd
import wx
import re
import sys
import shutil
import math
from vamas import Vamas
from openpyxl import Workbook
import numpy as np
from scipy.interpolate import interp1d
import pandas as pd
from openpyxl.styles import Alignment

from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
from libraries.FileMenu.Save import update_undo_redo_state, save_state
from libraries.Sheet_Operations import on_sheet_selected
from libraries.Grid_Operations import populate_results_grid


class ExcelDropTarget(wx.FileDropTarget):
    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, filenames):

        # Check all files are valid first
        for file in filenames:
            if not any(file.lower().endswith(ext) for ext in ['.xlsx', '.xls', '.vms', '.kal',
                                                              '.avg', '.spe', '.mrs', '.1', '.asc']):
                wx.MessageBox(f"Only .xlsx/.xls (Khervefitting or Avantage), .vms (Vamas), "
                              f".kal (Kratos), .avg (Thermo), .mrs, .1 (VG-Microtech), .asc and .spe "
                              f"(Phi) files can be dropped.", "Invalid File Type",
                              wx.OK | wx.ICON_ERROR)
                return False

        # If single file, use original logic
        if len(filenames) == 1:
            file = filenames[0]
            return self._process_single_dropped_file(file)

        # Multiple files - group by type and process
        return self._process_multiple_dropped_files(filenames)

    def _process_single_dropped_file_OLD(self, file):
        """Process single dropped file using original logic"""
        from libraries.FileMenu.Open import open_xlsx_file, open_vamas_file, open_kal_file, open_avg_file_direct, open_spe_file, \
            open_mrs_file, open_vg_microtech_file
        from libraries.FileMenu.Open import import_avantage_file_direct, import_avantage_file_direct_xls

        if file.lower().endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)
                if "Titles" in wb.sheetnames:
                    wx.CallAfter(import_avantage_file_direct, self.window, file)
                else:
                    wx.CallAfter(open_xlsx_file, self.window, file)
                wb.close()
                return True
            except Exception:
                return False
        elif file.lower().endswith('.xls'):
            try:
                wb = xlrd.open_workbook(file)
                if "Titles" in wb.sheet_names():
                    wx.CallAfter(import_avantage_file_direct_xls, self.window, file)
                else:
                    wx.CallAfter(open_xlsx_file, self.window, file)
                wb.close()
                return True
            except Exception:
                print("Error opening Excel file:", sys.exc_info())
                return False
        elif file.lower().endswith('.vms'):
            wx.CallAfter(open_vamas_file, self.window, file)
            return True
        elif file.lower().endswith('.kal'):
            wx.CallAfter(open_kal_file, self.window, file)
            return True
        elif file.lower().endswith('.avg'):
            wx.CallAfter(open_avg_file_direct, self.window, file)
            return True
        elif file.lower().endswith('.spe'):
            wx.CallAfter(open_spe_file, self.window, file)
            return True
        elif file.lower().endswith('.mrs'):
            wx.CallAfter(open_mrs_file, self.window, file)
            return True
        elif file.lower().endswith('.1'):
            wx.CallAfter(open_vg_microtech_file, self.window, file)
            return True
        elif file.lower().endswith('.asc'):
            from libraries.FileMenu.Open import import_xps_asc_file_direct
            wx.CallAfter(import_xps_asc_file_direct, self.window, file)
            return True
        return False

    def _process_single_dropped_file(self, file):
        """Process single dropped file using original logic"""
        from libraries.FileMenu.Open import (open_xlsx_file, open_vamas_file, open_kal_file,
                                             open_avg_file_direct, open_spe_file, open_mrs_file,
                                             open_vg_microtech_file, import_avantage_file_direct,
                                             import_avantage_file_direct_xls, import_generic_excel_file)

        if file.lower().endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)

                # Check for Avantage format
                if "Titles" in wb.sheetnames:
                    wx.CallAfter(import_avantage_file_direct, self.window, file)
                    wb.close()
                    return True

                # Check if it's standard kFitting format
                # Must have: proper headers OR proper sheet names (non-generic, non-special)
                is_kfitting = False
                has_generic_sheet_names = False
                has_proper_sheet_names = False

                for sheet_name in wb.sheetnames:
                    # Skip special sheets
                    if sheet_name.lower() in ["results table", "experimental description"]:
                        continue

                    # Check if sheet name is generic (Sheet1, Sheet2, etc.)
                    if re.match(r'^Sheet\d+$', sheet_name, re.IGNORECASE):
                        has_generic_sheet_names = True
                        break
                    else:
                        # If it's not generic and not a special sheet, it's a proper name
                        has_proper_sheet_names = True

                    sheet = wb[sheet_name]
                    if sheet.max_row > 0:
                        # Check first row for proper kFitting headers
                        header_a = sheet.cell(row=1, column=1).value
                        header_b = sheet.cell(row=1, column=2).value

                        if header_a and header_b:
                            header_a_lower = str(header_a).lower().strip()
                            header_b_lower = str(header_b).lower().strip()

                            # Valid kFitting headers
                            valid_headers_a = ['binding energy', 'raman shift', 'wavelength', 'depth']
                            valid_headers_b = ['raw data', 'intensity', 'cps', 'counts']

                            if header_a_lower in valid_headers_a and header_b_lower in valid_headers_b:
                                is_kfitting = True
                                break

                wb.close()

                # Treat as kFitting if it has proper headers OR proper sheet names (and no generic names)
                if (is_kfitting or has_proper_sheet_names) and not has_generic_sheet_names:
                    wx.CallAfter(open_xlsx_file, self.window, file)
                else:
                    # Has generic sheet names or no proper headers/names - use interactive import
                    wx.CallAfter(import_generic_excel_file, self.window, file)

                return True

            except Exception as e:
                print(f"Error processing Excel file: {e}")
                return False

        elif file.lower().endswith('.xls'):
            try:
                wb = xlrd.open_workbook(file)

                # Check for Avantage format
                if "Titles" in wb.sheet_names():
                    wx.CallAfter(import_avantage_file_direct_xls, self.window, file)
                    wb.release_resources()
                    return True

                # Check if kFitting format with proper headers AND proper sheet names
                is_kfitting = False
                has_generic_sheet_names = False

                for sheet_idx in range(wb.nsheets):
                    sheet = wb.sheet_by_index(sheet_idx)

                    if sheet.name.lower() in ["results table", "experimental description"]:
                        continue

                    # Check if sheet name is generic
                    if re.match(r'^Sheet\d+$', sheet.name, re.IGNORECASE):
                        has_generic_sheet_names = True
                        break

                    if sheet.nrows > 0 and sheet.ncols >= 2:
                        header_a = sheet.cell(0, 0).value
                        header_b = sheet.cell(0, 1).value

                        if header_a and header_b:
                            header_a_lower = str(header_a).lower().strip()
                            header_b_lower = str(header_b).lower().strip()

                            valid_headers_a = ['binding energy', 'raman shift', 'wavelength', 'depth']
                            valid_headers_b = ['raw data', 'intensity', 'cps', 'counts']

                            if header_a_lower in valid_headers_a and header_b_lower in valid_headers_b:
                                is_kfitting = True
                                break

                wb.release_resources()

                # Only treat as kFitting if it has proper headers AND no generic sheet names
                if is_kfitting and not has_generic_sheet_names:
                    wx.CallAfter(open_xlsx_file, self.window, file)
                else:
                    # Has generic sheet names or no proper headers - use interactive import
                    wx.CallAfter(import_generic_excel_file, self.window, file)

                return True

            except Exception:
                print("Error opening Excel file:", sys.exc_info())
                return False

        elif file.lower().endswith('.vms'):
            wx.CallAfter(open_vamas_file, self.window, file)
            return True
        elif file.lower().endswith('.kal'):
            wx.CallAfter(open_kal_file, self.window, file)
            return True
        elif file.lower().endswith('.avg'):
            wx.CallAfter(open_avg_file_direct, self.window, file)
            return True
        elif file.lower().endswith('.spe'):
            wx.CallAfter(open_spe_file, self.window, file)
            return True
        elif file.lower().endswith('.mrs'):
            wx.CallAfter(open_mrs_file, self.window, file)
            return True
        elif file.lower().endswith('.1'):
            wx.CallAfter(open_vg_microtech_file, self.window, file)
            return True
        elif file.lower().endswith('.asc'):
            from libraries.FileMenu.Open import import_xps_asc_file_direct
            wx.CallAfter(import_xps_asc_file_direct, self.window, file)
            return True
        return False

    def _is_numeric(self, value):
        """Check if value is numeric."""
        if value is None or value == '':
            return False
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False



    def _process_multiple_dropped_files_OLD(self, filenames):
        """Process multiple dropped files by grouping them by type"""
        # Group files by type
        file_groups = {
            'khervefitting_xlsx': [],
            'khervefitting_xls': [],
            'avantage_xlsx': [],
            'avantage_xls': [],
            'vms': [],
            'kal': [],
            'avg': [],
            'spe': [],
            'mrs': [],
            'vg': []
        }

        # Categorize files
        for file in filenames:
            file_lower = file.lower()

            if file_lower.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(file)
                    if "Titles" in wb.sheetnames:
                        file_groups['avantage_xlsx'].append(file)
                    else:
                        file_groups['khervefitting_xlsx'].append(file)
                    wb.close()
                except Exception:
                    file_groups['khervefitting_xlsx'].append(file)

            elif file_lower.endswith('.xls'):
                try:
                    wb = xlrd.open_workbook(file)
                    if "Titles" in wb.sheet_names():
                        file_groups['avantage_xls'].append(file)
                    else:
                        file_groups['khervefitting_xls'].append(file)
                    wb.close()
                except Exception:
                    file_groups['khervefitting_xls'].append(file)

            elif file_lower.endswith('.vms'):
                file_groups['vms'].append(file)
            elif file_lower.endswith('.kal'):
                file_groups['kal'].append(file)
            elif file_lower.endswith('.avg'):
                file_groups['avg'].append(file)
            elif file_lower.endswith('.spe'):
                file_groups['spe'].append(file)
            elif file_lower.endswith('.mrs'):
                file_groups['mrs'].append(file)
            elif file_lower.endswith('.1'):
                file_groups['vg'].append(file)

        # Process each group
        success = True

        # Process KherveFitting files
        all_khervefitting = file_groups['khervefitting_xlsx'] + file_groups['khervefitting_xls']
        if len(all_khervefitting) > 1:
            wx.CallAfter(self._import_multiple_khervefitting_direct, all_khervefitting)
        elif len(all_khervefitting) == 1:
            success &= self._process_single_dropped_file(all_khervefitting[0])

        # Process Avantage files
        all_avantage = file_groups['avantage_xlsx'] + file_groups['avantage_xls']
        if len(all_avantage) > 1:
            wx.CallAfter(self._import_multiple_avantage_direct, all_avantage)
        elif len(all_avantage) == 1:
            success &= self._process_single_dropped_file(all_avantage[0])

        # Process MRS files
        if len(file_groups['mrs']) > 1:
            wx.CallAfter(self._import_multiple_mrs_direct, file_groups['mrs'])
        elif len(file_groups['mrs']) == 1:
            success &= self._process_single_dropped_file(file_groups['mrs'][0])

        # Process AVG files
        if len(file_groups['avg']) > 1:
            wx.CallAfter(self._import_multiple_avg_direct, file_groups['avg'])
        elif len(file_groups['avg']) == 1:
            success &= self._process_single_dropped_file(file_groups['avg'][0])

        # Process VG-Microtech files
        if len(file_groups['vg']) > 1:
            wx.CallAfter(self._import_multiple_vg_direct, file_groups['vg'])
        elif len(file_groups['vg']) == 1:
            success &= self._process_single_dropped_file(file_groups['vg'][0])

        # Process other file types individually (no batch import available)
        for file_type in ['vms', 'kal', 'spe']:
            for file in file_groups[file_type]:
                success &= self._process_single_dropped_file(file)

        return success

    def _process_multiple_dropped_files(self, filenames):
        """Process multiple dropped files by grouping them by type"""
        # Group files by type
        file_groups = {
            'khervefitting_xlsx': [],
            'khervefitting_xls': [],
            'avantage_xlsx': [],
            'avantage_xls': [],
            'generic_xlsx': [],
            'generic_xls': [],
            'vms': [],
            'kal': [],
            'avg': [],
            'spe': [],
            'mrs': [],
            'vg': [],
            'asc': []
        }

        # Categorize files
        for file in filenames:
            file_lower = file.lower()

            if file_lower.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(file)

                    # Check for Avantage
                    if "Titles" in wb.sheetnames:
                        file_groups['avantage_xlsx'].append(file)
                    else:
                        # Check if kFitting format with proper headers AND sheet names
                        is_kfitting = False
                        has_generic_sheet_names = False

                        for sheet_name in wb.sheetnames:
                            if sheet_name.lower() in ["results table", "experimental description"]:
                                continue

                            # Check for generic sheet names
                            if re.match(r'^Sheet\d+$', sheet_name, re.IGNORECASE):
                                has_generic_sheet_names = True
                                break

                            sheet = wb[sheet_name]
                            if sheet.max_row > 0:
                                header_a = sheet.cell(row=1, column=1).value
                                header_b = sheet.cell(row=1, column=2).value

                                if header_a and header_b:
                                    header_a_lower = str(header_a).lower().strip()
                                    header_b_lower = str(header_b).lower().strip()

                                    valid_headers_a = ['binding energy', 'raman shift', 'wavelength', 'depth']
                                    valid_headers_b = ['raw data', 'intensity', 'cps', 'counts']

                                    if header_a_lower in valid_headers_a and header_b_lower in valid_headers_b:
                                        is_kfitting = True
                                        break

                        if is_kfitting and not has_generic_sheet_names:
                            file_groups['khervefitting_xlsx'].append(file)
                        else:
                            file_groups['generic_xlsx'].append(file)

                    wb.close()
                except Exception:
                    file_groups['khervefitting_xlsx'].append(file)

            elif file_lower.endswith('.xls'):
                try:
                    wb = xlrd.open_workbook(file)

                    # Check for Avantage
                    if "Titles" in wb.sheet_names():
                        file_groups['avantage_xls'].append(file)
                    else:
                        # Check if kFitting format
                        is_kfitting = False
                        has_generic_sheet_names = False

                        for sheet_idx in range(wb.nsheets):
                            sheet = wb.sheet_by_index(sheet_idx)

                            if sheet.name.lower() in ["results table", "experimental description"]:
                                continue

                            if re.match(r'^Sheet\d+$', sheet.name, re.IGNORECASE):
                                has_generic_sheet_names = True
                                break

                            if sheet.nrows > 0 and sheet.ncols >= 2:
                                header_a = sheet.cell(0, 0).value
                                header_b = sheet.cell(0, 1).value

                                if header_a and header_b:
                                    header_a_lower = str(header_a).lower().strip()
                                    header_b_lower = str(header_b).lower().strip()

                                    valid_headers_a = ['binding energy', 'raman shift', 'wavelength', 'depth']
                                    valid_headers_b = ['raw data', 'intensity', 'cps', 'counts']

                                    if header_a_lower in valid_headers_a and header_b_lower in valid_headers_b:
                                        is_kfitting = True
                                        break

                        if is_kfitting and not has_generic_sheet_names:
                            file_groups['khervefitting_xls'].append(file)
                        else:
                            file_groups['generic_xls'].append(file)

                    wb.release_resources()
                except Exception:
                    file_groups['khervefitting_xls'].append(file)

            elif file_lower.endswith('.vms'):
                file_groups['vms'].append(file)
            elif file_lower.endswith('.kal'):
                file_groups['kal'].append(file)
            elif file_lower.endswith('.avg'):
                file_groups['avg'].append(file)
            elif file_lower.endswith('.spe'):
                file_groups['spe'].append(file)
            elif file_lower.endswith('.mrs'):
                file_groups['mrs'].append(file)
            elif file_lower.endswith('.1'):
                file_groups['vg'].append(file)
            elif file_lower.endswith('.asc'):
                file_groups['asc'].append(file)

        # Process each group
        success = True

        # Process KherveFitting files
        all_khervefitting = file_groups['khervefitting_xlsx'] + file_groups['khervefitting_xls']
        if len(all_khervefitting) > 1:
            wx.CallAfter(self._import_multiple_khervefitting_direct, all_khervefitting)
        elif len(all_khervefitting) == 1:
            success &= self._process_single_dropped_file(all_khervefitting[0])

        # Process Avantage files
        all_avantage = file_groups['avantage_xlsx'] + file_groups['avantage_xls']
        if len(all_avantage) > 1:
            wx.CallAfter(self._import_multiple_avantage_direct, all_avantage)
        elif len(all_avantage) == 1:
            success &= self._process_single_dropped_file(all_avantage[0])

        # Process generic Excel files individually (they need user interaction)
        all_generic = file_groups['generic_xlsx'] + file_groups['generic_xls']
        for file in all_generic:
            success &= self._process_single_dropped_file(file)

        # Process MRS files
        if len(file_groups['mrs']) > 1:
            wx.CallAfter(self._import_multiple_mrs_direct, file_groups['mrs'])
        elif len(file_groups['mrs']) == 1:
            success &= self._process_single_dropped_file(file_groups['mrs'][0])

        # Process AVG files
        if len(file_groups['avg']) > 1:
            wx.CallAfter(self._import_multiple_avg_direct, file_groups['avg'])
        elif len(file_groups['avg']) == 1:
            success &= self._process_single_dropped_file(file_groups['avg'][0])

        # Process VG-Microtech files
        if len(file_groups['vg']) > 1:
            wx.CallAfter(self._import_multiple_vg_direct, file_groups['vg'])
        elif len(file_groups['vg']) == 1:
            success &= self._process_single_dropped_file(file_groups['vg'][0])

        # Process ASC files
        if len(file_groups['asc']) > 1:
            wx.CallAfter(self._import_multiple_asc_direct, file_groups['asc'])
        elif len(file_groups['asc']) == 1:
            success &= self._process_single_dropped_file(file_groups['asc'][0])

        # Process other file types individually (no batch import available)
        for file_type in ['vms', 'kal', 'spe']:
            for file in file_groups[file_type]:
                success &= self._process_single_dropped_file(file)

        return success

    def _import_multiple_khervefitting_direct(self, file_list):
        """Import multiple KherveFitting files directly including their JSON peak fitting data"""
        try:
            import openpyxl
            import os
            import json
            from libraries.FileMenu.Open import process_kfitting_file_with_sample_number_with_mapping

            # Sort files alphabetically
            file_list.sort(key=lambda x: os.path.basename(x))

            # Prompt user to choose save location and filename
            first_file_dir = os.path.dirname(file_list[0])
            default_filename = "Combined_KherveFitting_Files.xlsx"

            with wx.FileDialog(self.window, "Save combined KherveFitting file as...",
                               defaultDir=first_file_dir,
                               defaultFile=default_filename,
                               wildcard="Excel files (*.xlsx)|*.xlsx",
                               style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

                if fileDialog.ShowModal() == wx.ID_CANCEL:
                    return  # User cancelled

                combined_file_path = fileDialog.GetPath()

                # Ensure .xlsx extension
                if not combined_file_path.lower().endswith('.xlsx'):
                    combined_file_path += '.xlsx'

            # Create combined workbook
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            combined_json_data = {
                'Core levels': {},
                'SampleNames': {}
            }

            # Process each file with proper row numbering and sample names
            current_row = 0
            combined_sample_names = {}
            all_sheet_mappings = {}  # Track original -> new sheet name mappings

            for sample_idx, file_path in enumerate(file_list):
                filename = os.path.splitext(os.path.basename(file_path))[0]

                # Load original JSON file to get existing sample names
                original_sample_names = {}
                json_file_path = os.path.splitext(file_path)[0] + '.json'
                if os.path.exists(json_file_path):
                    try:
                        with open(json_file_path, 'r') as json_file:
                            individual_json_data = json.load(json_file)
                            if 'SampleNames' in individual_json_data:
                                original_sample_names = individual_json_data['SampleNames']
                    except Exception as e:
                        print(f"Warning: Could not load sample names from {json_file_path}: {e}")

                # Process Excel file
                wb = openpyxl.load_workbook(file_path)

                # Get original sheet names before processing
                original_sheets = []
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower() not in ["results table", "experimental description"]:
                        original_sheets.append(sheet_name)

                # Modified function call to get sheet mappings
                current_row, sample_names_for_rows, sheet_mappings = process_kfitting_file_with_sample_number_with_mapping(
                    wb, combined_wb, current_row, original_sample_names, filename)

                # Store sheet mappings for this file
                all_sheet_mappings.update(sheet_mappings)

                # Update combined sample names
                combined_sample_names.update(sample_names_for_rows)

                wb.close()

                # Load corresponding JSON file with peak fitting data
                if os.path.exists(json_file_path):
                    try:
                        with open(json_file_path, 'r') as json_file:
                            individual_json_data = json.load(json_file)

                        # Process each core level in the JSON data using proper mappings
                        if 'Core levels' in individual_json_data:
                            for original_sheet_name in original_sheets:
                                if original_sheet_name in individual_json_data['Core levels']:
                                    # Use the tracked mapping to get the correct new sheet name
                                    if original_sheet_name in sheet_mappings:
                                        new_sheet_name = sheet_mappings[original_sheet_name]

                                        # Copy the fitting data to the combined JSON structure
                                        combined_json_data['Core levels'][new_sheet_name] = \
                                            individual_json_data['Core levels'][original_sheet_name].copy()

                                        # Update peak names to include file index for uniqueness
                                        if ('Fitting' in combined_json_data['Core levels'][new_sheet_name] and
                                                'Peaks' in combined_json_data['Core levels'][new_sheet_name][
                                                    'Fitting']):

                                            peaks_data = combined_json_data['Core levels'][new_sheet_name]['Fitting'][
                                                'Peaks']
                                            updated_peaks = {}

                                            for peak_name, peak_data in peaks_data.items():
                                                # Add file index to peak names for uniqueness
                                                new_peak_name = f"{peak_name}_f{sample_idx}"
                                                updated_peaks[new_peak_name] = peak_data

                                            combined_json_data['Core levels'][new_sheet_name]['Fitting'][
                                                'Peaks'] = updated_peaks

                    except Exception as e:
                        print(f"Warning: Could not load JSON file for {filename}: {e}")
                else:
                    print(f"Warning: No JSON file found for {filename}")

            # Use the combined sample names
            sample_names_dict = combined_sample_names

            # Save combined Excel file
            combined_wb.save(combined_file_path)
            combined_wb.close()

            # Save combined JSON file with all peak fitting data
            combined_json_path = os.path.splitext(combined_file_path)[0] + '.json'
            combined_json_data['SampleNames'] = sample_names_dict

            # Add other necessary data structure
            combined_json_data['Number of Core levels'] = len(combined_json_data['Core levels'])
            combined_json_data['FilePath'] = combined_file_path

            # Convert to serializable format and save
            from libraries.FileMenu.Save import convert_to_serializable_and_round
            serializable_data = convert_to_serializable_and_round(combined_json_data)

            with open(combined_json_path, 'w') as json_file:
                json.dump(serializable_data, json_file, indent=2)

            # Open the combined file
            from libraries.FileMenu.Open import open_xlsx_file
            open_xlsx_file(self.window, combined_file_path)

            # Update window.Data with the combined JSON data
            if 'SampleNames' not in self.window.Data:
                self.window.Data['SampleNames'] = {}
            self.window.Data['SampleNames'].update(sample_names_dict)

            # Merge the peak fitting data into window.Data
            if 'Core levels' not in self.window.Data:
                self.window.Data['Core levels'] = {}
            self.window.Data['Core levels'].update(combined_json_data['Core levels'])

            # Get just the filename for the success message
            filename = os.path.basename(combined_file_path)
            self.window.show_popup_message2("Success",
                                            f"Combined {len(file_list)} KherveFitting files with peak fitting data.\nSaved as: {filename}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            wx.MessageBox(f"Error importing multiple KherveFitting files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def _import_multiple_avantage_direct(self, file_list):
        """Import multiple Avantage files directly"""
        try:
            from libraries.FileMenu.Open import import_avantage_file_direct, import_avantage_file_direct_xls

            processed_count = 0
            for file_path in file_list:
                try:
                    if file_path.lower().endswith('.xlsx'):
                        import_avantage_file_direct(self.window, file_path)
                    elif file_path.lower().endswith('.xls'):
                        import_avantage_file_direct_xls(self.window, file_path)
                    processed_count += 1
                except Exception as e:
                    print(f"Error processing {os.path.basename(file_path)}: {e}")

            self.window.show_popup_message2("Success",
                                            f"Processed {processed_count} of {len(file_list)} Avantage files.")

        except Exception as e:
            wx.MessageBox(f"Error importing multiple Avantage files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def _import_multiple_mrs_direct(self, file_list):
        """Import multiple MRS files directly"""
        try:
            from libraries.FileMenu.Open import open_mrs_file

            processed_count = 0
            for file_path in file_list:
                try:
                    if open_mrs_file(self.window, file_path):
                        processed_count += 1
                except Exception as e:
                    print(f"Error processing {os.path.basename(file_path)}: {e}")

            self.window.show_popup_message2("Success", f"Processed {processed_count} of {len(file_list)} MRS files.")

        except Exception as e:
            wx.MessageBox(f"Error importing multiple MRS files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def _import_multiple_avg_direct(self, file_list):
        """Import multiple AVG files directly"""
        try:
            from libraries.FileMenu.Open import open_avg_file_direct

            processed_count = 0
            for file_path in file_list:
                try:
                    open_avg_file_direct(self.window, file_path)
                    processed_count += 1
                except Exception as e:
                    print(f"Error processing {os.path.basename(file_path)}: {e}")

            self.window.show_popup_message2("Success", f"Processed {processed_count} of {len(file_list)} AVG files.")

        except Exception as e:
            wx.MessageBox(f"Error importing multiple AVG files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def _import_multiple_vg_direct(self, file_list):
        """Import multiple VG-Microtech files directly"""
        try:
            from libraries.FileMenu.Open import open_vg_microtech_file

            processed_count = 0
            for file_path in file_list:
                try:
                    open_vg_microtech_file(self.window, file_path)
                    processed_count += 1
                except Exception as e:
                    print(f"Error processing {os.path.basename(file_path)}: {e}")

            self.window.show_popup_message2("Success",
                                            f"Processed {processed_count} of {len(file_list)} VG-Microtech files.")

        except Exception as e:
            wx.MessageBox(f"Error importing multiple VG-Microtech files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def _import_multiple_asc_direct(self, file_list):
        """Import multiple ASC files directly"""
        try:
            import os
            import openpyxl
            from libraries.FileMenu.Open import open_xlsx_file

            # Create a single Excel workbook
            base_dir = os.path.dirname(file_list[0])
            excel_path = os.path.join(base_dir, "XPS_Data.xlsx")

            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            processed_count = 0
            for file_path in file_list:
                try:
                    base_filename = os.path.splitext(os.path.basename(file_path))[0]
                    sheet_name = base_filename

                    # Ensure sheet name is valid (max 31 chars)
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]

                    # Handle duplicate sheet names
                    count = 1
                    original_name = sheet_name
                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{original_name[:27]}_{count}"
                        count += 1

                    # Read data from asc file
                    data = []
                    with open(file_path, 'r') as f:
                        for line in f:
                            if line.strip() and not line.startswith('#'):
                                # Try different separators
                                if ';' in line:
                                    parts = line.strip().split(';')
                                elif ',' in line:
                                    parts = line.strip().split(',')
                                else:
                                    # Fall back to whitespace separator
                                    parts = line.strip().split()

                                if len(parts) >= 2:
                                    try:
                                        binding_energy = float(parts[0])
                                        intensity = float(parts[1])
                                        data.append([binding_energy, intensity])
                                    except ValueError:
                                        continue

                    if not data:
                        continue

                    # Create sheet
                    ws = wb.create_sheet(sheet_name)

                    # Add headers
                    ws["A1"] = "BE"
                    ws["B1"] = "Raw Data"

                    # Add data with .2f formatting
                    for i, (binding_energy, intensity) in enumerate(data, start=2):
                        ws[f"A{i}"] = f"{binding_energy:.2f}"
                        ws[f"B{i}"] = f"{intensity:.2f}"

                    processed_count += 1

                except Exception as e:
                    print(f"Error processing {os.path.basename(file_path)}: {e}")

            # Save and open if any sheets were created
            if len(wb.sheetnames) > 0:
                wb.save(excel_path)
                open_xlsx_file(self.window, excel_path)
                self.window.show_popup_message2("Success",
                                                f"Processed {processed_count} of {len(file_list)} ASC files.")
            else:
                self.window.show_popup_message2("Information",
                                                "No valid data found in any of the ASC files.")

        except Exception as e:
            import traceback
            traceback.print_exc()
            wx.MessageBox(f"Error importing multiple ASC files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)





def load_library_data_WITHEXCEL():
   wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
   sheet = wb['Library']
   data = {}
   for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
       element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
       key = (element, orbital)
       if key not in data:
           data[key] = {}
       data[key][instrument] = {
           'position': position,
           'ds': ds,
           'rsf': rsf,
           'row': row_idx
       }
   return data


def load_library_data_JSON_ONLY():
    with open('KherveFitting_library.json', 'r') as f:
        json_data = json.load(f)

    # Convert string keys back to tuples
    data = {}
    for k, v in json_data.items():
        element, orbital = k.split('_')
        data[(element, orbital)] = v
    return data


def load_library_data():
    """Load library data from parquet or json file"""
    import os

    parquet_file = 'KherveFitting_library.parquet'
    json_file = 'KherveFitting_library.json'

    if os.path.exists(parquet_file):
        print("Read Parquet Database")
        import pandas as pd
        df = pd.read_parquet(parquet_file)

        # Convert back to nested dict structure
        data = {}
        for _, row in df.iterrows():
            key = (row['element'], row['orbital'])
            if key not in data:
                data[key] = {}
            data[key][row['instrument']] = {
                'position': row['position'],
                'ds': row['ds'],
                'rsf': row['rsf'],
                'full_name': row['full_name'],
                'auger': row['auger'],
                'ke_be': row['ke_be']
            }
        return data

    elif os.path.exists(json_file):
        print("Convert Json Database")
        import json
        with open(json_file, 'r') as f:
            json_data = json.load(f)

        # Convert string keys back to tuples
        data = {}
        for k, v in json_data.items():
            element, orbital = k.split('_')
            data[(element, orbital)] = v
        return data

    else:
        raise FileNotFoundError("Neither parquet nor json library file found")




def load_library_data_NEWBUTNO():
    wb = openpyxl.load_workbook('KherveFitting_library.xlsx')
    sheet = wb['Library']
    data = {}
    instruments = set()  # Create set for unique instruments

    for row in sheet.iter_rows(min_row=2, values_only=True):
        element, orbital, full_name, auger, ke_be, position, ds, rsf, instrument = row
        instruments.add(instrument)  # Add each instrument to set
        key = (element, orbital)
        if key not in data:
            data[key] = {}
        data[key][instrument] = {
            'position': position,
            'ds': ds,
            'rsf': rsf
        }
    return data, sorted(list(instruments))  # Return data and sorted instruments list

def load_recent_files_from_config(window):
    config = window.load_config()
    window.recent_files = config.get('recent_files', [])
    update_recent_files_menu(window)

def update_recent_files(window, file_path):
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    window.save_config()  # Call save_config directly on the window object


def open_spe_file_OLD_PHI(window, file_path):
    try:
        from yadg.extractors.phi.spe import extract
        import openpyxl

        # Extract data from SPE file
        data = extract(fn=file_path)

        # Create new Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Process each core level
        for core_level in data:
            # Create sheet with core level name
            ws = wb.create_sheet(title=core_level)

            # Get energy and intensity values
            energy = data[core_level].coords["E"].values
            intensity = data[core_level].data_vars["y"].values

            # Set column headers
            ws["A1"] = "BE"
            ws["B1"] = "Corrected Data"
            ws["C1"] = "Raw Data"
            ws["D1"] = "Transmission"

            # Fill data
            for i, (e, inten) in enumerate(zip(energy, intensity), start=2):
                ws[f"A{i}"] = e
                ws[f"B{i}"] = inten
                ws[f"C{i}"] = inten
                ws[f"D{i}"] = 1.0

        # Create Experimental Description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Read header information using binary mode
        with open(file_path, 'rb') as f:
            header_lines = []
            in_header = False
            for line in f:
                try:
                    decoded_line = line.decode('utf-8', errors='ignore').strip()
                    if decoded_line == 'SOFH':
                        in_header = True
                        continue
                    if decoded_line == 'EOFH':
                        break
                    if in_header and decoded_line:
                        header_lines.append(decoded_line)
                except UnicodeDecodeError:
                    continue

        # Write header info to exp sheet
        for i, line in enumerate(header_lines, start=1):
            if ':' in line:
                key, value = line.split(':', 1)
                exp_sheet[f"A{i}"] = key.strip()
                exp_sheet[f"B{i}"] = value.strip()

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the created Excel file using existing function
        open_xlsx_file(window, excel_path)

    except Exception as e:
        # wx.MessageBox(f"Error processing SPE file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")


def open_spe_file(window, file_path):
    """Process a PHI SPE file and convert it to Excel format"""
    import numpy as np
    import openpyxl
    import re
    import os
    import struct

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Read file content
        with open(file_path, 'rb') as f:
            content = f.read()

        # Extract header
        header_match = re.search(rb'SOFH(.*?)EOFH', content, re.DOTALL)
        if not header_match:
            raise ValueError("Cannot find header section (SOFH...EOFH)")

        header_text = header_match.group(1).decode('utf-8', errors='ignore')
        header_lines = [line.strip() for line in header_text.strip().split('\n')]

        # Extract intensity calibration coefficients
        a, b = 31.826, 0.229  # Default values
        for line in header_lines:
            if 'IntensityCalCoeff:' in line:
                _, coeffs = line.split(':', 1)
                a, b = map(float, coeffs.strip().split())
                print(f"Extracted transmission function coefficients: a={a}, b={b}")
                break

        # Find active regions
        regions = []
        for line in header_lines:
            if line.startswith('SpectralRegDef:'):
                parts = line.split()
                if len(parts) >= 9 and int(parts[2]) == 1:  # Active region
                    region = {
                        'number': int(parts[1]),
                        'name': parts[3],
                        'atomic_number': int(parts[4]),
                        'num_points': int(parts[5]),
                        'start_energy': float(parts[7]),
                        'end_energy': float(parts[8])
                    }

                    # Normalize region name
                    if region['name'] == 'Su1s':
                        region['name'] = 'Survey'

                    regions.append(region)

        # Extract metadata for experimental description
        metadata = {
            'Sample ID': os.path.basename(file_path),
            'Date': '1970/1/1',
            'Time': '0:0:0',
            'Technique': 'XPS',
            'Species & Transition': '',
            'Number of scans': '1',
            'Source Label': 'Al',
            'Source Energy': '1486.6',
            'Source width X': '100',
            'Source width Y': '100',
            'Pass Energy': '224',
            'Work Function': '4.339',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': 'Kinetic Energy',
            'X Units': 'eV',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '1',
            'Collection Time': '0.16',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Update metadata from header
        for line in header_lines:
            if 'FileDateTime:' in line:
                parts = line.split(':', 1)[1].strip().split()
                if len(parts) >= 2:
                    date_part = parts[0]
                    time_part = parts[1]
                    metadata['Date'] = date_part
                    metadata['Time'] = time_part
            elif 'PassEnergy:' in line:
                metadata['Pass Energy'] = line.split(':', 1)[1].strip()
            elif 'WorkFunction:' in line:
                metadata['Work Function'] = line.split(':', 1)[1].strip()
            elif 'Comments:' in line:
                metadata['Block Comment'] = line.split(':', 1)[1].strip()
                metadata['# Comment Lines'] = '1'

        # Get binary data section
        data_start = content.find(b'EOFH') + 4
        binary_data = content[data_start:]

        # Implementation of FindFirstValidOffset
        def find_first_valid_offset(binary_data, region):
            num_points = region['num_points']

            # Iterate through possible offsets
            for offset in range(0, min(5000, len(binary_data) - num_points * 4), 2):
                try:
                    # Check first value - must be significant
                    first_val = struct.unpack('<f', binary_data[offset:offset + 4])[0]
                    if first_val < 0.5:
                        continue

                    # Try reading all points
                    values = []
                    for i in range(num_points):
                        data_offset = offset + (i * 4)
                        if data_offset + 4 <= len(binary_data):
                            val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                            if 0 <= val < 1e6:  # Reasonable value check
                                values.append(val)
                            else:
                                values = []
                                break
                        else:
                            values = []
                            break

                    # Check if we have the correct number of values with meaningful variation
                    if len(values) == num_points:
                        max_val = max(values)
                        min_val = min(values)
                        if max_val - min_val > 50:  # Meaningful variation
                            return offset
                except Exception:
                    pass

            # Try looking for marker patterns
            for offset in range(0, 5000, 2):
                try:
                    # Look for a marker (value near zero)
                    marker_val = struct.unpack('<f', binary_data[offset:offset + 4])[0]

                    # If found possible marker, check data after it
                    if 0 <= marker_val < 1e-6:
                        data_offset = offset + 4

                        # Validate the first actual data value
                        first_data_val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]

                        # Data must start with significant value
                        if first_data_val > 100:
                            # Check all values can be read
                            values = []
                            valid = True
                            for i in range(num_points):
                                point_offset = data_offset + (i * 4)
                                if point_offset + 4 <= len(binary_data):
                                    val = struct.unpack('<f', binary_data[point_offset:point_offset + 4])[0]
                                    if 0 <= val < 1e6:
                                        values.append(val)
                                    else:
                                        valid = False
                                        break
                                else:
                                    valid = False
                                    break

                            if valid and len(values) == num_points:
                                return data_offset
                except Exception:
                    pass

            # Try common offset 114 as a last resort
            try:
                values = []
                valid = True
                for i in range(num_points):
                    data_offset = 114 + (i * 4)
                    if data_offset + 4 <= len(binary_data):
                        val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                        if 0 <= val < 1e6:
                            values.append(val)
                        else:
                            valid = False
                            break
                    else:
                        valid = False
                        break

                if valid and len(values) == num_points:
                    max_val = max(values)
                    min_val = min(values)
                    if max_val - min_val > 100:
                        return 114
            except Exception:
                pass

            # If all else fails
            return -1

        # Implementation of CalculateAllOffsets
        def calculate_all_offsets(binary_data, regions, first_offset):
            if first_offset < 0:
                return []

            offsets = [first_offset]
            current_offset = first_offset

            # Calculate offsets for remaining regions
            for i in range(len(regions) - 1):
                # Next offset = current + (current region points * 4)
                current_offset += regions[i]['num_points'] * 4
                offsets.append(current_offset)

                # Verify this offset works
                valid = True
                for j in range(regions[i + 1]['num_points']):
                    data_offset = current_offset + (j * 4)
                    if data_offset + 4 <= len(binary_data):
                        try:
                            val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                            if val < 0 or val >= 1e6:
                                valid = False
                                break
                        except:
                            valid = False
                            break
                    else:
                        valid = False
                        break

                # If not valid, scan for a better offset
                if not valid:
                    test_offset = current_offset - 20
                    while test_offset < current_offset + 100:
                        test_valid = True
                        for j in range(regions[i + 1]['num_points']):
                            data_offset = test_offset + (j * 4)
                            if data_offset + 4 <= len(binary_data):
                                try:
                                    val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                                    if val < 0 or val >= 1e6:
                                        test_valid = False
                                        break
                                except:
                                    test_valid = False
                                    break
                            else:
                                test_valid = False
                                break

                        if test_valid:
                            offsets[i + 1] = test_offset
                            current_offset = test_offset
                            break

                        test_offset += 4

            return offsets

        # Find first valid offset
        first_offset = find_first_valid_offset(binary_data, regions[0])
        print(f"Found first valid offset at: {first_offset}")

        if first_offset < 0:
            raise ValueError("Could not find valid data offset in file")

        # Calculate offsets for all regions
        region_offsets = calculate_all_offsets(binary_data, regions, first_offset)
        print(f"Calculated offsets for all regions: {region_offsets}")

        if not region_offsets or len(region_offsets) != len(regions):
            raise ValueError("Failed to calculate valid offsets for all regions")

        # Process each region
        for i, region in enumerate(regions):
            sheet_name = region['name']
            num_points = region['num_points']
            offset = region_offsets[i]

            # Create sheet
            ws = wb.create_sheet(title=sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy (eV)"
            ws["B1"] = "Corrected Data"
            ws["C1"] = "Raw Data"
            ws["D1"] = "Transmission"

            # Update region-specific metadata
            region_metadata = metadata.copy()
            region_metadata['Species & Transition'] = sheet_name
            region_metadata['X Start'] = str(region['start_energy'])
            region_metadata['X Step'] = str((region['end_energy'] - region['start_energy']) / (num_points - 1))
            region_metadata['Num Y Values'] = str(num_points)

            # Read intensity values
            intensity_values = []
            for j in range(num_points):
                data_offset = offset + (j * 4)
                if data_offset + 4 <= len(binary_data):
                    val = struct.unpack('<f', binary_data[data_offset:data_offset + 4])[0]
                    if 0 <= val < 1e6:  # Reasonable value check
                        intensity_values.append(val)
                    else:
                        intensity_values.append(0.0)  # Use 0 for invalid values
                else:
                    intensity_values.append(0.0)

            # Make sure we have the expected number of points
            if len(intensity_values) < num_points:
                intensity_values.extend([0.0] * (num_points - len(intensity_values)))

            # Calculate energy values and transmission
            energy_values = np.linspace(region['start_energy'], region['end_energy'], num_points)
            ke_values = 1486.6 - energy_values  # Al K-alpha energy
            transmission = a * np.power(ke_values, -b)
            corrected_intensity = np.array(intensity_values) / transmission

            # Write data
            for j, (e, ci, ri, t) in enumerate(zip(energy_values, corrected_intensity, intensity_values, transmission),
                                               start=2):
                ws[f"A{j}"] = float(e)
                ws[f"B{j}"] = float(ci)
                ws[f"C{j}"] = float(ri)
                ws[f"D{j}"] = float(t)

            # Add experimental description at column AX (50)
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            # Add metadata fields
            row = 2
            for key, value in region_metadata.items():
                ws.cell(row=row, column=exp_col, value=key)
                ws.cell(row=row, column=exp_col + 1, value=value)
                row += 1

            # Set column widths for experimental description
            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40

        # Create Experimental description sheet
        exp_sheet = wb.create_sheet("Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata to the experimental description sheet
        row = 1
        exp_sheet.cell(row=row, column=1, value="Experimental Description")
        row += 1

        for key, value in metadata.items():
            exp_sheet.cell(row=row, column=1, value=key)
            exp_sheet.cell(row=row, column=2, value=value)
            row += 1

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the Excel file
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing SPE file: {str(e)}")
        return False




def open_spe_file_dialog(window):
    with wx.FileDialog(window, "Open SPE file", wildcard="PHI files (*.spe)|*.spe",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_spe_file(window, file_path)



def validate_data(df):
    """Validate the processed data"""
    issues = []

    if df['Binding Energy'].isna().any():
        issues.append("Missing binding energy values detected")

    if df['Raw Data'].isna().any():
        issues.append("Missing intensity values detected")

    if not np.allclose(df['Raw Data'], df['Corrected Data']):
        issues.append("Mismatch between raw and corrected data")

    if not np.allclose(df['Transmission'], 1.0):
        issues.append("Transmission values not all set to 1.0")

    return issues


def get_core_level_from_filename(filename):
    """
    Extract core level name from MRS filename.
    Examples:
    - xxx_C.MRS → C1s
    - xxx_CL.MRS → Cl2p
    - xxx_SU.MRS → Survey
    - xxx_O.MRS → O1s
    """
    import re

    # Define mapping of filename suffixes to standard core level names
    core_level_map = {
        'C': 'C1s',
        'N': 'N1s',
        'O': 'O1s',
        'F': 'F1s',
        'S': 'S2p',
        'CL': 'Cl2p',
        'BR': 'Br3d',
        'I': 'I3d',
        'SI': 'Si2p',
        'P': 'P2p',
        'B': 'B1s',
        'LI': 'Li1s',
        'NA': 'Na1s',
        'K': 'K2p',
        'CA': 'Ca2p',
        'MG': 'Mg2p',
        'AL': 'Al2p',
        'TI': 'Ti2p',
        'V': 'V2p',
        'CR': 'Cr2p',
        'MN': 'Mn2p',
        'FE': 'Fe2p',
        'CO': 'Co2p',
        'NI': 'Ni2p',
        'CU': 'Cu2p',
        'ZN': 'Zn2p',
        'GA': 'Ga3d',
        'GE': 'Ge3d',
        'AS': 'As3d',
        'SE': 'Se3d',
        'RB': 'Rb3d',
        'SR': 'Sr3d',
        'Y': 'Y3d',
        'ZR': 'Zr3d',
        'NB': 'Nb3d',
        'MO': 'Mo3d',
        'RU': 'Ru3d',
        'RH': 'Rh3d',
        'PD': 'Pd3d',
        'AG': 'Ag3d',
        'CD': 'Cd3d',
        'IN': 'In3d',
        'SN': 'Sn3d',
        'SB': 'Sb3d',
        'TE': 'Te3d',
        'CS': 'Cs3d',
        'BA': 'Ba3d',
        'LA': 'La3d',
        'CE': 'Ce3d',
        'HF': 'Hf4f',
        'TA': 'Ta4f',
        'W': 'W4f',
        'RE': 'Re4f',
        'OS': 'Os4f',
        'IR': 'Ir4f',
        'PT': 'Pt4f',
        'AU': 'Au4f',
        'HG': 'Hg4f',
        'PB': 'Pb4f',
        'BI': 'Bi4f',
        'TH': 'Th4f',
        'U': 'U4f',
        'SU': 'Survey',
        'VB': 'VB'  # Valence Band
    }

    # Extract the suffix after the last underscore and before the extension
    base_name = os.path.basename(filename)
    match = re.search(r'_([A-Z]+)\.mrs$', base_name.upper())

    if match:
        suffix = match.group(1)
        return core_level_map.get(suffix, suffix)
    # If no underscore or no match in mapping, try to use the whole filename
    # This is a fallback for unusual naming conventions
    base_name_without_ext = os.path.splitext(base_name)[0]
    for key, value in core_level_map.items():
        if key in base_name_without_ext.upper():
            return value

    # Last resort: just return the base filename without extension
    return base_name_without_ext


def open_mrs_file(window, file_path):
    """
    Import a Physical Electronics MRS file (XPS data) and convert it to Excel format
    suitable for KherveFitting.
    """
    try:
        # Read the MRS file content
        with open(file_path, 'r', errors='ignore') as f:
            content = f.read()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Extract sample info
        sample_name = os.path.splitext(os.path.basename(file_path))[0]
        desc_match = re.search(r'desc=(.*?)[\r\n]', content)
        sample_description = desc_match.group(1).strip() if desc_match else "Unknown"

        # Determine sheet name from filename using the new function
        sheet_name = get_core_level_from_filename(file_path)

        # Create sheet
        ws = wb.create_sheet(title=sheet_name)

        # Find data section
        data_section_match = re.search(
            r'array_size=(\d+).*?data=Data Array.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
            content, re.DOTALL)

        # If not found, try with Auto Survey
        if not data_section_match:
            data_section_match = re.search(
                r'array_size=(\d+).*?data=Auto Survey.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                content, re.DOTALL)

        # If still not found, try a more general pattern
        if not data_section_match:
            data_section_match = re.search(r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                                           content, re.DOTALL)

        if data_section_match:
            array_size = int(data_section_match.group(1))
            be_start = float(data_section_match.group(2))
            be_end = float(data_section_match.group(3))
            data_text = data_section_match.group(4).strip()

            # Extract numeric values
            data_values = []
            for line in data_text.split('\n'):
                line = line.strip()
                if line and line.isdigit():
                    data_values.append(int(line))

            if data_values:
                # Calculate binding energy values
                num_points = len(data_values)
                step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                be_values = [be_end - i * step_size for i in range(num_points)]

                # Add headers
                ws["A1"] = "Binding Energy (eV)"
                ws["B1"] = "Corrected Data"
                ws["C1"] = "Raw Data"
                ws["D1"] = "Transmission"

                # Add data
                for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                    ws[f"A{i}"] = be
                    ws[f"B{i}"] = intensity
                    ws[f"C{i}"] = intensity
                    ws[f"D{i}"] = 1.0  # Default transmission value

                # Create experimental description sheet
                exp_sheet = wb.create_sheet("Experimental description")
                exp_sheet.column_dimensions['A'].width = 30
                exp_sheet.column_dimensions['B'].width = 50

                # Add basic info to experimental sheet
                exp_sheet["A1"] = "File"
                exp_sheet["B1"] = os.path.basename(file_path)
                exp_sheet["A2"] = "Energy Range"
                exp_sheet["B2"] = f"{be_start} - {be_end} eV"
                exp_sheet["A3"] = "Number of Points"
                exp_sheet["B3"] = str(num_points)
                exp_sheet["A4"] = "Description"
                exp_sheet["B4"] = sample_description
                exp_sheet["A5"] = "Core Level"
                exp_sheet["B5"] = sheet_name

                # Extract additional metadata if available
                metadata_matches = {
                    "Scan Count": re.search(r'scan_total=(\d+)', content),
                    "Resolution": re.search(r'res=(\d+)', content),
                    "Spot Size": re.search(r'spot=(\d+)', content),
                    "Date": re.search(r'time_stamp=(.*?)[\r\n]', content),
                    "Operator": re.search(r'oper=(.*?)[\r\n]', content)
                }

                row = 6
                for key, match in metadata_matches.items():
                    if match:
                        exp_sheet[f"A{row}"] = key
                        exp_sheet[f"B{row}"] = match.group(1).strip()
                        row += 1

                # Save Excel file
                excel_path = os.path.splitext(file_path)[0] + ".xlsx"
                wb.save(excel_path)

                # Open the created Excel file
                from libraries.FileMenu.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
                return True
            else:
                window.show_popup_message2("Error", "No valid data found in the MRS file.")
                return False
        else:
            window.show_popup_message2("Error", "Could not locate data section in MRS file.")
            return False

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS file: {str(e)}")
        return False


def import_multiple_mrs_files(window):
    """
    Import multiple Physical Electronics MRS files from a folder.
    """
    with wx.DirDialog(window, "Choose a directory containing MRS files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all MRS files in the directory
        mrs_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.mrs')]

        if not mrs_files:
            window.show_popup_message2("Information", "No MRS files found in the selected folder.")
            return

        # Ask if user wants individual Excel files or one combined file
        dlg = wx.MessageDialog(window,
                               "Do you want to create individual Excel files for each MRS file or combine them into one Excel file?",
                               "Import Options",
                               wx.YES_NO | wx.ICON_QUESTION)
        dlg.SetYesNoLabels("Individual Files", "One Combined File")

        result = dlg.ShowModal()
        individual_files = (result == wx.ID_YES)
        dlg.Destroy()

        if individual_files:
            # Process each MRS file individually
            processed_count = 0
            for mrs_file in mrs_files:
                mrs_path = os.path.join(folder_path, mrs_file)
                if open_mrs_file(window, mrs_path):
                    processed_count += 1

            window.show_popup_message2("Success", f"Processed {processed_count} of {len(mrs_files)} MRS files.")
        else:
            # Combine all MRS files into one Excel file
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            processed_count = 0
            for mrs_file in mrs_files:
                mrs_path = os.path.join(folder_path, mrs_file)

                try:
                    # Read the MRS file content
                    with open(mrs_path, 'r', errors='ignore') as f:
                        content = f.read()

                    # Determine sheet name
                    sheet_name = get_core_level_from_filename(mrs_file)

                    # Ensure unique sheet name
                    suffix = 1
                    original_name = sheet_name
                    while sheet_name in combined_wb.sheetnames:
                        sheet_name = f"{original_name}{suffix}"
                        suffix += 1

                    # Ensure unique sheet name
                    suffix = 1
                    original_name = sheet_name
                    while sheet_name in combined_wb.sheetnames:
                        sheet_name = f"{original_name}{suffix}"
                        suffix += 1

                    # Create sheet
                    ws = combined_wb.create_sheet(title=sheet_name)

                    # Find and extract data
                    data_section_match = re.search(
                        r'array_size=(\d+).*?(?:data=Data Array|data=Auto Survey).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                        content, re.DOTALL)

                    if not data_section_match:
                        data_section_match = re.search(
                            r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                            content, re.DOTALL)

                    if data_section_match:
                        array_size = int(data_section_match.group(1))
                        be_start = float(data_section_match.group(2))
                        be_end = float(data_section_match.group(3))
                        data_text = data_section_match.group(4).strip()

                        # Extract numeric values
                        data_values = []
                        for line in data_text.split('\n'):
                            line = line.strip()
                            if line and line.isdigit():
                                data_values.append(int(line))

                        if data_values:
                            # Calculate binding energy values
                            num_points = len(data_values)
                            step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                            be_values = [be_end - i * step_size for i in range(num_points)]

                            # Add headers
                            ws["A1"] = "Binding Energy (eV)"
                            ws["B1"] = "Corrected Data"
                            ws["C1"] = "Raw Data"
                            ws["D1"] = "Transmission"

                            # Add data
                            for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                                ws[f"A{i}"] = be
                                ws[f"B{i}"] = intensity
                                ws[f"C{i}"] = intensity
                                ws[f"D{i}"] = 1.0  # Default transmission value

                            processed_count += 1
                except Exception as e:
                    print(f"Error processing {mrs_file}: {str(e)}")

            # Create experimental description sheet
            exp_sheet = combined_wb.create_sheet("Experimental description")
            exp_sheet.column_dimensions['A'].width = 30
            exp_sheet.column_dimensions['B'].width = 50

            exp_sheet["A1"] = "Source Folder"
            exp_sheet["B1"] = folder_path
            exp_sheet["A2"] = "Processed Files"
            exp_sheet["B2"] = f"{processed_count} of {len(mrs_files)}"

            # Save combined Excel file
            folder_name = os.path.basename(folder_path)
            excel_path = os.path.join(folder_path, f"{folder_name}_combined.xlsx")
            combined_wb.save(excel_path)

            # Open the combined Excel file
            if processed_count > 0:
                from libraries.FileMenu.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
                window.show_popup_message2("Success",
                                           f"Created combined Excel file with {processed_count} sheets from MRS files.")
            else:
                window.show_popup_message2("Error", "No valid data found in any MRS file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS files: {str(e)}")


def import_mrs_file(window):
    """
    Import a Physical Electronics MRS file (XPS data) and convert it to Excel format
    suitable for KherveFitting.
    """
    import wx
    import os
    import re
    from openpyxl import Workbook

    with wx.FileDialog(window, "Open MRS file", wildcard="MRS files (*.mrs)|*.mrs",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Read the MRS file content
        with open(file_path, 'r', errors='ignore') as f:
            content = f.read()

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Extract sample info
        sample_name = os.path.splitext(os.path.basename(file_path))[0]
        desc_match = re.search(r'desc=(.*?)[\r\n]', content)
        if desc_match:
            sample_description = desc_match.group(1).strip()

        # Determine sheet name from filename
        # base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Look for core level identifier in filename
        sheet_name = get_core_level_from_filename(file_path)
        if "_" in sheet_name:
            core_level = sheet_name.split("_")[1]
            if core_level.lower() in ["c1s", "o1s", "n1s", "s2p", "su", "vb"]:
                if core_level.lower() == "su":
                    sheet_name = "Survey"
                else:
                    sheet_name = core_level.upper()

        # Create sheet
        ws = wb.create_sheet(title=sheet_name)

        # Find data section
        data_section_match = re.search(
            r'array_size=(\d+).*?data=Data Array.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
            content, re.DOTALL)

        # If not found, try with Auto Survey
        if not data_section_match:
            data_section_match = re.search(
                r'array_size=(\d+).*?data=Auto Survey.*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                content, re.DOTALL)

        # If still not found, try a more general pattern
        if not data_section_match:
            data_section_match = re.search(r'array_size=(\d+).*?lo_be=([0-9.-]+).*?up_be=([0-9.-]+).*?!(.*?)!',
                                           content, re.DOTALL)

        if data_section_match:
            array_size = int(data_section_match.group(1))
            be_start = float(data_section_match.group(2))
            be_end = float(data_section_match.group(3))
            data_text = data_section_match.group(4).strip()

            # Extract numeric values
            data_values = []
            for line in data_text.split('\n'):
                line = line.strip()
                if line and line.isdigit():
                    data_values.append(int(line))

            if data_values:
                # Calculate binding energy values
                num_points = len(data_values)
                step_size = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0
                be_values = [be_end - i * step_size for i in range(num_points)]

                # Add headers
                ws["A1"] = "Binding Energy (eV)"
                ws["B1"] = "Raw Data"

                # Add data
                for i, (be, intensity) in enumerate(zip(be_values, data_values), start=2):
                    ws[f"A{i}"] = be
                    ws[f"B{i}"] = intensity

                # Create experimental description sheet
                exp_sheet = wb.create_sheet("Experimental description")
                exp_sheet.column_dimensions['A'].width = 30
                exp_sheet.column_dimensions['B'].width = 50

                # Add basic info to experimental sheet
                exp_sheet["A1"] = "File"
                exp_sheet["B1"] = os.path.basename(file_path)
                exp_sheet["A2"] = "Energy Range"
                exp_sheet["B2"] = f"{be_start} - {be_end} eV"
                exp_sheet["A3"] = "Number of Points"
                exp_sheet["B3"] = str(num_points)

                if desc_match:
                    exp_sheet["A4"] = "Description"
                    exp_sheet["B4"] = sample_description

                # Save and open Excel file
                excel_path = os.path.splitext(file_path)[0] + ".xlsx"
                wb.save(excel_path)

                # Open the created Excel file
                from libraries.FileMenu.Open import open_xlsx_file
                open_xlsx_file(window, excel_path)
            else:
                window.show_popup_message2("Error", "No valid data found in the MRS file.")
        else:
            window.show_popup_message2("Error", "Could not locate data section in MRS file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing MRS file: {str(e)}")


def extract_acquisition_parameters(sheet):
    """Extract acquisition parameters from Avantage sheet"""
    parameters = {}

    # Debug: Print sheet info
    print(f"Extracting parameters from sheet with {sheet.max_row} rows and {sheet.max_column} columns")

    # Search for "Acquisition Parameters" text in the first few rows across all columns
    acquisition_col = None
    for row in range(1, min(5, sheet.max_row + 1)):  # Check first 4 rows
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and "Acquisition Parameters" in str(cell_value):
                acquisition_col = col
                print(f"Found 'Acquisition Parameters' at row {row}, column {col}")
                break
        if acquisition_col:
            break

    if acquisition_col is None:
        print("No 'Acquisition Parameters' found - trying alternative search")
        # Alternative search for just "Parameters"
        for row in range(1, min(5, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and "Parameters" in str(cell_value):
                    acquisition_col = col
                    print(f"Found 'Parameters' at row {row}, column {col}")
                    break
            if acquisition_col:
                break

    if acquisition_col is None:
        print("No parameters column found")
        return parameters

    # Extract parameters starting from row 4, but also check nearby rows
    for row in range(3, min(15, sheet.max_row + 1)):  # Extended range
        # Parameter name is in acquisition_col
        param_name = sheet.cell(row=row, column=acquisition_col).value
        # Parameter value is in acquisition_col + 1
        param_value = sheet.cell(row=row, column=acquisition_col + 1).value
        # Sometimes additional value in acquisition_col + 2
        param_value_j = sheet.cell(row=row, column=acquisition_col + 2).value

        if param_name is None:
            continue

        param_name_str = str(param_name).strip()
        if not param_name_str or param_name_str in ['Parameter', 'Parameters', '']:
            continue

        # Build the parameter value
        value_parts = []
        if param_value is not None and str(param_value).strip():
            value_parts.append(str(param_value).strip())
        if param_value_j is not None and str(param_value_j).strip():
            value_parts.append(str(param_value_j).strip())

        if value_parts:
            final_value = " ".join(value_parts)
            parameters[param_name_str] = final_value
            print(f"Found parameter: {param_name_str} = {final_value}")

    print(f"Total parameters found: {len(parameters)}")
    return parameters


def convert_avantage_peak_name(avantage_name):
    """
    Convert Avantage peak names to KherveFitting format.
    Examples:
        "Sr3d5 Sr-F2" -> "Sr3d5/2 Sr-F2"
        "Sr3d3 Sr-F2" -> "Sr3d3/2 Sr-F2"
        "Ti2p3 ?" -> "Ti2p3/2 ?"
        "O1s O-Ti" -> "O1s O-Ti" (no change for s orbitals)
    """
    import re

    # Pattern to match element + orbital + incomplete spin notation
    # Matches: Ti2p3, Sr3d5, Au4f7, etc.
    pattern = r'([A-Z][a-z]?)(\d+)([pdf])(\d)(\s|$)'

    def replace_orbital(match):
        element = match.group(1)  # Ti, Sr, Au, etc.
        shell = match.group(2)  # 2, 3, 4, etc.
        orbital_type = match.group(3)  # p, d, or f
        j_value = match.group(4)  # 3, 5, 7, 1, etc.
        space = match.group(5)  # space or end of string

        # Convert based on orbital type and j value
        if orbital_type == 'p':
            if j_value == '3':
                j_full = '3/2'
            elif j_value == '1':
                j_full = '1/2'
            else:
                j_full = j_value  # Keep as is if unexpected
        elif orbital_type == 'd':
            if j_value == '5':
                j_full = '5/2'
            elif j_value == '3':
                j_full = '3/2'
            else:
                j_full = j_value
        elif orbital_type == 'f':
            if j_value == '7':
                j_full = '7/2'
            elif j_value == '5':
                j_full = '5/2'
            else:
                j_full = j_value
        else:
            j_full = j_value

        return f"{element}{shell}{orbital_type}{j_full}{space}"

    converted_name = re.sub(pattern, replace_orbital, avantage_name)
    return converted_name


def extract_peak_fitting_from_peak_table(wb):
    """
    Extract peak fitting parameters from Peak Table sheet in Avantage file.
    Returns a dictionary mapping scan names to their peak parameters.
    Each peak has 2 rows: values row and constraints row.
    Converts Avantage ref letters (L, M, N) to KherveFitting letters (A, B, C).
    """
    peak_fitting_data = {}

    # Check if Peak Table sheet exists
    if "Peak Table" not in wb.sheetnames:
        print("No Peak Table sheet found")
        return peak_fitting_data

    sheet = wb["Peak Table"]
    print(f"Parsing Peak Table sheet with {sheet.max_row} rows")

    current_scan = None
    header_row = None
    ref_letter_to_index = {}  # Maps L→A, M→B, etc.

    def convert_constraint_references(constraint_str, letter_mapping):
        """Convert constraint references from Avantage letters to KherveFitting letters"""
        if not constraint_str or not isinstance(constraint_str, str):
            return constraint_str

        import re
        # Pattern to match letter references (L, M, N, etc.) in constraints
        # Matches: L*1, M+1.00, L+1.00 (+0.2), etc.
        for avantage_letter, kherve_letter in letter_mapping.items():
            # Replace the letter at word boundaries or followed by operators
            constraint_str = re.sub(rf'\b{avantage_letter}\b', kherve_letter, constraint_str)

        return constraint_str

    def clean_range_constraint(constraint_str):
        """Remove :00 suffix from range constraints (e.g., '12:37:00' -> '12:37')"""
        if not constraint_str or not isinstance(constraint_str, str):
            return constraint_str

        # Check if it's a range constraint with :00 at the end
        if ':' in constraint_str and constraint_str.endswith(':00'):
            return constraint_str[:-3]  # Remove last 3 characters (:00)

        return constraint_str

    row_idx = 1
    while row_idx <= sheet.max_row:
        # Check for Peak Fit Table headers like "Peak Fit Table : O1s Scan"
        cell_value = sheet.cell(row=row_idx, column=1).value

        if cell_value and isinstance(cell_value, str) and "Peak Fit Table :" in cell_value:
            # Extract scan name (e.g., "O1s")
            parts = cell_value.split(":")
            if len(parts) >= 2:
                scan_part = parts[1].strip()
                # Extract element name like O1s, Sr3d, Ti2p from "O1s Scan"
                scan_name = scan_part.split()[0].strip()
                current_scan = scan_name
                peak_fitting_data[current_scan] = []
                header_row = row_idx + 1  # Next row should be headers
                ref_letter_to_index = {}  # Reset mapping for each scan
                print(f"\nFound Peak Fit Table for: {current_scan}")
                row_idx += 2  # Skip header row
                continue

        # Parse peak data rows (must have current_scan set)
        if current_scan and header_row and row_idx > header_row:
            # Check if this is a data row (has Ref value)
            ref_value = sheet.cell(row=row_idx, column=1).value

            # Stop if we hit another section or empty ref
            if not ref_value:
                row_idx += 1
                continue

            if "Peak Fit Table" in str(ref_value):
                current_scan = None
                header_row = None
                continue

            # This is a data row - extract values
            name_value = sheet.cell(row=row_idx, column=2).value
            peak_be = sheet.cell(row=row_idx, column=3).value
            height_cps = sheet.cell(row=row_idx, column=4).value
            area_cps_ev = sheet.cell(row=row_idx, column=6).value
            fwhm_param = sheet.cell(row=row_idx, column=8).value
            lg_mix = sheet.cell(row=row_idx, column=9).value

            # Convert ref letter (L, M, N...) to sequential (A, B, C...)
            ref_str = str(ref_value).strip()
            if ref_str not in ref_letter_to_index:
                ref_letter_to_index[ref_str] = chr(65 + len(ref_letter_to_index))  # A, B, C...

            peak_letter = ref_letter_to_index[ref_str]

            # Get constraints from the next row (constraint row)
            constraint_row_idx = row_idx + 1
            position_constraint = None
            fwhm_constraint = None
            lg_constraint = None

            if constraint_row_idx <= sheet.max_row:
                # Position constraint (column 3)
                pos_const_val = sheet.cell(row=constraint_row_idx, column=3).value
                if pos_const_val:
                    position_constraint = convert_constraint_references(
                        str(pos_const_val).strip(), ref_letter_to_index)
                    position_constraint = clean_range_constraint(position_constraint)

                # FWHM constraint (column 8)
                fwhm_const_val = sheet.cell(row=constraint_row_idx, column=8).value
                if fwhm_const_val:
                    fwhm_constraint = convert_constraint_references(
                        str(fwhm_const_val).strip(), ref_letter_to_index)
                    fwhm_constraint = clean_range_constraint(fwhm_constraint)

                # L/G constraint (column 9)
                lg_const_val = sheet.cell(row=constraint_row_idx, column=9).value
                if lg_const_val:
                    lg_constraint = convert_constraint_references(
                        str(lg_const_val).strip(), ref_letter_to_index)
                    lg_constraint = clean_range_constraint(lg_constraint)

            # Validate and convert values
            try:
                peak_be_val = float(peak_be) if peak_be and not isinstance(peak_be, str) else 0.0
                height_val = float(height_cps) if height_cps else 0.0
                area_val = float(area_cps_ev) if area_cps_ev else 0.0
                fwhm_val = float(fwhm_param) if fwhm_param and not isinstance(fwhm_param, str) else 1.5
                lg_val = float(lg_mix) if lg_mix and not isinstance(lg_mix, str) else 20.0

                # Convert Avantage peak name to KherveFitting format
                raw_name = str(name_value).strip() if name_value else f"Peak_{peak_letter}"
                converted_name = convert_avantage_peak_name(raw_name)

                peak_data = {
                    'ref': peak_letter,
                    'name': converted_name,
                    'position': peak_be_val,
                    'height': height_val,
                    'area': area_val,
                    'fwhm': fwhm_val,
                    'lg_ratio': lg_val,
                    'position_constraint': position_constraint,
                    'fwhm_constraint': fwhm_constraint,
                    'lg_constraint': lg_constraint
                }

                peak_fitting_data[current_scan].append(peak_data)
                print(f"  Peak {peak_letter}: {peak_data['name']} at {peak_be_val:.2f} eV")
                if position_constraint:
                    print(f"    Position constraint: {position_constraint}")
                if fwhm_constraint:
                    print(f"    FWHM constraint: {fwhm_constraint}")
                if lg_constraint:
                    print(f"    L/G constraint: {lg_constraint}")

            except (ValueError, TypeError) as e:
                print(f"Error parsing peak data at row {row_idx}: {e}")

            # Skip constraint row
            row_idx += 2
            continue

        row_idx += 1

    return peak_fitting_data

def import_avantage_file_direct(window, file_path):
    import re
    import openpyxl
    import math
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(file_path)
    new_file_path = os.path.splitext(file_path)[0] + "_Kfitting.xlsx"
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    # Group sheets by sample based on letter suffix
    sample_groups = {}
    for sheet_name in wb.sheetnames:
        if "Survey" in sheet_name or "Scan" in sheet_name:
            # Detect sample suffix (single letter at the end after space)
            parts = sheet_name.split()
            if len(parts) >= 2 and len(parts[-1]) == 1 and parts[-1].isalpha():
                suffix = parts[-1]
            else:
                suffix = None  # No suffix = first sample

            if suffix not in sample_groups:
                sample_groups[suffix] = []
            sample_groups[suffix].append(sheet_name)

    # Sort samples: None (no suffix) first, then alphabetically
    sorted_samples = sorted(sample_groups.keys(), key=lambda x: (x is not None, x))

    # Extract peak fitting data from Peak Table sheet
    peak_fitting_by_scan = extract_peak_fitting_from_peak_table(wb)
    print(f"Extracted peak fitting data for {len(peak_fitting_by_scan)} scans")

    # Process each sample group
    sample_counter = 0
    for sample_suffix in sorted_samples:
        sheets_to_process = sample_groups[sample_suffix]

        # Extract Title sheet information for this sample
        title_info = extract_title_sheet_info(wb, sample_suffix)

        # Track VB, Fermi and Cut-Off counts for numbering within each sample
        vb_count = 0
        fermi_count = 0
        cutoff_count = 0
        used_names = set()

        for sheet_name in sheets_to_process:
            sheet = wb[sheet_name]

            # Extract acquisition parameters
            acquisition_params = extract_acquisition_parameters(sheet)

            # Extract element name and handle special cases
            lower_name = sheet_name.lower()
            if "Survey" in sheet_name or "survey" in sheet_name:
                base_name = "Survey"
            elif any(term in lower_name for term in
                     ['valence', 'valence band', 'valence scan', 'vb scan', 'vb', 'valence band scan']):
                if vb_count == 0:
                    base_name = "VB"
                else:
                    base_name = f"VB{vb_count + 1}"
                vb_count += 1
            elif any(term in lower_name for term in ['fermi', 'fermi scan']):
                if fermi_count == 0:
                    base_name = "Fermi"
                else:
                    base_name = f"Fermi{fermi_count + 1}"
                fermi_count += 1
            elif any(term in lower_name for term in ['cut-off', 'cutoff', 'cut off']):
                if cutoff_count == 0:
                    base_name = "Cut-Off"
                else:
                    base_name = f"Cut-Off{cutoff_count + 1}"
                cutoff_count += 1
            else:
                parts = sheet_name.split()
                base_name = parts[0]  # Get element (C1s, O1s, etc.)

            # Apply sample numbering for multi-sample files
            if sample_counter == 0:
                final_name = base_name
            else:
                final_name = f"{base_name}{sample_counter}"

            # Check if multi-sample (D8 cell not empty and > 1)
            multi_sample = False
            num_samples = 1
            if sheet.cell(row=8, column=4).value is not None:
                try:
                    num_samples = int(sheet.cell(row=8, column=4).value)
                    if num_samples > 1:
                        multi_sample = True
                except (ValueError, TypeError):
                    pass

            # Find data start row (default to 19 for Thermo files)
            start_row = 19
            for row_idx in range(1, sheet.max_row + 1):
                if sheet.cell(row=row_idx, column=1).value == "eV":
                    start_row = row_idx + 1
                    break

            if multi_sample:
                # Process each sample into its own sheet
                for sample_idx in range(num_samples):
                    if sample_counter == 0:
                        sample_name = f"{base_name}{sample_idx if sample_idx > 0 else ''}"
                    else:
                        if sample_idx == 0:
                            sample_name = f"{base_name}{sample_counter}"
                        else:
                            sample_name = f"{base_name}{sample_counter}_{sample_idx}"

                    new_sheet = new_wb.create_sheet(sample_name)
                    new_sheet['A1'] = "Binding Energy"
                    new_sheet['B1'] = "Raw Data"

                    # A column is binding energy (col 1), data is in col C+sample_idx (col 3+sample_idx)
                    data_col = 3 + sample_idx  # C is 3, D is 4, etc.

                    # Copy data
                    row_new = 2
                    for row in range(start_row, sheet.max_row + 1):

                        be_value = sheet.cell(row=row, column=1).value
                        intensity_value = sheet.cell(row=row, column=data_col).value

                        # Simple validation - check if values can be converted to float
                        try:
                            be_val = float(be_value)
                            raw_val = float(intensity_value)
                        except (ValueError, TypeError):
                            print("Skipping invalid data row (simple float conversion):", be_value, intensity_value)
                            continue

                        new_sheet.cell(row=row_new, column=1, value=f"{be_value:.2f}")
                        new_sheet.cell(row=row_new, column=2, value=f"{intensity_value}")
                        row_new += 1

                    # Add experimental description with Title sheet info and acquisition parameters
                    if acquisition_params or title_info:
                        exp_col = 45  # Column AS
                        new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                        current_row = 2

                        # Add Title sheet information first
                        for param_name, param_value in title_info.items():
                            new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                            new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                            current_row += 1

                        # Add acquisition parameters
                        for param_name, param_value in acquisition_params.items():
                            new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                            new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                            current_row += 1

                        # Set column widths
                        new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                        new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

            else:
                # Single sample - process as before
                new_name = final_name
                # Extract number if present in format like "C1s Scan (2)"
                number_match = re.search(r'\((\d+)\)', sheet_name)
                if number_match:
                    number = number_match.group(1)
                    if sample_counter == 0:
                        new_name = f"{base_name}{number}"
                    else:
                        new_name = f"{base_name}{sample_counter}_{number}"

                new_sheet = new_wb.create_sheet(new_name)
                new_sheet['A1'] = "Binding Energy"
                new_sheet['B1'] = "Raw Data"

                for row in range(start_row, sheet.max_row + 1):
                    be_value = sheet.cell(row=row, column=1).value
                    intensity_value = sheet.cell(row=row, column=3).value  # Column C

                    # Simple validation - check if values can be converted to float
                    try:
                        be_val = float(be_value)
                        raw_val = float(intensity_value)
                    except (ValueError, TypeError):
                        print("Skipping invalid data row (simple float conversion):", be_value, intensity_value)
                        continue

                    new_sheet['A{}'.format(row - start_row + 2)] = f"{be_value:.2f}"
                    new_sheet['B{}'.format(row - start_row + 2)] = f"{intensity_value:.2f}"

                # Add experimental description with Title sheet info and acquisition parameters
                if acquisition_params or title_info:
                    exp_col = 45  # Column AS
                    new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                    current_row = 2

                    # Add Title sheet information first
                    for param_name, param_value in title_info.items():
                        new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                        new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                        current_row += 1

                    # Add acquisition parameters
                    for param_name, param_value in acquisition_params.items():
                        new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                        new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                        current_row += 1


                    # Set column widths
                    new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                    new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        sample_counter += 1

    sample_counter += 1

    new_wb.save(new_file_path)

    # Open the file first to let it create the proper data structure
    open_xlsx_file(window, new_file_path)

    # Add peak fitting data directly
    add_peak_fitting_data_after_load(window, peak_fitting_by_scan, new_file_path)


def add_peak_fitting_data_after_load(window, peak_fitting_by_scan, file_path):
    """Add peak fitting data after file is fully loaded"""
    if not peak_fitting_by_scan:
        return

    import json
    import re
    import numpy as np
    from libraries.FileMenu.Save import convert_to_serializable_and_round
    from libraries.Peak_Functions import BackgroundCalculations

    print(f"Adding peak fitting data for {len(peak_fitting_by_scan)} scans to window.Data")

    # Add peak fitting to each matching sheet in window.Data
    for sheet_name in window.Data['Core levels'].keys():
        # Extract base name (e.g., "O1s" from "O1s0")
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        base_name = match.group(1) if match else sheet_name

        if base_name in peak_fitting_by_scan:
            print(f"  Adding peaks to {sheet_name}")

            # Initialize Fitting structure if not exists
            if 'Fitting' not in window.Data['Core levels'][sheet_name]:
                window.Data['Core levels'][sheet_name]['Fitting'] = {}

            window.Data['Core levels'][sheet_name]['Fitting']['Peaks'] = {}
            window.Data['Core levels'][sheet_name]['Fitting']['Fitting Model'] = 'SGL (Area)'

            for peak in peak_fitting_by_scan[base_name]:
                peak_name = peak['name']

                # Build constraints
                position_const = peak.get('position_constraint')
                if position_const:
                    pos_constraint = position_const
                else:
                    pos_constraint = f"{peak['position'] - 2:.2f}:{peak['position'] + 2:.2f}"

                fwhm_const = peak.get('fwhm_constraint')
                if fwhm_const:
                    fwhm_constraint = fwhm_const
                else:
                    fwhm_constraint = "0.30:3.50"

                lg_const = peak.get('lg_constraint')
                if lg_const:
                    lg_constraint = lg_const
                else:
                    lg_constraint = "0.00:100.00"

                window.Data['Core levels'][sheet_name]['Fitting']['Peaks'][peak_name] = {
                    'Position': float(f"{peak['position']:.2f}"),
                    'Height': float(f"{peak['height']:.2f}"),
                    'FWHM': float(f"{peak['fwhm']:.2f}"),
                    'L/G': float(f"{peak['lg_ratio']:.2f}"),
                    'Area': float(f"{peak['area']:.2f}"),
                    'Sigma': 0.00,
                    'Gamma': 0.00,
                    'Skew': 0.00,
                    'Fitting Model': 'SGL (Area)',
                    'Constraints': {
                        'Position': pos_constraint,
                        'Height': f"{max(0, peak['height'] * 0.5):.2f}:{peak['height'] * 1.5:.2f}",
                        'FWHM': fwhm_constraint,
                        'L/G': lg_constraint,
                        'Area': f"{max(0, peak['area'] * 0.5):.2f}:{peak['area'] * 1.5:.2f}",
                        'Sigma': "0.10:1.00",
                        'Gamma': "0.10:1.00",
                        'Skew': "0.01:2.00"
                    }
                }

            # Create smart background with region from 10% to 90% of range
            print(f"  Creating smart background for {sheet_name}")

            # Get data
            x_values = np.array(window.Data['Core levels'][sheet_name]['B.E.'])
            y_values = np.array(window.Data['Core levels'][sheet_name]['Raw Data'])

            # Calculate 10% to 90% range
            x_min = np.min(x_values)
            x_max = np.max(x_values)
            x_range = x_max - x_min

            # 10% from min, 90% from min (usually BE decreases, so max is "left", min is "right")
            min_range = float(f"{(x_min + 0.1 * x_range):.2f}")
            max_range = float(f"{(x_min + 0.9 * x_range):.2f}")

            # Ensure proper order for BE scale
            if min_range > max_range:
                min_range, max_range = max_range, min_range

            print(f"    Background range: {min_range:.2f} to {max_range:.2f} eV")

            # Initialize background as raw data
            current_background = y_values.copy()

            # Calculate smart background
            try:
                smart_background = BackgroundCalculations.calculate_adaptive_single_smart_background(
                    x_values, y_values, (min_range, max_range), current_background, 0.0, 0.0
                )

                # Initialize Background structure
                if 'Background' not in window.Data['Core levels'][sheet_name]:
                    window.Data['Core levels'][sheet_name]['Background'] = {}

                # Store background data
                window.Data['Core levels'][sheet_name]['Background']['Bkg Y'] = smart_background.tolist()
                window.Data['Core levels'][sheet_name]['Background']['Bkg Low'] = min_range
                window.Data['Core levels'][sheet_name]['Background']['Bkg High'] = max_range

                # Store region with offset 0
                window.Data['Core levels'][sheet_name]['Background']['Recorded_Ranges'] = [
                    (0.00, 0.00, min_range, max_range)
                ]

                print(f"    Smart background created successfully")

            except Exception as e:
                print(f"    Error creating smart background: {e}")
                # Fallback: store raw data as background
                window.Data['Core levels'][sheet_name]['Background'] = {
                    'Bkg Y': y_values.tolist(),
                    'Bkg Low': min_range,
                    'Bkg High': max_range,
                    'Recorded_Ranges': [(0.00, 0.00, min_range, max_range)]
                }

    # Now save the complete JSON with spectral data, peak fitting, and background
    json_file_path = os.path.splitext(file_path)[0] + '.json'
    json_data = convert_to_serializable_and_round(window.Data)
    with open(json_file_path, 'w') as f:
        json.dump(json_data, f, indent=2)
    print(f"Saved complete data with peak fitting and background to {json_file_path}")

    # Refresh the current sheet to load peak fitting into grid
    current_sheet = window.sheet_combobox.GetValue()
    if current_sheet:
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(window, current_sheet)

def add_peak_fitting_data_after_load_OLD(window, peak_fitting_by_scan, file_path):
    """Add peak fitting data after file is fully loaded"""
    if not peak_fitting_by_scan:
        return

    import json
    import re
    from libraries.FileMenu.Save import convert_to_serializable_and_round

    print(f"Adding peak fitting data for {len(peak_fitting_by_scan)} scans to window.Data")

    # Add peak fitting to each matching sheet in window.Data
    for sheet_name in window.Data['Core levels'].keys():
        # Extract base name (e.g., "O1s" from "O1s0")
        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', sheet_name)
        base_name = match.group(1) if match else sheet_name

        if base_name in peak_fitting_by_scan:
            print(f"  Adding peaks to {sheet_name}")

            # Initialize Fitting structure if not exists
            if 'Fitting' not in window.Data['Core levels'][sheet_name]:
                window.Data['Core levels'][sheet_name]['Fitting'] = {}

            window.Data['Core levels'][sheet_name]['Fitting']['Peaks'] = {}
            window.Data['Core levels'][sheet_name]['Fitting']['Fitting Model'] = 'SGL (Area)'

            for peak in peak_fitting_by_scan[base_name]:
                peak_name = peak['name']

                # Build constraints
                position_const = peak.get('position_constraint')
                if position_const:
                    pos_constraint = position_const
                else:
                    pos_constraint = f"{peak['position'] - 2:.2f}:{peak['position'] + 2:.2f}"

                fwhm_const = peak.get('fwhm_constraint')
                if fwhm_const:
                    fwhm_constraint = fwhm_const
                else:
                    fwhm_constraint = "0.30:3.50"

                lg_const = peak.get('lg_constraint')
                if lg_const:
                    lg_constraint = lg_const
                else:
                    lg_constraint = "0.00:100.00"

                window.Data['Core levels'][sheet_name]['Fitting']['Peaks'][peak_name] = {
                    'Position': float(f"{peak['position']:.2f}"),
                    'Height': float(f"{peak['height']:.2f}"),
                    'FWHM': float(f"{peak['fwhm']:.2f}"),
                    'L/G': float(f"{peak['lg_ratio']:.2f}"),
                    'Area': float(f"{peak['area']:.2f}"),
                    'Sigma': 0.00,
                    'Gamma': 0.00,
                    'Skew': 0.00,
                    'Fitting Model': 'SGL (Area)',
                    'Constraints': {
                        'Position': pos_constraint,
                        'Height': f"{max(0, peak['height'] * 0.5):.2f}:{peak['height'] * 1.5:.2f}",
                        'FWHM': fwhm_constraint,
                        'L/G': lg_constraint,
                        'Area': f"{max(0, peak['area'] * 0.5):.2f}:{peak['area'] * 1.5:.2f}",
                        'Sigma': "0.10:1.00",
                        'Gamma': "0.10:1.00",
                        'Skew': "0.01:2.00"
                    }
                }

    # Now save the complete JSON with both spectral data and peak fitting
    json_file_path = os.path.splitext(file_path)[0] + '.json'
    json_data = convert_to_serializable_and_round(window.Data)
    with open(json_file_path, 'w') as f:
        json.dump(json_data, f, indent=2)
    print(f"Saved complete data with peak fitting to {json_file_path}")

    # Refresh the current sheet to load peak fitting into grid
    current_sheet = window.sheet_combobox.GetValue()
    if current_sheet:
        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(window, current_sheet)

def import_avantage_file_direct_xls(window, file_path):
    import xlrd
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb_xls = xlrd.open_workbook(file_path)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    # Group sheets by sample based on letter suffix
    sample_groups = {}
    for sheet_name in wb_xls.sheet_names():
        if "Survey" in sheet_name or "Scan" in sheet_name or any(term in sheet_name.lower() for term in
                                                                 ['valence', 'valence band', 'valence scan', 'vb scan',
                                                                  'vb', 'valence band scan', 'fermi', 'fermi scan',
                                                                  'cut-off', 'cutoff', 'cut off']):
            # Detect sample suffix (single letter at the end after space)
            parts = sheet_name.split()
            if len(parts) >= 2 and len(parts[-1]) == 1 and parts[-1].isalpha():
                suffix = parts[-1]
            else:
                suffix = None  # No suffix = first sample

            if suffix not in sample_groups:
                sample_groups[suffix] = []
            sample_groups[suffix].append(sheet_name)

    # Sort samples: None (no suffix) first, then alphabetically
    sorted_samples = sorted(sample_groups.keys(), key=lambda x: (x is not None, x))

    # Process each sample group
    sample_counter = 0
    for sample_suffix in sorted_samples:
        sheets_to_process = sample_groups[sample_suffix]

        # Extract Title sheet information for this sample
        title_info = extract_title_sheet_info_xls(wb_xls, sample_suffix)

        # Track VB, Fermi and Cut-Off counts for numbering within each sample
        vb_count = 0
        fermi_count = 0
        cutoff_count = 0

        for sheet_name in sheets_to_process:
            sheet = wb_xls.sheet_by_name(sheet_name)

            # Create temporary openpyxl sheet for parameter extraction
            temp_wb = openpyxl.Workbook()
            temp_sheet = temp_wb.active

            # Copy all relevant cells for parameter extraction (first 25 rows, first 20 columns)
            for row in range(min(25, sheet.nrows)):
                for col in range(min(20, sheet.ncols)):
                    cell_value = sheet.cell_value(row, col)
                    if cell_value is not None and str(cell_value).strip():
                        temp_sheet.cell(row=row + 1, column=col + 1, value=cell_value)

            # Extract acquisition parameters
            acquisition_params = extract_acquisition_parameters(temp_sheet)

            lower_name = sheet_name.lower()

            # Handle Survey sheets
            if "Survey" in sheet_name or "survey" in sheet_name:
                number_match = re.search(r'\((\d+)\)', sheet_name)
                if number_match:
                    number = number_match.group(1)
                    if sample_counter == 0:
                        new_name = f"Survey{number}"
                    else:
                        new_name = f"Survey{sample_counter}_{number}"
                else:
                    if sample_counter == 0:
                        new_name = "Survey"
                    else:
                        new_name = f"Survey{sample_counter}"
            # Handle Valence Band sheets
            elif any(term in lower_name for term in
                     ['valence', 'valence band', 'valence scan', 'vb scan', 'vb', 'valence band scan']):
                if vb_count == 0:
                    base_name = "VB"
                else:
                    base_name = f"VB{vb_count + 1}"
                vb_count += 1

                if sample_counter == 0:
                    new_name = base_name
                else:
                    new_name = f"{base_name}{sample_counter}"
            # Handle Fermi sheets
            elif any(term in lower_name for term in ['fermi', 'fermi scan']):
                if fermi_count == 0:
                    base_name = "Fermi"
                else:
                    base_name = f"Fermi{fermi_count + 1}"
                fermi_count += 1

                if sample_counter == 0:
                    new_name = base_name
                else:
                    new_name = f"{base_name}{sample_counter}"
            # Handle Cut-Off sheets
            elif any(term in lower_name for term in ['cut-off', 'cutoff', 'cut off']):
                if cutoff_count == 0:
                    base_name = "Cut-Off"
                else:
                    base_name = f"Cut-Off{cutoff_count + 1}"
                cutoff_count += 1

                if sample_counter == 0:
                    new_name = base_name
                else:
                    new_name = f"{base_name}{sample_counter}"
            else:
                # Extract element name and number for patterns like "C1s Scan (1)"
                parts = sheet_name.split()
                element = parts[0]
                number_match = re.search(r'\((\d+)\)', sheet_name)
                if number_match:
                    number = number_match.group(1)
                    if sample_counter == 0:
                        new_name = f"{element}{number}"
                    else:
                        new_name = f"{element}{sample_counter}_{number}"
                else:
                    if sample_counter == 0:
                        new_name = element
                    else:
                        new_name = f"{element}{sample_counter}"

            wb_new.create_sheet(new_name)
            new_sheet = wb_new[new_name]
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"

            start_row = 17
            for row_idx in range(sheet.nrows):
                if sheet.cell_value(row_idx, 0) == "eV":
                    start_row = row_idx + 1
                    break

            for row_idx in range(start_row, sheet.nrows):
                row_values = sheet.row_values(row_idx)

                if len(row_values) < 3:  # Need at least BE and intensity columns
                    continue

                be_value = row_values[0]
                intensity_value = row_values[2]  # Column C (index 2)

                # Simple validation - check if values can be converted to float
                try:
                    be_val = float(be_value)
                    raw_val = float(intensity_value)
                except (ValueError, TypeError):
                    print("Skipping invalid data row (simple float conversion):", be_value, intensity_value)
                    continue

                # Check for xlrd-specific empty cell types
                if (be_value == xlrd.empty_cell.value or
                        intensity_value == xlrd.empty_cell.value):
                    continue

                new_sheet.append([f"{row_values[0]:.2f}", f"{row_values[2]:.2f}"])

            for col in new_sheet.iter_cols(min_col=3, max_col=24):
                for cell in col:
                    cell.value = None

            # Add experimental description with acquisition parameters (same as xlsx version)
            if acquisition_params:
                exp_col = 45  # Column AS
                new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                current_row = 2

                # Add Title sheet information first
                for param_name, param_value in title_info.items():
                    new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                    new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                    current_row += 1

                # Add acquisition parameters
                for param_name, param_value in acquisition_params.items():
                    new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                    new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                    current_row += 1

                # Set column widths
                new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        sample_counter += 1

    new_file_path = os.path.splitext(file_path)[0] + "_Kfitting.xlsx"
    wb_new.save(new_file_path)
    open_xlsx_file(window, new_file_path)

def import_multiple_avantage_files(window):
    """
    Import multiple Avantage files from a folder in alphabetical order.
    Creates a single Excel file with numbered core levels (C1s0, O1s0, C1s1, O1s1...).
    Filenames are stored in SampleNames section of window.data JSON.
    """
    with wx.DirDialog(window, "Choose a directory containing Avantage files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all Avantage files in the directory
        avantage_files = [f for f in os.listdir(folder_path)
                          if f.lower().endswith(('.xlsx', '.xls'))]

        if not avantage_files:
            window.show_popup_message2("Information", "No Avantage files found in the selected folder.")
            return

        # Sort files alphabetically
        avantage_files.sort()

        # Create single combined workbook
        combined_file_path = os.path.join(folder_path, "Combined_Samples_Kfitting.xlsx")
        combined_wb = openpyxl.Workbook()
        combined_wb.remove(combined_wb.active)

        # Store sample names as dictionary for JSON
        sample_names_dict = {}

        # Process each file as a separate sample
        for sample_idx, avantage_file in enumerate(avantage_files):
            file_path = os.path.join(folder_path, avantage_file)

            # Remove file extension for cleaner sample names
            sample_name = os.path.splitext(avantage_file)[0]
            sample_names_dict[str(sample_idx)] = sample_name

            # Load the workbook
            if avantage_file.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path)
                process_avantage_xlsx_with_sample_number(wb, combined_wb, sample_idx)
            else:
                wb = xlrd.open_workbook(file_path)
                process_avantage_xls_with_sample_number(wb, combined_wb, sample_idx)

        # Save the combined file
        combined_wb.save(combined_file_path)

        window.show_popup_message2("Success", f"Combined {len(avantage_files)} Avantage files into single Excel file.")

        # Open the combined file
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, combined_file_path)

        # Update SampleNames in window.Data
        if not hasattr(window, 'Data') or window.Data is None:
            from libraries.ConfigFile import Init_Measurement_Data
            window.Data = Init_Measurement_Data(window)

        # Add SampleNames as dictionary to the data structure
        window.Data['SampleNames'] = sample_names_dict

        # Save the updated JSON file
        json_file_path = os.path.splitext(combined_file_path)[0] + '.json'
        from libraries.FileMenu.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

    except Exception as e:
        window.show_popup_message2("Error", f"Error processing Avantage files: {str(e)}")


def process_avantage_xlsx_with_sample_number(wb, combined_wb, sample_idx):
    """Process xlsx Avantage file and add numbered core levels to combined workbook"""
    from openpyxl.utils import get_column_letter  # ADDED

    sheets_to_process = []
    for sheet_name in wb.sheetnames:
        if "Survey" in sheet_name or "Scan" in sheet_name:
            sheets_to_process.append(sheet_name)

    for sheet_name in sheets_to_process:
        sheet = wb[sheet_name]

        # ADDED: Extract acquisition parameters
        acquisition_params = extract_acquisition_parameters(sheet)

        # Extract peak fitting data from Peak Table sheet
        peak_fitting_by_scan = extract_peak_fitting_from_peak_table(wb)
        print(f"Extracted peak fitting data for {len(peak_fitting_by_scan)} scans")

        # Extract element name
        if "Survey" in sheet_name or "survey" in sheet_name:
            base_name = "Survey"
        else:
            parts = sheet_name.split()
            base_name = parts[0]  # Get element (C1s, O1s, etc.)

        # Check if multi-sample (D8 cell not empty and > 1)
        multi_sample = False
        num_samples = 1
        if sheet.cell(row=8, column=4).value is not None:
            try:
                num_samples = int(sheet.cell(row=8, column=4).value)
                if num_samples > 1:
                    multi_sample = True
            except (ValueError, TypeError):
                pass

        # Find data start row
        start_row = 19
        for row_idx in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == "eV":
                start_row = row_idx + 1
                break

        if multi_sample:
            # Process each sample within the file
            for sub_sample_idx in range(num_samples):
                new_sheet_name = f"{base_name}{sample_idx}"
                if sub_sample_idx > 0:
                    new_sheet_name = f"{base_name}{sample_idx}_{sub_sample_idx}"

                new_sheet = combined_wb.create_sheet(new_sheet_name)
                new_sheet['A1'] = "Binding Energy"
                new_sheet['B1'] = "Raw Data"

                # Data is in column C+sub_sample_idx
                data_col = 3 + sub_sample_idx
                copy_sheet_data(sheet, new_sheet, start_row, 1, data_col, 2)

                # ADDED: Add experimental description with acquisition parameters
                if acquisition_params:
                    exp_col = 45  # Column AS
                    new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                    current_row = 2
                    for param_name, param_value in acquisition_params.items():
                        new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                        new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                        current_row += 1

                    # Set column widths
                    new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                    new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        else:
            # Single sample - create sheet with sample number
            new_sheet_name = f"{base_name}{sample_idx}"
            new_sheet = combined_wb.create_sheet(new_sheet_name)
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"

            # Copy data from columns A (BE) and C (intensity)
            copy_sheet_data(sheet, new_sheet, start_row, 1, 3, 2)

            # ADDED: Add experimental description with acquisition parameters
            if acquisition_params:
                exp_col = 45  # Column AS
                new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                current_row = 2
                for param_name, param_value in acquisition_params.items():
                    new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                    new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                    current_row += 1

                # Set column widths
                new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40


def process_avantage_xls_with_sample_number(wb_xls, combined_wb, sample_idx):
    """Process xls Avantage file and add numbered core levels to combined workbook"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    for sheet_name in wb_xls.sheet_names():
        if "Survey" in sheet_name or "Scan" in sheet_name:
            sheet = wb_xls.sheet_by_name(sheet_name)

            # Create temporary openpyxl sheet for parameter extraction
            temp_wb = openpyxl.Workbook()
            temp_sheet = temp_wb.active

            # Copy all relevant cells for parameter extraction (first 25 rows, first 20 columns)
            for row in range(min(25, sheet.nrows)):
                for col in range(min(20, sheet.ncols)):
                    cell_value = sheet.cell_value(row, col)
                    if cell_value is not None and str(cell_value).strip():
                        temp_sheet.cell(row=row + 1, column=col + 1, value=cell_value)

            # Extract acquisition parameters
            acquisition_params = extract_acquisition_parameters(temp_sheet)

            # Extract element name
            if "Survey" in sheet_name or "survey" in sheet_name:
                base_name = "Survey"
            else:
                parts = sheet_name.split()
                base_name = parts[0]  # Get element (C1s, O1s, etc.)

            # Create sheet with sample number
            new_sheet_name = f"{base_name}{sample_idx}"
            new_sheet = combined_wb.create_sheet(new_sheet_name)
            new_sheet['A1'] = "Binding Energy"
            new_sheet['B1'] = "Raw Data"

            # Copy data starting from row 19 (18 in 0-indexed)
            start_row = 18
            row_new = 2
            for row in range(start_row, sheet.nrows):
                be_value = sheet.cell_value(row, 0)  # Column A
                intensity_value = sheet.cell_value(row, 2)  # Column C

                if be_value and intensity_value:
                    new_sheet.cell(row=row_new, column=1, value=f"{be_value:.2f}")
                    new_sheet.cell(row=row_new, column=2, value=f"{intensity_value:.2f}")
                    row_new += 1

            # Add experimental description with acquisition parameters
            if acquisition_params:
                exp_col = 45  # Column AS
                new_sheet.cell(row=1, column=exp_col, value="Experimental Description")

                current_row = 2
                for param_name, param_value in acquisition_params.items():
                    new_sheet.cell(row=current_row, column=exp_col, value=param_name)
                    new_sheet.cell(row=current_row, column=exp_col + 1, value=param_value)
                    current_row += 1

                # Set column widths
                new_sheet.column_dimensions[get_column_letter(exp_col)].width = 25
                new_sheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40


def extract_title_sheet_info(wb, sample_suffix=None):
    """Extract information from Title sheet (or Title A, Title B, etc.)"""
    title_info = {}

    # Determine the correct Title sheet name based on sample suffix
    if sample_suffix is None:
        title_sheet_name = "Titles"
    else:
        title_sheet_name = f"Titles {sample_suffix}"

    # Try to find the Title sheet
    if title_sheet_name in wb.sheetnames:
        title_sheet = wb[title_sheet_name]

        # Extract values from A1, A3, A5
        cell_a1 = title_sheet.cell(row=1, column=1).value
        cell_a3 = title_sheet.cell(row=3, column=1).value
        cell_a5 = title_sheet.cell(row=5, column=1).value

        if cell_a1:
            title_info['Sample Info'] = str(cell_a1).strip()
        if cell_a3:
            title_info['Additional Info'] = str(cell_a3).strip()
        if cell_a5:
            title_info['Notes'] = str(cell_a5).strip()

    return title_info


def extract_title_sheet_info_xls(wb_xls, sample_suffix=None):
    """Extract information from Title sheet for .xls files"""
    title_info = {}

    # Determine the correct Title sheet name based on sample suffix
    if sample_suffix is None:
        title_sheet_name = "Titles"
    else:
        title_sheet_name = f"Titles {sample_suffix}"

    # Try to find the Title sheet
    if title_sheet_name in wb_xls.sheet_names():
        title_sheet = wb_xls.sheet_by_name(title_sheet_name)

        # Extract values from A1, A3, A5 (0-indexed: row 0, 2, 4; column 0)
        try:
            cell_a1 = title_sheet.cell_value(0, 0)
            if cell_a1:
                title_info['Sample Info'] = str(cell_a1).strip()
        except:
            pass

        try:
            cell_a3 = title_sheet.cell_value(2, 0)
            if cell_a3:
                title_info['Additional Info'] = str(cell_a3).strip()
        except:
            pass

        try:
            cell_a5 = title_sheet.cell_value(4, 0)
            if cell_a5:
                title_info['Notes'] = str(cell_a5).strip()
        except:
            pass

    return title_info

def copy_sheet_data(source_sheet, target_sheet, start_row, source_be_col, source_intensity_col, target_row_start):
    """Helper function to copy data from source sheet to target sheet"""
    row_new = target_row_start
    for row in range(start_row, source_sheet.max_row + 1):
        be_value = source_sheet.cell(row=row, column=source_be_col).value
        intensity_value = source_sheet.cell(row=row, column=source_intensity_col).value

        if be_value is None or intensity_value is None:
            continue

        target_sheet.cell(row=row_new, column=1, value=be_value)
        target_sheet.cell(row=row_new, column=2, value=intensity_value)
        row_new += 1



def open_avg_file_direct(window, avg_file_path):

    # Get the basename without extension
    raw_sheet_name = os.path.basename(avg_file_path).split('.')[0]

    # Normalize the sheet name to follow KherveFitting conventions
    sheet_name = normalize_sheet_name(raw_sheet_name)

    photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)
    be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

    df = pd.DataFrame({
        'BE': be_values,
        'Intensity': y_values[:num_points]
    })

    output_path = avg_file_path.rsplit('.', 1)[0] + '.xlsx'

    # Extract metadata
    metadata = extract_metadata_from_avg(avg_file_path)

    # Fields in the order shown in the example
    field_order = [
        'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
        'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
        'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
        'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
        'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
        'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
        'Collection Time', 'Time Correction', 'Y Unit'
    ]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Use the normalized sheet name instead of the raw file name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add metadata to the main sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Add experimental description at column 50
        exp_col = 50
        worksheet.cell(row=1, column=exp_col, value="Experimental Description")

        for i, field in enumerate(field_order, start=2):
            worksheet.cell(row=i, column=exp_col, value=field)
            worksheet.cell(row=i, column=exp_col + 1, value=metadata.get(field, "N/A"))

        # Set column widths
        from openpyxl.utils import get_column_letter
        worksheet.column_dimensions[get_column_letter(exp_col)].width = 25
        worksheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        # Create separate experimental description sheet
        exp_sheet = workbook.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        exp_sheet['A1'] = "Experimental Description"

        for row, field in enumerate(field_order, start=2):
            exp_sheet[f'A{row}'] = field
            exp_sheet[f'B{row}'] = metadata.get(field, "N/A")

    return output_path

def import_avantage_file(window):
    # Opens a file dialog for user to select file
    with wx.FileDialog(window, "Open Avantage Excel file", wildcard="Excel files (*.xlsx;*.xls)|*.xlsx;*.xls",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

        # Call appropriate function based on extension
        if file_path.lower().endswith('.xlsx'):
            import_avantage_file_direct(window, file_path)
        elif file_path.lower().endswith('.xls'):
            import_avantage_file_direct_xls(window, file_path)


def parse_avg_file_OLD(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    photon_energy = float(re.search(r'DS_SOPROPID_ENERGY\s+:\s+VT_R4\s+=\s+(\d+\.\d+)', content).group(1))
    start_energy, width, num_points = map(float, re.search(r'\$SPACEAXES=1\s+0=\s+(\d+\.\d+),\s+(\d+\.\d+),\s+(\d+),',
                                                           content).groups())

    # Extract acquisition time and periods for intensity correction
    acq_time_match = re.search(r'DS_ACPROPID_ACQ_TIME\s+:\s+VT_R4\s+=\s+(\d+\.\d+)', content)
    periods_match = re.search(r'DS_ACPROPID_PERIODS\s+:\s+VT_I4\s+=\s+(\d+)', content)

    acq_time = float(acq_time_match.group(1)) if acq_time_match else 0.05
    periods = int(periods_match.group(1)) if periods_match else 1

    # Extract all intensity values using a different approach
    y_values = []
    in_list_section = False

    for line in content.split('\n'):
        if line.strip().startswith('LIST@'):
            in_list_section = True
            values_part = line.split('=', 1)[1].strip()
            for val in values_part.split(','):
                if val.strip():
                    try:
                        y_values.append(float(val.strip()))
                    except ValueError:
                        pass
        elif in_list_section:
            if ',' in line and not line.strip().startswith('$') and not line.strip().startswith('DS_'):
                for val in line.split(','):
                    if val.strip():
                        try:
                            y_values.append(float(val.strip()))
                        except ValueError:
                            pass
            else:
                in_list_section = False

    # Ensure we have the correct number of points
    if len(y_values) > int(num_points):
        y_values = y_values[:int(num_points)]

    if len(y_values) < int(num_points):
        print(f"Warning: Expected {int(num_points)} points but found only {len(y_values)}.")

    # Apply intensity correction: raw_intensity / (acq_time * periods)
    correction_factor = acq_time * periods
    if correction_factor > 0:
        y_values = [y / correction_factor for y in y_values]
        print(f"Applied intensity correction: divided by ({acq_time} * {periods}) = {correction_factor}")

    return photon_energy, start_energy, width, int(num_points), y_values


def parse_avg_file(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    # Check for photon energy with error handling
    energy_match = re.search(r'DS_SOPROPID_ENERGY\s+:\s+VT_R4\s+=\s+(\d+\.?\d*)', content)
    if energy_match:
        photon_energy = float(energy_match.group(1))
    else:
        print(f"Warning: Could not find photon energy in {file_path}, using default UPS energy 21.22 eV")
        photon_energy = 21.22  # Default HeI energy for UPS

    # Try multiple patterns for space axes information
    start_energy = None
    width = None
    num_points = None

    # Pattern 1: XPS format - $SPACEAXES=1 0= start, width, points,
    axes_match = re.search(r'\$SPACEAXES=1\s+0=\s+(\d+\.?\d*),\s+(\d+\.?\d*),\s+(\d+)', content)
    if axes_match:
        start_energy, width, num_points = map(float, axes_match.groups())
        print(f"Found XPS format: start={start_energy}, width={width}, points={int(num_points)}")

    # Pattern 2: UPS format - $SPACEAXES=3 with axis 0 being energy
    if not axes_match:
        # Look for the energy axis (axis 0) in multi-axis format
        axes_match = re.search(r'\$SPACEAXES=3.*?0=\s+(\d+\.?\d*),\s+(\d+\.?\d*),\s+(\d+),\s+ENERGY', content, re.DOTALL)
        if axes_match:
            start_energy, width, num_points = map(float, axes_match.groups())
            print(f"Found UPS format: start={start_energy}, width={width}, points={int(num_points)}")

    # Pattern 3: General multi-axis format
    if not axes_match:
        axes_match = re.search(r'\$SPACEAXES=\d+.*?0=\s+(\d+\.?\d*),\s+(\d+\.?\d*),\s+(\d+)', content, re.DOTALL)
        if axes_match:
            start_energy, width, num_points = map(float, axes_match.groups())
            print(f"Found general format: start={start_energy}, width={width}, points={int(num_points)}")

    if start_energy is None:
        # Debug output
        print(f"DEBUG: Could not find energy range in {file_path}")
        spaceaxes_section = re.search(r'\$SPACEAXES=.*?(?=\$|\Z)', content, re.DOTALL)
        if spaceaxes_section:
            print("SPACEAXES section found:")
            print(spaceaxes_section.group(0))
        raise ValueError(f"Could not find energy range information in {file_path}")

    # Extract acquisition time and periods for intensity correction
    acq_time_match = re.search(r'DS_ACPROPID_ACQ_TIME\s+:\s+VT_R4\s+=\s+(\d+\.?\d*)', content)
    periods_match = re.search(r'DS_ACPROPID_PERIODS\s+:\s+VT_I4\s+=\s+(\d+)', content)

    acq_time = float(acq_time_match.group(1)) if acq_time_match else 0.05
    periods = int(periods_match.group(1)) if periods_match else 1

    # Extract all intensity values
    y_values = []
    in_list_section = False

    for line in content.split('\n'):
        if line.strip().startswith('LIST@'):
            in_list_section = True
            values_part = line.split('=', 1)[1].strip()
            for val in values_part.split(','):
                if val.strip():
                    try:
                        y_values.append(float(val.strip()))
                    except ValueError:
                        pass
        elif in_list_section:
            if ',' in line and not line.strip().startswith('$') and not line.strip().startswith('DS_'):
                for val in line.split(','):
                    if val.strip():
                        try:
                            y_values.append(float(val.strip()))
                        except ValueError:
                            pass
            else:
                in_list_section = False

    # Ensure we have the correct number of points
    if len(y_values) > int(num_points):
        y_values = y_values[:int(num_points)]

    if len(y_values) < int(num_points):
        print(f"Warning: Expected {int(num_points)} points but found only {len(y_values)}.")

    # Apply intensity correction: raw_intensity / (acq_time * periods)
    correction_factor = acq_time * periods
    if correction_factor > 0:
        y_values = [y / correction_factor for y in y_values]
        print(f"Applied intensity correction: divided by ({acq_time} * {periods}) = {correction_factor}")

    return photon_energy, start_energy, width, int(num_points), y_values

def create_excel_from_avg(avg_file_path):
    from openpyxl.utils import get_column_letter

    # Extract the raw sheet name from the file name
    raw_sheet_name = os.path.basename(avg_file_path).split()[0]

    # Normalize the sheet name to follow KherveFitting conventions
    sheet_name = normalize_sheet_name(raw_sheet_name)

    photon_energy, start_energy, width, num_points, y_values = parse_avg_file(avg_file_path)

    be_values = [photon_energy - (start_energy + i * width) for i in range(num_points)]

    df = pd.DataFrame({
        'BE': be_values,
        'Intensity': y_values[:num_points]
    })

    output_path = avg_file_path.rsplit('.', 1)[0] + '.xlsx'

    # Extract metadata
    metadata = extract_metadata_from_avg(avg_file_path)

    # Fields in the order shown in the example
    field_order = [
        'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
        'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
        'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
        'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
        'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
        'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
        'Collection Time', 'Time Correction', 'Y Unit'
    ]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Use normalized sheet name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add metadata to the main sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Add experimental description at column 50
        exp_col = 50
        worksheet.cell(row=1, column=exp_col, value="Experimental Description")

        for i, field in enumerate(field_order, start=2):
            worksheet.cell(row=i, column=exp_col, value=field)
            worksheet.cell(row=i, column=exp_col + 1, value=metadata.get(field, "N/A"))

        # Set column widths
        worksheet.column_dimensions[get_column_letter(exp_col)].width = 25
        worksheet.column_dimensions[get_column_letter(exp_col + 1)].width = 40

        # Create separate experimental description sheet
        exp_sheet = workbook.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        exp_sheet['A1'] = "Experimental Description"

        for row, field in enumerate(field_order, start=2):
            exp_sheet[f'A{row}'] = field
            exp_sheet[f'B{row}'] = metadata.get(field, "N/A")

    return output_path


def extract_metadata_from_avg(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    metadata = {
        'Sample ID': "Group",
        'Date': "N/A",
        'Time': "N/A",
        'Technique': "XPS",
        'Species & Transition': "C 1s",
        'Number of scans': "20",
        'Source Label': "KA1486",
        'Source Energy': "1486.68",
        'Source width X': "1E+37",
        'Source width Y': "1E+37",
        'Pass Energy': "20",
        'Work Function': "1E+37",
        'Analyzer Mode': "FAT",
        'Sputtering Energy': "N/A",
        'Take-off Polar Angle': "1E+37",
        'Take-off Azimuth': "1E+37",
        'Target Bias': "1E+37",
        'Analysis Width X': "1E+37",
        'Analysis Width Y': "1E+37",
        'X Label': "Kinetic Energy",
        'X Units': "eV",
        'X Start': "1188.6",
        'X Step': "0.1",
        'Num Y Values': "382",
        'Num Scans': "20",
        'Collection Time': "0.05",
        'Time Correction': "1E+37",
        'Y Unit': "d"
    }

    # Extract acquisition time and periods
    acq_time_match = re.search(r'DS_ACPROPID_ACQ_TIME\s+:\s+VT_R4\s+=\s+(\d+\.\d+)', content)
    periods_match = re.search(r'DS_ACPROPID_PERIODS\s+:\s+VT_I4\s+=\s+(\d+)', content)

    if acq_time_match:
        metadata['Collection Time'] = acq_time_match.group(1)
    if periods_match:
        metadata['Num Scans'] = periods_match.group(1)
        metadata['Number of scans'] = periods_match.group(1)

    # Extract other values from file where available
    created_match = re.search(r'DS_EXT_SUPROPID_CREATED\s+:\s+VT_DATE\s+=\s+(\d+)/(\d+)/(\d+)\s+(\d+):(\d+):(\d+)',
                              content)
    if created_match:
        day, month, year = created_match.group(1), created_match.group(2), created_match.group(3)
        hour, minute, second = created_match.group(4), created_match.group(5), created_match.group(6)
        metadata['Date'] = f"{year}/{month}/{day}"
        metadata['Time'] = f"{hour}:{minute}:{second}"

    return metadata


def open_avg_file(window):
    with wx.FileDialog(window, "Open AVG file", wildcard="AVG files (*.avg)|*.avg",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return

        avg_file_path = fileDialog.GetPath()

    try:
        # Use create_excel_from_avg which we need to modify to normalize sheet names
        excel_file_path = create_excel_from_avg(avg_file_path)

        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, excel_file_path)
    except Exception as e:
        wx.MessageBox(f"Error processing AVG file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def import_multiple_avg_files(window):
    with wx.DirDialog(window, "Choose a directory containing AVG files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        root_folder_path = dirDialog.GetPath()

    try:
        # Function to process a single folder
        def process_folder(folder_path):
            folder_name = os.path.basename(folder_path)
            avg_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.avg')]

            if not avg_files:
                return None  # No AVG files in this folder

            # Create Excel file for this folder
            excel_file_path = os.path.join(folder_path, f"{folder_name}.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Process each AVG file in this folder
            for avg_file in avg_files:
                avg_file_path = os.path.join(folder_path, avg_file)

                # Extract raw sheet name from file name (without extension)
                raw_sheet_name = os.path.splitext(avg_file)[0]

                # Normalize the sheet name according to KherveFitting nomenclature
                sheet_name = normalize_sheet_name(raw_sheet_name)

                # Handle duplicate sheet names if needed
                suffix = 1
                original_sheet_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_sheet_name}{suffix}"
                    suffix += 1

                # Parse AVG file
                try:
                    photon_energy, start_energy, width, expected_points, y_values = parse_avg_file(avg_file_path)

                    # Use the actual number of points we retrieved, not the expected number
                    actual_points = len(y_values)

                    # If we got fewer points than expected, adjust the width of the energy range
                    if actual_points < expected_points:
                        print(f"Warning: {avg_file} - Expected {expected_points} points but found {actual_points}.")
                        # Generate BE values based on the actual number of points we have
                        be_values = [photon_energy - (start_energy + i * width) for i in range(actual_points)]
                    else:
                        be_values = [photon_energy - (start_energy + i * width) for i in range(actual_points)]

                    # Create new sheet
                    ws = wb.create_sheet(title=sheet_name)

                    # Add headers
                    ws.append(["BE", "Intensity"])

                    # Add data - only use the points we actually have
                    for be, intensity in zip(be_values, y_values):
                        ws.append([be, intensity])

                    # Extract metadata if available
                    metadata = extract_metadata_from_avg(avg_file_path)

                    # Add metadata starting at column 50
                    exp_col = 50
                    ws.cell(row=1, column=exp_col, value="Experimental Description")

                    for i, (key, value) in enumerate(metadata.items(), start=2):
                        ws.cell(row=i, column=exp_col, value=key)
                        ws.cell(row=i, column=exp_col + 1, value=value)

                except Exception as e:
                    print(f"Error processing {avg_file}: {str(e)}")
                    continue

            # Save the Excel file if any sheets were created
            if len(wb.sheetnames) > 0:
                wb.save(excel_file_path)
                return excel_file_path
            else:
                return None

        # Process the root folder and all subfolders
        excel_files = []

        # Walk through directory tree
        for dirpath, dirnames, filenames in os.walk(root_folder_path):
            # Process current folder
            excel_path = process_folder(dirpath)
            if excel_path:
                excel_files.append(excel_path)

        # Show results
        if excel_files:
            message = f"Created {len(excel_files)} Excel files:\n" + "\n".join(excel_files)
            wx.MessageBox(message, "Success", wx.OK | wx.ICON_INFORMATION)

            # Open the first Excel file
            from libraries.FileMenu.Open import open_xlsx_file
            open_xlsx_file(window, excel_files[0])
        else:
            wx.MessageBox("No AVG files found in the selected folder or subfolders.",
                          "Information", wx.OK | wx.ICON_INFORMATION)

    except Exception as e:
        wx.MessageBox(f"Error processing AVG files: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def open_xlsx_file(window, file_path=None):
    if file_path is None:
        with wx.FileDialog(window, "Open XLSX file", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
            else:
                return

    # Store file manager state
    file_manager_was_open = False
    file_manager_position = None
    try:
        if hasattr(window, 'file_manager') and window.file_manager is not None and window.file_manager.IsShown():
            file_manager_was_open = True
            file_manager_position = window.file_manager.GetPosition()
            window.file_manager.Close()
            window.file_manager = None
    except RuntimeError:
        window.file_manager = None

    # Close fitting and export windows to refresh with new core levels
    try:
        if hasattr(window, 'fitting_screen') and window.fitting_screen is not None:
            try:
                window.fitting_screen.GetSize()
                window.fitting_screen.Close()
            except RuntimeError:
                pass
            window.fitting_screen = None
    except:
        pass

    try:
        if hasattr(window, 'areafit_screen') and window.areafit_screen is not None:
            try:
                window.areafit_screen.GetSize()
                window.areafit_screen.Close()
            except RuntimeError:
                pass
            window.areafit_screen = None
    except:
        pass

    try:
        if hasattr(window, 'profile_creator_window') and window.profile_creator_window is not None:
            try:
                window.profile_creator_window.GetSize()
                window.profile_creator_window.Close()
            except RuntimeError:
                pass
            window.profile_creator_window = None
    except:
        pass

    try:
        if hasattr(window, 'export_results_window') and window.export_results_window is not None:
            try:
                window.export_results_window.GetSize()
                window.export_results_window.Close()
            except RuntimeError:
                pass
            window.export_results_window = None
    except:
        pass

    try:
        # Create console window centered on parent
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Loading Excel File", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

        update_console("Initializing...")
        window.SetStatusText(f"Selected File: {file_path}", 0)

        # Clear history and results
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)
        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Check for JSON file
        json_file = os.path.splitext(file_path)[0] + '.json'
        json_data_loaded = False
        if os.path.exists(json_file):
            update_console("Found corresponding JSON file - loading saved data...")
            with open(json_file, 'r') as f:
                loaded_data = json.load(f)
            window.Data = convert_from_serializable(loaded_data)
            json_data_loaded = True
            populate_results_grid(window)
        else:
            update_console("Initializing new data structure...")
            window.Data = Init_Measurement_Data(window)

        # Read Excel file
        update_console("Reading Excel file structure...")
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        # Validation
        invalid_sheets = [name for name in sheet_names if name.startswith('Sheet')]
        if invalid_sheets:
            console_frame.Close()
            wx.MessageBox(f"File contains invalid sheet names: {', '.join(invalid_sheets)}", "Invalid Sheet Names",
                          wx.OK | wx.ICON_WARNING)
            return

        for sheet_name in sheet_names:
            # Skip validation for zzProfile sheets - they have different column structure
            if sheet_name.startswith('zzProfile'):
                continue

            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            col1_value = str(df.iloc[0, 0]).strip().upper()
            col2_value = str(df.iloc[0, 1]).strip().upper()

            xps_valid = ('BE' in col1_value or 'BINDING' in col1_value) and \
                        ('RAW DATA' in col2_value or 'CORRECTED DATA' in col2_value or 'INTENSITY' in col2_value)
            raman_valid = ('WAVENUMBER' in col1_value or 'CM-1' in col1_value) and \
                          ('RAW DATA' in col2_value or 'INTENSITY' in col2_value)

            if not (xps_valid or raman_valid):
                console_frame.Close()
                wx.MessageBox(f"Sheet '{sheet_name}' has invalid column labels", "Invalid Column Labels",
                              wx.OK | wx.ICON_WARNING)
                return

        window.Data['FilePath'] = file_path
        update_console(f"Found {len(sheet_names)} sheets to process...")

        # Load core level data
        if not json_data_loaded and ('Core levels' not in window.Data or not window.Data['Core levels']):
            window.Data['Number of Core levels'] = 0
            for i, sheet_name in enumerate(sheet_names, 1):
                update_console(f"Loading sheet {i}/{len(sheet_names)}: {sheet_name}")
                # window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)
                # Special handling for zzProfile sheets
                if sheet_name.startswith('zzProfile'):
                    # Load profile sheet data
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

                    first_data_col = df.columns[1] if len(df.columns) > 1 else 'Number'

                    window.Data['Core levels'][sheet_name] = {
                        'Name': sheet_name,
                        'B.E.': df['Number'].tolist() if 'Number' in df.columns else [],
                        'Raw Data': df[first_data_col].tolist() if first_data_col in df.columns else [],
                        'Profile Data': df.to_dict('list'),
                        'Y_Axis_Label': 'Atomic Concentration (%)',
                        'X_Axis_Label': 'Number',
                        'Profile_Type': 'Atomic_Concentration',
                        'Background': {
                            'Bkg Type': '',
                            'Bkg Low': '',
                            'Bkg High': '',
                            'Bkg Offset Low': '',
                            'Bkg Offset High': '',
                            'Bkg Y': df[first_data_col].tolist() if first_data_col in df.columns else []
                        }
                    }
                else:
                    window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)
        else:
            update_console("Data loaded from JSON file...")
            update_console("Available sheets:")
            for i, sheet_name in enumerate(sheet_names, 1):
                update_console(f"  {i}/{len(sheet_names)}: {sheet_name}")

        update_console("Loading BE corrections...")
        window.load_be_correction()

        update_console("Setting up interface...")
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        first_sheet = sheet_names[0]
        window.sheet_combobox.SetValue(first_sheet)

        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(first_sheet)
        window.plot_config.plot_limits.clear()
        on_sheet_selected(window, event)

        save_state(window)
        update_recent_files(window, file_path)

        if hasattr(window, 'setup_backup_timer'):
            window.setup_backup_timer()


        # Create backup before opening file
        from libraries.Utilities import perform_auto_backup
        update_console("Performing auto backup...")
        perform_auto_backup(window)

        # Refresh Sheet
        update_console("Refreshing sheets...")
        from libraries.FileMenu.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected, update_console)

        update_console("File loaded successfully!")
        wx.CallLater(500, console_frame.Close)

        # Restore file manager
        if file_manager_was_open:
            from libraries.ViewMenu.FileManager import FileManagerWindow
            window.file_manager = FileManagerWindow(window)
            if file_manager_position:
                window.file_manager.SetPosition(file_manager_position)
            window.file_manager.Show()

    except Exception as e:
        wx.MessageBox(f"Error reading file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def convert_from_serializable(obj):
    """
    Recursively converts a serializable object (list or dict) back into its original structure.
    This is used to restore complex data structures that were serialized to JSON.
    """
    if isinstance(obj, list):
        return [convert_from_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_from_serializable(v) for k, v in obj.items()}
    else:
        return obj

def update_recent_files(window, file_path):
    """
    Updates the list of recently opened files, adding the new file to the top and removing duplicates.
    """
    if file_path in window.recent_files:
        window.recent_files.remove(file_path)
    window.recent_files.insert(0, file_path)
    window.recent_files = window.recent_files[:window.max_recent_files]
    update_recent_files_menu(window)
    save_recent_files_to_config(window)

def update_recent_files_menu(window):
    """
    Refreshes the 'Recent Files' menu with the updated list of recently opened files.
    """
    if hasattr(window, 'recent_files_menu'):
        # Remove all existing menu items
        for item in window.recent_files_menu.GetMenuItems():
            window.recent_files_menu.DestroyItem(item)

        # Add new menu items for recent files
        for i, file_path in enumerate(window.recent_files):
            item = window.recent_files_menu.Append(wx.ID_ANY, os.path.basename(file_path))
            window.Bind(wx.EVT_MENU, lambda evt, fp=file_path: open_xlsx_file(window, fp), item)

def save_recent_files_to_config(window):
    """
    Saves the updated list of recently opened files to the application's configuration.
    """
    window.recent_files = window.recent_files[:window.max_recent_files]
    window.save_config()


def normalize_sheet_name(name):
    """Normalize core level name to standard format."""
    import re
    new_name = name

    lower_name = name.lower()
    if 'xps survey' in lower_name:
        new_name = 'Survey'
    elif 'survey' in lower_name:
        new_name = 'Survey'
    elif 'survey scan' in lower_name:
        new_name = 'Survey'
    elif 'wide' in lower_name:
        new_name = 'Wide'
    elif 'wide scan' in lower_name:
        new_name = 'Wide'
    elif 'su1s' in lower_name or '_su' in lower_name or name.lower().endswith('_su'):
        new_name = 'Survey'
    elif any(term in lower_name for term in ['valence', 'valence band', 'valence scan', 'vb scan', 'vb', 'valence band scan']):
        new_name = 'VB'
    elif any(term in lower_name for term in ['fermi', 'fermi scan']):
        new_name = 'Fermi'
    else:
        # Remove spaces between element and orbital (e.g., "C 1s" → "C1s")
        match = re.search(r'([A-Z][a-z]?)\s+(\d+[spdf])', name)
        if match:
            element, orbital = match.groups()
            new_name = f"{element}{orbital}"

        # Simplify names like "C1s Scan" to just "C1s"
        match = re.search(r'([A-Z][a-z]?\d+[spdf])', new_name)
        if match and len(new_name) > len(match.group(1)):
            new_name = match.group(1)

    # Preserve sample number suffix if it exists
    suffix_match = re.search(r'(\d+)$', name)
    if suffix_match and not re.search(r'\d+$', new_name):
        new_name = f"{new_name}{suffix_match.group(1)}"

    return new_name

def open_vamas_file_dialog(window):
    """
    Open a file dialog for selecting a VAMAS file and process it.

    This function displays a file dialog for the user to select a VAMAS file,
    then calls open_vamas_file to process the selected file.

    Args:
    window: The main application window object.
    """
    with wx.FileDialog(window, "Open VAMAS file", wildcard="VAMAS files (*.vms)|*.vms",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_vamas_file(window, file_path)

def extract_transmission_reference(block):
    """Extract transmission reference value from VAMAS block"""
    try:
        # Check if the block has the transmission reference in its properties
        # This might be in block attributes or in the block structure
        if hasattr(block, 'additional_parameters'):
            # Look for transmission reference in additional parameters
            for param in block.additional_parameters:
                if 'transmission' in str(param).lower():
                    # Extract the number after transmission
                    pass

        # Alternative: parse from string representation or other block properties
        # Since VAMAS structure can vary, we might need to check different attributes
        if hasattr(block, 'block_identifier') or hasattr(block, 'comment'):
            # Parse block information for transmission reference
            pass

        # For now, return None if we can't find it
        return None

    except Exception as e:
        print(f"Error extracting transmission reference: {e}")
        return None


def normalize_auger_and_valence_names(sheet_name):
    """
    Normalize Auger transition names and valence band variations.

    Examples:
    - Fe LM2 -> Felmm
    - C KL1 -> Ckll
    - V.B. -> VB
    - V.B -> VB
    """
    import re

    # Handle valence band variations first
    vb_pattern = r'^V\.?B\.?\s*$'
    if re.match(vb_pattern, sheet_name.strip(), re.IGNORECASE):
        return "VB"

    # Handle Auger transitions
    # Pattern: Element symbol + space + Auger notation (letters + numbers)
    auger_pattern = r'^([A-Z][a-z]?)\s+([A-Z]+\d*)$'
    match = re.match(auger_pattern, sheet_name.strip())

    if match:
        element = match.group(1)  # e.g., "Fe", "C"
        auger_part = match.group(2)  # e.g., "LM2", "KL1"

        # Remove numbers and convert letters to lowercase
        auger_letters = re.sub(r'\d+', '', auger_part).lower()

        # Duplicate letters based on the original notation
        # For LM2 -> lm -> lmm (duplicate the last letter)
        # For KL1 -> kl -> kll (duplicate the last letter)
        if len(auger_letters) >= 2:
            # Add extra letter (typically duplicate the last one)
            auger_letters += auger_letters[-1]

        return f"{element}{auger_letters}"

    return sheet_name


def open_vamas_file(window, file_path):
    """
    Open and process a VAMAS file, converting it to an Excel file format.
    This function reads a VAMAS file, extracts its data and metadata,
    and creates a new Excel file with multiple sheets for each data block
    and an additional sheet for experimental description.

    Args:
    window: The main application window object.
    file_path: The path to the VAMAS file to be opened.
    """
    try:
        # Clear undo and redo history
        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")

        # Copy VAMAS file to current working directory
        vamas_filename = os.path.basename(file_path)
        destination_path = os.path.join(os.getcwd(), vamas_filename)
        shutil.copy2(file_path, destination_path)

        # Read VAMAS file
        vamas_data = Vamas(vamas_filename)

        # Check for Casa peak fitting information
        has_casa_fitting = check_for_casa_fitting(vamas_data)
        import_fitting = False

        if has_casa_fitting:
            dlg = wx.MessageDialog(window,
                                   "Peak fitting information detected in VAMAS file.\n\nDo you want to import the "
                                   "peak fitting data?\n"
                                   "Note that this feature is still in beta testing and may not work perfectly.\n",
                                   "Import Peak Fitting [Beta testing]",
                                   wx.YES_NO | wx.ICON_QUESTION)
            result = dlg.ShowModal()
            import_fitting = (result == wx.ID_YES)
            dlg.Destroy()

        # Create new Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        exp_data = []

        # Create console window centered on parent
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Processing VAMAS File", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

        num_blocks = len(vamas_data.blocks)
        update_console(f"Found {num_blocks} core levels to process...")
        update_console("Starting conversion to Excel format...")

        # Process each block
        for i, block in enumerate(vamas_data.blocks, start=1):
            # Skip blocks with 0 scans
            if block.num_scans_to_compile_block == 0:
                update_console(f"Skipping block {i}: {block.species_label} - 0 scans")
                continue
            if block.species_label.lower() == "wide" or block.transition_or_charge_state_label.lower() == "none":
                raw_sheet_name = block.species_label
            else:
                raw_sheet_name = f"{block.species_label}{block.transition_or_charge_state_label}"
            raw_sheet_name = raw_sheet_name.replace("/", "_")

            # Apply Auger and valence band normalization
            raw_sheet_name = normalize_auger_and_valence_names(raw_sheet_name)

            sheet_name = normalize_sheet_name(raw_sheet_name)

            if sheet_name in wb.sheetnames:
                count = 1
                while f"{sheet_name}{count}" in wb.sheetnames:
                    count += 1
                sheet_name = f"{sheet_name}{count}"

            update_console(f"Processing {i}/{num_blocks}: {sheet_name}")

            # Create new sheet with normalized name
            ws = wb.create_sheet(title=sheet_name)

            # Extract and process data
            num_points = block.num_y_values
            x_start = block.x_start
            x_step = block.x_step
            x_values = [x_start + i * x_step for i in range(num_points)]
            y_values = block.corresponding_variables[0].y_values
            y_unit = block.corresponding_variables[0].unit
            num_scans = block.num_scans_to_compile_block


            try:
                collection_time = getattr(block, 'signal_collection_time', None)
                if collection_time is None:
                    collection_time = getattr(block, 'dwell_time', None)
                if collection_time is None:
                    collection_time = 1.0  # Default fallback
            except (AttributeError, TypeError):
                collection_time = 1.0  # Default fallback
            print(f"Collection time: {collection_time}, Num scans: {num_scans}, Y unit: {y_unit}")
            print(f'Number of Scans: {num_scans}')
            num_scan = num_scans  # To stop division by number of scan quickly if needed
            if y_unit != "c/s" and collection_time > 0:
                y_values = [y / (num_scan * collection_time) for y in y_values]
                print(f'Maximum Y value after conversion: {max(y_values)}')
            elif y_unit != "c/s":
                y_values = [y / num_scan for y in y_values]

            # # Convert to Binding Energy if necessary
            # if block.x_label.lower() in ["kinetic energy", "ke"]:
            #     x_values = [window.photons - x - window.workfunction for x in x_values]

            # Convert to Binding Energy if necessary
            if block.x_label.lower() in ["kinetic energy", "ke"]:
                # Use photon energy from VAMAS block, not hardcoded window.photons
                photon_energy = block.analysis_source_characteristic_energy
                x_values = [photon_energy - x - window.workfunction for x in x_values]
                print(f"KE to BE conversion using photon energy: {photon_energy:.2f} eV")
                x_label = "Binding Energy"
            else:
                x_label = block.x_label

            # Write data to sheet
            ws.append([x_label, "Corrected Data", "Raw Data", "Transmission"])

            # Get transmission data and correct it by dividing by (num_scans * collection_time)
            if hasattr(block, 'corresponding_variables') and len(block.corresponding_variables) > 1:
                raw_transmission_data = block.corresponding_variables[1].y_values
                # Simple correction: divide by num_scans * collection_time
                if collection_time > 0:
                    transmission_data = [t / (num_scans * collection_time) for t in raw_transmission_data]
                    print(f"DEBUG: Transmission corrected by dividing by ({num_scans} * {collection_time}) = {num_scans * collection_time}")
                else:
                    transmission_data = [t / num_scans for t in raw_transmission_data]
                    print(f"DEBUG: Transmission corrected by dividing by num_scans = {num_scans}")
            else:
                if collection_time > 0:
                    transmission_data = [1.0 / (num_scans * collection_time)] * len(y_values)
                else:
                    transmission_data = [1.0 / num_scans] * len(y_values)

            # Write data row by row
            for j, (x, y) in enumerate(zip(x_values, y_values)):
                trans = transmission_data[j] if j < len(transmission_data) else 1.0
                # Multiply by dwell time and number of scans for corrected data
                if collection_time > 0:
                    corrected_y = (y / abs(trans)) / (num_scans * collection_time)
                else:
                    corrected_y = (y / abs(trans)) / num_scans
                ws.append([x, corrected_y, y, trans])

            if import_fitting:
                if update_console:
                    update_console(f"Processing fitting data for {sheet_name}...")

                # Parse Casa fitting information
                casa_data = parse_casa_peak_fitting(block.block_comment, block.num_scans_to_compile_block,
                                                    window.photons, transmission_data, collection_time, sheet_name)

                if import_fitting and casa_data:
                    print(f"DEBUG: Casa data keys: {list(casa_data.keys())}")
                    print(f"DEBUG: Background data: {casa_data.get('Background', 'No Background key')}")

                    num_peaks = len(casa_data['Peaks'])
                    if num_peaks > 0:
                        peak_names = list(casa_data['Peaks'].keys())
                        if update_console:
                            update_console(f"  Found {num_peaks} peaks: {', '.join(peak_names)}")

                    # Store fitting data in a way that will be transferred to window.Data
                    if not hasattr(wb, '_fitting_data'):
                        wb._fitting_data = {}

                    wb._fitting_data[sheet_name] = {
                        'Fitting': {
                            'Peaks': casa_data['Peaks']
                        }
                    }

                    if casa_data['Background']:
                        print(f"DEBUG: Background exists, storing it")
                        wb._fitting_data[sheet_name]['Background'] = casa_data['Background']
                        wb._fitting_data[sheet_name]['Background']['Bkg X'] = x_values  # Full length X data

                        # DEBUG: Print array lengths
                        print(f"DEBUG: x_values length: {len(x_values)}")
                        print(f"DEBUG: y_values length: {len(y_values)}")
                        print(
                            f"DEBUG: transmission_data length: {len(transmission_data) if transmission_data else 'None'}")

                        # Calculate corrected_y for FULL LENGTH
                        corrected_y_values = []
                        for j, y in enumerate(y_values):
                            trans = transmission_data[j] if j < len(transmission_data) else 1.0
                            if collection_time > 0:
                                corrected_y = (y / abs(trans)) / (num_scans * collection_time)
                            else:
                                corrected_y = (y / abs(trans)) / num_scans
                            corrected_y_values.append(corrected_y)

                        print(f"DEBUG: corrected_y_values length: {len(corrected_y_values)}")
                        wb._fitting_data[sheet_name]['Background']['Bkg Y'] = corrected_y_values

                        if update_console:
                            update_console(
                                f"  Background: {casa_data['Background'].get('Bkg Type')} from {casa_data['Background'].get('Bkg Low'):.1f} to {casa_data['Background'].get('Bkg High'):.1f} eV")
                    else:
                        print(f"DEBUG: No background data in casa_data or background is empty")
                        print(f"DEBUG: casa_data['Background'] = {casa_data.get('Background', 'KEY NOT FOUND')}")
                else:
                    if update_console:
                        update_console(f"  No Casa fitting data found for {sheet_name}")

            # Store experimental setup data
            block_exp_data = [
                f"Block {i}",
                block.sample_identifier,
                f"{block.year}/{block.month}/{block.day}",
                f"{block.hour}:{block.minute}:{block.second}",
                block.technique,
                f"{block.species_label} {block.transition_or_charge_state_label}",
                block.num_scans_to_compile_block,
                block.analysis_source_label,
                block.analysis_source_characteristic_energy,
                block.analysis_source_beam_width_x,
                block.analysis_source_beam_width_y,
                block.analyzer_pass_energy_or_retard_ratio_or_mass_res,
                block.analyzer_work_function_or_acceptance_energy,
                block.analyzer_mode,
                block.sputtering_source_energy if hasattr(block, 'sputtering_source_energy') else 'N/A',
                block.analyzer_axis_take_off_polar_angle,
                block.analyzer_axis_take_off_azimuth,
                block.target_bias,
                block.analysis_width_x,
                block.analysis_width_y,
                block.x_label,
                block.x_units,
                block.x_start,
                block.x_step,
                block.num_y_values,
                block.num_scans_to_compile_block,
                block.signal_collection_time,
                block.signal_time_correction,
                y_unit,
                block.num_lines_block_comment,
                block.block_comment
            ]
            exp_data.append(block_exp_data)

            # Add experimental description data to this sheet starting at column 50
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            exp_labels = [
                "Sample ID", "Date", "Time", "Technique", "Species & Transition", "Number of scans",
                "Source Label", "Source Energy", "Source width X", "Source width Y", "Pass Energy", "Work Function",
                "Analyzer Mode", "Sputtering Energy", "Take-off Polar Angle", "Take-off Azimuth", "Target Bias",
                "Analysis Width X", "Analysis Width Y", "X Label", "X Units", "X Start", "X Step", "Num Y Values",
                "Num Scans", "Collection Time", "Time Correction", "Y Unit", "# Comment Lines", "Block Comment"
            ]

            for j, (label, value) in enumerate(zip(exp_labels, block_exp_data[1:])):
                ws.cell(row=j + 2, column=exp_col, value=label)
                ws.cell(row=j + 2, column=exp_col + 1, value=value)

            # Set column width for experimental data
            ws.column_dimensions[chr(64 + exp_col)].width = 25
            ws.column_dimensions[chr(64 + exp_col + 1)].width = 40

            # Store experimental info in sheet data structure for JSON
            experimental_info = {}
            for j, (label, value) in enumerate(zip(exp_labels, block_exp_data[1:])):
                experimental_info[label] = value

            # This needs to be stored when sheet data is created in window.Data
            # Add to the sheet data dictionary when it's created

        update_console("Creating experimental description sheet...")

        # Create "Experimental description" sheet (keep this for backward compatibility)
        exp_sheet = wb.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 50
        exp_sheet.column_dimensions['B'].width = 100
        left_aligned = Alignment(horizontal='left')

        # Add VAMAS header information
        exp_sheet.append(["VAMAS Header Information"])
        for item in [
            ("Format Identifier", vamas_data.header.format_identifier),
            ("Institution Identifier", vamas_data.header.institution_identifier),
            ("Instrument Model", vamas_data.header.instrument_model_identifier),
            ("Operator Identifier", vamas_data.header.operator_identifier),
            ("Experiment Identifier", vamas_data.header.experiment_identifier),
            ("Number of Comment Lines", vamas_data.header.num_lines_comment),
            ("Comment", vamas_data.header.comment),
            ("Experiment Mode", vamas_data.header.experiment_mode),
            ("Scan Mode", vamas_data.header.scan_mode),
            ("Number of Spectral Regions", vamas_data.header.num_spectral_regions),
            ("Number of Analysis Positions", vamas_data.header.num_analysis_positions),
            ("Number of Discrete X Coordinates", vamas_data.header.num_discrete_x_coords_in_full_map),
            ("Number of Discrete Y Coordinates", vamas_data.header.num_discrete_y_coords_in_full_map)
        ]:
            exp_sheet.append(item)

        exp_sheet.append([])  # Add a blank row for separation

        # Define the order of block information
        block_info_order = [
            "Sample ID", "Year/Month/Day", "Time HH,MM,SS", "Technique", "Species & Transition", "Number of scans",
            "Source Label", "Source Energy", "Source width X", "Source width Y", "Pass Energy", "Work Function",
            "Analyzer Mode", "Sputtering Energy", "Take-off Polar Angle", "Take-off Azimuth", "Target Bias",
            "Analysis Width X", "Analysis Width Y", "X Label", "X Units", "X Start", "X Step", "Num Y Values",
            "Num Scans", "Collection Time", "Time Correction", "Y Unit", "# Comment Lines", "Block Comment"
        ]

        # Add block information
        for i, block_data in enumerate(exp_data, start=1):
            exp_sheet.append([f"Block {i}", ""])
            for j, info in enumerate(block_info_order):
                exp_sheet.append([info, block_data[j + 1]])
            exp_sheet.append([])  # Add a blank row between blocks

        # Set alignment for all cells in column B
        for row in exp_sheet.iter_rows(min_row=1, max_row=exp_sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = left_aligned

        update_console("Saving Excel file...")
        excel_filename = os.path.splitext(vamas_filename)[0] + ".xlsx"
        excel_path = os.path.join(os.path.dirname(file_path), excel_filename)
        wb.save(excel_path)

        # Save fitting data to JSON file immediately after Excel file is saved
        if import_fitting and hasattr(wb, '_fitting_data'):
            print(f"VAMAS: Saving fitting data for {len(wb._fitting_data)} sheets")
            import json
            fitting_file = excel_path.replace('.xlsx', '_fitting.json')
            with open(fitting_file, 'w') as f:
                json.dump(wb._fitting_data, f)
            print(f"VAMAS: Fitting data saved to: {fitting_file}")

        update_console("Excel file created successfully!")
        update_console("Loading Excel file into KherveFitting...")

        os.remove(destination_path)
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        # Pass console to next function
        open_xlsx_file_vamas(window, excel_path, console_frame, update_console, import_fitting)

    except Exception as e:
        wx.MessageBox(f"Error processing VAMAS file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def parse_casa_peak_fitting(block_comment, num_scans=1, photon_energy=1486.67, transmission_data=None,
                           collection_time=1.0, core_level_name=None):
    """
    Parse Casa XPS peak fitting information from VAMAS block comment.

    Args:
        block_comment (str): The block comment containing Casa Info
        num_scans (int): Number of scans to divide area by
        photon_energy (float): X-ray photon energy for KE to BE conversion
        transmission_data (list): Transmission values for area correction

    Returns:
        dict: Dictionary containing background and peak fitting information
    """
    if "Casa Info Follows" not in block_comment:
        return None

    # Calculate average transmission if available
    avg_transmission = 1.0
    if transmission_data and len(transmission_data) > 0:
        avg_transmission = np.mean(transmission_data)

    lines = block_comment.split('\n')
    casa_start = -1

    # Find the start of Casa Info
    for i, line in enumerate(lines):
        if "Casa Info Follows" in line:
            casa_start = i + 1
            break

    if casa_start == -1:
        return None

    fitting_data = {
        'Background': {},
        'Peaks': {}
    }

    # Track peak names for numbering duplicates
    peak_name_counts = {}

    # Parse the lines after "Casa Info Follows"
    i = casa_start
    while i < len(lines):
        line = lines[i].strip()

        if line.startswith('CASA region'):
            # Parse background information
            parts = line.split()
            if len(parts) >= 5:
                # Extract background type - look for second (*...*) pattern
                bg_matches = re.findall(r'\(\*([^*]+)\*\)', line)
                bg_type = bg_matches[1] if len(bg_matches) > 1 else "Shirley"

                # Map Casa background types to KherveFitting types
                if bg_type.lower() == "shirley":
                    kf_bg_type = "Shirley"
                else:
                    kf_bg_type = bg_type

                # Extract energy range and convert from KE to BE
                try:
                    low_ke = float(parts[5])
                    high_ke = float(parts[6])
                    # Convert KE to BE
                    low_energy = photon_energy - high_ke
                    high_energy = photon_energy - low_ke
                    print(f"DEBUG: Parsed background - Type: {kf_bg_type}, Low: {low_energy}, High: {high_energy}")

                    fitting_data['Background'] = {
                        'Bkg Type': kf_bg_type,
                        'Bkg Low': round(low_energy,3),
                        'Bkg High': round(high_energy,3),
                        'Bkg Offset Low': '0',
                        'Bkg Offset High': '0',
                        'Bkg Y': []
                    }
                except (ValueError, IndexError):
                    pass
        elif line.startswith('CASA comp'):
            # Parse peak information
            # Extract peak name (first pattern)
            peak_matches = re.findall(r'\(\*([^*]+)\*\)', line)
            raw_peak_name = peak_matches[0] if peak_matches else f"Peak_{len(fitting_data['Peaks']) + 1}"

            # Remove spaces from core level names (e.g., "C 1s" -> "C1s", "Fe 2p" -> "Fe2p")
            if raw_peak_name:
                # Pattern matches: Element (1-2 letters) + space + orbital (number + letter)
                core_level_pattern = r'([A-Z][a-z]?)\s+(\d+[spdfghi])'
                raw_peak_name = re.sub(core_level_pattern, r'\1\2', raw_peak_name)

            # Handle duplicate peak names, especially when peak name equals core level name
            if core_level_name and raw_peak_name == core_level_name:
                # If peak name equals core level name, number them as p1, p2, etc.
                if raw_peak_name not in peak_name_counts:
                    peak_name_counts[raw_peak_name] = 0
                peak_name_counts[raw_peak_name] += 1
                peak_name = f"{raw_peak_name} p{peak_name_counts[raw_peak_name]}"
            else:
                # For other cases, check if name already exists
                if raw_peak_name in fitting_data['Peaks']:
                    if raw_peak_name not in peak_name_counts:
                        peak_name_counts[raw_peak_name] = 1
                    peak_name_counts[raw_peak_name] += 1
                    peak_name = f"{raw_peak_name} p{peak_name_counts[raw_peak_name]}"
                else:
                    peak_name = raw_peak_name


            # Extract model (second pattern)
            model_str = peak_matches[1] if len(peak_matches) > 1 else "GL(30)"

            # Parse model and determine KherveFitting equivalent
            model = "GL (Area)"
            lg_value = 30
            sigma_value = 0.6
            gamma_value = 0.4

            if 'SGL(' in model_str:
                lg_match = re.search(r'SGL\((\d+)\)', model_str)
                lg_value = float(lg_match.group(1)) if lg_match else 30
                model = "SGL (Area)"
            elif 'GL(' in model_str:
                lg_match = re.search(r'GL\((\d+)\)', model_str)
                lg_value = float(lg_match.group(1)) if lg_match else 30
                model = "GL (Area)"
            elif 'LA(' in model_str:
                # Extract LA parameters: LA(sigma, gamma, w_g)
                la_params = re.search(r'LA\(([\d.]+),\s*([\d.]+),\s*([\d.]+)\)', model_str)
                if la_params:
                    sigma_value = float(la_params.group(1))
                    gamma_value = float(la_params.group(2))
                    w_g = float(la_params.group(3))

                    if w_g > 100:
                        model = "LA*G (Area, σ/γ, γ)"
                    else:
                        model = "LA (Area, σ, γ)"

                    # Calculate L/G ratio from sigma and gamma
                    lg_value = 100 * sigma_value / (sigma_value + gamma_value)
                else:
                    # Check for two parameter format: LA(param1, param2)
                    two_param = re.search(r'LA\(([\d.]+),\s*([\d.]+)\)', model_str)
                    if two_param:
                        param1 = float(two_param.group(1))
                        param2 = float(two_param.group(2))

                        # If second parameter > 100, treat as LA*G with sigma=gamma=param1, w_g=param2
                        if param2 > 100:
                            sigma_value = param1  # 1.53
                            gamma_value = param1  # 1.53 (same as sigma)
                            w_g = param2  # 243
                            model = "LA*G (Area, σ/γ, γ)"
                            lg_value = 50  # 50% since sigma = gamma
                            print(
                                f"DEBUG: LA(2-param) large second -> LA*G: σ={sigma_value}, γ={gamma_value}, w_g={w_g}")
                        else:
                            # Standard two-parameter LA(sigma, gamma)
                            sigma_value = param1
                            gamma_value = param2
                            model = "LA (Area, σ, γ)"
                            lg_value = 100 * sigma_value / (sigma_value + gamma_value)
                            print(f"DEBUG: LA(2-param) standard -> LA: σ={sigma_value}, γ={gamma_value}")
                    else:
                        # Check for single parameter LA(number) format
                        single_param = re.search(r'LA\((\d+)\)', model_str)
                        if single_param:
                            ratio_value = int(single_param.group(1))

                            # Mapping for gamma values based on sigma/gamma ratio
                            gamma_mapping = {
                                20: 15, 30: 12, 40: 10, 50: 8, 60: 5,
                                70: 3, 80: 2.5, 90: 1.7, 100: 1.0
                            }

                            gamma_value = gamma_mapping.get(ratio_value, 2.0)  # Default to 2.0 if not found
                            sigma_value = gamma_value  # sigma = (sigma/gamma) * gamma
                            lg_value = ratio_value  # L/G ratio is the number itself
                            model = "LA (Area, σ/γ, γ)"
                        else:
                            model = "LA (Area, σ, γ)"
                            sigma_value = 0.6
                            gamma_value = 0.4
                            lg_value = 100 * sigma_value / (sigma_value + gamma_value)

            # Extract parameters with constraints
            area_match = re.search(r'Area\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)
            fwhm_match = re.search(r'MFWHM\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)
            pos_match = re.search(r'Position\s+([\d.e-]+)\s+([\d.e-]+)\s+([\d.e-]+)\s+(-?\d+)\s+([\d.e-]+)', line)

            # Calculate area (divide by transmission AND correction factor like raw data)
            if collection_time > 0:
                correction_factor = avg_transmission * num_scans * collection_time
            else:
                correction_factor = avg_transmission * num_scans

            area_value = float(area_match.group(1)) / correction_factor if area_match else 1000
            area_value = round(area_value, 2)
            # Handle area constraints
            if area_match:
                area_constrained_peak = int(area_match.group(4))
                area_constraint_value = float(area_match.group(5))

                if area_constrained_peak >= 0:
                    peak_letter = chr(65 + area_constrained_peak)  # A=0, B=1, etc.
                    area_constraint = f"{peak_letter}*{area_constraint_value}"
                else:
                    area_min = float(area_match.group(2)) if float(area_match.group(2)) > 1e-19 else 1
                    area_max = float(area_match.group(3)) / avg_transmission
                    area_constraint = f"{area_min}:{area_max}"
            else:
                area_constraint = "1:1e7"

            # Convert position from KE to BE
            pos_ke = float(pos_match.group(1)) if pos_match else 800
            position_be = photon_energy - pos_ke

            # Handle position constraints
            if pos_match:
                pos_constrained_peak = int(pos_match.group(4))
                pos_constraint_value = float(pos_match.group(5))

                if pos_constrained_peak >= 0:
                    peak_letter = chr(65 + pos_constrained_peak)
                    pos_constraint = f"{peak_letter}+0.0"  # Position typically uses +
                else:
                    pos_ke_min = float(pos_match.group(2))
                    pos_ke_max = float(pos_match.group(3))
                    pos_be_min = photon_energy - pos_ke_max  # Inverted
                    pos_be_max = photon_energy - pos_ke_min  # Inverted
                    pos_constraint = f"{pos_be_min:.2f}:{pos_be_max:.2f}"
            else:
                pos_constraint = "1:1000"

            # Handle FWHM constraints
            fwhm_value = float(fwhm_match.group(1)) if fwhm_match else 1.5
            fwhm_value = round(fwhm_value,2)
            if fwhm_match:
                fwhm_constrained_peak = int(fwhm_match.group(4))
                fwhm_constraint_value = float(fwhm_match.group(5))

                if fwhm_constrained_peak >= 0:
                    peak_letter = chr(65 + fwhm_constrained_peak)
                    fwhm_constraint = f"{peak_letter}*{fwhm_constraint_value}"
                else:
                    fwhm_constraint = f"{fwhm_match.group(2)}:{fwhm_match.group(3)}"
            else:
                fwhm_constraint = "0.3:3.5"

            # Calculate height from area and FWHM
            if 'LA' in model:
                height = area_value / (fwhm_value * 1.5)  # Approximation
            else:
                height = area_value / (fwhm_value * np.sqrt(np.pi / (4 * np.log(2))))

            peak_data = {
                'Position': position_be,
                'Height': round(height,2),
                'FWHM': round(fwhm_value,2),
                'L/G': lg_value,
                'Area': round(area_value,2),
                'Sigma': sigma_value,
                'Gamma': gamma_value,
                'Skew': 0.1,
                'Fitting Model': model,
                'Bkg Type': fitting_data['Background'].get('Bkg Type', ''),
                'Bkg Low': fitting_data['Background'].get('Bkg Low', '0'),
                'Bkg High': fitting_data['Background'].get('Bkg High', '0'),
                'Bkg Offset Low': '0',
                'Bkg Offset High': '0'
            }

            # Set constraints
            if 'LA' in model:
                lg_constraint = "Fixed"
            else:
                lg_constraint = "Fixed" if any(x in model_str for x in ['GL(', 'SGL(']) else "5:80"

            constraints = {
                'Position': pos_constraint,
                'Height': "1:1e7",
                'FWHM': fwhm_constraint,
                'L/G': lg_constraint,
                'Area': area_constraint,
                'Sigma': "0.01:10" if 'LA' in model else "0.3:3",
                'Gamma': "0.01:10" if 'LA' in model else "0.3:3",
                'Skew': "0.01:2"
            }

            peak_data['Constraints'] = constraints
            fitting_data['Peaks'][peak_name] = peak_data

        i += 1

    return fitting_data if (fitting_data['Background'] or fitting_data['Peaks']) else None


def check_for_casa_fitting(vamas_data):
    """
    Check if any blocks contain Casa peak fitting information.

    Args:
        vamas_data: VAMAS data object

    Returns:
        bool: True if Casa fitting info found
    """
    for block in vamas_data.blocks:
        if hasattr(block, 'block_comment') and "Casa Info Follows" in block.block_comment:
            return True
    return False


def open_xlsx_file_vamas(window, file_path, console_frame=None, update_console=None, has_fitting=False):
    """
    Open and process an Excel file created from a VAMAS file.

    This function initializes the data structure, reads the Excel file,
    populates the window.Data dictionary with core level information,
    updates the GUI elements, and plots the data for the first sheet.

    Args:
    window: The main application window object.
    file_path: The path to the Excel file to be opened.
    console_frame: Console window for progress updates.
    update_console: Function to update console output.
    has_fitting: Whether fitting data should be loaded.
    """
    try:
        if update_console:
            update_console("Reading Excel file structure...")

        window.SetStatusText(f"Selected File: {file_path}", 0)
        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = file_path

        # Load fitting data if available
        fitting_data = {}
        if has_fitting:
            fitting_file = file_path.replace('.xlsx', '_fitting.json')
            if os.path.exists(fitting_file):
                if update_console:
                    update_console("Loading peak fitting data...")

                with open(fitting_file, 'r') as f:
                    fitting_data = json.load(f)

                # Update console with peak information
                for sheet_name, data in fitting_data.items():
                    num_peaks = len(data.get('Fitting', {}).get('Peaks', {}))
                    if num_peaks > 0:
                        peak_names = list(data.get('Fitting', {}).get('Peaks', {}).keys())
                        if update_console:
                            update_console(f"  {sheet_name}: {num_peaks} peaks - {', '.join(peak_names)}")

                # Clean up temporary file
                os.remove(fitting_file)

        excel_file = pd.ExcelFile(file_path)
        sheet_names = [name for name in excel_file.sheet_names if name != "Experimental description"]
        window.Data['Number of Core levels'] = 0

        # After loading JSON fitting data, add this debug:
        if hasattr(window, 'Data') and 'Core levels' in window.Data:
            for core_level, data in window.Data['Core levels'].items():
                if 'Background' in data:
                    bg_y = data['Background'].get('Bkg Y', [])
                    print(f"DEBUG: Core level {core_level} background length: {len(bg_y)}")
                if 'B.E.' in data:
                    be_data = data['B.E.']
                    print(f"DEBUG: Core level {core_level} B.E. length: {len(be_data)}")

        if update_console:
            update_console(f"Found {len(sheet_names)} sheets to load...")

        for i, sheet_name in enumerate(sheet_names, 1):
            if update_console:
                update_console(f"Loading sheet {i}/{len(sheet_names)}: {sheet_name}")

            window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

            # Add fitting data if available for this sheet
            if fitting_data and sheet_name in fitting_data:
                if 'Core levels' in window.Data and sheet_name in window.Data['Core levels']:
                    # Add the fitting data to the core level
                    window.Data['Core levels'][sheet_name].update(fitting_data[sheet_name])

        if update_console:
            update_console("Updating interface...")

        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)
        window.sheet_combobox.SetValue(sheet_names[0])

        # Refresh peak fitting grid to show imported peaks
        from libraries.Sheet_Operations import on_sheet_selected
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(sheet_names[0])
        on_sheet_selected(window, event)

        # Calculate backgrounds for sheets with fitting data
        for sheet_name in sheet_names:
            if (fitting_data and sheet_name in fitting_data and
                    'Background' in fitting_data[sheet_name]):
                window.sheet_combobox.SetValue(sheet_name)
                bg_data = fitting_data[sheet_name]['Background']
                window.bg_min_energy = bg_data.get('Bkg Low')  # Add print here
                window.bg_max_energy = bg_data.get('Bkg High')  # Add print here
                print(f"DEBUG: Setting bg range: {window.bg_min_energy} to {window.bg_max_energy}")
                window.background_method = bg_data.get('Bkg Type', 'Multi-Regions Smart')
                window.plot_manager.plot_background(window, use_smoothing=False)

        # Set back to first sheet
        window.sheet_combobox.SetValue(sheet_names[0])

        window.plot_manager.plot_data(window)

        window.plot_manager.clear_and_replot(window)

        if update_console:
            update_console("Loading complete!")
            wx.CallLater(500, console_frame.Close)

    except Exception as e:
        wx.MessageBox(f"Error reading Excel file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def extract_transmission_data(block):
    """
    Extract transmission function data from a Kratos .kal file block.

    Args:
        block (str): Text block containing transmission data

    Returns:
        tuple: (ke_trans, trans_values) arrays of kinetic energies and transmission values
    """
    if 'Transmission Function Object (ke,t)' in block:
        trans_section = block.split('Transmission Function Object (ke,t)')[1].split('Dataset filename')[0]

        ke_line = trans_section.split('Transmission Function Kinetic Energy =')[1].split('\n')[0]
        ke_str = ke_line.strip().strip('{}').strip()
        ke_trans = np.array([float(x.strip()) for x in ke_str.split(',')])

        val_line = trans_section.split('Transmission Function Value          =')[1].split('\n')[0]
        val_str = val_line.strip().strip('{}').strip()
        trans_values = np.array([float(x.strip()) for x in val_str.split(',')])

        return ke_trans, trans_values
    return None, None


def convert_kal_to_excel(file_path):
    """
    Convert Kratos .kal file to Excel format suitable for KherveFitting.

    Args:
        file_path (str): Path to the .kal file

    Returns:
        str: Path to the created Excel file
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    with open(file_path, 'r') as f:
        content = f.read()

    blocks = content.split('Dataset filename')
    spectra = {}
    PHOTON_ENERGY = 1486.67

    # Extract sample ID
    sample_id = "Unknown"
    for block in blocks:
        for line in block.split('\n'):
            if 'Stage Position Name' in line:
                sample_id = line.split('=')[1].strip()
                break
        if sample_id != "Unknown":
            break

    # Process each spectrum block
    for block in blocks:
        if 'Ordinate values' not in block or 'Object name' not in block:
            continue

        # Get spectrum name and number
        object_name_line = next((line for line in block.split('\n') if 'Object name' in line), None)
        if not object_name_line:
            continue

        name_parts = object_name_line.split('=')[1].strip().split('/')
        name = name_parts[0].strip()
        spectrum_id = name_parts[1].strip() if len(name_parts) > 1 else ""

        # Skip non-spectrum blocks
        if 'Sample Position' in name or 'Counter' in name:
            continue

        # Initialize metadata dictionary with default values
        metadata = {
            'Sample ID': sample_id,
            'Date': '',
            'Time': '',
            'Technique': 'XPS',
            'Species & Transition': '',
            'Number of scans': '',
            'Source Label': 'Al mono',
            'Source Energy': '1486.6',
            'Source width X': '1E+37',
            'Source width Y': '1E+37',
            'Pass Energy': '',
            'Work Function': '1E+37',
            'Analyzer Mode': 'FAT',
            'Sputtering Energy': 'N/A',
            'Take-off Polar Angle': '1E+37',
            'Take-off Azimuth': '1E+37',
            'Target Bias': '1E+37',
            'Analysis Width X': '1E+37',
            'Analysis Width Y': '1E+37',
            'X Label': '',
            'X Units': '',
            'X Start': '',
            'X Step': '',
            'Num Y Values': '',
            'Num Scans': '',
            'Collection Time': '',
            'Time Correction': '1E+37',
            'Y Unit': 'd',
            '# Comment Lines': '0',
            'Block Comment': ''
        }

        # Extract metadata from block
        for line in block.split('\n'):
            line = line.strip()

            # Date and Time
            if 'Date Acquired' in line:
                parts = line.split('=')[1].strip().split()
                if len(parts) >= 2:
                    metadata['Date'] = parts[0]
                    metadata['Time'] = parts[1]

            # Species & Transition
            elif 'Chemical symbol or formula' in line:
                metadata['species'] = line.split('=')[1].strip()
            elif 'Transition or charge state' in line:
                metadata['transition'] = line.split('=')[1].strip()

            # Number of scans
            elif '# Sweeps completed' in line:
                num_scans = line.split('=')[1].strip()
                metadata['Number of scans'] = num_scans
                metadata['Num Scans'] = num_scans

            # Pass Energy
            elif 'Pass energy' in line:
                metadata['Pass Energy'] = line.split('=')[1].strip()

            # X Label, Units, Start, Step
            elif 'Abscissa label' in line:
                metadata['X Label'] = line.split('=')[1].strip()
            elif 'Abscissa units' in line:
                metadata['X Units'] = line.split('=')[1].strip()
            elif 'Spectrum scan start' in line:
                metadata['X Start'] = line.split('=')[1].split('eV')[0].strip()
            elif 'Spectrum scan step size' in line:
                metadata['X Step'] = line.split('=')[1].split('eV')[0].strip()

            # Collection Time
            elif 'Dwell time' in line:
                metadata['Collection Time'] = line.split('=')[1].replace('seconds', '').strip()

        # Combine species and transition
        if 'species' in metadata and 'transition' in metadata:
            metadata['Species & Transition'] = f"{metadata['species']} {metadata['transition']}"
            metadata.pop('species')
            metadata.pop('transition')

        # Count Y values
        if 'Ordinate values' in block:
            try:
                values_str = block.split('Ordinate values')[1].split('=')[1].split('}')[0].strip().strip('{').strip()
                values = values_str.split(',')
                metadata['Num Y Values'] = str(len(values))
            except:
                pass

        # Extract block comment
        comment_lines = []
        in_comment = False
        for line in block.split('\n'):
            if 'Casa Info Follows' in line:
                in_comment = True
                comment_lines.append(line.strip())
            elif in_comment and line.strip():
                comment_lines.append(line.strip())

        if comment_lines:
            metadata['# Comment Lines'] = str(len(comment_lines))
            metadata['Block Comment'] = '"' + '\n'.join(comment_lines) + '"'

        # Process spectral data
        ke_trans, trans_values = extract_transmission_data(block)
        if ke_trans is not None and trans_values is not None:
            # Create interpolation function
            trans_func = interp1d(ke_trans, trans_values, kind='linear', bounds_error=False,
                                  fill_value='extrapolate')

            start_ke = float(metadata['X Start'])
            step = float(metadata['X Step'])
            raw_str = block.split('Ordinate values')[1].split('=')[1].split('}')[0].strip().strip('{').strip()
            raw_data = np.array([float(x.strip()) for x in raw_str.split(',')])

            num_points = len(raw_data)
            ke_values = np.linspace(start_ke, start_ke + (num_points - 1) * step, num_points)
            be_values = PHOTON_ENERGY - ke_values

            transmission = trans_func(ke_values)
            corrected_data = raw_data / transmission

            df = pd.DataFrame({
                'BE': be_values,
                'Corrected Data': corrected_data,
                'Raw Data': raw_data,
                'Transmission': transmission
            })

            spectra[name] = {
                'data': df,
                'metadata': metadata,
                'id': spectrum_id
            }

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets for each spectrum
    for name, info in spectra.items():
        sheet_name = name.replace(' ', '')
        df = info['data']
        metadata = info['metadata']

        # Create sheet and write data
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data values
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add experimental description (column 50)
        exp_col = 50
        ws.cell(row=1, column=exp_col, value="Experimental Description")

        # Metadata fields in specified order
        fields = [
            'Sample ID', 'Date', 'Time', 'Technique', 'Species & Transition',
            'Number of scans', 'Source Label', 'Source Energy', 'Source width X',
            'Source width Y', 'Pass Energy', 'Work Function', 'Analyzer Mode',
            'Sputtering Energy', 'Take-off Polar Angle', 'Take-off Azimuth',
            'Target Bias', 'Analysis Width X', 'Analysis Width Y', 'X Label',
            'X Units', 'X Start', 'X Step', 'Num Y Values', 'Num Scans',
            'Collection Time', 'Time Correction', 'Y Unit', '# Comment Lines',
            'Block Comment'
        ]

        # Write metadata in order
        for i, field in enumerate(fields, 2):
            ws.cell(row=i, column=exp_col, value=field)
            ws.cell(row=i, column=exp_col + 1, value=metadata.get(field, ''))

        # Set column widths
        ws.column_dimensions[get_column_letter(exp_col)].width = 25
        ws.column_dimensions[get_column_letter(exp_col + 1)].width = 50

    # Create Experimental description sheet
    exp_sheet = wb.create_sheet("Experimental description")
    exp_sheet.column_dimensions['A'].width = 30
    exp_sheet.column_dimensions['B'].width = 100
    left_aligned = Alignment(horizontal='left')

    # Write metadata to Experimental description sheet
    row = 1
    for name, info in spectra.items():
        metadata = info['metadata']
        spectrum_id = info['id']

        # Add spectrum header
        header = exp_sheet.cell(row=row, column=1, value=f"Spectrum: {name}/{spectrum_id}")
        header.alignment = left_aligned
        row += 1

        # Write metadata fields
        for field in fields:
            cell_a = exp_sheet.cell(row=row, column=1, value=field)
            cell_b = exp_sheet.cell(row=row, column=2, value=metadata.get(field, ''))

            cell_a.alignment = left_aligned
            cell_b.alignment = left_aligned

            row += 1

        # Add space between spectra
        row += 1

    # Save Excel file
    output_file = file_path.replace('.kal', '.xlsx')
    wb.save(output_file)

    return output_file

def open_kal_file_dialog(window):
    """Open a file dialog for selecting a Kratos .kal file"""
    with wx.FileDialog(window, "Open Kratos file", wildcard="Kratos files (*.kal)|*.kal",
                      style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_kal_file(window, file_path)

def open_kal_file(window, file_path):
    """Process a Kratos .kal file and convert it to Excel format"""

    try:
        output_excel = convert_kal_to_excel(file_path)
        open_xlsx_file(window, output_excel)
    except Exception as e:
        wx.MessageBox(f"Error processing Kratos file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_raman_txt_file(window):
    import wx
    import os
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.FileDialog(window, "Open Raman text file", wildcard="Text files (*.txt)|*.txt",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = f"Raman_{base_filename}"

        # Read data from txt file
        data = []
        with open(file_path, 'r') as f:
            for line in f:
                if line.strip() and not line.startswith('#'):
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        try:
                            wavenumber = float(parts[0])
                            intensity = float(parts[1])
                            data.append([wavenumber, intensity])
                        except ValueError:
                            continue

        if not data:
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "Wavenumber (cm-1)"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (wavenumber, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = wavenumber
            ws[f"B{i}"] = intensity

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")


def import_multiple_raman_files(window):
    import wx
    import os
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing Raman text files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all txt files in the directory
        txt_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.txt')]

        if not txt_files:
            window.show_popup_message2("Information", "No .txt files found in the selected folder.")
            return

        # Create an Excel file for each group of files
        excel_path = os.path.join(dir_path, "Raman_Data.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each txt file
        for txt_file in txt_files:
            file_path = os.path.join(dir_path, txt_file)
            base_filename = os.path.splitext(txt_file)[0]
            sheet_name = f"Raman_{base_filename}"

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from txt file
            data = []
            with open(file_path, 'r') as f:
                for line in f:
                    if line.strip() and not line.startswith('#'):
                        parts = line.strip().split()
                        if len(parts) >= 2:
                            try:
                                wavenumber = float(parts[0])
                                intensity = float(parts[1])
                                data.append([wavenumber, intensity])
                            except ValueError:
                                continue

            if not data:
                continue  # Skip files with no valid data

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "Wavenumber (cm-1)"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (wavenumber, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = wavenumber
                ws[f"B{i}"] = intensity

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} Raman data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the text files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")

def import_xps_asc_file(window):
    import wx
    import os
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.FileDialog(window, "Open XPS ASC file", wildcard="ASC files (*.asc)|*.asc",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = base_filename

        # Read data from asc file
        data = []
        with open(file_path, 'r') as f:
            for line in f:
                if line.strip() and not line.startswith('#'):
                    # Try different separators
                    if ';' in line:
                        parts = line.strip().split(';')
                    elif ',' in line:
                        parts = line.strip().split(',')
                    else:
                        # Fall back to whitespace separator
                        parts = line.strip().split()

                    if len(parts) >= 2:
                        try:
                            binding_energy = float(parts[0])
                            intensity = float(parts[1])
                            data.append([binding_energy, intensity])
                        except ValueError:
                            continue

        if not data:
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "BE"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (binding_energy, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = f"{binding_energy:.2f}"
            ws[f"B{i}"] = f"{intensity:.2f}"

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")


def import_multiple_xps_asc_files(window):
    import wx
    import os
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing XPS ASC files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all asc files in the directory
        asc_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.asc')]

        if not asc_files:
            window.show_popup_message2("Information", "No .asc files found in the selected folder.")
            return

        # Create an Excel file for each group of files
        excel_path = os.path.join(dir_path, "XPS_Data.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each asc file
        for asc_file in asc_files:
            file_path = os.path.join(dir_path, asc_file)
            base_filename = os.path.splitext(asc_file)[0]
            sheet_name = base_filename

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from asc file
            data = []
            with open(file_path, 'r') as f:
                for line in f:
                    if line.strip() and not line.startswith('#'):
                        # Try different separators
                        if ';' in line:
                            parts = line.strip().split(';')
                        elif ',' in line:
                            parts = line.strip().split(',')
                        else:
                            # Fall back to whitespace separator
                            parts = line.strip().split()

                        if len(parts) >= 2:
                            try:
                                binding_energy = float(parts[0])
                                intensity = float(parts[1])
                                data.append([binding_energy, intensity])
                            except ValueError:
                                continue
            if not data:
                continue  # Skip files with no valid data

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "BE"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (binding_energy, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = binding_energy
                ws[f"B{i}"] = intensity

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} XPS data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the ASC files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")


def import_xps_asc_file_direct(window, file_path):
    import os
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    try:
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = base_filename

        data = []
        with open(file_path, 'r') as f:
            for line in f:
                if line.strip() and not line.startswith('#'):
                    # Try different separators
                    if ';' in line:
                        parts = line.strip().split(';')
                    elif ',' in line:
                        parts = line.strip().split(',')
                    else:
                        # Fall back to whitespace separator
                        parts = line.strip().split()

                    if len(parts) >= 2:
                        try:
                            binding_energy = float(parts[0])
                            intensity = float(parts[1])
                            data.append([binding_energy, intensity])
                        except ValueError:
                            continue

        if not data:
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = "BE"
        ws["B1"] = "Raw Data"

        for i, (binding_energy, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = f"{binding_energy:.2f}"
            ws[f"B{i}"] = f"{intensity:.2f}"

        wb.save(excel_path)
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")

def import_xps_csv_file(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.FileDialog(window, "Open XPS CSV file", wildcard="CSV files (*.csv)|*.csv",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()

    try:
        # Get filename without extension for sheet name
        base_filename = os.path.basename(file_path).split('.')[0]
        sheet_name = base_filename

        # Read data from csv file
        df = pd.read_csv(file_path, delimiter=',', header=None)
        if df.shape[1] < 2:
            # Try other common separators if comma doesn't work
            for sep in [';', '\t', ' ']:
                test_df = pd.read_csv(file_path, delimiter=sep, header=None)
                if test_df.shape[1] >= 2:
                    df = test_df
                    break

        if df.shape[1] < 2:
            window.show_popup_message2("Error", "CSV file must contain at least two columns (BE and Intensity).")
            return

        # Extract the first two columns (assuming BE and intensity)
        data = df.iloc[:, :2].values

        if len(data) == 0:  # Use len() instead of direct boolean check
            window.show_popup_message2("Error", "No valid data found in the file.")
            return

        # Create new Excel file if needed
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_filename}.xlsx")

        if os.path.exists(excel_path):
            # If Excel file exists, load it
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create new Excel file
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create or replace the sheet
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws["A1"] = "Binding Energy (eV)"
        ws["B1"] = "Raw Data"

        # Add data
        for i, (binding_energy, intensity) in enumerate(data, start=2):
            ws[f"A{i}"] = float(binding_energy)
            ws[f"B{i}"] = float(intensity)

        # Save Excel file
        wb.save(excel_path)

        # Open the Excel file in Khervefitting
        open_xlsx_file(window, excel_path)

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing file: {str(e)}")

def import_multiple_xps_csv_files(window):
    import wx
    import os
    import pandas as pd
    import openpyxl
    from libraries.FileMenu.Open import open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing XPS CSV files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return
        dir_path = dirDialog.GetPath()

    try:
        # Find all csv files in the directory
        csv_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.csv')]

        if not csv_files:
            window.show_popup_message2("Information", "No .csv files found in the selected folder.")
            return

        # Create an Excel file for all files
        excel_path = os.path.join(dir_path, "XPS_Data_CSV.xlsx")
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Process each csv file
        for csv_file in csv_files:
            file_path = os.path.join(dir_path, csv_file)
            base_filename = os.path.splitext(csv_file)[0]
            sheet_name = base_filename

            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # Handle duplicate sheet names
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:27]}_{count}"
                count += 1

            # Read data from csv file
            try:
                df = pd.read_csv(file_path, delimiter=',', header=None)
                if df.shape[1] < 2:
                    # Try other common separators if comma doesn't work
                    for sep in [';', '\t', ' ']:
                        test_df = pd.read_csv(file_path, delimiter=sep, header=None)
                        if test_df.shape[1] >= 2:
                            df = test_df
                            break

                if df.shape[1] < 2:
                    continue  # Skip files without at least two columns

                # Extract the first two columns (assuming BE and intensity)
                data = df.iloc[:, :2].values

                if len(data) == 0:  # Use len() instead of direct boolean check
                    continue  # Skip files with no valid data
            except Exception as e:
                print(f"Error reading {csv_file}: {e}")
                continue

            # Create sheet
            ws = wb.create_sheet(sheet_name)

            # Add headers
            ws["A1"] = "Binding Energy (eV)"
            ws["B1"] = "Raw Data"

            # Add data
            for i, (binding_energy, intensity) in enumerate(data, start=2):
                ws[f"A{i}"] = float(binding_energy)
                ws[f"B{i}"] = float(intensity)

        # Save Excel file if sheets were created
        if len(wb.sheetnames) > 0:
            wb.save(excel_path)
            open_xlsx_file(window, excel_path)
            window.show_popup_message2("Success", f"Created Excel file with {len(wb.sheetnames)} XPS data sheets.")
        else:
            window.show_popup_message2("Information", "No valid data found in any of the CSV files.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing files: {str(e)}")

def open_file_location(window):
    if 'FilePath' in window.Data:
        file_path = window.Data['FilePath']
        folder_path = os.path.dirname(file_path)
        if os.path.exists(folder_path):
            if sys.platform == 'win32':
                os.startfile(folder_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{folder_path}"')
            else:
                os.system(f'xdg-open "{folder_path}"')


def open_vg_microtech_file(window, file_path):
    """
    Process a VG-Microtech .1 file and convert it to Excel format suitable for KherveFitting.

    Args:
        window: The main application window
        file_path: Path to the .1 file

    Returns:
        bool: Success status
    """
    import numpy as np
    import openpyxl
    import os

    try:
        # Read the file content
        with open(file_path, 'r') as f:
            lines = f.readlines()

        # Parse header information
        header = lines[1].strip().split()
        print(f'Header: {header}')
        be_start = float(header[0])
        be_end = float(header[1])
        energy_step = float(header[2])
        unknown_param = float(header[3])
        dwell_time = float(header[4])
        num_points = int(header[5])
        pass_energy = float(header[6])
        photon_energy = abs(float(header[7]))  # Absolute value for negative photon energies

        # Get measurement type
        measurement_type = lines[2].strip()

        # Extract intensity values
        intensity_values = []
        for i in range(3, len(lines)):
            if lines[i].strip():
                try:
                    intensity_values.append(float(lines[i].strip()))
                except ValueError:
                    continue

        # Ensure we have the right number of intensity values
        if len(intensity_values) < num_points:
            window.show_popup_message2("Warning",
                                       f"Expected {num_points} data points but found only {len(intensity_values)}.")
        elif len(intensity_values) > num_points:
            intensity_values = intensity_values[:num_points]

        # Calculate binding energy values
        be_values = np.linspace(be_start, be_end, num_points)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheet with normalized measurement type as name
        from libraries.FileMenu.Open import normalize_sheet_name
        sheet_name = normalize_sheet_name(measurement_type)
        ws = wb.create_sheet(title=sheet_name)

        # Add column headers
        ws["A1"] = "Binding Energy (eV)"
        ws["B1"] = "Corrected Data"
        ws["C1"] = "Raw Data"
        ws["D1"] = "Transmission"

        # Add data rows
        for i, (be, intensity) in enumerate(zip(be_values, intensity_values), start=2):
            ws[f"A{i}"] = be
            ws[f"B{i}"] = intensity
            ws[f"C{i}"] = intensity
            ws[f"D{i}"] = 1.0  # Default transmission value

        # Create experimental description sheet
        exp_sheet = wb.create_sheet(title="Experimental description")
        exp_sheet.column_dimensions['A'].width = 30
        exp_sheet.column_dimensions['B'].width = 50

        # Add metadata
        exp_sheet["A1"] = "Sample ID"
        exp_sheet["B1"] = os.path.basename(file_path)
        exp_sheet["A2"] = "BE Start"
        exp_sheet["B2"] = str(be_start)
        exp_sheet["A3"] = "BE End"
        exp_sheet["B3"] = str(be_end)
        exp_sheet["A4"] = "Energy Step"
        exp_sheet["B4"] = str(energy_step)
        exp_sheet["A5"] = "Dwell Time"
        exp_sheet["B5"] = str(dwell_time)
        exp_sheet["A6"] = "Number of Points"
        exp_sheet["B6"] = str(num_points)
        exp_sheet["A7"] = "Pass Energy"
        exp_sheet["B7"] = str(pass_energy)
        exp_sheet["A8"] = "Photon Energy"
        exp_sheet["B8"] = str(photon_energy)
        exp_sheet["A9"] = "Technique"
        exp_sheet["B9"] = "XPS"
        exp_sheet["A10"] = "Species & Transition"
        exp_sheet["B10"] = measurement_type

        # Save Excel file
        excel_path = os.path.splitext(file_path)[0] + ".xlsx"
        wb.save(excel_path)

        # Open the created Excel file
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, excel_path)
        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing VG-Microtech file: {str(e)}")
        return False


def open_vg_microtech_file_dialog(window):
    """
    Open a file dialog for selecting a VG-Microtech .1 file and process it.
    """
    import wx

    with wx.FileDialog(window, "Open VG-Microtech file", wildcard="VG-Microtech files (*.1)|*.1",
                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        file_path = fileDialog.GetPath()
        open_vg_microtech_file(window, file_path)


def import_multiple_vg_microtech_files(window):
    """
    Import multiple VG-Microtech .1 files from a folder.
    """
    import wx
    import os
    import numpy as np
    import openpyxl
    from libraries.FileMenu.Open import normalize_sheet_name, open_xlsx_file

    with wx.DirDialog(window, "Choose a directory containing VG-Microtech .1 files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all .1 files in the directory
        vg_files = [f for f in os.listdir(folder_path) if f.endswith('.1')]

        if not vg_files:
            window.show_popup_message2("Information", "No .1 files found in the selected folder.")
            return

        # Ask if user wants individual Excel files or one combined file
        dlg = wx.MessageDialog(window,
                               "Do you want to create individual Excel files for each .1 file or combine them into one Excel file?",
                               "Import Options",
                               wx.YES_NO | wx.ICON_QUESTION)
        dlg.SetYesNoLabels("Individual Files", "One Combined File")

        result = dlg.ShowModal()
        individual_files = (result == wx.ID_YES)
        dlg.Destroy()

        if individual_files:
            # Process each .1 file individually
            processed_count = 0
            for vg_file in vg_files:
                vg_path = os.path.join(folder_path, vg_file)
                if open_vg_microtech_file(window, vg_path):
                    processed_count += 1

            window.show_popup_message2("Success", f"Processed {processed_count} of {len(vg_files)} VG-Microtech files.")
        else:
            # Combine all .1 files into one Excel file
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            processed_count = 0
            exp_sheet = combined_wb.create_sheet(title="Experimental description")
            exp_sheet.column_dimensions['A'].width = 30
            exp_sheet.column_dimensions['B'].width = 50

            row = 1

            for vg_file in vg_files:
                vg_path = os.path.join(folder_path, vg_file)

                try:
                    # Read the file content
                    with open(vg_path, 'r') as f:
                        lines = f.readlines()

                    # Parse header information
                    header = lines[1].strip().split()
                    be_start = float(header[0])
                    be_end = float(header[1])
                    energy_step = float(header[2])
                    unknown_param = float(header[3])
                    dwell_time = float(header[4])
                    num_points = int(header[5])
                    pass_energy = float(header[6])
                    photon_energy = abs(float(header[7]))

                    # Get measurement type
                    measurement_type = lines[2].strip()

                    # Create a unique sheet name
                    sheet_name = normalize_sheet_name(measurement_type)
                    if sheet_name in combined_wb.sheetnames:
                        # If sheet name already exists, append a number
                        count = 1
                        while f"{sheet_name}{count}" in combined_wb.sheetnames:
                            count += 1
                        sheet_name = f"{sheet_name}{count}"

                    # Create sheet
                    ws = combined_wb.create_sheet(title=sheet_name)

                    # Add column headers
                    ws["A1"] = "Binding Energy (eV)"
                    ws["B1"] = "Corrected Data"
                    ws["C1"] = "Raw Data"
                    ws["D1"] = "Transmission"

                    # Extract intensity values
                    intensity_values = []
                    for i in range(3, len(lines)):
                        if lines[i].strip():
                            try:
                                intensity_values.append(float(lines[i].strip()))
                            except ValueError:
                                continue

                    # Ensure we have the right number of intensity values
                    if len(intensity_values) > num_points:
                        intensity_values = intensity_values[:num_points]

                    # Calculate binding energy values
                    be_values = np.linspace(be_start, be_end, num_points)

                    # Match be_values and intensity_values lengths
                    actual_points = min(len(be_values), len(intensity_values))
                    be_values = be_values[:actual_points]
                    intensity_values = intensity_values[:actual_points]

                    # Add data rows
                    for i, (be, intensity) in enumerate(zip(be_values, intensity_values), start=2):
                        ws[f"A{i}"] = be
                        ws[f"B{i}"] = intensity
                        ws[f"C{i}"] = intensity
                        ws[f"D{i}"] = 1.0  # Default transmission value

                    # Add metadata to experimental description sheet
                    exp_sheet[f"A{row}"] = f"File: {vg_file}"
                    exp_sheet[f"B{row}"] = sheet_name
                    row += 1

                    exp_sheet[f"A{row}"] = "BE Start"
                    exp_sheet[f"B{row}"] = str(be_start)
                    row += 1

                    exp_sheet[f"A{row}"] = "BE End"
                    exp_sheet[f"B{row}"] = str(be_end)
                    row += 1

                    exp_sheet[f"A{row}"] = "Energy Step"
                    exp_sheet[f"B{row}"] = str(energy_step)
                    row += 1

                    exp_sheet[f"A{row}"] = "Dwell Time"
                    exp_sheet[f"B{row}"] = str(dwell_time)
                    row += 1

                    exp_sheet[f"A{row}"] = "Number of Points"
                    exp_sheet[f"B{row}"] = str(num_points)
                    row += 1

                    exp_sheet[f"A{row}"] = "Pass Energy"
                    exp_sheet[f"B{row}"] = str(pass_energy)
                    row += 1

                    exp_sheet[f"A{row}"] = "Photon Energy"
                    exp_sheet[f"B{row}"] = str(photon_energy)
                    row += 1

                    # Add a blank row for separation
                    row += 1

                    processed_count += 1

                except Exception as e:
                    print(f"Error processing {vg_file}: {str(e)}")

            # Save combined Excel file
            if processed_count > 0:
                combined_excel_path = os.path.join(folder_path, "VG_Microtech_combined.xlsx")
                combined_wb.save(combined_excel_path)

                # Open the combined Excel file
                open_xlsx_file(window, combined_excel_path)

                window.show_popup_message2("Success",
                                           f"Created combined Excel file with {processed_count} sheets from VG-Microtech files.")
            else:
                window.show_popup_message2("Error", "No valid data found in any VG-Microtech file.")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Error", f"Error processing VG-Microtech files: {str(e)}")


def import_multiple_kfitting_files(window):
    """
    Import multiple KherveFitting Excel files from a folder in alphabetical order.
    Creates a single Excel file with numbered core levels (C1s0, O1s0, C1s1, O1s1...).
    Filenames are stored in SampleNames section of window.data JSON.
    """
    with wx.DirDialog(window, "Choose a directory containing KherveFitting Excel files",
                      style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:

        if dirDialog.ShowModal() == wx.ID_CANCEL:
            return

        folder_path = dirDialog.GetPath()

    try:
        # Find all Excel files in the directory
        excel_files = [f for f in os.listdir(folder_path)
                       if f.lower().endswith('.xlsx') and not f.startswith('~')]

        if not excel_files:
            window.show_popup_message2("Information", "No Excel files found in the selected folder.")
            return

        # Prompt user to choose save location and filename
        default_filename = "Combined_KherveFitting_Files.xlsx"

        with wx.FileDialog(window, "Save combined KherveFitting file as...",
                           defaultDir=folder_path,
                           defaultFile=default_filename,
                           wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return  # User cancelled

            combined_file_path = fileDialog.GetPath()

            # Ensure .xlsx extension
            if not combined_file_path.lower().endswith('.xlsx'):
                combined_file_path += '.xlsx'

        # Sort files alphabetically
        excel_files.sort()

        # # Create single combined workbook
        # combined_file_path = os.path.join(folder_path, "Combined_KherveFitting_Files.xlsx")
        # combined_wb = openpyxl.Workbook()
        # combined_wb.remove(combined_wb.active)
        #
        # # Store sample names as dictionary for JSON
        # sample_names_dict = {}
        #
        # # Process each file as a separate sample
        # for sample_idx, excel_file in enumerate(excel_files):
        #     file_path = os.path.join(folder_path, excel_file)
        #
        #     # Remove file extension for cleaner sample names
        #     sample_name = os.path.splitext(excel_file)[0]
        #     sample_names_dict[str(sample_idx)] = sample_name
        #
        #     # Load the workbook
        #     wb = openpyxl.load_workbook(file_path)
        #     process_kfitting_file_with_sample_number(wb, combined_wb, sample_idx)

        # Create single combined workbook
        combined_file_path = os.path.join(folder_path, "Combined_KherveFitting_Files.xlsx")
        combined_wb = openpyxl.Workbook()
        combined_wb.remove(combined_wb.active)

        # Initialize tracking variables
        current_row = 0
        combined_sample_names = {}
        all_sheet_mappings = {}

        # Process each file with proper row numbering and sample names
        for sample_idx, excel_file in enumerate(excel_files):
            file_path = os.path.join(folder_path, excel_file)
            filename = os.path.splitext(excel_file)[0]

            # Load original JSON file to get existing sample names
            original_sample_names = {}
            json_file_path = os.path.splitext(file_path)[0] + '.json'
            if os.path.exists(json_file_path):
                try:
                    with open(json_file_path, 'r') as json_file:
                        individual_json_data = json.load(json_file)
                        if 'SampleNames' in individual_json_data:
                            original_sample_names = individual_json_data['SampleNames']
                except Exception as e:
                    print(f"Warning: Could not load sample names from {json_file_path}: {e}")

            # Process Excel file
            wb = openpyxl.load_workbook(file_path)

            # Modified function call to get sheet mappings
            current_row, sample_names_for_rows, sheet_mappings = process_kfitting_file_with_sample_number_with_mapping(
                wb, combined_wb, current_row, original_sample_names, filename)

            # Store sheet mappings for this file
            all_sheet_mappings.update(sheet_mappings)

            # Update combined sample names
            combined_sample_names.update(sample_names_for_rows)

            wb.close()

        # Save the combined file
        combined_wb.save(combined_file_path)

        window.show_popup_message2("Success",
                                   f"Combined {len(excel_files)} KherveFitting files into single Excel file.")

        # Open the combined file
        from libraries.FileMenu.Open import open_xlsx_file
        open_xlsx_file(window, combined_file_path)

        # Update SampleNames in window.Data
        if not hasattr(window, 'Data') or window.Data is None:
            from libraries.ConfigFile import Init_Measurement_Data
            window.Data = Init_Measurement_Data(window)

        # Add SampleNames as dictionary to the data structure
        window.Data['SampleNames'] = sample_names_dict

        # Save the updated JSON file
        json_file_path = os.path.splitext(combined_file_path)[0] + '.json'
        from libraries.FileMenu.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

    except Exception as e:
        window.show_popup_message2("Error", f"Error processing KherveFitting files: {str(e)}")



def process_kfitting_file_with_sample_number_with_mapping(wb, combined_wb, starting_row, original_sample_names,
                                                          filename):
    """Process KherveFitting Excel file and return mapping of original to new sheet names"""
    import re

    # Get all sheet names except Results Table and Experimental description
    sheets_to_process = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() not in ["results table", "experimental description"]:
            sheets_to_process.append(sheet_name)

    # Group sheets by their row number using simpler regex
    sheet_groups = {}
    for sheet_name in sheets_to_process:
        # Simple regex to extract trailing number
        match = re.search(r'(\d+)$', sheet_name)
        if match:
            row_number = int(match.group(1))
            base_name = sheet_name[:-len(match.group(1))]  # Remove the trailing number
        else:
            row_number = 0
            base_name = sheet_name

        if row_number not in sheet_groups:
            sheet_groups[row_number] = []
        sheet_groups[row_number].append((sheet_name, base_name))

    # Process each row group and return sample names mapping and sheet mappings
    current_row = starting_row
    sample_names_for_rows = {}
    sheet_mappings = {}  # original_name -> new_name

    for row_number in sorted(sheet_groups.keys()):
        # Determine sample name for this row
        if str(row_number) in original_sample_names and original_sample_names[str(row_number)].strip():
            # Use existing sample name from original file
            sample_name = original_sample_names[str(row_number)]
        else:
            # Use filename if no existing sample name
            sample_name = filename

        sample_names_for_rows[str(current_row)] = sample_name

        for sheet_name, base_name in sheet_groups[row_number]:
            sheet = wb[sheet_name]

            # Create new sheet name with current row number
            new_sheet_name = f"{base_name}{current_row}"

            # Track the mapping
            sheet_mappings[sheet_name] = new_sheet_name

            # Create new sheet in combined workbook
            new_sheet = combined_wb.create_sheet(new_sheet_name)

            # Copy headers from row 1
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=1, column=col).value
                if header_value:
                    new_sheet.cell(row=1, column=col, value=header_value)

            # Copy data starting from row 2
            for row in range(2, sheet.max_row + 1):
                row_has_data = False
                for col in range(1, min(sheet.max_column + 1, 5)):  # Copy first 4 columns typically
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_has_data = True
                    new_sheet.cell(row=row, column=col, value=cell_value)

                # Stop copying if we hit empty rows
                if not row_has_data:
                    break

        current_row += 1

    return current_row, sample_names_for_rows, sheet_mappings

def normalize_core_level_name(sheet_name):
    """
    Normalize core level names by removing numbers and common suffixes
    Examples: C1s1 -> C1s, Survey2 -> Survey, O1s_old -> O1s
    """
    import re

    # Remove trailing numbers
    name = re.sub(r'\d+$', '', sheet_name)

    # Remove common suffixes
    suffixes_to_remove = ['_old', '_new', '_copy', '_backup']
    for suffix in suffixes_to_remove:
        if name.endswith(suffix):
            name = name[:-len(suffix)]
            break

    # Handle Survey variations
    if 'survey' in name.lower():
        return 'Survey'

    # Handle Wide scan variations
    if 'wide' in name.lower():
        return 'Wide'

    return name


def is_numeric(value):
    """Check if value is numeric - .2f compatible."""
    if value is None or value == '':
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def import_generic_excel_file(window, file_path=None):
    """
    Interactive import for non-standard Excel files.
    Asks user to identify data type and reformats to kFitting format.
    """
    if not file_path:
        with wx.FileDialog(window, "Open Generic Excel File",
                           wildcard="Excel files (*.xlsx;*.xls)|*.xlsx;*.xls",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            file_path = fileDialog.GetPath()

    try:
        # Load workbook
        if file_path.lower().endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path)
        else:
            # Convert .xls to .xlsx
            from openpyxl import Workbook
            old_wb = xlrd.open_workbook(file_path)
            new_wb = Workbook()
            new_wb.remove(new_wb.active)

            for sheet_idx in range(old_wb.nsheets):
                old_sheet = old_wb.sheet_by_index(sheet_idx)
                new_sheet = new_wb.create_sheet(old_sheet.name)

                for row_idx in range(old_sheet.nrows):
                    for col_idx in range(old_sheet.ncols):
                        cell = old_sheet.cell(row_idx, col_idx)
                        new_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)

            temp_path = file_path.replace('.xls', '_temp.xlsx')
            new_wb.save(temp_path)
            old_wb.release_resources()
            wb = openpyxl.load_workbook(temp_path)
            file_path = temp_path

        # Analyze sheets
        sheet_info = analyze_excel_sheets(wb)

        # Show configuration dialog
        converted_wb = show_sheet_config_dialog(window, wb, sheet_info, file_path)

        if converted_wb:
            # Save converted file
            output_path = file_path.rsplit('.', 1)[0] + '_kFitting.xlsx'
            converted_wb.save(output_path)
            converted_wb.close()

            # Open the converted file
            open_xlsx_file(window, output_path)
            wx.MessageBox(f"File converted and opened successfully!\nSaved as: {output_path}",
                          "Success", wx.OK | wx.ICON_INFORMATION)

        wb.close()

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error importing file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def analyze_excel_sheets(wb):
    """Analyze sheets to detect data structure with .2f precision."""
    sheet_info = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        info = {
            'name': sheet_name,
            'data_start_row': 1,
            'has_headers': False,
            'num_cols': 0,
            'sample_data': []
        }

        # Check if first row has headers or numeric data
        cell1_row1 = sheet.cell(row=1, column=1).value
        cell2_row1 = sheet.cell(row=1, column=2).value

        # If first row is numeric, data starts at row 1 (no headers)
        if is_numeric(cell1_row1) and is_numeric(cell2_row1):
            info['data_start_row'] = 1
            info['has_headers'] = False
        else:
            # Check if row 1 has text headers
            if cell1_row1 and cell2_row1 and not is_numeric(cell1_row1):
                info['has_headers'] = True
                info['data_start_row'] = 2
            else:
                # Find first row with numeric data
                for row_idx in range(1, min(50, sheet.max_row + 1)):
                    cell1 = sheet.cell(row=row_idx, column=1).value
                    cell2 = sheet.cell(row=row_idx, column=2).value

                    if is_numeric(cell1) and is_numeric(cell2):
                        info['data_start_row'] = row_idx
                        info['has_headers'] = row_idx > 1
                        break

        # Count columns with data
        if info['data_start_row'] <= sheet.max_row:
            row = info['data_start_row']
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                val = sheet.cell(row=row, column=col_idx).value
                if val is not None:
                    info['num_cols'] += 1

        # Get sample data (show from data start row)
        for row_idx in range(info['data_start_row'], min(info['data_start_row'] + 5, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(info['num_cols'] + 1, 5)):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if is_numeric(val):
                    row_data.append(f"{float(val):.2f}")
                else:
                    row_data.append(str(val) if val else "")
            info['sample_data'].append(row_data)

        sheet_info[sheet_name] = info

    return sheet_info


class SheetConfigDialog(wx.Dialog):
    """Dialog for configuring sheet import settings."""

    def __init__(self, parent, wb, sheet_info):
        super().__init__(parent, title="Configure Excel Import",
                         size=(400, 550),
                         style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER)

        self.wb = wb
        self.sheet_info = sheet_info
        self.sheet_configs = {}

        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Instructions with more details
        instr_text = wx.StaticText(
            self,
            label="Configure each sheet below. Specify the data type and starting row.\n"
                  "Note that for XPS:\n"
                  "- Sheet Name must be in a single word  e.g., C1s, O1s, N1s, Fe2p, Au4f)\n"
        )
        instr_text.Wrap(780)
        main_sizer.Add(instr_text, 0, wx.ALL | wx.EXPAND, 10)

        # Notebook for sheets
        notebook = wx.Notebook(self)
        main_sizer.Add(notebook, 1, wx.ALL | wx.EXPAND, 5)

        # Create page for each sheet
        for sheet_name, info in sheet_info.items():
            panel = self.create_sheet_config_panel(notebook, sheet_name, info)
            notebook.AddPage(panel, sheet_name)

        # Buttons
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        ok_btn = wx.Button(self, wx.ID_OK, "Convert and Import")
        cancel_btn = wx.Button(self, wx.ID_CANCEL, "Cancel")
        btn_sizer.Add(ok_btn, 0, wx.ALL, 5)
        btn_sizer.Add(cancel_btn, 0, wx.ALL, 5)
        main_sizer.Add(btn_sizer, 0, wx.ALIGN_CENTER | wx.ALL, 10)

        self.SetSizer(main_sizer)
        self.Centre()

    def create_sheet_config_panel(self, parent, sheet_name, info):
        """Create configuration panel for a single sheet."""
        panel = wx.Panel(parent)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create grid sizer for controls (2 columns: label, control)
        grid = wx.FlexGridSizer(rows=5, cols=2, hgap=10, vgap=5)

        # Row 1: Data Type
        type_label = wx.StaticText(panel, label="Data Type:")
        type_choice = wx.Choice(panel, choices=[
            "XPS Core Level",
            "Raman Spectrum",
            "zzProfile",
            "Skip This Sheet"
        ], size=(150, -1))
        type_choice.SetSelection(0)

        grid.Add(type_label, 0, wx.ALIGN_CENTER_VERTICAL)
        grid.Add(type_choice, 0, wx.ALIGN_CENTER_VERTICAL)

        # Row 2: Sheet Name
        name_label = wx.StaticText(panel, label="Core level (e.g. C1s):")
        name_text = wx.TextCtrl(panel, value=sheet_name, size=(150, -1))

        grid.Add(name_label, 0, wx.ALIGN_CENTER_VERTICAL)
        grid.Add(name_text, 0, wx.ALIGN_CENTER_VERTICAL)

        # Row 3: Data Starts at Row
        row_label = wx.StaticText(panel, label="Data Starts at Row:")
        row_spin = wx.SpinCtrl(panel, value=str(info['data_start_row']),
                               min=1, max=1000, initial=info['data_start_row'],
                               size=(80, -1))

        grid.Add(row_label, 0, wx.ALIGN_CENTER_VERTICAL)
        grid.Add(row_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        # Row 4: X-Axis Column
        x_label = wx.StaticText(panel, label="X-Axis Column:")
        x_spin = wx.SpinCtrl(panel, value="1", min=1, max=50, initial=1,
                             size=(80, -1))

        grid.Add(x_label, 0, wx.ALIGN_CENTER_VERTICAL)
        grid.Add(x_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        # Row 5: Y-Axis Column
        y_label = wx.StaticText(panel, label="Y-Axis Column:")
        y_spin = wx.SpinCtrl(panel, value="2", min=1, max=50, initial=2,
                             size=(80, -1))

        grid.Add(y_label, 0, wx.ALIGN_CENTER_VERTICAL)
        grid.Add(y_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        main_sizer.Add(grid, 0, wx.ALL, 10)

        # Preview section
        preview_label = wx.StaticText(panel, label="Data Preview:")
        main_sizer.Add(preview_label, 0, wx.ALL, 5)

        # Always show 6 columns and up to 20 rows
        max_col_to_show = 6
        max_rows_to_show = 20

        preview_grid = wx.grid.Grid(panel)
        preview_grid.CreateGrid(max_rows_to_show, max_col_to_show)
        preview_grid.EnableEditing(False)
        preview_grid.SetColLabelSize(25)
        preview_grid.SetRowLabelSize(25)

        # Set size for all columns
        for col_idx in range(max_col_to_show):
            preview_grid.SetColSize(col_idx, 80)

        # Populate grid - read first 20 rows from source data
        sheet = self.wb[sheet_name]
        for row_idx in range(max_rows_to_show):
            actual_row = info['data_start_row'] + row_idx
            if actual_row > sheet.max_row:
                break
            for col_idx in range(max_col_to_show):
                val = sheet.cell(row=actual_row, column=col_idx + 1).value
                if val is not None:
                    if is_numeric(val):
                        preview_grid.SetCellValue(row_idx, col_idx, f"{float(val):.2f}")
                    else:
                        preview_grid.SetCellValue(row_idx, col_idx, str(val))
                else:
                    preview_grid.SetCellValue(row_idx, col_idx, "")

        main_sizer.Add(preview_grid, 1, wx.ALL | wx.EXPAND, 5)

        # Store controls
        self.sheet_configs[sheet_name] = {
            'type_choice': type_choice,
            'name_text': name_text,
            'row_spin': row_spin,
            'x_spin': x_spin,
            'y_spin': y_spin,
            'preview_grid': preview_grid
        }

        panel.SetSizer(main_sizer)  # Fixed: was 'sizer', now 'main_sizer'
        return panel


def show_sheet_config_dialog(window, wb, sheet_info, file_path):
    """Show dialog and process sheets based on user configuration."""
    from openpyxl import Workbook

    dialog = SheetConfigDialog(window, wb, sheet_info)

    if dialog.ShowModal() == wx.ID_OK:
        new_wb = Workbook()
        new_wb.remove(new_wb.active)

        # Track used sheet names
        used_sheet_names = []

        # Process each sheet
        for sheet_name in wb.sheetnames:
            config = dialog.sheet_configs.get(sheet_name)
            if not config:
                continue

            data_type = config['type_choice'].GetStringSelection()

            if data_type == "Skip This Sheet":
                continue

            new_sheet_name = config['name_text'].GetValue().strip()
            data_start_row = config['row_spin'].GetValue()
            x_col = config['x_spin'].GetValue()
            y_col = config['y_spin'].GetValue()

            # Validate and fix sheet name based on data type
            if data_type == "XPS Core Level":
                # Check if sheet name is generic or contains spaces
                if re.match(r'^Sheet\d+$', new_sheet_name, re.IGNORECASE) or ' ' in new_sheet_name:
                    # Prompt user for core level name
                    dlg = wx.TextEntryDialog(
                        window,
                        f"Sheet '{sheet_name}' needs a proper core level name.\n\n"
                        "Enter core level name as ONE WORD (no spaces):\n\n"
                        "Examples:\n"
                        "  C1s, O1s, N1s, Si2p, Fe2p, Au4f, Cu2p3, etc.\n\n"
                        "Format: ElementOrbital (e.g., C1s = Carbon 1s)",
                        "Core Level Name Required",
                        ""
                    )

                    while True:
                        if dlg.ShowModal() == wx.ID_OK:
                            new_sheet_name = dlg.GetValue().strip()

                            # Validate: no spaces, not empty
                            if not new_sheet_name:
                                wx.MessageBox("Core level name cannot be empty!",
                                              "Invalid Name", wx.OK | wx.ICON_WARNING)
                                continue
                            elif ' ' in new_sheet_name:
                                wx.MessageBox("Core level name must be ONE WORD (no spaces)!\n"
                                              "Examples: C1s, O1s, Fe2p",
                                              "Invalid Name", wx.OK | wx.ICON_WARNING)
                                continue
                            else:
                                break
                        else:
                            # User cancelled
                            dlg.Destroy()
                            new_sheet_name = None
                            break

                    dlg.Destroy()

                    if not new_sheet_name:
                        continue

            elif data_type == "Raman Spectrum":
                # Auto-name as Raman
                new_sheet_name = "Raman"
                # If already used, add number suffix
                if new_sheet_name in used_sheet_names:
                    counter = 1
                    while f"{new_sheet_name}{counter}" in used_sheet_names:
                        counter += 1
                    new_sheet_name = f"{new_sheet_name}{counter}"

            elif data_type == "zzProfile":
                # Auto-name as zzProfile
                new_sheet_name = "zzProfile"
                # If already used, add number suffix
                if new_sheet_name in used_sheet_names:
                    counter = 1
                    while f"{new_sheet_name}{counter}" in used_sheet_names:
                        counter += 1
                    new_sheet_name = f"{new_sheet_name}{counter}"

            # Check if name already used in this conversion
            if new_sheet_name in used_sheet_names:
                wx.MessageBox(f"Sheet name '{new_sheet_name}' already used. Skipping duplicate.",
                              "Duplicate Name", wx.OK | wx.ICON_WARNING)
                continue

            # Convert sheet
            converted_data = convert_sheet_to_kfitting(
                wb[sheet_name],
                new_sheet_name,
                data_start_row,
                x_col,
                y_col,
                data_type
            )

            if converted_data:
                new_wb.create_sheet(new_sheet_name)
                target_sheet = new_wb[new_sheet_name]
                used_sheet_names.append(new_sheet_name)

                # Add headers in first row based on data type
                if data_type == "XPS Core Level":
                    target_sheet.cell(row=1, column=1, value="Binding Energy")
                    target_sheet.cell(row=1, column=2, value="Raw Data")
                elif data_type == "Raman Spectrum":
                    target_sheet.cell(row=1, column=1, value="Raman Shift")
                    target_sheet.cell(row=1, column=2, value="Intensity")
                elif data_type == "zzProfile":
                    target_sheet.cell(row=1, column=1, value="Depth")
                    target_sheet.cell(row=1, column=2, value="Intensity")

                # Copy data with .2f formatting starting from row 2
                for row_idx, (x_val, y_val) in enumerate(converted_data, start=2):
                    target_sheet.cell(row=row_idx, column=1, value=f"{x_val:.2f}")
                    target_sheet.cell(row=row_idx, column=2, value=f"{y_val:.2f}")

                # Clear columns 3-24 for all rows including header
                for row_idx in range(1, len(converted_data) + 2):
                    for col_idx in range(3, 25):
                        target_sheet.cell(row=row_idx, column=col_idx, value=None)

        dialog.Destroy()
        return new_wb

    dialog.Destroy()
    return None


def convert_sheet_to_kfitting(sheet, new_name, start_row, x_col, y_col, data_type):
    """Convert sheet data to kFitting format with .2f precision."""
    data = []

    for row_idx in range(start_row, sheet.max_row + 1):
        x_val = sheet.cell(row=row_idx, column=x_col).value
        y_val = sheet.cell(row=row_idx, column=y_col).value

        if not is_numeric(x_val) or not is_numeric(y_val):
            continue

        try:
            x = float(x_val)
            y = float(y_val)
            data.append((x, y))
        except (ValueError, TypeError):
            continue

    return data