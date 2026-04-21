# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
VGD File Import Module for KherveFitting
Thermo Scientific / VG Scienta Binary Format (OLE2 Compound File)

VGD files store single XPS core level scans in Microsoft Compound File format.
Data is stored as Kinetic Energy and converted to Binding Energy using: BE = Source_Energy - KE
"""

import os
import struct
import numpy as np
import openpyxl
import wx
import json
import re

try:
    import olefile
except ImportError:
    olefile = None


def extract_core_level_name(filename):
    """
    Extract clean core level name from filename.

    Examples:
        'O1s_Scan.VGD' -> 'O1s'
        'O1s Scan.VGD' -> 'O1s'
        'Sr3d Scan.VGD' -> 'Sr3d'
        'Ni2p core level.VGD' -> 'Ni2p'
        'C1s_Region.VGD' -> 'C1s'
    """
    base_name = os.path.splitext(filename)[0]

    # Remove common suffixes (case insensitive)
    suffixes_to_remove = [
        '_Scan', '_scan', '_SCAN',
        ' Scan', ' scan', ' SCAN',
        '_Region', '_region', '_REGION',
        ' Region', ' region', ' REGION',
        '_core level', '_Core Level', '_Core level',
        ' core level', ' Core Level', ' Core level',
        '_spectrum', '_Spectrum', ' spectrum', ' Spectrum'
    ]

    result = base_name
    for suffix in suffixes_to_remove:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break

    # Also try regex to catch variations like "O1s scan" or "Ni2p_scan"
    # Pattern: element + orbital at the start, followed by separator and text
    match = re.match(r'^([A-Z][a-z]?\d+[spdfgh]\d*)[\s_]', result + ' ')
    if match:
        result = match.group(1)

    return result.strip()


def parse_vgd_file(file_path):
    """
    Parse a VGD file and return extracted data.
    Handles single-spectrum, multi-spectrum, and 3D area scan VGD files.

    For 3D area scans (ndims == 3), the data has dimensions:
        [n_x_pixels, n_y_pixels, n_ke_points]
    Each Y-pixel row is averaged across all X-pixels to produce one spectrum,
    yielding n_y_pixels spectra total (one sheet per Y row).

    Returns:
        dict with keys: intensities, ke_start, ke_step, num_points, source_energy,
                       txf_coeffs, pass_energy, work_fn, dwell_time, periods, metadata,
                       num_spectra, points_per_spectrum, is_area_scan
    """
    if olefile is None:
        raise ImportError("olefile library is required")

    ole = olefile.OleFileIO(file_path)

    # Get metadata from OLE properties
    metadata = ole.get_metadata()
    title = metadata.title.split('\x00')[0] if metadata.title else "Unknown"
    subject = metadata.subject.split('\x00')[0] if metadata.subject else ""
    author = metadata.author.split('\x00')[0] if metadata.author else ""
    create_time = str(metadata.create_time) if metadata.create_time else ""
    saved_time = str(metadata.last_saved_time) if metadata.last_saved_time else ""

    # Read intensity data from VGData stream
    vgdata = ole.openstream('VGData').read()
    total_points = len(vgdata) // 8

    # Check VGDataAxes for dimensionality info
    # offset 4  = number of dimensions (ndims)
    # offset 12 = dim1 - 1  (KE points)
    # offset 28 = dim2 - 1  (spectra count, or X pixels for area scan)
    # offset 44 = dim3 - 1  (Y pixels, only present for 3D area scans)
    data_axes = ole.openstream('VGDataAxes').read()

    num_spectra = 1
    points_per_spectrum = total_points
    is_area_scan = False

    if len(data_axes) >= 8:
        ndims = struct.unpack('<i', data_axes[4:8])[0]

        if ndims == 3 and len(data_axes) >= 56:
            # 3D area scan: [n_x_pixels, n_y_pixels, n_ke_points]
            n_ke  = struct.unpack('<i', data_axes[12:16])[0] + 1
            n_x   = struct.unpack('<i', data_axes[28:32])[0] + 1
            n_y   = struct.unpack('<i', data_axes[44:48])[0] + 1

            if n_ke * n_x * n_y == total_points:
                is_area_scan = True
                points_per_spectrum = n_ke
                num_spectra = n_y  # one spectrum per Y-pixel row (averaged over X)

        elif ndims <= 2 and len(data_axes) >= 32:
            # Standard 2D: [n_spectra, n_ke_points]
            dim1 = struct.unpack('<i', data_axes[12:16])[0] + 1
            dim2 = struct.unpack('<i', data_axes[28:32])[0] + 1
            if dim1 * dim2 == total_points and dim2 > 1:
                num_spectra = dim2
                points_per_spectrum = dim1

    # Parse all raw intensities
    all_intensities = []
    for i in range(total_points):
        val = struct.unpack('<d', vgdata[i * 8:i * 8 + 8])[0]
        all_intensities.append(val)

    # Reshape intensities
    if is_area_scan:
        # Reshape to [n_x, n_y, n_ke] then average over X to get [n_y, n_ke]
        import numpy as _np
        arr = _np.array(all_intensities).reshape(n_x, n_y, n_ke)
        intensities = [arr[:, y, :].mean(axis=0).tolist() for y in range(n_y)]
    elif num_spectra > 1:
        intensities_2d = []
        for s in range(num_spectra):
            start_idx = s * points_per_spectrum
            end_idx = start_idx + points_per_spectrum
            intensities_2d.append(all_intensities[start_idx:end_idx])
        intensities = intensities_2d  # List of lists
    else:
        intensities = all_intensities  # Single list

    # Extract KE range from VGSpaceAxes
    space_axes = ole.openstream('VGSpaceAxes').read()
    ke_start = struct.unpack('<d', space_axes[30:38])[0] if len(space_axes) >= 46 else None
    ke_step = struct.unpack('<d', space_axes[38:46])[0] if len(space_axes) >= 46 else None

    # Extract acquisition parameters from property stream
    prop_stream_name = '\x05Q5nw4m3lIjudbfwyAayojlptCa'
    prop_data = ole.openstream(prop_stream_name).read()

    # Extract X-ray source energy
    source_energy = None
    for i in range(0, len(prop_data) - 4, 4):
        val = struct.unpack('<f', prop_data[i:i + 4])[0]
        if 1480 < val < 1490:
            source_energy = val
            break
    if source_energy is None:
        source_energy = 1486.68

    # Extract Pass Energy and use its offset as reference for other parameters
    pass_energy = None
    pe_offset = None
    for i in range(0, len(prop_data) - 4, 4):
        val = struct.unpack('<f', prop_data[i:i + 4])[0]
        if val in [10.0, 20.0, 35.0, 50.0, 100.0, 160.0, 200.0]:
            pass_energy = val
            pe_offset = i
            break

    # Dwell Time is at PE offset + 8
    dwell_time = None
    if pe_offset is not None and pe_offset + 12 <= len(prop_data):
        dwell_time = struct.unpack('<f', prop_data[pe_offset + 8:pe_offset + 12])[0]
        if not (0.001 < dwell_time < 10.0):  # Sanity check
            dwell_time = None

    # Periods is at PE offset - 32
    periods = None
    if pe_offset is not None and pe_offset >= 32:
        periods = struct.unpack('<i', prop_data[pe_offset - 32:pe_offset - 28])[0]
        if not (1 <= periods <= 1000):  # Sanity check
            periods = None

    # Work Function is at PE offset + 16
    work_fn = None
    if pe_offset is not None and pe_offset + 20 <= len(prop_data):
        work_fn = struct.unpack('<f', prop_data[pe_offset + 16:pe_offset + 20])[0]
        if not (3.0 < work_fn < 6.0):  # Sanity check
            work_fn = None

    # Extract Transmission Function coefficients
    # TXF coefficients are stored with 8-byte spacing (float + 4-byte padding)
    # First coefficient 'a' is typically 4.0-4.5
    txf_coeffs = []
    for start in range(3100, min(3800, len(prop_data) - 32), 4):
        val = struct.unpack('<f', prop_data[start:start + 4])[0]
        if 4.0 < val < 4.5:
            vals = []
            valid = True
            for j in range(4):
                off = start + j * 8
                if off + 4 <= len(prop_data):
                    v = struct.unpack('<f', prop_data[off:off + 4])[0]
                    vals.append(v)
                else:
                    valid = False
                    break
            if valid and len(vals) == 4 and 0.5 < vals[1] < 1.0:
                txf_coeffs = vals
                break

    ole.close()

    return {
        'intensities': intensities,
        'ke_start': ke_start,
        'ke_step': ke_step,
        'num_points': points_per_spectrum,
        'total_points': total_points,
        'num_spectra': num_spectra,
        'is_area_scan': is_area_scan,
        'source_energy': source_energy,
        'txf_coeffs': txf_coeffs,
        'pass_energy': pass_energy,
        'work_fn': work_fn,
        'dwell_time': dwell_time,
        'periods': periods,
        'metadata': {
            'title': title,
            'subject': subject,
            'author': author,
            'create_time': create_time,
            'saved_time': saved_time
        }
    }


def calculate_vgd_data(parsed_data, spectrum_index=0):
    """
    Calculate BE values, transmission, and corrected data from parsed VGD data.

    Args:
        parsed_data: Output from parse_vgd_file
        spectrum_index: For multi-spectrum files, which spectrum to process (0-indexed)

    Returns:
        dict with keys: be_values, ke_values, transmission_values, corrected_data,
                       be_start, be_end, be_step
    """
    num_spectra = parsed_data.get('num_spectra', 1)

    # Get intensities for the specified spectrum
    if num_spectra > 1:
        intensities = parsed_data['intensities'][spectrum_index]
    else:
        intensities = parsed_data['intensities']

    ke_start = parsed_data['ke_start']
    ke_step = parsed_data['ke_step']
    num_points = parsed_data['num_points']
    source_energy = parsed_data['source_energy']
    txf_coeffs = parsed_data['txf_coeffs']

    # Calculate KE values (increasing)
    ke_values = [ke_start + i * ke_step for i in range(num_points)]

    # Convert to BE (decreasing - high to low, correct for XPS)
    be_values = [source_energy - ke for ke in ke_values]

    be_start = be_values[0]
    be_end = be_values[-1]
    be_step = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0

    # Calculate corrected data: Raw Data / (Periods * Dwell Time)
    dwell_time = parsed_data.get('dwell_time')
    periods = parsed_data.get('periods')
    txf_valid = True

    if dwell_time and periods and dwell_time > 0 and periods > 0:
        correction_factor = periods * dwell_time
        corrected_data = [intensity / correction_factor for intensity in intensities]
        transmission_values = [correction_factor] * num_points
    else:
        # No correction available
        corrected_data = list(intensities)
        transmission_values = [1.0] * num_points
        txf_valid = False

    return {
        'be_values': be_values,
        'ke_values': ke_values,
        'transmission_values': transmission_values,
        'corrected_data': corrected_data,
        'intensities': list(intensities),
        'be_start': be_start,
        'be_end': be_end,
        'be_step': be_step,
        'txf_valid': txf_valid
    }


def import_vgd_file(window, file_path=None, show_message=False):
    """
    Import Thermo VGD binary file (single or multi-spectrum).

    Args:
        window: Main KherveFitting window instance
        file_path: Path to .vgd file (if None, shows file dialog)
        show_message: If True, show success message box
    """
    if olefile is None:
        wx.MessageBox(
            "The 'olefile' library is required to import VGD files.\n\n"
            "Install it with:\n  pip install olefile",
            "Missing Library",
            wx.OK | wx.ICON_ERROR
        )
        return

    if file_path is None:
        with wx.FileDialog(window, "Open VGD file",
                           wildcard="VGD files (*.vgd;*.VGD)|*.vgd;*.VGD",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            file_path = fileDialog.GetPath()

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing VGD File", size=(400, 350))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 400) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        # Parse VGD file
        update_console(f"Opening: {os.path.basename(file_path)}")
        parsed_data = parse_vgd_file(file_path)

        num_spectra = parsed_data.get('num_spectra', 1)
        is_area_scan = parsed_data.get('is_area_scan', False)
        update_console(f"  Points per spectrum: {parsed_data['num_points']}")
        if is_area_scan:
            update_console(f"  Type: Area scan ({num_spectra} Y-pixel rows)")
        else:
            update_console(f"  Number of spectra: {num_spectra}")
        update_console(f"  Source Energy: {parsed_data['source_energy']:.2f} eV")
        if parsed_data['pass_energy']:
            update_console(f"  Pass Energy: {parsed_data['pass_energy']:.1f} eV")
        if parsed_data['dwell_time'] and parsed_data['periods']:
            update_console(f"  Dwell: {parsed_data['dwell_time']:.4f} s, Periods: {parsed_data['periods']}")

        # Extract core level name from filename
        # For area scans, prefer the file title from metadata
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        meta = parsed_data['metadata']
        if is_area_scan and meta['title'] and meta['title'] != "Unknown":
            core_level = meta['title'].replace(' ', '_')
        else:
            core_level = extract_core_level_name(os.path.basename(file_path))
        update_console(f"  Core level / label: {core_level}")

        source_energy = parsed_data['source_energy']
        source_label = "Al K-alpha Monochromated" if abs(source_energy - 1486.68) < 0.1 else f"X-ray {source_energy:.2f} eV"

        # Create Excel file with same name as VGD file
        output_dir = os.path.dirname(file_path)
        excel_path = os.path.join(output_dir, f"{base_name}.xlsx")

        update_console(f"\nCreating Excel file...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []

        # Process each spectrum
        for spectrum_idx in range(num_spectra):
            calc_data = calculate_vgd_data(parsed_data, spectrum_idx)

            # Create sheet name: Pt4f, Pt4f1, Pt4f2, etc. (first has no number)
            if spectrum_idx == 0:
                sheet_name = f"{core_level}"
            else:
                sheet_name = f"{core_level}{spectrum_idx}"
            sheet_names.append(sheet_name)

            # Update console for all sheets
            update_console(f"  Processing: {sheet_name}")

            ws = wb.create_sheet(sheet_name)

            # Write headers
            ws.cell(row=1, column=1, value='BE')
            ws.cell(row=1, column=2, value='Corrected Data')
            ws.cell(row=1, column=3, value='Raw Data')
            ws.cell(row=1, column=4, value='Transmission')

            # Write data
            for i, (be, corr, raw, trans) in enumerate(zip(
                    calc_data['be_values'],
                    calc_data['corrected_data'],
                    calc_data['intensities'],
                    calc_data['transmission_values']), start=2):
                ws.cell(row=i, column=1, value=round(be, 2))
                ws.cell(row=i, column=2, value=round(corr, 2))
                ws.cell(row=i, column=3, value=round(raw, 2))
                ws.cell(row=i, column=4, value=round(trans, 2))

            # Add experimental info in column 50 (AX)
            exp_col = 50
            ws.cell(row=1, column=exp_col, value="Experimental Description")

            exp_metadata = {
                'Sample ID': meta['subject'] if meta['subject'] else base_name,
                'Title': meta['title'],
                'Author': meta['author'],
                'Date Created': meta['create_time'].split()[0] if meta['create_time'] else '',
                'Time Created': meta['create_time'].split()[1] if meta['create_time'] and len(meta['create_time'].split()) > 1 else '',
                'Date Saved': meta['saved_time'].split()[0] if meta['saved_time'] else '',
                'Time Saved': meta['saved_time'].split()[1] if meta['saved_time'] and len(meta['saved_time'].split()) > 1 else '',
                'Technique': 'XPS',
                'Species & Transition': core_level,
                'Spectrum Index': str(spectrum_idx),
                'Total Spectra': str(num_spectra),
                'Area Scan': 'Yes' if is_area_scan else 'No',
                'Source Label': source_label,
                'Source Energy': f"{source_energy:.2f}",
                'Pass Energy': f"{parsed_data['pass_energy']:.2f}" if parsed_data['pass_energy'] else 'Unknown',
                'Work Function': f"{parsed_data['work_fn']:.2f}" if parsed_data['work_fn'] else 'Unknown',
                'Dwell Time': f"{parsed_data['dwell_time']:.4f}" if parsed_data['dwell_time'] else 'Unknown',
                'Periods': str(parsed_data['periods']) if parsed_data['periods'] else 'Unknown',
                'Number of Points': str(parsed_data['num_points']),
                'BE Start': f"{calc_data['be_start']:.2f}",
                'BE End': f"{calc_data['be_end']:.2f}",
                'BE Step': f"{abs(calc_data['be_step']):.4f}",
                'KE Start': f"{parsed_data['ke_start']:.2f}" if parsed_data['ke_start'] else 'Unknown',
                'KE Step': f"{parsed_data['ke_step']:.4f}" if parsed_data['ke_step'] else 'Unknown',
                'TXF Applied': 'Yes' if calc_data['txf_valid'] else 'No',
                'TXF Coefficients': ', '.join([f"{c:.6f}" for c in parsed_data['txf_coeffs']]) if parsed_data['txf_coeffs'] else 'N/A',
            }

            row = 2
            for key, value in exp_metadata.items():
                ws.cell(row=row, column=exp_col, value=key)
                ws.cell(row=row, column=exp_col + 1, value=str(value))
                row += 1

            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
            ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40

        update_console(f"Saving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        # Import into KherveFitting
        from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
        from libraries.FileMenu.Save import update_undo_redo_state, save_state, convert_to_serializable_and_round
        from libraries.Sheet_Operations import on_sheet_selected

        update_console(f"\nImporting to KherveFitting...")

        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Add all core levels/spectra
        for idx, sheet_name in enumerate(sheet_names):
            add_core_level_Data(window.Data, window, excel_path, sheet_name)
            update_console(f"  Imported: {sheet_name}")

        sample_name = meta['subject'] if meta['subject'] else base_name
        window.Data['SampleNames'] = {0: sample_name}

        update_console(f"Creating JSON file...")
        json_path = excel_path.replace('.xlsx', '.json')
        serializable_data = convert_to_serializable_and_round(window.Data)
        with open(json_path, 'w') as f:
            json.dump(serializable_data, f, indent=4)

        window.sheet_combobox.Clear()
        for sheet_name in sheet_names:
            window.sheet_combobox.Append(sheet_name)
        window.sheet_combobox.SetSelection(0)
        window.current_sheet = sheet_names[0]

        on_sheet_selected(window, None)
        save_state(window)

        # Refresh sheets
        update_console(f"Refreshing sheets...")
        from libraries.FileMenu.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected, update_console)

        # Refresh Sample Manager if it's open
        file_manager_was_open = False
        file_manager_position = None
        if hasattr(window, 'file_manager') and window.file_manager is not None:
            try:
                window.file_manager.GetSize()  # Check if window still exists
                file_manager_was_open = True
                file_manager_position = window.file_manager.GetPosition()
                window.file_manager.Close()
                window.file_manager = None
            except RuntimeError:
                window.file_manager = None

        update_console(f"\nImport complete!")
        if is_area_scan:
            update_console(f"  {core_level}: area scan, {num_spectra} Y-row spectra imported")
        elif num_spectra > 1:
            update_console(f"  {core_level}: {num_spectra} spectra imported")
        else:
            update_console(f"  {core_level}: 1 spectrum imported")

        # Close console after delay
        wx.CallLater(1500, console_frame.Close)

        # Restore file manager
        if file_manager_was_open:
            from libraries.ViewMenu.FileManager import FileManagerWindow
            window.file_manager = FileManagerWindow(window)
            if file_manager_position:
                window.file_manager.SetPosition(file_manager_position)
            window.file_manager.Show()

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing VGD file:\n\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def import_multiple_vgd_files(window, file_paths=None, show_message=False):
    """
    Import multiple VGD files into a single Excel workbook.

    Args:
        window: Main KherveFitting window instance
        file_paths: List of paths to .vgd files (if None, shows file dialog)
        show_message: If True, show success message box
    """
    if olefile is None:
        wx.MessageBox(
            "The 'olefile' library is required to import VGD files.\n\n"
            "Install it with:\n  pip install olefile",
            "Missing Library",
            wx.OK | wx.ICON_ERROR
        )
        return

    if file_paths is None:
        with wx.FileDialog(window, "Select VGD files",
                           wildcard="VGD files (*.vgd;*.VGD)|*.vgd;*.VGD",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            file_paths = fileDialog.GetPaths()

    if not file_paths:
        return

    # Ask for output filename first (before creating console)
    output_dir = os.path.dirname(file_paths[0])
    first_base = os.path.splitext(os.path.basename(file_paths[0]))[0]

    if len(file_paths) > 1:
        default_name = f"{first_base}_combined"
    else:
        default_name = first_base

    with wx.TextEntryDialog(window, "Enter name for the output file:",
                            "Output File Name", default_name) as dlg:
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        output_name = dlg.GetValue().strip()
        if not output_name:
            output_name = default_name

    excel_path = os.path.join(output_dir, f"{output_name}.xlsx")

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Importing VGD Files", size=(400, 350),
                             style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 400) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console(f"Importing {len(file_paths)} VGD file(s)...")
        update_console(f"Output file: {output_name}.xlsx")

        # output_dir = os.path.dirname(file_paths[0])
        #
        # first_base = os.path.splitext(os.path.basename(file_paths[0]))[0]
        #
        # if len(file_paths) > 1:
        #     default_name = f"{first_base}_combined"
        # else:
        #     default_name = first_base
        #
        # with wx.TextEntryDialog(window, "Enter name for the output file:",
        #                         "Output File Name", default_name) as dlg:
        #     if dlg.ShowModal() == wx.ID_CANCEL:
        #         return
        #     output_name = dlg.GetValue().strip()
        #     if not output_name:
        #         output_name = default_name
        #
        # excel_path = os.path.join(output_dir, f"{output_name}.xlsx")
        #
        # update_console(f"Importing {len(file_paths)} VGD file(s)...")
        # update_console(f"Output file: {output_name}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_names = []
        first_sample_name = None

        for file_idx, file_path in enumerate(file_paths):
            try:
                update_console(f"\nOpening ({file_idx + 1}/{len(file_paths)}): {os.path.basename(file_path)}")
                parsed_data = parse_vgd_file(file_path)

                num_spectra = parsed_data.get('num_spectra', 1)
                is_area_scan = parsed_data.get('is_area_scan', False)
                if is_area_scan:
                    update_console(f"  Type: Area scan ({num_spectra} Y-pixel rows)")
                else:
                    update_console(f"  Spectra in file: {num_spectra}")

                base_name = os.path.splitext(os.path.basename(file_path))[0]
                meta = parsed_data['metadata']
                if is_area_scan and meta['title'] and meta['title'] != "Unknown":
                    core_level = meta['title'].replace(' ', '_')
                else:
                    core_level = extract_core_level_name(os.path.basename(file_path))

                source_energy = parsed_data['source_energy']
                source_label = "Al K-alpha Monochromated" if abs(source_energy - 1486.68) < 0.1 else f"X-ray {source_energy:.2f} eV"

                if first_sample_name is None:
                    first_sample_name = meta['subject'] if meta['subject'] else base_name

                # Process each spectrum in the file
                for spectrum_idx in range(num_spectra):
                    calc_data = calculate_vgd_data(parsed_data, spectrum_idx)

                    # Create unique sheet name (first has no number, rest have 1, 2, 3...)
                    if f"{core_level}" not in sheet_names:
                        sheet_name = f"{core_level}"
                    else:
                        counter = 1
                        sheet_name = f"{core_level}{counter}"
                        while sheet_name in sheet_names:
                            counter += 1
                            sheet_name = f"{core_level}{counter}"
                    sheet_names.append(sheet_name)

                    update_console(f"  Processing: {sheet_name}")

                    ws = wb.create_sheet(sheet_name)

                    # Write headers
                    ws.cell(row=1, column=1, value='BE')
                    ws.cell(row=1, column=2, value='Corrected Data')
                    ws.cell(row=1, column=3, value='Raw Data')
                    ws.cell(row=1, column=4, value='Transmission')

                    # Write data
                    for i, (be, corr, raw, trans) in enumerate(zip(
                            calc_data['be_values'],
                            calc_data['corrected_data'],
                            calc_data['intensities'],
                            calc_data['transmission_values']), start=2):
                        ws.cell(row=i, column=1, value=round(be, 2))
                        ws.cell(row=i, column=2, value=round(corr, 2))
                        ws.cell(row=i, column=3, value=round(raw, 2))
                        ws.cell(row=i, column=4, value=round(trans, 2))

                    # Add experimental info
                    exp_col = 50
                    ws.cell(row=1, column=exp_col, value="Experimental Description")

                    exp_metadata = {
                        'Sample ID': meta['subject'] if meta['subject'] else base_name,
                        'Title': meta['title'],
                        'Author': meta['author'],
                        'Date Created': meta['create_time'].split()[0] if meta['create_time'] else '',
                        'Time Created': meta['create_time'].split()[1] if meta['create_time'] and len(meta['create_time'].split()) > 1 else '',
                        'Date Saved': meta['saved_time'].split()[0] if meta['saved_time'] else '',
                        'Time Saved': meta['saved_time'].split()[1] if meta['saved_time'] and len(meta['saved_time'].split()) > 1 else '',
                        'Technique': 'XPS',
                        'Species & Transition': core_level,
                        'Spectrum Index': str(spectrum_idx),
                        'Total Spectra': str(num_spectra),
                        'Area Scan': 'Yes' if is_area_scan else 'No',
                        'Source Label': source_label,
                        'Source Energy': f"{source_energy:.2f}",
                        'Pass Energy': f"{parsed_data['pass_energy']:.2f}" if parsed_data['pass_energy'] else 'Unknown',
                        'Work Function': f"{parsed_data['work_fn']:.2f}" if parsed_data['work_fn'] else 'Unknown',
                        'Dwell Time': f"{parsed_data['dwell_time']:.4f}" if parsed_data['dwell_time'] else 'Unknown',
                        'Periods': str(parsed_data['periods']) if parsed_data['periods'] else 'Unknown',
                        'Number of Points': str(parsed_data['num_points']),
                        'BE Start': f"{calc_data['be_start']:.2f}",
                        'BE End': f"{calc_data['be_end']:.2f}",
                        'BE Step': f"{abs(calc_data['be_step']):.4f}",
                        'KE Start': f"{parsed_data['ke_start']:.2f}" if parsed_data['ke_start'] else 'Unknown',
                        'KE Step': f"{parsed_data['ke_step']:.4f}" if parsed_data['ke_step'] else 'Unknown',
                        'TXF Applied': 'Yes' if calc_data['txf_valid'] else 'No',
                        'TXF Coefficients': ', '.join([f"{c:.6f}" for c in parsed_data['txf_coeffs']]) if parsed_data['txf_coeffs'] else 'N/A',
                    }

                    row = 2
                    for key, value in exp_metadata.items():
                        ws.cell(row=row, column=exp_col, value=key)
                        ws.cell(row=row, column=exp_col + 1, value=str(value))
                        row += 1

                    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
                    ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40

            except Exception as e:
                update_console(f"  Error: {e}")
                continue

        if not sheet_names:
            update_console("No VGD files could be processed")
            wx.CallLater(2000, console_frame.Close)
            wx.MessageBox("No VGD files could be processed", "Error", wx.OK | wx.ICON_ERROR)
            return

        update_console(f"\nSaving: {os.path.basename(excel_path)}")
        wb.save(excel_path)
        update_console(f"  Created {len(sheet_names)} sheet(s)")

        # Import into KherveFitting
        from libraries.ConfigFile import Init_Measurement_Data, add_core_level_Data
        from libraries.FileMenu.Save import update_undo_redo_state, save_state, convert_to_serializable_and_round
        from libraries.Sheet_Operations import on_sheet_selected

        update_console(f"\nImporting to KherveFitting...")

        window.Data = Init_Measurement_Data(window)
        window.Data['FilePath'] = excel_path

        window.history = []
        window.redo_stack = []
        update_undo_redo_state(window)

        window.results_grid.ClearGrid()
        if window.results_grid.GetNumberRows() > 0:
            window.results_grid.DeleteRows(0, window.results_grid.GetNumberRows())

        # Add all core levels
        for idx, sheet_name in enumerate(sheet_names):
            add_core_level_Data(window.Data, window, excel_path, sheet_name)
            update_console(f"  Imported: {sheet_name}")

        window.Data['SampleNames'] = {0: first_sample_name}

        update_console(f"Creating JSON file...")
        json_path = excel_path.replace('.xlsx', '.json')
        serializable_data = convert_to_serializable_and_round(window.Data)
        with open(json_path, 'w') as f:
            json.dump(serializable_data, f, indent=4)

        # Update UI
        window.sheet_combobox.Clear()
        for sheet_name in sheet_names:
            window.sheet_combobox.Append(sheet_name)
        window.sheet_combobox.SetSelection(0)
        window.current_sheet = sheet_names[0]

        on_sheet_selected(window, None)
        save_state(window)

        # Refresh sheets
        update_console(f"Refreshing sheets...")
        from libraries.FileMenu.Save import refresh_sheets
        refresh_sheets(window, on_sheet_selected, update_console)

        # Refresh Sample Manager if it's open
        file_manager_was_open = False
        file_manager_position = None
        if hasattr(window, 'file_manager') and window.file_manager is not None:
            try:
                window.file_manager.GetSize()  # Check if window still exists
                file_manager_was_open = True
                file_manager_position = window.file_manager.GetPosition()
                window.file_manager.Close()
                window.file_manager = None
            except RuntimeError:
                window.file_manager = None

        update_console(f"\nImport complete!")
        update_console(f"  {len(sheet_names)} core level(s) imported")

        # Close console after delay
        wx.CallLater(1500, console_frame.Close)

        # Restore file manager
        if file_manager_was_open:
            from libraries.ViewMenu.FileManager import FileManagerWindow
            window.file_manager = FileManagerWindow(window)
            if file_manager_position:
                window.file_manager.SetPosition(file_manager_position)
            window.file_manager.Show()

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"\nError: {str(e)}")
        wx.CallLater(3000, console_frame.Close)
        wx.MessageBox(f"Error importing VGD files:\n\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)