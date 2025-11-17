import wx
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.figure import Figure
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import NMF, PCA
import pandas as pd
from matplotlib.ticker import ScalarFormatter, FuncFormatter
from libraries.Peak_Functions import BackgroundCalculations
import re
import json
import os


class PCAnalysisWindow(wx.Frame):
    def __init__(self, parent):
        """
        Initialize PCA Analysis Window

        Args:
            parent (wx.Window): Parent window with Data structure
        """
        super().__init__(parent, title="Principal Component Analysis", size=(1080, 750), style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))

        # Store parent reference
        self.parent = parent

        # Analysis parameters
        self.n_components_find = 2
        self.n_components_use = 2
        self.offset_mode = "Minimum Value"
        self.nonneg_iterations = 1000
        self.nonneg_convergence = 0.0001

        # Analysis results
        self.nmf_model = None
        self.transformed_data = None
        self.components = None
        self.eigenvalues = None

        # Selected core levels data
        self.selected_core_levels = {}
        self.spectra_data = None
        self.common_be_grid = None
        self.sheet_names = []
        self.etch_times = []

        # Create main panel
        self.panel = wx.Panel(self)
        self.main_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Create control panel (LEFT side)
        self.create_control_panel()

        # Create plots panel (RIGHT side)
        self.create_plots_panel()

        # Set up layout
        self.panel.SetSizer(self.main_sizer)
        self.panel.Layout()

        # Center on parent and show
        self.CenterOnParent()
        self.Show()

        # Populate core level list after window is shown
        wx.CallAfter(self.populate_core_level_list)

    def natural_sort_key(self, sheet_name):
        """
        Natural sorting key to handle numbered core levels correctly
        e.g., O1s1, O1s2, ..., O1s10, O1s11
        """
        # Split the name into parts
        parts = re.split(r'(\d+)', sheet_name)
        # Convert numeric parts to integers for proper sorting
        return [int(part) if part.isdigit() else part.lower() for part in parts]

    def create_control_panel(self):
        """Create the left control panel with core level selection"""
        control_panel = wx.Panel(self.panel, style=wx.BORDER_RAISED)
        control_panel.SetMinSize((200, -1))
        control_sizer = wx.BoxSizer(wx.VERTICAL)

        # # Title
        # title = wx.StaticText(control_panel, label="Principal Component Analysis")
        # title_font = title.GetFont()
        # title_font.PointSize += 2
        # title_font = title_font.Bold()
        # title.SetFont(title_font)
        # control_sizer.Add(title, 0, wx.ALL | wx.ALIGN_CENTER, 0)

        # Core Level Selection section
        selection_box = wx.StaticBox(control_panel, label="Select Core Levels:")
        selection_sizer = wx.StaticBoxSizer(selection_box, wx.VERTICAL)

        # Selection buttons
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        select_all_btn = wx.Button(control_panel, label="Select All", size=(100, -1))
        select_all_btn.Bind(wx.EVT_BUTTON, self.on_select_all)
        button_sizer.Add(select_all_btn, 0, wx.ALL, 2)

        unselect_all_btn = wx.Button(control_panel, label="Unselect All", size=(100, -1))
        unselect_all_btn.Bind(wx.EVT_BUTTON, self.on_unselect_all)
        button_sizer.Add(unselect_all_btn, 0, wx.ALL, 2)

        selection_sizer.Add(button_sizer, 0, wx.ALL, 0)

        # CheckListBox for core levels
        self.core_level_list = wx.CheckListBox(control_panel, size=(-1, 200))
        self.core_level_list.Bind(wx.EVT_CHECKLISTBOX, self.on_core_level_checked)
        self.core_level_list.Bind(wx.EVT_RIGHT_DOWN, self.on_list_right_click)
        selection_sizer.Add(self.core_level_list, 1, wx.EXPAND | wx.ALL, 0)

        # Info text
        self.info_text = wx.StaticText(control_panel, label="0 core levels selected")
        selection_sizer.Add(self.info_text, 0, wx.ALL, 2)

        control_sizer.Add(selection_sizer, 0, wx.EXPAND | wx.ALL, 0)

        # Offset options
        offset_box = wx.StaticBox(control_panel, label="Offset")
        offset_sizer = wx.StaticBoxSizer(offset_box, wx.VERTICAL)
        self.offset_min = wx.RadioButton(control_panel, label="Minimum Value", style=wx.RB_GROUP)
        self.offset_smart = wx.RadioButton(control_panel, label="Smart Background")
        self.offset_min.SetValue(True)
        # Bind radio buttons to replot
        self.offset_min.Bind(wx.EVT_RADIOBUTTON, self.on_offset_changed)
        self.offset_smart.Bind(wx.EVT_RADIOBUTTON, self.on_offset_changed)
        offset_sizer.Add(self.offset_min, 0, wx.ALL, 5)
        offset_sizer.Add(self.offset_smart, 0, wx.ALL, 5)
        control_sizer.Add(offset_sizer, 0, wx.EXPAND | wx.ALL, 0)

        # Normalization options
        norm_box = wx.StaticBox(control_panel, label="Normalization")
        norm_sizer = wx.StaticBoxSizer(norm_box, wx.VERTICAL)
        self.norm_none = wx.RadioButton(control_panel, label="None", style=wx.RB_GROUP)
        self.norm_area = wx.RadioButton(control_panel, label="Area")
        self.norm_height = wx.RadioButton(control_panel, label="Max Height")
        self.norm_area.SetValue(True)
        # Bind radio buttons to replot
        self.norm_none.Bind(wx.EVT_RADIOBUTTON, self.on_norm_changed)
        self.norm_area.Bind(wx.EVT_RADIOBUTTON, self.on_norm_changed)
        self.norm_height.Bind(wx.EVT_RADIOBUTTON, self.on_norm_changed)
        # norm_sizer.Add(wx.StaticText(control_panel, label="Normalization:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        norm_sizer.Add(self.norm_none, 0, wx.ALL, 5)
        norm_sizer.Add(self.norm_area, 0, wx.ALL, 5)
        norm_sizer.Add(self.norm_height, 0, wx.ALL, 5)
        control_sizer.Add(norm_sizer, 0, wx.EXPAND | wx.ALL, 0)

        # Non-Negativity Fitting parameters
        nonneg_box = wx.StaticBox(control_panel, label="Non-Negativity Fitting")
        nonneg_sizer = wx.StaticBoxSizer(nonneg_box, wx.VERTICAL)

        # Iterations
        iter_sizer = wx.BoxSizer(wx.HORIZONTAL)
        iter_sizer.Add(wx.StaticText(control_panel, label="Iterations"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.iter_spin = wx.SpinCtrl(control_panel, value="1000", min=10, max=10000, size=(80, -1))
        iter_sizer.Add(self.iter_spin, 0)
        nonneg_sizer.Add(iter_sizer, 0, wx.ALL, 0)

        # Convergence
        conv_sizer = wx.BoxSizer(wx.HORIZONTAL)
        conv_sizer.Add(wx.StaticText(control_panel, label="Convergence"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.conv_spin = wx.SpinCtrlDouble(control_panel, value="0.0001", min=0.00001, max=1.0, inc=0.0001, size=(80, -1))
        self.conv_spin.SetDigits(5)
        conv_sizer.Add(self.conv_spin, 0)
        nonneg_sizer.Add(conv_sizer, 0, wx.ALL, 5)

        control_sizer.Add(nonneg_sizer, 0, wx.EXPAND | wx.ALL, 0)

        # Components section
        comp_box = wx.StaticBox(control_panel, label="Components")
        comp_sizer = wx.StaticBoxSizer(comp_box, wx.VERTICAL)

        # Find components
        find_sizer = wx.BoxSizer(wx.HORIZONTAL)
        find_sizer.Add(wx.StaticText(control_panel, label="Find"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.find_spin = wx.SpinCtrl(control_panel, value="2", min=1, max=28, size=(60, -1))
        find_sizer.Add(self.find_spin, 0, wx.RIGHT, 10)
        find_btn = wx.Button(control_panel, label="Analyse", size=(80, -1))
        find_btn.Bind(wx.EVT_BUTTON, self.on_analyse)
        find_sizer.Add(find_btn, 0)
        comp_sizer.Add(find_sizer, 0, wx.ALL, 0)

        # Result range text
        self.result_text = wx.StaticText(control_panel, label="1 to 28")
        comp_sizer.Add(self.result_text, 0, wx.ALL, 5)

        # Use components
        use_sizer = wx.BoxSizer(wx.HORIZONTAL)
        use_sizer.Add(wx.StaticText(control_panel, label="Use"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        self.use_spin = wx.SpinCtrl(control_panel, value="2", min=1, max=28, size=(60, -1))
        use_sizer.Add(self.use_spin, 0, wx.RIGHT, 10)
        redisplay_btn = wx.Button(control_panel, label="Re-Display", size=(80, -1))
        redisplay_btn.Bind(wx.EVT_BUTTON, self.on_redisplay)
        use_sizer.Add(redisplay_btn, 0)
        comp_sizer.Add(use_sizer, 0, wx.ALL, 5)

        control_sizer.Add(comp_sizer, 0, wx.EXPAND | wx.ALL, 0)

        # Bottom buttons
        control_sizer.AddStretchSpacer()
        # Bottom buttons
        bottom_sizer = wx.BoxSizer(wx.VERTICAL)

        # Add some spacing
        bottom_sizer.AddSpacer(10)

        # Create Core Levels button
        create_cl_btn = wx.Button(control_panel, label="Create Core Levels from PCA")
        create_cl_btn.Bind(wx.EVT_BUTTON, self.on_create_core_levels)
        bottom_sizer.Add(create_cl_btn, 0, wx.EXPAND | wx.ALL, 5)

        # Export and Close buttons
        buttons_sizer = wx.BoxSizer(wx.HORIZONTAL)

        export_btn = wx.Button(control_panel, label="Export Results")
        export_btn.Bind(wx.EVT_BUTTON, self.on_export_results)
        buttons_sizer.Add(export_btn, 1, wx.ALL, 5)

        cancel_btn = wx.Button(control_panel, label="Close")
        cancel_btn.Bind(wx.EVT_BUTTON, self.on_cancel)
        buttons_sizer.Add(cancel_btn, 1, wx.ALL, 5)

        bottom_sizer.Add(buttons_sizer, 0, wx.EXPAND)

        control_sizer.Add(bottom_sizer, 0, wx.EXPAND | wx.ALL, 5)

        control_panel.SetSizer(control_sizer)
        self.main_sizer.Add(control_panel, 0, wx.EXPAND | wx.ALL, 0)

    def create_plots_panel(self):
        """Create the right panel with 4 plots arranged in 2x2 grid"""
        plots_panel = wx.Panel(self.panel, style=wx.BORDER_RAISED)
        plots_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create figure with 4 subplots
        self.fig = Figure(figsize=(8, 7))
        self.canvas = FigureCanvas(plots_panel, -1, self.fig)

        # Create the 4 subplots
        self.ax_spectrum = self.fig.add_subplot(221)
        self.ax_eigenvalues = self.fig.add_subplot(222)
        self.ax_eigenvectors = self.fig.add_subplot(223)
        self.ax_profiles = self.fig.add_subplot(224)

        # Set titles
        self.ax_spectrum.set_title('Selected Spectra')
        self.ax_spectrum.set_xlabel('Binding Energy (eV)')
        self.ax_spectrum.set_ylabel('Counts / s')

        self.ax_eigenvalues.set_title('PCA Initial Eigen Values')
        self.ax_eigenvalues.set_xlabel('PCA Component')
        self.ax_eigenvalues.set_ylabel('Intensity')

        self.ax_eigenvectors.set_title('PCA Eigen Vectors Rotated NonNeg')
        self.ax_eigenvectors.set_xlabel('Binding Energy (eV)')
        self.ax_eigenvectors.set_ylabel('Counts / s')

        self.ax_profiles.set_title('PCA Profiles Rotated NonNeg')
        self.ax_profiles.set_xlabel('Etch Time (s)')
        self.ax_profiles.set_ylabel('Intensity')

        self.fig.tight_layout(pad=0.5, h_pad=0.5, w_pad=0.5)

        plots_sizer.Add(self.canvas, 1, wx.EXPAND | wx.ALL, 0)
        plots_panel.SetSizer(plots_sizer)

        self.main_sizer.Add(plots_panel, 1, wx.EXPAND | wx.ALL, 0)

    def on_norm_changed(self, event):
        """Handle normalization radio button change"""
        if self.spectra_data is not None and len(self.spectra_data) > 0:
            self.plot_all_spectra()

    def on_offset_changed(self, event):
        """Handle offset radio button change"""
        if self.spectra_data is not None and len(self.spectra_data) > 0:
            self.plot_all_spectra()

    def populate_core_level_list(self):
        """Populate the core level list from parent Data"""
        if not hasattr(self.parent, 'Data') or 'Core levels' not in self.parent.Data:
            return

        core_levels = []
        for sheet_name in self.parent.Data['Core levels'].keys():
            if 'B.E.' in self.parent.Data['Core levels'][sheet_name] and \
                    'Raw Data' in self.parent.Data['Core levels'][sheet_name]:
                core_levels.append(sheet_name)

        # Sort using natural sorting
        sorted_core_levels = sorted(core_levels, key=self.natural_sort_key)
        self.core_level_list.SetItems(sorted_core_levels)

    def on_select_all(self, event):
        """Select all core levels"""
        for i in range(self.core_level_list.GetCount()):
            self.core_level_list.Check(i, True)
        self.update_selection()

    def on_unselect_all(self, event):
        """Unselect all core levels"""
        for i in range(self.core_level_list.GetCount()):
            self.core_level_list.Check(i, False)
        self.update_selection()

    def on_list_right_click(self, event):
        """Show context menu for selecting core levels by element"""
        # Get all unique elements from the list
        elements = set()
        has_survey = False

        for i in range(self.core_level_list.GetCount()):
            name = self.core_level_list.GetString(i)

            # Check if it's a survey scan
            if 'survey' in name.lower() or 'surv' in name.lower():
                has_survey = True
                continue

            # Extract element (e.g., "C1s" -> "C1s", "O1s_0" -> "O1s")
            if '_' in name:
                element = name.split('_')[0]
            else:
                element = name
            # Get base element without number suffix
            match = re.match(r'([A-Za-z]+\d*[spdfg]*)', element)
            if match:
                elements.add(match.group(1))

        # Create context menu
        menu = wx.Menu()

        # Add survey option if any survey scans exist
        if has_survey:
            survey_item = menu.Append(wx.ID_ANY, "Select all Survey")
            self.Bind(wx.EVT_MENU, lambda evt: self.select_by_element("Survey"), survey_item)
            menu.AppendSeparator()

        for element in sorted(elements):
            item = menu.Append(wx.ID_ANY, f"Select all {element}")
            self.Bind(wx.EVT_MENU, lambda evt, el=element: self.select_by_element(el), item)

        menu.AppendSeparator()
        unselect_item = menu.Append(wx.ID_ANY, "Unselect All")
        self.Bind(wx.EVT_MENU, self.on_unselect_all, unselect_item)

        self.PopupMenu(menu)
        menu.Destroy()

    def select_by_element(self, element):
        """Select all core levels of a specific element"""
        element_lower = element.lower()

        for i in range(self.core_level_list.GetCount()):
            name = self.core_level_list.GetString(i)

            # Handle survey selection
            if element_lower == "survey":
                if 'survey' in name.lower() or 'surv' in name.lower():
                    self.core_level_list.Check(i, True)
            else:
                # Handle regular elements
                if name.startswith(element):
                    self.core_level_list.Check(i, True)

        self.update_selection()

    def on_core_level_checked(self, event):
        """Handle core level checkbox change"""
        self.update_selection()

    def update_selection(self):
        """Update selected core levels and replot"""
        self.selected_core_levels = {}
        checked_items = []

        for i in range(self.core_level_list.GetCount()):
            if self.core_level_list.IsChecked(i):
                sheet_name = self.core_level_list.GetString(i)
                checked_items.append(sheet_name)
                if sheet_name in self.parent.Data['Core levels']:
                    core_level = self.parent.Data['Core levels'][sheet_name]
                    self.selected_core_levels[sheet_name] = {
                        'B.E.': core_level['B.E.'],
                        'Raw Data': core_level['Raw Data']
                    }

        # Update info text
        self.info_text.SetLabel(f"{len(self.selected_core_levels)} core levels selected")

        # Prepare and plot data if we have selections
        if len(self.selected_core_levels) > 0:
            self.prepare_data()
        else:
            # Clear plots
            self.ax_spectrum.clear()
            self.ax_spectrum.set_title('Selected Spectra')
            self.ax_spectrum.set_xlabel('Binding Energy (eV)')
            self.ax_spectrum.set_ylabel('Counts / s')
            self.canvas.draw()

    def prepare_data(self):
        """Prepare data for PCA/NMF analysis"""
        be_ranges = []
        raw_intensities = []
        be_values_list = []
        self.sheet_names = []
        self.etch_times = []

        # Sort selected core levels naturally
        sorted_sheet_names = sorted(self.selected_core_levels.keys(), key=self.natural_sort_key)

        for sheet_name in sorted_sheet_names:
            core_level_data = self.selected_core_levels[sheet_name]
            try:
                be_values = np.array(core_level_data['B.E.'], dtype=float)
                raw_data = np.array(core_level_data['Raw Data'], dtype=float)

                # Store original BE values and intensities
                be_values_list.append(be_values)
                raw_intensities.append(raw_data)
                be_ranges.append((be_values.min(), be_values.max()))
                self.sheet_names.append(sheet_name)

                # Extract etch time/level from sheet name
                parts = sheet_name.split('_')
                if len(parts) > 1 and parts[-1].isdigit():
                    self.etch_times.append(int(parts[-1]) * 10)
                else:
                    self.etch_times.append(len(self.etch_times) * 10)

            except Exception as e:
                print(f"Error processing {sheet_name}: {e}")

        if not raw_intensities:
            return

        # Determine common binding energy range
        min_be = max(range[0] for range in be_ranges)
        max_be = min(range[1] for range in be_ranges)

        # Create common binding energy grid (High to Low BE)
        num_points = 1000
        self.common_be_grid = np.linspace(max_be, min_be, num_points)

        # Interpolate all data to the common grid
        interpolated_data = []
        for be_values, intensities in zip(be_values_list, raw_intensities):
            from scipy.interpolate import interp1d

            try:
                interp_func = interp1d(be_values, intensities,
                                       kind='linear', fill_value='extrapolate')
                interpolated_intensities = interp_func(self.common_be_grid)
                interpolated_data.append(interpolated_intensities)
            except Exception as e:
                print(f"Interpolation error: {e}")
                continue

        if not interpolated_data:
            return

        self.spectra_data = np.array(interpolated_data)

        # Plot all spectra
        self.plot_all_spectra()

    def plot_all_spectra(self):
        """Plot all selected spectra overlaid with normalization"""
        self.ax_spectrum.clear()

        # Get normalization mode
        if self.norm_none.GetValue():
            norm_mode = "None"
        elif self.norm_area.GetValue():
            norm_mode = "Area"
        else:
            norm_mode = "Max Height"

        # Get offset mode
        offset_mode = "Minimum Value" if self.offset_min.GetValue() else "Smart Background"

        # Copy data for plotting
        plot_data = np.abs(self.spectra_data.copy())

        # Apply offset (background subtraction)
        if offset_mode == "Minimum Value":
            plot_data = plot_data - np.min(plot_data, axis=1, keepdims=True)
        elif offset_mode == "Smart Background":
            # Apply smart background using the full spectrum for each
            for i in range(len(plot_data)):
                spectrum = plot_data[i]
                x_data = self.common_be_grid

                # Determine if background is ascending or descending
                if spectrum[0] > spectrum[-1]:
                    # Descending - use Shirley background
                    bg = BackgroundCalculations.calculate_shirley_background(
                        x_data, spectrum, start_offset=0.0, end_offset=0.0, num_points=5
                    )
                else:
                    # Ascending - use Linear background
                    bg = BackgroundCalculations.calculate_linear_background(
                        x_data, spectrum, start_offset=0.0, end_offset=0.0, num_points=5
                    )

                # Subtract background
                plot_data[i] = spectrum - bg

                # Ensure no negative values after subtraction
                plot_data[i] = np.maximum(plot_data[i], 0)

        # Ensure all values are non-negative
        plot_data = np.maximum(plot_data, 0)

        # Apply normalization to plot data
        if norm_mode == "Area":
            # Set lowest BE to 0, then normalize by area
            for i in range(len(plot_data)):
                plot_data[i] = plot_data[i] - plot_data[i][-1]
                plot_data[i] = np.maximum(plot_data[i], 0)
                area = np.trapz(plot_data[i], self.common_be_grid)
                if area > 0:
                    plot_data[i] = plot_data[i] / area
        elif norm_mode == "Max Height":
            # Set lowest BE to 0, then normalize by maximum
            for i in range(len(plot_data)):
                plot_data[i] = plot_data[i] - plot_data[i][-1]
                plot_data[i] = np.maximum(plot_data[i], 0)
                max_val = np.max(plot_data[i])
                if max_val > 0:
                    plot_data[i] = plot_data[i] / max_val

        colors = plt.cm.Greens(np.linspace(0.4, 0.9, len(plot_data)))

        for idx, (spectrum, sheet_name) in enumerate(zip(plot_data, self.sheet_names)):
            self.ax_spectrum.plot(self.common_be_grid, spectrum,
                                  label=f'{sheet_name}',
                                  color=colors[idx],
                                  linewidth=1)

        # Update title to show offset and normalization mode
        title = f'{len(plot_data)} Spectra Selected'
        if offset_mode == "Smart Background":
            title += ' (Smart Bkg)'
        if norm_mode != "None":
            title += f' (Norm: {norm_mode})'
        self.ax_spectrum.set_title(title)

        self.ax_spectrum.set_xlabel('Binding Energy (eV)')
        self.ax_spectrum.set_ylabel('Counts / s')

        # Invert x-axis to show High BE to Low BE
        self.ax_spectrum.invert_xaxis()

        # if len(plot_data) > 1:
        #     self.ax_spectrum.legend(fontsize=8, loc='best')

        # Ensure scientific notation is maintained
        formatter = ScalarFormatter(useMathText=True)
        formatter.set_scientific(True)
        formatter.set_powerlimits((-2, 2))
        self.ax_spectrum.yaxis.set_major_formatter(formatter)

        self.fig.tight_layout(pad=0.5, h_pad=0.5, w_pad=0.5)
        self.canvas.draw()

    def on_analyse(self, event):
        """Perform Non-Negative PCA analysis"""
        if self.spectra_data is None or len(self.spectra_data) == 0:
            wx.MessageBox("Please select core levels first", "No Data", wx.OK | wx.ICON_WARNING)
            return

        n_components = self.find_spin.GetValue()
        self.n_components_find = n_components

        # Get parameters
        self.offset_mode = "Minimum Value" if self.offset_min.GetValue() else "Smart Background"
        self.nonneg_iterations = self.iter_spin.GetValue()
        self.nonneg_convergence = self.conv_spin.GetValue()

        # Get normalization mode
        if self.norm_none.GetValue():
            norm_mode = "None"
        elif self.norm_area.GetValue():
            norm_mode = "Area"
        else:
            norm_mode = "Max Height"

        # Preprocess data - make a copy to avoid modifying original
        data_for_analysis = np.abs(self.spectra_data.copy())

        # Apply offset
        if self.offset_mode == "Minimum Value":
            data_for_analysis = data_for_analysis - np.min(data_for_analysis, axis=1, keepdims=True)
        elif self.offset_mode == "Smart Background":
            # Apply smart background using the full spectrum for each
            for i in range(len(data_for_analysis)):
                spectrum = data_for_analysis[i]
                x_data = self.common_be_grid

                # Determine if background is ascending or descending
                # If first point > last point: descending, use Shirley
                # If first point <= last point: ascending, use Linear
                if spectrum[0] > spectrum[-1]:
                    # Descending - use Shirley background
                    bg = BackgroundCalculations.calculate_shirley_background(
                        x_data, spectrum, start_offset=0.0, end_offset=0.0, num_points=5
                    )
                else:
                    # Ascending - use Linear background
                    bg = BackgroundCalculations.calculate_linear_background(
                        x_data, spectrum, start_offset=0.0, end_offset=0.0, num_points=5
                    )

                # Subtract background
                data_for_analysis[i] = spectrum - bg

                # Ensure no negative values after subtraction
                data_for_analysis[i] = np.maximum(data_for_analysis[i], 0)

        # Final safety check: ensure all values are non-negative
        data_for_analysis = np.maximum(data_for_analysis, 0)

        # Apply normalization
        if norm_mode == "Area":
            # Set lowest BE to 0, then normalize by area
            for i in range(len(data_for_analysis)):
                data_for_analysis[i] = data_for_analysis[i] - data_for_analysis[i][-1]
                data_for_analysis[i] = np.maximum(data_for_analysis[i], 0)
                area = np.trapz(data_for_analysis[i], self.common_be_grid)
                if area > 0:
                    data_for_analysis[i] = data_for_analysis[i] / area
        elif norm_mode == "Max Height":
            # Set lowest BE to 0, then normalize by maximum
            for i in range(len(data_for_analysis)):
                data_for_analysis[i] = data_for_analysis[i] - data_for_analysis[i][-1]
                data_for_analysis[i] = np.maximum(data_for_analysis[i], 0)
                max_val = np.max(data_for_analysis[i])
                if max_val > 0:
                    data_for_analysis[i] = data_for_analysis[i] / max_val

        # Add small epsilon to avoid exact zeros (helps NMF stability)
        data_for_analysis = data_for_analysis + 1e-10

        # Use Non-Negative Matrix Factorization
        try:
            self.nmf_model = NMF(n_components=n_components, init='random', random_state=0,
                                 max_iter=self.nonneg_iterations, tol=self.nonneg_convergence)
            self.transformed_data = self.nmf_model.fit_transform(data_for_analysis)
            self.components = self.nmf_model.components_
            self.eigenvalues = np.sum(self.transformed_data, axis=0)
        except Exception as e:
            wx.MessageBox(f"NMF analysis failed: {e}\n\nTry using 'Minimum Value' offset instead.",
                          "Analysis Error", wx.OK | wx.ICON_ERROR)
            return

        # Plot Eigen Values
        self.ax_eigenvalues.clear()
        colors = plt.cm.Greens(np.linspace(0.4, 0.9, n_components))
        self.ax_eigenvalues.bar(range(1, n_components + 1), self.eigenvalues, color=colors)
        self.ax_eigenvalues.set_title('PCA Initial Eigen Values')
        self.ax_eigenvalues.set_xlabel('PCA Component')
        self.ax_eigenvalues.set_ylabel('Intensity')

        # Plot Eigenvectors
        self.ax_eigenvectors.clear()
        colors = plt.cm.Greens(np.linspace(0.4, 0.9, n_components))
        for i in range(n_components):
            self.ax_eigenvectors.plot(
                self.common_be_grid,
                self.components[i],
                label=f'PCA {i + 1}',
                color=colors[i],
                linewidth=1
            )
        self.ax_eigenvectors.set_title('PCA Eigen Vectors Rotated NonNeg')
        self.ax_eigenvectors.set_xlabel('Binding Energy (eV)')
        self.ax_eigenvectors.set_ylabel('Counts / s')

        # Ensure scientific notation on eigenvectors
        formatter = ScalarFormatter(useMathText=True)
        formatter.set_scientific(True)
        formatter.set_powerlimits((-2, 2))
        self.ax_eigenvectors.yaxis.set_major_formatter(formatter)

        # Invert x-axis for eigenvectors too
        self.ax_eigenvectors.invert_xaxis()
        self.ax_eigenvectors.legend()

        # Plot profiles
        self.ax_profiles.clear()
        plot_numbers = list(range(len(self.etch_times)))
        colors = plt.cm.Greens(np.linspace(0.4, 0.9, n_components))
        for i in range(n_components):
            self.ax_profiles.plot(
                plot_numbers,
                self.transformed_data[:, i],
                marker='o',
                label=f'PCA {i + 1}',
                color=colors[i],
                linewidth=1,
                markersize=3
            )
        self.ax_profiles.set_title('PCA Profiles Rotated NonNeg')
        self.ax_profiles.set_xlabel('Plot Number')
        self.ax_profiles.set_ylabel('Intensity')
        self.ax_profiles.legend()

        self.fig.tight_layout(pad=0.5, h_pad=0.5, w_pad=0.5)
        self.canvas.draw()

        # Update result range
        self.result_text.SetLabel(f"1 to {n_components}")


    def on_redisplay(self, event):
        """Re-display with different number of components"""
        n_use = self.use_spin.GetValue()

        if self.components is None:
            wx.MessageBox("Please run Analyse first", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Re-plot with selected number of components
        self.ax_eigenvectors.clear()
        colors = plt.cm.tab10(np.linspace(0, 1, n_use))
        for i in range(min(n_use, len(self.components))):
            self.ax_eigenvectors.plot(
                self.common_be_grid,
                self.components[i],
                label=f'PCA {i + 1}',
                color=colors[i],
                linewidth=1
            )
        self.ax_eigenvectors.set_title('PCA Eigen Vectors Rotated NonNeg')
        self.ax_eigenvectors.set_xlabel('Binding Energy (eV)')
        self.ax_eigenvectors.set_ylabel('Counts / s')
        # Invert x-axis
        self.ax_eigenvectors.invert_xaxis()
        self.ax_eigenvectors.legend()

        self.ax_profiles.clear()
        plot_numbers = list(range(len(self.etch_times)))
        for i in range(min(n_use, self.transformed_data.shape[1])):
            self.ax_profiles.plot(
                plot_numbers,
                self.transformed_data[:, i],
                marker='o',
                label=f'PCA {i + 1}',
                color=colors[i],
                linewidth=1,
                markersize=3
            )
        self.ax_profiles.set_title('PCA Profiles Rotated NonNeg')
        self.ax_profiles.set_xlabel('Plot Number')
        self.ax_profiles.set_ylabel('Intensity')
        self.ax_profiles.legend()

        # Ensure scientific notation
        formatter = ScalarFormatter(useMathText=True)
        formatter.set_scientific(True)
        formatter.set_powerlimits((-2, 2))
        self.ax_eigenvectors.yaxis.set_major_formatter(formatter)

        self.fig.tight_layout(pad=0.5, h_pad=0.5, w_pad=0.5)
        self.canvas.draw()

    def on_export_results(self, event):
        """Export PCA results to JSON or Excel"""
        if self.components is None:
            wx.MessageBox("Please run Analyse first", "No Results", wx.OK | wx.ICON_WARNING)
            return

        # Create dialog to choose format
        dlg = wx.SingleChoiceDialog(
            self,
            'Choose export format:',
            'Export PCA Results',
            ['Single Entity (JSON)', 'Excel (Multiple Sheets)']
        )

        if dlg.ShowModal() == wx.ID_OK:
            choice = dlg.GetSelection()
            dlg.Destroy()

            if choice == 0:  # Single Entity
                self.export_as_single_entity()
            elif choice == 1:  # Full JSON
                self.export_to_json()
            else:  # Excel export
                self.export_to_excel()
        else:
            dlg.Destroy()

    def export_to_excel(self):
        """Export PCA results to Excel (original method)"""
        wildcard = "Excel Files (*.xlsx)|*.xlsx"
        dlg = wx.FileDialog(self, "Export PCA Results to Excel", wildcard=wildcard,
                            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)

        if dlg.ShowModal() == wx.ID_OK:
            filepath = dlg.GetPath()

            try:
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    # Write binding energy grid and components
                    components_data = {'Binding Energy (eV)': self.common_be_grid}
                    for i in range(len(self.components)):
                        components_data[f'PCA {i + 1}'] = self.components[i]
                    df_components = pd.DataFrame(components_data)
                    df_components.to_excel(writer, sheet_name='PCA Components', index=False)

                    # Write profiles
                    plot_numbers = list(range(len(self.etch_times)))
                    profiles_data = {'Plot Number': plot_numbers, 'Sheet Name': self.sheet_names}
                    for i in range(self.transformed_data.shape[1]):
                        profiles_data[f'PCA {i + 1}'] = self.transformed_data[:, i]
                    df_profiles = pd.DataFrame(profiles_data)
                    df_profiles.to_excel(writer, sheet_name='PCA Profiles', index=False)

                    # Write eigenvalues
                    df_eigen = pd.DataFrame({
                        'Component': range(1, len(self.eigenvalues) + 1),
                        'Eigenvalue': self.eigenvalues
                    })
                    df_eigen.to_excel(writer, sheet_name='Eigenvalues', index=False)

                wx.MessageBox("PCA results exported to Excel successfully!", "Export Complete", wx.OK)
            except Exception as e:
                wx.MessageBox(f"Error exporting to Excel: {e}", "Export Error", wx.OK | wx.ICON_ERROR)

        dlg.Destroy()

    def export_to_json(self):
        """Export full PCA analysis as JSON file"""
        import json

        wildcard = "JSON Files (*.json)|*.json"
        dlg = wx.FileDialog(self, "Export Full PCA Analysis", wildcard=wildcard,
                            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)

        if dlg.ShowModal() == wx.ID_OK:
            filepath = dlg.GetPath()

            try:
                # Helper function to convert numpy to list
                def convert_numpy_to_list(data):
                    if isinstance(data, np.ndarray):
                        return data.tolist()
                    return data

                # Create comprehensive data structure
                pca_data = {
                    'analysis_parameters': {
                        'n_components_find': self.n_components_find,
                        'n_components_use': self.n_components_use,
                        'offset_mode': self.offset_mode,
                        'iterations': self.nonneg_iterations,
                        'convergence': self.nonneg_convergence,
                        'n_iterations_actual': int(self.nmf_model.n_iter_) if self.nmf_model else None,
                        'reconstruction_error': float(self.nmf_model.reconstruction_err_) if self.nmf_model else None
                    },
                    'input_data': {
                        'sheet_names': self.sheet_names,
                        'plot_numbers': list(range(len(self.sheet_names))),
                        'binding_energy_grid': convert_numpy_to_list(self.common_be_grid),
                        'spectra_data': convert_numpy_to_list(self.spectra_data)
                    },
                    'results': {
                        'eigenvalues': convert_numpy_to_list(self.eigenvalues),
                        'components': {
                            'binding_energy': convert_numpy_to_list(self.common_be_grid),
                            'pca_components': [
                                {
                                    'component_id': i + 1,
                                    'intensity': convert_numpy_to_list(self.components[i])
                                }
                                for i in range(len(self.components))
                            ]
                        },
                        'profiles': {
                            'plot_numbers': list(range(len(self.sheet_names))),
                            'sheet_names': self.sheet_names,
                            'pca_profiles': [
                                {
                                    'component_id': i + 1,
                                    'intensity': convert_numpy_to_list(self.transformed_data[:, i])
                                }
                                for i in range(self.transformed_data.shape[1])
                            ]
                        }
                    }
                }

                # Write to JSON file with nice formatting
                with open(filepath, 'w') as f:
                    json.dump(pca_data, f, indent=2)

                wx.MessageBox("Full PCA analysis exported to JSON successfully!", "Export Complete", wx.OK)
            except Exception as e:
                wx.MessageBox(f"Error exporting to JSON: {e}", "Export Error", wx.OK | wx.ICON_ERROR)

        dlg.Destroy()

    def export_as_single_entity(self):
        """Export PCA components as Single Entity JSON file"""
        if self.components is None:
            wx.MessageBox("Please run Analyse first", "No Results", wx.OK | wx.ICON_WARNING)
            return

        import json

        wildcard = "JSON Files (*.json)|*.json"
        dlg = wx.FileDialog(self, "Export PCA as Single Entity", wildcard=wildcard,
                            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)

        if dlg.ShowModal() == wx.ID_OK:
            filepath = dlg.GetPath()

            try:
                # Get number of components to use
                n_use = self.use_spin.GetValue()

                # Helper function to convert numpy to list
                def convert_numpy_to_list(data):
                    if isinstance(data, np.ndarray):
                        return data.tolist()
                    return data

                # Create envelope peaks for each PCA component
                envelope_peaks = {}

                # Get base name from selected core levels
                base_name = "PCA"
                if len(self.sheet_names) > 0:
                    first_sheet = self.sheet_names[0]
                    if '_' in first_sheet:
                        base_name = first_sheet.split('_')[0]
                    else:
                        import re
                        match = re.match(r'([A-Za-z]+\d*[spdfg]*)', first_sheet)
                        if match:
                            base_name = match.group(1)

                for i in range(min(n_use, len(self.components))):
                    component_name = f"{base_name} PCA{i + 1}"

                    # Get component data
                    x_data = self.common_be_grid
                    y_data = self.components[i]

                    # Calculate properties
                    y_max = float(np.max(y_data))
                    y_max_index = np.argmax(y_data)
                    x_center = float(x_data[y_max_index])

                    # Sort for integration
                    sorted_indices = np.argsort(x_data)
                    x_sorted = x_data[sorted_indices]
                    y_sorted = y_data[sorted_indices]

                    # Calculate area
                    y_area = float(abs(np.trapz(y_sorted, x_sorted)))

                    # Create envelope entry
                    envelope_peaks[component_name] = {
                        "Position": round(x_center, 2),
                        "Height": round(y_max, 2),
                        "FWHM": 0.0,
                        "L/G": round(y_area, 2),
                        "Area": round(y_area, 2),
                        "Sigma": 0.00,
                        "Gamma": 1.00,
                        "Skew": 0.0,
                        "Fitting Model": "SingleEntity",
                        "x_data": convert_numpy_to_list(x_data),
                        "y_data": convert_numpy_to_list(y_data),
                        "Original_Position": round(x_center, 2),
                        "Original_Area": round(y_area, 2),
                        "Constraints": {
                            "Position": f"{float(np.min(x_data)):.2f},{float(np.max(x_data)):.2f}",
                            "Sigma": "-10:10",
                            "Gamma": "0.01:1000"
                        }
                    }

                # Create the main data structure following SingleEntity format
                peaks_data = {
                    'type': 'envelope_pca',
                    'Core levels': {
                        'PCA_Components': {
                            'Fitting': {
                                'Peaks': envelope_peaks,
                                'Model': 'SingleEntity'
                            }
                        }
                    }
                }

                # Write to JSON file
                with open(filepath, 'w') as f:
                    json.dump(peaks_data, f, indent=2)

                wx.MessageBox("PCA exported as Single Entity successfully!", "Export Complete", wx.OK)
            except Exception as e:
                wx.MessageBox(f"Error exporting as Single Entity: {e}", "Export Error", wx.OK | wx.ICON_ERROR)

        dlg.Destroy()

    def on_create_core_levels(self, event):
        """Create core levels from PCA eigenvectors"""
        if self.components is None:
            wx.MessageBox("Please run Analyse first", "No Results", wx.OK | wx.ICON_WARNING)
            return

        # Get number of components to use
        n_use = self.use_spin.GetValue()

        # Find the next available number for core levels
        if not hasattr(self.parent, 'Data') or 'Core levels' not in self.parent.Data:
            wx.MessageBox("No data structure found", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get file path
        file_path = self.parent.Data.get('FilePath', '')
        if not file_path:
            wx.MessageBox("No file path found. Please save your data first.", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Import necessary modules
        import pandas as pd
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows

        existing_sheets = list(self.parent.Data['Core levels'].keys())

        # Create core levels for each PCA component
        created_sheets = []

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)

            for i in range(min(n_use, len(self.components))):
                sheet_name = f"PCA{i + 1}"

                # Create the core level data structure
                self.parent.Data['Core levels'][sheet_name] = {
                    'B.E.': self.common_be_grid.tolist(),
                    'Raw Data': self.components[i].tolist(),
                    'Background': {
                        'Bkg Y': self.components[i].tolist(),
                        'Type': 'None',
                        'Bkg Low': float(self.common_be_grid[-1]),
                        'Bkg High': float(self.common_be_grid[0])
                    }
                }

                # Create DataFrame for Excel
                data_length = len(self.common_be_grid)
                column_names = ['Binding Energy', 'Corrected Data', 'Raw Data', 'Transmission']

                data_dict = {
                    column_names[0]: self.common_be_grid.tolist(),
                    column_names[1]: self.components[i].tolist(),
                    column_names[2]: self.components[i].tolist(),
                    column_names[3]: [1.0] * data_length
                }

                df = pd.DataFrame(data_dict)

                # Create new sheet in Excel
                if sheet_name in wb.sheetnames:
                    # Remove existing sheet if it exists
                    del wb[sheet_name]

                ws = wb.create_sheet(sheet_name)

                # Write DataFrame to worksheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                created_sheets.append(sheet_name)

            # Update number of core levels
            self.parent.Data['Number of Core levels'] = len(self.parent.Data['Core levels'])

            # Save workbook
            wb.save(file_path)

            # Update JSON file
            import json

            def convert_numpy_to_serializable(obj):
                """Convert numpy types to Python native types"""
                import numpy as np
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                elif isinstance(obj, np.integer):
                    return int(obj)
                elif isinstance(obj, np.floating):
                    return float(obj)
                elif isinstance(obj, dict):
                    return {key: convert_numpy_to_serializable(value) for key, value in obj.items()}
                elif isinstance(obj, list):
                    return [convert_numpy_to_serializable(item) for item in obj]
                return obj

            json_file_path = os.path.splitext(file_path)[0] + '.json'
            json_data = convert_numpy_to_serializable(self.parent.Data)
            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

        except Exception as e:
            wx.MessageBox(f"Error creating Excel sheets: {e}", "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()
            return

        # Update the sheet combobox if it exists
        if hasattr(self.parent, 'sheet_combobox'):
            current_sheets = [self.parent.sheet_combobox.GetString(i)
                              for i in range(self.parent.sheet_combobox.GetCount())]

            # Add new sheets to combobox
            for sheet_name in created_sheets:
                if sheet_name not in current_sheets:
                    self.parent.sheet_combobox.Append(sheet_name)

            # Select the first created sheet
            if created_sheets:
                self.parent.sheet_combobox.SetStringSelection(created_sheets[0])
                # Load the new sheet
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(self.parent, created_sheets[0])

        wx.MessageBox(f"Created {len(created_sheets)} core levels:\n" + "\n".join(created_sheets),
                      "Core Levels Created", wx.OK | wx.ICON_INFORMATION)

    def on_cancel(self, event):
        """Close the PCA analysis window"""
        self.Close()

    def on_close(self, event):
        """Clear parent reference when closing"""
        if hasattr(self.parent, 'pca_analysis_window'):
            self.parent.pca_analysis_window = None
        event.Skip()  # Allow window to close


def launch_pca_analysis(parent):
    """
    Launch PCA Analysis Window

    Args:
        parent (wx.Window): Parent window with Data structure
    """
    pca_window = PCAnalysisWindow(parent)
    return pca_window