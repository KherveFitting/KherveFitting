# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
EDX_SEM_Analysis.py
Module for Energy Dispersive X-ray Spectroscopy (EDX) and Scanning Electron Microscopy (SEM) analysis
Uses HyperSpy/ExSpy library for data import and analysis
"""

import wx
import numpy as np
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.backends.backend_wx import NavigationToolbar2Wx as NavigationToolbar
from matplotlib.widgets import RectangleSelector, SpanSelector
import matplotlib.patches as patches


# EDX Utilities - standalone replacements for HyperSpy/ExSpy
try:
    from libraries.EDX_Utilities import (
        elements as edx_elements,
        get_xray_lines_near_energy,
        get_element_xray_lines,
        find_peaks1D_ohaver,
        Signal1D,
        get_kfactor,
        DEFAULT_KFACTORS
    )
    from libraries.BCF_Reader import load_bcf, BCFData, BCFSpectrum, BCFMap
    EDX_UTILITIES_AVAILABLE = True
except ImportError:
    EDX_UTILITIES_AVAILABLE = False
    edx_elements = None

def get_hdf5_path_from_filepath(filepath):
    """
    Get HDF5 file path from Excel/JSON filepath.
    HDF5 file is always named same as Excel file but with .hdf5 extension.
    """
    if not filepath:
        return None
    base_path = os.path.splitext(filepath)[0]
    hdf5_path = f"{base_path}.hdf5"
    return hdf5_path if os.path.exists(hdf5_path) else None


class EDXSEMWindow(wx.Frame):
    """Main window for EDX/SEM data analysis"""

    def __init__(self, parent, title="EDX HeatMap"):
        super().__init__(parent, title=title, size=(590, 700), style = wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)

        # Store reference in parent for access from other modules
        if parent is not None:
            parent.edx_window = self

        self.parent = parent
        self.edx_data = None
        self.sem_data = None
        self.current_data = None

        # Selection modes
        self.selection_mode = None  # 'point', 'area', 'line'
        self.rect_selector = None
        self.line_selector = None
        self.selected_points = []
        self.selected_areas = []
        self.selected_lines = []

        # Default excluded elements (rare, noble gases, radioactive)
        default_excluded = {'Ra', 'Os', 'Dy', 'Lv', 'Db', 'Rg', 'Ho', 'Ac', 'Sm', 'Og', 'Po',
                            'Np', 'Nd', 'Hf', 'Rb', 'U', 'Rn', 'Th', 'Am', 'Lr', 'Pm', 'Bh',
                            'Ne', 'Cf', 'Xe', 'Lu', 'Bk', 'Cm', 'Cn', 'H', 'Ds', 'Eu', 'Pa',
                            'Sg', 'Fl', 'Tb', 'Es', 'Fr', 'No', 'Mc', 'Nh', 'Md', 'Hs', 'Yb',
                            'Kr', 'Pu', 'Tm', 'Mt', 'Rf', 'Er', 'At', 'Ts', 'Ar', 'He', 'Fm'}

        self.include_elements = set()  # elements to include (green)
        self.exclude_elements = default_excluded.copy()  # elements to exclude (red)
        self.selected_elements = []

        self.point_size = 1  # Size in pixels (1 = single pixel)
        self.line_width = 1  # Width of line selection in pixels

        # Store last positions for arrow button movement
        self.last_line_start = None
        self.last_line_end = None
        self.last_line_width = 1  # Store last line width for arrow movement

        self.loaded_signals = []
        self.data_browser_window = None
        self.info_window = None
        self.current_colorbar = None

        self.zoom_selector = None

        self.init_ui()
        self.Centre()

        # Load element preferences from data if available
        self.load_element_preferences_from_data()

        # Bind close event to clear parent reference
        self.Bind(wx.EVT_CLOSE, self.on_close)



    def init_ui(self):
        """Initialize user interface"""
        panel = wx.Panel(self, style=wx.BORDER_RAISED)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # # Create menubar
        # self.create_menubar()

        # Toolbar
        toolbar_panel = self.create_toolbar(panel)
        main_sizer.Add(toolbar_panel, 0, wx.EXPAND | wx.ALL, 2)

        # Main content - just the map
        map_panel = wx.Panel(panel)
        map_sizer = wx.BoxSizer(wx.VERTICAL)

        # Map display only
        self.map_figure = plt.Figure(figsize=(4, 4))
        self.map_figure.set_tight_layout(True)
        self.map_figure.subplots_adjust(left=0.1, right=0.95, top=0.95, bottom=0.1)

        self.map_canvas = FigureCanvas(map_panel, -1, self.map_figure)
        self.map_ax = self.map_figure.add_subplot(111)

        # Connect to matplotlib draw event to keep buttons visible
        self.map_figure.canvas.mpl_connect('draw_event', self._on_mpl_draw)

        # Add canvas to sizer
        map_sizer.Add(self.map_canvas, 1, wx.EXPAND)
        map_panel.SetSizer(map_sizer)

        # Create arrow button panel that will be positioned on bottom-left
        arrow_panel = wx.Panel(map_panel)
        arrow_panel.SetBackgroundColour(wx.Colour(240, 240, 240, 220))  # Semi-transparent gray
        arrow_sizer = wx.GridBagSizer(0, 0)  # 2px spacing

        btn_size = (30, 30)
        btn_size_wide = (60, 30)  # Width for 2 columns

        # Row 0: Up button spanning 2 columns
        self.arrow_up_btn = wx.Button(arrow_panel, label="↑", size=btn_size_wide)
        self.arrow_up_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_up_btn, pos=(0, 0), span=(1, 2), flag=wx.ALIGN_CENTER)

        # Row 1: Left and Right buttons
        self.arrow_left_btn = wx.Button(arrow_panel, label="←", size=btn_size)
        self.arrow_left_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_left_btn, pos=(1, 0), flag=wx.ALIGN_CENTER)

        self.arrow_right_btn = wx.Button(arrow_panel, label="→", size=btn_size)
        self.arrow_right_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_right_btn, pos=(1, 1), flag=wx.ALIGN_CENTER)

        # Row 2: Down button spanning 2 columns
        self.arrow_down_btn = wx.Button(arrow_panel, label="↓", size=btn_size_wide)
        self.arrow_down_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_down_btn, pos=(2, 0), span=(1, 2), flag=wx.ALIGN_CENTER)

        arrow_panel.SetSizer(arrow_sizer)
        arrow_panel.Fit()

        # Store reference for positioning later
        self.arrow_control_panel = arrow_panel

        # Set map panel sizer
        map_panel.SetSizer(map_sizer)

        # Bind arrow button events
        self.arrow_up_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('up'))
        self.arrow_down_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('down'))
        self.arrow_left_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('left'))
        self.arrow_right_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('right'))

        # Bind size event to reposition arrow panel
        map_panel.Bind(wx.EVT_SIZE, self.on_map_panel_resize)

        # Bind paint event to keep buttons visible
        self.map_canvas.Bind(wx.EVT_PAINT, self.on_canvas_paint)

        map_panel.SetSizer(map_sizer)

        main_sizer.Add(map_panel, 1, wx.EXPAND)

        panel.SetSizer(main_sizer)
        main_sizer.Layout()
        panel.Layout()

        # Bind canvas events
        self.map_canvas.mpl_connect('button_press_event', self.on_map_click)
        self.map_canvas.mpl_connect('motion_notify_event', self.on_map_motion)
        self.map_canvas.mpl_connect('scroll_event', self.on_map_scroll)
        self._line_preview = None

        # Right-click menu
        self.map_canvas.Bind(wx.EVT_RIGHT_DOWN, self.on_right_click)

        # Store colorbar reference
        self.current_colorbar = None
        self.current_cmap = 'Greens'

        wx.CallAfter(self.Layout)

    def create_menubar(self):
        """Create menu bar"""
        menubar = wx.MenuBar()

        # File menu
        file_menu = wx.Menu()

        # Import submenu
        import_menu = wx.Menu()
        import_sem = import_menu.Append(wx.ID_ANY, "SEM Image...", "Import SEM image")
        import_edx_spectrum = import_menu.Append(wx.ID_ANY, "EDX Spectrum...", "Import single EDX spectrum")
        import_edx_map = import_menu.Append(wx.ID_ANY, "EDX Map/Spectrum Image...", "Import EDX map or spectrum image")
        import_menu.AppendSeparator()
        import_bcf = import_menu.Append(wx.ID_ANY, "Bruker BCF File...", "Import Bruker BCF file")

        file_menu.AppendSubMenu(import_menu, "Import")

        # Export submenu
        export_menu = wx.Menu()
        export_excel = export_menu.Append(wx.ID_ANY, "Export to Excel...", "Export data to Excel")
        export_csv = export_menu.Append(wx.ID_ANY, "Export to CSV...", "Export spectrum to CSV")
        export_hdf5 = export_menu.Append(wx.ID_ANY, "Export to HDF5...", "Export to HDF5 format")
        export_menu.AppendSeparator()
        export_map_image = export_menu.Append(wx.ID_ANY, "Export Map as Image...", "Save map as PNG/TIFF")

        file_menu.AppendSubMenu(export_menu, "Export")

        file_menu.AppendSeparator()
        close_item = file_menu.Append(wx.ID_CLOSE, "Close\tCtrl+W", "Close window")

        menubar.Append(file_menu, "&File")

        # Analysis menu
        analysis_menu = wx.Menu()
        quantify_item = analysis_menu.Append(wx.ID_ANY, "Quantification...", "Perform EDX quantification")
        background_item = analysis_menu.Append(wx.ID_ANY, "Background Subtraction...", "Subtract background")

        menubar.Append(analysis_menu, "&Analysis")

        self.SetMenuBar(menubar)

        # Bind menu events
        self.Bind(wx.EVT_MENU, self.on_import_sem, import_sem)
        self.Bind(wx.EVT_MENU, self.on_import_edx_spectrum, import_edx_spectrum)
        self.Bind(wx.EVT_MENU, self.on_import_edx_map, import_edx_map)
        self.Bind(wx.EVT_MENU, self.on_import_bcf, import_bcf)
        self.Bind(wx.EVT_MENU, self.on_export_excel, export_excel)
        self.Bind(wx.EVT_MENU, self.on_export_csv, export_csv)
        self.Bind(wx.EVT_MENU, self.on_export_hdf5, export_hdf5)
        self.Bind(wx.EVT_MENU, self.on_export_map_image, export_map_image)
        self.Bind(wx.EVT_MENU, lambda e: self.Close(), close_item)
        self.Bind(wx.EVT_MENU, self.on_quantification, quantify_item)

    def create_toolbar(self, parent):
        """Create toolbar with icon buttons only"""
        toolbar_panel = wx.Panel(parent)
        toolbar_sizer = wx.BoxSizer(wx.HORIZONTAL)

        btn_size = (25, 25)

        # Get icon path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(current_dir, "Icons")

        # Navigation tools
        self.zoom_in_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.zoom_out_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        self.pan_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)

        # Load zoom icons with fallback
        zoom_in_path = os.path.join(icon_path, "ZoomIN-3.png")
        if os.path.exists(zoom_in_path):
            self.zoom_in_btn.SetBitmap(wx.Bitmap(zoom_in_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_in_btn.SetBitmap(self.create_icon_bitmap('zoom_in'))

        zoom_out_path = os.path.join(icon_path, "ZoomOUT-3.png")
        if os.path.exists(zoom_out_path):
            self.zoom_out_btn.SetBitmap(wx.Bitmap(zoom_out_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_out_btn.SetBitmap(self.create_icon_bitmap('zoom_out'))

        pan_path = os.path.join(icon_path, "Drag-25.png")
        if os.path.exists(pan_path):
            self.pan_btn.SetBitmap(wx.Bitmap(pan_path, wx.BITMAP_TYPE_PNG))
        else:
            self.pan_btn.SetBitmap(self.create_icon_bitmap('pan'))

        self.zoom_in_btn.SetToolTip("Zoom In")
        self.zoom_out_btn.SetToolTip("Zoom Out / Home")
        self.pan_btn.SetToolTip("Pan")

        toolbar_sizer.Add(self.zoom_in_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.zoom_out_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.pan_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL), 0, wx.EXPAND | wx.ALL, 3)

        # Selection tools - use created icons only
        self.point_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.area_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.line_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.clear_btn = wx.BitmapButton(toolbar_panel, size=btn_size)

        self.point_btn.SetBitmap(self.create_icon_bitmap('point'))
        self.area_btn.SetBitmap(self.create_icon_bitmap('area'))
        self.line_btn.SetBitmap(self.create_icon_bitmap('line'))
        self.clear_btn.SetBitmap(self.create_icon_bitmap('clear'))

        self.point_btn.SetToolTip("Point Selection")
        self.area_btn.SetToolTip("Area Selection")
        self.line_btn.SetToolTip("Line Selection")
        self.clear_btn.SetToolTip("Clear Selections")

        toolbar_sizer.Add(self.point_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.area_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.line_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.clear_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL), 0, wx.EXPAND | wx.ALL, 3)

        toolbar_panel.SetSizer(toolbar_sizer)

        # Plot Maps button
        self.plot_maps_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        self.plot_maps_btn.SetBitmap(self.create_icon_bitmap('intensity'))
        self.plot_maps_btn.SetToolTip("Plot Element Maps")
        toolbar_sizer.Add(self.plot_maps_btn, 0, wx.ALL, 2)

        # Intensity Map button
        self.intensity_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        heatmap_path = os.path.join(icon_path, "heatmap-3.png")
        self.intensity_btn.SetBitmap(wx.Bitmap(heatmap_path, wx.BITMAP_TYPE_PNG))
        self.intensity_btn.SetToolTip("Show Intensity Map")
        toolbar_sizer.Add(self.intensity_btn, 0, wx.ALL, 2)

        # Add spacer to push remaining buttons to the right
        toolbar_sizer.AddStretchSpacer()

        # Set Elements button
        self.set_elements_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        id_path = os.path.join(icon_path, "ID-3.png")
        if os.path.exists(id_path):
            self.set_elements_btn.SetBitmap(wx.Bitmap(id_path, wx.BITMAP_TYPE_PNG))
        else:
            self.set_elements_btn.SetBitmap(self.create_icon_bitmap('point'))
        self.set_elements_btn.SetToolTip("Set Elements")
        toolbar_sizer.Add(self.set_elements_btn, 0, wx.ALL, 2)

        # Sensitivity/Display controls button
        self.sensitivity_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        settings_path = os.path.join(icon_path, "Settings-3.png")
        if os.path.exists(settings_path):
            self.sensitivity_btn.SetBitmap(wx.Bitmap(settings_path, wx.BITMAP_TYPE_PNG))
        else:
            self.sensitivity_btn.SetBitmap(self.create_icon_bitmap('sensitivity'))
        self.sensitivity_btn.SetToolTip("Display & Sensitivity Controls")
        self.sensitivity_btn.Bind(wx.EVT_BUTTON, self.on_sensitivity_controls)
        toolbar_sizer.Add(self.sensitivity_btn, 0, wx.ALL, 2)



        toolbar_panel.SetSizer(toolbar_sizer)

        # Bind events
        self.zoom_in_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_zoom_in)
        self.zoom_out_btn.Bind(wx.EVT_BUTTON, self.on_zoom_out)
        self.pan_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_pan)
        self.point_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_point_mode)
        self.area_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_area_mode)
        self.line_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_line_mode)
        self.clear_btn.Bind(wx.EVT_BUTTON, self.on_clear_selections)
        self.set_elements_btn.Bind(wx.EVT_BUTTON, self.on_set_elements)
        self.plot_maps_btn.Bind(wx.EVT_BUTTON, self.on_plot_maps)
        self.intensity_btn.Bind(wx.EVT_BUTTON, self.on_intensity_map)

        return toolbar_panel

    def create_icon_bitmap(self, icon_type):
        """Create simple icon bitmaps"""
        size = 20
        bmp = wx.Bitmap(size, size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.Brush(wx.Colour(240, 240, 240)))
        dc.Clear()

        dc.SetPen(wx.Pen(wx.Colour(50, 50, 50), 2))
        dc.SetBrush(wx.Brush(wx.Colour(100, 100, 100)))

        if icon_type == 'zoom_in':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(8, 10, 12, 10)
            dc.DrawLine(10, 8, 10, 12)
            dc.DrawLine(14, 14, 18, 18)
        elif icon_type == 'zoom_out':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(8, 10, 12, 10)
            dc.DrawLine(14, 14, 18, 18)
        elif icon_type == 'pan':
            dc.DrawLine(10, 3, 10, 17)
            dc.DrawLine(3, 10, 17, 10)
            dc.DrawLine(10, 3, 7, 6)
            dc.DrawLine(10, 3, 13, 6)
            dc.DrawLine(10, 17, 7, 14)
            dc.DrawLine(10, 17, 13, 14)
        elif icon_type == 'point':
            dc.SetBrush(wx.Brush(wx.Colour(79, 190, 159)))
            dc.DrawCircle(10, 10, 4)
        elif icon_type == 'area':
            dc.SetBrush(wx.TRANSPARENT_BRUSH)
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            dc.DrawRectangle(3, 3, 14, 14)
        elif icon_type == 'line':
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            dc.DrawLine(3, 17, 17, 3)
        elif icon_type == 'clear':
            dc.SetPen(wx.Pen(wx.Colour(200, 50, 50), 2))
            dc.DrawLine(5, 5, 15, 15)
            dc.DrawLine(15, 5, 5, 15)
        elif icon_type == 'intensity' or icon_type == 'heatmap':
            # Draw gradient-like squares for heatmap
            dc.SetPen(wx.Pen(wx.Colour(50, 50, 50), 1))
            dc.SetBrush(wx.Brush(wx.Colour(0, 0, 255)))
            dc.DrawRectangle(2, 2, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(0, 255, 0)))
            dc.DrawRectangle(11, 2, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(255, 255, 0)))
            dc.DrawRectangle(2, 11, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(255, 0, 0)))
            dc.DrawRectangle(11, 11, 7, 7)
        elif icon_type == 'sensitivity':
            # Slider icon
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            # Horizontal line (slider track)
            dc.DrawLine(5, 10, 15, 10)
            # Slider knob
            dc.SetBrush(wx.Brush(wx.Colour(79, 190, 159)))
            dc.DrawCircle(10, 10, 3)
            # Small lines above and below for adjustment
            dc.DrawLine(10, 4, 10, 7)
            dc.DrawLine(10, 13, 10, 16)

        dc.SelectObject(wx.NullBitmap)
        return bmp



    # ==================== TOOLBAR HANDLERS ====================

    def on_sensitivity_controls(self, event):
        """Open sensitivity and display controls window"""
        if hasattr(self, 'sensitivity_window') and self.sensitivity_window and not self.sensitivity_window.IsBeingDeleted():
            # Window already exists, bring to front
            self.sensitivity_window.Raise()
        else:
            # Create new window
            self.sensitivity_window = EDXSensitivityWindow(self)
            self.sensitivity_window.Show()

    def on_zoom_in(self, event):
        """Toggle zoom in mode using rectangle selector"""
        if self.zoom_in_btn.GetValue():
            self.pan_btn.SetValue(False)
            self.deactivate_selection_modes()

            # Create zoom rectangle selector
            if not hasattr(self, 'zoom_selector') or self.zoom_selector is None:
                self.zoom_selector = RectangleSelector(
                    self.map_ax,
                    self.on_zoom_select,
                    useblit=True,
                    props=dict(facecolor='blue', edgecolor='blue', alpha=0.2, fill=True),
                    button=[1],
                    minspanx=5,
                    minspany=5,
                    spancoords='pixels',
                    interactive=False
                )
            else:
                self.zoom_selector.set_active(True)
        else:
            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

    def on_zoom_select(self, eclick, erelease):
        """Handle zoom rectangle selection"""
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        x_min, x_max = min(x1, x2), max(x1, x2)
        y_min, y_max = min(y1, y2), max(y1, y2)

        self.map_ax.set_xlim(x_min, x_max)
        self.map_ax.set_ylim(y_max, y_min)  # Inverted for image

        # Update scale bar for new zoom level
        self._update_scale_bar()

        self.map_canvas.draw()

        # Deselect zoom button after use
        self.zoom_in_btn.SetValue(False)
        if hasattr(self, 'zoom_selector') and self.zoom_selector:
            self.zoom_selector.set_active(False)

    def on_zoom_out(self, event):
        """Zoom out / reset view"""
        self.map_ax.autoscale()

        # Update scale bar for full view
        self._update_scale_bar()

        self.map_canvas.draw()

        # Deselect buttons
        self.zoom_in_btn.SetValue(False)
        self.pan_btn.SetValue(False)
        if hasattr(self, 'zoom_selector') and self.zoom_selector:
            self.zoom_selector.set_active(False)

    def on_pan(self, event):
        """Toggle pan mode"""
        if self.pan_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.deactivate_selection_modes()

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

            # Enable pan via mouse drag
            self._pan_start = None
            self._pan_cid_press = self.map_canvas.mpl_connect('button_press_event', self._on_pan_press)
            self._pan_cid_release = self.map_canvas.mpl_connect('button_release_event', self._on_pan_release)
            self._pan_cid_motion = self.map_canvas.mpl_connect('motion_notify_event', self._on_pan_motion)
        else:
            # Disconnect pan events
            if hasattr(self, '_pan_cid_press'):
                self.map_canvas.mpl_disconnect(self._pan_cid_press)
            if hasattr(self, '_pan_cid_release'):
                self.map_canvas.mpl_disconnect(self._pan_cid_release)
            if hasattr(self, '_pan_cid_motion'):
                self.map_canvas.mpl_disconnect(self._pan_cid_motion)

    def _on_pan_press(self, event):
        if event.inaxes == self.map_ax and event.button == 1:
            self._pan_start = (event.xdata, event.ydata)

    def _on_pan_release(self, event):
        self._pan_start = None
        self._add_scale_bar()
        self.map_canvas.draw_idle()

    def _on_pan_motion(self, event):
        if self._pan_start is None or event.inaxes != self.map_ax:
            return

        dx = self._pan_start[0] - event.xdata
        dy = self._pan_start[1] - event.ydata

        xlim = self.map_ax.get_xlim()
        ylim = self.map_ax.get_ylim()

        self.map_ax.set_xlim(xlim[0] + dx, xlim[1] + dx)
        self.map_ax.set_ylim(ylim[0] + dy, ylim[1] + dy)

        self.map_canvas.draw_idle()

    def on_home(self, event):
        """Reset view to home"""
        self.map_toolbar.home()
        self.map_canvas.draw()

    def deactivate_selection_modes(self):
        """Deactivate all selection modes"""
        self.point_btn.SetValue(False)
        self.area_btn.SetValue(False)
        self.line_btn.SetValue(False)
        self.selection_mode = None

        if self.rect_selector:
            self.rect_selector.set_active(False)
        if self.line_selector:
            self.line_selector.set_active(False)

    def on_point_mode(self, event):
        """Toggle point selection mode"""
        if self.point_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.area_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'point'

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)
            if self.rect_selector:
                self.rect_selector.set_active(False)

            # Clear previous selections
            self.clear_selection_markers()
            self.map_canvas.draw()
        else:
            self.selection_mode = None

    def on_area_mode(self, event):
        """Toggle area selection mode with rotatable rectangle"""
        if self.area_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.point_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'area'

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

            # Clear previous selections
            self.clear_selection_markers()
            self.map_canvas.draw()

            # Deactivate old rectangle selector
            if self.rect_selector:
                self.rect_selector.set_active(False)

            # Create or activate rotatable rectangle
            if not hasattr(self, 'rotatable_rect') or self.rotatable_rect is None:
                self.rotatable_rect = RotatableRectangle(self.map_ax, self.on_rotated_area_complete)

            # Connect click event to start new selection
            if not hasattr(self, '_area_click_cid'):
                self._area_click_cid = self.map_canvas.mpl_connect('button_press_event',
                                                                   lambda e: self.rotatable_rect.start_selection(e) if self.area_btn.GetValue() and e.button == 1 else None)
        else:
            self.selection_mode = None
            if self.rect_selector:
                self.rect_selector.set_active(False)
            if hasattr(self, 'rotatable_rect') and self.rotatable_rect:
                self.rotatable_rect.clear()
            if hasattr(self, '_area_click_cid'):
                self.map_canvas.mpl_disconnect(self._area_click_cid)
                self._area_click_cid = None

    def on_rotated_area_complete(self, center, width, height, angle):
        """Handle completion of rotated area selection with proper pixel extraction"""
        if self.current_data is None:
            return

        data = self.current_data.data
        if len(data.shape) != 3:
            return

        # Get dimensions
        map_height, map_width = data.shape[0], data.shape[1]
        cx, cy = center
        hw, hh = width / 2, height / 2

        # Create rotation matrix
        angle_rad = np.radians(-angle)  # Negative for inverse transform
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        # Create mask for pixels inside rotated rectangle
        y_indices, x_indices = np.meshgrid(np.arange(map_height), np.arange(map_width), indexing='ij')

        # Transform all pixels to rectangle's local coordinate system
        local_x = (x_indices - cx) * cos_a - (y_indices - cy) * sin_a
        local_y = (x_indices - cx) * sin_a + (y_indices - cy) * cos_a

        # Create mask for pixels inside rectangle
        mask = (np.abs(local_x) <= hw) & (np.abs(local_y) <= hh)

        # Extract spectra from masked pixels
        masked_spectra = data[mask]

        if len(masked_spectra) == 0:
            print("No pixels in selected area")
            return

        # Sum spectra
        summed_spectrum = np.sum(masked_spectra, axis=0)

        # Get energy axis
        energy_axis = self.current_data.axes_manager[-1]
        energy = energy_axis.axis

        # Plot to parent window
        if self.parent is not None:
            # Create EDX~Plot sheet if it doesn't exist
            if 'EDX~Plot' not in self.parent.Data['Core levels']:
                self.parent.Data['Core levels']['EDX~Plot'] = {}

            # Store data
            self.parent.Data['Core levels']['EDX~Plot']['Energy_keV'] = energy.tolist()
            self.parent.Data['Core levels']['EDX~Plot']['Intensity'] = summed_spectrum.tolist()

            # Update sheet combobox
            self.parent.sheet_combobox.SetValue('EDX~Plot')

            # Plot using common styled method
            title = f'EDX Area Spectrum (Rotated, {len(masked_spectra)} pixels)'
            grid_label = f"Rotated Area ({len(masked_spectra)} px)"
            self._plot_edx_spectrum_styled(energy, summed_spectrum, title, grid_label)

        print(f"Rotated area: center=({cx:.0f},{cy:.0f}), size=({width:.0f}x{height:.0f}), angle={angle:.1f}°")
        print(f"Extracted {len(masked_spectra)} pixels")

    def on_line_mode(self, event):
        """Toggle line selection mode"""
        if self.line_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.point_btn.SetValue(False)
            self.area_btn.SetValue(False)
            self.selection_mode = 'line'
            self.line_start = None
            self.line_end = None

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)
            if self.rect_selector:
                self.rect_selector.set_active(False)

            # Clear previous selections
            self.clear_selection_markers()
            self.map_canvas.draw()
        else:
            self.selection_mode = None

    def on_clear_selections(self, event=None):
        """Clear all selections"""
        self.selected_points = []
        self.selected_areas = []
        self.selected_lines = []
        self.line_start = None
        self.line_end = None
        self.last_line_start = None
        self.last_line_end = None

        # Clear rotatable rectangle if it exists
        if hasattr(self, 'rotatable_rect') and self.rotatable_rect:
            self.rotatable_rect.clear()

        # Clear selection markers
        self.clear_selection_markers()

        # Clear saved selection markers (black elements on heatmap)
        self._clear_saved_selection()

        self.map_canvas.draw()

    def on_colormap_change(self, event):
        """Change colormap"""
        if self.current_data is not None:
            self.plot_current_map()


    # =================== RIGHT CLICK HANDLERS ===================

    def on_map_right_click(self, event):
        """Handle right-click on map for context menu"""
        if event.button != 3:  # Right click
            return
        if event.inaxes != self.map_ax:
            return

        # Create popup menu
        menu = wx.Menu()

        save_plot_item = menu.Append(wx.ID_ANY, "Save Current EDX Plot...")
        self.Bind(wx.EVT_MENU, self.on_save_edx_plot, save_plot_item)

        # Show menu at mouse position
        self.PopupMenu(menu)
        menu.Destroy()

    def on_save_edx_plot(self, event):
        """Save current EDX plot to a numbered sheet in Excel and window.Data"""
        if self.parent is None:
            return

        # Check if Excel file exists
        if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
            wx.MessageBox("No Excel file found. Please ensure the EDX data was imported correctly.",
                          "No File", wx.OK | wx.ICON_WARNING)
            return

        # Find next available EDX~Plot number
        existing_sheets = list(self.parent.Data.get('Core levels', {}).keys())
        plot_num = 1
        while f'EDX~Plot{plot_num}' in existing_sheets:
            plot_num += 1

        sheet_name = f'EDX~Plot{plot_num}'

        # Gather current selection info
        selection_info = self._get_current_selection_info()

        if selection_info is None:
            wx.MessageBox("No selection to save. Please select a point, line, or area first.",
                          "No Selection", wx.OK | wx.ICON_WARNING)
            return

        # Get current spectrum data from parent
        if not hasattr(self.parent, 'ax') or len(self.parent.ax.lines) == 0:
            wx.MessageBox("No EDX plot data to save.", "No Data", wx.OK | wx.ICON_WARNING)
            return

        # Get spectrum data from plot
        line = self.parent.ax.lines[0]
        energy = line.get_xdata()
        intensity = line.get_ydata()

        # Get grid data
        grid_data = self._get_grid_data()

        # Create sheet data - USE SAME STRUCTURE AS XPS SHEETS
        import datetime
        import json
        sheet_data = {
            'Name': sheet_name,
            'B.E.': list(energy),
            'Raw Data': list(intensity),
            '_EDX_display_max': 20,
            '_EDX_type': 'plot',
            '_EDX_selection': selection_info,
            '_EDX_grid_data': grid_data,
            '_EDX_save_time': datetime.datetime.now().isoformat(),
            'Background': {}
        }

        # Save to parent Data
        if 'Core levels' not in self.parent.Data:
            self.parent.Data['Core levels'] = {}

        self.parent.Data['Core levels'][sheet_name] = sheet_data

        # Update sheet combobox
        if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                              for i in range(self.parent.sheet_combobox.GetCount())]:
            self.parent.sheet_combobox.Append(sheet_name)

        # ========== Save to Excel file ==========
        try:
            import pandas as pd
            import openpyxl

            file_path = self.parent.Data['FilePath']

            # Create DataFrame for this plot
            edx_df = pd.DataFrame({
                'Energy (keV)': [f"{v:.2f}" for v in energy],
                'Intensity': [f"{v:.2f}" for v in intensity]
            })

            # Append to Excel file
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                edx_df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Saved {sheet_name} to Excel: {file_path}")

            # Also update JSON
            json_path = file_path.replace('.xlsx', '.json')
            if os.path.exists(json_path):
                with open(json_path, 'r') as f:
                    json_data = json.load(f)

                if 'Core levels' not in json_data:
                    json_data['Core levels'] = {}

                json_data['Core levels'][sheet_name] = {
                    'Name': sheet_name,
                    'B.E.': [float(f"{v:.2f}") for v in energy],
                    'Raw Data': [float(f"{v:.2f}") for v in intensity],
                    '_EDX_display_max': 20,
                    '_EDX_type': 'plot',
                    '_EDX_selection': selection_info,
                    '_EDX_save_time': datetime.datetime.now().isoformat()
                }

                with open(json_path, 'w') as f:
                    json.dump(json_data, f, indent=2)

                print(f"Updated JSON: {json_path}")

            wx.MessageBox(f"EDX plot saved as '{sheet_name}'\n\nSaved to Excel and JSON files.",
                          "Saved", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Saved to memory but error saving to Excel:\n{str(e)}\n\nUse 'Save' from toolbar to save all data.",
                          "Partial Save", wx.OK | wx.ICON_WARNING)
            import traceback
            traceback.print_exc()

    def _get_current_selection_info(self):
        """Get information about current selection"""
        info = {}

        # Check for point selection
        if self.selected_points:
            point = self.selected_points[-1]
            info['type'] = 'point'
            info['x'] = point[0]
            info['y'] = point[1]
            info['size'] = point[2] if len(point) > 2 else 1
            return info

        # Check for line selection
        if hasattr(self, 'line_start') and self.line_start and hasattr(self, 'line_end') and self.line_end:
            info['type'] = 'line'
            info['x1'] = self.line_start[0]
            info['y1'] = self.line_start[1]
            info['x2'] = self.line_end[0]
            info['y2'] = self.line_end[1]
            return info

        # Check for rotatable rectangle
        if hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.active:
            rect = self.rotatable_rect
            info['type'] = 'rectangle'
            info['center_x'] = rect.center[0]
            info['center_y'] = rect.center[1]
            info['width'] = rect.width
            info['height'] = rect.height
            info['angle'] = rect.angle
            return info

        # Check for area selection
        if self.selected_areas:
            area = self.selected_areas[-1]
            info['type'] = 'area'
            info['x1'] = area[0]
            info['y1'] = area[1]
            info['x2'] = area[2]
            info['y2'] = area[3]
            return info

        return None

    def _get_grid_data(self):
        """Get current peak fitting grid data"""
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return []

        grid = self.parent.peak_params_grid
        data = []

        for row in range(0, grid.GetNumberRows(), 2):  # Only data rows, skip constraint rows
            row_data = {
                'id': grid.GetCellValue(row, 0),
                'label': grid.GetCellValue(row, 1),
                'position': grid.GetCellValue(row, 2),
                'height': grid.GetCellValue(row, 3),
                'fwhm': grid.GetCellValue(row, 4),
                'area': grid.GetCellValue(row, 6),
                'concentration': grid.GetCellValue(row, 10),
            }
            data.append(row_data)

        return data

    def draw_saved_selection(self, selection_info):
        """Draw a saved selection on the map"""
        # Clear previous saved selection markers
        self._clear_saved_selection()

        if selection_info is None:
            return

        sel_type = selection_info.get('type')

        if sel_type == 'point':
            x, y = selection_info['x'], selection_info['y']
            size = selection_info.get('size', 1)

            if size == 1:
                marker, = self.map_ax.plot(x, y, 'k+', markersize=15, markeredgewidth=2)
                marker._is_saved_selection = True
            else:
                from matplotlib.patches import Rectangle
                half_size = size / 2
                rect = Rectangle((x - half_size, y - half_size), size, size,
                                 linewidth=2, edgecolor='black', facecolor='none')
                rect._is_saved_selection = True
                self.map_ax.add_patch(rect)

        elif sel_type == 'line':
            line, = self.map_ax.plot([selection_info['x1'], selection_info['x2']],
                                     [selection_info['y1'], selection_info['y2']],
                                     'k-', linewidth=2)
            line._is_saved_selection = True

        elif sel_type == 'rectangle':
            from matplotlib.patches import Rectangle
            from matplotlib.transforms import Affine2D

            cx, cy = selection_info['center_x'], selection_info['center_y']
            w, h = selection_info['width'], selection_info['height']
            angle = selection_info['angle']

            rect = Rectangle((-w / 2, -h / 2), w, h,
                             linewidth=2, edgecolor='black', facecolor='none')
            t = Affine2D().rotate_deg(angle).translate(cx, cy) + self.map_ax.transData
            rect.set_transform(t)
            rect._is_saved_selection = True
            self.map_ax.add_patch(rect)

        elif sel_type == 'area':
            from matplotlib.patches import Rectangle
            x1, y1 = selection_info['x1'], selection_info['y1']
            x2, y2 = selection_info['x2'], selection_info['y2']
            rect = Rectangle((x1, y1), x2 - x1, y2 - y1,
                             linewidth=2, edgecolor='black', facecolor='none')
            rect._is_saved_selection = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw_idle()

    def _clear_saved_selection(self):
        """Clear saved selection markers"""
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_saved_selection') and artist._is_saved_selection:
                try:
                    artist.remove()
                except:
                    pass
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_saved_selection') and artist._is_saved_selection:
                try:
                    artist.remove()
                except:
                    pass

    # ==================== SELECTION HANDLERS ====================

    def on_map_scroll(self, event):
        """Handle mouse scroll on map"""
        if event.inaxes != self.map_ax:
            return

        if self.selection_mode == 'point':
            # Adjust point size with scroll wheel
            if event.button == 'up':
                self.point_size = min(50, self.point_size + 1)
            elif event.button == 'down':
                self.point_size = max(1, self.point_size - 1)

            # Update point preview if we have a position
            self._update_point_preview(event.xdata, event.ydata)

        elif self.selection_mode == 'line':
            # Adjust line width with scroll wheel
            if event.button == 'up':
                self.line_width = min(50, self.line_width + 1)
            elif event.button == 'down':
                self.line_width = max(1, self.line_width - 1)

            # Update line preview if we have a start point
            if self.line_start is not None and event.xdata is not None and event.ydata is not None:
                self._clear_line_preview()
                self._draw_line_preview_with_width(self.line_start,
                                                   (int(event.xdata), int(event.ydata)),
                                                   self.line_width)
                self.map_canvas.draw_idle()


    def _update_point_preview(self, x, y):
        """Update point size preview rectangle"""
        # Remove existing preview
        self._clear_point_preview()

        if x is None or y is None:
            return

        if self.point_size == 1:
            # Single pixel - show as cross
            marker, = self.map_ax.plot(x, y, 'r+', markersize=15, markeredgewidth=2, alpha=0.5)
            marker._is_point_preview = True
        else:
            # Show rectangle preview
            from matplotlib.patches import Rectangle
            half_size = self.point_size / 2
            rect = Rectangle((x - half_size, y - half_size), self.point_size, self.point_size,
                             linewidth=2, edgecolor='red', facecolor='salmon', alpha=0.3)
            rect._is_point_preview = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw_idle()

    def _clear_point_preview(self):
        """Clear point preview"""
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_point_preview') and artist._is_point_preview:
                try:
                    artist.remove()
                except:
                    pass
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_point_preview') and artist._is_point_preview:
                try:
                    artist.remove()
                except:
                    pass

    def on_map_click(self, event):
        """Handle click on map"""
        if event.inaxes != self.map_ax:
            return
        if self.current_data is None:
            return

        x, y = int(event.xdata), int(event.ydata)

        # Validate coordinates
        data_shape = self.current_data.data.shape
        if len(data_shape) < 2:
            return

        max_x = data_shape[1] if len(data_shape) >= 2 else data_shape[0]
        max_y = data_shape[0]

        if x < 0 or x >= max_x or y < 0 or y >= max_y:
            return

        if self.selection_mode == 'point':
            # Clear previous markers and preview
            self.clear_selection_markers()
            self._clear_point_preview()

            # Store point info with size
            self.selected_points = [(x, y, self.point_size)]

            if self.point_size == 1:
                # Single pixel - show as cross
                marker, = self.map_ax.plot(x, y, 'r+', markersize=15, markeredgewidth=2)
                marker._is_selection_marker = True
            else:
                # Show rectangle for multi-pixel selection
                from matplotlib.patches import Rectangle
                half_size = self.point_size / 2
                rect = Rectangle((x - half_size, y - half_size), self.point_size, self.point_size,
                                 linewidth=2, edgecolor='darkred', facecolor='red', alpha=0.3)
                rect._is_selection_marker = True
                self.map_ax.add_patch(rect)

            self.map_canvas.draw()

            # Extract and plot spectrum
            self.plot_point_spectrum(x, y)

        elif self.selection_mode == 'line':
            if self.line_start is None:
                # First click - start of line
                # Clear any previous line markers first (for starting a new line)
                self.clear_selection_markers()
                self._clear_line_preview()

                self.line_start = (x, y)
                # Draw start marker (red)
                marker, = self.map_ax.plot(x, y, 'ro', markersize=8, markeredgecolor='darkred', markeredgewidth=2)
                marker._is_selection_marker = True
                self.map_canvas.draw()
            else:
                # Second click - end of line
                self.line_end = (x, y)

                # Clear preview line
                self._clear_line_preview()

                # Clear previous markers and redraw final line with width (red)
                self.clear_selection_markers()
                self._draw_line_with_width(self.line_start, self.line_end, self.line_width)

                self.map_canvas.draw()

                # Create line profile (zzProfile) with atomic concentrations
                self.create_line_profile(self.line_start[0], self.line_start[1],
                                         self.line_end[0], self.line_end[1], self.line_width)

                # Store positions for arrow button movement
                self.last_line_start = self.line_start
                self.last_line_end = self.line_end
                self.last_line_width = self.line_width

                # Reset for next line
                self.line_start = None
                self.line_end = None


    def on_map_motion(self, event):
        """Handle mouse motion on map for line preview"""
        if self.selection_mode != 'line' or self.line_start is None:
            return
        if event.inaxes != self.map_ax:
            return
        if event.xdata is None or event.ydata is None:
            return

        x, y = int(event.xdata), int(event.ydata)

        # Update line preview with width indicator (red)
        self._clear_line_preview()
        self._draw_line_preview_with_width(self.line_start, (x, y), self.line_width)
        self.map_canvas.draw_idle()

    def _clear_line_preview(self):
        """Clear the line preview"""
        if self._line_preview is not None:
            try:
                self._line_preview.remove()
            except (ValueError, AttributeError):
                pass
            self._line_preview = None

        # Also clear width preview patches
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_line_preview') and artist._is_line_preview:
                try:
                    artist.remove()
                except (ValueError, AttributeError):
                    pass
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_line_preview') and artist._is_line_preview:
                try:
                    artist.remove()
                except (ValueError, AttributeError):
                    pass

    def _draw_line_preview_with_width(self, start, end, width):
        """Draw line preview with width indicator"""
        from matplotlib.patches import Polygon

        x1, y1 = start
        x2, y2 = end

        # Draw center line (dashed, red)
        self._line_preview, = self.map_ax.plot([x1, x2], [y1, y2],
                                               'r--', linewidth=1.5, alpha=0.7)
        self._line_preview._is_line_preview = True

        if width > 1:
            # Calculate perpendicular direction
            dx = x2 - x1
            dy = y2 - y1
            length = np.sqrt(dx ** 2 + dy ** 2)
            if length > 0:
                # Perpendicular unit vector
                px = -dy / length
                py = dx / length

                half_width = width / 2

                # Create polygon for width preview
                corners = [
                    (x1 + px * half_width, y1 + py * half_width),
                    (x2 + px * half_width, y2 + py * half_width),
                    (x2 - px * half_width, y2 - py * half_width),
                    (x1 - px * half_width, y1 - py * half_width),
                ]
                poly = Polygon(corners, closed=True, fill=True,
                               facecolor='red', edgecolor='darkred',
                               alpha=0.2, linewidth=1)
                poly._is_line_preview = True
                self.map_ax.add_patch(poly)

    def _draw_line_with_width(self, start, end, width):
        """Draw final line selection with width indicator"""
        from matplotlib.patches import Polygon

        x1, y1 = start
        x2, y2 = end

        # Draw center line (solid, red)
        line, = self.map_ax.plot([x1, x2], [y1, y2], 'r-', linewidth=2)
        line._is_selection_marker = True

        # Draw end points (red)
        marker1, = self.map_ax.plot(x1, y1, 'ro', markersize=8,
                                    markeredgecolor='darkred', markeredgewidth=2)
        marker1._is_selection_marker = True
        marker2, = self.map_ax.plot(x2, y2, 'ro', markersize=8,
                                    markeredgecolor='darkred', markeredgewidth=2)
        marker2._is_selection_marker = True

        if width > 1:
            # Calculate perpendicular direction
            dx = x2 - x1
            dy = y2 - y1
            length = np.sqrt(dx ** 2 + dy ** 2)
            if length > 0:
                # Perpendicular unit vector
                px = -dy / length
                py = dx / length

                half_width = width / 2

                # Create polygon for width indicator
                corners = [
                    (x1 + px * half_width, y1 + py * half_width),
                    (x2 + px * half_width, y2 + py * half_width),
                    (x2 - px * half_width, y2 - py * half_width),
                    (x1 - px * half_width, y1 - py * half_width),
                ]
                poly = Polygon(corners, closed=True, fill=True,
                               facecolor='red', edgecolor='darkred',
                               alpha=0.3, linewidth=1.5)
                poly._is_selection_marker = True
                self.map_ax.add_patch(poly)


    def clear_selection_markers(self):
        """Clear all selection markers from map"""
        # Remove marked lines and points
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_selection_marker') and artist._is_selection_marker:
                artist.remove()
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_selection_marker') and artist._is_selection_marker:
                artist.remove()

    def on_area_select(self, eclick, erelease):
        """Handle area selection"""
        if self.current_data is None:
            return

        x1, y1 = int(eclick.xdata), int(eclick.ydata)
        x2, y2 = int(erelease.xdata), int(erelease.ydata)

        # Ensure proper order
        x_min, x_max = min(x1, x2), max(x1, x2)
        y_min, y_max = min(y1, y2), max(y1, y2)

        self.selected_areas.append((x_min, y_min, x_max, y_max))

        # Plot summed spectrum from area
        self.plot_area_spectrum(x_min, y_min, x_max, y_max)

    def on_line_select(self, xmin, xmax):
        """Handle line selection (horizontal profile)"""
        if self.current_data is None:
            return

        x1, x2 = int(xmin), int(xmax)
        self.selected_lines.append((x1, x2))

        # Plot line profile spectrum
        self.plot_line_spectrum(x1, x2)

    # ==================== SPECTRUM PLOTTING ====================


    def plot_point_spectrum(self, x, y):
        """Plot spectrum from point or area around point"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for point spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get point size
        point_size = getattr(self, 'point_size', 1)

        if point_size == 1:
            # Single pixel
            spectrum = data[y, x, :]
            title = f'EDX Spectrum at Point ({x}, {y})'
            title = ' '
        else:
            # Average over area
            half_size = point_size // 2
            y1 = max(0, y - half_size)
            y2 = min(data.shape[0], y + half_size + 1)
            x1 = max(0, x - half_size)
            x2 = min(data.shape[1], x + half_size + 1)

            spectrum = np.mean(data[y1:y2, x1:x2, :], axis=(0, 1))
            title = f'EDX Spectrum at Point ({x}, {y}) [{point_size}×{point_size} px]'
            title = ' '

        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        # Plot using common EDX plotting method
        self._plot_edx_spectrum_styled(energy, spectrum, title)

    def _plot_spectrum_with_style(self, ax, energy, spectrum):
        """Plot spectrum using saved style settings from parent window"""
        if self.parent is None:
            ax.plot(energy, spectrum, 'k-', linewidth=0.8)
            return

        # Get style from parent window (loaded from config.json)
        plot_style = getattr(self.parent, 'plot_style', 'scatter')
        line_color = getattr(self.parent, 'line_color', 'black')
        line_width = getattr(self.parent, 'line_width', 1.0)
        line_alpha = getattr(self.parent, 'line_alpha', 1.0)
        scatter_color = getattr(self.parent, 'scatter_color', 'black')
        scatter_size = getattr(self.parent, 'scatter_size', 10)
        scatter_marker = getattr(self.parent, 'scatter_marker', 'o')
        raw_data_linestyle = getattr(self.parent, 'raw_data_linestyle', '-')
        if plot_style == 'scatter':
            ax.scatter(energy, spectrum,
                       c=scatter_color,
                       s=scatter_size,
                       marker=scatter_marker,
                       alpha=line_alpha)
        elif plot_style == 'line':
            ax.plot(energy, spectrum,
                    color=line_color,
                    linewidth=line_width,
                    linestyle=raw_data_linestyle,
                    alpha=line_alpha)
        elif plot_style == 'both':
            ax.plot(energy, spectrum,
                    color=line_color,
                    linewidth=line_width,
                    linestyle=raw_data_linestyle,
                    alpha=line_alpha)
            ax.scatter(energy, spectrum,
                       c=scatter_color,
                       s=scatter_size,
                       marker=scatter_marker,
                       alpha=line_alpha)
        else:
            # Default fallback
            ax.plot(energy, spectrum,
                    color=line_color,
                    linewidth=line_width,
                    alpha=line_alpha)

    def _plot_edx_spectrum_styled(self, energy, spectrum, title, grid_label=None):
        """Common method to plot EDX spectrum with style from Plot_Operations.apply_edx_plot_style

        Args:
            energy: energy axis array
            spectrum: intensity array
            title: plot title
            grid_label: optional label for grid (defaults to title if None)
        """
        if self.parent is None or not hasattr(self.parent, 'ax'):
            return

        if grid_label is None:
            grid_label = title

        self.parent.ax.clear()

        # Get stored style preference from window config
        style = getattr(self.parent, 'edx_plot_style', 'black')

        # Use Plot_Operations.apply_edx_plot_style if available
        if hasattr(self.parent, 'apply_edx_plot_style'):
            label_color = self.parent.apply_edx_plot_style(
                self.parent.ax, self.parent.figure, energy, spectrum, style
            )
        else:
            # Fallback: apply style directly (same logic as Plot_Operations)
            label_color = self._apply_edx_style(energy, spectrum, style)

        self.parent.ax.set_title(title, color=label_color)

        # Set Y-axis to scientific format
        self.parent.ax.ticklabel_format(axis='y', style='scientific', scilimits=(0, 0))

        # Add element peak labels
        self.add_peak_labels(self.parent.ax, energy, spectrum)

        # Apply label color to all text elements
        for text in self.parent.ax.texts:
            text.set_color(label_color)
            text.set_fontweight('bold')

        # Get stored X max from any EDX~Plot sheet or default to 20
        display_x_max = 20
        if 'Core levels' in self.parent.Data:
            for sname in self.parent.Data['Core levels']:
                if sname == 'EDX~Plot' or sname.startswith('EDX~Plot'):
                    if '_EDX_display_max' in self.parent.Data['Core levels'][sname]:
                        display_x_max = self.parent.Data['Core levels'][sname]['_EDX_display_max']
                        break

        self.parent.ax.set_xlim(0, display_x_max)
        self.parent.ax.set_ylim(np.min(spectrum) * 0.95, np.max(spectrum) * 1.1)

        self.parent.canvas.draw()

        # Populate peak fitting grid with quantification
        self._populate_edx_grid(energy, spectrum, grid_label)

    def _apply_edx_style(self, energy, spectrum, style='black'):
        """Fallback method to apply EDX plot style (mirrors Plot_Operations.apply_edx_plot_style)"""
        ax = self.parent.ax
        figure = self.parent.figure

        if style == 'black':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.plot(energy, spectrum, 'k-', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        elif style == 'blue_yellow':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('#1e3a5f')
            ax.fill_between(energy, spectrum, color='yellow', alpha=0.8)
            ax.plot(energy, spectrum, 'yellow', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'yellow'

        elif style == 'white_green':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.fill_between(energy, spectrum, color='#16A085', alpha=0.6)
            ax.plot(energy, spectrum, '#16A085', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        elif style == 'white_red':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.fill_between(energy, spectrum, color='red', alpha=0.6)
            ax.plot(energy, spectrum, 'red', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        elif style == 'white_blue':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.fill_between(energy, spectrum, color='blue', alpha=0.6)
            ax.plot(energy, spectrum, 'blue', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        elif style == 'white_purple':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.fill_between(energy, spectrum, color='purple', alpha=0.6)
            ax.plot(energy, spectrum, 'purple', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        elif style == 'white_pink':
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.fill_between(energy, spectrum, color='pink', alpha=0.8)
            ax.plot(energy, spectrum, 'purple', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'

        else:
            # Default fallback
            figure.patch.set_facecolor('white')
            ax.set_facecolor('white')
            ax.plot(energy, spectrum, 'k-', linewidth=1.0)
            ax.set_xlabel('Energy (keV)', color='black')
            ax.set_ylabel('Counts', color='black')
            ax.tick_params(colors='black', labelcolor='black')
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
            return 'black'


    def plot_area_spectrum(self, x1, y1, x2, y2):
        """Plot summed spectrum from area in KherveFitting main window"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for area spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Sum spectrum over area
        spectrum = np.sum(data[y1:y2 + 1, x1:x2 + 1, :], axis=(0, 1))
        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        title = f'Summed EDX Spectrum - Area: {(x2 - x1 + 1) * (y2 - y1 + 1)} pixels'
        title = ' '
        grid_label = f"Area ({x2 - x1 + 1}×{y2 - y1 + 1})"

        # Plot using common EDX plotting method
        self._plot_edx_spectrum_styled(energy, spectrum, title, grid_label)


    def plot_line_spectrum_OLD(self, x1, y1, x2, y2):
        """Plot summed spectrum along line from (x1,y1) to (x2,y2) in KherveFitting main window"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for line spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get points along line using Bresenham-like algorithm
        num_points = max(abs(x2 - x1), abs(y2 - y1)) + 1
        x_coords = np.linspace(x1, x2, num_points).astype(int)
        y_coords = np.linspace(y1, y2, num_points).astype(int)

        # Sum spectrum along line
        spectrum = np.zeros(data.shape[2])
        for px, py in zip(x_coords, y_coords):
            if 0 <= px < data.shape[1] and 0 <= py < data.shape[0]:
                spectrum += data[py, px, :]

        energy = self.get_energy_axis()
        if energy is None:
            energy = np.arange(len(spectrum))

        title = f'EDX Line Spectrum ({x1},{y1}) to ({x2},{y2}) - {num_points} pts'
        grid_label = f"Line ({num_points} pts)"

        # Plot using common EDX plotting method
        self._plot_edx_spectrum_styled(energy, spectrum, title, grid_label)

    def plot_line_spectrum(self, x1, y1, x2, y2):
        """Plot summed spectrum along line from (x1,y1) to (x2,y2) in KherveFitting main window"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for line spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get points along line using Bresenham-like algorithm
        num_points = max(abs(x2 - x1), abs(y2 - y1)) + 1
        x_coords = np.linspace(x1, x2, num_points).astype(int)
        y_coords = np.linspace(y1, y2, num_points).astype(int)

        # Sum spectrum along line
        spectrum = np.zeros(data.shape[2])
        for px, py in zip(x_coords, y_coords):
            if 0 <= px < data.shape[1] and 0 <= py < data.shape[0]:
                spectrum += data[py, px, :]

        energy = self.get_energy_axis()
        if energy is None:
            energy = np.arange(len(spectrum))

        title = f'EDX Line Spectrum ({x1},{y1}) to ({x2},{y2}) - {num_points} pts'
        grid_label = f"Line ({num_points} pts)"

        # Plot using common EDX plotting method
        self._plot_edx_spectrum_styled(energy, spectrum, title, grid_label)

    def create_line_profile(self, x1, y1, x2, y2, line_width=1):
        """
        Create a zzProfile of atomic concentration along a line.
        Each position along the line sums pixels within the specified width,
        then quantifies elements to create a concentration profile.
        Positions are calculated in nm using the data's scale information.
        """
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for line profile extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get scale from navigation axes (nm per pixel)
        scale_per_pixel = 1.0  # Default to 1 if not available
        scale_unit = 'nm'

        if hasattr(self.current_data, 'axes_manager'):
            nav_axes = self.current_data.axes_manager.navigation_axes
            if len(nav_axes) > 0:
                axis = nav_axes[0]
                scale_per_pixel = axis.scale if axis.scale and axis.scale > 0 else 1.0
                scale_unit = axis.units if hasattr(axis, 'units') and axis.units else 'nm'

        # Get number of positions along the line
        num_positions = max(abs(x2 - x1), abs(y2 - y1)) + 1
        x_coords = np.linspace(x1, x2, num_positions)
        y_coords = np.linspace(y1, y2, num_positions)

        # Calculate distance in nm for each position along the line
        line_length_pixels = np.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
        positions_nm = np.linspace(0, line_length_pixels * scale_per_pixel, num_positions)

        # Calculate perpendicular direction for width
        dx = x2 - x1
        dy = y2 - y1
        length = np.sqrt(dx ** 2 + dy ** 2)
        if length > 0:
            px = -dy / length  # Perpendicular unit vector
            py = dx / length
        else:
            px, py = 0, 1

        half_width = line_width / 2

        # Get energy axis
        energy = self.get_energy_axis()
        if energy is None:
            energy = np.arange(data.shape[2])

        # Collect spectra and quantify for each position along the line
        profile_results = []
        all_elements = set()

        print(f"Creating line profile: ({x1},{y1}) to ({x2},{y2}), width={line_width}, positions={num_positions}")
        print(f"Scale: {scale_per_pixel:.4f} {scale_unit}/pixel, Total length: {line_length_pixels * scale_per_pixel:.2f} {scale_unit}")

        for i, (cx, cy, pos_nm) in enumerate(zip(x_coords, y_coords, positions_nm)):
            # Sum pixels within the width at this position
            spectrum = np.zeros(data.shape[2])
            pixel_count = 0

            if line_width == 1:
                # Single pixel
                ix, iy = int(round(cx)), int(round(cy))
                if 0 <= ix < data.shape[1] and 0 <= iy < data.shape[0]:
                    spectrum = data[iy, ix, :].copy()
                    pixel_count = 1
            else:
                # Sum pixels across the width
                for w in np.linspace(-half_width, half_width, max(2, line_width)):
                    ix = int(round(cx + px * w))
                    iy = int(round(cy + py * w))
                    if 0 <= ix < data.shape[1] and 0 <= iy < data.shape[0]:
                        spectrum += data[iy, ix, :]
                        pixel_count += 1

            if pixel_count == 0:
                profile_results.append({'position': i, 'position_nm': pos_nm, 'elements': {}})
                continue

            # Quantify this spectrum
            quant_results = self._quantify_spectrum_for_profile(energy, spectrum)
            all_elements.update(quant_results.keys())
            profile_results.append({'position': i, 'position_nm': pos_nm, 'elements': quant_results})

        if not all_elements:
            wx.MessageBox("No elements detected along the line profile.",
                          "No Elements", wx.OK | wx.ICON_WARNING)
            return

        # Create zzProfile sheet and plot
        self._create_profile_sheet(profile_results, all_elements, x1, y1, x2, y2, line_width, scale_unit)

    def _quantify_spectrum_for_profile(self, energy, spectrum):
        """Quantify a single spectrum and return element atomic percentages.
        Uses only the green-selected elements from the periodic table.
        """
        results = {}

        try:
            # Get the green-selected elements from periodic table
            include_elements = getattr(self, 'include_elements', set())

            if not include_elements:
                # Fallback to selected_elements if include_elements is empty
                include_elements = set(getattr(self, 'selected_elements', []))

            if not include_elements:
                print("No elements selected (green) in periodic table for quantification")
                return results

            # Convert to list for quantification
            elements_list = list(include_elements)

            # Calculate quantification using the selected elements
            quant = self._calculate_quantification(energy, spectrum, elements_list)

            if quant:
                for item in quant:
                    element = item.get('element', item.get('Element', ''))
                    # Get concentration - check multiple possible keys
                    conc = item.get('concentration', item.get('at_percent', item.get('At%', 0)))
                    if element and conc > 0:
                        results[element] = float(f"{conc:.2f}")

        except Exception as e:
            print(f"Error quantifying spectrum for profile: {e}")
            import traceback
            traceback.print_exc()

        return results

    def _create_profile_sheet(self, profile_results, all_elements, x1, y1, x2, y2, line_width, scale_unit='nm'):
        """Create a zzProfile sheet from profile results and plot it.
        Uses proper zzProfile structure compatible with ProfileEditor.py
        Positions are in nm (or whatever scale unit is provided).
        """
        import datetime
        import json
        import os

        if self.parent is None:
            return

        # Find next available profile number
        existing_sheets = list(self.parent.Data.get('Core levels', {}).keys())
        profile_num = 1
        while f'zzProfile{profile_num}' in existing_sheets:
            profile_num += 1

        sheet_name = f'zzProfile{profile_num}'

        # Build profile data structure with x_values and y_values format
        profile_data = {}

        # Number column (position along line in nm)
        positions_nm = [r['position_nm'] for r in profile_results]
        profile_data['Number'] = {
            'x_values': list(range(len(positions_nm))),
            'y_values': [float(f"{p:.2f}") for p in positions_nm]
        }

        # Element columns
        sorted_elements = sorted(all_elements)
        first_element_col = None
        first_element_values = []

        for element in sorted_elements:
            col_name = f"{element} At(%)"
            x_values = []
            y_values = []

            for i, result in enumerate(profile_results):
                if element in result['elements']:
                    x_values.append(i)
                    y_values.append(float(f"{result['elements'][element]:.2f}"))

            profile_data[col_name] = {
                'x_values': x_values,
                'y_values': y_values
            }

            # Store first element for B.E. and Raw Data
            if first_element_col is None and y_values:
                first_element_col = col_name
                # Build full array for Raw Data
                first_element_values = [0.0] * len(positions_nm)
                for x, y in zip(x_values, y_values):
                    if 0 <= x < len(first_element_values):
                        first_element_values[x] = y

        # Create sheet data structure matching ProfileEditor.py format
        sheet_data = {
            'Name': sheet_name,
            'B.E.': [float(f"{p:.2f}") for p in positions_nm],  # Use positions in nm as B.E.
            'Raw Data': first_element_values if first_element_values else [0.0] * len(positions_nm),
            'Profile Data': profile_data,
            'Y_Axis_Label': 'Atomic Concentration (%)',
            'X_Axis_Label': f'Position ({scale_unit})',
            'Profile_Type': 'EDX Line Profile',
            'Background': {
                'Bkg Type': '',
                'Bkg Low': '',
                'Bkg High': '',
                'Bkg Offset Low': '',
                'Bkg Offset High': '',
                'Bkg Y': first_element_values if first_element_values else [0.0] * len(positions_nm)
            },
            # EDX-specific metadata
            '_EDX_type': 'line_profile',
            '_EDX_line_start': [x1, y1],
            '_EDX_line_end': [x2, y2],
            '_EDX_line_width': line_width,
            '_EDX_num_positions': len(profile_results),
            '_EDX_elements': list(sorted_elements),
            '_EDX_scale_unit': scale_unit,
            '_EDX_save_time': datetime.datetime.now().isoformat()
        }

        # Store in parent Data
        if 'Core levels' not in self.parent.Data:
            self.parent.Data['Core levels'] = {}

        self.parent.Data['Core levels'][sheet_name] = sheet_data

        # Update sheet combobox
        if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                              for i in range(self.parent.sheet_combobox.GetCount())]:
            self.parent.sheet_combobox.Append(sheet_name)

        # Select the new sheet
        self.parent.sheet_combobox.SetValue(sheet_name)

        # Plot the profile
        if hasattr(self.parent, 'plot_manager'):
            self.parent.plot_manager.plot_profile(self.parent)
            self.parent.canvas.draw()

        print(f"Created line profile: {sheet_name}")
        print(f"  Line: ({x1},{y1}) to ({x2},{y2}), width={line_width}")
        print(f"  Positions: {len(profile_results)}, Elements: {sorted_elements}")
        print(f"  X-axis: Position ({scale_unit})")

        # Save to Excel
        self._save_profile_to_excel(sheet_name, profile_data, sorted_elements)

        # Save to JSON
        self._save_profile_to_json()

    def _save_profile_to_json(self):
        """Save the current Data structure to JSON file"""
        try:
            import json
            import os

            if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
                return

            file_path = self.parent.Data['FilePath']
            json_file_path = os.path.splitext(file_path)[0] + '.json'

            # Import the serialization function from Save module
            from libraries.FileMenu.Save import convert_to_serializable_and_round

            json_data = convert_to_serializable_and_round(self.parent.Data)

            with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)

            print(f"Saved profile to JSON: {json_file_path}")

        except Exception as e:
            print(f"Error saving profile to JSON: {e}")
            import traceback
            traceback.print_exc()

    def _save_profile_to_excel(self, sheet_name, profile_data, elements):
        """Save profile data to Excel file"""
        try:
            import pandas as pd

            if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
                return

            file_path = self.parent.Data['FilePath']

            # Build DataFrame
            df_data = {}

            # Get positions
            if 'Number' in profile_data:
                num_info = profile_data['Number']
                max_pos = max(num_info['x_values']) + 1 if num_info['x_values'] else 0

                # Create position array
                positions = [None] * max_pos
                for x, y in zip(num_info['x_values'], num_info['y_values']):
                    if 0 <= x < max_pos:
                        positions[x] = f"{y:.2f}"
                df_data['Number'] = positions

                # Add element columns
                for element in elements:
                    col_name = f"{element} At(%)"
                    if col_name in profile_data:
                        col_info = profile_data[col_name]
                        values = [None] * max_pos
                        for x, y in zip(col_info['x_values'], col_info['y_values']):
                            if 0 <= x < max_pos:
                                values[x] = f"{y:.2f}"
                        df_data[col_name] = values

            df = pd.DataFrame(df_data)

            # Append to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a',
                                if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Saved profile to Excel: {sheet_name}")

        except Exception as e:
            print(f"Error saving profile to Excel: {e}")

    def _add_scale_bar(self):
        """Add scale bar to the map based on current view and axes scale information"""
        if self.current_data is None:
            return

        # Remove existing scale bar elements
        self._remove_scale_bar()

        # Get scale from navigation axes
        scale_per_pixel = None
        scale_unit = 'nm'

        if hasattr(self.current_data, 'axes_manager'):
            nav_axes = self.current_data.axes_manager.navigation_axes
            if len(nav_axes) > 0:
                axis = nav_axes[0]
                scale_per_pixel = axis.scale
                scale_unit = axis.units if hasattr(axis, 'units') and axis.units else 'nm'

        if scale_per_pixel is None or scale_per_pixel <= 0:
            return

        # Get current view limits (for zoomed view)
        xlim = self.map_ax.get_xlim()
        ylim = self.map_ax.get_ylim()

        # Calculate visible width in pixels and nm
        visible_width_pixels = abs(xlim[1] - xlim[0])
        visible_height_pixels = abs(ylim[1] - ylim[0])
        visible_width_nm = visible_width_pixels * scale_per_pixel

        # Choose a nice round number for scale bar (~20% of visible width)
        target_length_nm = visible_width_nm * 0.2
        nice_lengths = [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000, 10000]
        scale_bar_nm = min(nice_lengths, key=lambda x: abs(x - target_length_nm))

        # Convert to pixels
        scale_bar_pixels = scale_bar_nm / scale_per_pixel

        # Position: bottom right corner of current view with margin
        margin_x = visible_width_pixels * 0.05
        margin_y = visible_height_pixels * 0.05

        # Handle inverted Y axis
        y_min, y_max = min(ylim), max(ylim)
        x_min, x_max = min(xlim), max(xlim)

        x_start = x_max - margin_x - scale_bar_pixels
        y_pos = y_max - margin_y  # Bottom of view

        # Draw scale bar
        bar_height = visible_height_pixels * 0.02
        from matplotlib.patches import Rectangle
        import matplotlib.patheffects as path_effects

        # Black outline
        outline = Rectangle((x_start - 1, y_pos - bar_height / 2 - 1),
                            scale_bar_pixels + 2, bar_height + 2,
                            facecolor='black', edgecolor='none', zorder=100)
        outline._is_scale_bar = True
        self.map_ax.add_patch(outline)

        # White bar
        bar = Rectangle((x_start, y_pos - bar_height / 2),
                        scale_bar_pixels, bar_height,
                        facecolor='white', edgecolor='none', zorder=101)
        bar._is_scale_bar = True
        self.map_ax.add_patch(bar)

        # Format scale label
        if scale_bar_nm >= 1000:
            label = f'{scale_bar_nm / 1000:.0f} µm'
        else:
            label = f'{scale_bar_nm:.0f} {scale_unit}'

        # Add label above scale bar
        text = self.map_ax.text(x_start + scale_bar_pixels / 2, y_pos - bar_height - margin_y * 0.5,
                                label, ha='center', va='top', fontsize=9, fontweight='bold',
                                color='white', zorder=102,
                                path_effects=[path_effects.withStroke(linewidth=2, foreground='black')])
        text._is_scale_bar = True

    def _remove_scale_bar(self):
        """Remove existing scale bar elements"""
        # Remove patches
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_scale_bar') and artist._is_scale_bar:
                artist.remove()
        # Remove texts
        for artist in self.map_ax.texts[:]:
            if hasattr(artist, '_is_scale_bar') and artist._is_scale_bar:
                artist.remove()

    def _update_scale_bar(self):
        """Update scale bar after view change"""
        self._add_scale_bar()
        self.map_canvas.draw_idle()


    def get_energy_axis(self):
        """Get energy axis from current data"""
        if self.current_data is None:
            return None

        if hasattr(self.current_data, 'axes_manager'):
            # Find the signal axis (energy)
            if len(self.current_data.axes_manager.signal_axes) > 0:
                return self.current_data.axes_manager.signal_axes[0].axis

        # Fallback: create channel numbers
        if len(self.current_data.data.shape) == 3:
            return np.arange(self.current_data.data.shape[2])
        elif len(self.current_data.data.shape) == 1:
            return np.arange(len(self.current_data.data))

        return None

    # ==================== FILE IMPORT ====================

    def on_import_sem(self, event):
        """Import SEM image"""
        with wx.FileDialog(self, "Open SEM Image",
                          wildcard="All files (*.*)|*.*|TIFF files (*.tif;*.tiff)|*.tif;*.tiff|"
                                  "HDF5 files (*.hdf5;*.h5)|*.hdf5;*.h5|"
                                  "BMP files (*.bmp)|*.bmp|PNG files (*.png)|*.png",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            file_path = fileDialog.GetPath()

        self.load_file(file_path, 'sem')

    def on_import_edx_spectrum(self, event):
        """Import single EDX spectrum"""
        with wx.FileDialog(self, "Open EDX Spectrum",
                          wildcard="All files (*.*)|*.*|EMSA files (*.emsa;*.ems)|*.emsa;*.ems|"
                                  "SPD files (*.spd)|*.spd|MSA files (*.msa)|*.msa",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            file_path = fileDialog.GetPath()

        self.load_file(file_path, 'edx_spectrum')

    def on_import_edx_map(self, event):
        """Import EDX map or spectrum image"""
        wildcard = "All supported files|*.hdf5;*.h5;*.bcf;*.rpl;*.raw;*.emd;*.ser|" \
                   "HDF5 files (*.hdf5)|*.hdf5|" \
                   "HDF5 files (*.h5)|*.h5|" \
                   "Bruker BCF files (*.bcf)|*.bcf|" \
                   "All files (*.*)|*.*"

        with wx.FileDialog(self, "Open EDX Map file",
                           wildcard=wildcard,
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
                self.load_file(file_path, 'EDX Map')

                # Create excel file and add to window.data
                if self.current_data is not None:
                    self.create_edx_map_output(file_path)

    def create_edx_map_output(self, file_path):
        """Create JSON and HDF5 files, and add EDX map data to parent window.Data"""
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import shutil
        import json

        try:
            # ========== Initialize parent Data structure if needed ==========
            if self.parent is not None:
                if not hasattr(self.parent, 'Data'):
                    from libraries.ConfigFile import Init_Measurement_Data
                    self.parent.Data = Init_Measurement_Data(self.parent)

                if 'Core levels' not in self.parent.Data:
                    self.parent.Data['Core levels'] = {}

            base_name = os.path.splitext(os.path.basename(file_path))[0]
            excel_path = os.path.join(os.path.dirname(file_path), f"{base_name}_EDX.xlsx")
            json_path = os.path.join(os.path.dirname(file_path), f"{base_name}_EDX.json")
            hdf5_copy_path = os.path.join(os.path.dirname(file_path), f"{base_name}_EDX.hdf5")

            # Set FilePath EARLY so it's available for other operations
            if self.parent is not None and hasattr(self.parent, 'Data'):
                self.parent.Data['FilePath'] = excel_path

                # Update current_file_path
                if hasattr(self.parent, 'current_file_path'):
                    self.parent.current_file_path = excel_path

                # Update Working_directory
                if hasattr(self.parent, 'Working_directory'):
                    self.parent.Working_directory = os.path.dirname(excel_path)

            # Copy original HDF5 file
            if file_path.lower().endswith(('.hdf5', '.h5')):
                shutil.copy2(file_path, hdf5_copy_path)
                print(f"HDF5 copy saved to: {hdf5_copy_path}")

            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Get energy range
            energy_axis = self.get_energy_axis()
            if energy_axis is not None:
                energy_min = f"{np.min(energy_axis):.2f}"
                energy_max = f"{np.max(energy_axis):.2f}"
                energy_range = f"{energy_min} - {energy_max} keV"
            else:
                energy_range = "N/A"

            # Create sum spectrum
            sum_signal = self.current_data.sum()
            spectrum_data = sum_signal.data

            # EDX~Plot sheet
            ws_plot = wb.create_sheet("EDX~Plot")
            ws_plot.append(['Energy (keV)', 'Intensity', f'Range: {energy_range}'])

            for i, intensity in enumerate(spectrum_data):
                if energy_axis is not None:
                    ws_plot.append([f"{energy_axis[i]:.2f}", f"{intensity:.2f}"])
                else:
                    ws_plot.append([f"{i:.2f}", f"{intensity:.2f}"])

            # EDX~Map sheet
            ws_map = wb.create_sheet("EDX~Map")
            map_data = np.sum(self.current_data.data, axis=2)

            ws_map.append([f'EDX Intensity Map - Range: {energy_range}'])
            ws_map.append([''] * (map_data.shape[1] + 1))

            for row in map_data:
                ws_map.append([f"{val:.2f}" for val in row])

            # Create and save map image at 100 dpi
            fig, ax = plt.subplots(figsize=(map_data.shape[1] / 100, map_data.shape[0] / 100), dpi=100)
            im = ax.imshow(map_data, cmap=self.current_cmap)
            ax.set_title(f'EDX Map - {energy_range}')
            plt.colorbar(im, ax=ax)
            ax.axis('off')

            img_buffer = BytesIO()
            fig.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close(fig)

            img = OpenpyxlImage(img_buffer)
            ws_map.add_image(img, f'A{map_data.shape[0] + 5}')

            # Save Excel file
            wb.save(excel_path)
            print(f"EDX data exported to: {excel_path}")

            # ========== Add to parent window.Data['Core levels'] ==========
            if self.parent is not None and hasattr(self.parent, 'Data'):
                if 'Core levels' not in self.parent.Data:
                    self.parent.Data['Core levels'] = {}

                energy_values = energy_axis if energy_axis is not None else np.arange(len(spectrum_data))

                # Get element preferences
                include_elements = list(getattr(self, 'include_elements', set()))
                exclude_elements = list(getattr(self, 'exclude_elements', set()))

                # Add EDX~Plot
                self.parent.Data['Core levels']['EDX~Plot'] = {
                    'Name': 'EDX~Plot',
                    'B.E.': list(energy_values),
                    'Raw Data': list(spectrum_data),
                    '_EDX_display_max': 20,
                    '_EDX_type': 'plot',
                    '_EDX_include_elements': include_elements,
                    '_EDX_exclude_elements': exclude_elements,
                    'Background': {}
                }

                # Add EDX~Map (HDF5 path is derived from FilePath, not stored)
                self.parent.Data['Core levels']['EDX~Map'] = {
                    'Name': 'EDX~Map',
                    'Map_Intensity': map_data.tolist(),
                    'Map_Shape': list(map_data.shape),
                    'Energy_Range': energy_range,
                    '_EDX_type': 'map',
                    '_EDX_include_elements': include_elements,
                    '_EDX_exclude_elements': exclude_elements
                }

                # FilePath was already set at the beginning, now update UI
                # Update status bar
                if hasattr(self.parent, 'SetStatusText'):
                    self.parent.SetStatusText(f"Working Directory: {os.path.dirname(excel_path)}", 0)

                # Update window title
                if hasattr(self.parent, 'SetTitle'):
                    self.parent.SetTitle(f"KherveFitting - {os.path.basename(excel_path)}")

                # Add sheets to combobox
                for sheet_name in ['EDX~Plot', 'EDX~Map']:
                    if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                                          for i in range(self.parent.sheet_combobox.GetCount())]:
                        self.parent.sheet_combobox.Append(sheet_name)

            # ========== Create JSON file ==========
            json_data = {
                'FilePath': excel_path,
                'Core levels': {}
            }

            # Get element preferences
            include_elements = list(getattr(self, 'include_elements', set()))
            exclude_elements = list(getattr(self, 'exclude_elements', set()))

            # Add EDX~Plot to JSON
            json_data['Core levels']['EDX~Plot'] = {
                'Name': 'EDX~Plot',
                'B.E.': [float(f"{v:.2f}") for v in energy_values],
                'Raw Data': [float(f"{v:.2f}") for v in spectrum_data],
                '_EDX_display_max': 20,
                '_EDX_type': 'plot',
                '_EDX_include_elements': include_elements,
                '_EDX_exclude_elements': exclude_elements
            }

            # Add EDX~Map to JSON (HDF5 path is derived from FilePath, not stored)
            json_data['Core levels']['EDX~Map'] = {
                'Name': 'EDX~Map',
                'Map_Intensity': [[float(f"{val:.2f}") for val in row] for row in map_data],
                'Map_Shape': list(map_data.shape),
                'Energy_Range': energy_range,
                '_EDX_type': 'map',
                '_EDX_include_elements': include_elements,
                '_EDX_exclude_elements': exclude_elements
            }

            with open(json_path, 'w') as jf:
                json.dump(json_data, jf, indent=2)
            print(f"JSON data saved to: {json_path}")

            wx.MessageBox(f"EDX data exported to:\n{excel_path}\n\nJSON: {json_path}\n" +
                          (f"HDF5: {hdf5_copy_path}" if os.path.exists(hdf5_copy_path) else ""),
                          "Export Complete", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Error creating EDX output:\n{str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def load_element_preferences_from_data(self):
        """Load element preferences from parent Data structure if available"""
        try:
            if self.parent is None or not hasattr(self.parent, 'Data'):
                return

            if 'Core levels' not in self.parent.Data:
                return

            # Check EDX sheets for saved preferences
            preferences_found = False
            for sheet_name in ['EDX~Plot', 'EDX~Map']:
                if sheet_name in self.parent.Data['Core levels']:
                    sheet_data = self.parent.Data['Core levels'][sheet_name]

                    if '_EDX_include_elements' in sheet_data or '_EDX_exclude_elements' in sheet_data:
                        preferences_found = True

                        if '_EDX_include_elements' in sheet_data:
                            self.include_elements = set(sheet_data['_EDX_include_elements'])
                            print(f"Loaded include elements: {self.include_elements}")

                        if '_EDX_exclude_elements' in sheet_data:
                            self.exclude_elements = set(sheet_data['_EDX_exclude_elements'])
                            print(f"Loaded exclude elements: {self.exclude_elements}")

                        # Also set legacy selected_elements
                        self.selected_elements = list(self.include_elements)
                        break

            # If no saved preferences found, save the defaults to data
            if not preferences_found:
                print("No saved element preferences found, using defaults")
                self._save_element_preferences()

        except Exception as e:
            print(f"Could not load element preferences: {e}")

    def on_import_bcf(self, event):
        """Import Bruker BCF file"""
        with wx.FileDialog(self, "Open Bruker BCF File",
                          wildcard="BCF files (*.bcf)|*.bcf",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            file_path = fileDialog.GetPath()

        self.load_file(file_path, 'bcf')

    def load_file_OLD(self, file_path, data_type):
        """Generic file loader with automatic reader selection"""
        try:
            loaded_data = None

            # List of readers to try
            readers_to_try = ['HSPY', 'USID', 'Delmic', 'EMSA', 'Bruker']

            # Try each reader in order
            for reader in readers_to_try:
                try:
                    loaded_data = hs.load(file_path, reader=reader)
                    print(f"Successfully loaded with {reader} reader")
                    break
                except Exception as e:
                    print(f"Failed with {reader} reader: {e}")
                    continue

            # If all readers failed, try without specifying reader
            if loaded_data is None:
                try:
                    loaded_data = hs.load(file_path)
                    print("Successfully loaded with default reader")
                except Exception as e:
                    wx.MessageBox(f"Could not load file with any available reader.\nError: {str(e)}",
                                  "Load Error", wx.OK | wx.ICON_ERROR)
                    return

            # Handle list of signals
            if isinstance(loaded_data, list):
                if len(loaded_data) == 1:
                    loaded_data = loaded_data[0]
                else:
                    self.add_multiple_signals_to_tree(loaded_data, file_path, data_type)
                    return

            # Store signal
            self.loaded_signals.append({
                'data': loaded_data,
                'filename': os.path.basename(file_path),
                'path': file_path,
                'type': data_type
            })

            self.current_data = loaded_data

            # Update data browser if open
            if hasattr(self, 'data_browser_window') and self.data_browser_window is not None:
                self.data_browser_window.refresh_tree()

            self.plot_current_map()
            self.update_info_text(loaded_data)



            # Extract elements if available
            self.extract_elements(loaded_data)

            # Load saved element preferences
            self.load_element_preferences_from_data()

        except Exception as e:
            wx.MessageBox(f"Error loading file:\n{str(e)}",
                         "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def load_file(self, file_path, data_type):
        """Generic file loader - uses standalone readers"""
        try:
            loaded_data = None
            ext = os.path.splitext(file_path)[1].lower()

            # Try standalone readers first
            if ext == '.bcf':
                try:
                    from libraries.BCF_Reader import load_bcf
                    bcf_data = load_bcf(file_path)
                    if bcf_data.maps:
                        loaded_data = self._convert_bcf_to_signal(bcf_data.maps[0])
                    elif bcf_data.spectra:
                        loaded_data = self._convert_bcf_to_signal(bcf_data.spectra[0])
                    print("Successfully loaded with BCF_Reader")
                except Exception as e:
                    print(f"BCF_Reader failed: {e}")
                    loaded_data = None

            elif ext in ['.emsa', '.msa']:
                try:
                    from libraries.EDX_Utilities import load_emsa
                    emsa_data = load_emsa(file_path)
                    loaded_data = self._convert_dict_to_signal(emsa_data)
                    print("Successfully loaded with EDX_Utilities")
                except Exception as e:
                    print(f"EMSA reader failed: {e}")
                    loaded_data = None

            elif ext in ['.hdf5', '.h5']:
                try:
                    loaded_data = self._load_hdf5_edx(file_path)
                    print("Successfully loaded HDF5 file")
                except Exception as e:
                    print(f"HDF5 reader failed: {e}")
                    loaded_data = None

            if loaded_data is None:
                wx.MessageBox(f"Could not load file. Supported formats: BCF, EMSA/MSA, HDF5",
                              "Load Error", wx.OK | wx.ICON_ERROR)
                return

            # Handle list of signals
            if isinstance(loaded_data, list):
                if len(loaded_data) == 1:
                    loaded_data = loaded_data[0]
                else:
                    self.add_multiple_signals_to_tree(loaded_data, file_path, data_type)
                    return

            # Store signal
            self.loaded_signals.append({
                'data': loaded_data,
                'filename': os.path.basename(file_path),
                'path': file_path,
                'type': data_type
            })

            self.current_data = loaded_data

            if hasattr(self, 'data_browser_window') and self.data_browser_window is not None:
                self.data_browser_window.refresh_tree()

            self.plot_current_map()
            self.update_info_text(loaded_data)
            self.extract_elements(loaded_data)
            self.load_element_preferences_from_data()

        except Exception as e:
            wx.MessageBox(f"Error loading file:\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def _load_hdf5_edx(self, file_path):
        """Load EDX data from HDF5 file directly"""
        import h5py
        from libraries.EDX_Utilities import Signal1D

        with h5py.File(file_path, 'r') as f:
            # Common HDF5 structures for EDX data
            data_paths = [
                'EDX/data', 'edx/data', 'Data/data',
                'Experiments/EDX/data', 'entry/data/data',
                'spectrum', 'data', 'counts'
            ]

            energy_paths = [
                'EDX/energy', 'edx/energy', 'Data/energy',
                'Experiments/EDX/energy', 'entry/data/energy',
                'energy', 'axis', 'x'
            ]

            data = None
            energy_axis = None

            # Find data
            for path in data_paths:
                if path in f:
                    data = np.array(f[path])
                    print(f"Found data at: {path}")
                    break

            # If not found, search for largest dataset
            if data is None:
                def find_largest_dataset(group):
                    largest = None
                    largest_size = 0
                    for key in group.keys():
                        item = group[key]
                        if isinstance(item, h5py.Dataset):
                            if item.size > largest_size and item.ndim >= 1:
                                largest = np.array(item)
                                largest_size = item.size
                        elif isinstance(item, h5py.Group):
                            sub_largest = find_largest_dataset(item)
                            if sub_largest is not None and sub_largest.size > largest_size:
                                largest = sub_largest
                                largest_size = sub_largest.size
                    return largest

                data = find_largest_dataset(f)

            if data is None:
                raise ValueError("No data found in HDF5 file")

            # Find energy axis
            for path in energy_paths:
                if path in f:
                    energy_axis = np.array(f[path])
                    print(f"Found energy at: {path}")
                    break

            # Generate energy axis if not found
            if energy_axis is None:
                if data.ndim == 3:
                    n_channels = data.shape[2]
                else:
                    n_channels = data.shape[-1]
                # Default 10 eV per channel (0.01 keV), typical for EDX with ~4000 channels
                # This gives 0-20 keV range for 4000 channels
                energy_axis = np.arange(n_channels) * 0.01
                print(f"Generated default energy axis: 0 - {(n_channels - 1) * 0.01:.2f} keV ({n_channels} channels)")

            # Check if energy axis needs scaling to keV
            max_energy = np.max(energy_axis)
            if max_energy > 1000:
                energy_axis = energy_axis / 1000.0
                print(f"Divided by 1000 (eV to keV)")
            elif max_energy > 100:
                energy_axis = energy_axis / 10.0
                print(f"Divided by 10")

            print(f"Final energy range: {energy_axis.min():.2f} - {energy_axis.max():.2f} keV")

            # Create Signal1D object
            signal = Signal1D(data)

            # IMPORTANT: Set the energy axis scale and offset
            # For 3D data, energy is the LAST axis (index -1 or signal_axes[0])
            scale = energy_axis[1] - energy_axis[0] if len(energy_axis) > 1 else 0.01
            offset = energy_axis[0]

            # Get the index of the last axis (energy axis)
            last_axis_idx = len(signal.axes_manager) - 1

            signal.axes_manager[last_axis_idx].scale = scale
            signal.axes_manager[last_axis_idx].offset = offset
            signal.axes_manager[last_axis_idx].units = 'keV'
            signal.axes_manager[last_axis_idx].name = 'Energy'

            # Store energy axis in metadata for direct access
            signal.metadata['energy_axis'] = energy_axis

            print(f"Signal created with scale={scale:.6f}, offset={offset:.2f}, data shape={data.shape}")

            return signal

    def _convert_bcf_to_signal_OLD(self, bcf_obj):
        """Convert BCF data to signal-like object"""
        from libraries.EDX_Utilities import Signal1D

        if hasattr(bcf_obj, 'energy') and bcf_obj.energy is not None:
            signal = Signal1D(bcf_obj.data)

            # Set up energy axis on the LAST axis
            if len(bcf_obj.energy) > 1:
                scale = bcf_obj.energy[1] - bcf_obj.energy[0]
                offset = bcf_obj.energy[0]
            else:
                scale = 0.005
                offset = 0

            last_axis_idx = len(signal.axes_manager) - 1
            signal.axes_manager[last_axis_idx].scale = scale
            signal.axes_manager[last_axis_idx].offset = offset
            signal.axes_manager[last_axis_idx].units = 'keV'
            signal.axes_manager[last_axis_idx].name = 'Energy'

            signal.metadata = bcf_obj.metadata
            return signal
        else:
            return Signal1D(bcf_obj.data)

    def _convert_bcf_to_signal(self, bcf_obj):
        """Convert BCF data to signal-like object"""
        from libraries.EDX_Utilities import Signal1D

        if hasattr(bcf_obj, 'energy') and bcf_obj.energy is not None:
            signal = Signal1D(bcf_obj.data)

            # Set up energy axis on the LAST axis
            if len(bcf_obj.energy) > 1:
                scale = bcf_obj.energy[1] - bcf_obj.energy[0]
                offset = bcf_obj.energy[0]
            else:
                scale = 0.005
                offset = 0

            last_axis_idx = len(signal.axes_manager) - 1
            signal.axes_manager[last_axis_idx].scale = scale
            signal.axes_manager[last_axis_idx].offset = offset
            signal.axes_manager[last_axis_idx].units = 'keV'
            signal.axes_manager[last_axis_idx].name = 'Energy'

            # Set navigation axes scale from BCF pixel_size (µm -> nm)
            # BCF pixel_size is in µm, we want nm for display
            if signal.data.ndim == 3:  # 3D hyperspectral data
                nav_axes = signal.axes_manager.navigation_axes
                # Get pixel sizes from BCF (default 1.0 µm)
                pixel_size_x = getattr(bcf_obj, 'pixel_size_x', 1.0)
                pixel_size_y = getattr(bcf_obj, 'pixel_size_y', 1.0)

                # Convert µm to nm (1 µm = 1000 nm)
                scale_nm_x = pixel_size_x * 1000.0
                scale_nm_y = pixel_size_y * 1000.0

                if len(nav_axes) >= 2:
                    nav_axes[0].scale = scale_nm_y  # Y axis (first nav axis)
                    nav_axes[0].units = 'nm'
                    nav_axes[0].name = 'Y'
                    nav_axes[1].scale = scale_nm_x  # X axis (second nav axis)
                    nav_axes[1].units = 'nm'
                    nav_axes[1].name = 'X'
                elif len(nav_axes) == 1:
                    nav_axes[0].scale = scale_nm_x
                    nav_axes[0].units = 'nm'
                    nav_axes[0].name = 'X'

            signal.metadata = bcf_obj.metadata
            return signal
        else:
            return Signal1D(bcf_obj.data)

    def _convert_dict_to_signal(self, data_dict):
        """Convert dict with 'data' and 'energy' to signal-like object"""
        from libraries.EDX_Utilities import Signal1D

        signal = Signal1D(data_dict['data'])
        energy = data_dict.get('energy')

        if energy is not None and len(energy) > 1:
            scale = energy[1] - energy[0]
            offset = energy[0]
        else:
            scale = 0.005
            offset = 0

        # Set on LAST axis (energy axis)
        last_axis_idx = len(signal.axes_manager) - 1
        signal.axes_manager[last_axis_idx].scale = scale
        signal.axes_manager[last_axis_idx].offset = offset
        signal.axes_manager[last_axis_idx].units = 'keV'
        signal.axes_manager[last_axis_idx].name = 'Energy'

        signal.metadata = data_dict.get('metadata', {})
        return signal

    def add_signal_to_tree(self, signal, file_path, data_type):
        """Add a signal to the data tree"""
        filename = os.path.basename(file_path)

        # Get signal title if available
        title = filename
        if hasattr(signal, 'metadata') and hasattr(signal.metadata, 'General'):
            if hasattr(signal.metadata.General, 'title') and signal.metadata.General.title:
                title = signal.metadata.General.title

        # Create tree item
        item_text = f"{title} ({signal.data.shape})"
        item = self.data_tree.AppendItem(self.tree_root, item_text)
        self.data_tree.SetItemData(item, {'type': data_type, 'data': signal, 'path': file_path})

        # Add metadata as children
        self.add_metadata_to_tree(item, signal)

        # Expand tree
        self.data_tree.Expand(self.tree_root)
        self.data_tree.SelectItem(item)

    def add_multiple_signals_to_tree(self, signals, file_path, data_type):
        """Add multiple signals to tree"""
        filename = os.path.basename(file_path)

        # Create parent item for file
        file_item = self.data_tree.AppendItem(self.tree_root, filename)
        self.data_tree.SetItemData(file_item, {'type': 'container', 'path': file_path})

        for i, signal in enumerate(signals):
            title = f"Signal {i}"
            if hasattr(signal, 'metadata') and hasattr(signal.metadata, 'General'):
                if hasattr(signal.metadata.General, 'title') and signal.metadata.General.title:
                    title = signal.metadata.General.title

            item_text = f"{title} ({signal.data.shape})"
            item = self.data_tree.AppendItem(file_item, item_text)
            self.data_tree.SetItemData(item, {'type': data_type, 'data': signal, 'path': file_path})

            # Add metadata
            self.add_metadata_to_tree(item, signal)

        self.data_tree.Expand(self.tree_root)
        self.data_tree.Expand(file_item)

        # Select first signal
        if signals:
            self.current_data = signals[0]
            self.plot_current_map()
            self.update_info_text(signals[0])
            self.extract_elements(signals[0])

    def add_metadata_to_tree(self, parent_item, signal):
        """Add signal metadata as tree children"""
        if not hasattr(signal, 'metadata'):
            return

        # Add shape info
        shape_item = self.data_tree.AppendItem(parent_item, f"Shape: {signal.data.shape}")

        # Add data type
        dtype_item = self.data_tree.AppendItem(parent_item, f"Dtype: {signal.data.dtype}")

        # Add axes info
        if hasattr(signal, 'axes_manager'):
            axes_item = self.data_tree.AppendItem(parent_item, "Axes")
            for i, axis in enumerate(signal.axes_manager.signal_axes):
                axis_text = f"Signal {i}: {axis.name} [{axis.units}] - {axis.size} pts"
                self.data_tree.AppendItem(axes_item, axis_text)
            for i, axis in enumerate(signal.axes_manager.navigation_axes):
                axis_text = f"Nav {i}: {axis.name} [{axis.units}] - {axis.size} pts"
                self.data_tree.AppendItem(axes_item, axis_text)

        # Add elements if available
        if hasattr(signal, 'metadata') and hasattr(signal.metadata, 'Sample'):
            if hasattr(signal.metadata.Sample, 'elements'):
                elements = signal.metadata.Sample.elements
                elem_item = self.data_tree.AppendItem(parent_item, f"Elements: {', '.join(elements)}")

    def extract_elements(self, signal):
        """Extract element information from signal"""
        if hasattr(signal, 'metadata') and hasattr(signal.metadata, 'Sample'):
            if hasattr(signal.metadata.Sample, 'elements'):
                self.selected_elements = list(signal.metadata.Sample.elements)
            else:
                self.selected_elements = []
        else:
            self.selected_elements = []

    def update_info_text(self, signal):
        """Update info window if open"""
        if hasattr(self, 'info_window') and self.info_window is not None:
            try:
                self.info_window.update_info(signal)
            except:
                pass

    # ==================== MAP PLOTTING ====================

    def plot_current_map(self):
        """Plot the current data as a map"""
        if self.current_data is None:
            return

        # Clear previous colorbar
        if self.current_colorbar is not None:
            try:
                self.current_colorbar.remove()
            except:
                pass
            self.current_colorbar = None

        self.map_ax.clear()
        data = self.current_data.data
        cmap = self.current_cmap

        if len(data.shape) == 2:
            sum_image = data
            title = 'SEM/EDX Image'
        elif len(data.shape) == 3:
            sum_image = np.sum(data, axis=2)
            title = f'Sum Image ({data.shape[0]}×{data.shape[1]} px, {data.shape[2]} ch)'
        elif len(data.shape) == 4:
            sum_image = np.sum(data, axis=(2, 3))
            title = f'Sum Image (4D: {data.shape})'
        else:
            self.map_ax.text(0.5, 0.5, f'Unsupported shape: {data.shape}',
                             ha='center', va='center', transform=self.map_ax.transAxes)
            self.map_canvas.draw()
            return

        im = self.map_ax.imshow(sum_image, cmap=cmap, origin='upper')

        # Remove axis labels and title - use scale bar instead
        self.map_ax.set_xlabel('')
        self.map_ax.set_ylabel('')
        self.map_ax.set_title('')
        self.map_ax.set_xticks([])
        self.map_ax.set_yticks([])

        # Add scale bar if scale information is available
        # self._add_scale_bar(sum_image.shape)
        self._add_scale_bar()

        # Add colorbar matching heatmap style (full height on right side)
        from mpl_toolkits.axes_grid1 import make_axes_locatable
        divider = make_axes_locatable(self.map_ax)
        cax = divider.append_axes("right", size="5%", pad=0.05)
        self.current_colorbar = self.map_figure.colorbar(im, cax=cax)

        self.map_figure.tight_layout(pad=0.5)

        # Add scale bar
        self._add_scale_bar()

        self.map_canvas.draw()

        # Plot sum spectrum in parent KherveFitting window
        self.plot_sum_spectrum_to_parent()

        # # Populate peak fitting grid with quantification
        # self._populate_edx_grid(energy, spectrum, "Sum Spectrum")

        self.reinitialize_selectors()

    def plot_sum_spectrum(self):
        """Plot sum spectrum from current data"""
        if self.current_data is None:
            return

        data = self.current_data.data
        if len(data.shape) != 3:
            return

        # Sum over spatial dimensions
        spectrum = np.sum(data, axis=(0, 1))
        energy = self.get_energy_axis()

        self.spectrum_ax.clear()
        self.spectrum_ax.plot(energy, spectrum, 'b-', linewidth=0.8, label='Sum Spectrum')
        self.spectrum_ax.set_xlabel('Energy (keV)')
        self.spectrum_ax.set_ylabel('Counts')
        self.spectrum_ax.set_title('Total Sum Spectrum')
        self.spectrum_ax.grid(True, alpha=0.3)
        self.spectrum_figure.tight_layout()
        self.spectrum_canvas.draw()

    def on_tree_select(self, event):
        """Handle tree selection"""
        item = event.GetItem()
        data = self.data_tree.GetItemData(item)

        if data is None:
            return

        if data.get('type') == 'container':
            return

        signal = data.get('data')
        if signal is not None:
            self.current_data = signal
            self.plot_current_map()
            self.update_info_text(signal)
            self.extract_elements(signal)

    def on_set_elements(self, event):
        """Set elements using periodic table dialog with three states"""
        # Get current element preferences
        include_elements = getattr(self, 'include_elements', set())
        exclude_elements = getattr(self, 'exclude_elements', set())

        dlg = PeriodicTableDialog(self, include_elements, exclude_elements)

        if dlg.ShowModal() == wx.ID_OK:
            self.include_elements = dlg.get_include_elements()
            self.exclude_elements = dlg.get_exclude_elements()
            # Legacy compatibility
            self.selected_elements = list(self.include_elements)

            print(f"Include elements: {self.include_elements}")
            print(f"Exclude elements: {self.exclude_elements}")

            # Save to parent Data structure for persistence
            self._save_element_preferences()

        dlg.Destroy()

    def _save_element_preferences(self):
        """Save element preferences to parent Data structure"""
        try:
            if self.parent is None or not hasattr(self.parent, 'Data'):
                return

            if 'Core levels' not in self.parent.Data:
                return

            include_list = list(self.include_elements)
            exclude_list = list(self.exclude_elements)

            # Save to all EDX sheets
            for sheet_name in self.parent.Data['Core levels']:
                if sheet_name.startswith('EDX~'):
                    self.parent.Data['Core levels'][sheet_name]['_EDX_include_elements'] = include_list
                    self.parent.Data['Core levels'][sheet_name]['_EDX_exclude_elements'] = exclude_list

            print(f"Saved element preferences to Data structure")

        except Exception as e:
            print(f"Could not save element preferences: {e}")

    def on_plot_maps(self, event):
        """Plot element maps as RGB composite overlay"""
        if not hasattr(self, 'selected_elements') or not self.selected_elements:
            wx.MessageBox("No elements selected.\nUse 'Set Elements' first.",
                          "No Selection", wx.OK | wx.ICON_WARNING)
            return

        if self.current_data is None:
            wx.MessageBox("No data loaded", "No Data", wx.OK | wx.ICON_WARNING)
            return

        selected_elements = self.selected_elements

        if len(selected_elements) > 7:
            wx.MessageBox("Maximum 7 elements can be displayed.\nPlease select fewer elements.",
                          "Too Many Elements", wx.OK | wx.ICON_WARNING)
            return

        try:
            element_maps = {}
            for element in selected_elements:
                element_map = self.get_element_map(element)
                if element_map is not None:
                    element_maps[element] = element_map

            if not element_maps:
                wx.MessageBox("Could not extract any element maps", "Error", wx.OK | wx.ICON_ERROR)
                return

            self.create_composite_map(element_maps)

        except Exception as e:
            wx.MessageBox(f"Error plotting element maps:\n{str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def get_element_map(self, element):
        """Get element map using HyperSpy or manual integration"""
        element_map = None

        # Method 1: Try HyperSpy's get_lines_intensity
        if hasattr(self.current_data, 'get_lines_intensity'):
            try:
                # Set element if not already set
                if hasattr(self.current_data, 'set_elements'):
                    current_elements = []
                    if hasattr(self.current_data.metadata, 'Sample') and \
                            hasattr(self.current_data.metadata.Sample, 'elements'):
                        current_elements = list(self.current_data.metadata.Sample.elements)

                    if element not in current_elements:
                        current_elements.append(element)
                        self.current_data.set_elements(current_elements)
                        self.current_data.add_lines()

                # Try different X-ray lines
                for line in ['Ka', 'La', 'Ma', 'Kb', 'Lb']:
                    try:
                        line_name = f"{element}_{line}"
                        intensity_maps = self.current_data.get_lines_intensity([line_name])
                        if intensity_maps and len(intensity_maps) > 0:
                            element_map = intensity_maps[0].data
                            break
                    except:
                        continue
            except Exception as e:
                print(f"HyperSpy method failed for {element}: {e}")

        # Method 2: Manual energy window integration
        if element_map is None and len(self.current_data.data.shape) == 3:
            element_map = self.get_element_map_manual(element)

        return element_map

    def get_element_map_manual(self, element):
        """Manually integrate around element's characteristic X-ray energy"""
        xray_energies = {
            'C': 0.277, 'N': 0.392, 'O': 0.525, 'F': 0.677, 'Na': 1.041, 'Mg': 1.254,
            'Al': 1.487, 'Si': 1.740, 'P': 2.013, 'S': 2.307, 'Cl': 2.622, 'K': 3.313,
            'Ca': 3.691, 'Ti': 4.510, 'V': 4.952, 'Cr': 5.414, 'Mn': 5.898, 'Fe': 6.403,
            'Co': 6.930, 'Ni': 7.478, 'Cu': 8.048, 'Zn': 8.638, 'Ga': 9.251, 'Ge': 9.886,
            'As': 10.543, 'Se': 11.222, 'Br': 11.924, 'Rb': 13.395, 'Sr': 14.165,
            'Y': 14.958, 'Zr': 15.775, 'Nb': 16.615, 'Mo': 17.479, 'Ag': 22.163,
            'Cd': 23.174, 'In': 24.210, 'Sn': 25.271, 'Sb': 26.359, 'Te': 27.472,
            'I': 28.612, 'Ba': 32.194, 'La': 33.442, 'Ce': 34.720, 'W': 59.318,
            'Pt': 66.831, 'Au': 68.803, 'Pb': 74.969, 'Bi': 77.107
        }

        if element not in xray_energies:
            return None

        target_energy = xray_energies[element]
        energy_axis = self.get_energy_axis()
        if energy_axis is None:
            return None

        # Find energy window (±0.15 keV around peak)
        window = 0.15
        mask = (energy_axis >= target_energy - window) & (energy_axis <= target_energy + window)

        if not np.any(mask):
            window = 0.3
            mask = (energy_axis >= target_energy - window) & (energy_axis <= target_energy + window)

        if not np.any(mask):
            return None

        data = self.current_data.data
        element_map = np.sum(data[:, :, mask], axis=2)

        return element_map

    def create_composite_map(self, element_maps):
        """Create RGB composite map with legend - no colorbar"""
        # Clear previous colorbar
        if self.current_colorbar is not None:
            try:
                self.current_colorbar.remove()
            except:
                pass
            self.current_colorbar = None

        colors = [
            (1.0, 0.0, 0.0),  # Red
            (0.0, 1.0, 0.0),  # Green
            (0.0, 0.0, 1.0),  # Blue
            (1.0, 1.0, 0.0),  # Yellow
            (1.0, 0.0, 1.0),  # Magenta
            (0.0, 1.0, 1.0),  # Cyan
            (1.0, 0.5, 0.0),  # Orange
        ]

        elements = list(element_maps.keys())
        first_map = list(element_maps.values())[0]
        height, width = first_map.shape

        rgb_image = np.zeros((height, width, 3), dtype=np.float64)
        self.element_colors = {}

        for idx, element in enumerate(elements):
            element_map = element_maps[element]
            color = colors[idx % len(colors)]
            self.element_colors[element] = color

            map_min = np.min(element_map)
            map_max = np.max(element_map)
            if map_max > map_min:
                normalized_map = (element_map - map_min) / (map_max - map_min)
            else:
                normalized_map = np.zeros_like(element_map)

            rgb_image[:, :, 0] += normalized_map * color[0]
            rgb_image[:, :, 1] += normalized_map * color[1]
            rgb_image[:, :, 2] += normalized_map * color[2]

        for c in range(3):
            channel_max = np.max(rgb_image[:, :, c])
            if channel_max > 0:
                rgb_image[:, :, c] /= channel_max

        rgb_image = np.clip(rgb_image, 0, 1)

        self.map_ax.clear()
        self.map_ax.imshow(rgb_image, origin='upper')

        # Remove axis labels and title - use scale bar instead
        self.map_ax.set_xlabel('')
        self.map_ax.set_ylabel('')
        self.map_ax.set_xticks([])
        self.map_ax.set_yticks([])

        # Add scale bar
        self._add_scale_bar()

        # Legend only - no colorbar for composite
        legend_elements = []
        for element, color in self.element_colors.items():
            from matplotlib.patches import Patch
            legend_elements.append(Patch(facecolor=color, edgecolor='black', label=element))

        self.map_ax.legend(handles=legend_elements, loc='upper right', fontsize=8, framealpha=0.8)
        self.map_ax.set_title(f'Composite: {", ".join(elements)}')

        self.map_figure.tight_layout()
        self.map_canvas.draw()

        self.reinitialize_selectors()

    def plot_sum_spectrum_to_parent(self):
        """Plot sum spectrum in KherveFitting main window with element labels"""
        if self.current_data is None:
            return

        data = self.current_data.data
        if len(data.shape) != 3:
            return

        spectrum = np.sum(data, axis=(0, 1))
        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        # Plot using common styled method
        self._plot_edx_spectrum_styled(energy, spectrum, ' ', 'Sum Spectrum')

    def add_peak_labels(self, ax, energy, spectrum):
        """
        Add element peak labels above peaks with improved identification.
        Stores identified peaks for use by quantification.
        Respects user element selections (green=include, red=exclude).
        """
        try:
            from libraries.EDX_Utilities import (
                elements, get_xray_lines_near_energy, Signal1D
            )

            # DEBUG: Print energy range
            print(f"DEBUG add_peak_labels: energy range = {energy[0]:.4f} - {energy[-1]:.4f}, len={len(energy)}")
            print(f"DEBUG add_peak_labels: scale would be = {energy[1] - energy[0]:.6f}")

            # Create signal with proper energy axis
            temp_signal = Signal1D(spectrum)
            temp_signal.axes_manager[0].scale = energy[1] - energy[0] if len(energy) > 1 else 0.01
            temp_signal.axes_manager[0].offset = energy[0]
            temp_signal.axes_manager[0].units = 'keV'
            temp_signal.axes_manager[0].name = 'Energy'

            # Get threshold
            threshold = 0.02  # Default 2%
            if hasattr(self, 'parent') and self.parent and hasattr(self.parent, 'Data'):
                current_sheet = self.parent.sheet_combobox.GetValue()
                if current_sheet == 'EDX~Plot' and 'Core levels' in self.parent.Data:
                    if current_sheet in self.parent.Data['Core levels']:
                        stored_threshold = self.parent.Data['Core levels'][current_sheet].get('_EDX_sensitivity')
                        if stored_threshold is not None:
                            threshold = stored_threshold

            if hasattr(self, 'peak_threshold'):
                threshold = self.peak_threshold

            amp_thresh = np.max(spectrum) * threshold
            print(f"Peak detection with threshold: {threshold:.4f}, amp_thresh: {amp_thresh:.2f}")

            peak_data = temp_signal.find_peaks1D_ohaver(
                maxpeakn=100,
                medfilt_radius=5,
                peakgroup=10,
                amp_thresh=amp_thresh,
                slope_thresh=0
            )

            y_min, y_max = ax.get_ylim()
            y_range = y_max - y_min

            if peak_data is None or not isinstance(peak_data, np.ndarray) or len(peak_data) == 0:
                self.identified_peaks = []
                return

            # Handle nested array format
            if peak_data.ndim > 1:
                peaks = peak_data[0]
            else:
                peaks = peak_data

            # Handle empty inner array
            if len(peaks) == 0 or (isinstance(peaks, np.ndarray) and peaks.size == 0):
                self.identified_peaks = []
                return

            print(f"Found {len(peaks)} peaks")

            # Build peak database with energies and heights
            peak_list = []
            for peak_tuple in peaks:
                try:
                    peak_energy = float(peak_tuple[0])
                    peak_height = float(peak_tuple[1])
                    peak_list.append((peak_energy, peak_height))
                except (TypeError, ValueError, IndexError):
                    continue

            # Sort by intensity (highest first)
            peak_list.sort(key=lambda x: x[1], reverse=True)

            # Get user element preferences (green=include, red=exclude, grey=neutral)
            include_elements = getattr(self, 'include_elements', set())
            exclude_elements = getattr(self, 'exclude_elements', set())

            # Identify peaks using intensity-aware matching with element preferences
            identified_peaks = self._identify_peaks_with_ratios(
                peak_list, elements, include_elements, exclude_elements
            )

            # Store ALL identified peaks for quantification (not just Ka)
            self.identified_peaks = identified_peaks

            # Add labels to plot
            for peak_energy, element, line_type in identified_peaks:
                # Format line type with Greek letters
                line_display = str(line_type)
                line_display = line_display.replace('Ka', 'Kα').replace('Kb', 'Kβ')
                line_display = line_display.replace('La', 'Lα').replace('Lb', 'Lβ')
                line_display = line_display.replace('Ma', 'Mα').replace('Mb', 'Mβ')

                label = f"{element}\n{line_display}"

                # Get peak height at this energy
                idx = np.argmin(np.abs(energy - peak_energy))
                spectrum_height = spectrum[idx]

                # Position label
                peak_pct = (spectrum_height - y_min) / y_range
                label_y = y_min + (peak_pct + 0.02) * y_range

                # Use plot style color if available
                label_color = 'black'
                if hasattr(self, 'parent') and self.parent:
                    label_color = getattr(self.parent, 'line_color', 'black')
                    if label_color in ['white', '#FFFFFF', (1, 1, 1), (255, 255, 255)]:
                        label_color = 'black'  # Ensure visibility

                ax.text(peak_energy, label_y, label,
                        rotation=0,
                        verticalalignment='bottom',
                        horizontalalignment='center',
                        fontsize=7,
                        color=label_color,
                        fontweight='bold')

        except Exception as e:
            print(f"Error adding peak labels: {e}")
            import traceback
            traceback.print_exc()
            self.identified_peaks = []

    def _identify_peaks_with_ratios(self, peak_list, elements, include_elements=None, exclude_elements=None):
        """
        Identify peaks with proper multi-line verification:
        1. If a strong Ka is found, the Kb position MUST be from the same element
        2. User-selected elements reserve ALL their line positions
        3. Rare elements are heavily penalized unless user-selected
        """
        from libraries.EDX_Utilities import get_xray_lines_near_energy

        include_elements = include_elements or set()
        exclude_elements = exclude_elements or set()

        identified = []
        used_energies = set()
        reserved_energies = {}  # Maps energy -> element that reserves it

        # Complete X-ray line energies database
        XRAY_ENERGIES = {
            # Light elements (Z=3-10)
            'Li_Ka': 0.054, 'Be_Ka': 0.109, 'B_Ka': 0.183, 'C_Ka': 0.277,
            'N_Ka': 0.392, 'O_Ka': 0.525, 'F_Ka': 0.677, 'Ne_Ka': 0.849,
            # Period 3 (Z=11-18)
            'Na_Ka': 1.041, 'Na_Kb': 1.071, 'Mg_Ka': 1.254, 'Mg_Kb': 1.302,
            'Al_Ka': 1.487, 'Al_Kb': 1.557, 'Si_Ka': 1.740, 'Si_Kb': 1.836,
            'P_Ka': 2.013, 'P_Kb': 2.139, 'S_Ka': 2.307, 'S_Kb': 2.464,
            'Cl_Ka': 2.622, 'Cl_Kb': 2.816, 'Ar_Ka': 2.957, 'Ar_Kb': 3.190,
            # Period 4 K-lines (Z=19-36)
            'K_Ka': 3.313, 'K_Kb': 3.590, 'Ca_Ka': 3.691, 'Ca_Kb': 4.012,
            'Sc_Ka': 4.090, 'Sc_Kb': 4.460, 'Ti_Ka': 4.510, 'Ti_Kb': 4.931,
            'V_Ka': 4.952, 'V_Kb': 5.427, 'Cr_Ka': 5.414, 'Cr_Kb': 5.947,
            'Mn_Ka': 5.898, 'Mn_Kb': 6.490, 'Fe_Ka': 6.403, 'Fe_Kb': 7.058,
            'Co_Ka': 6.930, 'Co_Kb': 7.649, 'Ni_Ka': 7.478, 'Ni_Kb': 8.265,
            'Cu_Ka': 8.048, 'Cu_Kb': 8.905, 'Zn_Ka': 8.638, 'Zn_Kb': 9.572,
            'Ga_Ka': 9.251, 'Ga_Kb': 10.264, 'Ge_Ka': 9.886, 'Ge_Kb': 10.982,
            'As_Ka': 10.543, 'As_Kb': 11.726, 'Se_Ka': 11.222, 'Se_Kb': 12.496,
            'Br_Ka': 11.924, 'Br_Kb': 13.291, 'Kr_Ka': 12.649, 'Kr_Kb': 14.112,
            # Period 4 L-lines
            'Sc_La': 0.395, 'Ti_La': 0.452, 'Ti_Lb': 0.458,
            'V_La': 0.511, 'V_Lb': 0.519, 'Cr_La': 0.573, 'Cr_Lb': 0.583,
            'Mn_La': 0.637, 'Mn_Lb': 0.649, 'Fe_La': 0.705, 'Fe_Lb': 0.718,
            'Co_La': 0.776, 'Co_Lb': 0.791, 'Ni_La': 0.851, 'Ni_Lb': 0.869,
            'Cu_La': 0.930, 'Cu_Lb': 0.950, 'Zn_La': 1.012, 'Zn_Lb': 1.035,
            'Ga_La': 1.098, 'Ga_Lb': 1.125, 'Ge_La': 1.188, 'Ge_Lb': 1.218,
            'As_La': 1.282, 'As_Lb': 1.317, 'Se_La': 1.379, 'Se_Lb': 1.419,
            'Br_La': 1.480, 'Br_Lb': 1.526,
            # Period 5 (Z=37-54)
            'Rb_Ka': 13.395, 'Rb_Kb': 14.961, 'Rb_La': 1.694, 'Rb_Lb': 1.752,
            'Sr_Ka': 14.165, 'Sr_Kb': 15.835, 'Sr_La': 1.806, 'Sr_Lb': 1.872,
            'Y_Ka': 14.958, 'Y_Kb': 16.737, 'Y_La': 1.922, 'Y_Lb': 1.996,
            'Zr_Ka': 15.775, 'Zr_Kb': 17.667, 'Zr_La': 2.042, 'Zr_Lb': 2.124,
            'Nb_Ka': 16.615, 'Nb_Kb': 18.622, 'Nb_La': 2.166, 'Nb_Lb': 2.257,
            'Mo_Ka': 17.479, 'Mo_Kb': 19.608, 'Mo_La': 2.293, 'Mo_Lb': 2.395,
            'Tc_Ka': 18.367, 'Tc_Kb': 20.619, 'Tc_La': 2.424, 'Tc_Lb': 2.538,
            'Ru_Ka': 19.279, 'Ru_Kb': 21.657, 'Ru_La': 2.558, 'Ru_Lb': 2.683,
            'Rh_Ka': 20.216, 'Rh_Kb': 22.724, 'Rh_La': 2.696, 'Rh_Lb': 2.834,
            'Pd_Ka': 21.177, 'Pd_Kb': 23.819, 'Pd_La': 2.838, 'Pd_Lb': 2.990,
            'Ag_Ka': 22.163, 'Ag_Kb': 24.942, 'Ag_La': 2.984, 'Ag_Lb': 3.151,
            'Cd_Ka': 23.174, 'Cd_Kb': 26.095, 'Cd_La': 3.133, 'Cd_Lb': 3.316,
            'In_Ka': 24.210, 'In_Kb': 27.276, 'In_La': 3.287, 'In_Lb': 3.487,
            'Sn_Ka': 25.271, 'Sn_Kb': 28.486, 'Sn_La': 3.444, 'Sn_Lb': 3.663,
            'Sb_Ka': 26.359, 'Sb_Kb': 29.725, 'Sb_La': 3.605, 'Sb_Lb': 3.843,
            'Te_Ka': 27.472, 'Te_Kb': 30.995, 'Te_La': 3.769, 'Te_Lb': 4.029,
            'I_Ka': 28.612, 'I_Kb': 32.294, 'I_La': 3.937, 'I_Lb': 4.221,
            'Xe_Ka': 29.779, 'Xe_Kb': 33.624, 'Xe_La': 4.109, 'Xe_Lb': 4.422,
            # Period 6 (Z=55-86)
            'Cs_Ka': 30.973, 'Cs_Kb': 34.987, 'Cs_La': 4.286, 'Cs_Lb': 4.620,
            'Ba_Ka': 32.194, 'Ba_Kb': 36.378, 'Ba_La': 4.466, 'Ba_Lb': 4.828,
            'La_La': 4.651, 'La_Lb': 5.042, 'La_Lg': 5.383,
            'Ce_La': 4.840, 'Ce_Lb': 5.262, 'Ce_Lg': 5.613,
            'Pr_La': 5.034, 'Pr_Lb': 5.489, 'Pr_Lg': 5.850,
            'Nd_La': 5.230, 'Nd_Lb': 5.722, 'Nd_Lg': 6.090,
            'Pm_La': 5.432, 'Pm_Lb': 5.961, 'Pm_Lg': 6.339,
            'Sm_La': 5.636, 'Sm_Lb': 6.206, 'Sm_Lg': 6.586,
            'Eu_La': 5.846, 'Eu_Lb': 6.456, 'Eu_Lg': 6.843,
            'Gd_La': 6.057, 'Gd_Lb': 6.713, 'Gd_Lg': 7.103,
            'Tb_La': 6.273, 'Tb_Lb': 6.978, 'Tb_Lg': 7.368,
            'Dy_La': 6.495, 'Dy_Lb': 7.247, 'Dy_Lg': 7.635,
            'Ho_La': 6.720, 'Ho_Lb': 7.525, 'Ho_Lg': 7.911,
            'Er_La': 6.949, 'Er_Lb': 7.810, 'Er_Lg': 8.189,
            'Tm_La': 7.180, 'Tm_Lb': 8.101, 'Tm_Lg': 8.468,
            'Yb_La': 7.416, 'Yb_Lb': 8.401, 'Yb_Lg': 8.758,
            'Lu_La': 7.655, 'Lu_Lb': 8.709, 'Lu_Lg': 9.048,
            'Hf_La': 7.899, 'Hf_Lb': 9.022, 'Hf_Lg': 9.347,
            'Ta_La': 8.146, 'Ta_Lb': 9.343, 'Ta_Lg': 9.651,
            'W_La': 8.398, 'W_Lb': 9.672, 'W_Lg': 9.961,
            'Re_La': 8.652, 'Re_Lb': 10.010, 'Re_Lg': 10.275,
            'Os_La': 8.911, 'Os_Lb': 10.355, 'Os_Lg': 10.598,
            'Ir_La': 9.175, 'Ir_Lb': 10.708, 'Ir_Lg': 10.920,
            'Pt_La': 9.442, 'Pt_Lb': 11.070, 'Pt_Lg': 11.250,
            'Au_La': 9.713, 'Au_Lb': 11.442, 'Au_Lg': 11.584,
            'Hg_La': 9.988, 'Hg_Lb': 11.822, 'Hg_Lg': 11.924,
            'Tl_La': 10.268, 'Tl_Lb': 12.213, 'Tl_Lg': 12.271,
            'Pb_La': 10.551, 'Pb_Lb': 12.614, 'Pb_Lg': 12.622,
            'Bi_La': 10.839, 'Bi_Lb': 13.023, 'Bi_Lg': 12.979,
            'Po_La': 11.131, 'Po_Lb': 13.447,
            'At_La': 11.427, 'At_Lb': 13.876,
            'Rn_La': 11.727, 'Rn_Lb': 14.316,
            # M-lines for heavy elements
            'Hf_Ma': 1.644, 'Hf_Mb': 1.700, 'Ta_Ma': 1.709, 'Ta_Mb': 1.766,
            'W_Ma': 1.775, 'W_Mb': 1.835, 'Re_Ma': 1.842, 'Re_Mb': 1.903,
            'Os_Ma': 1.910, 'Os_Mb': 1.978, 'Ir_Ma': 1.979, 'Ir_Mb': 2.053,
            'Pt_Ma': 2.048, 'Pt_Mb': 2.127, 'Au_Ma': 2.120, 'Au_Mb': 2.204,
            'Hg_Ma': 2.195, 'Hg_Mb': 2.282, 'Tl_Ma': 2.271, 'Tl_Mb': 2.362,
            'Pb_Ma': 2.346, 'Pb_Mb': 2.443, 'Bi_Ma': 2.423, 'Bi_Mb': 2.525,
            # Actinides
            'Th_La': 12.968, 'Th_Lb': 15.623, 'Th_Lg': 16.202, 'Th_Ma': 2.996, 'Th_Mb': 3.144,
            'Pa_La': 13.291, 'Pa_Lb': 16.024, 'Pa_Ma': 3.082, 'Pa_Mb': 3.240,
            'U_La': 13.614, 'U_Lb': 16.428, 'U_Lg': 17.220, 'U_Ma': 3.171, 'U_Mb': 3.336,
        }

        # Build ELEMENT_LINES: element -> {line_type: energy}
        ELEMENT_LINES = {}
        for line_key, energy in XRAY_ENERGIES.items():
            parts = line_key.split('_')
            if len(parts) == 2:
                elem, line_type = parts
                if elem not in ELEMENT_LINES:
                    ELEMENT_LINES[elem] = {}
                ELEMENT_LINES[elem][line_type] = energy

        # Element categories for scoring
        VERY_COMMON = ['C', 'N', 'O', 'Na', 'Mg', 'Al', 'Si', 'P', 'S', 'Cl', 'K', 'Ca',
                       'Ti', 'V', 'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn']
        COMMON_ELEMENTS = ['F', 'Ga', 'Ge', 'As', 'Se', 'Br', 'Sr', 'Y', 'Zr', 'Nb', 'Mo',
                           'Ag', 'Cd', 'In', 'Sn', 'Sb', 'Te', 'I', 'Ba', 'La', 'Ce',
                           'W', 'Pt', 'Au', 'Pb', 'Bi', 'Th', 'U']
        RARE_ELEMENTS = ['Li', 'Be', 'B', 'Ne', 'Ar', 'Kr', 'Xe', 'Rn',
                         'Tc', 'Pm', 'Po', 'At', 'Fr', 'Ra', 'Ac',
                         'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu',
                         'Hf', 'Ta', 'Re', 'Os', 'Ir', 'Hg', 'Tl',
                         'Rb', 'Cs', 'Pa', 'Np', 'Pu', 'Am', 'Cm',
                         'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'Ds', 'Rg', 'Cn',
                         'Nh', 'Fl', 'Mc', 'Lv', 'Ts', 'Og']

        def get_atomic_number(elem):
            try:
                elem_obj = elements[elem]
                return elem_obj.General_properties.Z
            except:
                return 26

        def get_k_ratio_range(Z):
            if Z < 20:
                return (7.0, 12.0)
            elif Z < 30:
                return (6.0, 9.0)
            elif Z < 40:
                return (5.5, 8.0)
            else:
                return (5.0, 7.5)

        def get_l_ratio_range(Z):
            if Z < 50:
                return (2.5, 6.0)
            elif Z < 70:
                return (2.0, 5.0)
            else:
                return (1.8, 4.5)

        def find_peak_near(target_energy, tolerance=0.12):
            for pe, ph in peak_list:
                if abs(pe - target_energy) < tolerance:
                    return pe, ph
            return None, 0

        def is_energy_reserved(energy, tolerance=0.10):
            """Check if energy is reserved by another element"""
            for reserved_e, elem in reserved_energies.items():
                if abs(energy - reserved_e) < tolerance:
                    return elem
            return None

        # ============================================================
        # STEP 1: Reserve ALL line positions for user-selected elements
        # ============================================================
        for elem in include_elements:
            if elem in ELEMENT_LINES:
                for line_type, energy in ELEMENT_LINES[elem].items():
                    reserved_energies[energy] = elem
                    # print(f"Reserved {energy:.3f} keV for {elem} {line_type}")

        # ============================================================
        # STEP 2: Find strong Ka peaks and reserve their Kb positions
        # ============================================================
        strong_ka_elements = {}  # element -> (ka_peak, ka_height)

        # First pass: identify strong Ka peaks
        for elem, lines in ELEMENT_LINES.items():
            if elem in exclude_elements:
                continue
            if 'Ka' not in lines:
                continue

            ka_energy = lines['Ka']
            ka_peak, ka_height = find_peak_near(ka_energy)

            if ka_peak is None:
                continue

            # Is this a significant peak? (top 50% of peak heights)
            max_height = peak_list[0][1] if peak_list else 1
            if ka_height < max_height * 0.05:  # At least 5% of max
                continue

            # Check if this energy is reserved for a different element
            reserved_by = is_energy_reserved(ka_energy)
            if reserved_by and reserved_by != elem:
                continue

            # Store this as a potential Ka
            if elem not in strong_ka_elements:
                strong_ka_elements[elem] = (ka_peak, ka_height, lines.get('Kb'))

        # For strong Ka peaks, reserve the Kb position
        for elem, (ka_peak, ka_height, kb_energy) in strong_ka_elements.items():
            if kb_energy:
                # Check if Kb position has a peak
                kb_peak, kb_height = find_peak_near(kb_energy)
                if kb_peak:
                    # Verify ratio is reasonable
                    Z = get_atomic_number(elem)
                    ratio_range = get_k_ratio_range(Z)
                    if kb_height > 0:
                        ratio = ka_height / kb_height
                        # If ratio is in reasonable range, reserve this Kb for this element
                        if ratio_range[0] * 0.5 <= ratio <= ratio_range[1] * 1.5:
                            if kb_energy not in reserved_energies:
                                reserved_energies[kb_energy] = elem
                                # print(f"Strong Ka found: reserving Kb at {kb_energy:.3f} keV for {elem}")

        # ============================================================
        # STEP 3: Similarly for strong La peaks, reserve Lb positions
        # ============================================================
        strong_la_elements = {}

        for elem, lines in ELEMENT_LINES.items():
            if elem in exclude_elements:
                continue
            if 'La' not in lines:
                continue

            la_energy = lines['La']
            la_peak, la_height = find_peak_near(la_energy)

            if la_peak is None:
                continue

            max_height = peak_list[0][1] if peak_list else 1
            if la_height < max_height * 0.05:
                continue

            reserved_by = is_energy_reserved(la_energy)
            if reserved_by and reserved_by != elem:
                continue

            if elem not in strong_la_elements:
                strong_la_elements[elem] = (la_peak, la_height, lines.get('Lb'))

        for elem, (la_peak, la_height, lb_energy) in strong_la_elements.items():
            if lb_energy:
                lb_peak, lb_height = find_peak_near(lb_energy)
                if lb_peak:
                    Z = get_atomic_number(elem)
                    ratio_range = get_l_ratio_range(Z)
                    if lb_height > 0:
                        ratio = la_height / lb_height
                        if ratio_range[0] * 0.5 <= ratio <= ratio_range[1] * 1.5:
                            if lb_energy not in reserved_energies:
                                reserved_energies[lb_energy] = elem
                                # print(f"Strong La found: reserving Lb at {lb_energy:.3f} keV for {elem}")

        # ============================================================
        # STEP 4: Score all elements considering reservations
        # ============================================================
        element_scores = {}

        for elem, lines in ELEMENT_LINES.items():
            if elem in exclude_elements:
                continue

            score = 0
            matched_lines = []
            Z = get_atomic_number(elem)

            # Check K-series
            if 'Ka' in lines:
                ka_energy = lines['Ka']
                ka_peak, ka_height = find_peak_near(ka_energy)

                # Check reservation
                reserved_by = is_energy_reserved(ka_energy)
                if reserved_by and reserved_by != elem:
                    # This energy belongs to another element
                    pass
                elif ka_peak:
                    kb_energy = lines.get('Kb')
                    if kb_energy:
                        kb_peak, kb_height = find_peak_near(kb_energy)

                        # Check if Kb is reserved for us or available
                        kb_reserved_by = is_energy_reserved(kb_energy)

                        if kb_peak and kb_height > 0:
                            ratio = ka_height / kb_height
                            ratio_range = get_k_ratio_range(Z)

                            if kb_reserved_by == elem:
                                # Kb is reserved for this element - strong match
                                score += 5.0
                                matched_lines.extend([('Ka', ka_peak), ('Kb', kb_peak)])
                            elif kb_reserved_by is None:
                                # Kb not reserved - check ratio
                                if ratio_range[0] <= ratio <= ratio_range[1]:
                                    score += 4.0
                                    matched_lines.extend([('Ka', ka_peak), ('Kb', kb_peak)])
                                elif ratio_range[0] * 0.7 <= ratio <= ratio_range[1] * 1.3:
                                    score += 2.5
                                    matched_lines.extend([('Ka', ka_peak), ('Kb', kb_peak)])
                            # If Kb reserved for different element, only count Ka
                            else:
                                score += 0.5
                                matched_lines.append(('Ka', ka_peak))
                        else:
                            # No Kb peak found
                            score += 0.5
                            matched_lines.append(('Ka', ka_peak))
                    else:
                        score += 0.5
                        matched_lines.append(('Ka', ka_peak))

            # Check L-series
            if 'La' in lines:
                la_energy = lines['La']
                la_peak, la_height = find_peak_near(la_energy)

                reserved_by = is_energy_reserved(la_energy)
                if reserved_by and reserved_by != elem:
                    pass
                elif la_peak:
                    lb_energy = lines.get('Lb')
                    if lb_energy:
                        lb_peak, lb_height = find_peak_near(lb_energy)
                        lb_reserved_by = is_energy_reserved(lb_energy)

                        if lb_peak and lb_height > 0:
                            ratio = la_height / lb_height
                            ratio_range = get_l_ratio_range(Z)

                            has_k_match = any(line[0] in ['Ka', 'Kb'] for line in matched_lines)

                            if lb_reserved_by == elem:
                                if has_k_match:
                                    score += 4.0
                                else:
                                    score += 3.0
                                matched_lines.extend([('La', la_peak), ('Lb', lb_peak)])
                            elif lb_reserved_by is None:
                                if ratio_range[0] <= ratio <= ratio_range[1]:
                                    if has_k_match:
                                        score += 3.0
                                    else:
                                        score += 2.0
                                    matched_lines.extend([('La', la_peak), ('Lb', lb_peak)])
                                elif ratio_range[0] * 0.7 <= ratio <= ratio_range[1] * 1.3:
                                    if has_k_match:
                                        score += 2.0
                                    else:
                                        score += 1.0
                                    matched_lines.extend([('La', la_peak), ('Lb', lb_peak)])
                            else:
                                if has_k_match:
                                    score += 1.0
                                    matched_lines.append(('La', la_peak))
                        else:
                            has_k_match = any(line[0] in ['Ka', 'Kb'] for line in matched_lines)
                            if has_k_match:
                                score += 1.5
                                matched_lines.append(('La', la_peak))
                            else:
                                score += 0.3
                                matched_lines.append(('La', la_peak))

            # Check M-series for heavy elements
            if 'Ma' in lines and Z > 70:
                ma_energy = lines['Ma']
                ma_peak, ma_height = find_peak_near(ma_energy)

                reserved_by = is_energy_reserved(ma_energy)
                if reserved_by and reserved_by != elem:
                    pass
                elif ma_peak:
                    mb_energy = lines.get('Mb')
                    if mb_energy:
                        mb_peak, mb_height = find_peak_near(mb_energy)
                        if mb_peak and mb_height > 0:
                            ratio = ma_height / mb_height
                            if 1.5 <= ratio <= 4.0:
                                if len(matched_lines) > 0:
                                    score += 1.5
                                else:
                                    score += 1.0
                                matched_lines.extend([('Ma', ma_peak), ('Mb', mb_peak)])

            # ============================================================
            # Apply element preference modifiers
            # ============================================================
            if elem in include_elements:
                score *= 10.0  # MASSIVE bonus for user-selected
            elif elem in VERY_COMMON:
                score *= 2.0
            elif elem in COMMON_ELEMENTS:
                score *= 1.5
            elif elem in RARE_ELEMENTS:
                score *= 0.05  # 95% penalty for rare elements

            # Additional penalty when user has made specific selections
            if len(include_elements) > 0 and elem not in include_elements:
                score *= 0.2

            if score > 0.3 and len(matched_lines) >= 1:
                element_scores[elem] = {
                    'score': score,
                    'lines': matched_lines,
                    'Z': Z
                }

        # ============================================================
        # STEP 5: Assign peaks by score, respecting reservations
        # ============================================================
        sorted_elements = sorted(element_scores.items(), key=lambda x: x[1]['score'], reverse=True)

        for elem, data in sorted_elements:
            lines_to_add = []
            conflict = False

            for line_type, peak_energy in data['lines']:
                if peak_energy in used_energies:
                    # Check if it's reserved for this element
                    reserved_by = is_energy_reserved(peak_energy)
                    if reserved_by == elem:
                        # It's reserved for us but already used - skip
                        pass
                    conflict = True
                    break

                # Check reservation
                reserved_by = is_energy_reserved(peak_energy)
                if reserved_by is not None and reserved_by != elem:
                    conflict = True
                    break

                lines_to_add.append((peak_energy, elem, line_type))

            if not conflict and lines_to_add:
                for peak_energy, elem, line_type in lines_to_add:
                    identified.append((peak_energy, elem, line_type))
                    used_energies.add(peak_energy)
                    # print(f"Identified: {elem} {line_type} at {peak_energy:.3f} keV (score: {data['score']:.2f})")

        # ============================================================
        # STEP 6: Handle reserved but unmatched peaks
        # ============================================================
        for reserved_energy, elem in reserved_energies.items():
            if reserved_energy in used_energies:
                continue

            # Find which line this is
            if elem in ELEMENT_LINES:
                for line_type, line_energy in ELEMENT_LINES[elem].items():
                    if abs(line_energy - reserved_energy) < 0.05:
                        # Check if there's a peak here
                        peak, height = find_peak_near(reserved_energy)
                        if peak and height > 0:
                            identified.append((peak, elem, line_type))
                            used_energies.add(peak)
                            # print(f"Identified (reserved): {elem} {line_type} at {peak:.3f} keV")
                        break

        # ============================================================
        # STEP 7: Remaining unidentified significant peaks
        # ============================================================
        for peak_energy, peak_height in peak_list:
            if peak_energy in used_energies:
                continue

            # Skip weak peaks
            if peak_height < 0.10 * peak_list[0][1]:
                continue

            # Check if reserved
            reserved_by = is_energy_reserved(peak_energy)
            if reserved_by:
                continue  # Already handled above

            best_match = None
            best_diff = 0.08
            best_score = 0

            for line_key, line_energy in XRAY_ENERGIES.items():
                if '_Kb' in line_key or '_Lb' in line_key or '_Mb' in line_key or '_Lg' in line_key:
                    continue

                diff = abs(peak_energy - line_energy)
                if diff < best_diff:
                    elem = line_key.split('_')[0]
                    if elem in exclude_elements:
                        continue

                    score = 1.0 - diff / 0.08

                    if elem in include_elements:
                        score *= 10.0
                    elif elem in VERY_COMMON:
                        score *= 2.0
                    elif elem in COMMON_ELEMENTS:
                        score *= 1.5
                    elif elem in RARE_ELEMENTS:
                        score *= 0.05

                    if len(include_elements) > 0 and elem not in include_elements:
                        score *= 0.2

                    if score > best_score:
                        best_score = score
                        best_diff = diff
                        best_match = line_key

            if best_match and best_score > 0.2:
                elem, line_type = best_match.split('_')
                identified.append((peak_energy, elem, line_type))
                used_energies.add(peak_energy)
                # print(f"Identified (unpaired): {elem} {line_type} at {peak_energy:.3f} keV")

        return identified



    def reinitialize_selectors(self):
        """Reinitialize selection tools after axes change"""
        # Reset selectors
        self.rect_selector = None
        self.line_selector = None

        # Recreate if mode is active
        if self.selection_mode == 'area':
            self.rect_selector = RectangleSelector(
                self.map_ax,
                self.on_area_select,
                useblit=True,
                props=dict(facecolor='white', edgecolor='white', alpha=0.3, fill=True),
                button=[1],
                minspanx=5,
                minspany=5,
                spancoords='pixels',
                interactive=True
            )
        elif self.selection_mode == 'line':
            self.line_selector = SpanSelector(
                self.map_ax,
                self.on_line_select,
                'horizontal',
                useblit=True,
                props=dict(facecolor='white', alpha=0.3),
                interactive=True
            )

    # ==================== EXPORT ====================

    def on_export_excel(self, event):
        """Export to Excel"""
        if self.current_data is None:
            wx.MessageBox("No data to export", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Export to Excel", wildcard="Excel files (*.xlsx)|*.xlsx",
                          style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

        self.export_to_excel(file_path)

    def on_export_csv(self, event):
        """Export to CSV"""
        if self.current_data is None:
            wx.MessageBox("No data to export", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Export to CSV", wildcard="CSV files (*.csv)|*.csv",
                          style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

        self.export_to_csv(file_path)

    def on_export_hdf5(self, event):
        """Export to HDF5"""
        if self.current_data is None:
            wx.MessageBox("No data to export", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Export to HDF5", wildcard="HDF5 files (*.hdf5)|*.hdf5",
                          style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

        try:
            self.current_data.save(file_path)
            wx.MessageBox(f"Data exported to:\n{file_path}", "Success", wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            wx.MessageBox(f"Error exporting:\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def on_export_map_image(self, event):
        """Export current map as image"""
        with wx.FileDialog(self, "Export Map Image",
                          wildcard="PNG files (*.png)|*.png|TIFF files (*.tiff)|*.tiff",
                          style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            file_path = dlg.GetPath()

        try:
            self.map_figure.savefig(file_path, dpi=300, bbox_inches='tight')
            wx.MessageBox(f"Map exported to:\n{file_path}", "Success", wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            wx.MessageBox(f"Error exporting:\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def export_to_excel(self, file_path):
        """Export data to Excel"""
        import pandas as pd

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data = self.current_data.data

                if len(data.shape) == 3:
                    # Export sum spectrum
                    spectrum = np.sum(data, axis=(0, 1))
                    energy = self.get_energy_axis()

                    df = pd.DataFrame({
                        'Energy (keV)': [f"{e:.2f}" for e in energy],
                        'Intensity (Counts)': [f"{i:.2f}" for i in spectrum]
                    })
                    df.to_excel(writer, sheet_name='Sum_Spectrum', index=False)

                    # Export sum image stats
                    sum_image = np.sum(data, axis=2)
                    stats_df = pd.DataFrame({
                        'Parameter': ['Width (px)', 'Height (px)', 'Channels', 'Min', 'Max', 'Mean'],
                        'Value': [
                            f"{data.shape[1]:.2f}",
                            f"{data.shape[0]:.2f}",
                            f"{data.shape[2]:.2f}",
                            f"{np.min(sum_image):.2f}",
                            f"{np.max(sum_image):.2f}",
                            f"{np.mean(sum_image):.2f}"
                        ]
                    })
                    stats_df.to_excel(writer, sheet_name='Statistics', index=False)

                elif len(data.shape) == 2:
                    stats_df = pd.DataFrame({
                        'Parameter': ['Width (px)', 'Height (px)', 'Min', 'Max', 'Mean', 'Std'],
                        'Value': [
                            f"{data.shape[1]:.2f}",
                            f"{data.shape[0]:.2f}",
                            f"{np.min(data):.2f}",
                            f"{np.max(data):.2f}",
                            f"{np.mean(data):.2f}",
                            f"{np.std(data):.2f}"
                        ]
                    })
                    stats_df.to_excel(writer, sheet_name='Statistics', index=False)

            wx.MessageBox(f"Data exported to:\n{file_path}", "Success", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Error exporting:\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def export_to_csv(self, file_path):
        """Export spectrum to CSV"""
        import pandas as pd

        try:
            data = self.current_data.data

            if len(data.shape) == 3:
                spectrum = np.sum(data, axis=(0, 1))
            elif len(data.shape) == 1:
                spectrum = data
            else:
                wx.MessageBox("Cannot export 2D image to CSV spectrum format",
                             "Error", wx.OK | wx.ICON_ERROR)
                return

            energy = self.get_energy_axis()

            df = pd.DataFrame({
                'Energy_keV': [f"{e:.2f}" for e in energy],
                'Intensity_Counts': [f"{i:.2f}" for i in spectrum]
            })
            df.to_csv(file_path, index=False)

            wx.MessageBox(f"Spectrum exported to:\n{file_path}", "Success", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Error exporting:\n{str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def on_right_click(self, event):
        """Show right-click context menu"""
        menu = wx.Menu()

        # Save EDX Plot option at top
        save_plot_item = menu.Append(wx.ID_ANY, "Save Current EDX Plot...")
        self.Bind(wx.EVT_MENU, self.on_save_edx_plot, save_plot_item)

        menu.AppendSeparator()

        # HeatMap submenu
        heatmap_menu = wx.Menu()

        colormaps = ['Greens', 'plasma', 'viridis', 'inferno', 'magma', 'hot', 'cool', 'gray', 'jet',
                     'rainbow', 'turbo', 'cividis', 'Spectral', 'coolwarm', 'RdYlBu', 'RdBu',
                     'Blues', 'Reds', 'Oranges', 'Purples', 'YlOrRd', 'YlGnBu', 'RdPu', 'BuPu',
                     'GnBu', 'PuBu', 'YlGn', 'binary', 'bone', 'copper', 'autumn', 'winter',
                     'spring', 'summer', 'twilight', 'hsv', 'nipy_spectral', 'terrain', 'ocean',
                     'gist_earth', 'seismic', 'bwr', 'BrBG', 'PRGn', 'PiYG', 'RdGy', 'RdYlGn']

        for cmap in colormaps:
            item = heatmap_menu.AppendRadioItem(wx.ID_ANY, cmap)
            if cmap == self.current_cmap:
                item.Check(True)
            self.Bind(wx.EVT_MENU, lambda evt, c=cmap: self.on_change_colormap(c), item)

        menu.AppendSubMenu(heatmap_menu, "HeatMap")

        menu.AppendSeparator()

        data_browser_item = menu.Append(wx.ID_ANY, "Data Browser...")
        info_item = menu.Append(wx.ID_ANY, "Info...")
        menu.AppendSeparator()
        clear_item = menu.Append(wx.ID_ANY, "Clear Selections")

        self.Bind(wx.EVT_MENU, self.on_show_data_browser, data_browser_item)
        self.Bind(wx.EVT_MENU, self.on_show_info, info_item)
        self.Bind(wx.EVT_MENU, self.on_clear_selections, clear_item)

        self.PopupMenu(menu)
        menu.Destroy()

    def on_quantification(self, event):
        """Perform EDX quantification and display atomic percentages"""
        if self.current_data is None:
            wx.MessageBox("No EDX data loaded.", "Quantification Error", wx.OK | wx.ICON_WARNING)
            return

        if not self.selected_elements:
            wx.MessageBox("Please select elements first using the periodic table.",
                          "Quantification Error", wx.OK | wx.ICON_WARNING)
            return

        try:
            # Get sum spectrum
            data = self.current_data.data
            if len(data.shape) == 3:
                spectrum = np.sum(data, axis=(0, 1))
            elif len(data.shape) == 1:
                spectrum = data
            else:
                wx.MessageBox("Unsupported data shape for quantification.",
                              "Quantification Error", wx.OK | wx.ICON_WARNING)
                return

            energy = self.get_energy_axis()
            if energy is None:
                wx.MessageBox("Could not get energy axis.",
                              "Quantification Error", wx.OK | wx.ICON_WARNING)
                return

            # Calculate atomic percentages using peak intensities
            results = self._calculate_atomic_percent(energy, spectrum, self.selected_elements)

            if results:
                # Display results in a dialog
                self._show_quantification_results(results)

                # Add to peak fitting grid in parent window if available
                if self.parent is not None and hasattr(self.parent, 'peak_params_grid'):
                    self._add_results_to_grid(results)

        except Exception as e:
            wx.MessageBox(f"Quantification error: {str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def _calculate_atomic_percent(self, energy, spectrum, elements_list):
        """Calculate atomic percentages from peak intensities using identified Ka peaks"""
        from libraries.EDX_Utilities import elements as edx_elements

        results = []
        total_intensity = 0
        element_intensities = {}

        # Use identified Ka peaks if available
        if hasattr(self, 'identified_ka_peaks') and self.identified_ka_peaks:
            print(f"Using {len(self.identified_ka_peaks)} identified Ka peaks for quantification")

            for peak_energy, element, line_type in self.identified_ka_peaks:
                try:
                    if element not in edx_elements:
                        continue
                    elem_obj = edx_elements[element]

                    # Find peak intensity at this energy
                    idx = np.argmin(np.abs(energy - peak_energy))

                    # Integrate around peak
                    window = 5  # channels
                    start_idx = max(0, idx - window)
                    end_idx = min(len(spectrum), idx + window)
                    peak_intensity = np.sum(spectrum[start_idx:end_idx])

                    # Get atomic weight for normalization
                    atomic_weight = elem_obj.General_properties.atomic_weight

                    element_intensities[element] = {
                        'intensity': peak_intensity,
                        'line': line_type,
                        'energy': peak_energy,
                        'atomic_weight': atomic_weight
                    }
                    total_intensity += peak_intensity

                except (KeyError, AttributeError) as e:
                    print(f"Could not process element {element}: {e}")
                    continue
        else:
            # Fallback to old method using selected elements
            print("No identified Ka peaks found, using selected elements")
            for element in elements_list:
                try:
                    if element not in edx_elements:
                        continue
                    elem_obj = edx_elements[element]

                    if elem_obj.Atomic_properties is None:
                        continue
                    xray_lines = elem_obj.Atomic_properties.Xray_lines
                    if xray_lines is None:
                        continue

                    # Find the strongest line (Ka preferred, then La)
                    best_line = None
                    best_energy = None
                    for line_type in ['Ka', 'La', 'Ma']:
                        line_obj = getattr(xray_lines, line_type, None)
                        if line_obj is not None:
                            best_line = line_type
                            best_energy = line_obj.energy_keV
                            break

                    if best_energy is not None:
                        # Find peak intensity at this energy
                        idx = np.argmin(np.abs(energy - best_energy))

                        # Integrate around peak (simple approach)
                        window = 5  # channels
                        start_idx = max(0, idx - window)
                        end_idx = min(len(spectrum), idx + window)
                        peak_intensity = np.sum(spectrum[start_idx:end_idx])

                        # Get atomic weight for normalization
                        atomic_weight = elem_obj.General_properties.atomic_weight

                        element_intensities[element] = {
                            'intensity': peak_intensity,
                            'line': best_line,
                            'energy': best_energy,
                            'atomic_weight': atomic_weight
                        }
                        total_intensity += peak_intensity

                except (KeyError, AttributeError) as e:
                    print(f"Could not process element {element}: {e}")
                    continue

        # Calculate atomic percentages (simplified - without k-factors)
        if total_intensity > 0:
            for element, data in element_intensities.items():
                # Simple normalized intensity (not true atomic %)
                percent = (data['intensity'] / total_intensity) * 100
                results.append({
                    'element': element,
                    'line': data['line'],
                    'energy': data['energy'],
                    'atomic_percent': percent,
                    'intensity': data['intensity']
                })

        return results

    def _calculate_atomic_percent_OLD(self, energy, spectrum, elements):
        """Calculate atomic percentages from peak intensities using identified Ka peaks"""
        from exspy.material import elements as exspy_elements

        results = []
        total_intensity = 0
        element_intensities = {}

        # Use identified Ka peaks if available
        if hasattr(self, 'identified_ka_peaks') and self.identified_ka_peaks:
            print(f"Using {len(self.identified_ka_peaks)} identified Ka peaks for quantification")

            for peak_energy, element, line_type in self.identified_ka_peaks:
                try:
                    elem_obj = exspy_elements[element]

                    # Find peak intensity at this energy
                    idx = np.argmin(np.abs(energy - peak_energy))

                    # Integrate around peak
                    window = 5  # channels
                    start_idx = max(0, idx - window)
                    end_idx = min(len(spectrum), idx + window)
                    peak_intensity = np.sum(spectrum[start_idx:end_idx])

                    # Get atomic weight for normalization
                    atomic_weight = elem_obj.General_properties.atomic_weight

                    element_intensities[element] = {
                        'intensity': peak_intensity,
                        'line': line_type,
                        'energy': peak_energy,
                        'atomic_weight': atomic_weight
                    }
                    total_intensity += peak_intensity

                except (KeyError, AttributeError) as e:
                    print(f"Could not process element {element}: {e}")
                    continue
        else:
            # Fallback to old method using selected elements
            print("No identified Ka peaks found, using selected elements")
            for element in elements:
                try:
                    elem_obj = exspy_elements[element]
                    if hasattr(elem_obj, 'Atomic_properties') and hasattr(elem_obj.Atomic_properties, 'Xray_lines'):
                        xray_lines = elem_obj.Atomic_properties.Xray_lines

                        # Find the strongest line (Ka preferred, then La)
                        best_line = None
                        best_energy = None
                        for line_type in ['Ka', 'La', 'Ma']:
                            if line_type in xray_lines:
                                best_line = line_type
                                best_energy = xray_lines[line_type].energy_keV
                                break

                        if best_energy is not None:
                            # Find peak intensity at this energy
                            idx = np.argmin(np.abs(energy - best_energy))

                            # Integrate around peak (simple approach)
                            window = 5  # channels
                            start_idx = max(0, idx - window)
                            end_idx = min(len(spectrum), idx + window)
                            peak_intensity = np.sum(spectrum[start_idx:end_idx])

                            # Get atomic weight for normalization
                            atomic_weight = elem_obj.General_properties.atomic_weight

                            element_intensities[element] = {
                                'intensity': peak_intensity,
                                'line': best_line,
                                'energy': best_energy,
                                'atomic_weight': atomic_weight
                            }
                            total_intensity += peak_intensity

                except (KeyError, AttributeError) as e:
                    print(f"Could not process element {element}: {e}")
                    continue

        # Calculate atomic percentages (simplified - without k-factors)
        if total_intensity > 0:
            for element, data in element_intensities.items():
                # Simple normalized intensity (not true atomic %)
                # For accurate results, k-factors would be needed
                normalized = (data['intensity'] / total_intensity) * 100
                results.append({
                    'element': element,
                    'line': data['line'],
                    'energy': data['energy'],
                    'intensity': data['intensity'],
                    'atomic_percent': normalized,
                    'atomic_weight': data['atomic_weight']
                })

        return results

    def _show_quantification_results(self, results):
        """Display quantification results in a dialog"""
        dlg = wx.Dialog(self, title="EDX Quantification Results", size=(400, 300))
        panel = wx.Panel(dlg)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Header
        header = wx.StaticText(panel, label="Element Composition (Relative %)")
        header.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        sizer.Add(header, 0, wx.ALL | wx.ALIGN_CENTER, 10)

        # Note about simplified calculation
        note = wx.StaticText(panel, label="Note: Simplified calculation without k-factors.\nFor accurate results, use instrument-specific k-factors.")
        note.SetForegroundColour(wx.Colour(100, 100, 100))
        sizer.Add(note, 0, wx.ALL, 5)

        # Results list
        list_ctrl = wx.ListCtrl(panel, style=wx.LC_REPORT | wx.LC_HRULES)
        list_ctrl.InsertColumn(0, "Element", width=80)
        list_ctrl.InsertColumn(1, "Line", width=60)
        list_ctrl.InsertColumn(2, "Energy (keV)", width=100)
        list_ctrl.InsertColumn(3, "Rel. At.%", width=80)

        for i, res in enumerate(results):
            list_ctrl.InsertItem(i, res['element'])
            list_ctrl.SetItem(i, 1, res['line'])
            list_ctrl.SetItem(i, 2, f"{res['energy']:.2f}")
            list_ctrl.SetItem(i, 3, f"{res['atomic_percent']:.2f}")

        sizer.Add(list_ctrl, 1, wx.EXPAND | wx.ALL, 10)

        # Close button
        close_btn = wx.Button(panel, wx.ID_OK, "Close")
        sizer.Add(close_btn, 0, wx.ALL | wx.ALIGN_CENTER, 10)

        panel.SetSizer(sizer)
        dlg.ShowModal()
        dlg.Destroy()

    def _add_results_to_grid(self, results):
        """Add quantification results to the peak fitting grid"""
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return

        grid = self.parent.peak_params_grid

        # Clear existing grid
        if grid.GetNumberRows() > 0:
            grid.DeleteRows(0, grid.GetNumberRows())

        # Add results
        for i, res in enumerate(results):
            grid.AppendRows(2)  # Data row + constraint row
            row = i * 2

            # Format line name with Greek letters
            line_type = res['line']
            line_type = line_type.replace('Ka', 'Kα').replace('Kb', 'Kβ')
            line_type = line_type.replace('La', 'Lα').replace('Lb', 'Lβ')
            line_type = line_type.replace('Ma', 'Mα').replace('Mb', 'Mβ')

            # Peak label
            label = f"{res['element']} {line_type}"
            grid.SetCellValue(row, 1, label)

            # Position (energy)
            grid.SetCellValue(row, 2, f"{res['energy']:.2f}")

            # Intensity
            grid.SetCellValue(row, 3, f"{res['intensity']:.2f}")

            # Atomic percent in a relevant column (using column 6 for Area or similar)
            grid.SetCellValue(row, 6, f"{res['atomic_percent']:.2f}")

        self.parent.canvas.draw()

    def _populate_edx_grid_OLD(self, energy, spectrum, title="EDX"):
        """Calculate quantification and populate peak fitting grid"""
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return

        # Use identified Ka peaks if available (from add_peak_labels)
        if hasattr(self, 'identified_ka_peaks') and self.identified_ka_peaks:
            elements = [elem for _, elem, _ in self.identified_ka_peaks]
        else:
            # Fallback: detect elements fresh from the current spectrum
            elements = self._detect_elements_from_spectrum(energy, spectrum)

            # If no elements detected, try using selected elements
            if not elements and self.selected_elements:
                elements = self.selected_elements

        if not elements:
            # Clear grid if no elements
            grid = self.parent.peak_params_grid
            if grid.GetNumberRows() > 0:
                grid.DeleteRows(0, grid.GetNumberRows())
            grid.ForceRefresh()
            return

        # Calculate quantification results
        results = self._calculate_quantification(energy, spectrum, elements)

        if not results:
            # Clear grid if no results
            grid = self.parent.peak_params_grid
            if grid.GetNumberRows() > 0:
                grid.DeleteRows(0, grid.GetNumberRows())
            grid.ForceRefresh()
            return

        # Populate grid
        self._update_peak_params_grid(results, title)

    def _populate_edx_grid(self, energy, spectrum, title="EDX"):
        """
        Populate peak fitting grid using the SAME peaks identified by add_peak_labels.
        This ensures labels and quantification match.
        """
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return

        # Initialize quantification settings if not present
        if not hasattr(self, 'quant_settings'):
            self.quant_settings = EDXQuantificationSettings()

        # Use the peaks identified by add_peak_labels
        identified_peaks = getattr(self, 'identified_peaks', [])

        if not identified_peaks:
            # No peaks identified - clear grid
            grid = self.parent.peak_params_grid
            if grid.GetNumberRows() > 0:
                grid.DeleteRows(0, grid.GetNumberRows())
            grid.ForceRefresh()
            return

        # Group peaks by element for quantification (use primary line: Ka > La > Ma)
        element_peaks = {}
        for peak_energy, element, line_type in identified_peaks:
            if element not in element_peaks:
                element_peaks[element] = []
            element_peaks[element].append((peak_energy, line_type))

        # Calculate quantification using identified peaks
        results = self._calculate_quantification_from_identified(energy, spectrum, element_peaks)

        if results:
            self._update_peak_params_grid_hyperspy(results, self.quant_settings.composition_units)

    def _calculate_quantification_from_identified_OLD(self, energy, spectrum, element_peaks):
        """
        Calculate quantification using the peaks that were identified during labeling.
        This ensures consistency between plot labels and grid values.
        """
        try:
            from exspy.material import elements as exspy_elements

            settings = self.quant_settings if hasattr(self, 'quant_settings') else EDXQuantificationSettings()

            results = []
            intensities = []
            valid_elements = []
            kfactors = []

            for element, peaks in element_peaks.items():
                # Use the primary line (prefer Ka, then La, then Ma)
                primary_peak = None
                primary_line = None

                for peak_energy, line_type in peaks:
                    if line_type == 'Ka':
                        primary_peak = peak_energy
                        primary_line = 'Ka'
                        break
                    elif line_type == 'La' and primary_line != 'Ka':
                        primary_peak = peak_energy
                        primary_line = 'La'
                    elif line_type == 'Ma' and primary_line is None:
                        primary_peak = peak_energy
                        primary_line = 'Ma'

                if primary_peak is None:
                    # Use first available
                    primary_peak, primary_line = peaks[0]

                # Calculate intensity at this peak
                window = 0.15  # keV
                mask = (energy >= primary_peak - window) & (energy <= primary_peak + window)

                if not np.any(mask):
                    continue

                peak_spectrum = spectrum[mask]
                peak_energy_vals = energy[mask]

                # Background subtraction
                if settings.use_background_subtraction:
                    bg_width = settings.background_window_width
                    bg_left_mask = (energy >= primary_peak - window - bg_width) & (energy < primary_peak - window)
                    bg_right_mask = (energy > primary_peak + window) & (energy <= primary_peak + window + bg_width)

                    bg_left = spectrum[bg_left_mask].mean() if np.any(bg_left_mask) else 0
                    bg_right = spectrum[bg_right_mask].mean() if np.any(bg_right_mask) else 0
                    background = (bg_left + bg_right) / 2

                    net_intensity = np.sum(peak_spectrum) - background * len(peak_spectrum)
                else:
                    background = np.min(peak_spectrum)
                    net_intensity = np.sum(peak_spectrum - background)

                if net_intensity <= 0:
                    continue

                # Get k-factor
                xray_line = f"{element}_{primary_line}"
                kf = settings.get_kfactor(xray_line)

                # Peak parameters
                peak_height = float(np.max(peak_spectrum) - background)

                # FWHM estimation
                half_max = background + peak_height / 2
                above_half = peak_spectrum >= half_max
                if np.any(above_half):
                    indices = np.where(above_half)[0]
                    if len(indices) > 1:
                        fwhm_kev = peak_energy_vals[indices[-1]] - peak_energy_vals[indices[0]]
                    else:
                        fwhm_kev = 0.1
                else:
                    fwhm_kev = 0.1

                intensities.append(net_intensity)
                valid_elements.append(element)
                kfactors.append(kf)

                results.append({
                    'element': element,
                    'line': primary_line,
                    'energy': primary_peak,
                    'height': peak_height,
                    'area': net_intensity,
                    'fwhm': fwhm_kev * 1000,  # eV
                    'kfactor': kf,
                    'all_lines': peaks  # Store all identified lines for this element
                })

            # Calculate composition using Cliff-Lorimer
            if len(intensities) > 0:
                intensities_arr = np.array(intensities)
                kfactors_arr = np.array(kfactors)

                # Corrected intensities
                corrected = intensities_arr / kfactors_arr
                total_corrected = np.sum(corrected)

                if total_corrected > 0:
                    atomic_percent = (corrected / total_corrected) * 100

                    # Convert to weight percent if requested
                    if settings.composition_units == 'weight':
                        weight_factors = []
                        for elem in valid_elements:
                            try:
                                elem_obj = exspy_elements[elem]
                                weight_factors.append(elem_obj.General_properties.atomic_weight)
                            except:
                                weight_factors.append(1.0)

                        weight_factors = np.array(weight_factors)
                        weight_values = atomic_percent * weight_factors
                        composition = (weight_values / np.sum(weight_values)) * 100
                    else:
                        composition = atomic_percent

                    # Update results with composition
                    for i, res in enumerate(results):
                        res['concentration'] = composition[i]
                        res['units'] = 'wt%' if settings.composition_units == 'weight' else 'at%'

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _calculate_quantification_from_identified(self, energy, spectrum, element_peaks):
        """
        Calculate quantification using the peaks that were identified during labeling.
        This ensures consistency between plot labels and grid values.

        element_peaks is a dict: {element: [(peak_energy, line_type), ...]}
        """
        try:
            from libraries.EDX_Utilities import elements as edx_elements, DEFAULT_KFACTORS

            results = []
            intensities = []
            kfactors = []

            # element_peaks is a dict: {element: [(peak_energy, line_type), ...]}
            for element, peaks_list in element_peaks.items():
                try:
                    if element not in edx_elements:
                        continue

                    elem_obj = edx_elements[element]

                    # Use the first (primary) peak for this element
                    if not peaks_list:
                        continue

                    peak_energy, line_type = peaks_list[0]

                    # Find peak intensity
                    idx = np.argmin(np.abs(energy - peak_energy))
                    window = 5
                    start_idx = max(0, idx - window)
                    end_idx = min(len(spectrum), idx + window)

                    peak_spectrum = spectrum[start_idx:end_idx]
                    peak_height = float(np.max(peak_spectrum))
                    area = float(np.sum(peak_spectrum))

                    # Get k-factor
                    kfactor_key = f"{element}_{line_type}"
                    kfactor = DEFAULT_KFACTORS.get(kfactor_key, 1.0)

                    intensities.append(area)
                    kfactors.append(kfactor)

                    results.append({
                        'element': element,
                        'line': line_type,
                        'energy': peak_energy,
                        'height': peak_height,
                        'area': area,
                        'kfactor': kfactor
                    })

                except Exception as e:
                    print(f"Error processing {element}: {e}")
                    continue

            # Calculate atomic percentages using Cliff-Lorimer
            if len(intensities) > 0:
                intensities_arr = np.array(intensities)
                kfactors_arr = np.array(kfactors)

                corrected = intensities_arr / kfactors_arr
                total = np.sum(corrected)

                if total > 0:
                    for i, result in enumerate(results):
                        result['atomic_percent'] = (corrected[i] / total) * 100

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _detect_elements_from_spectrum(self, energy, spectrum):
        """Auto-detect elements from spectrum peaks"""
        try:
            # from exspy.utils.eds import get_xray_lines_near_energy

            from libraries.EDX_Utilities import (
                elements, get_xray_lines_near_energy, Signal1D,
                find_peaks1D_ohaver, DEFAULT_KFACTORS
            )
            from scipy.signal import find_peaks

            detected_elements = []
            detected_energies = {}

            # Find peaks in spectrum
            threshold = np.max(spectrum) * 0.03  # 3% threshold
            peaks, properties = find_peaks(spectrum, height=threshold, distance=10, prominence=threshold * 0.5)

            for peak_idx in peaks:
                if peak_idx >= len(energy):
                    continue
                peak_energy = energy[peak_idx]
                peak_height = spectrum[peak_idx]

                try:
                    lines_list = get_xray_lines_near_energy(peak_energy, width=0.12)
                    if lines_list:
                        # Get the first (best) match
                        line_str = lines_list[0]
                        parts = line_str.split('_')
                        if len(parts) == 2:
                            element = parts[0]
                            # Only add if not already detected or if this peak is stronger
                            if element not in detected_elements:
                                detected_elements.append(element)
                                detected_energies[element] = (peak_energy, peak_height)
                            elif peak_height > detected_energies[element][1]:
                                detected_energies[element] = (peak_energy, peak_height)
                except Exception:
                    continue

            # Limit to top 15 elements by peak height
            if len(detected_elements) > 15:
                sorted_elements = sorted(detected_elements,
                                         key=lambda e: detected_energies.get(e, (0, 0))[1],
                                         reverse=True)
                detected_elements = sorted_elements[:15]

            return detected_elements

        except Exception as e:
            print(f"Element detection error: {e}")
            return []

    def _detect_elements_from_spectrum(self, energy, spectrum):
        """Auto-detect elements from spectrum peaks"""
        try:
            from libraries.EDX_Utilities import get_xray_lines_near_energy, Signal1D

            # Create signal for peak finding
            temp_signal = Signal1D(spectrum)
            temp_signal.axes_manager[0].scale = energy[1] - energy[0] if len(energy) > 1 else 0.01
            temp_signal.axes_manager[0].offset = energy[0]

            # Find peaks
            amp_thresh = np.max(spectrum) * 0.02
            peak_data = temp_signal.find_peaks1D_ohaver(
                maxpeakn=50,
                medfilt_radius=5,
                amp_thresh=amp_thresh
            )

            if peak_data is None or len(peak_data) == 0:
                return []

            # Handle nested array
            if peak_data.ndim > 1:
                peaks = peak_data[0]
            else:
                peaks = peak_data

            if len(peaks) == 0:
                return []

            # Identify elements from peaks
            detected = set()
            for peak in peaks:
                try:
                    peak_energy = float(peak[0])
                    matches = get_xray_lines_near_energy(peak_energy, tolerance=0.15)
                    if matches:
                        # Take the closest match
                        detected.add(matches[0][0])  # Element symbol
                except (TypeError, IndexError):
                    continue

            return list(detected)

        except Exception as e:
            print(f"Error detecting elements: {e}")
            return []



    def _calculate_quantification_hyperspy_OLD(self, energy, spectrum):
        """
        Calculate EDX quantification using HyperSpy/ExSpy Cliff-Lorimer method.
        This provides accurate atomic% using proper k-factors.
        """
        import hyperspy.api as hs

        try:
            # Get quantification settings
            if not hasattr(self, 'quant_settings'):
                self.quant_settings = EDXQuantificationSettings()

            settings = self.quant_settings

            # Create a HyperSpy EDS signal from the spectrum
            s = hs.signals.Signal1D(spectrum.astype(float))

            # Set up the energy axis
            if len(energy) > 1:
                scale = energy[1] - energy[0]
            else:
                scale = 0.01
            s.axes_manager[0].scale = scale
            s.axes_manager[0].offset = energy[0]
            s.axes_manager[0].units = 'keV'
            s.axes_manager[0].name = 'Energy'

            # Try to set signal type to EDS_SEM
            try:
                s.set_signal_type("EDS_SEM")
            except:
                pass

            # Detect elements from spectrum
            elements = self._detect_elements_from_spectrum(energy, spectrum)

            if not elements:
                print("No elements detected for quantification")
                return []

            # Sort elements alphabetically (HyperSpy requirement)
            elements = sorted(elements)

            # Get k-factors for detected elements
            kfactors, xray_lines = settings.get_all_kfactors_for_elements(elements)

            print(f"Quantification - Elements: {elements}")
            print(f"X-ray lines: {xray_lines}")
            print(f"K-factors: {kfactors}")

            # Calculate intensities for each element
            from exspy.material import elements as exspy_elements

            intensities = []
            valid_elements = []
            valid_kfactors = []
            valid_lines = []

            for i, element in enumerate(elements):
                try:
                    elem_obj = exspy_elements[element]
                    if not hasattr(elem_obj, 'Atomic_properties'):
                        continue
                    if not hasattr(elem_obj.Atomic_properties, 'Xray_lines'):
                        continue

                    xray_props = elem_obj.Atomic_properties.Xray_lines

                    # Find the line energy
                    line_type = xray_lines[i].split('_')[1] if i < len(xray_lines) else 'Ka'
                    if line_type in xray_props:
                        line_energy = xray_props[line_type].energy_keV
                    else:
                        # Try to find any available line
                        for lt in ['Ka', 'La', 'Ma']:
                            if lt in xray_props:
                                line_type = lt
                                line_energy = xray_props[lt].energy_keV
                                break
                        else:
                            continue

                    # Define integration window
                    window = 0.15  # keV
                    mask = (energy >= line_energy - window) & (energy <= line_energy + window)

                    if not np.any(mask):
                        continue

                    peak_spectrum = spectrum[mask]

                    # Background subtraction if enabled
                    if settings.use_background_subtraction:
                        bg_width = settings.background_window_width
                        bg_left_mask = (energy >= line_energy - window - bg_width) & (energy < line_energy - window)
                        bg_right_mask = (energy > line_energy + window) & (energy <= line_energy + window + bg_width)

                        bg_left = spectrum[bg_left_mask].mean() if np.any(bg_left_mask) else 0
                        bg_right = spectrum[bg_right_mask].mean() if np.any(bg_right_mask) else 0
                        background = (bg_left + bg_right) / 2

                        net_intensity = np.sum(peak_spectrum) - background * len(peak_spectrum)
                    else:
                        net_intensity = np.sum(peak_spectrum)

                    if net_intensity > 0:
                        intensities.append(net_intensity)
                        valid_elements.append(element)
                        valid_kfactors.append(kfactors[i] if i < len(kfactors) else 1.0)
                        valid_lines.append(f"{element}_{line_type}")

                except Exception as e:
                    print(f"Error processing element {element}: {e}")
                    continue

            if len(intensities) < 2:
                print("Not enough valid elements for quantification")
                # Fall back to normalized intensity method
                return self._calculate_quantification(energy, spectrum, elements)

            # Calculate composition using Cliff-Lorimer method
            intensities_arr = np.array(intensities)
            kfactors_arr = np.array(valid_kfactors)

            # Cliff-Lorimer: C_A/C_B = k_AB * (I_A/I_B)
            # For multiple elements, normalize to sum to 100%

            # Corrected intensities (divide by k-factor)
            corrected = intensities_arr / kfactors_arr

            # Normalize to 100%
            total_corrected = np.sum(corrected)
            atomic_percent = (corrected / total_corrected) * 100

            # Convert to weight percent if requested
            if settings.composition_units == 'weight':
                weight_percent = []
                total_weight = 0

                for i, element in enumerate(valid_elements):
                    try:
                        elem_obj = exspy_elements[element]
                        atomic_weight = elem_obj.General_properties.atomic_weight
                        weight_percent.append(atomic_percent[i] * atomic_weight)
                        total_weight += atomic_percent[i] * atomic_weight
                    except:
                        weight_percent.append(atomic_percent[i])
                        total_weight += atomic_percent[i]

                weight_percent = np.array(weight_percent)
                weight_percent = (weight_percent / total_weight) * 100
                composition = weight_percent
                units = 'wt%'
            else:
                composition = atomic_percent
                units = 'at%'

            # Build results list
            results = []
            for i, element in enumerate(valid_elements):
                try:
                    elem_obj = exspy_elements[element]
                    xray_props = elem_obj.Atomic_properties.Xray_lines

                    line_type = valid_lines[i].split('_')[1]
                    if line_type in xray_props:
                        line_energy = xray_props[line_type].energy_keV
                    else:
                        line_energy = 0

                    # Get peak parameters
                    window = 0.15
                    mask = (energy >= line_energy - window) & (energy <= line_energy + window)
                    peak_spectrum = spectrum[mask]
                    peak_height = float(np.max(peak_spectrum)) if len(peak_spectrum) > 0 else 0

                    # Estimate FWHM
                    if len(peak_spectrum) > 0:
                        half_max = np.min(peak_spectrum) + (np.max(peak_spectrum) - np.min(peak_spectrum)) / 2
                        above_half = peak_spectrum >= half_max
                        if np.any(above_half):
                            indices = np.where(above_half)[0]
                            if len(indices) > 1:
                                peak_energy_vals = energy[mask]
                                fwhm_kev = peak_energy_vals[indices[-1]] - peak_energy_vals[indices[0]]
                            else:
                                fwhm_kev = 0.1
                        else:
                            fwhm_kev = 0.1
                    else:
                        fwhm_kev = 0.1

                    # Area calculation
                    area = intensities[i]

                    results.append({
                        'element': element,
                        'line': line_type,
                        'energy': line_energy,
                        'height': peak_height,
                        'area': area,
                        'fwhm': fwhm_kev * 1000,  # Convert to eV
                        'concentration': composition[i],
                        'units': units,
                        'kfactor': valid_kfactors[i]
                    })

                except Exception as e:
                    print(f"Error building result for {element}: {e}")
                    continue

            print(f"\nQuantification Results ({units}):")
            for res in results:
                print(f"  {res['element']} ({res['line']}): {res['concentration']:.2f} {units}")

            # Update grid with results
            self._update_peak_params_grid_hyperspy(results, units)

            return results

        except Exception as e:
            print(f"HyperSpy quantification error: {e}")
            import traceback
            traceback.print_exc()
            # Fall back to simple method
            return self._calculate_quantification(energy, spectrum, elements if 'elements' in dir() else [])

    def _calculate_quantification_hyperspy(self, energy, spectrum):
        """
        Calculate EDX quantification using Cliff-Lorimer method.
        This provides accurate atomic% using proper k-factors.
        Now uses standalone EDX_Utilities instead of HyperSpy.
        """
        try:
            from libraries.EDX_Utilities import elements as edx_elements, Signal1D, DEFAULT_KFACTORS

            # Get quantification settings
            if not hasattr(self, 'quant_settings'):
                self.quant_settings = EDXQuantificationSettings()

            settings = self.quant_settings

            # Detect elements from spectrum
            detected_elements = self._detect_elements_from_spectrum(energy, spectrum)

            if not detected_elements:
                print("No elements detected for quantification")
                return []

            # Sort elements alphabetically
            detected_elements = sorted(detected_elements)

            # Get k-factors for detected elements
            kfactors, xray_lines = settings.get_all_kfactors_for_elements(detected_elements)

            print(f"Quantification - Elements: {detected_elements}")
            print(f"X-ray lines: {xray_lines}")
            print(f"K-factors: {kfactors}")

            # Calculate intensities for each element
            intensities = []
            valid_elements = []
            valid_kfactors = []
            valid_lines = []

            for i, element in enumerate(detected_elements):
                try:
                    if element not in edx_elements:
                        continue

                    elem_obj = edx_elements[element]
                    if elem_obj.Atomic_properties is None:
                        continue
                    if elem_obj.Atomic_properties.Xray_lines is None:
                        continue

                    xray_props = elem_obj.Atomic_properties.Xray_lines

                    # Find the line energy
                    line_type = xray_lines[i].split('_')[1] if i < len(xray_lines) else 'Ka'

                    line_obj = getattr(xray_props, line_type, None)
                    if line_obj is not None:
                        line_energy = line_obj.energy_keV
                    else:
                        # Try to find any available line
                        line_energy = None
                        for lt in ['Ka', 'La', 'Ma']:
                            line_obj = getattr(xray_props, lt, None)
                            if line_obj is not None:
                                line_type = lt
                                line_energy = line_obj.energy_keV
                                break
                        if line_energy is None:
                            continue

                    # Define integration window
                    window = 0.15  # keV
                    mask = (energy >= line_energy - window) & (energy <= line_energy + window)

                    if not np.any(mask):
                        continue

                    peak_spectrum = spectrum[mask]

                    # Background subtraction if enabled
                    if settings.use_background_subtraction:
                        bg_width = settings.background_window_width
                        bg_left_mask = (energy >= line_energy - window - bg_width) & (energy < line_energy - window)
                        bg_right_mask = (energy > line_energy + window) & (energy <= line_energy + window + bg_width)

                        bg_left = spectrum[bg_left_mask].mean() if np.any(bg_left_mask) else 0
                        bg_right = spectrum[bg_right_mask].mean() if np.any(bg_right_mask) else 0
                        background = (bg_left + bg_right) / 2

                        net_intensity = np.sum(peak_spectrum) - background * len(peak_spectrum)
                    else:
                        net_intensity = np.sum(peak_spectrum)

                    if net_intensity > 0:
                        intensities.append(net_intensity)
                        valid_elements.append(element)
                        valid_kfactors.append(kfactors[i] if i < len(kfactors) else 1.0)
                        valid_lines.append(f"{element}_{line_type}")

                except Exception as e:
                    print(f"Error processing element {element}: {e}")
                    continue

            if len(intensities) < 2:
                print("Not enough valid elements for quantification")
                return self._calculate_quantification(energy, spectrum, detected_elements)

            # Calculate composition using Cliff-Lorimer method
            intensities_arr = np.array(intensities)
            kfactors_arr = np.array(valid_kfactors)

            # Corrected intensities (divide by k-factor)
            corrected = intensities_arr / kfactors_arr

            # Normalize to 100%
            total_corrected = np.sum(corrected)
            atomic_percent = (corrected / total_corrected) * 100

            # Convert to weight percent if requested
            if settings.composition_units == 'weight':
                weight_percent = []
                total_weight = 0

                for i, element in enumerate(valid_elements):
                    try:
                        elem_obj = edx_elements[element]
                        atomic_weight = elem_obj.General_properties.atomic_weight
                        weight_percent.append(atomic_percent[i] * atomic_weight)
                        total_weight += atomic_percent[i] * atomic_weight
                    except:
                        weight_percent.append(atomic_percent[i])
                        total_weight += atomic_percent[i]

                weight_percent = np.array(weight_percent)
                weight_percent = (weight_percent / total_weight) * 100
                composition = weight_percent
                units = 'wt%'
            else:
                composition = atomic_percent
                units = 'at%'

            # Build results list
            results = []
            for i, element in enumerate(valid_elements):
                try:
                    elem_obj = edx_elements[element]
                    xray_props = elem_obj.Atomic_properties.Xray_lines

                    line_type = valid_lines[i].split('_')[1]
                    line_obj = getattr(xray_props, line_type, None)
                    line_energy = line_obj.energy_keV if line_obj else 0

                    # Get peak parameters
                    window = 0.15
                    mask = (energy >= line_energy - window) & (energy <= line_energy + window)
                    peak_spectrum = spectrum[mask]
                    peak_height = float(np.max(peak_spectrum)) if len(peak_spectrum) > 0 else 0

                    # Estimate FWHM
                    if len(peak_spectrum) > 0:
                        half_max = np.min(peak_spectrum) + (np.max(peak_spectrum) - np.min(peak_spectrum)) / 2
                        above_half = peak_spectrum >= half_max
                        if np.any(above_half):
                            indices = np.where(above_half)[0]
                            if len(indices) > 1:
                                peak_energy_vals = energy[mask]
                                fwhm_kev = peak_energy_vals[indices[-1]] - peak_energy_vals[indices[0]]
                            else:
                                fwhm_kev = 0.1
                        else:
                            fwhm_kev = 0.1
                    else:
                        fwhm_kev = 0.1

                    # Area calculation
                    area = intensities[i]

                    results.append({
                        'element': element,
                        'line': line_type,
                        'energy': line_energy,
                        'height': peak_height,
                        'fwhm': fwhm_kev,
                        'area': area,
                        'atomic_percent': atomic_percent[i],
                        'composition': composition[i],
                        'units': units,
                        'kfactor': valid_kfactors[i]
                    })

                except Exception as e:
                    print(f"Error building result for {element}: {e}")
                    continue

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _update_peak_params_grid_hyperspy(self, results, units="at%"):
        """Update peak fitting grid with HyperSpy quantification results"""
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return

        grid = self.parent.peak_params_grid

        # Clear existing grid
        if grid.GetNumberRows() > 0:
            grid.DeleteRows(0, grid.GetNumberRows())

        for i, res in enumerate(results):
            grid.AppendRows(2)
            row = i * 2
            constraint_row = row + 1

            # Format line name with Greek letters
            line_type = res['line']
            line_display = line_type.replace('Ka', 'Kα').replace('Kb', 'Kβ')
            line_display = line_display.replace('La', 'Lα').replace('Lb', 'Lβ')
            line_display = line_display.replace('Ma', 'Mα').replace('Mb', 'Mβ')

            # Column 0: ID
            letter_id = chr(65 + i)
            grid.SetCellValue(row, 0, letter_id)

            # Column 1: Peak label
            label = f"{res['element']} {line_display}"
            grid.SetCellValue(row, 1, label)

            # Column 2: Position (energy in eV)
            grid.SetCellValue(row, 2, f"{res['energy'] * 1000:.2f}")
            grid.SetCellValue(constraint_row, 2, "fixed")

            # Column 3: Height
            grid.SetCellValue(row, 3, f"{res['height']:.2f}")
            grid.SetCellValue(constraint_row, 3, "fixed")

            # Column 4: FWHM (in eV)
            grid.SetCellValue(row, 4, f"{res['fwhm']:.2f}")
            grid.SetCellValue(constraint_row, 4, "fixed")

            # Column 6: Area
            grid.SetCellValue(row, 6, f"{res['area']:.2f}")
            grid.SetCellValue(constraint_row, 6, "fixed")

            # Column 10: Concentration (%) - with units indicator
            conc_str = f"{res['concentration']:.2f}"
            grid.SetCellValue(row, 10, conc_str)

            # Column 11: K-factor (if column exists)
            if grid.GetNumberCols() > 11:
                grid.SetCellValue(row, 11, f"{res.get('kfactor', 1.0):.4f}")

            # Apply formatting
            for col in range(grid.GetNumberCols()):
                grid.SetCellBackgroundColour(row, col, wx.WHITE)
                grid.SetReadOnly(row, col, True)
                grid.SetCellBackgroundColour(constraint_row, col, wx.Colour(200, 245, 228))
                grid.SetReadOnly(constraint_row, col, True)

        grid.ForceRefresh()

        # Store grid data for persistence
        if hasattr(self.parent, 'Data') and 'Core levels' in self.parent.Data:
            current_sheet = self.parent.sheet_combobox.GetValue()
            if current_sheet in self.parent.Data['Core levels']:
                grid_data = []
                for row in range(grid.GetNumberRows()):
                    row_data = [grid.GetCellValue(row, col) for col in range(grid.GetNumberCols())]
                    grid_data.append(row_data)
                self.parent.Data['Core levels'][current_sheet]['_EDX_grid_data'] = grid_data

    def _calculate_quantification_OLD(self, energy, spectrum, elements):
        """Calculate quantification with peak fitting parameters"""
        try:
            from exspy.material import elements as exspy_elements

            results = []
            total_intensity = 0

            for element in elements:
                try:
                    elem_obj = exspy_elements[element]
                    if not hasattr(elem_obj, 'Atomic_properties'):
                        continue
                    if not hasattr(elem_obj.Atomic_properties, 'Xray_lines'):
                        continue

                    xray_lines = elem_obj.Atomic_properties.Xray_lines

                    # Find the strongest line (Ka preferred, then La, Ma)
                    best_line = None
                    best_energy_val = None
                    for line_type in ['Ka', 'La', 'Ma']:
                        if line_type in xray_lines:
                            best_line = line_type
                            best_energy_val = xray_lines[line_type].energy_keV
                            break

                    if best_energy_val is None:
                        continue

                    # Find peak in spectrum near this energy
                    idx = np.argmin(np.abs(energy - best_energy_val))

                    # Define integration window (±0.15 keV typical for EDX)
                    energy_window = 0.15  # keV
                    mask = (energy >= best_energy_val - energy_window) & (energy <= best_energy_val + energy_window)

                    if not np.any(mask):
                        continue

                    # Extract peak region
                    peak_energy = energy[mask]
                    peak_spectrum = spectrum[mask]

                    if len(peak_spectrum) == 0:
                        continue

                    # Calculate peak parameters
                    peak_height = float(np.max(peak_spectrum))

                    # Estimate background as minimum in the window
                    background = float(np.min(peak_spectrum))

                    # Net peak height
                    net_height = peak_height - background

                    # Estimate FWHM from the peak shape
                    half_max = background + net_height / 2
                    above_half = peak_spectrum >= half_max
                    if np.any(above_half):
                        indices = np.where(above_half)[0]
                        if len(indices) > 1:
                            fwhm_kev = peak_energy[indices[-1]] - peak_energy[indices[0]]
                        else:
                            fwhm_kev = 0.1  # Default 100 eV
                    else:
                        fwhm_kev = 0.1

                    # Calculate area using Gaussian approximation: Area ≈ Height × FWHM × 1.064
                    # This is more accurate than trapezoid integration for peaks
                    peak_area = net_height * (fwhm_kev * 1000) * 1.064  # Convert FWHM to eV for area calc

                    # Get atomic weight
                    atomic_weight = elem_obj.General_properties.atomic_weight

                    results.append({
                        'element': element,
                        'line': best_line,
                        'energy': best_energy_val,
                        'height': net_height,
                        'area': peak_area,
                        'fwhm': fwhm_kev * 1000,  # Store in eV
                        'atomic_weight': atomic_weight
                    })
                    total_intensity += peak_area

                except (KeyError, AttributeError) as e:
                    print(f"Could not process element {element}: {e}")
                    continue

            # Calculate atomic percentages
            if total_intensity > 0:
                for res in results:
                    res['concentration'] = (res['area'] / total_intensity) * 100

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _calculate_quantification(self, energy, spectrum, elements):
        """Calculate quantification using EDX_Utilities (no exspy/hyperspy).

        Args:
            energy: Energy axis array (keV)
            spectrum: Spectrum intensity array
            elements: List of element symbols to quantify

        Returns:
            List of dicts with element, line, energy, height, area, concentration, atomic_percent
        """
        try:
            from libraries.EDX_Utilities import (
                elements as edx_elements,
                get_element_xray_lines,
                DEFAULT_KFACTORS
            )

            results = []
            intensities = []
            kfactors = []

            for element in elements:
                try:
                    if element not in edx_elements:
                        continue

                    # Get X-ray line energies for this element using helper function
                    xray_lines = get_element_xray_lines(element)

                    if not xray_lines:
                        continue

                    # Find the best line (Ka preferred, then La, Ma)
                    best_line = None
                    best_energy_val = None

                    for line_type in ['Ka', 'La', 'Ma']:
                        if line_type in xray_lines and xray_lines[line_type] is not None:
                            best_line = line_type
                            best_energy_val = xray_lines[line_type]
                            break

                    if best_energy_val is None:
                        continue

                    # Find peak in spectrum near this energy
                    idx = np.argmin(np.abs(energy - best_energy_val))

                    # Define integration window (±0.15 keV typical for EDX)
                    energy_window = 0.15  # keV
                    mask = (energy >= best_energy_val - energy_window) & (energy <= best_energy_val + energy_window)

                    if not np.any(mask):
                        continue

                    # Extract peak region
                    peak_energy = energy[mask]
                    peak_spectrum = spectrum[mask]

                    if len(peak_spectrum) == 0:
                        continue

                    # Calculate peak parameters
                    peak_height = float(np.max(peak_spectrum))

                    # Estimate background as minimum in the window
                    background = float(np.min(peak_spectrum))

                    # Net peak height
                    net_height = peak_height - background

                    # Estimate FWHM from the peak shape
                    half_max = background + net_height / 2
                    above_half = peak_spectrum >= half_max
                    if np.any(above_half):
                        indices = np.where(above_half)[0]
                        if len(indices) > 1:
                            fwhm_kev = peak_energy[indices[-1]] - peak_energy[indices[0]]
                        else:
                            fwhm_kev = 0.1  # Default 100 eV
                    else:
                        fwhm_kev = 0.1

                    # Calculate area using Gaussian approximation
                    peak_area = net_height * (fwhm_kev * 1000) * 1.064

                    # Get k-factor
                    kfactor_key = f"{element}_{best_line}"
                    kfactor = DEFAULT_KFACTORS.get(kfactor_key, 1.0)

                    results.append({
                        'element': element,
                        'line': best_line,
                        'energy': best_energy_val,
                        'height': net_height,
                        'area': peak_area,
                        'fwhm': fwhm_kev * 1000,  # Store in eV
                        'kfactor': kfactor
                    })
                    intensities.append(peak_area)
                    kfactors.append(kfactor)

                except Exception as e:
                    print(f"Could not process element {element}: {e}")
                    continue

            # Calculate atomic percentages using Cliff-Lorimer
            if len(intensities) > 0:
                intensities_arr = np.array(intensities)
                kfactors_arr = np.array(kfactors)

                corrected = intensities_arr / kfactors_arr
                total = np.sum(corrected)

                if total > 0:
                    for i, res in enumerate(results):
                        at_percent = (corrected[i] / total) * 100
                        res['concentration'] = at_percent
                        res['atomic_percent'] = at_percent
                        res['at_percent'] = at_percent
                        res['At%'] = at_percent

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []


    def _calculate_quantification_from_identified(self, energy, spectrum, element_peaks):
        """
        Calculate quantification using the peaks that were identified during labeling.
        This ensures consistency between plot labels and grid values.

        element_peaks is a dict: {element: [(peak_energy, line_type), ...]}
        """
        try:
            from libraries.EDX_Utilities import elements as edx_elements, DEFAULT_KFACTORS

            results = []
            total_intensity = 0

            # element_peaks is a dict: {element: [(peak_energy, line_type), ...]}
            for element, peaks_list in element_peaks.items():
                try:
                    if element not in edx_elements:
                        continue

                    elem_obj = edx_elements[element]

                    # Use the first (primary) peak for this element
                    if not peaks_list:
                        continue

                    peak_energy, line_type = peaks_list[0]

                    # Find peak in spectrum near this energy
                    idx = np.argmin(np.abs(energy - peak_energy))

                    # Define integration window (±0.15 keV typical for EDX)
                    energy_window = 0.15  # keV
                    mask = (energy >= peak_energy - energy_window) & (energy <= peak_energy + energy_window)

                    if not np.any(mask):
                        continue

                    # Extract peak region
                    peak_energy_arr = energy[mask]
                    peak_spectrum = spectrum[mask]

                    if len(peak_spectrum) == 0:
                        continue

                    # Calculate peak parameters
                    peak_height = float(np.max(peak_spectrum))

                    # Estimate background as minimum in the window
                    background = float(np.min(peak_spectrum))

                    # Net peak height
                    net_height = peak_height - background

                    # Estimate FWHM from the peak shape
                    half_max = background + net_height / 2
                    above_half = peak_spectrum >= half_max
                    if np.any(above_half):
                        indices = np.where(above_half)[0]
                        if len(indices) > 1:
                            fwhm_kev = peak_energy_arr[indices[-1]] - peak_energy_arr[indices[0]]
                        else:
                            fwhm_kev = 0.1  # Default 100 eV
                    else:
                        fwhm_kev = 0.1

                    # Calculate area using Gaussian approximation: Area ≈ Height × FWHM × 1.064
                    peak_area = net_height * (fwhm_kev * 1000) * 1.064  # Convert FWHM to eV for area calc

                    # Get atomic weight
                    atomic_weight = elem_obj.General_properties.atomic_weight

                    # Get k-factor
                    kfactor_key = f"{element}_{line_type}"
                    kfactor = DEFAULT_KFACTORS.get(kfactor_key, 1.0)

                    results.append({
                        'element': element,
                        'line': line_type,
                        'energy': peak_energy,
                        'height': net_height,
                        'area': peak_area,
                        'fwhm': fwhm_kev * 1000,  # Store in eV
                        'atomic_weight': atomic_weight,
                        'kfactor': kfactor
                    })
                    total_intensity += peak_area

                except Exception as e:
                    print(f"Error processing {element}: {e}")
                    continue

            # Calculate concentration (atomic percentages)
            if total_intensity > 0:
                for res in results:
                    res['concentration'] = (res['area'] / total_intensity) * 100
                    res['atomic_percent'] = res['concentration']  # Alias for compatibility

            return results

        except Exception as e:
            print(f"Quantification error: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _update_peak_params_grid(self, results, title="EDX"):
        """Update peak fitting grid with EDX results"""
        if self.parent is None or not hasattr(self.parent, 'peak_params_grid'):
            return

        grid = self.parent.peak_params_grid

        # Clear existing grid completely
        if grid.GetNumberRows() > 0:
            grid.DeleteRows(0, grid.GetNumberRows())

        # Add results
        for i, res in enumerate(results):
            grid.AppendRows(2)  # Data row + constraint row
            row = i * 2
            constraint_row = row + 1

            # Format line name with Greek letters
            line_type = res['line']
            line_display = line_type.replace('Ka', 'Kα').replace('Kb', 'Kβ')
            line_display = line_display.replace('La', 'Lα').replace('Lb', 'Lβ')
            line_display = line_display.replace('Ma', 'Mα').replace('Mb', 'Mβ')

            # Column 0: ID (letter)
            letter_id = chr(65 + i)  # A, B, C, ...
            grid.SetCellValue(row, 0, letter_id)

            # Column 1: Peak label
            label = f"{res['element']} {line_display}"
            grid.SetCellValue(row, 1, label)

            # Column 2: Position (energy in keV displayed as eV * 1000)
            grid.SetCellValue(row, 2, f"{res['energy'] * 1000:.2f}")
            grid.SetCellValue(constraint_row, 2, "fixed")

            # Column 3: Height
            grid.SetCellValue(row, 3, f"{res['height']:.2f}")
            grid.SetCellValue(constraint_row, 3, "fixed")

            # Column 4: FWHM (in eV)
            grid.SetCellValue(row, 4, f"{res['fwhm']:.2f}")
            grid.SetCellValue(constraint_row, 4, "fixed")

            # Column 6: Area
            grid.SetCellValue(row, 6, f"{res['area']:.2f}")
            grid.SetCellValue(constraint_row, 6, "fixed")

            # Column 10: Concentration (%)
            grid.SetCellValue(row, 10, f"{res.get('concentration', 0):.2f}")

            # Apply formatting - Data row (white background, read-only)
            for col in range(grid.GetNumberCols()):
                grid.SetCellBackgroundColour(row, col, wx.WHITE)
                grid.SetReadOnly(row, col, True)

            # Constraint row - light green background, read-only
            for col in range(grid.GetNumberCols()):
                grid.SetCellBackgroundColour(constraint_row, col, wx.Colour(200, 245, 228))
                grid.SetReadOnly(constraint_row, col, True)

        grid.ForceRefresh()

    def on_change_colormap(self, cmap):
        """Change colormap and replot"""
        self.current_cmap = cmap
        self.plot_current_map()

    def on_show_data_browser(self, event):
        """Show data browser window"""
        if not hasattr(self, 'data_browser_window') or self.data_browser_window is None:
            self.data_browser_window = DataBrowserWindow(self)
        self.data_browser_window.Show()
        self.data_browser_window.Raise()

    def on_show_info(self, event):
        """Show info window"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "Info", wx.OK | wx.ICON_INFORMATION)
            return

        if not hasattr(self, 'info_window') or self.info_window is None:
            self.info_window = InfoWindow(self)
        self.info_window.update_info(self.current_data)
        self.info_window.Show()
        self.info_window.Raise()

    def on_intensity_map(self, event):
        """Show intensity/sum map"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "No Data", wx.OK | wx.ICON_WARNING)
            return
        self.plot_current_map()

    def on_map_panel_resize(self, event):
        """Position arrow control panel on bottom-left of map canvas"""
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            # Get canvas size and position
            canvas_rect = self.map_canvas.GetRect()

            # Get arrow panel size
            arrow_size = self.arrow_control_panel.GetBestSize()

            # Position on bottom-left with 10px margins
            x = canvas_rect.x
            y = canvas_rect.y + canvas_rect.height - arrow_size.height

            self.arrow_control_panel.SetPosition((x, y))
            self.arrow_control_panel.SetSize(arrow_size)
            self.arrow_control_panel.Raise()  # Bring to front
            self.arrow_control_panel.Show()

        event.Skip()

    def on_canvas_paint(self, event):
        """Keep arrow buttons visible when canvas redraws"""
        event.Skip()  # Let the paint event continue

        # Re-raise the arrow panel to keep it on top after a short delay
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            wx.CallLater(10, self._reposition_arrow_panel)

    def _reposition_arrow_panel(self):
        """Helper to reposition arrow panel"""
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            canvas_rect = self.map_canvas.GetRect()
            arrow_size = self.arrow_control_panel.GetBestSize()

            x = canvas_rect.x
            y = canvas_rect.y + canvas_rect.height - arrow_size.height

            self.arrow_control_panel.SetPosition((x, y))
            self.arrow_control_panel.Show()
            self.arrow_control_panel.Raise()
            self.arrow_control_panel.Refresh()
            self.arrow_control_panel.Update()

    def _on_mpl_draw(self, event):
        """Called after matplotlib draws - keep buttons on top"""
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            wx.CallLater(5, self._reposition_arrow_panel)

    def on_arrow_move(self, direction):
        """Handle arrow button clicks to move selected elements"""
        if self.current_data is None:
            return

        # Get the actual data array from HyperSpy signal
        if hasattr(self.current_data, 'data'):
            data_array = self.current_data.data
        else:
            data_array = self.current_data

        # If nothing is selected, clear saved selections and return
        if (not self.selected_points and
            not self.last_line_start and
            not (hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.center is not None) and
            not self.selected_areas):
            self._clear_saved_selection()
            self.map_canvas.draw()
            return

        # Get plot dimensions for calculating movement step
        # Use 0.5% of plot dimension as movement step
        height, width = data_array.shape[:2]
        step_x = int(width * 0.005)
        step_y = int(height * 0.005)

        # Ensure minimum movement of 1 pixel
        step_x = max(1, step_x)
        step_y = max(1, step_y)

        # Determine movement direction
        dx, dy = 0, 0
        if direction == 'left':
            dx = -step_x
        elif direction == 'right':
            dx = step_x
        elif direction == 'up':
            dy = -step_y
        elif direction == 'down':
            dy = step_y

        moved = False

        # Move selected points
        if self.selected_points:
            moved = self._move_and_replot_point(dx, dy, width, height)

        # Move selected lines (use last positions)
        elif self.last_line_start and self.last_line_end:
            moved = self._move_and_replot_line(dx, dy, width, height)

        # Move rotatable rectangle if it exists (priority over selected_areas)
        elif hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.center is not None:
            moved = self._move_and_replot_rotated(dx, dy, width, height)

        # Move selected areas (rectangles)
        elif self.selected_areas:
            moved = self._move_and_replot_area(dx, dy, width, height)

        if moved:
            print(f"Moved selection by ({dx}, {dy})")

    def _move_and_replot_point(self, dx, dy, width, height):
        """Move point and replot"""
        if not self.selected_points:
            return False

        new_points = []
        for point in self.selected_points:
            if len(point) == 3:
                x, y, size = point
            else:
                x, y = point
                size = 1

            new_x = int(np.clip(x + dx, 0, width - 1))
            new_y = int(np.clip(y + dy, 0, height - 1))
            new_points.append((new_x, new_y, size))

        self.selected_points = new_points

        # Redraw marker (red)
        self.clear_selection_markers()
        point = self.selected_points[-1]
        x, y, size = point

        if size == 1:
            marker, = self.map_ax.plot(x, y, 'r+', markersize=15, markeredgewidth=2)
            marker._is_selection_marker = True
        else:
            from matplotlib.patches import Rectangle
            half_size = size / 2
            rect = Rectangle((x - half_size, y - half_size), size, size,
                             linewidth=2, edgecolor='darkred', facecolor='red', alpha=0.3)
            rect._is_selection_marker = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw()
        self.map_canvas.Refresh()

        # Replot spectrum
        self.plot_point_spectrum(x, y)

        # Keep arrow buttons visible
        wx.CallLater(10, self._reposition_arrow_panel)

        return True

    def _move_and_replot_line(self, dx, dy, width, height):
        """Move line and replot"""
        if not self.last_line_start or not self.last_line_end:
            return False

        x1, y1 = self.last_line_start
        x2, y2 = self.last_line_end

        # Move both endpoints
        new_x1 = int(np.clip(x1 + dx, 0, width - 1))
        new_y1 = int(np.clip(y1 + dy, 0, height - 1))
        new_x2 = int(np.clip(x2 + dx, 0, width - 1))
        new_y2 = int(np.clip(y2 + dy, 0, height - 1))

        self.last_line_start = (new_x1, new_y1)
        self.last_line_end = (new_x2, new_y2)

        # Get stored line width
        line_width = getattr(self, 'last_line_width', self.line_width)

        # Redraw markers with width (red)
        self.clear_selection_markers()
        self._draw_line_with_width((new_x1, new_y1), (new_x2, new_y2), line_width)

        self.map_canvas.draw()

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        # Create line profile
        self.create_line_profile(new_x1, new_y1, new_x2, new_y2, line_width)

        return True

    def _move_and_replot_area(self, dx, dy, width, height):
        """Move area and replot"""
        if not self.selected_areas:
            return False

        x1, y1, x2, y2 = self.selected_areas[-1]

        # Calculate rectangle dimensions
        rect_width = x2 - x1
        rect_height = y2 - y1

        # Move and clip to boundaries
        new_x1 = int(np.clip(x1 + dx, 0, width - rect_width))
        new_y1 = int(np.clip(y1 + dy, 0, height - rect_height))
        new_x2 = new_x1 + rect_width
        new_y2 = new_y1 + rect_height

        self.selected_areas[-1] = (new_x1, new_y1, new_x2, new_y2)

        # Redraw marker (red)
        self.clear_selection_markers()

        from matplotlib.patches import Rectangle
        rect = Rectangle((new_x1, new_y1), rect_width, rect_height,
                         linewidth=2, edgecolor='red', facecolor='red', alpha=0.2)
        rect._is_selection_marker = True
        self.map_ax.add_patch(rect)

        self.map_canvas.draw()

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        # Replot spectrum
        self.plot_area_spectrum(new_x1, new_y1, new_x2, new_y2)

        return True

    def _move_and_replot_rotated(self, dx, dy, width, height):
        """Move rotatable rectangle and replot"""
        if not hasattr(self, 'rotatable_rect') or not self.rotatable_rect:
            return False

        if self.rotatable_rect.center is None:
            return False

        # Get current center
        cx, cy = self.rotatable_rect.center

        # Calculate new center with boundary checking
        new_cx = float(np.clip(cx + dx, 0, width - 1))
        new_cy = float(np.clip(cy + dy, 0, height - 1))

        # Update center
        self.rotatable_rect.center = (new_cx, new_cy)

        # Redraw the rectangle using the RotatableRectangle's draw method
        self.rotatable_rect.draw_rectangle()

        self.map_canvas.draw()

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        # Replot spectrum for rotated area using SAME method as on_rotated_area_complete
        data = self.current_data.data
        if len(data.shape) != 3:
            return False

        # Get dimensions
        map_height, map_width = data.shape[0], data.shape[1]
        w = self.rotatable_rect.width
        h = self.rotatable_rect.height
        angle = self.rotatable_rect.angle
        hw, hh = w / 2, h / 2

        # Create rotation matrix
        angle_rad = np.radians(-angle)  # Negative for inverse transform
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        # Create mask for pixels inside rotated rectangle
        y_indices, x_indices = np.meshgrid(np.arange(map_height), np.arange(map_width), indexing='ij')

        # Transform all pixels to rectangle's local coordinate system
        local_x = (x_indices - new_cx) * cos_a - (y_indices - new_cy) * sin_a
        local_y = (x_indices - new_cx) * sin_a + (y_indices - new_cy) * cos_a

        # Create mask for pixels inside rectangle
        mask = (np.abs(local_x) <= hw) & (np.abs(local_y) <= hh)

        # Extract spectra from masked pixels
        masked_spectra = data[mask]

        if len(masked_spectra) == 0:
            print("No pixels in selected area")
            return False

        # Sum spectra
        summed_spectrum = np.sum(masked_spectra, axis=0)

        # Get energy axis
        energy_axis = self.current_data.axes_manager[-1]
        energy = energy_axis.axis

        # Plot to parent window using common styled method
        if self.parent is not None:
            title = f'EDX Area Spectrum (Rotated, {len(masked_spectra)} pixels)'
            grid_label = f"Rotated Area ({len(masked_spectra)} px)"
            self._plot_edx_spectrum_styled(energy, summed_spectrum, title, grid_label)

        print(f"Moved rotated area: center=({new_cx:.0f},{new_cy:.0f}), size=({w:.0f}x{h:.0f}), angle={angle:.1f}Â°")
        print(f"Extracted {len(masked_spectra)} pixels")

        return True

    def on_close(self, event):
        """Clear parent reference when closing"""
        if hasattr(self.parent, 'edx_window'):
            self.parent.edx_window = None
        event.Skip()  # Allow normal close


class DataBrowserWindow(wx.Frame):
    """Separate window for data browser"""

    def __init__(self, parent):
        super().__init__(parent, title="Data Browser", size=(400, 500))
        self.parent = parent

        panel = wx.Panel(self)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Data tree
        self.data_tree = wx.TreeCtrl(panel, style=wx.TR_DEFAULT_STYLE | wx.TR_HAS_BUTTONS)
        self.tree_root = self.data_tree.AddRoot("Loaded Data")
        sizer.Add(self.data_tree, 1, wx.EXPAND | wx.ALL, 5)

        # Refresh from parent
        self.refresh_tree()

        panel.SetSizer(sizer)

        self.data_tree.Bind(wx.EVT_TREE_SEL_CHANGED, self.on_tree_select)
        self.Bind(wx.EVT_CLOSE, self.on_close)

    def refresh_tree(self):
        """Refresh tree from parent data"""
        self.data_tree.DeleteChildren(self.tree_root)

        if hasattr(self.parent, 'loaded_signals'):
            for item in self.parent.loaded_signals:
                signal = item['data']
                filename = item['filename']

                title = filename
                if hasattr(signal, 'metadata') and hasattr(signal.metadata, 'General'):
                    if hasattr(signal.metadata.General, 'title') and signal.metadata.General.title:
                        title = signal.metadata.General.title

                item_text = f"{title} ({signal.data.shape})"
                tree_item = self.data_tree.AppendItem(self.tree_root, item_text)
                self.data_tree.SetItemData(tree_item, item)

                # Add metadata children
                self.add_metadata_to_tree(tree_item, signal)

        self.data_tree.Expand(self.tree_root)

    def add_metadata_to_tree(self, parent_item, signal):
        """Add signal metadata as tree children"""
        self.data_tree.AppendItem(parent_item, f"Shape: {signal.data.shape}")
        self.data_tree.AppendItem(parent_item, f"Dtype: {signal.data.dtype}")

        if hasattr(signal, 'axes_manager'):
            axes_item = self.data_tree.AppendItem(parent_item, "Axes")
            for i, axis in enumerate(signal.axes_manager.signal_axes):
                axis_text = f"Signal {i}: {axis.name} [{axis.units}] - {axis.size} pts"
                self.data_tree.AppendItem(axes_item, axis_text)
            for i, axis in enumerate(signal.axes_manager.navigation_axes):
                axis_text = f"Nav {i}: {axis.name} [{axis.units}] - {axis.size} pts"
                self.data_tree.AppendItem(axes_item, axis_text)

    def on_tree_select(self, event):
        """Handle tree selection"""
        item = event.GetItem()
        data = self.data_tree.GetItemData(item)

        if data is not None and 'data' in data:
            self.parent.current_data = data['data']
            self.parent.plot_current_map()
            self.parent.extract_elements(data['data'])

    def on_close(self, event):
        """Handle close"""
        self.Hide()


class InfoWindow(wx.Frame):
    """Separate window for data info"""

    def __init__(self, parent):
        super().__init__(parent, title="Data Info", size=(350, 400))
        self.parent = parent

        panel = wx.Panel(self)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.info_text = wx.TextCtrl(panel, style=wx.TE_MULTILINE | wx.TE_READONLY)
        sizer.Add(self.info_text, 1, wx.EXPAND | wx.ALL, 5)

        panel.SetSizer(sizer)
        self.Bind(wx.EVT_CLOSE, self.on_close)

    def update_info(self, signal):
        """Update info display"""
        info = []
        info.append(f"Shape: {signal.data.shape}")
        info.append(f"Data type: {signal.data.dtype}")
        info.append(f"Min: {np.min(signal.data):.2f}")
        info.append(f"Max: {np.max(signal.data):.2f}")
        info.append(f"Mean: {np.mean(signal.data):.2f}")

        if hasattr(signal, 'axes_manager'):
            info.append("")
            info.append("Axes:")
            for axis in signal.axes_manager.signal_axes:
                info.append(f"  {axis.name}: {axis.size} pts, {axis.scale:.4f} {axis.units}/pt")

            # Navigation axes with size calculation
            for axis in signal.axes_manager.navigation_axes:
                scale_str = f"{axis.scale:.2f}" if axis.scale >= 0.01 else f"{axis.scale:.4f}"
                info.append(f"  {axis.name}: {axis.size} pts, {scale_str} {axis.units}/pixel")
                # Show total dimension
                total_size = axis.size * axis.scale
                if axis.units == 'nm':
                    if total_size >= 1000:
                        info.append(f"    Total: {total_size / 1000:.2f} µm")
                    else:
                        info.append(f"    Total: {total_size:.2f} nm")
                elif axis.units:
                    info.append(f"    Total: {total_size:.2f} {axis.units}")

            # Summary for map data
            if signal.data.ndim == 3:
                nav_axes = signal.axes_manager.navigation_axes
                if len(nav_axes) >= 2:
                    info.append("")
                    info.append("Map dimensions:")
                    x_total = nav_axes[1].size * nav_axes[1].scale if len(nav_axes) > 1 else 0
                    y_total = nav_axes[0].size * nav_axes[0].scale
                    units = nav_axes[0].units if nav_axes[0].units else 'pixels'
                    if units == 'nm' and x_total >= 1000:
                        info.append(f"  {x_total / 1000:.2f} × {y_total / 1000:.2f} µm")
                    else:
                        info.append(f"  {x_total:.2f} × {y_total:.2f} {units}")
                    info.append(f"  ({nav_axes[1].size if len(nav_axes) > 1 else 1} × {nav_axes[0].size} pixels)")

        self.info_text.SetValue('\n'.join(info))

    def on_close(self, event):
        self.Hide()


class PeriodicTableDialog(wx.Dialog):
    """
    Periodic table element selector with three states:
    - Green: Include (prioritize this element)
    - Red: Exclude (never identify as this element)
    - Grey: Neutral (no preference)
    """

    def __init__(self, parent, include_elements=None, exclude_elements=None):
        super().__init__(parent, title="Select Elements", size=(570, 350))

        self.include_elements = set(include_elements or [])
        self.exclude_elements = set(exclude_elements or [])
        self.element_buttons = {}
        self.element_states = {}  # 0=grey, 1=green, 2=red

        # Initialize states
        for elem in self.include_elements:
            self.element_states[elem] = 1  # Green
        for elem in self.exclude_elements:
            self.element_states[elem] = 2  # Red

        panel = wx.Panel(self, style=wx.BORDER_RAISED)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Legend
        legend_sizer = wx.BoxSizer(wx.HORIZONTAL)
        legend_sizer.Add(wx.StaticText(panel, label="Click to cycle: "), 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)

        grey_label = wx.StaticText(panel, label=" Grey (neutral) ")
        grey_label.SetBackgroundColour(wx.Colour(220, 220, 220))
        legend_sizer.Add(grey_label, 0, wx.ALL, 2)

        legend_sizer.Add(wx.StaticText(panel, label="→"), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)

        green_label = wx.StaticText(panel, label=" Green (include) ")
        green_label.SetBackgroundColour(wx.Colour(100, 200, 100))
        legend_sizer.Add(green_label, 0, wx.ALL, 2)

        legend_sizer.Add(wx.StaticText(panel, label="→"), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)

        red_label = wx.StaticText(panel, label=" Red (exclude) ")
        red_label.SetBackgroundColour(wx.Colour(255, 100, 100))
        legend_sizer.Add(red_label, 0, wx.ALL, 2)

        legend_sizer.Add(wx.StaticText(panel, label="→ Grey..."), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)

        main_sizer.Add(legend_sizer, 0, wx.EXPAND | wx.ALL, 5)

        # Periodic table layout
        table_panel = wx.Panel(panel)
        grid_sizer = wx.GridBagSizer(0, 0)

        # Complete periodic table structure (row, col, symbol)
        # Main table - periods 1-7
        elements = [
            # Period 1
            (0, 0, 'H'), (0, 17, 'He'),
            # Period 2
            (1, 0, 'Li'), (1, 1, 'Be'), (1, 12, 'B'), (1, 13, 'C'), (1, 14, 'N'),
            (1, 15, 'O'), (1, 16, 'F'), (1, 17, 'Ne'),
            # Period 3
            (2, 0, 'Na'), (2, 1, 'Mg'), (2, 12, 'Al'), (2, 13, 'Si'), (2, 14, 'P'),
            (2, 15, 'S'), (2, 16, 'Cl'), (2, 17, 'Ar'),
            # Period 4
            (3, 0, 'K'), (3, 1, 'Ca'), (3, 2, 'Sc'), (3, 3, 'Ti'), (3, 4, 'V'),
            (3, 5, 'Cr'), (3, 6, 'Mn'), (3, 7, 'Fe'), (3, 8, 'Co'), (3, 9, 'Ni'),
            (3, 10, 'Cu'), (3, 11, 'Zn'), (3, 12, 'Ga'), (3, 13, 'Ge'), (3, 14, 'As'),
            (3, 15, 'Se'), (3, 16, 'Br'), (3, 17, 'Kr'),
            # Period 5
            (4, 0, 'Rb'), (4, 1, 'Sr'), (4, 2, 'Y'), (4, 3, 'Zr'), (4, 4, 'Nb'),
            (4, 5, 'Mo'), (4, 6, 'Tc'), (4, 7, 'Ru'), (4, 8, 'Rh'), (4, 9, 'Pd'),
            (4, 10, 'Ag'), (4, 11, 'Cd'), (4, 12, 'In'), (4, 13, 'Sn'), (4, 14, 'Sb'),
            (4, 15, 'Te'), (4, 16, 'I'), (4, 17, 'Xe'),
            # Period 6 (La placeholder at col 2, then Hf onwards)
            (5, 0, 'Cs'), (5, 1, 'Ba'), (5, 3, 'Hf'), (5, 4, 'Ta'),
            (5, 5, 'W'), (5, 6, 'Re'), (5, 7, 'Os'), (5, 8, 'Ir'), (5, 9, 'Pt'),
            (5, 10, 'Au'), (5, 11, 'Hg'), (5, 12, 'Tl'), (5, 13, 'Pb'), (5, 14, 'Bi'),
            (5, 15, 'Po'), (5, 16, 'At'), (5, 17, 'Rn'),
            # Period 7 (Ac placeholder at col 2, then Rf onwards)
            (6, 0, 'Fr'), (6, 1, 'Ra'), (6, 3, 'Rf'), (6, 4, 'Db'),
            (6, 5, 'Sg'), (6, 6, 'Bh'), (6, 7, 'Hs'), (6, 8, 'Mt'), (6, 9, 'Ds'),
            (6, 10, 'Rg'), (6, 11, 'Cn'), (6, 12, 'Nh'), (6, 13, 'Fl'), (6, 14, 'Mc'),
            (6, 15, 'Lv'), (6, 16, 'Ts'), (6, 17, 'Og'),
        ]

        # Lanthanides (row 8) - La through Lu
        lanthanides = [
            (8, 2, 'La'), (8, 3, 'Ce'), (8, 4, 'Pr'), (8, 5, 'Nd'), (8, 6, 'Pm'),
            (8, 7, 'Sm'), (8, 8, 'Eu'), (8, 9, 'Gd'), (8, 10, 'Tb'), (8, 11, 'Dy'),
            (8, 12, 'Ho'), (8, 13, 'Er'), (8, 14, 'Tm'), (8, 15, 'Yb'), (8, 16, 'Lu'),
        ]

        # Actinides (row 9) - Ac through Lr
        actinides = [
            (9, 2, 'Ac'), (9, 3, 'Th'), (9, 4, 'Pa'), (9, 5, 'U'), (9, 6, 'Np'),
            (9, 7, 'Pu'), (9, 8, 'Am'), (9, 9, 'Cm'), (9, 10, 'Bk'), (9, 11, 'Cf'),
            (9, 12, 'Es'), (9, 13, 'Fm'), (9, 14, 'Md'), (9, 15, 'No'), (9, 16, 'Lr'),
        ]

        # Add placeholder indicators for lanthanides/actinides in main table
        la_indicator = wx.StaticText(table_panel, label="La-Lu→")
        la_indicator.SetFont(wx.Font(7, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_NORMAL))
        grid_sizer.Add(la_indicator, pos=(5, 2), flag=wx.ALIGN_CENTER)

        ac_indicator = wx.StaticText(table_panel, label="Ac-Lr→")
        ac_indicator.SetFont(wx.Font(7, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_NORMAL))
        grid_sizer.Add(ac_indicator, pos=(6, 2), flag=wx.ALIGN_CENTER)

        # Add row labels for lanthanides and actinides
        ln_label = wx.StaticText(table_panel, label="Ln")
        ln_label.SetFont(wx.Font(8, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_BOLD))
        grid_sizer.Add(ln_label, pos=(8, 0), span=(1, 2), flag=wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_RIGHT | wx.RIGHT, border=5)

        an_label = wx.StaticText(table_panel, label="An")
        an_label.SetFont(wx.Font(8, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_BOLD))
        grid_sizer.Add(an_label, pos=(9, 0), span=(1, 2), flag=wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_RIGHT | wx.RIGHT, border=5)

        # Add spacer row between main table and lanthanides/actinides
        spacer = wx.StaticText(table_panel, label="")
        grid_sizer.Add(spacer, pos=(7, 0), flag=wx.ALL, border=5)

        # Combine all elements
        all_elements = elements + lanthanides + actinides

        for row, col, symbol in all_elements:
            btn = wx.Button(table_panel, label=symbol, size=(30, 22))
            btn.SetFont(wx.Font(8, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

            # Set initial state
            state = self.element_states.get(symbol, 0)
            self._apply_button_color(btn, state)

            btn.Bind(wx.EVT_BUTTON, lambda evt, s=symbol, b=btn: self.on_element_click(evt, s, b))
            grid_sizer.Add(btn, pos=(row, col), flag=wx.ALL, border=0)
            self.element_buttons[symbol] = btn

        table_panel.SetSizer(grid_sizer)
        main_sizer.Add(table_panel, 1, wx.EXPAND | wx.ALL, 0)

        # Buttons
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)

        clear_btn = wx.Button(panel, label="Clear All")
        clear_btn.Bind(wx.EVT_BUTTON, self.on_clear_all)
        btn_sizer.Add(clear_btn, 0, wx.ALL, 5)

        btn_sizer.AddStretchSpacer()

        ok_btn = wx.Button(panel, wx.ID_OK, "OK")
        cancel_btn = wx.Button(panel, wx.ID_CANCEL, "Cancel")

        btn_sizer.Add(ok_btn, 0, wx.ALL, 5)
        btn_sizer.Add(cancel_btn, 0, wx.ALL, 5)

        main_sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 5)

        panel.SetSizer(main_sizer)
        self.Centre()

    def _apply_button_color(self, btn, state):
        """Apply color based on state: 0=grey, 1=green, 2=red"""
        if state == 1:  # Green - include
            btn.SetBackgroundColour(wx.Colour(100, 200, 100))
        elif state == 2:  # Red - exclude
            btn.SetBackgroundColour(wx.Colour(255, 100, 100))
        else:  # Grey - neutral
            btn.SetBackgroundColour(wx.Colour(220, 220, 220))
        btn.Refresh()

    def on_element_click(self, event, symbol, btn):
        """Cycle through states: grey -> green -> red -> grey"""
        current_state = self.element_states.get(symbol, 0)
        new_state = (current_state + 1) % 3

        self.element_states[symbol] = new_state
        self._apply_button_color(btn, new_state)

        # Update sets
        if new_state == 1:  # Green
            self.include_elements.add(symbol)
            self.exclude_elements.discard(symbol)
        elif new_state == 2:  # Red
            self.include_elements.discard(symbol)
            self.exclude_elements.add(symbol)
        else:  # Grey
            self.include_elements.discard(symbol)
            self.exclude_elements.discard(symbol)

    def on_clear_all(self, event):
        """Reset all elements to grey (neutral)"""
        self.include_elements = set()
        self.exclude_elements = set()
        self.element_states = {}

        for symbol, btn in self.element_buttons.items():
            self._apply_button_color(btn, 0)

    def get_include_elements(self):
        """Return set of elements to include (green)"""
        return self.include_elements

    def get_exclude_elements(self):
        """Return set of elements to exclude (red)"""
        return self.exclude_elements

    def get_selected_elements(self):
        """Legacy method - returns include elements as list"""
        return list(self.include_elements)


# ==================== EDX QUANTIFICATION SETTINGS ====================

# Complete k-factors database for EDX quantification
# K-factors are relative to Si Ka = 1.0
# Covers elements Z=3 (Li) to Z=92 (U), excluding radioactive/dangerous elements
# Values based on theoretical calculations and empirical measurements

EDX_KFACTORS_DATABASE = {
    'Generic SEM': {
        'description': 'Generic SEM EDX (standardless approximation)',
        'beam_energy': 15.0,
        'kfactors': {
            # Period 2 (Z=3-10)
            'Li_Ka': 15.50, 'Be_Ka': 8.20, 'B_Ka': 5.45, 'C_Ka': 3.78,
            'N_Ka': 3.02, 'O_Ka': 1.98, 'F_Ka': 1.52, 'Ne_Ka': 1.25,
            # Period 3 (Z=11-18)
            'Na_Ka': 0.97, 'Mg_Ka': 0.85, 'Al_Ka': 0.87, 'Si_Ka': 1.00,
            'P_Ka': 1.09, 'S_Ka': 1.08, 'Cl_Ka': 1.12, 'Ar_Ka': 1.18,
            # Period 4 (Z=19-36)
            'K_Ka': 1.07, 'Ca_Ka': 1.09, 'Sc_Ka': 1.05, 'Ti_Ka': 1.01,
            'V_Ka': 1.00, 'Cr_Ka': 1.02, 'Mn_Ka': 1.04, 'Fe_Ka': 1.05,
            'Co_Ka': 1.08, 'Ni_Ka': 1.12, 'Cu_Ka': 1.18, 'Zn_Ka': 1.28,
            'Ga_Ka': 1.40, 'Ge_Ka': 1.55, 'As_Ka': 1.72, 'Se_Ka': 1.92,
            'Br_Ka': 2.15, 'Kr_Ka': 2.40,
            # Period 5 (Z=37-54)
            'Rb_Ka': 2.68, 'Sr_Ka': 2.98, 'Y_Ka': 3.30, 'Zr_Ka': 3.65,
            'Nb_Ka': 4.02, 'Mo_Ka': 3.28, 'Ru_Ka': 4.85, 'Rh_Ka': 5.30,
            'Pd_Ka': 5.78, 'Ag_Ka': 6.28, 'Cd_Ka': 6.82, 'In_Ka': 7.38,
            'Sn_Ka': 7.98, 'Sb_Ka': 8.60, 'Te_Ka': 9.25, 'I_Ka': 9.95,
            'Xe_Ka': 10.68,
            # Period 5 L-lines (used when K-lines too high energy)
            'Rb_La': 1.45, 'Sr_La': 1.52, 'Y_La': 1.60, 'Zr_La': 1.68,
            'Nb_La': 1.77, 'Mo_La': 1.86, 'Ru_La': 2.05, 'Rh_La': 2.16,
            'Pd_La': 2.28, 'Ag_La': 2.40, 'Cd_La': 2.53, 'In_La': 2.67,
            'Sn_La': 2.10, 'Sb_La': 2.00, 'Te_La': 2.12, 'I_La': 2.25,
            'Xe_La': 2.38,
            # Period 6 (Z=55-86) - L-lines primarily
            'Cs_La': 1.78, 'Ba_La': 1.85, 'La_La': 1.75, 'Ce_La': 1.82,
            'Pr_La': 1.89, 'Nd_La': 1.96, 'Sm_La': 2.10, 'Eu_La': 2.18,
            'Gd_La': 2.26, 'Tb_La': 2.34, 'Dy_La': 2.43, 'Ho_La': 2.52,
            'Er_La': 2.61, 'Tm_La': 2.70, 'Yb_La': 2.80, 'Lu_La': 2.90,
            'Hf_La': 2.02, 'Ta_La': 2.08, 'W_La': 2.15, 'Re_La': 2.22,
            'Os_La': 2.30, 'Ir_La': 2.38, 'Pt_La': 2.45, 'Au_La': 2.55,
            'Hg_La': 2.28, 'Tl_La': 2.35, 'Pb_La': 2.35, 'Bi_La': 2.40,
            # Period 6 M-lines (for heavy elements)
            'Hf_Ma': 1.65, 'Ta_Ma': 1.70, 'W_Ma': 1.75, 'Re_Ma': 1.80,
            'Os_Ma': 1.85, 'Ir_Ma': 1.90, 'Pt_Ma': 1.95, 'Au_Ma': 2.00,
            'Hg_Ma': 2.05, 'Tl_Ma': 2.10, 'Pb_Ma': 2.15, 'Bi_Ma': 2.20,
            # Period 7 (Actinides) - L and M lines
            'Th_La': 2.55, 'Th_Ma': 2.30, 'U_La': 2.65, 'U_Ma': 2.40,
        }
    },
    'Bruker XFlash': {
        'description': 'Bruker XFlash SDD detector',
        'beam_energy': 15.0,
        'kfactors': {
            # Period 2
            'Li_Ka': 16.20, 'Be_Ka': 8.55, 'B_Ka': 5.68, 'C_Ka': 4.21,
            'N_Ka': 3.35, 'O_Ka': 2.18, 'F_Ka': 1.65, 'Ne_Ka': 1.32,
            # Period 3
            'Na_Ka': 1.02, 'Mg_Ka': 0.88, 'Al_Ka': 0.89, 'Si_Ka': 1.00,
            'P_Ka': 1.12, 'S_Ka': 1.10, 'Cl_Ka': 1.14, 'Ar_Ka': 1.20,
            # Period 4
            'K_Ka': 1.09, 'Ca_Ka': 1.11, 'Sc_Ka': 1.06, 'Ti_Ka': 1.03,
            'V_Ka': 1.02, 'Cr_Ka': 1.04, 'Mn_Ka': 1.06, 'Fe_Ka': 1.07,
            'Co_Ka': 1.10, 'Ni_Ka': 1.14, 'Cu_Ka': 1.20, 'Zn_Ka': 1.30,
            'Ga_Ka': 1.42, 'Ge_Ka': 1.58, 'As_Ka': 1.75, 'Se_Ka': 1.95,
            'Br_Ka': 2.18, 'Kr_Ka': 2.44,
            # Period 5
            'Rb_Ka': 2.72, 'Sr_Ka': 3.02, 'Y_Ka': 3.35, 'Zr_Ka': 3.70,
            'Nb_Ka': 4.08, 'Mo_Ka': 3.35, 'Ru_Ka': 4.92, 'Rh_Ka': 5.38,
            'Pd_Ka': 5.86, 'Ag_Ka': 6.38, 'Cd_Ka': 6.92, 'In_Ka': 7.48,
            'Sn_Ka': 8.10, 'Sb_Ka': 8.74, 'Te_Ka': 9.42, 'I_Ka': 10.12,
            'Xe_Ka': 10.86,
            # L-lines
            'Rb_La': 1.48, 'Sr_La': 1.55, 'Y_La': 1.63, 'Zr_La': 1.72,
            'Nb_La': 1.81, 'Mo_La': 1.90, 'Ag_La': 2.45, 'Sn_La': 2.15,
            'Sb_La': 2.05, 'Te_La': 2.18, 'I_La': 2.30,
            # Period 6
            'Cs_La': 1.82, 'Ba_La': 1.89, 'La_La': 1.79, 'Ce_La': 1.86,
            'Pr_La': 1.93, 'Nd_La': 2.00, 'Sm_La': 2.15, 'Eu_La': 2.23,
            'Gd_La': 2.31, 'Tb_La': 2.39, 'Dy_La': 2.48, 'Ho_La': 2.57,
            'Er_La': 2.66, 'Tm_La': 2.76, 'Yb_La': 2.86, 'Lu_La': 2.96,
            'Hf_La': 2.06, 'Ta_La': 2.12, 'W_La': 2.20, 'Re_La': 2.27,
            'Os_La': 2.35, 'Ir_La': 2.43, 'Pt_La': 2.50, 'Au_La': 2.60,
            'Hg_La': 2.33, 'Tl_La': 2.40, 'Pb_La': 2.40, 'Bi_La': 2.45,
            # M-lines
            'Hf_Ma': 1.68, 'Ta_Ma': 1.73, 'W_Ma': 1.78, 'Pt_Ma': 1.98,
            'Au_Ma': 2.05, 'Pb_Ma': 2.20, 'Bi_Ma': 2.25,
            # Actinides
            'Th_La': 2.60, 'Th_Ma': 2.35, 'U_La': 2.70, 'U_Ma': 2.45,
        }
    },
    'Oxford X-Max': {
        'description': 'Oxford Instruments X-Max SDD',
        'beam_energy': 15.0,
        'kfactors': {
            # Period 2
            'Li_Ka': 15.80, 'Be_Ka': 8.35, 'B_Ka': 5.55, 'C_Ka': 3.92,
            'N_Ka': 3.15, 'O_Ka': 2.05, 'F_Ka': 1.58, 'Ne_Ka': 1.28,
            # Period 3
            'Na_Ka': 0.99, 'Mg_Ka': 0.86, 'Al_Ka': 0.88, 'Si_Ka': 1.00,
            'P_Ka': 1.10, 'S_Ka': 1.09, 'Cl_Ka': 1.13, 'Ar_Ka': 1.19,
            # Period 4
            'K_Ka': 1.08, 'Ca_Ka': 1.10, 'Sc_Ka': 1.05, 'Ti_Ka': 1.02,
            'V_Ka': 1.01, 'Cr_Ka': 1.03, 'Mn_Ka': 1.05, 'Fe_Ka': 1.06,
            'Co_Ka': 1.09, 'Ni_Ka': 1.13, 'Cu_Ka': 1.19, 'Zn_Ka': 1.29,
            'Ga_Ka': 1.41, 'Ge_Ka': 1.56, 'As_Ka': 1.73, 'Se_Ka': 1.93,
            'Br_Ka': 2.16, 'Kr_Ka': 2.42,
            # Period 5
            'Rb_Ka': 2.70, 'Sr_Ka': 3.00, 'Y_Ka': 3.32, 'Zr_Ka': 3.68,
            'Nb_Ka': 4.05, 'Mo_Ka': 3.30, 'Ru_Ka': 4.88, 'Rh_Ka': 5.34,
            'Pd_Ka': 5.82, 'Ag_Ka': 6.32, 'Cd_Ka': 6.86, 'In_Ka': 7.42,
            'Sn_Ka': 8.02, 'Sb_Ka': 8.66, 'Te_Ka': 9.32, 'I_Ka': 10.02,
            # L-lines
            'Ag_La': 2.42, 'Sn_La': 2.12, 'Sb_La': 2.02, 'Te_La': 2.15,
            'I_La': 2.28,
            # Period 6
            'Cs_La': 1.80, 'Ba_La': 1.87, 'La_La': 1.77, 'Ce_La': 1.84,
            'Pr_La': 1.91, 'Nd_La': 1.98, 'Sm_La': 2.12, 'Eu_La': 2.20,
            'Gd_La': 2.28, 'Tb_La': 2.36, 'Dy_La': 2.45, 'Ho_La': 2.54,
            'Er_La': 2.63, 'Tm_La': 2.73, 'Yb_La': 2.82, 'Lu_La': 2.92,
            'Hf_La': 2.04, 'Ta_La': 2.10, 'W_La': 2.17, 'Re_La': 2.24,
            'Os_La': 2.32, 'Ir_La': 2.40, 'Pt_La': 2.47, 'Au_La': 2.57,
            'Hg_La': 2.30, 'Tl_La': 2.37, 'Pb_La': 2.37, 'Bi_La': 2.42,
            # M-lines
            'Hf_Ma': 1.66, 'Ta_Ma': 1.71, 'W_Ma': 1.76, 'Pt_Ma': 1.96,
            'Au_Ma': 2.02, 'Pb_Ma': 2.17, 'Bi_Ma': 2.22,
            # Actinides
            'Th_La': 2.57, 'Th_Ma': 2.32, 'U_La': 2.67, 'U_Ma': 2.42,
        }
    },
    'JEOL JED': {
        'description': 'JEOL JED Series detector',
        'beam_energy': 15.0,
        'kfactors': {
            # Period 2
            'Li_Ka': 16.00, 'Be_Ka': 8.45, 'B_Ka': 5.60, 'C_Ka': 4.05,
            'N_Ka': 3.22, 'O_Ka': 2.10, 'F_Ka': 1.60, 'Ne_Ka': 1.30,
            # Period 3
            'Na_Ka': 1.00, 'Mg_Ka': 0.87, 'Al_Ka': 0.88, 'Si_Ka': 1.00,
            'P_Ka': 1.11, 'S_Ka': 1.09, 'Cl_Ka': 1.13, 'Ar_Ka': 1.19,
            # Period 4
            'K_Ka': 1.08, 'Ca_Ka': 1.10, 'Sc_Ka': 1.06, 'Ti_Ka': 1.02,
            'V_Ka': 1.01, 'Cr_Ka': 1.03, 'Mn_Ka': 1.05, 'Fe_Ka': 1.06,
            'Co_Ka': 1.09, 'Ni_Ka': 1.13, 'Cu_Ka': 1.19, 'Zn_Ka': 1.29,
            'Ga_Ka': 1.41, 'Ge_Ka': 1.57, 'As_Ka': 1.74, 'Se_Ka': 1.94,
            'Br_Ka': 2.17, 'Kr_Ka': 2.43,
            # Period 5
            'Rb_Ka': 2.71, 'Sr_Ka': 3.01, 'Y_Ka': 3.33, 'Zr_Ka': 3.69,
            'Nb_Ka': 4.06, 'Mo_Ka': 3.32, 'Ru_Ka': 4.90, 'Rh_Ka': 5.36,
            'Pd_Ka': 5.84, 'Ag_Ka': 6.34, 'Cd_Ka': 6.88, 'In_Ka': 7.44,
            'Sn_Ka': 8.05, 'Sb_Ka': 8.68, 'Te_Ka': 9.35, 'I_Ka': 10.05,
            # L-lines
            'Ag_La': 2.43, 'Sn_La': 2.13, 'Sb_La': 2.03, 'Te_La': 2.16,
            'I_La': 2.29,
            # Period 6
            'Cs_La': 1.81, 'Ba_La': 1.88, 'La_La': 1.78, 'Ce_La': 1.85,
            'Nd_La': 1.99, 'Sm_La': 2.13, 'Gd_La': 2.29, 'Dy_La': 2.46,
            'Er_La': 2.64, 'Yb_La': 2.83, 'Hf_La': 2.05, 'Ta_La': 2.11,
            'W_La': 2.18, 'Pt_La': 2.48, 'Au_La': 2.58, 'Pb_La': 2.38,
            'Bi_La': 2.43,
            # M-lines
            'W_Ma': 1.77, 'Pt_Ma': 1.97, 'Au_Ma': 2.03, 'Pb_Ma': 2.18,
            # Actinides
            'Th_La': 2.58, 'U_La': 2.68,
        }
    },
    'Thermo Noran': {
        'description': 'Thermo Scientific Noran System',
        'beam_energy': 15.0,
        'kfactors': {
            # Period 2
            'Li_Ka': 15.60, 'Be_Ka': 8.25, 'B_Ka': 5.48, 'C_Ka': 3.85,
            'N_Ka': 3.08, 'O_Ka': 2.00, 'F_Ka': 1.55, 'Ne_Ka': 1.26,
            # Period 3
            'Na_Ka': 0.98, 'Mg_Ka': 0.85, 'Al_Ka': 0.87, 'Si_Ka': 1.00,
            'P_Ka': 1.10, 'S_Ka': 1.08, 'Cl_Ka': 1.12, 'Ar_Ka': 1.18,
            # Period 4
            'K_Ka': 1.07, 'Ca_Ka': 1.09, 'Sc_Ka': 1.05, 'Ti_Ka': 1.01,
            'V_Ka': 1.00, 'Cr_Ka': 1.02, 'Mn_Ka': 1.04, 'Fe_Ka': 1.05,
            'Co_Ka': 1.08, 'Ni_Ka': 1.12, 'Cu_Ka': 1.18, 'Zn_Ka': 1.28,
            'Ga_Ka': 1.40, 'Ge_Ka': 1.55, 'As_Ka': 1.72, 'Se_Ka': 1.92,
            'Br_Ka': 2.15, 'Kr_Ka': 2.40,
            # Period 5
            'Rb_Ka': 2.68, 'Sr_Ka': 2.98, 'Y_Ka': 3.30, 'Zr_Ka': 3.65,
            'Nb_Ka': 4.02, 'Mo_Ka': 3.28, 'Ru_Ka': 4.85, 'Rh_Ka': 5.30,
            'Pd_Ka': 5.78, 'Ag_Ka': 6.28, 'Cd_Ka': 6.82, 'In_Ka': 7.38,
            'Sn_Ka': 7.98, 'Sb_Ka': 8.60, 'Te_Ka': 9.25, 'I_Ka': 9.95,
            # L-lines
            'Ag_La': 2.40, 'Sn_La': 2.10, 'Sb_La': 2.00, 'Te_La': 2.12,
            'I_La': 2.25,
            # Period 6
            'Cs_La': 1.78, 'Ba_La': 1.85, 'La_La': 1.75, 'Ce_La': 1.82,
            'Nd_La': 1.96, 'Sm_La': 2.10, 'Gd_La': 2.26, 'Dy_La': 2.43,
            'Er_La': 2.61, 'Yb_La': 2.80, 'Hf_La': 2.02, 'Ta_La': 2.08,
            'W_La': 2.15, 'Pt_La': 2.45, 'Au_La': 2.55, 'Pb_La': 2.35,
            'Bi_La': 2.40,
            # M-lines
            'W_Ma': 1.75, 'Pt_Ma': 1.95, 'Au_Ma': 2.00, 'Pb_Ma': 2.15,
            # Actinides
            'Th_La': 2.55, 'U_La': 2.65,
        }
    },
    'TEM Generic': {
        'description': 'Generic TEM EDX (200 keV)',
        'beam_energy': 200.0,
        'kfactors': {
            # Period 2
            'Li_Ka': 12.50, 'Be_Ka': 6.80, 'B_Ka': 4.55, 'C_Ka': 2.98,
            'N_Ka': 2.55, 'O_Ka': 1.82, 'F_Ka': 1.45, 'Ne_Ka': 1.18,
            # Period 3
            'Na_Ka': 0.94, 'Mg_Ka': 0.83, 'Al_Ka': 0.86, 'Si_Ka': 1.00,
            'P_Ka': 1.08, 'S_Ka': 1.06, 'Cl_Ka': 1.10, 'Ar_Ka': 1.15,
            # Period 4
            'K_Ka': 1.05, 'Ca_Ka': 1.07, 'Sc_Ka': 1.02, 'Ti_Ka': 0.99,
            'V_Ka': 0.98, 'Cr_Ka': 1.00, 'Mn_Ka': 1.02, 'Fe_Ka': 1.03,
            'Co_Ka': 1.06, 'Ni_Ka': 1.10, 'Cu_Ka': 1.15, 'Zn_Ka': 1.24,
            'Ga_Ka': 1.35, 'Ge_Ka': 1.48, 'As_Ka': 1.63, 'Se_Ka': 1.80,
            'Br_Ka': 2.00, 'Kr_Ka': 2.22,
            # Period 5
            'Rb_Ka': 2.46, 'Sr_Ka': 2.72, 'Y_Ka': 3.00, 'Zr_Ka': 3.30,
            'Nb_Ka': 3.62, 'Mo_Ka': 2.95, 'Ru_Ka': 4.32, 'Rh_Ka': 4.70,
            'Pd_Ka': 5.10, 'Ag_Ka': 5.52, 'Cd_Ka': 5.98, 'In_Ka': 6.46,
            'Sn_Ka': 6.98, 'Sb_Ka': 7.52, 'Te_Ka': 8.10, 'I_Ka': 8.70,
            # L-lines
            'Ag_La': 2.20, 'Sn_La': 1.92, 'Sb_La': 1.82, 'Te_La': 1.94,
            'I_La': 2.06,
            # Period 6
            'Cs_La': 1.62, 'Ba_La': 1.68, 'La_La': 1.58, 'Ce_La': 1.65,
            'Nd_La': 1.78, 'Sm_La': 1.91, 'Gd_La': 2.05, 'Dy_La': 2.20,
            'Er_La': 2.36, 'Yb_La': 2.52, 'Hf_La': 1.82, 'Ta_La': 1.88,
            'W_La': 1.94, 'Pt_La': 2.20, 'Au_La': 2.28, 'Pb_La': 2.12,
            'Bi_La': 2.16,
            # M-lines
            'W_Ma': 1.58, 'Pt_Ma': 1.75, 'Au_Ma': 1.80, 'Pb_Ma': 1.92,
            # Actinides
            'Th_La': 2.30, 'U_La': 2.38,
        }
    },
    'Hitachi': {
        'description': 'Hitachi SEM with EDX',
        'beam_energy': 15.0,
        'kfactors': {
            # Light elements
            'B_Ka': 5.52, 'C_Ka': 3.90, 'N_Ka': 3.12, 'O_Ka': 2.04,
            'F_Ka': 1.56,
            # Period 3
            'Na_Ka': 0.98, 'Mg_Ka': 0.86, 'Al_Ka': 0.88, 'Si_Ka': 1.00,
            'P_Ka': 1.10, 'S_Ka': 1.08, 'Cl_Ka': 1.12,
            # Period 4
            'K_Ka': 1.07, 'Ca_Ka': 1.09, 'Ti_Ka': 1.01, 'V_Ka': 1.00,
            'Cr_Ka': 1.02, 'Mn_Ka': 1.04, 'Fe_Ka': 1.05, 'Co_Ka': 1.08,
            'Ni_Ka': 1.12, 'Cu_Ka': 1.18, 'Zn_Ka': 1.28, 'Ga_Ka': 1.40,
            'Ge_Ka': 1.55, 'As_Ka': 1.72, 'Se_Ka': 1.92, 'Br_Ka': 2.15,
            # Period 5
            'Rb_Ka': 2.68, 'Sr_Ka': 2.98, 'Y_Ka': 3.30, 'Zr_Ka': 3.65,
            'Nb_Ka': 4.02, 'Mo_Ka': 3.28, 'Ag_Ka': 6.28,
            # L-lines
            'Ag_La': 2.40, 'Sn_La': 2.10, 'Sb_La': 2.00, 'Ba_La': 1.85,
            'La_La': 1.75, 'W_La': 2.15, 'Pt_La': 2.45, 'Au_La': 2.55,
            'Pb_La': 2.35, 'Bi_La': 2.40,
        }
    },
}

# Quantification methods available
EDX_QUANT_METHODS = {
    'CL': 'Cliff-Lorimer (k-factors)',
    'normalized': 'Normalized Intensity (no k-factors)',
    'ZAF': 'ZAF Correction (SEM only)'
}


class EDXQuantificationSettings:
    """Stores and manages EDX quantification settings"""

    def __init__(self):
        self.instrument = 'Generic SEM'
        self.method = 'CL'  # Cliff-Lorimer
        self.beam_energy = 15.0  # keV
        self.custom_kfactors = {}
        self.composition_units = 'atomic'  # 'atomic' or 'weight'
        self.use_background_subtraction = True
        self.background_window_width = 1.0  # keV on each side

    def get_kfactor(self, xray_line):
        """Get k-factor for an X-ray line"""
        # Check custom k-factors first
        if xray_line in self.custom_kfactors:
            return self.custom_kfactors[xray_line]

        # Get from instrument database
        if self.instrument in EDX_KFACTORS_DATABASE:
            kfactors = EDX_KFACTORS_DATABASE[self.instrument]['kfactors']
            if xray_line in kfactors:
                return kfactors[xray_line]

        # Default to 1.0 if not found
        return 1.0

    def get_all_kfactors_for_elements(self, elements):
        """Get k-factors for a list of elements, determining best X-ray line"""
        kfactors = []
        lines = []

        for element in elements:
            # Try Ka, then La, then Ma
            for line_type in ['Ka', 'La', 'Ma']:
                xray_line = f"{element}_{line_type}"
                kf = self.get_kfactor(xray_line)
                if kf != 1.0 or line_type == 'Ma':  # Use value if found or default on last try
                    kfactors.append(kf)
                    lines.append(xray_line)
                    break

        return kfactors, lines

    def to_dict(self):
        """Convert settings to dictionary for storage"""
        return {
            'instrument': self.instrument,
            'method': self.method,
            'beam_energy': self.beam_energy,
            'custom_kfactors': self.custom_kfactors.copy(),
            'composition_units': self.composition_units,
            'use_background_subtraction': self.use_background_subtraction,
            'background_window_width': self.background_window_width
        }

    def from_dict(self, data):
        """Load settings from dictionary"""
        if data:
            self.instrument = data.get('instrument', 'Generic SEM')
            self.method = data.get('method', 'CL')
            self.beam_energy = data.get('beam_energy', 15.0)
            self.custom_kfactors = data.get('custom_kfactors', {})
            self.composition_units = data.get('composition_units', 'atomic')
            self.use_background_subtraction = data.get('use_background_subtraction', True)
            self.background_window_width = data.get('background_window_width', 1.0)


class EDXSensitivityWindow(wx.Frame):
    """Popup window for EDX sensitivity, display controls, and quantification settings"""

    def __init__(self, parent):
        super().__init__(parent, title="EDX Settings & Quantification",
                         size=(400, 550),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)

        self.parent = parent

        # Initialize quantification settings
        if not hasattr(self.parent, 'quant_settings'):
            self.parent.quant_settings = EDXQuantificationSettings()

        # Load saved settings if available
        self._load_saved_settings()

        self.init_ui()
        self.Centre()

    def _load_saved_settings(self):
        """Load quantification settings from parent Data if available"""
        try:
            if (self.parent.parent is not None and
                    hasattr(self.parent.parent, 'Data') and
                    'Core levels' in self.parent.parent.Data):
                for sheet_name in self.parent.parent.Data['Core levels']:
                    if sheet_name.startswith('EDX~'):
                        sheet_data = self.parent.parent.Data['Core levels'][sheet_name]
                        if '_quant_settings' in sheet_data:
                            self.parent.quant_settings.from_dict(sheet_data['_quant_settings'])
                            break
        except Exception as e:
            print(f"Could not load saved quant settings: {e}")

    def init_ui(self):
        """Initialize the control interface with quantification options"""
        panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create notebook for tabbed interface
        notebook = wx.Notebook(panel)

        # ==================== DISPLAY TAB ====================
        display_panel = wx.Panel(notebook)
        display_sizer = wx.BoxSizer(wx.VERTICAL)
        display_sizer.AddSpacer(10)

        # X Max control
        x_max_box = wx.BoxSizer(wx.HORIZONTAL)
        x_max_label = wx.StaticText(display_panel, label="X Max (keV):")
        x_max_box.Add(x_max_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)

        self.x_max_spin = wx.SpinCtrl(display_panel, value="20", min=1, max=50000, size=(80, -1))
        self.x_max_spin.SetToolTip("Set maximum X-axis energy (0 to X)")
        self.x_max_spin.Bind(wx.EVT_SPINCTRL, self.on_x_max_change)
        x_max_box.Add(self.x_max_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        display_sizer.Add(x_max_box, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 10)

        # Y Max control
        y_max_box = wx.BoxSizer(wx.HORIZONTAL)
        y_max_label = wx.StaticText(display_panel, label="Y Max (counts):")
        y_max_box.Add(y_max_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)

        self.y_max_spin = wx.SpinCtrlDouble(display_panel, value="10000", min=100, max=1000000,
                                            inc=1000, size=(100, -1))
        self.y_max_spin.SetToolTip("Set maximum Y-axis intensity")
        self.y_max_spin.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_y_max_change)
        y_max_box.Add(self.y_max_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        display_sizer.Add(y_max_box, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 10)

        # Sensitivity slider
        sensitivity_box = wx.BoxSizer(wx.VERTICAL)
        sensitivity_label = wx.StaticText(display_panel, label="Peak Detection Sensitivity:")
        sensitivity_box.Add(sensitivity_label, 0, wx.BOTTOM, 5)

        slider_box = wx.BoxSizer(wx.HORIZONTAL)
        low_label = wx.StaticText(display_panel, label="Low")
        slider_box.Add(low_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)

        self.sensitivity_slider = wx.Slider(display_panel, value=50, minValue=1, maxValue=100,
                                            style=wx.SL_HORIZONTAL, size=(150, -1))
        self.sensitivity_slider.SetToolTip("Adjust peak detection sensitivity\n(Higher = more peaks detected)")
        self.sensitivity_slider.Bind(wx.EVT_SLIDER, self.on_sensitivity_change)
        slider_box.Add(self.sensitivity_slider, 1, wx.ALIGN_CENTER_VERTICAL)

        high_label = wx.StaticText(display_panel, label="High")
        slider_box.Add(high_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)

        sensitivity_box.Add(slider_box, 0, wx.EXPAND)

        self.sensitivity_value_label = wx.StaticText(display_panel, label="Value: 50")
        sensitivity_box.Add(self.sensitivity_value_label, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.TOP, 5)

        display_sizer.Add(sensitivity_box, 0, wx.ALL | wx.EXPAND, 10)

        display_panel.SetSizer(display_sizer)
        notebook.AddPage(display_panel, "Display")

        # ==================== QUANTIFICATION TAB ====================
        quant_panel = wx.Panel(notebook)
        quant_sizer = wx.BoxSizer(wx.VERTICAL)
        quant_sizer.AddSpacer(10)

        # Instrument selection
        inst_box = wx.StaticBox(quant_panel, label="Instrument / Detector")
        inst_sizer = wx.StaticBoxSizer(inst_box, wx.VERTICAL)

        self.instrument_choice = wx.Choice(quant_panel,
                                           choices=list(EDX_KFACTORS_DATABASE.keys()))
        # Set current selection
        current_inst = self.parent.quant_settings.instrument
        if current_inst in EDX_KFACTORS_DATABASE:
            idx = list(EDX_KFACTORS_DATABASE.keys()).index(current_inst)
            self.instrument_choice.SetSelection(idx)
        else:
            self.instrument_choice.SetSelection(0)

        self.instrument_choice.Bind(wx.EVT_CHOICE, self.on_instrument_change)
        inst_sizer.Add(self.instrument_choice, 0, wx.EXPAND | wx.ALL, 5)

        # Instrument description
        self.inst_desc_label = wx.StaticText(quant_panel, label="")
        self._update_instrument_description()
        inst_sizer.Add(self.inst_desc_label, 0, wx.ALL, 5)

        quant_sizer.Add(inst_sizer, 0, wx.EXPAND | wx.ALL, 10)

        # Quantification method
        method_box = wx.StaticBox(quant_panel, label="Quantification Method")
        method_sizer = wx.StaticBoxSizer(method_box, wx.VERTICAL)

        self.method_choice = wx.Choice(quant_panel,
                                       choices=list(EDX_QUANT_METHODS.values()))
        # Set current selection
        current_method = self.parent.quant_settings.method
        method_keys = list(EDX_QUANT_METHODS.keys())
        if current_method in method_keys:
            idx = method_keys.index(current_method)
            self.method_choice.SetSelection(idx)
        else:
            self.method_choice.SetSelection(0)

        self.method_choice.Bind(wx.EVT_CHOICE, self.on_method_change)
        method_sizer.Add(self.method_choice, 0, wx.EXPAND | wx.ALL, 5)

        quant_sizer.Add(method_sizer, 0, wx.EXPAND | wx.ALL, 10)

        # Composition units
        units_box = wx.StaticBox(quant_panel, label="Composition Units")
        units_sizer = wx.StaticBoxSizer(units_box, wx.HORIZONTAL)

        self.atomic_radio = wx.RadioButton(quant_panel, label="Atomic %", style=wx.RB_GROUP)
        self.weight_radio = wx.RadioButton(quant_panel, label="Weight %")

        if self.parent.quant_settings.composition_units == 'weight':
            self.weight_radio.SetValue(True)
        else:
            self.atomic_radio.SetValue(True)

        self.atomic_radio.Bind(wx.EVT_RADIOBUTTON, self.on_units_change)
        self.weight_radio.Bind(wx.EVT_RADIOBUTTON, self.on_units_change)

        units_sizer.Add(self.atomic_radio, 0, wx.ALL, 5)
        units_sizer.Add(self.weight_radio, 0, wx.ALL, 5)

        quant_sizer.Add(units_sizer, 0, wx.EXPAND | wx.ALL, 10)

        # Background subtraction option
        self.bg_subtract_cb = wx.CheckBox(quant_panel, label="Background Subtraction")
        self.bg_subtract_cb.SetValue(self.parent.quant_settings.use_background_subtraction)
        self.bg_subtract_cb.Bind(wx.EVT_CHECKBOX, self.on_bg_subtract_change)
        quant_sizer.Add(self.bg_subtract_cb, 0, wx.ALL, 10)

        # Beam energy
        beam_box = wx.BoxSizer(wx.HORIZONTAL)
        beam_label = wx.StaticText(quant_panel, label="Beam Energy (keV):")
        beam_box.Add(beam_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)

        self.beam_energy_spin = wx.SpinCtrlDouble(quant_panel,
                                                  value=str(self.parent.quant_settings.beam_energy),
                                                  min=0.5, max=300.0, inc=0.5, size=(80, -1))
        self.beam_energy_spin.Bind(wx.EVT_SPINCTRLDOUBLE, self.on_beam_energy_change)
        beam_box.Add(self.beam_energy_spin, 0, wx.ALIGN_CENTER_VERTICAL)

        quant_sizer.Add(beam_box, 0, wx.ALL, 10)

        quant_panel.SetSizer(quant_sizer)
        notebook.AddPage(quant_panel, "Quantification")

        # ==================== K-FACTORS TAB ====================
        kfactor_panel = wx.Panel(notebook)
        kfactor_sizer = wx.BoxSizer(wx.VERTICAL)
        kfactor_sizer.AddSpacer(10)

        # K-factor grid
        kf_label = wx.StaticText(kfactor_panel, label="K-factors for selected instrument:")
        kfactor_sizer.Add(kf_label, 0, wx.ALL, 5)

        self.kfactor_list = wx.ListCtrl(kfactor_panel, style=wx.LC_REPORT | wx.LC_SINGLE_SEL,
                                        size=(-1, 200))
        self.kfactor_list.InsertColumn(0, "X-ray Line", width=100)
        self.kfactor_list.InsertColumn(1, "K-factor", width=80)
        self.kfactor_list.InsertColumn(2, "Source", width=80)

        self._populate_kfactor_list()

        kfactor_sizer.Add(self.kfactor_list, 1, wx.EXPAND | wx.ALL, 5)

        # Edit k-factor button
        edit_btn = wx.Button(kfactor_panel, label="Edit Selected K-factor...")
        edit_btn.Bind(wx.EVT_BUTTON, self.on_edit_kfactor)
        kfactor_sizer.Add(edit_btn, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)

        kfactor_panel.SetSizer(kfactor_sizer)
        notebook.AddPage(kfactor_panel, "K-factors")

        main_sizer.Add(notebook, 1, wx.EXPAND | wx.ALL, 5)

        # ==================== BUTTONS ====================
        button_box = wx.BoxSizer(wx.HORIZONTAL)

        reset_btn = wx.Button(panel, label="Reset to Default")
        reset_btn.Bind(wx.EVT_BUTTON, self.on_reset)
        button_box.Add(reset_btn, 0, wx.RIGHT, 5)

        apply_btn = wx.Button(panel, label="Apply && Recalculate")
        apply_btn.Bind(wx.EVT_BUTTON, self.on_apply)
        button_box.Add(apply_btn, 0)

        main_sizer.Add(button_box, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 10)

        panel.SetSizer(main_sizer)

        # Initialize values from current plot
        self.initialize_from_plot()

    def _update_instrument_description(self):
        """Update the instrument description label"""
        inst = self.instrument_choice.GetStringSelection()
        if inst in EDX_KFACTORS_DATABASE:
            desc = EDX_KFACTORS_DATABASE[inst]['description']
            beam = EDX_KFACTORS_DATABASE[inst]['beam_energy']
            self.inst_desc_label.SetLabel(f"{desc}\nDefault beam: {beam:.1f} keV")

    def _populate_kfactor_list(self):
        """Populate the k-factor list control"""
        self.kfactor_list.DeleteAllItems()

        inst = self.parent.quant_settings.instrument
        if inst in EDX_KFACTORS_DATABASE:
            kfactors = EDX_KFACTORS_DATABASE[inst]['kfactors']

            # Sort by element (periodic table order approximation)
            sorted_lines = sorted(kfactors.keys(),
                                  key=lambda x: (len(x.split('_')[0]), x))

            for line in sorted_lines:
                kf = kfactors[line]
                source = "Instrument"

                # Check if custom override exists
                if line in self.parent.quant_settings.custom_kfactors:
                    kf = self.parent.quant_settings.custom_kfactors[line]
                    source = "Custom"

                idx = self.kfactor_list.InsertItem(self.kfactor_list.GetItemCount(), line)
                self.kfactor_list.SetItem(idx, 1, f"{kf:.4f}")
                self.kfactor_list.SetItem(idx, 2, source)

    def on_instrument_change(self, event):
        """Handle instrument selection change"""
        inst = self.instrument_choice.GetStringSelection()
        self.parent.quant_settings.instrument = inst
        self._update_instrument_description()
        self._populate_kfactor_list()

        # Update beam energy to instrument default
        if inst in EDX_KFACTORS_DATABASE:
            beam = EDX_KFACTORS_DATABASE[inst]['beam_energy']
            self.beam_energy_spin.SetValue(beam)
            self.parent.quant_settings.beam_energy = beam

    def on_method_change(self, event):
        """Handle quantification method change"""
        method_idx = self.method_choice.GetSelection()
        method_keys = list(EDX_QUANT_METHODS.keys())
        self.parent.quant_settings.method = method_keys[method_idx]

    def on_units_change(self, event):
        """Handle composition units change"""
        if self.atomic_radio.GetValue():
            self.parent.quant_settings.composition_units = 'atomic'
        else:
            self.parent.quant_settings.composition_units = 'weight'

    def on_bg_subtract_change(self, event):
        """Handle background subtraction checkbox"""
        self.parent.quant_settings.use_background_subtraction = self.bg_subtract_cb.GetValue()

    def on_beam_energy_change(self, event):
        """Handle beam energy change"""
        self.parent.quant_settings.beam_energy = self.beam_energy_spin.GetValue()

    def on_edit_kfactor(self, event):
        """Edit a custom k-factor"""
        selected = self.kfactor_list.GetFirstSelected()
        if selected == -1:
            wx.MessageBox("Please select a k-factor to edit.", "No Selection",
                          wx.OK | wx.ICON_INFORMATION)
            return

        line = self.kfactor_list.GetItemText(selected, 0)
        current_kf = float(self.kfactor_list.GetItemText(selected, 1))

        dlg = wx.TextEntryDialog(self, f"Enter k-factor for {line}:",
                                 "Edit K-factor", str(current_kf))

        if dlg.ShowModal() == wx.ID_OK:
            try:
                new_kf = float(dlg.GetValue())
                self.parent.quant_settings.custom_kfactors[line] = new_kf
                self._populate_kfactor_list()
            except ValueError:
                wx.MessageBox("Invalid number entered.", "Error", wx.OK | wx.ICON_ERROR)

        dlg.Destroy()

    def initialize_from_plot(self):
        """Set initial values from current plot state"""
        self.x_max_spin.SetValue(20)
        self.y_max_spin.SetValue(10000)

        if self.parent.parent is not None and hasattr(self.parent.parent, 'ax'):
            try:
                xlim = self.parent.parent.ax.get_xlim()
                if xlim[1] > 0:
                    self.x_max_spin.SetValue(int(xlim[1]))

                ylim = self.parent.parent.ax.get_ylim()
                if ylim[1] > 0:
                    self.y_max_spin.SetValue(ylim[1])
            except:
                pass

    def on_x_max_change(self, event):
        """Handle X max spin control change"""
        if self.parent.parent is not None and hasattr(self.parent.parent, 'ax'):
            x_max = self.x_max_spin.GetValue()
            self.parent.parent.ax.set_xlim(0, x_max)

            if 'Core levels' in self.parent.parent.Data:
                for sheet_name in self.parent.parent.Data['Core levels']:
                    if sheet_name == 'EDX~Plot' or sheet_name.startswith('EDX~Plot'):
                        self.parent.parent.Data['Core levels'][sheet_name]['_EDX_display_max'] = x_max
                print(f"Stored _EDX_display_max={x_max:.2f} for all EDX~Plot sheets")

            self.parent.parent.canvas.draw()

    def on_y_max_change(self, event):
        """Handle Y max spin control change"""
        if self.parent.parent is not None and hasattr(self.parent.parent, 'ax'):
            y_max = self.y_max_spin.GetValue()
            self.parent.parent.ax.set_ylim(0, y_max)
            self.parent.parent.canvas.draw()

    def on_sensitivity_change(self, event):
        """Handle sensitivity slider change"""
        sensitivity = self.sensitivity_slider.GetValue()
        self.sensitivity_value_label.SetLabel(f"Value: {sensitivity}")

    def on_apply(self, event):
        """Apply all settings and recalculate quantification"""
        # Save settings to Data structure
        self._save_settings()

        if self.parent.parent is not None and hasattr(self.parent.parent, 'sheet_combobox'):
            sensitivity = self.sensitivity_slider.GetValue()
            threshold = 0.11 - (sensitivity / 1000.0)

            print(f"Applying settings - Instrument: {self.parent.quant_settings.instrument}, "
                  f"Method: {self.parent.quant_settings.method}")

            self.parent.peak_threshold = threshold

            # Clear current plot
            self.parent.parent.ax.clear()

            # Re-plot with new settings
            current_sheet = self.parent.parent.sheet_combobox.GetValue()
            if current_sheet == 'EDX~Plot' or current_sheet.startswith('EDX~Plot'):
                if 'Core levels' in self.parent.parent.Data and current_sheet in self.parent.parent.Data['Core levels']:
                    sheet_data = self.parent.parent.Data['Core levels'][current_sheet]
                    sheet_data['_EDX_sensitivity'] = threshold
                    sheet_data['_quant_settings'] = self.parent.quant_settings.to_dict()

                # Recalculate quantification with new settings
                if hasattr(self.parent, 'current_data') and self.parent.current_data is not None:
                    energy = self.parent.get_energy_axis()
                    if energy is not None:
                        spectrum = self.parent.current_data.sum().data if hasattr(self.parent.current_data, 'sum') else None
                        if spectrum is not None:
                            self.parent._calculate_quantification_hyperspy(energy, spectrum)

                # Refresh the plot using the main window's plot manager
                if hasattr(self.parent.parent, 'plot_manager') and self.parent.parent.plot_manager:
                    self.parent.parent.plot_manager.plot_edx_data(self.parent.parent, current_sheet)
                elif hasattr(self.parent.parent, 'PlotManager') and self.parent.parent.PlotManager:
                    self.parent.parent.PlotManager.plot_edx_data(self.parent.parent, current_sheet)
                else:
                    # Fallback: just redraw the canvas
                    self.parent.parent.canvas.draw_idle()

    def _save_settings(self):
        """Save quantification settings to Data structure"""
        try:
            if (self.parent.parent is not None and
                    hasattr(self.parent.parent, 'Data') and
                    'Core levels' in self.parent.parent.Data):
                for sheet_name in self.parent.parent.Data['Core levels']:
                    if sheet_name.startswith('EDX~'):
                        sheet_data = self.parent.parent.Data['Core levels'][sheet_name]
                        sheet_data['_quant_settings'] = self.parent.quant_settings.to_dict()
                print(f"Saved quantification settings to Data structure")
        except Exception as e:
            print(f"Could not save quant settings: {e}")

    def on_reset(self, event):
        """Reset all values to defaults"""
        self.x_max_spin.SetValue(20)
        self.y_max_spin.SetValue(10000)
        self.sensitivity_slider.SetValue(50)
        self.sensitivity_value_label.SetLabel("Value: 50")

        # Reset quantification settings
        self.parent.quant_settings = EDXQuantificationSettings()
        self.instrument_choice.SetSelection(0)
        self.method_choice.SetSelection(0)
        self.atomic_radio.SetValue(True)
        self.bg_subtract_cb.SetValue(True)
        self.beam_energy_spin.SetValue(15.0)
        self._update_instrument_description()
        self._populate_kfactor_list()

        if self.parent.parent is not None and hasattr(self.parent.parent, 'ax'):
            self.parent.parent.ax.set_xlim(0, 20)
            self.parent.parent.ax.set_ylim(0, 10000)
            self.parent.parent.canvas.draw()


class RotatableRectangle:
    """Custom rotatable rectangle selector for area selection"""

    def __init__(self, ax, callback):
        self.ax = ax
        self.callback = callback
        self.active = False
        self.angle = 0

        self.center = None
        self.width = None
        self.height = None

        self.rectangle = None
        self.rotation_handle = None
        self.resize_handles = []

        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None
        self.resize_corner = None
        self.angle_text = None

        self._stored_xlim = None
        self._stored_ylim = None

        self.cid_press = ax.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.cid_release = ax.figure.canvas.mpl_connect('button_release_event', self.on_release)
        self.cid_motion = ax.figure.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.cid_scroll = ax.figure.canvas.mpl_connect('scroll_event', self.on_scroll)

    def start_selection(self, event):
        """Start creating a new rectangle"""
        if event.inaxes != self.ax or event.button != 1:
            return

        # If not active, create new rectangle at click position
        if not self.active:
            # Store axis limits BEFORE any drawing
            self._stored_xlim = self.ax.get_xlim()
            self._stored_ylim = self.ax.get_ylim()

            self.center = (event.xdata, event.ydata)
            self.width = 10
            self.height = 10
            self.angle = 0
            self.active = True
            self.draw_rectangle()

    def draw_rectangle(self):
        """Draw the rotatable rectangle and handles"""
        # Clear existing elements safely
        if self.rectangle:
            try:
                self.rectangle.remove()
            except (ValueError, AttributeError):
                pass
            self.rectangle = None

        for handle in self.resize_handles:
            try:
                handle.remove()
            except (ValueError, AttributeError):
                pass
        self.resize_handles = []

        if self.rotation_handle:
            try:
                self.rotation_handle.remove()
            except (ValueError, AttributeError):
                pass
            self.rotation_handle = None

        if self.angle_text:
            try:
                self.angle_text.remove()
            except (ValueError, AttributeError):
                pass
            self.angle_text = None

        if self.center is None:
            self.ax.figure.canvas.draw_idle()
            return

        from matplotlib.patches import Rectangle
        from matplotlib.transforms import Affine2D

        rect = Rectangle((-self.width / 2, -self.height / 2), self.width, self.height,
                         linewidth=2, edgecolor='red', facecolor='red', alpha=0.3)

        t = Affine2D().rotate_deg(self.angle).translate(*self.center) + self.ax.transData
        rect.set_transform(t)
        rect._is_selection_marker = True

        self.rectangle = self.ax.add_patch(rect)

        # Rotation handle (red)
        handle_dist = max(self.width, self.height) / 2 + 10
        angle_rad = np.radians(self.angle)
        handle_x = self.center[0] + handle_dist * np.sin(angle_rad)
        handle_y = self.center[1] + handle_dist * np.cos(angle_rad)

        self.rotation_handle = self.ax.plot(handle_x, handle_y, 'ro', markersize=10,
                                            markeredgecolor='darkred', markeredgewidth=2)[0]
        self.rotation_handle._is_selection_marker = True

        # Show angle text when rotating
        if self.rotating:
            self.angle_text = self.ax.text(handle_x, handle_y + 5, f'{self.angle:.1f}°',
                                           fontsize=9, color='white', fontweight='bold',
                                           ha='center', va='bottom',
                                           bbox=dict(boxstyle='round,pad=0.2', facecolor='darkred', alpha=0.7))
            self.angle_text._is_selection_marker = True

        # Resize handles at corners (red)
        corners = self.get_corners()
        for corner in corners:
            handle = self.ax.plot(corner[0], corner[1], 'rs', markersize=8,
                                  markeredgecolor='darkred', markeredgewidth=1)[0]
            handle._is_selection_marker = True
            self.resize_handles.append(handle)

        # ALWAYS restore stored axis limits to prevent expansion
        if self._stored_xlim is not None and self._stored_ylim is not None:
            self.ax.set_xlim(self._stored_xlim)
            self.ax.set_ylim(self._stored_ylim)

        self.ax.figure.canvas.draw_idle()

    def get_corners(self):
        """Get the four corners of the rotated rectangle"""
        angle_rad = np.radians(self.angle)
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        hw = self.width / 2
        hh = self.height / 2
        corners_local = [(-hw, -hh), (hw, -hh), (hw, hh), (-hw, hh)]

        corners = []
        for x, y in corners_local:
            rx = x * cos_a - y * sin_a + self.center[0]
            ry = x * sin_a + y * cos_a + self.center[1]
            corners.append((rx, ry))

        return corners

    def on_scroll(self, event):
        """Handle mouse scroll to rotate rectangle"""
        if not self.active or event.inaxes != self.ax:
            return

        # Check if mouse is near the rectangle
        if self.center is None:
            return

        # Rotate by 5 degrees per scroll step
        if event.button == 'up':
            self.angle += 5
        elif event.button == 'down':
            self.angle -= 5

        # Normalize angle to -180 to 180
        self.angle = ((self.angle + 180) % 360) - 180

        self.draw_rectangle()

        # Trigger callback
        if self.callback:
            self.callback(self.center, self.width, self.height, self.angle)

    def on_press(self, event):
        """Handle mouse press"""
        if not self.active or event.inaxes != self.ax or event.button != 1:
            return

        # Store current limits
        if self._stored_xlim is None:
            self._stored_xlim = self.ax.get_xlim()
            self._stored_ylim = self.ax.get_ylim()

        # Check rotation handle first (highest priority)
        if self.rotation_handle:
            handle_data = self.rotation_handle.get_data()
            dist = np.sqrt((event.xdata - handle_data[0][0]) ** 2 + (event.ydata - handle_data[1][0]) ** 2)
            if dist < 8:
                self.rotating = True
                self.drag_start = (event.xdata, event.ydata)
                return

        # Check resize handles - but only if very close (within 5 pixels)
        closest_handle_dist = float('inf')
        closest_handle_idx = None
        for i, handle in enumerate(self.resize_handles):
            handle_data = handle.get_data()
            dist = np.sqrt((event.xdata - handle_data[0][0]) ** 2 + (event.ydata - handle_data[1][0]) ** 2)
            if dist < closest_handle_dist:
                closest_handle_dist = dist
                closest_handle_idx = i

        # Only resize if very close to handle (< 5 pixels)
        if closest_handle_dist < 5:
            self.resizing = True
            self.resize_corner = closest_handle_idx
            self.drag_start = (event.xdata, event.ydata)
            return

        # Check if inside rectangle for dragging (most common operation)
        if self.point_in_rectangle(event.xdata, event.ydata):
            self.dragging = True
            self.drag_start = (event.xdata, event.ydata)
            return

        # If not inside but close to resize handle (< 10 pixels), allow resize
        if closest_handle_dist < 10:
            self.resizing = True
            self.resize_corner = closest_handle_idx
            self.drag_start = (event.xdata, event.ydata)

    def on_release(self, event):
        """Handle mouse release"""
        was_interacting = self.dragging or self.resizing or self.rotating

        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None

        # Restore limits after any operation
        if self._stored_xlim is not None and self._stored_ylim is not None:
            self.ax.set_xlim(self._stored_xlim)
            self.ax.set_ylim(self._stored_ylim)

        # Trigger callback after interaction
        if was_interacting and self.callback and self.center:
            self.callback(self.center, self.width, self.height, self.angle)

    def on_motion(self, event):
        """Handle mouse motion"""
        if event.inaxes != self.ax or self.drag_start is None:
            return

        if event.xdata is None or event.ydata is None:
            return

        dx = event.xdata - self.drag_start[0]
        dy = event.ydata - self.drag_start[1]

        if self.dragging:
            new_x = self.center[0] + dx
            new_y = self.center[1] + dy

            # Clamp to stored bounds
            if self._stored_xlim is not None:
                new_x = max(self._stored_xlim[0], min(self._stored_xlim[1], new_x))
            if self._stored_ylim is not None:
                y_min = min(self._stored_ylim[0], self._stored_ylim[1])
                y_max = max(self._stored_ylim[0], self._stored_ylim[1])
                new_y = max(y_min, min(y_max, new_y))

            self.center = (new_x, new_y)
            self.drag_start = (event.xdata, event.ydata)
            self.draw_rectangle()

        elif self.rotating:
            angle_to_center = np.arctan2(event.xdata - self.center[0], event.ydata - self.center[1])
            self.angle = np.degrees(angle_to_center)
            self.draw_rectangle()

        elif self.resizing:
            angle_rad = np.radians(-self.angle)
            cos_a = np.cos(angle_rad)
            sin_a = np.sin(angle_rad)

            local_x = (event.xdata - self.center[0]) * cos_a - (event.ydata - self.center[1]) * sin_a
            local_y = (event.xdata - self.center[0]) * sin_a + (event.ydata - self.center[1]) * cos_a

            # Allow minimum size of 1 pixel
            self.width = max(1, abs(local_x) * 2)
            self.height = max(1, abs(local_y) * 2)
            self.draw_rectangle()

    def point_in_rectangle(self, x, y):
        """Check if point is inside the rotated rectangle"""
        if self.center is None:
            return False

        angle_rad = np.radians(-self.angle)
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        local_x = (x - self.center[0]) * cos_a - (y - self.center[1]) * sin_a
        local_y = (x - self.center[0]) * sin_a + (y - self.center[1]) * cos_a

        return abs(local_x) <= self.width / 2 and abs(local_y) <= self.height / 2

    def clear(self):
        """Clear the rectangle and handles, allow new selection"""
        if self.rectangle:
            try:
                self.rectangle.remove()
            except (ValueError, AttributeError):
                pass
            self.rectangle = None

        for handle in self.resize_handles:
            try:
                handle.remove()
            except (ValueError, AttributeError):
                pass
        self.resize_handles = []

        if self.rotation_handle:
            try:
                self.rotation_handle.remove()
            except (ValueError, AttributeError):
                pass
            self.rotation_handle = None

        if self.angle_text:
            try:
                self.angle_text.remove()
            except (ValueError, AttributeError):
                pass
            self.angle_text = None

        # Reset state to allow new selection
        self.center = None
        self.width = None
        self.height = None
        self.angle = 0
        self.active = False
        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None

        self.ax.figure.canvas.draw_idle()

    def disconnect(self):
        """Disconnect event handlers"""
        try:
            self.ax.figure.canvas.mpl_disconnect(self.cid_press)
            self.ax.figure.canvas.mpl_disconnect(self.cid_release)
            self.ax.figure.canvas.mpl_disconnect(self.cid_motion)
            self.ax.figure.canvas.mpl_disconnect(self.cid_scroll)
        except (ValueError, AttributeError):
            pass

def open_edx_sem_window(parent):
    """Open EDX/SEM analysis window"""
    window = EDXSEMWindow(parent)
    window.Show()
    return window