# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
EELS_Analysis.py
Module for Electron Energy Loss Spectroscopy (EELS) data analysis
Uses standalone EELS_Utilities for data import and analysis (no HyperSpy dependency)
Follows same structure as EDX_SEM_Analysis.py but without peak labelling/quantification
"""

import wx
import numpy as np
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.widgets import RectangleSelector
import matplotlib.patches as patches
import matplotlib.patheffects as path_effects

# EELS Utilities - standalone replacements for HyperSpy
try:
    from libraries.EELS_Utilities import (
        Signal1D,
        load_eels,
        load_hdf5_eels,
        create_eels_signal
    )
    EELS_UTILITIES_AVAILABLE = True
except ImportError:
    EELS_UTILITIES_AVAILABLE = False
    Signal1D = None

class EELSWindow(wx.Frame):
    """Main window for EELS data analysis"""

    def __init__(self, parent, title="EELS HeatMap"):
        super().__init__(parent, title=title, size=(590, 700),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)

        self.parent = parent
        self.eels_data = None
        self.current_data = None

        # Selection modes
        self.selection_mode = None  # 'point', 'area', 'line'
        self.rect_selector = None
        self.selected_points = []
        self.selected_areas = []
        self.line_start = None
        self.line_end = None

        self.point_size = 1  # Size in pixels (1 = single pixel)

        self.loaded_signals = []
        self.current_colorbar = None
        self.zoom_selector = None

        # Windows
        self.info_window = None
        self.data_browser_window = None

        # Rotatable rectangle for area selection
        self.rotatable_rect = None

        self.init_ui()
        self.Centre()

        # Bind close event
        self.Bind(wx.EVT_CLOSE, self.on_close)

    def init_ui(self):
        """Initialize user interface"""
        panel = wx.Panel(self, style=wx.BORDER_RAISED)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Toolbar
        toolbar_panel = self.create_toolbar(panel)
        main_sizer.Add(toolbar_panel, 0, wx.EXPAND | wx.ALL, 2)

        # Main content - just the map
        map_panel = wx.Panel(panel)
        map_sizer = wx.BoxSizer(wx.VERTICAL)

        # Map display only
        self.map_figure = plt.Figure(figsize=(4, 4))
        self.map_figure.set_tight_layout(True)
        self.map_figure.subplots_adjust(left=0.1, right=0.95, top=0.95, bottom=0.1)

        self.map_canvas = FigureCanvas(map_panel, -1, self.map_figure)
        self.map_ax = self.map_figure.add_subplot(111)

        map_sizer.Add(self.map_canvas, 1, wx.EXPAND)

        # ========== Arrow control panel (like EDX) ==========
        arrow_panel = wx.Panel(map_panel)
        arrow_panel.SetBackgroundColour(wx.Colour(240, 240, 240, 200))
        arrow_sizer = wx.GridBagSizer(0, 0)

        btn_size = (30, 30)
        btn_size_wide = (60, 30)

        # Row 0: Up button spanning 2 columns
        self.arrow_up_btn = wx.Button(arrow_panel, label="↑", size=btn_size_wide)
        self.arrow_up_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_up_btn, pos=(0, 0), span=(1, 2), flag=wx.ALIGN_CENTER)

        # Row 1: Left and Right buttons
        self.arrow_left_btn = wx.Button(arrow_panel, label="←", size=btn_size)
        self.arrow_left_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_left_btn, pos=(1, 0), flag=wx.ALIGN_CENTER)

        self.arrow_right_btn = wx.Button(arrow_panel, label="→", size=btn_size)
        self.arrow_right_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_right_btn, pos=(1, 1), flag=wx.ALIGN_CENTER)

        # Row 2: Down button spanning 2 columns
        self.arrow_down_btn = wx.Button(arrow_panel, label="↓", size=btn_size_wide)
        self.arrow_down_btn.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        arrow_sizer.Add(self.arrow_down_btn, pos=(2, 0), span=(1, 2), flag=wx.ALIGN_CENTER)

        arrow_panel.SetSizer(arrow_sizer)
        arrow_panel.Fit()

        # Store reference for positioning later
        self.arrow_control_panel = arrow_panel

        # Set map panel sizer
        map_panel.SetSizer(map_sizer)

        # Bind arrow button events
        self.arrow_up_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('up'))
        self.arrow_down_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('down'))
        self.arrow_left_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('left'))
        self.arrow_right_btn.Bind(wx.EVT_BUTTON, lambda e: self.on_arrow_move('right'))

        # Bind size event to reposition arrow panel
        map_panel.Bind(wx.EVT_SIZE, self.on_map_panel_resize)

        # Bind paint event to keep buttons visible
        self.map_canvas.Bind(wx.EVT_PAINT, self.on_canvas_paint)

        map_panel.SetSizer(map_sizer)

        main_sizer.Add(map_panel, 1, wx.EXPAND)

        panel.SetSizer(main_sizer)
        main_sizer.Layout()
        panel.Layout()

        # Bind canvas events
        self.map_canvas.mpl_connect('button_press_event', self.on_map_click)
        self.map_canvas.mpl_connect('motion_notify_event', self.on_map_motion)
        self.map_canvas.mpl_connect('scroll_event', self.on_map_scroll)
        self._line_preview = None

        # Right-click menu
        self.map_canvas.Bind(wx.EVT_RIGHT_DOWN, self.on_right_click)

        # Store colorbar reference
        self.current_colorbar = None
        self.current_cmap = 'plasma'

        wx.CallAfter(self.Layout)

    def create_toolbar(self, parent):
        """Create toolbar with icon buttons only"""
        toolbar_panel = wx.Panel(parent)
        toolbar_sizer = wx.BoxSizer(wx.HORIZONTAL)

        btn_size = (25, 25)

        # Get icon path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(current_dir, "Icons")

        # Navigation tools
        self.zoom_in_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.zoom_out_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        self.pan_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)

        # Load zoom icons with fallback
        zoom_in_path = os.path.join(icon_path, "ZoomIN-3.png")
        if os.path.exists(zoom_in_path):
            self.zoom_in_btn.SetBitmap(wx.Bitmap(zoom_in_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_in_btn.SetBitmap(self.create_icon_bitmap('zoom_in'))

        zoom_out_path = os.path.join(icon_path, "ZoomOUT-3.png")
        if os.path.exists(zoom_out_path):
            self.zoom_out_btn.SetBitmap(wx.Bitmap(zoom_out_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_out_btn.SetBitmap(self.create_icon_bitmap('zoom_out'))

        pan_path = os.path.join(icon_path, "Drag-25.png")
        if os.path.exists(pan_path):
            self.pan_btn.SetBitmap(wx.Bitmap(pan_path, wx.BITMAP_TYPE_PNG))
        else:
            self.pan_btn.SetBitmap(self.create_icon_bitmap('pan'))

        self.zoom_in_btn.SetToolTip("Zoom In")
        self.zoom_out_btn.SetToolTip("Zoom Out / Home")
        self.pan_btn.SetToolTip("Pan")

        toolbar_sizer.Add(self.zoom_in_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.zoom_out_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.pan_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL),
                          0, wx.EXPAND | wx.ALL, 3)

        # Selection tools
        self.point_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.area_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.line_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.clear_btn = wx.BitmapButton(toolbar_panel, size=btn_size)

        self.point_btn.SetBitmap(self.create_icon_bitmap('point'))
        self.area_btn.SetBitmap(self.create_icon_bitmap('area'))
        self.line_btn.SetBitmap(self.create_icon_bitmap('line'))
        self.clear_btn.SetBitmap(self.create_icon_bitmap('clear'))

        self.point_btn.SetToolTip("Point Selection")
        self.area_btn.SetToolTip("Area Selection")
        self.line_btn.SetToolTip("Line Selection")
        self.clear_btn.SetToolTip("Clear Selections")

        toolbar_sizer.Add(self.point_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.area_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.line_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.clear_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL),
                          0, wx.EXPAND | wx.ALL, 3)

        # Intensity Map button
        self.intensity_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        heatmap_path = os.path.join(icon_path, "heatmap-3.png")
        if os.path.exists(heatmap_path):
            self.intensity_btn.SetBitmap(wx.Bitmap(heatmap_path, wx.BITMAP_TYPE_PNG))
        else:
            self.intensity_btn.SetBitmap(self.create_icon_bitmap('heatmap'))
        self.intensity_btn.SetToolTip("Show Intensity Map")
        toolbar_sizer.Add(self.intensity_btn, 0, wx.ALL, 2)

        # Add spacer
        toolbar_sizer.AddStretchSpacer()

        # Sensitivity/Display controls button - use Settings PNG icon
        self.sensitivity_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        settings_path = os.path.join(icon_path, "Settings-3.png")
        if os.path.exists(settings_path):
            self.sensitivity_btn.SetBitmap(wx.Bitmap(settings_path, wx.BITMAP_TYPE_PNG))
        else:
            self.sensitivity_btn.SetBitmap(self.create_icon_bitmap('sensitivity'))
        self.sensitivity_btn.SetToolTip("Display & Sensitivity Controls")
        self.sensitivity_btn.Bind(wx.EVT_BUTTON, self.on_sensitivity_controls)
        toolbar_sizer.Add(self.sensitivity_btn, 0, wx.ALL, 2)

        toolbar_panel.SetSizer(toolbar_sizer)

        # Bind events
        self.zoom_in_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_zoom_in)
        self.zoom_out_btn.Bind(wx.EVT_BUTTON, self.on_zoom_out)
        self.pan_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_pan)
        self.point_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_point_mode)
        self.area_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_area_mode)
        self.line_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_line_mode)
        self.clear_btn.Bind(wx.EVT_BUTTON, self.on_clear_selections)
        self.intensity_btn.Bind(wx.EVT_BUTTON, self.on_intensity_map)

        return toolbar_panel

    def create_icon_bitmap(self, icon_type):
        """Create simple icon bitmaps"""
        size = 20
        bmp = wx.Bitmap(size, size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.Brush(wx.Colour(240, 240, 240)))
        dc.Clear()

        dc.SetPen(wx.Pen(wx.Colour(50, 50, 50), 2))
        dc.SetBrush(wx.Brush(wx.Colour(100, 100, 100)))

        if icon_type == 'zoom_in':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(8, 10, 12, 10)
            dc.DrawLine(10, 8, 10, 12)
            dc.DrawLine(14, 14, 18, 18)
        elif icon_type == 'zoom_out':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(8, 10, 12, 10)
            dc.DrawLine(14, 14, 18, 18)
        elif icon_type == 'pan':
            dc.DrawLine(10, 3, 10, 17)
            dc.DrawLine(3, 10, 17, 10)
            dc.DrawLine(10, 3, 7, 6)
            dc.DrawLine(10, 3, 13, 6)
            dc.DrawLine(10, 17, 7, 14)
            dc.DrawLine(10, 17, 13, 14)
        elif icon_type == 'point':
            dc.SetBrush(wx.Brush(wx.Colour(79, 190, 159)))
            dc.DrawCircle(10, 10, 4)
        elif icon_type == 'area':
            dc.SetBrush(wx.TRANSPARENT_BRUSH)
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            dc.DrawRectangle(3, 3, 14, 14)
        elif icon_type == 'line':
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            dc.DrawLine(3, 17, 17, 3)
        elif icon_type == 'clear':
            dc.SetPen(wx.Pen(wx.Colour(200, 50, 50), 2))
            dc.DrawLine(5, 5, 15, 15)
            dc.DrawLine(15, 5, 5, 15)
        elif icon_type == 'heatmap':
            dc.SetPen(wx.Pen(wx.Colour(50, 50, 50), 1))
            dc.SetBrush(wx.Brush(wx.Colour(0, 0, 255)))
            dc.DrawRectangle(2, 2, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(0, 255, 0)))
            dc.DrawRectangle(11, 2, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(255, 255, 0)))
            dc.DrawRectangle(2, 11, 7, 7)
            dc.SetBrush(wx.Brush(wx.Colour(255, 0, 0)))
            dc.DrawRectangle(11, 11, 7, 7)
        elif icon_type == 'sensitivity':
            dc.SetPen(wx.Pen(wx.Colour(79, 190, 159), 2))
            dc.DrawLine(5, 10, 15, 10)
            dc.SetBrush(wx.Brush(wx.Colour(79, 190, 159)))
            dc.DrawCircle(10, 10, 3)
            dc.DrawLine(10, 4, 10, 7)
            dc.DrawLine(10, 13, 10, 16)

        dc.SelectObject(wx.NullBitmap)
        return bmp

    # ==================== TOOLBAR HANDLERS ====================

    def on_sensitivity_controls(self, event):
        """Open sensitivity and display controls window"""
        if hasattr(self, 'sensitivity_window') and self.sensitivity_window:
            if not self.sensitivity_window.IsBeingDeleted():
                self.sensitivity_window.Raise()
                return
        self.sensitivity_window = EELSSensitivityWindow(self)
        self.sensitivity_window.Show()

    def on_zoom_in(self, event):
        """Toggle zoom in mode using rectangle selector"""
        if self.zoom_in_btn.GetValue():
            self.pan_btn.SetValue(False)
            self.deactivate_selection_modes()

            if not hasattr(self, 'zoom_selector') or self.zoom_selector is None:
                self.zoom_selector = RectangleSelector(
                    self.map_ax,
                    self.on_zoom_select,
                    useblit=True,
                    props=dict(facecolor='blue', edgecolor='blue', alpha=0.2, fill=True),
                    button=[1],
                    minspanx=5,
                    minspany=5,
                    spancoords='pixels',
                    interactive=False
                )
            else:
                self.zoom_selector.set_active(True)
        else:
            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

    def on_zoom_select(self, eclick, erelease):
        """Handle zoom rectangle selection"""
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        x_min, x_max = min(x1, x2), max(x1, x2)
        y_min, y_max = min(y1, y2), max(y1, y2)

        self.map_ax.set_xlim(x_min, x_max)
        self.map_ax.set_ylim(y_max, y_min)  # Inverted for image
        self.map_canvas.draw()

        self.zoom_in_btn.SetValue(False)
        if hasattr(self, 'zoom_selector') and self.zoom_selector:
            self.zoom_selector.set_active(False)

    def on_zoom_out(self, event):
        """Zoom out / reset view"""
        self.map_ax.autoscale()
        self.map_canvas.draw()

        self.zoom_in_btn.SetValue(False)
        self.pan_btn.SetValue(False)
        if hasattr(self, 'zoom_selector') and self.zoom_selector:
            self.zoom_selector.set_active(False)

    def on_pan(self, event):
        """Toggle pan mode"""
        if self.pan_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.deactivate_selection_modes()

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

            self._pan_start = None
            self._pan_cid_press = self.map_canvas.mpl_connect(
                'button_press_event', self._on_pan_press)
            self._pan_cid_release = self.map_canvas.mpl_connect(
                'button_release_event', self._on_pan_release)
            self._pan_cid_motion = self.map_canvas.mpl_connect(
                'motion_notify_event', self._on_pan_motion)
        else:
            if hasattr(self, '_pan_cid_press'):
                self.map_canvas.mpl_disconnect(self._pan_cid_press)
            if hasattr(self, '_pan_cid_release'):
                self.map_canvas.mpl_disconnect(self._pan_cid_release)
            if hasattr(self, '_pan_cid_motion'):
                self.map_canvas.mpl_disconnect(self._pan_cid_motion)

    def _on_pan_press(self, event):
        if event.inaxes == self.map_ax and event.button == 1:
            self._pan_start = (event.xdata, event.ydata)

    def _on_pan_release(self, event):
        self._pan_start = None
        self._add_scale_bar()
        self.map_canvas.draw_idle()

    def _on_pan_motion(self, event):
        if self._pan_start is None or event.inaxes != self.map_ax:
            return

        dx = self._pan_start[0] - event.xdata
        dy = self._pan_start[1] - event.ydata

        xlim = self.map_ax.get_xlim()
        ylim = self.map_ax.get_ylim()

        self.map_ax.set_xlim(xlim[0] + dx, xlim[1] + dx)
        self.map_ax.set_ylim(ylim[0] + dy, ylim[1] + dy)

        self.map_canvas.draw_idle()

    def deactivate_selection_modes(self):
        """Deactivate all selection modes"""
        self.point_btn.SetValue(False)
        self.area_btn.SetValue(False)
        self.line_btn.SetValue(False)
        self.selection_mode = None

        if self.rect_selector:
            self.rect_selector.set_active(False)

    def on_point_mode(self, event):
        """Toggle point selection mode"""
        if self.point_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.area_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'point'

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)
            if self.rect_selector:
                self.rect_selector.set_active(False)

            self.clear_selection_markers()
            self.map_canvas.draw()
        else:
            self.selection_mode = None

    def on_area_mode(self, event):
        """Toggle area selection mode with rotatable rectangle"""
        if self.area_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.point_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'area'

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)

            # Deactivate old rectangle selector
            if self.rect_selector:
                self.rect_selector.set_active(False)

            # Create or activate rotatable rectangle
            if not hasattr(self, 'rotatable_rect') or self.rotatable_rect is None:
                self.rotatable_rect = RotatableRectangle(self.map_ax, self.on_rotated_area_complete)

            self.clear_selection_markers()

            # Connect to canvas for starting new selections
            if not hasattr(self, '_area_press_cid'):
                self._area_press_cid = self.map_canvas.mpl_connect('button_press_event',
                                                                   lambda e: self.rotatable_rect.start_selection(e) if self.area_btn.GetValue() and e.button == 1 else None)
        else:
            self.selection_mode = None
            if hasattr(self, 'rotatable_rect') and self.rotatable_rect:
                self.rotatable_rect.clear()
            if self.rect_selector:
                self.rect_selector.set_active(False)

    def on_rotated_area_complete(self, center, width, height, angle):
        """Handle completion of rotated area selection with proper pixel extraction"""
        if self.current_data is None:
            return

        data = self.current_data.data
        if len(data.shape) != 3:
            return

        cx, cy = center

        # Create rotation matrix
        angle_rad = np.radians(-angle)  # Negative for inverse transform
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        # Get all pixel coordinates
        y_indices, x_indices = np.meshgrid(np.arange(data.shape[0]), np.arange(data.shape[1]), indexing='ij')

        # Transform all pixels to rectangle's local coordinate system
        local_x = (x_indices - cx) * cos_a - (y_indices - cy) * sin_a
        local_y = (x_indices - cx) * sin_a + (y_indices - cy) * cos_a

        # Create mask for pixels inside rectangle
        mask = (np.abs(local_x) <= width / 2) & (np.abs(local_y) <= height / 2)

        if not np.any(mask):
            return

        # Extract spectra from all pixels inside rectangle
        selected_spectra = data[mask]
        summed_spectrum = np.sum(selected_spectra, axis=0)

        # Get energy axis
        energy_axis = self.get_energy_axis()
        if energy_axis is None:
            energy_axis = np.arange(len(summed_spectrum))

        # Store selection info
        self.selected_areas = [(cx, cy, width, height, angle)]

        print(f"Rotated area: center=({cx:.0f},{cy:.0f}), size=({width:.0f}x{height:.0f}), angle={angle:.1f}°")
        print(f"Selected {np.sum(mask)} pixels")

        # Plot spectrum in parent window
        if self.parent is not None and hasattr(self.parent, 'ax'):
            self.parent.ax.clear()
            self.parent.ax.plot(energy_axis, summed_spectrum, 'k-', linewidth=0.8)
            self.parent.ax.set_xlabel('Energy Loss (eV)')
            self.parent.ax.set_ylabel('Intensity')
            self.parent.ax.set_title(f'EELS - Rotated Area ({width:.0f}×{height:.0f} px, {angle:.1f}°)')
            self.parent.canvas.draw()

    def on_line_mode(self, event):
        """Toggle line selection mode"""
        if self.line_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.point_btn.SetValue(False)
            self.area_btn.SetValue(False)
            self.selection_mode = 'line'
            self.line_start = None
            self.line_end = None

            if hasattr(self, 'zoom_selector') and self.zoom_selector:
                self.zoom_selector.set_active(False)
            if self.rect_selector:
                self.rect_selector.set_active(False)

            self.clear_selection_markers()
            self.map_canvas.draw()
        else:
            self.selection_mode = None

    def on_clear_selections(self, event):
        """Clear all selections"""
        self.selected_points = []
        self.selected_areas = []
        self.line_start = None
        self.line_end = None

        self.clear_selection_markers()
        self._clear_line_preview()

        # Clear rotatable rectangle if it exists
        if hasattr(self, 'rotatable_rect') and self.rotatable_rect:
            self.rotatable_rect.clear()

        self.map_canvas.draw()

    # ==================== ARROW MOVEMENT ====================

    def on_map_panel_resize(self, event):
        """Position arrow control panel on bottom-left of map canvas"""
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            # Get canvas size and position
            canvas_rect = self.map_canvas.GetRect()

            # Get arrow panel size
            arrow_size = self.arrow_control_panel.GetBestSize()

            # Position on bottom-left with 10px margins
            x = canvas_rect.x
            y = canvas_rect.y + canvas_rect.height - arrow_size.height

            self.arrow_control_panel.SetPosition((x, y))
            self.arrow_control_panel.SetSize(arrow_size)
            self.arrow_control_panel.Raise()
            self.arrow_control_panel.Show()

        event.Skip()

    def on_canvas_paint(self, event):
        """Keep arrow buttons visible when canvas redraws"""
        event.Skip()
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            wx.CallLater(10, self._reposition_arrow_panel)

    def _reposition_arrow_panel(self):
        """Helper to reposition arrow panel"""
        if hasattr(self, 'arrow_control_panel') and self.arrow_control_panel:
            canvas_rect = self.map_canvas.GetRect()
            arrow_size = self.arrow_control_panel.GetBestSize()

            x = canvas_rect.x
            y = canvas_rect.y + canvas_rect.height - arrow_size.height

            self.arrow_control_panel.SetPosition((x, y))
            self.arrow_control_panel.Show()
            self.arrow_control_panel.Raise()
            self.arrow_control_panel.Refresh()
            self.arrow_control_panel.Update()

    def on_arrow_move(self, direction):
        """Handle arrow button clicks to move selected elements"""
        if self.current_data is None:
            return

        # Get the actual data array
        if hasattr(self.current_data, 'data'):
            data_array = self.current_data.data
        else:
            data_array = self.current_data

        # If nothing is selected, return
        if (not self.selected_points and
            not self.line_start and
            not (hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.center is not None) and
            not self.selected_areas):
            return

        # Get plot dimensions for calculating movement step (0.5% of dimension)
        height, width = data_array.shape[:2]
        step_x = max(1, int(width * 0.005))
        step_y = max(1, int(height * 0.005))

        # Determine movement direction
        dx, dy = 0, 0
        if direction == 'left':
            dx = -step_x
        elif direction == 'right':
            dx = step_x
        elif direction == 'up':
            dy = -step_y
        elif direction == 'down':
            dy = step_y

        moved = False

        # Move selected points
        if self.selected_points:
            moved = self._move_and_replot_point(dx, dy, width, height)

        # Move selected lines
        elif self.line_start and self.line_end:
            moved = self._move_and_replot_line(dx, dy, width, height)

        # Move rotatable rectangle
        elif hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.center is not None:
            moved = self._move_and_replot_rotated(dx, dy, width, height)

        # Move selected areas
        elif self.selected_areas:
            moved = self._move_and_replot_area(dx, dy, width, height)

        if moved:
            print(f"Moved selection by ({dx}, {dy})")

    def _move_and_replot_point(self, dx, dy, width, height):
        """Move point and replot"""
        if not self.selected_points:
            return False

        new_points = []
        for point in self.selected_points:
            if len(point) == 3:
                x, y, size = point
            else:
                x, y = point
                size = 1

            new_x = int(np.clip(x + dx, 0, width - 1))
            new_y = int(np.clip(y + dy, 0, height - 1))
            new_points.append((new_x, new_y, size))

        self.selected_points = new_points

        # Redraw marker (red)
        self.clear_selection_markers()
        point = self.selected_points[-1]
        x, y, size = point

        if size == 1:
            marker, = self.map_ax.plot(x, y, 'r+', markersize=15, markeredgewidth=2)
            marker._is_selection_marker = True
        else:
            from matplotlib.patches import Rectangle
            half_size = size / 2
            rect = Rectangle((x - half_size, y - half_size), size, size,
                             linewidth=2, edgecolor='darkred', facecolor='red', alpha=0.3)
            rect._is_selection_marker = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw()
        self.map_canvas.Refresh()

        # Replot spectrum
        self.plot_point_spectrum(x, y)

        # Keep arrow buttons visible
        wx.CallLater(10, self._reposition_arrow_panel)

        return True

    def _move_and_replot_line(self, dx, dy, width, height):
        """Move line and replot"""
        if not self.line_start or not self.line_end:
            return False

        x1, y1 = self.line_start
        x2, y2 = self.line_end

        # Move both endpoints
        new_x1 = int(np.clip(x1 + dx, 0, width - 1))
        new_y1 = int(np.clip(y1 + dy, 0, height - 1))
        new_x2 = int(np.clip(x2 + dx, 0, width - 1))
        new_y2 = int(np.clip(y2 + dy, 0, height - 1))

        self.line_start = (new_x1, new_y1)
        self.line_end = (new_x2, new_y2)

        # Redraw markers (red)
        self.clear_selection_markers()

        line, = self.map_ax.plot([new_x1, new_x2], [new_y1, new_y2], 'r-', linewidth=2)
        line._is_selection_marker = True

        marker1, = self.map_ax.plot(new_x1, new_y1, 'ro', markersize=8, markeredgecolor='darkred', markeredgewidth=2)
        marker1._is_selection_marker = True
        marker2, = self.map_ax.plot(new_x2, new_y2, 'ro', markersize=8, markeredgecolor='darkred', markeredgewidth=2)
        marker2._is_selection_marker = True

        self.map_canvas.draw()

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        # Replot spectrum
        self.plot_line_spectrum(new_x1, new_y1, new_x2, new_y2)

        return True

    def _move_and_replot_area(self, dx, dy, width, height):
        """Move area and replot"""
        if not self.selected_areas:
            return False

        area = self.selected_areas[-1]

        # Check if it's a rotated area (5 elements) or standard area (4 elements)
        if len(area) == 5:
            # Rotated area: (cx, cy, w, h, angle)
            cx, cy, w, h, angle = area
            new_cx = float(np.clip(cx + dx, 0, width - 1))
            new_cy = float(np.clip(cy + dy, 0, height - 1))
            self.selected_areas[-1] = (new_cx, new_cy, w, h, angle)

            # Update rotatable rectangle if exists
            if hasattr(self, 'rotatable_rect') and self.rotatable_rect:
                self.rotatable_rect.center = (new_cx, new_cy)
                self.rotatable_rect.draw_rectangle()
        else:
            # Standard area: (x1, y1, x2, y2)
            x1, y1, x2, y2 = area
            rect_width = x2 - x1
            rect_height = y2 - y1

            new_x1 = int(np.clip(x1 + dx, 0, width - rect_width))
            new_y1 = int(np.clip(y1 + dy, 0, height - rect_height))
            new_x2 = new_x1 + rect_width
            new_y2 = new_y1 + rect_height

            self.selected_areas[-1] = (new_x1, new_y1, new_x2, new_y2)

            # Redraw marker (red)
            self.clear_selection_markers()

            from matplotlib.patches import Rectangle
            rect = Rectangle((new_x1, new_y1), rect_width, rect_height,
                             linewidth=2, edgecolor='red', facecolor='red', alpha=0.2)
            rect._is_selection_marker = True
            self.map_ax.add_patch(rect)

            self.map_canvas.draw()

            # Replot spectrum
            self.plot_area_spectrum(new_x1, new_y1, new_x2, new_y2)

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        return True

    def _move_and_replot_rotated(self, dx, dy, width, height):
        """Move rotatable rectangle and replot"""
        if not hasattr(self, 'rotatable_rect') or not self.rotatable_rect:
            return False

        if self.rotatable_rect.center is None:
            return False

        # Get current center
        cx, cy = self.rotatable_rect.center

        # Calculate new center with boundary checking
        new_cx = float(np.clip(cx + dx, 0, width - 1))
        new_cy = float(np.clip(cy + dy, 0, height - 1))

        # Update center
        self.rotatable_rect.center = (new_cx, new_cy)

        # Redraw the rectangle
        self.rotatable_rect.draw_rectangle()

        self.map_canvas.draw()

        # Keep arrow buttons visible
        if hasattr(self, 'arrow_control_panel'):
            self.arrow_control_panel.Raise()

        # Replot spectrum for rotated area
        self.on_rotated_area_complete(
            (new_cx, new_cy),
            self.rotatable_rect.width,
            self.rotatable_rect.height,
            self.rotatable_rect.angle
        )

        return True

    def on_intensity_map(self, event):
        """Show intensity map"""
        if self.current_data is None:
            return
        self.plot_current_map()

    def on_right_click(self, event):
        """Show right-click context menu"""
        menu = wx.Menu()

        # Save EELS Plot option at top
        save_plot_item = menu.Append(wx.ID_ANY, "Save Current EELS Plot...")
        self.Bind(wx.EVT_MENU, self.on_save_eels_plot, save_plot_item)

        menu.AppendSeparator()

        # HeatMap submenu
        heatmap_menu = wx.Menu()
        colormaps = ['Greens', 'plasma', 'viridis', 'inferno', 'magma', 'hot', 'cool', 'gray', 'jet',
                     'rainbow', 'turbo', 'cividis', 'Spectral', 'coolwarm', 'RdYlBu', 'RdBu',
                     'Blues', 'Reds', 'Oranges', 'Purples', 'YlOrRd', 'YlGnBu', 'RdPu', 'BuPu',
                     'GnBu', 'PuBu', 'YlGn', 'binary', 'bone', 'copper', 'autumn', 'winter',
                     'spring', 'summer', 'twilight', 'hsv', 'nipy_spectral', 'terrain', 'ocean',
                     'gist_earth', 'seismic', 'bwr', 'BrBG', 'PRGn', 'PiYG', 'RdGy', 'RdYlGn']

        for cmap in colormaps:
            item = heatmap_menu.AppendRadioItem(wx.ID_ANY, cmap)
            if cmap == self.current_cmap:
                item.Check(True)
            self.Bind(wx.EVT_MENU, lambda evt, c=cmap: self.change_colormap(c), item)

        menu.AppendSubMenu(heatmap_menu, "HeatMap")

        menu.AppendSeparator()

        # Export submenu
        export_menu = wx.Menu()
        export_excel_item = export_menu.Append(wx.ID_ANY, "Export to Excel...")
        export_csv_item = export_menu.Append(wx.ID_ANY, "Export to CSV...")
        export_image_item = export_menu.Append(wx.ID_ANY, "Export Map Image...")

        self.Bind(wx.EVT_MENU, self.on_export_excel, export_excel_item)
        self.Bind(wx.EVT_MENU, self.on_export_csv, export_csv_item)
        self.Bind(wx.EVT_MENU, self.on_export_map_image, export_image_item)

        menu.AppendSubMenu(export_menu, "Export")

        menu.AppendSeparator()

        data_browser_item = menu.Append(wx.ID_ANY, "Data Browser...")
        info_item = menu.Append(wx.ID_ANY, "Info...")
        menu.AppendSeparator()
        clear_item = menu.Append(wx.ID_ANY, "Clear Selections")

        self.Bind(wx.EVT_MENU, self.on_show_data_browser, data_browser_item)
        self.Bind(wx.EVT_MENU, self.on_show_info, info_item)
        self.Bind(wx.EVT_MENU, self.on_clear_selections, clear_item)

        self.PopupMenu(menu)
        menu.Destroy()

    def change_colormap(self, cmap):
        """Change the colormap"""
        self.current_cmap = cmap
        self.plot_current_map()

    # ==================== SELECTION HANDLERS ====================

    def on_map_scroll(self, event):
        """Handle mouse scroll on map"""
        if event.inaxes != self.map_ax:
            return

        if self.selection_mode == 'point':
            # Adjust point size with scroll wheel
            if event.button == 'up':
                self.point_size = min(50, self.point_size + 1)
            elif event.button == 'down':
                self.point_size = max(1, self.point_size - 1)

            self._update_point_preview(event.xdata, event.ydata)

        elif self.selection_mode == 'area':
            # Rotation is handled by RotatableRectangle.on_scroll
            pass

    def _update_point_preview(self, x, y):
        """Update point size preview rectangle"""
        self._clear_point_preview()

        if x is None or y is None:
            return

        if self.point_size == 1:
            # Single pixel - show as red cross
            marker, = self.map_ax.plot(x, y, 'r+', markersize=15,
                                       markeredgewidth=2, alpha=0.5)
            marker._is_point_preview = True
        else:
            # Show red rectangle preview
            half_size = self.point_size / 2
            rect = patches.Rectangle((x - half_size, y - half_size),
                                      self.point_size, self.point_size,
                                      linewidth=2, edgecolor='red',
                                      facecolor='salmon', alpha=0.3)
            rect._is_point_preview = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw_idle()

    def _clear_point_preview(self):
        """Clear point preview"""
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_point_preview') and artist._is_point_preview:
                try:
                    artist.remove()
                except:
                    pass
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_point_preview') and artist._is_point_preview:
                try:
                    artist.remove()
                except:
                    pass

    def on_map_click(self, event):
        """Handle click on map"""
        if event.inaxes != self.map_ax:
            return
        if self.current_data is None:
            return

        x, y = int(event.xdata), int(event.ydata)

        # Validate coordinates
        data_shape = self.current_data.data.shape
        if len(data_shape) < 2:
            return

        max_x = data_shape[1] if len(data_shape) >= 2 else data_shape[0]
        max_y = data_shape[0]

        if x < 0 or x >= max_x or y < 0 or y >= max_y:
            return

        if self.selection_mode == 'point':
            # Clear previous markers and preview
            self.clear_selection_markers()
            self._clear_point_preview()

            # Store point info with size
            self.selected_points = [(x, y, self.point_size)]

            if self.point_size == 1:
                # Single pixel - show as cross (RED)
                marker, = self.map_ax.plot(x, y, 'r+', markersize=15, markeredgewidth=2)
                marker._is_selection_marker = True
            else:
                # Show rectangle for multi-pixel selection (RED)
                from matplotlib.patches import Rectangle
                half_size = self.point_size / 2
                rect = Rectangle((x - half_size, y - half_size), self.point_size, self.point_size,
                                 linewidth=2, edgecolor='darkred', facecolor='red', alpha=0.3)
                rect._is_selection_marker = True
                self.map_ax.add_patch(rect)

            self.map_canvas.draw()

            # Extract and plot spectrum
            self.plot_point_spectrum(x, y)

        elif self.selection_mode == 'line':
            if self.line_start is None:
                # First click - start of line
                # Clear any previous line markers first (for starting a new line)
                self.clear_selection_markers()
                self._clear_line_preview()

                self.line_start = (x, y)
                # Draw start marker (RED)
                marker, = self.map_ax.plot(x, y, 'ro', markersize=8, markeredgecolor='darkred', markeredgewidth=2)
                marker._is_selection_marker = True
                self.map_canvas.draw()
            else:
                # Second click - end of line
                self.line_end = (x, y)

                # Clear preview line
                self._clear_line_preview()

                # Clear previous markers and redraw final line (RED)
                self.clear_selection_markers()

                # Draw center line (solid, RED)
                line, = self.map_ax.plot([self.line_start[0], self.line_end[0]],
                                         [self.line_start[1], self.line_end[1]],
                                         'r-', linewidth=2)
                line._is_selection_marker = True

                # Draw end points (RED)
                marker1, = self.map_ax.plot(self.line_start[0], self.line_start[1],
                                            'ro', markersize=8,
                                            markeredgecolor='darkred', markeredgewidth=2)
                marker1._is_selection_marker = True
                marker2, = self.map_ax.plot(self.line_end[0], self.line_end[1],
                                            'ro', markersize=8,
                                            markeredgecolor='darkred', markeredgewidth=2)
                marker2._is_selection_marker = True

                self.map_canvas.draw()

                self.plot_line_spectrum(self.line_start[0], self.line_start[1],
                                        self.line_end[0], self.line_end[1])

                # Reset for next line
                self.line_start = None
                self.line_end = None

    def on_map_motion(self, event):
        """Handle mouse motion on map for line preview"""
        if self.selection_mode != 'line' or self.line_start is None:
            return
        if event.inaxes != self.map_ax:
            return
        if event.xdata is None or event.ydata is None:
            return

        x, y = int(event.xdata), int(event.ydata)

        # Update line preview (RED)
        self._clear_line_preview()
        self._line_preview, = self.map_ax.plot([self.line_start[0], x],
                                               [self.line_start[1], y],
                                               'r--', linewidth=1.5, alpha=0.7)
        self._line_preview._is_selection_marker = True
        self.map_canvas.draw_idle()

    def _clear_line_preview(self):
        """Clear the line preview"""
        if self._line_preview is not None:
            try:
                self._line_preview.remove()
            except (ValueError, AttributeError):
                pass
            self._line_preview = None

    def clear_selection_markers(self):
        """Clear all selection markers from map"""
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_selection_marker') and artist._is_selection_marker:
                artist.remove()
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_selection_marker') and artist._is_selection_marker:
                artist.remove()

    def on_area_select(self, eclick, erelease):
        """Handle area selection"""
        if self.current_data is None:
            return

        x1, y1 = int(eclick.xdata), int(eclick.ydata)
        x2, y2 = int(erelease.xdata), int(erelease.ydata)

        x_min, x_max = min(x1, x2), max(x1, x2)
        y_min, y_max = min(y1, y2), max(y1, y2)

        self.selected_areas.append((x_min, y_min, x_max, y_max))

        # Draw rectangle selection marker (RED)
        self.clear_selection_markers()
        rect = patches.Rectangle((x_min, y_min), x_max - x_min, y_max - y_min,
                                 linewidth=2, edgecolor='red', facecolor='red', alpha=0.2)
        rect._is_selection_marker = True
        self.map_ax.add_patch(rect)
        self.map_canvas.draw()

        self.plot_area_spectrum(x_min, y_min, x_max, y_max)

    # ==================== SPECTRUM PLOTTING ====================

    def plot_point_spectrum(self, x, y):
        """Plot spectrum from point or area around point"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D (x, y, energy) for spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        point_size = getattr(self, 'point_size', 1)

        if point_size == 1:
            spectrum = data[y, x, :]
            title = f'EELS Spectrum at Point ({x}, {y})'
        else:
            half_size = point_size // 2
            y1 = max(0, y - half_size)
            y2 = min(data.shape[0], y + half_size + 1)
            x1 = max(0, x - half_size)
            x2 = min(data.shape[1], x + half_size + 1)

            spectrum = np.mean(data[y1:y2, x1:x2, :], axis=(0, 1))
            title = f'EELS Spectrum at Point ({x}, {y}) [{point_size}×{point_size} px]'

        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        # Plot in parent KherveFitting window
        if self.parent is not None and hasattr(self.parent, 'ax'):
            self._plot_to_parent(energy, spectrum, title, 'point', x, y, point_size)

    def plot_area_spectrum(self, x1, y1, x2, y2):
        """Plot summed spectrum from selected area"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            wx.MessageBox("Data must be 3D for area spectrum extraction",
                          "Error", wx.OK | wx.ICON_ERROR)
            return

        # Sum spectra from area
        spectrum = np.sum(data[y1:y2+1, x1:x2+1, :], axis=(0, 1))
        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        num_pixels = (y2 - y1 + 1) * (x2 - x1 + 1)
        title = f'EELS Area Spectrum ({num_pixels} pixels)'

        if self.parent is not None and hasattr(self.parent, 'ax'):
            self._plot_to_parent(energy, spectrum, title, 'area',
                                 selection_info={'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})

    def plot_line_spectrum(self, x1, y1, x2, y2):
        """Plot summed spectrum from line profile"""
        if self.current_data is None:
            return

        data = self.current_data.data

        if len(data.shape) != 3:
            return

        # Get line coordinates using Bresenham's algorithm
        num_points = int(max(abs(x2 - x1), abs(y2 - y1))) + 1
        x_coords = np.linspace(x1, x2, num_points).astype(int)
        y_coords = np.linspace(y1, y2, num_points).astype(int)

        # Clamp to valid range
        x_coords = np.clip(x_coords, 0, data.shape[1] - 1)
        y_coords = np.clip(y_coords, 0, data.shape[0] - 1)

        # Sum spectra along line
        spectrum = np.sum(data[y_coords, x_coords, :], axis=0)
        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        title = f'EELS Line Spectrum ({num_points} pixels)'

        if self.parent is not None and hasattr(self.parent, 'ax'):
            self._plot_to_parent(energy, spectrum, title, 'line',
                                 selection_info={'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})

    def _plot_to_parent(self, energy, spectrum, title, sel_type,
                        x=None, y=None, size=None, selection_info=None):
        """Plot spectrum to parent KherveFitting window and store data"""
        if self.parent is None:
            return

        # Create EELS~Plot sheet if it doesn't exist
        if 'EELS~Plot' not in self.parent.Data['Core levels']:
            self.parent.Data['Core levels']['EELS~Plot'] = {}

        # Build selection info
        if selection_info is None:
            selection_info = {}
        if sel_type == 'point':
            selection_info = {'type': 'point', 'x': x, 'y': y, 'size': size}
        elif sel_type:
            selection_info['type'] = sel_type

        # Convert to lists with proper formatting
        energy_list = [float(f"{v:.2f}") for v in energy]
        spectrum_list = [float(f"{v:.2f}") for v in spectrum]

        # Store data in B.E./Raw Data format for compatibility - WITH COMPLETE Background
        self.parent.Data['Core levels']['EELS~Plot'] = {
            'Name': 'EELS~Plot',
            'B.E.': energy_list,
            'Raw Data': spectrum_list,
            '_EELS_selection': selection_info,
            '_EELS_type': 'plot',
            'Background': {
                'Bkg Type': '',
                'Bkg Low': '',
                'Bkg High': '',
                'Bkg Offset Low': '',
                'Bkg Offset High': '',
                'Bkg X': energy_list.copy(),
                'Bkg Y': spectrum_list.copy()
            }
        }

        # Update sheet combobox
        if 'EELS~Plot' not in [self.parent.sheet_combobox.GetString(i)
                               for i in range(self.parent.sheet_combobox.GetCount())]:
            self.parent.sheet_combobox.Append('EELS~Plot')
        self.parent.sheet_combobox.SetValue('EELS~Plot')

        # Plot
        self.parent.ax.clear()
        self.parent.ax.plot(energy, spectrum, 'k-', linewidth=0.8)
        self.parent.ax.set_xlabel('Energy Loss (eV)')
        self.parent.ax.set_ylabel('Counts')
        self.parent.ax.set_title(title)
        self.parent.ax.ticklabel_format(axis='y', style='scientific', scilimits=(0, 0))
        self.parent.ax.grid(False)

        self.parent.canvas.draw()

    def plot_sum_spectrum_to_parent(self):
        """Plot sum spectrum from all pixels to parent window"""
        if self.current_data is None or self.parent is None:
            return

        data = self.current_data.data
        if len(data.shape) != 3:
            return

        # Sum over spatial dimensions
        spectrum = np.sum(data, axis=(0, 1))
        energy = self.get_energy_axis()

        if energy is None:
            energy = np.arange(len(spectrum))

        title = f'EELS Sum Spectrum ({data.shape[0]}×{data.shape[1]} px)'

        self._plot_to_parent(energy, spectrum, title, 'sum')

    def get_energy_axis(self):
        """Get energy axis from current data"""
        if self.current_data is None:
            return None

        if hasattr(self.current_data, 'axes_manager'):
            if len(self.current_data.axes_manager.signal_axes) > 0:
                signal_axis = self.current_data.axes_manager.signal_axes[0]

                # For EELS, if offset is very negative (like -6250), ignore it
                # This typically means the calibration is for the absolute energy position
                # but we want to display energy loss starting from 0
                if signal_axis.offset < -1000:
                    print(f"EELS: Ignoring large negative offset ({signal_axis.offset:.1f} eV)")
                    print(f"  Displaying energy loss from 0 to {signal_axis.size * signal_axis.scale:.1f} eV")
                    return signal_axis.axis_no_offset

                return signal_axis.axis

        # Fallback: create channel numbers
        if len(self.current_data.data.shape) == 3:
            return np.arange(self.current_data.data.shape[2])
        elif len(self.current_data.data.shape) == 1:
            return np.arange(len(self.current_data.data))

        return None

    # ==================== FILE IMPORT ====================

    def on_import_dm3(self, event):
        """Import DM3 file"""
        self._import_file("DM3 files (*.dm3)|*.dm3|All files (*.*)|*.*")

    def on_import_dm4(self, event):
        """Import DM4 file"""
        self._import_file("DM4 files (*.dm4)|*.dm4|All files (*.*)|*.*")

    def on_import_hdf5(self, event):
        """Import HDF5 file"""
        self._import_file("HDF5 files (*.hdf5;*.h5)|*.hdf5;*.h5|All files (*.*)|*.*")

    def _import_file(self, wildcard):
        """Generic file import"""
        with wx.FileDialog(self, "Open EELS Map",
                           wildcard=wildcard,
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                file_path = dlg.GetPath()
                self.load_file(file_path)

                if self.current_data is not None:
                    self.create_eels_map_output(file_path)

    def load_file(self, file_path):
        """Load EELS file using standalone readers (no HyperSpy)"""
        try:
            if not EELS_UTILITIES_AVAILABLE:
                wx.MessageBox("EELS_Utilities not available. Cannot load EELS data.",
                              "Error", wx.OK | wx.ICON_ERROR)
                return

            loaded_data = None
            ext = os.path.splitext(file_path)[1].lower()

            # Try standalone readers
            if ext in ['.dm3', '.dm4', '.hdf5', '.h5', '.hspy']:
                try:
                    loaded_data = load_eels(file_path)
                    print(f"Successfully loaded {ext} file with EELS_Utilities")
                except Exception as e:
                    print(f"EELS_Utilities failed: {e}")
                    loaded_data = None

            if loaded_data is None:
                wx.MessageBox(f"Could not load file. Supported formats: DM3, DM4, HDF5",
                              "Load Error", wx.OK | wx.ICON_ERROR)
                return

            # Handle list of signals
            if isinstance(loaded_data, list):
                if len(loaded_data) == 1:
                    loaded_data = loaded_data[0]
                else:
                    # Use first 3D signal found
                    for sig in loaded_data:
                        if len(sig.data.shape) == 3:
                            loaded_data = sig
                            break
                    else:
                        loaded_data = loaded_data[0]

            self.loaded_signals.append({
                'data': loaded_data,
                'filename': os.path.basename(file_path),
                'path': file_path
            })

            self.current_data = loaded_data
            self.plot_current_map()

        except Exception as e:
            wx.MessageBox(f"Error loading file:\n{str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    def create_eels_map_output(self, file_path):
        """Create output files and add EELS data to parent window.Data"""
        import json
        import shutil

        try:
            # Initialize parent Data structure if needed
            if self.parent is not None:
                if not hasattr(self.parent, 'Data'):
                    from libraries.ConfigFile import Init_Measurement_Data
                    self.parent.Data = Init_Measurement_Data(self.parent)

                if 'Core levels' not in self.parent.Data:
                    self.parent.Data['Core levels'] = {}

            base_name = os.path.splitext(os.path.basename(file_path))[0]
            excel_path = os.path.join(os.path.dirname(file_path),
                                      f"{base_name}_EELS.xlsx")
            json_path = os.path.join(os.path.dirname(file_path),
                                     f"{base_name}_EELS.json")

            # Set FilePath
            if self.parent is not None and hasattr(self.parent, 'Data'):
                self.parent.Data['FilePath'] = excel_path

                if hasattr(self.parent, 'current_file_path'):
                    self.parent.current_file_path = excel_path

                if hasattr(self.parent, 'Working_directory'):
                    self.parent.Working_directory = os.path.dirname(excel_path)

            # Get energy range
            energy_axis = self.get_energy_axis()
            if energy_axis is not None:
                energy_min = f"{np.min(energy_axis):.2f}"
                energy_max = f"{np.max(energy_axis):.2f}"
                energy_range = f"{energy_min} - {energy_max} eV"
            else:
                energy_range = "N/A"

            # Create sum spectrum
            sum_signal = self.current_data.sum()
            spectrum_data = sum_signal.data

            # Create Excel file
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # EELS~Plot sheet
            ws_plot = wb.create_sheet("EELS~Plot")
            ws_plot.append(['Energy Loss (eV)', 'Intensity', f'Range: {energy_range}'])

            for i, intensity in enumerate(spectrum_data):
                if energy_axis is not None:
                    ws_plot.append([f"{energy_axis[i]:.2f}", f"{intensity:.2f}"])
                else:
                    ws_plot.append([f"{i:.2f}", f"{intensity:.2f}"])

            # EELS~Map sheet
            ws_map = wb.create_sheet("EELS~Map")
            map_data = np.sum(self.current_data.data, axis=2)

            ws_map.append([f'EELS Intensity Map - Range: {energy_range}'])
            ws_map.append([''] * (map_data.shape[1] + 1))

            for row in map_data:
                ws_map.append([f"{val:.2f}" for val in row])

            wb.save(excel_path)
            print(f"EELS data exported to: {excel_path}")

            # Add to parent window.Data
            if self.parent is not None and hasattr(self.parent, 'Data'):
                energy_values = energy_axis if energy_axis is not None else np.arange(len(spectrum_data))

                self.parent.Data['Core levels']['EELS~Plot'] = {
                    'Name': 'EELS~Plot',
                    'B.E.': [float(f"{v:.2f}") for v in energy_values],
                    'Raw Data': [float(f"{v:.2f}") for v in spectrum_data],
                    '_EELS_type': 'plot'
                }

                self.parent.Data['Core levels']['EELS~Map'] = {
                    'Name': 'EELS~Map',
                    'Map_Intensity': map_data.tolist(),
                    'Map_Shape': list(map_data.shape),
                    'Energy_Range': energy_range,
                    '_EELS_type': 'map',
                    '_Source_Path': file_path
                }

                # Update UI
                if hasattr(self.parent, 'SetStatusText'):
                    self.parent.SetStatusText(
                        f"Working Directory: {os.path.dirname(excel_path)}", 0)

                if hasattr(self.parent, 'SetTitle'):
                    self.parent.SetTitle(
                        f"KherveFitting - {os.path.basename(excel_path)}")

                # Add sheets to combobox
                for sheet_name in ['EELS~Plot', 'EELS~Map']:
                    if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                                          for i in range(self.parent.sheet_combobox.GetCount())]:
                        self.parent.sheet_combobox.Append(sheet_name)

            # Create JSON file
            json_data = {
                'FilePath': excel_path,
                'Core levels': {
                    'EELS~Plot': {
                        'Name': 'EELS~Plot',
                        'B.E.': [float(f"{v:.2f}") for v in energy_values],
                        'Raw Data': [float(f"{v:.2f}") for v in spectrum_data],
                        '_EELS_type': 'plot'
                    },
                    'EELS~Map': {
                        'Name': 'EELS~Map',
                        'Map_Intensity': [[float(f"{val:.2f}") for val in row]
                                          for row in map_data],
                        'Map_Shape': list(map_data.shape),
                        'Energy_Range': energy_range,
                        '_EELS_type': 'map',
                        '_Source_Path': file_path
                    }
                }
            }

            with open(json_path, 'w') as jf:
                json.dump(json_data, jf, indent=2)
            print(f"JSON data saved to: {json_path}")

            wx.MessageBox(f"EELS data exported to:\n{excel_path}\n\nJSON: {json_path}",
                          "Export Complete", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Error creating EELS output:\n{str(e)}",
                          "Error", wx.OK | wx.ICON_ERROR)
            import traceback
            traceback.print_exc()

    # ==================== MAP PLOTTING ====================

    def plot_current_map(self):
        """Plot the current data as a map"""
        if self.current_data is None:
            return

        # Clear previous colorbar
        if self.current_colorbar is not None:
            try:
                self.current_colorbar.remove()
            except:
                pass
            self.current_colorbar = None

        self.map_ax.clear()
        data = self.current_data.data
        cmap = self.current_cmap

        # Debug: Print data info
        print(f"EELS data shape: {data.shape}")
        if hasattr(self.current_data, 'axes_manager'):
            print(f"Number of navigation axes: {len(self.current_data.axes_manager.navigation_axes)}")
            print(f"Number of signal axes: {len(self.current_data.axes_manager.signal_axes)}")
            for i, ax in enumerate(self.current_data.axes_manager._axes):
                print(f"  Axis {i}: {ax.name}, scale={ax.scale:.4f} {ax.units}, size={ax.size}")

        if len(data.shape) == 2:
            sum_image = data
            title = 'EELS Image'
        elif len(data.shape) == 3:
            # Sum across energy axis (last axis) to get intensity map
            sum_image = np.sum(data, axis=2)
            title = f'EELS Intensity Map ({data.shape[0]}×{data.shape[1]} px, {data.shape[2]} ch)'
            print(f"Sum image shape: {sum_image.shape}, min={np.min(sum_image):.2f}, max={np.max(sum_image):.2f}")
        elif len(data.shape) == 4:
            sum_image = np.sum(data, axis=(2, 3))
            title = f'Sum Image (4D: {data.shape})'
        else:
            self.map_ax.text(0.5, 0.5, f'Unsupported shape: {data.shape}',
                             ha='center', va='center', transform=self.map_ax.transAxes)
            self.map_canvas.draw()
            return

        im = self.map_ax.imshow(sum_image, cmap=cmap, origin='upper')

        # Remove axis labels - use scale bar instead
        self.map_ax.set_xlabel('')
        self.map_ax.set_ylabel('')
        self.map_ax.set_title('')
        self.map_ax.set_xticks([])
        self.map_ax.set_yticks([])

        # Add scale bar
        self._add_scale_bar()

        # Add colorbar
        from mpl_toolkits.axes_grid1 import make_axes_locatable
        divider = make_axes_locatable(self.map_ax)
        cax = divider.append_axes("right", size="5%", pad=0.05)
        self.current_colorbar = self.map_figure.colorbar(im, cax=cax)

        self.map_figure.tight_layout(pad=0.5)
        self._add_scale_bar()

        self.map_canvas.draw()

        # Plot sum spectrum in parent window
        self.plot_sum_spectrum_to_parent()

        self.reinitialize_selectors()

    def _add_scale_bar(self):
        """Add scale bar to map if scale information is available"""
        self._remove_scale_bar()

        if self.current_data is None:
            return

        # Try to get scale from axes_manager
        scale_value = None
        units = 'px'

        if hasattr(self.current_data, 'axes_manager'):
            nav_axes = self.current_data.axes_manager.navigation_axes
            print(f"Navigation axes available: {len(nav_axes)}")
            if nav_axes:
                axis = nav_axes[0]
                scale_value = axis.scale
                units = axis.units if axis.units else 'px'
                print(f"Scale bar: scale_value={scale_value} {units}")

        if scale_value is None or scale_value <= 0:
            print("No valid scale information available for scale bar")
            return

        # Get current view limits
        xlim = self.map_ax.get_xlim()
        ylim = self.map_ax.get_ylim()

        view_width = abs(xlim[1] - xlim[0])
        view_height = abs(ylim[1] - ylim[0])

        # Calculate appropriate scale bar length
        target_fraction = 0.2
        target_pixels = view_width * target_fraction
        target_physical = target_pixels * scale_value

        # Round to nice number
        nice_values = [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000]
        scale_bar_physical = min(nice_values, key=lambda x: abs(x - target_physical))
        scale_bar_pixels = scale_bar_physical / scale_value

        print(f"Scale bar: {scale_bar_physical} {units} = {scale_bar_pixels:.1f} pixels")

        # Position in lower right
        margin_x = view_width * 0.05
        margin_y = view_height * 0.05

        x_start = xlim[1] - margin_x - scale_bar_pixels
        y_pos = max(ylim[0], ylim[1]) - margin_y

        bar_height = view_height * 0.02

        # Draw scale bar
        rect = patches.Rectangle((x_start, y_pos - bar_height), scale_bar_pixels, bar_height,
                                  linewidth=1, edgecolor='white', facecolor='white',
                                  zorder=101)
        rect._is_scale_bar = True
        self.map_ax.add_patch(rect)

        # Add label
        if units == 'nm':
            if scale_bar_physical >= 1000:
                label = f'{scale_bar_physical/1000:.0f} µm'
            else:
                label = f'{scale_bar_physical:.0f} nm'
        else:
            label = f'{scale_bar_physical:.0f} {units}'

        text = self.map_ax.text(x_start + scale_bar_pixels / 2,
                                y_pos - bar_height - margin_y * 0.5,
                                label, ha='center', va='top', fontsize=9,
                                fontweight='bold', color='white', zorder=102,
                                path_effects=[path_effects.withStroke(
                                    linewidth=2, foreground='black')])
        text._is_scale_bar = True

    def _remove_scale_bar(self):
        """Remove existing scale bar elements"""
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_scale_bar') and artist._is_scale_bar:
                artist.remove()
        for artist in self.map_ax.texts[:]:
            if hasattr(artist, '_is_scale_bar') and artist._is_scale_bar:
                artist.remove()

    def draw_saved_selection(self, selection_info):
        """Draw a saved selection on the map"""
        self._clear_saved_selection()

        if selection_info is None:
            return

        sel_type = selection_info.get('type')

        if sel_type == 'point':
            x, y = selection_info['x'], selection_info['y']
            size = selection_info.get('size', 1)

            if size == 1:
                marker, = self.map_ax.plot(x, y, 'k+', markersize=15, markeredgewidth=2)
                marker._is_saved_selection = True
            else:
                half_size = size / 2
                rect = patches.Rectangle((x - half_size, y - half_size), size, size,
                                          linewidth=2, edgecolor='black', facecolor='none')
                rect._is_saved_selection = True
                self.map_ax.add_patch(rect)

        elif sel_type == 'line':
            line, = self.map_ax.plot([selection_info['x1'], selection_info['x2']],
                                     [selection_info['y1'], selection_info['y2']],
                                     'k-', linewidth=2)
            line._is_saved_selection = True

        elif sel_type == 'area':
            x1, y1 = selection_info['x1'], selection_info['y1']
            x2, y2 = selection_info['x2'], selection_info['y2']
            rect = patches.Rectangle((x1, y1), x2 - x1, y2 - y1,
                                      linewidth=2, edgecolor='black', facecolor='none')
            rect._is_saved_selection = True
            self.map_ax.add_patch(rect)

        self.map_canvas.draw_idle()

    def _clear_saved_selection(self):
        """Clear saved selection markers"""
        for artist in self.map_ax.patches[:]:
            if hasattr(artist, '_is_saved_selection') and artist._is_saved_selection:
                try:
                    artist.remove()
                except:
                    pass
        for artist in self.map_ax.lines[:]:
            if hasattr(artist, '_is_saved_selection') and artist._is_saved_selection:
                try:
                    artist.remove()
                except:
                    pass

    def reinitialize_selectors(self):
        """Reinitialize rectangle selector after redraw"""
        if self.rect_selector is not None:
            try:
                # Recreate selector
                self.rect_selector = RectangleSelector(
                    self.map_ax,
                    self.on_area_select,
                    useblit=True,
                    props=dict(facecolor='green', edgecolor='lime',
                               alpha=0.3, fill=True),
                    button=[1],
                    minspanx=2,
                    minspany=2,
                    spancoords='pixels',
                    interactive=False
                )
                if self.selection_mode != 'area':
                    self.rect_selector.set_active(False)
            except:
                pass

    # ==================== EXPORT ====================

    def on_export_excel(self, event):
        """Export data to Excel"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Save Excel File",
                           wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                self._export_to_excel(dlg.GetPath())

    def _export_to_excel(self, path):
        """Export current data to Excel"""
        import openpyxl

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EELS Spectrum"

            energy = self.get_energy_axis()
            spectrum = np.sum(self.current_data.data, axis=(0, 1))

            ws.append(['Energy Loss (eV)', 'Intensity'])
            for i, val in enumerate(spectrum):
                e = energy[i] if energy is not None else i
                ws.append([f"{e:.2f}", f"{val:.2f}"])

            wb.save(path)
            wx.MessageBox(f"Exported to: {path}", "Export Complete",
                          wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            wx.MessageBox(f"Export error: {e}", "Error", wx.OK | wx.ICON_ERROR)

    def on_export_csv(self, event):
        """Export spectrum to CSV"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Save CSV File",
                           wildcard="CSV files (*.csv)|*.csv",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                self._export_to_csv(dlg.GetPath())

    def _export_to_csv(self, path):
        """Export current spectrum to CSV"""
        try:
            energy = self.get_energy_axis()
            spectrum = np.sum(self.current_data.data, axis=(0, 1))

            with open(path, 'w') as f:
                f.write("Energy Loss (eV),Intensity\n")
                for i, val in enumerate(spectrum):
                    e = energy[i] if energy is not None else i
                    f.write(f"{e:.2f},{val:.2f}\n")

            wx.MessageBox(f"Exported to: {path}", "Export Complete",
                          wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            wx.MessageBox(f"Export error: {e}", "Error", wx.OK | wx.ICON_ERROR)

    def on_export_map_image(self, event):
        """Export map as image"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "Error", wx.OK | wx.ICON_ERROR)
            return

        with wx.FileDialog(self, "Save Map Image",
                           wildcard="PNG files (*.png)|*.png|TIFF files (*.tiff)|*.tiff",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                self.map_figure.savefig(dlg.GetPath(), dpi=300, bbox_inches='tight')
                wx.MessageBox(f"Map saved to: {dlg.GetPath()}",
                              "Export Complete", wx.OK | wx.ICON_INFORMATION)

    # ==================== WINDOW MANAGEMENT ====================

    def on_save_eels_plot(self, event):
        """Save current EELS plot to a numbered sheet in Excel and window.Data"""
        if self.parent is None:
            return

        # Check if Excel file exists
        if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
            wx.MessageBox("No Excel file found. Please ensure the EELS data was imported correctly.",
                          "No File", wx.OK | wx.ICON_WARNING)
            return

        # Find next available EELS~Plot number
        existing_sheets = list(self.parent.Data.get('Core levels', {}).keys())
        plot_num = 1
        while f'EELS~Plot{plot_num}' in existing_sheets:
            plot_num += 1

        sheet_name = f'EELS~Plot{plot_num}'

        # Gather current selection info
        selection_info = self._get_current_selection_info()

        if selection_info is None:
            wx.MessageBox("No selection to save. Please select a point, line, or area first.",
                          "No Selection", wx.OK | wx.ICON_WARNING)
            return

        # Get current spectrum data from parent
        if not hasattr(self.parent, 'ax') or len(self.parent.ax.lines) == 0:
            wx.MessageBox("No EELS plot data to save.", "No Data", wx.OK | wx.ICON_WARNING)
            return

        # Get spectrum data from plot
        line = self.parent.ax.lines[0]
        energy = line.get_xdata()
        intensity = line.get_ydata()

        # Convert to lists with proper formatting
        energy_list = [float(f"{v:.2f}") for v in energy]
        intensity_list = [float(f"{v:.2f}") for v in intensity]

        # Create sheet data - USE SAME STRUCTURE AS XPS SHEETS with complete Background
        import datetime
        import json
        sheet_data = {
            'Name': sheet_name,
            'B.E.': energy_list,
            'Raw Data': intensity_list,
            '_EELS_type': 'plot',
            '_EELS_selection': selection_info,
            '_EELS_save_time': datetime.datetime.now().isoformat(),
            'Background': {
                'Bkg Type': '',
                'Bkg Low': '',
                'Bkg High': '',
                'Bkg Offset Low': '',
                'Bkg Offset High': '',
                'Bkg X': energy_list.copy(),
                'Bkg Y': intensity_list.copy()
            }
        }

        # Save to parent Data
        if 'Core levels' not in self.parent.Data:
            self.parent.Data['Core levels'] = {}

        self.parent.Data['Core levels'][sheet_name] = sheet_data

        # Update sheet combobox
        if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                              for i in range(self.parent.sheet_combobox.GetCount())]:
            self.parent.sheet_combobox.Append(sheet_name)

        # Save to Excel file
        try:
            import pandas as pd

            file_path = self.parent.Data['FilePath']

            # Create DataFrame for this plot
            eels_df = pd.DataFrame({
                'Energy Loss (eV)': [f"{v:.2f}" for v in energy],
                'Intensity': [f"{v:.2f}" for v in intensity]
            })

            # Append to Excel file
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                eels_df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Saved {sheet_name} to Excel: {file_path}")

            # Also update JSON
            json_path = file_path.replace('.xlsx', '.json')
            if os.path.exists(json_path):
                with open(json_path, 'r') as f:
                    json_data = json.load(f)

                if 'Core levels' not in json_data:
                    json_data['Core levels'] = {}

                json_data['Core levels'][sheet_name] = {
                    'Name': sheet_name,
                    'B.E.': [float(f"{v:.2f}") for v in energy],
                    'Raw Data': [float(f"{v:.2f}") for v in intensity],
                    '_EELS_type': 'plot',
                    '_EELS_selection': selection_info
                }

                with open(json_path, 'w') as f:
                    json.dump(json_data, f, indent=2)

            # Short message
            wx.MessageBox(f"Saved as {sheet_name}", "Saved", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            print(f"Error saving to Excel: {e}")
            import traceback
            traceback.print_exc()
            wx.MessageBox(f"Error saving to Excel: {e}", "Error", wx.OK | wx.ICON_ERROR)

    def _get_current_selection_info(self):
        """Get current selection information"""
        # Check for rotatable rectangle first
        if hasattr(self, 'rotatable_rect') and self.rotatable_rect and self.rotatable_rect.active:
            rect = self.rotatable_rect
            return {
                'type': 'rectangle',
                'center_x': float(rect.center[0]),
                'center_y': float(rect.center[1]),
                'width': float(rect.width),
                'height': float(rect.height),
                'angle': float(rect.angle)
            }
        elif self.selected_points:
            point = self.selected_points[-1]
            if len(point) == 3:
                x, y, size = point
            else:
                x, y = point
                size = 1
            return {'type': 'point', 'x': int(x), 'y': int(y), 'size': int(size)}
        elif self.selected_areas:
            area = self.selected_areas[-1]
            if len(area) == 5:
                # Rotated area: (cx, cy, width, height, angle)
                cx, cy, w, h, angle = area
                return {'type': 'rectangle', 'center_x': float(cx), 'center_y': float(cy),
                        'width': float(w), 'height': float(h), 'angle': float(angle)}
            else:
                # Standard area: (x1, y1, x2, y2)
                x1, y1, x2, y2 = area
                return {'type': 'area', 'x1': int(x1), 'y1': int(y1), 'x2': int(x2), 'y2': int(y2)}
        return None

    def on_show_data_browser(self, event):
        """Show data browser window"""
        if not hasattr(self, 'data_browser_window') or self.data_browser_window is None:
            self.data_browser_window = EELSDataBrowserWindow(self)
        self.data_browser_window.Show()
        self.data_browser_window.Raise()

    def on_show_info(self, event):
        """Show info window"""
        if self.current_data is None:
            wx.MessageBox("No data loaded", "Info", wx.OK | wx.ICON_INFORMATION)
            return

        if not hasattr(self, 'info_window') or self.info_window is None:
            self.info_window = EELSInfoWindow(self)
        self.info_window.update_info(self.current_data)
        self.info_window.Show()
        self.info_window.Raise()

    def on_close(self, event):
        """Handle window close event"""
        # Clear parent reference
        if self.parent is not None and hasattr(self.parent, 'eels_window'):
            self.parent.eels_window = None

        # Destroy sensitivity window if open
        if hasattr(self, 'sensitivity_window') and self.sensitivity_window:
            try:
                self.sensitivity_window.Destroy()
            except:
                pass

        # Destroy info window if open
        if hasattr(self, 'info_window') and self.info_window:
            try:
                self.info_window.Destroy()
            except:
                pass

        # Destroy data browser window if open
        if hasattr(self, 'data_browser_window') and self.data_browser_window:
            try:
                self.data_browser_window.Destroy()
            except:
                pass

        self.Destroy()


class EELSSensitivityWindow(wx.Frame):
    """Window for EELS display controls"""

    def __init__(self, parent):
        super().__init__(parent, title="Display Controls", size=(300, 200),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)

        self.parent = parent
        self.init_ui()
        self.Centre()

    def init_ui(self):
        panel = wx.Panel(self)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Colormap selection
        cmap_sizer = wx.BoxSizer(wx.HORIZONTAL)
        cmap_label = wx.StaticText(panel, label="Colormap:")
        self.cmap_combo = wx.ComboBox(panel,
                                      choices=['plasma', 'viridis', 'inferno',
                                               'magma', 'hot', 'jet', 'gray'],
                                      style=wx.CB_READONLY)
        self.cmap_combo.SetValue(self.parent.current_cmap)
        self.cmap_combo.Bind(wx.EVT_COMBOBOX, self.on_cmap_change)

        cmap_sizer.Add(cmap_label, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)
        cmap_sizer.Add(self.cmap_combo, 1, wx.ALL | wx.EXPAND, 5)
        sizer.Add(cmap_sizer, 0, wx.EXPAND)

        # Brightness/Contrast (placeholder for future)
        sizer.Add(wx.StaticLine(panel), 0, wx.EXPAND | wx.ALL, 5)

        info_text = wx.StaticText(panel, label="Scroll wheel adjusts point size\n"
                                               "when in point selection mode.")
        sizer.Add(info_text, 0, wx.ALL, 10)

        panel.SetSizer(sizer)

    def on_cmap_change(self, event):
        self.parent.current_cmap = self.cmap_combo.GetValue()
        self.parent.plot_current_map()


class EELSInfoWindow(wx.Frame):
    """Window showing EELS data information"""

    def __init__(self, parent):
        super().__init__(parent, title="EELS Info", size=(400, 350),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
        self.parent = parent
        self.init_ui()
        self.Centre()

    def init_ui(self):
        panel = wx.Panel(self)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.info_text = wx.TextCtrl(panel, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)
        sizer.Add(self.info_text, 1, wx.EXPAND | wx.ALL, 10)

        close_btn = wx.Button(panel, wx.ID_CLOSE, "Close")
        close_btn.Bind(wx.EVT_BUTTON, lambda e: self.Hide())
        sizer.Add(close_btn, 0, wx.ALL | wx.ALIGN_CENTER, 5)

        panel.SetSizer(sizer)

    def update_info(self, signal):
        """Update info display with signal information"""
        info_lines = []
        info_lines.append("=" * 40)
        info_lines.append("EELS Data Information")
        info_lines.append("=" * 40)

        if hasattr(signal, 'data'):
            info_lines.append(f"\nData Shape: {signal.data.shape}")
            info_lines.append(f"Data Type: {signal.data.dtype}")
            info_lines.append(f"Min Value: {np.min(signal.data):.2f}")
            info_lines.append(f"Max Value: {np.max(signal.data):.2f}")
            info_lines.append(f"Mean Value: {np.mean(signal.data):.2f}")

        if hasattr(signal, 'axes_manager'):
            info_lines.append("\n" + "-" * 40)
            info_lines.append("Axes Information:")
            info_lines.append("-" * 40)

            for i, axis in enumerate(signal.axes_manager._axes):
                info_lines.append(f"\nAxis {i}: {axis.name}")
                info_lines.append(f"  Size: {axis.size}")
                info_lines.append(f"  Scale: {axis.scale:.6f}")
                info_lines.append(f"  Offset: {axis.offset:.6f}")
                info_lines.append(f"  Units: {axis.units}")

            info_lines.append("\n" + "-" * 40)
            info_lines.append(f"Navigation Axes: {len(signal.axes_manager.navigation_axes)}")
            info_lines.append(f"Signal Axes: {len(signal.axes_manager.signal_axes)}")

        self.info_text.SetValue("\n".join(info_lines))


class EELSDataBrowserWindow(wx.Frame):
    """Window for browsing EELS data"""

    def __init__(self, parent):
        super().__init__(parent, title="EELS Data Browser", size=(350, 400),
                         style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)
        self.parent = parent
        self.init_ui()
        self.Centre()
        self.refresh_tree()

    def init_ui(self):
        panel = wx.Panel(self)
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Tree control for data hierarchy
        self.tree = wx.TreeCtrl(panel, style=wx.TR_DEFAULT_STYLE | wx.TR_HIDE_ROOT)
        sizer.Add(self.tree, 1, wx.EXPAND | wx.ALL, 5)

        # Buttons
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        refresh_btn = wx.Button(panel, label="Refresh")
        refresh_btn.Bind(wx.EVT_BUTTON, lambda e: self.refresh_tree())
        btn_sizer.Add(refresh_btn, 0, wx.ALL, 5)

        close_btn = wx.Button(panel, wx.ID_CLOSE, "Close")
        close_btn.Bind(wx.EVT_BUTTON, lambda e: self.Hide())
        btn_sizer.Add(close_btn, 0, wx.ALL, 5)

        sizer.Add(btn_sizer, 0, wx.ALIGN_CENTER)

        panel.SetSizer(sizer)

    def refresh_tree(self):
        """Refresh the data tree"""
        self.tree.DeleteAllItems()
        root = self.tree.AddRoot("EELS Data")

        if self.parent.loaded_signals:
            for i, signal_info in enumerate(self.parent.loaded_signals):
                filename = signal_info.get('filename', f'Signal {i}')
                signal = signal_info.get('data')

                item = self.tree.AppendItem(root, filename)

                if signal is not None and hasattr(signal, 'data'):
                    shape_item = self.tree.AppendItem(item, f"Shape: {signal.data.shape}")
                    dtype_item = self.tree.AppendItem(item, f"Type: {signal.data.dtype}")

                    if hasattr(signal, 'axes_manager'):
                        axes_item = self.tree.AppendItem(item, "Axes")
                        for j, axis in enumerate(signal.axes_manager._axes):
                            ax_text = f"{axis.name}: {axis.size} pts, {axis.scale:.4f} {axis.units}"
                            self.tree.AppendItem(axes_item, ax_text)

        self.tree.ExpandAll()


class RotatableRectangle:
    """Custom rotatable rectangle selector for area selection"""

    def __init__(self, ax, callback):
        self.ax = ax
        self.callback = callback
        self.active = False
        self.angle = 0

        self.center = None
        self.width = None
        self.height = None

        self.rectangle = None
        self.rotation_handle = None
        self.resize_handles = []

        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None
        self.resize_corner = None
        self.angle_text = None

        self._stored_xlim = None
        self._stored_ylim = None

        self.cid_press = ax.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.cid_release = ax.figure.canvas.mpl_connect('button_release_event', self.on_release)
        self.cid_motion = ax.figure.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.cid_scroll = ax.figure.canvas.mpl_connect('scroll_event', self.on_scroll)

    def start_selection(self, event):
        """Start creating a new rectangle"""
        if event.inaxes != self.ax or event.button != 1:
            return

        # If not active, create new rectangle at click position
        if not self.active:
            # Store axis limits BEFORE any drawing
            self._stored_xlim = self.ax.get_xlim()
            self._stored_ylim = self.ax.get_ylim()

            self.center = (event.xdata, event.ydata)
            self.width = 10
            self.height = 10
            self.angle = 0
            self.active = True
            self.draw_rectangle()

    def draw_rectangle(self):
        """Draw the rotatable rectangle and handles"""
        # Clear existing elements safely
        if self.rectangle:
            try:
                self.rectangle.remove()
            except (ValueError, AttributeError):
                pass
            self.rectangle = None

        for handle in self.resize_handles:
            try:
                handle.remove()
            except (ValueError, AttributeError):
                pass
        self.resize_handles = []

        if self.rotation_handle:
            try:
                self.rotation_handle.remove()
            except (ValueError, AttributeError):
                pass
            self.rotation_handle = None

        if self.angle_text:
            try:
                self.angle_text.remove()
            except (ValueError, AttributeError):
                pass
            self.angle_text = None

        if self.center is None:
            self.ax.figure.canvas.draw_idle()
            return

        from matplotlib.patches import Rectangle
        from matplotlib.transforms import Affine2D

        rect = Rectangle((-self.width / 2, -self.height / 2), self.width, self.height,
                         linewidth=2, edgecolor='red', facecolor='red', alpha=0.3)

        t = Affine2D().rotate_deg(self.angle).translate(*self.center) + self.ax.transData
        rect.set_transform(t)
        rect._is_selection_marker = True

        self.rectangle = self.ax.add_patch(rect)

        # Rotation handle (red)
        handle_dist = max(self.width, self.height) / 2 + 10
        angle_rad = np.radians(self.angle)
        handle_x = self.center[0] + handle_dist * np.sin(angle_rad)
        handle_y = self.center[1] + handle_dist * np.cos(angle_rad)

        self.rotation_handle = self.ax.plot(handle_x, handle_y, 'ro', markersize=10,
                                            markeredgecolor='darkred', markeredgewidth=2)[0]
        self.rotation_handle._is_selection_marker = True

        # Show angle text when rotating
        if self.rotating:
            self.angle_text = self.ax.text(handle_x, handle_y + 5, f'{self.angle:.1f}°',
                                           fontsize=9, color='white', fontweight='bold',
                                           ha='center', va='bottom',
                                           bbox=dict(boxstyle='round,pad=0.2', facecolor='darkred', alpha=0.7))
            self.angle_text._is_selection_marker = True

        # Resize handles at corners (red)
        corners = self.get_corners()
        for corner in corners:
            handle = self.ax.plot(corner[0], corner[1], 'rs', markersize=8,
                                  markeredgecolor='darkred', markeredgewidth=1)[0]
            handle._is_selection_marker = True
            self.resize_handles.append(handle)

        # ALWAYS restore stored axis limits to prevent expansion
        if self._stored_xlim is not None and self._stored_ylim is not None:
            self.ax.set_xlim(self._stored_xlim)
            self.ax.set_ylim(self._stored_ylim)

        self.ax.figure.canvas.draw_idle()

    def get_corners(self):
        """Get the four corners of the rotated rectangle"""
        angle_rad = np.radians(self.angle)
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        hw = self.width / 2
        hh = self.height / 2
        corners_local = [(-hw, -hh), (hw, -hh), (hw, hh), (-hw, hh)]

        corners = []
        for x, y in corners_local:
            rx = x * cos_a - y * sin_a + self.center[0]
            ry = x * sin_a + y * cos_a + self.center[1]
            corners.append((rx, ry))

        return corners

    def on_scroll(self, event):
        """Handle mouse scroll to rotate rectangle"""
        if not self.active or event.inaxes != self.ax:
            return

        # Check if mouse is near the rectangle
        if self.center is None:
            return

        # Rotate by 5 degrees per scroll step
        if event.button == 'up':
            self.angle += 5
        elif event.button == 'down':
            self.angle -= 5

        # Normalize angle to -180 to 180
        self.angle = ((self.angle + 180) % 360) - 180

        self.draw_rectangle()

        # Trigger callback
        if self.callback:
            self.callback(self.center, self.width, self.height, self.angle)

    def on_press(self, event):
        """Handle mouse press"""
        if not self.active or event.inaxes != self.ax or event.button != 1:
            return

        # Store current limits
        if self._stored_xlim is None:
            self._stored_xlim = self.ax.get_xlim()
            self._stored_ylim = self.ax.get_ylim()

        # Check rotation handle first (highest priority)
        if self.rotation_handle:
            handle_data = self.rotation_handle.get_data()
            dist = np.sqrt((event.xdata - handle_data[0][0]) ** 2 + (event.ydata - handle_data[1][0]) ** 2)
            if dist < 8:
                self.rotating = True
                self.drag_start = (event.xdata, event.ydata)
                return

        # Check resize handles - but only if very close (within 5 pixels)
        closest_handle_dist = float('inf')
        closest_handle_idx = None
        for i, handle in enumerate(self.resize_handles):
            handle_data = handle.get_data()
            dist = np.sqrt((event.xdata - handle_data[0][0]) ** 2 + (event.ydata - handle_data[1][0]) ** 2)
            if dist < closest_handle_dist:
                closest_handle_dist = dist
                closest_handle_idx = i

        # Only resize if very close to handle (< 5 pixels)
        if closest_handle_dist < 5:
            self.resizing = True
            self.resize_corner = closest_handle_idx
            self.drag_start = (event.xdata, event.ydata)
            return

        # Check if inside rectangle for dragging (most common operation)
        if self.point_in_rectangle(event.xdata, event.ydata):
            self.dragging = True
            self.drag_start = (event.xdata, event.ydata)
            return

        # If not inside but close to resize handle (< 10 pixels), allow resize
        if closest_handle_dist < 10:
            self.resizing = True
            self.resize_corner = closest_handle_idx
            self.drag_start = (event.xdata, event.ydata)

    def on_release(self, event):
        """Handle mouse release"""
        was_interacting = self.dragging or self.resizing or self.rotating

        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None

        # Restore limits after any operation
        if self._stored_xlim is not None and self._stored_ylim is not None:
            self.ax.set_xlim(self._stored_xlim)
            self.ax.set_ylim(self._stored_ylim)

        # Trigger callback after interaction
        if was_interacting and self.callback and self.center:
            self.callback(self.center, self.width, self.height, self.angle)

    def on_motion(self, event):
        """Handle mouse motion"""
        if event.inaxes != self.ax or self.drag_start is None:
            return

        if event.xdata is None or event.ydata is None:
            return

        dx = event.xdata - self.drag_start[0]
        dy = event.ydata - self.drag_start[1]

        if self.dragging:
            new_x = self.center[0] + dx
            new_y = self.center[1] + dy

            # Clamp to stored bounds
            if self._stored_xlim is not None:
                new_x = max(self._stored_xlim[0], min(self._stored_xlim[1], new_x))
            if self._stored_ylim is not None:
                y_min = min(self._stored_ylim[0], self._stored_ylim[1])
                y_max = max(self._stored_ylim[0], self._stored_ylim[1])
                new_y = max(y_min, min(y_max, new_y))

            self.center = (new_x, new_y)
            self.drag_start = (event.xdata, event.ydata)
            self.draw_rectangle()

        elif self.rotating:
            angle_to_center = np.arctan2(event.xdata - self.center[0], event.ydata - self.center[1])
            self.angle = np.degrees(angle_to_center)
            self.draw_rectangle()

        elif self.resizing:
            angle_rad = np.radians(-self.angle)
            cos_a = np.cos(angle_rad)
            sin_a = np.sin(angle_rad)

            local_x = (event.xdata - self.center[0]) * cos_a - (event.ydata - self.center[1]) * sin_a
            local_y = (event.xdata - self.center[0]) * sin_a + (event.ydata - self.center[1]) * cos_a

            # Allow minimum size of 1 pixel
            self.width = max(1, abs(local_x) * 2)
            self.height = max(1, abs(local_y) * 2)
            self.draw_rectangle()

    def point_in_rectangle(self, x, y):
        """Check if point is inside the rotated rectangle"""
        if self.center is None:
            return False

        angle_rad = np.radians(-self.angle)
        cos_a = np.cos(angle_rad)
        sin_a = np.sin(angle_rad)

        local_x = (x - self.center[0]) * cos_a - (y - self.center[1]) * sin_a
        local_y = (x - self.center[0]) * sin_a + (y - self.center[1]) * cos_a

        return abs(local_x) <= self.width / 2 and abs(local_y) <= self.height / 2

    def clear(self):
        """Clear the rectangle and handles, allow new selection"""
        if self.rectangle:
            try:
                self.rectangle.remove()
            except (ValueError, AttributeError):
                pass
            self.rectangle = None

        for handle in self.resize_handles:
            try:
                handle.remove()
            except (ValueError, AttributeError):
                pass
        self.resize_handles = []

        if self.rotation_handle:
            try:
                self.rotation_handle.remove()
            except (ValueError, AttributeError):
                pass
            self.rotation_handle = None

        if self.angle_text:
            try:
                self.angle_text.remove()
            except (ValueError, AttributeError):
                pass
            self.angle_text = None

        # Reset state to allow new selection
        self.center = None
        self.width = None
        self.height = None
        self.angle = 0
        self.active = False
        self.dragging = False
        self.rotating = False
        self.resizing = False
        self.drag_start = None

        self.ax.figure.canvas.draw_idle()

    def disconnect(self):
        """Disconnect event handlers"""
        try:
            self.ax.figure.canvas.mpl_disconnect(self.cid_press)
            self.ax.figure.canvas.mpl_disconnect(self.cid_release)
            self.ax.figure.canvas.mpl_disconnect(self.cid_motion)
            self.ax.figure.canvas.mpl_disconnect(self.cid_scroll)
        except (ValueError, AttributeError):
            pass


def open_eels_window(parent):
    """Open EELS analysis window"""
    window = EELSWindow(parent)
    window.Show()
    return window