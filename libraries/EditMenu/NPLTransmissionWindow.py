import wx
import wx.grid
import json
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
# matplotlib.use('WXAgg')  # Ensure we're using the wx backend
import openpyxl
from openpyxl.styles import Font
import re
import pandas as pd


class NPLTransmissionWindow(wx.Frame):
    def __init__(self, parent):
        super().__init__(parent, title="NPL Transmission Function", size=(820, 674), style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX) | wx.STAY_ON_TOP)
        self.parent = parent

        # Current slot (1, 2, or 3)
        self.current_slot = 1

        # Initialize transmission parameters
        self.a0 = None
        self.a1 = None
        self.a2 = None
        self.a3 = None
        self.a4 = None
        self.b1 = None
        self.b2 = None
        self.b3 = None
        self.b4 = None
        self.min_ke = None
        self.max_ke = None
        self.photon_energy = 1486.69  # Default Al Ka

        # Initialize hover elements
        self.hover_point = None
        self.annotation = None
        self.ax2 = None
        self.hover_cid = None
        self.test_cid = None

        self.init_ui()
        self.load_parameters_from_config()
        self.Centre()

    def init_ui(self):
        panel = wx.Panel(self)
        panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        main_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Left panel for controls
        left_panel = wx.Panel(panel, style=wx.BORDER_RAISED)
        left_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        left_sizer = wx.BoxSizer(wx.VERTICAL)

        # TOP SIZER: Slot selection, VMS file, and parameters (with border)
        top_static_box = wx.StaticBox(left_panel, label="")
        top_sizer = wx.StaticBoxSizer(top_static_box, wx.VERTICAL)

        # Slot selection
        slot_label = wx.StaticText(left_panel, label="Saved NPL Transmission:")
        top_sizer.Add(slot_label, 0, wx.ALL, 5)

        slot_radio_sizer = wx.BoxSizer(wx.VERTICAL)
        self.slot_radio_1 = wx.RadioButton(left_panel, label="NPL Wide", style=wx.RB_GROUP)
        self.slot_radio_2 = wx.RadioButton(left_panel, label="NPL Narrow 1")
        self.slot_radio_3 = wx.RadioButton(left_panel, label="NPL Narrow 2")

        self.slot_radio_1.SetValue(True)
        self.slot_radio_1.Bind(wx.EVT_RADIOBUTTON, lambda e: self.on_slot_change(1))
        self.slot_radio_2.Bind(wx.EVT_RADIOBUTTON, lambda e: self.on_slot_change(2))
        self.slot_radio_3.Bind(wx.EVT_RADIOBUTTON, lambda e: self.on_slot_change(3))

        slot_radio_sizer.Add(self.slot_radio_1, 0, wx.ALL, 2)
        slot_radio_sizer.Add(self.slot_radio_2, 0, wx.ALL, 2)
        slot_radio_sizer.Add(self.slot_radio_3, 0, wx.ALL, 2)
        top_sizer.Add(slot_radio_sizer, 0, wx.ALL, 5)

        # VMS file drop zone
        drop_label = wx.StaticText(left_panel, label="Drop NPL Vamas File Here:")
        top_sizer.Add(drop_label, 0, wx.ALL, 5)

        self.vms_path_text = wx.TextCtrl(left_panel, style=wx.TE_READONLY, size=(200, 25))
        self.vms_path_text.SetDropTarget(VMSFileDropTarget(self))
        top_sizer.Add(self.vms_path_text, 0, wx.ALL | wx.EXPAND, 5)

        browse_btn = wx.Button(left_panel, label="Browse NPL Vamas File")
        browse_btn.Bind(wx.EVT_BUTTON, self.on_browse_vms)
        top_sizer.Add(browse_btn, 0, wx.ALL | wx.EXPAND, 5)

        # Parameters display
        param_label = wx.StaticText(left_panel, label="Saved Transmission Parameters:")
        top_sizer.Add(param_label, 0, wx.ALL | wx.TOP, 5)

        self.param_grid = wx.grid.Grid(left_panel)
        self.param_grid.CreateGrid(11, 2)
        self.param_grid.SetColLabelValue(0, "ID")
        self.param_grid.SetColLabelValue(1, "Value")
        self.param_grid.SetColSize(0, 60)
        self.param_grid.SetColSize(1, 110)
        self.param_grid.EnableEditing(False)
        self.param_grid.SetLabelBackgroundColour(wx.Colour(254, 254, 254))

        # Make header row smaller
        self.param_grid.SetColLabelSize(25)
        self.param_grid.SetRowLabelSize(25)

        # Set parameter names
        param_names = ["a0", "a1", "a2", "a3", "a4", "b1", "b2", "b3", "b4", "KEmin", "KEmax"]
        for i, name in enumerate(param_names):
            self.param_grid.SetCellValue(i, 0, name)
            self.param_grid.SetCellValue(i, 1, "")

        top_sizer.Add(self.param_grid, 1, wx.ALL | wx.EXPAND, 5)

        # BOTTOM SIZER: Sheet selection and apply button (with border)
        bottom_static_box = wx.StaticBox(left_panel, label="")
        bottom_sizer = wx.StaticBoxSizer(bottom_static_box, wx.VERTICAL)

        # Sheet selection - CheckListBox
        sheet_label = wx.StaticText(left_panel, label="Select Core Levels:")
        bottom_sizer.Add(sheet_label, 0, wx.ALL | wx.TOP, 5)

        self.sheet_checklist = wx.CheckListBox(left_panel, choices=[])
        self.sheet_checklist.SetSize((200, 150))
        self.sheet_checklist.SetMaxSize((200, 150))
        self.sheet_checklist.Bind(wx.EVT_CONTEXT_MENU, self.on_sheet_context_menu)
        bottom_sizer.Add(self.sheet_checklist, 0, wx.ALL | wx.EXPAND, 5)

        # Apply button
        apply_btn = wx.Button(left_panel, label="Write Transmission to\nSelected Core Levels")
        apply_btn.SetBackgroundColour(wx.Colour(200, 230, 201))
        apply_btn.SetForegroundColour(wx.Colour(0, 0, 0))
        apply_btn.SetMinSize((125, 35))
        apply_btn.Bind(wx.EVT_BUTTON, self.on_apply_transmission)
        bottom_sizer.Add(apply_btn, 0, wx.ALL | wx.EXPAND, 5)

        # Add both sizers to left_sizer
        left_sizer.Add(top_sizer, 1, wx.ALL | wx.EXPAND, 0)
        left_sizer.Add(bottom_sizer, 0, wx.ALL | wx.EXPAND, 0)

        left_panel.SetSizer(left_sizer)
        left_panel.SetMinSize((220, -1))
        main_sizer.Add(left_panel, 0, wx.ALL, 1)

        # Right panel for plot
        right_panel = wx.Panel(panel, style=wx.BORDER_RAISED)
        right_sizer = wx.BoxSizer(wx.VERTICAL)

        self.figure = plt.figure(figsize=(6, 4))
        # self.figure.tight_layout()
        self.ax = self.figure.add_subplot(111)
        self.canvas = FigureCanvas(right_panel, -1, self.figure)

        right_sizer.Add(self.canvas, 1, wx.ALL | wx.EXPAND, 0)

        right_panel.SetSizer(right_sizer)
        main_sizer.Add(right_panel, 1, wx.ALL | wx.EXPAND, 0)

        panel.SetSizer(main_sizer)

        # Update sheet checklist
        self.update_sheet_list()

    def on_slot_change(self, slot_number):
        """Handle slot selection change"""
        self.current_slot = slot_number
        self.load_parameters_from_config()
        self.update_parameter_display()
        self.plot_transmission_function()

    def update_sheet_list(self):
        """Update the checklist with available sheets"""
        self.sheet_checklist.Clear()
        if hasattr(self.parent, 'Data') and 'Core levels' in self.parent.Data:
            sheets = list(self.parent.Data['Core levels'].keys())
            if sheets:
                self.sheet_checklist.AppendItems(sheets)

    def on_browse_vms(self, event):
        wildcard = "VMS files (*.vms)|*.vms"
        dlg = wx.FileDialog(self, "Choose VMS file", wildcard=wildcard, style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)

        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            self.load_vms_file(path)

        dlg.Destroy()

    def load_vms_file(self, filepath):
        """Parse VMS file and extract transmission function parameters"""
        try:
            with open(filepath, 'r') as f:
                content = f.read()

            # Extract parameters from ASCII section
            params = {}
            for line in content.split('\n'):
                if 'Response Function Parameter a0' in line:
                    params['a0'] = float(line.split()[-1])
                elif 'Response Function Parameter a1' in line:
                    params['a1'] = float(line.split()[-1])
                elif 'Response Function Parameter a2' in line:
                    params['a2'] = float(line.split()[-1])
                elif 'Response Function Parameter a3' in line:
                    params['a3'] = float(line.split()[-1])
                elif 'Response Function Parameter a4' in line:
                    params['a4'] = float(line.split()[-1])
                elif 'Response Function Parameter b1' in line:
                    params['b1'] = float(line.split()[-1])
                elif 'Response Function Parameter b2' in line:
                    params['b2'] = float(line.split()[-1])
                elif 'Response Function Parameter b3' in line:
                    params['b3'] = float(line.split()[-1])
                elif 'Response Function Parameter b4' in line:
                    params['b4'] = float(line.split()[-1])
                elif 'Minimum Calibration Energy/eV' in line:
                    params['min_ke'] = float(line.split()[-1])
                elif 'Maximum Calibration Energy/eV' in line:
                    params['max_ke'] = float(line.split()[-1])

            if len(params) != 11:
                wx.MessageBox("Could not find all required parameters in VMS file", "Error", wx.OK | wx.ICON_ERROR)
                return

            # Update class attributes
            self.a0 = params['a0']
            self.a1 = params['a1']
            self.a2 = params['a2']
            self.a3 = params['a3']
            self.a4 = params['a4']
            self.b1 = params['b1']
            self.b2 = params['b2']
            self.b3 = params['b3']
            self.b4 = params['b4']
            self.min_ke = params['min_ke']
            self.max_ke = params['max_ke']

            # Update display
            self.vms_path_text.SetValue(filepath)
            self.update_parameter_display()
            self.save_parameters_to_config()
            self.plot_transmission_function()

            wx.MessageBox(f"VMS file loaded and saved to slot {self.current_slot}!", "Success", wx.OK | wx.ICON_INFORMATION)

        except Exception as e:
            wx.MessageBox(f"Error loading VMS file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def update_parameter_display(self):
        """Update parameter grid with current values"""
        values = [self.a0, self.a1, self.a2, self.a3, self.a4,
                  self.b1, self.b2, self.b3, self.b4, self.min_ke, self.max_ke]

        for i, val in enumerate(values):
            if val is not None:
                self.param_grid.SetCellValue(i, 1, f"{val:.6f}")
            else:
                self.param_grid.SetCellValue(i, 1, "")

    def calculate_transmission(self, kinetic_energy):
        """
        Calculate Q(E) transmission function at given kinetic energy
        Uses normalized energy: ε = (E - 1000 eV) / 1000 eV
        Q(E) = (a0 + a1*ε + a2*ε^2 + a3*ε^3 + a4*ε^4) / (1 + b1*ε + b2*ε^2 + b3*ε^3 + b4*ε^4)
        """
        if None in [self.a0, self.a1, self.a2, self.a3, self.a4, self.b1, self.b2, self.b3, self.b4]:
            return 1.0

        # Normalize the kinetic energy
        epsilon = (kinetic_energy - 1000.0) / 1000.0

        numerator = self.a0 + self.a1 * epsilon + self.a2 * epsilon ** 2 + self.a3 * epsilon ** 3 + self.a4 * epsilon ** 4
        denominator = 1.0 + self.b1 * epsilon + self.b2 * epsilon ** 2 + self.b3 * epsilon ** 3 + self.b4 * epsilon ** 4

        return numerator / denominator

    def plot_transmission_function(self):
        """Plot the transmission function with hover interaction"""
        if None in [self.a0, self.min_ke, self.max_ke]:
            # Clear plot if no data
            self.ax.clear()
            self.ax.set_xlabel('Kinetic Energy (eV)', fontsize=12)
            self.ax.set_ylabel('Q(E) Transmission', fontsize=12)
            self.ax.set_title('NPL Transmission Function', fontsize=14)
            self.canvas.draw()
            return

        # Create kinetic energy range
        ke_range = np.linspace(self.min_ke, self.max_ke, 500)
        transmission = [self.calculate_transmission(ke) for ke in ke_range]

        # Store data for hover
        self.ke_data = ke_range
        self.trans_data = np.array(transmission)

        # Clear and redraw
        self.ax.clear()

        # Remove old top axis if it exists
        if self.ax2 is not None:
            try:
                self.ax2.remove()
            except:
                pass
            self.ax2 = None

        # Plot the transmission line
        self.line, = self.ax.plot(ke_range, transmission, 'b-', linewidth=2)
        self.ax.set_xlabel('Kinetic Energy (eV)', fontsize=12)
        self.ax.set_ylabel('Q(E) Transmission', fontsize=12)
        self.ax.set_title('NPL Transmission Function', fontsize=14)
        self.ax.grid(True, alpha=0.3)

        # Add top axis for binding energy
        self.ax2 = self.ax.twiny()
        self.ax2.set_xlabel('Binding Energy (eV)', fontsize=12)

        # Calculate BE limits from KE limits
        be_max = self.photon_energy - self.min_ke
        be_min = self.photon_energy - self.max_ke
        self.ax2.set_xlim(be_max, be_min)

        # Create hover marker on the main axis (initially invisible)
        self.hover_point, = self.ax.plot([], [], 'go', markersize=10, visible=False, zorder=10)

        # Create annotation for displaying values
        if self.annotation is not None:
            try:
                self.annotation.remove()
            except:
                pass

        self.annotation = self.ax.annotate('', xy=(0, 0), xytext=(15, 15),
                                           textcoords='offset points',
                                           bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.9, edgecolor='black'),
                                           fontsize=10,
                                           visible=False,
                                           zorder=11)

        # Draw first
        self.canvas.draw()

        # NOW connect the motion event AFTER everything is drawn
        if hasattr(self, 'hover_cid') and self.hover_cid is not None:
            self.canvas.mpl_disconnect(self.hover_cid)

        print("Connecting hover event...")
        self.hover_cid = self.canvas.mpl_connect('motion_notify_event', self.on_hover)
        print(f"Hover event connected with ID: {self.hover_cid}")

    def on_hover(self, event):
        """Handle mouse hover events on the plot"""
        try:
            # Check if mouse is over any axis in the figure
            if event.inaxes is None:
                if self.hover_point is not None and self.hover_point.get_visible():
                    self.hover_point.set_visible(False)
                    if self.annotation is not None:
                        self.annotation.set_visible(False)
                    self.canvas.draw_idle()
                return

            # Check if we're on the main axis (not the top twin axis)
            # The main axis is the one with the blue line
            if event.inaxes != self.ax and event.inaxes != self.ax2:
                return

            # Check if we have valid data
            if not hasattr(self, 'ke_data') or not hasattr(self, 'trans_data'):
                return

            if self.ke_data is None or self.trans_data is None:
                return

            if len(self.ke_data) == 0:
                return

            # Get mouse position - use the x coordinate from the event
            x_mouse = event.xdata
            y_mouse = event.ydata

            if x_mouse is None or y_mouse is None:
                return

            # If on the top axis (BE), convert to KE
            if event.inaxes == self.ax2:
                # Convert BE to KE
                x_mouse = self.photon_energy - x_mouse

            # Find index of closest KE value
            idx = np.argmin(np.abs(self.ke_data - x_mouse))
            ke_closest = self.ke_data[idx]
            trans_closest = self.trans_data[idx]

            # Calculate binding energy
            be_closest = self.photon_energy - ke_closest

            # Update hover point position
            if self.hover_point is not None:
                self.hover_point.set_data([ke_closest], [trans_closest])
                self.hover_point.set_visible(True)

            # Update annotation
            if self.annotation is not None:
                self.annotation.xy = (ke_closest, trans_closest)
                self.annotation.set_text(f'KE: {ke_closest:.2f} eV\nBE: {be_closest:.2f} eV\nQ(E): {trans_closest:.2f}')
                self.annotation.set_visible(True)

            # Redraw
            self.canvas.draw_idle()

        except Exception as e:
            print(f"Hover error: {e}")
            import traceback
            traceback.print_exc()

    def on_sheet_context_menu(self, event):
        """Handle right-click on sheet checklist"""
        # Get all sheet names from the checklist
        all_sheets = []
        for i in range(self.sheet_checklist.GetCount()):
            all_sheets.append(self.sheet_checklist.GetString(i))

        if not all_sheets:
            return

        # Group sheets by core level type
        sheet_groups = {}
        for sheet in all_sheets:
            core_type = self.extract_core_type(sheet)
            if core_type not in sheet_groups:
                sheet_groups[core_type] = []
            sheet_groups[core_type].append(sheet)

        # Create context menu
        menu = wx.Menu()

        # Add Select All and Unselect All at the top
        select_all_item = menu.Append(wx.ID_ANY, "Select All")
        unselect_all_item = menu.Append(wx.ID_ANY, "Unselect All")

        self.Bind(wx.EVT_MENU, self.on_select_all_sheets, select_all_item)
        self.Bind(wx.EVT_MENU, self.on_unselect_all_sheets, unselect_all_item)

        # Add separator
        if len(sheet_groups) > 0:
            menu.AppendSeparator()

        # Sort core level types for consistent ordering
        sorted_core_types = sorted(sheet_groups.keys())

        for core_type in sorted_core_types:
            sheets_of_type = sheet_groups[core_type]

            # Create select and unselect items for this core level type
            select_item = menu.Append(wx.ID_ANY, f"Select All {core_type} ({len(sheets_of_type)})")
            unselect_item = menu.Append(wx.ID_ANY, f"Unselect All {core_type} ({len(sheets_of_type)})")

            # Bind events
            self.Bind(wx.EVT_MENU, lambda evt, ct=core_type: self.select_all_core_type(ct), select_item)
            self.Bind(wx.EVT_MENU, lambda evt, ct=core_type: self.unselect_all_core_type(ct), unselect_item)

            # Add separator after each core level type (except the last one)
            if core_type != sorted_core_types[-1]:
                menu.AppendSeparator()

        # Show the menu
        self.sheet_checklist.PopupMenu(menu)
        menu.Destroy()

    def on_select_all_sheets(self, event):
        """Select all sheets in the checklist"""
        for i in range(self.sheet_checklist.GetCount()):
            self.sheet_checklist.Check(i, True)

    def on_unselect_all_sheets(self, event):
        """Unselect all sheets in the checklist"""
        for i in range(self.sheet_checklist.GetCount()):
            self.sheet_checklist.Check(i, False)

    def select_all_core_type(self, core_type):
        """Select all sheets of a specific core level type"""
        for i in range(self.sheet_checklist.GetCount()):
            sheet = self.sheet_checklist.GetString(i)
            if self.extract_core_type(sheet) == core_type:
                self.sheet_checklist.Check(i, True)

    def unselect_all_core_type(self, core_type):
        """Unselect all sheets of a specific core level type"""
        for i in range(self.sheet_checklist.GetCount()):
            sheet = self.sheet_checklist.GetString(i)
            if self.extract_core_type(sheet) == core_type:
                self.sheet_checklist.Check(i, False)

    def extract_core_type(self, sheet_name):
        """Extract the core level type (element + orbital) from sheet name"""
        if not sheet_name:
            return ""

        import re

        # Remove common suffixes/prefixes that might indicate sample info
        cleaned_name = sheet_name.strip()

        # Look for pattern: Element + orbital (e.g., C1s, O1s, Au4f, Ti2p, etc.)
        pattern = r'^([A-Z][a-z]?\d+[a-z]*)(?:\d+)?.*'
        match = re.match(pattern, cleaned_name)

        if match:
            return match.group(1)

        # Fallback: look for first part before underscore, space, or dash
        separators = ['_', ' ', '-', '.']
        for sep in separators:
            if sep in cleaned_name:
                first_part = cleaned_name.split(sep)[0]
                # Check if first part looks like a core level (has both letters and numbers)
                if re.match(r'^[A-Z][a-z]?\d+[a-z]*\d*$', first_part):
                    # Remove trailing digits to get just the core level type
                    core_level_match = re.match(r'^([A-Z][a-z]?\d+[a-z]*)', first_part)
                    if core_level_match:
                        return core_level_match.group(1)
                break

        # Final fallback: extract core level pattern from anywhere in name
        core_level_match = re.search(r'([A-Z][a-z]?\d+[a-z]*)', cleaned_name)
        if core_level_match:
            return core_level_match.group(1)

        # Last resort: return first 4 characters or whole name if shorter
        return cleaned_name[:4] if len(cleaned_name) > 4 else cleaned_name

    def update_sheet_list(self):
        """Update sheet checklist with available sheets"""
        if 'FilePath' in self.parent.Data and self.parent.Data['FilePath']:
            try:
                wb = openpyxl.load_workbook(self.parent.Data['FilePath'])
                sheets = [s for s in wb.sheetnames if s not in ['Experimental description', 'Sheet']]
                self.sheet_checklist.Clear()
                self.sheet_checklist.AppendItems(sheets)
                # Check all sheets by default
                for i in range(len(sheets)):
                    self.sheet_checklist.Check(i)
                wb.close()
            except:
                pass

    def on_apply_transmission(self, event):
        """Apply transmission correction to selected sheets in Excel and JSON"""
        if None in [self.a0, self.a1, self.a2, self.a3, self.a4, self.b1, self.b2, self.b3, self.b4]:
            wx.MessageBox("Please load a VMS file first", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Get checked sheets
        checked_sheets = [self.sheet_checklist.GetString(i) for i in range(self.sheet_checklist.GetCount())
                          if self.sheet_checklist.IsChecked(i)]

        if not checked_sheets:
            wx.MessageBox("Please select at least one sheet", "Error", wx.OK | wx.ICON_ERROR)
            return

        if 'FilePath' not in self.parent.Data or not self.parent.Data['FilePath']:
            wx.MessageBox("No Excel file is open", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Show warning dialog
        dlg = wx.MessageDialog(self,
                               f"Warning: This will remove all background regions and peaks from the selected {len(checked_sheets)} sheet(s) before applying transmission correction.\n\n"
                               f"Selected sheets:\n" + "\n".join(checked_sheets) + "\n\n"
                                                                                   "Do you want to continue?",
                               "Confirm Remove Regions and Peaks",
                               wx.YES_NO | wx.ICON_WARNING)

        result = dlg.ShowModal()
        dlg.Destroy()

        if result != wx.ID_YES:
            return

        try:
            # Remove background and peaks from all selected sheets first
            for sheet_name in checked_sheets:
                if sheet_name in self.parent.Data['Core levels']:
                    # Clear background recorded ranges
                    if 'Background' in self.parent.Data['Core levels'][sheet_name]:
                        self.parent.Data['Core levels'][sheet_name]['Background']['Recorded_Ranges'] = []
                        self.parent.Data['Core levels'][sheet_name]['Background'].update({
                            'Bkg Type': '',
                            'Bkg Low': '',
                            'Bkg High': '',
                            'Bkg Offset Low': '',
                            'Bkg Offset High': ''
                        })

                    # Clear fitting data
                    if 'Fitting' in self.parent.Data['Core levels'][sheet_name]:
                        self.parent.Data['Core levels'][sheet_name]['Fitting'] = {}

            # Get photon energy from parent
            photon_energy = self.parent.photons

            # Load workbook
            wb = openpyxl.load_workbook(self.parent.Data['FilePath'])

            # Create bold font for headers
            bold_font = Font(bold=True)

            sheets_processed = []

            for sheet_name in checked_sheets:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # Write headers in row 1 with bold formatting
                headers = ["Binding Energy", "Corrected Data", "Raw Data", "Transmission"]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = bold_font

                # Process each row
                row = 2
                while ws.cell(row=row, column=1).value is not None:
                    be_value = ws.cell(row=row, column=1).value

                    if be_value is not None:
                        # Check if raw data exists in column C, if not copy from column B
                        raw_data = ws.cell(row=row, column=3).value
                        if raw_data is None:
                            # Copy data from column B to column C (save original as raw)
                            raw_data = ws.cell(row=row, column=2).value
                            if raw_data is not None:
                                ws.cell(row=row, column=3, value=float(f"{float(raw_data):.2f}"))

                        if raw_data is not None:
                            # Convert BE to KE
                            ke = photon_energy - float(be_value)

                            # Calculate transmission
                            transmission = self.calculate_transmission(ke)

                            # Write transmission to column D
                            ws.cell(row=row, column=4, value=float(f"{transmission:.2f}"))

                            # Calculate corrected data: raw_data / transmission
                            corrected = float(raw_data) / transmission
                            ws.cell(row=row, column=2, value=float(f"{corrected:.2f}"))

                    row += 1

                sheets_processed.append(sheet_name)

            # Save Excel workbook AFTER all sheets are processed
            wb.save(self.parent.Data['FilePath'])
            wb.close()

            # Store the file path
            file_path = self.parent.Data['FilePath']

            # First reopen - load the data fresh from Excel
            from libraries.FileMenu.Open import open_xlsx_file
            open_xlsx_file(self.parent, file_path)

            # Calculate min/max intensity for processed sheets
            for sheet_name in sheets_processed:
                if sheet_name in self.parent.Data['Core levels']:
                    intensity_data = self.parent.Data['Core levels'][sheet_name].get('Intensity', [])
                    if intensity_data:
                        self.parent.Data['Core levels'][sheet_name]['Min Intensity'] = float(f"{min(intensity_data):.2f}")
                        self.parent.Data['Core levels'][sheet_name]['Max Intensity'] = float(f"{max(intensity_data):.2f}")

            # Save to JSON with updated min/max
            json_path = os.path.splitext(file_path)[0] + '.json'
            with open(json_path, 'w') as f:
                json.dump(self.parent.Data, f, indent=2, default=str)

            # Second reopen - reload everything with updated min/max
            open_xlsx_file(self.parent, file_path)


        except Exception as e:
            wx.MessageBox(f"Error applying transmission: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    def update_json_file(self, sheet_names):
        """Update the JSON file with transmission-corrected data"""
        json_path = os.path.splitext(self.parent.Data['FilePath'])[0] + '.json'

        try:
            # Load Excel data
            wb = openpyxl.load_workbook(self.parent.Data['FilePath'])

            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # Read data from Excel
                be_data = []
                corrected_data = []
                raw_data = []
                transmission_data = []

                row = 2
                while ws.cell(row=row, column=1).value is not None:
                    be_val = ws.cell(row=row, column=1).value
                    corr_val = ws.cell(row=row, column=2).value
                    raw_val = ws.cell(row=row, column=3).value
                    trans_val = ws.cell(row=row, column=4).value

                    if be_val is not None:
                        be_data.append(float(be_val))
                        corrected_data.append(float(corr_val) if corr_val is not None else 0.0)
                        raw_data.append(float(raw_val) if raw_val is not None else 0.0)
                        transmission_data.append(float(trans_val) if trans_val is not None else 1.0)

                    row += 1

                # Update parent Data structure
                if 'Core levels' in self.parent.Data and sheet_name in self.parent.Data['Core levels']:
                    self.parent.Data['Core levels'][sheet_name]['BE'] = be_data
                    self.parent.Data['Core levels'][sheet_name]['Intensity'] = corrected_data
                    self.parent.Data['Core levels'][sheet_name]['Raw Data'] = raw_data
                    self.parent.Data['Core levels'][sheet_name]['Transmission'] = transmission_data

            wb.close()

            # Save JSON file
            with open(json_path, 'w') as f:
                json.dump(self.parent.Data, f, indent=2, default=str)

        except Exception as e:
            print(f"Error updating JSON: {e}")

    def save_parameters_to_config(self):
        """Save transmission parameters to parent and config.json"""
        try:
            # Create parameter dictionary
            params = {
                'a0': self.a0,
                'a1': self.a1,
                'a2': self.a2,
                'a3': self.a3,
                'a4': self.a4,
                'b1': self.b1,
                'b2': self.b2,
                'b3': self.b3,
                'b4': self.b4,
                'min_ke': self.min_ke,
                'max_ke': self.max_ke
            }

            # Update parent's npl_transmission attribute for selected slot
            if self.current_slot == 1:
                self.parent.npl_transmission_1 = params
            elif self.current_slot == 2:
                self.parent.npl_transmission_2 = params
            elif self.current_slot == 3:
                self.parent.npl_transmission_3 = params

            # Save using parent's save_config method
            if hasattr(self.parent, 'save_config'):
                self.parent.save_config()

        except Exception as e:
            print(f"Error saving config: {e}")

    def load_parameters_from_config(self):
        """Load transmission parameters from parent for current slot"""
        try:
            # Load from parent's npl_transmission attribute for selected slot
            params = None
            if self.current_slot == 1 and hasattr(self.parent, 'npl_transmission_1'):
                params = self.parent.npl_transmission_1
            elif self.current_slot == 2 and hasattr(self.parent, 'npl_transmission_2'):
                params = self.parent.npl_transmission_2
            elif self.current_slot == 3 and hasattr(self.parent, 'npl_transmission_3'):
                params = self.parent.npl_transmission_3

            if params:
                self.a0 = params.get('a0')
                self.a1 = params.get('a1')
                self.a2 = params.get('a2')
                self.a3 = params.get('a3')
                self.a4 = params.get('a4')
                self.b1 = params.get('b1')
                self.b2 = params.get('b2')
                self.b3 = params.get('b3')
                self.b4 = params.get('b4')
                self.min_ke = params.get('min_ke')
                self.max_ke = params.get('max_ke')
            else:
                # Clear parameters if slot is empty
                self.a0 = self.a1 = self.a2 = self.a3 = self.a4 = None
                self.b1 = self.b2 = self.b3 = self.b4 = None
                self.min_ke = self.max_ke = None

            # Update display and plot
            wx.CallAfter(self.update_parameter_display)
            wx.CallAfter(self.plot_transmission_function)

        except Exception as e:
            print(f"Error loading config: {e}")


class VMSFileDropTarget(wx.FileDropTarget):
    def __init__(self, window):
        super().__init__()
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        if filenames and filenames[0].endswith('.vms'):
            self.window.load_vms_file(filenames[0])
            return True
        return False