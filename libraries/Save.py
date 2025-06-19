

from scipy import interpolate
import json
import numpy as np
import wx
# import requests
import base64
import io
import os
import pandas as pd
from libraries.ConfigFile import add_core_level_Data
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl import load_workbook
from libraries.Sheet_Operations import on_sheet_selected
from copy import deepcopy
import shutil
import datetime
# from Functions import convert_to_serializable_and_round


def save_json_only(window):
    """Save only the JSON file with console updates"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Saving Data", size=(300, 200))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 300) // 2,
        parent_pos.y + (parent_size.height - 200) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console("Starting JSON save process...")

        # Save JSON file with entire window.Data
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        update_console("Preparing data for JSON export...")

        # Create a copy of window.Data to modify
        json_data = window.Data.copy()

        update_console("Converting data to serializable format...")
        json_data = convert_to_serializable_and_round(json_data)

        update_console("Writing JSON file...")
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        update_console("JSON file saved successfully!")
        wx.CallLater(500, console_frame.Close)

    except Exception as e:
        update_console(f"Error during save: {str(e)}")
        wx.CallLater(1000, console_frame.Close)
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving JSON: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_all_sheets_with_plots(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Saving All Sheets", size=(300, 350))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 300) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console("Starting save all sheets process...")
        total_sheets = len(window.Data['Core levels'].keys())
        update_console(f"Found {total_sheets} sheets to process")

        for i, sheet_name in enumerate(window.Data['Core levels'].keys(), 1):
            update_console(f"Processing sheet {i}/{total_sheets}: {sheet_name}")

            # Select the sheet
            window.sheet_combobox.SetValue(sheet_name)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, sheet_name)

            # Get fitting data
            update_console(f"Getting fitting data for {sheet_name}")
            fit_data = window.get_data_for_save()

            # Save fitting data to Excel
            update_console(f"Saving fitting data to Excel for {sheet_name}")
            save_to_excel(window, fit_data, file_path, sheet_name, update_console)

            # Save plot to Excel
            update_console(f"Saving plot to Excel for {sheet_name}")
            save_plot_to_excel(window, update_console)

        # Save results table
        update_console("Saving results table...")
        save_results_table(window)

        # Save JSON
        update_console("Saving JSON data...")
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        update_console("All sheets, plots, and results table have been saved successfully!")
        wx.CallLater(500, console_frame.Close)

    except Exception as e:
        update_console(f"Error saving sheets with plots: {str(e)}")
        wx.CallLater(1000, console_frame.Close)
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving sheets with plots: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_data_WorkingFILLED(window, data):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Saving Data", size=(300, 350))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 300) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    try:
        update_console("Starting save process...")

        # Save JSON file with entire window.Data
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        update_console("Preparing data for JSON export...")

        # Create a copy of window.Data to modify
        json_data = window.Data.copy()
        print('Created copy of json Data')

        update_console("Converting data to serializable format...")
        # Convert numpy arrays and other non-serializable types to lists, and round floats
        try:
            json_data = convert_to_serializable_and_round(json_data)
            print('Converted json data to serializable and rounded')
        except Exception as e:
            print(f"Error converting data: {str(e)}")
            for key in window.Data:
                try:
                    convert_to_serializable_and_round(window.Data[key])
                except Exception as sub_e:
                    print(f"Problem in section '{key}': {str(sub_e)}")
            raise

        update_console("Saving JSON file...")
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)
        print('Saved json file')

        # Save to Excel
        update_console(f"Saving fitting data to Excel sheet: {sheet_name}")
        try:
            save_to_excel(window, data, file_path, sheet_name, update_console)
            print('Saved to Excel')
        except Exception as e:
            print(f"Error in save_to_excel: {str(e)}")
            raise

        update_console("Saving plot to Excel...")
        try:
            save_plot_to_excel(window, update_console)
            print('Saved plot to Excel')
        except Exception as e:
            print(f"Error in save_plot_to_excel: {str(e)}")
            raise

        # Save results table
        update_console("Saving results table...")
        try:
            save_results_table(window)
            print('Saved results table')
        except Exception as e:
            print(f"Error in save_results_table: {str(e)}")
            raise

        update_console("Save process completed successfully!")
        wx.CallLater(500, console_frame.Close)
        print("Data Saved")

    except Exception as e:
        update_console(f"Error during save: {str(e)}")
        wx.CallLater(1000, console_frame.Close)
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving data: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_data(window, data):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    # Create console window centered on parent
    parent_pos = window.GetPosition()
    parent_size = window.GetSize()
    console_frame = wx.Frame(window, title="Saving Data", size=(300, 350))
    console_frame.SetPosition((
        parent_pos.x + (parent_size.width - 300) // 2,
        parent_pos.y + (parent_size.height - 350) // 2
    ))
    console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
    console_frame.Show()

    def update_console(message):
        console_text.AppendText(message + '\n')
        console_text.Update()
        wx.SafeYield()

    # Store original peak fill state
    original_fill_state = window.plot_manager.peak_fill_enabled

    try:
        update_console("Starting save process...")

        # FORCE PEAK FILLING for data extraction
        if not original_fill_state:
            update_console("Temporarily enabling peak filling for data extraction...")
            window.plot_manager.peak_fill_enabled = True

            # Force a replot to generate collections
            if hasattr(window, 'clear_and_replot'):
                window.clear_and_replot()
            else:
                window.plot_manager.plot_data(window)
                if window.show_fit:
                    window.plot_manager.plot_fitting_results(window)

        # Now get fresh data with filled peaks (should extract from collections)
        update_console("Extracting peak data (with filling enabled)...")
        fresh_data = window.get_data_for_save()

        # Save JSON file with entire window.Data
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        update_console("Preparing data for JSON export...")

        json_data = window.Data.copy()

        update_console("Converting data to serializable format...")
        try:
            json_data = convert_to_serializable_and_round(json_data)
        except Exception as e:
            for key in window.Data:
                try:
                    convert_to_serializable_and_round(window.Data[key])
                except Exception as sub_e:
                    print(f"Problem in section '{key}': {str(sub_e)}")
            raise

        update_console("Saving JSON file...")
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Save to Excel using fresh_data (with peak data from collections)
        update_console(f"Saving fitting data to Excel sheet: {sheet_name}")
        try:
            save_to_excel(window, fresh_data, file_path, sheet_name, update_console)
        except Exception as e:
            print(f"Error in save_to_excel: {str(e)}")
            raise

        # NOW restore original fill state for plot image
        if not original_fill_state:
            update_console("Restoring unfilled peak display for plot image...")
            window.plot_manager.peak_fill_enabled = False

            # Replot with unfilled peaks for plot image
            if hasattr(window, 'clear_and_replot'):
                window.clear_and_replot()
            else:
                window.plot_manager.plot_data(window)
                if window.show_fit:
                    window.plot_manager.plot_fitting_results(window)

        update_console("Saving plot to Excel...")
        try:
            save_plot_to_excel(window, update_console)
        except Exception as e:
            print(f"Error in save_plot_to_excel: {str(e)}")
            raise

        # Save results table
        update_console("Saving results table...")
        try:
            save_results_table(window)
        except Exception as e:
            print(f"Error in save_results_table: {str(e)}")
            raise

        update_console("Save process completed successfully!")
        wx.CallLater(500, console_frame.Close)

    except Exception as e:
        update_console(f"Error during save: {str(e)}")
        wx.CallLater(1000, console_frame.Close)
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving data: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

    finally:
        # Always restore original state
        window.plot_manager.peak_fill_enabled = original_fill_state

def convert_to_serializable_and_round2(obj, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            return {
                "rows": obj.GetNumberRows(),
                "cols": obj.GetNumberCols(),
                "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), decimal_places)
                          for col in range(obj.GetNumberCols())]
                         for row in range(obj.GetNumberRows())]
            }
        elif hasattr(obj, 'tolist'):  # This catches any object with a tolist method, like some lmfit results
            return convert_to_serializable_and_round(obj.tolist(), decimal_places)
        else:
            return obj
    except Exception as e:
        return str(obj)  # Return a string representation as a fallback

def convert_to_serializable_and_round_OLD(obj, window=None, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            if obj == window.results_grid:
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [{
                        "Label": f"Peak_{row}",
                        "Name": obj.GetCellValue(row, 0),
                        "Position": convert_to_serializable_and_round(obj.GetCellValue(row, 1), decimal_places),
                        "Height": convert_to_serializable_and_round(obj.GetCellValue(row, 2), decimal_places),
                        "FWHM": convert_to_serializable_and_round(obj.GetCellValue(row, 3), decimal_places),
                        "L/G": convert_to_serializable_and_round(obj.GetCellValue(row, 4), decimal_places),
                        "Area": convert_to_serializable_and_round(obj.GetCellValue(row, 5), decimal_places),
                        "at. %": convert_to_serializable_and_round(obj.GetCellValue(row, 6), decimal_places),
                        "Checkbox": obj.GetCellValue(row, 7),
                        "RSF": convert_to_serializable_and_round(obj.GetCellValue(row, 8), decimal_places),
                        "Fitting Model": obj.GetCellValue(row, 9),
                        "Rel. Area": convert_to_serializable_and_round(obj.GetCellValue(row, 10), decimal_places),
                        "Sigma": convert_to_serializable_and_round(obj.GetCellValue(row, 11), decimal_places),
                        "Gamma": convert_to_serializable_and_round(obj.GetCellValue(row, 12), decimal_places),
                        "Bkg Type": obj.GetCellValue(row, 13),
                        "Bkg Low": convert_to_serializable_and_round(obj.GetCellValue(row, 14), decimal_places),
                        "Bkg High": convert_to_serializable_and_round(obj.GetCellValue(row, 15), decimal_places),
                        "Bkg Offset Low": obj.GetCellValue(row, 16),
                        "Bkg Offset High": obj.GetCellValue(row, 17),
                        "Sheetname": obj.GetCellValue(row, 18),
                        "Pos. Constraint": obj.GetCellValue(row, 19),
                        "Height Constraint": obj.GetCellValue(row, 20),
                        "FWHM Constraint": obj.GetCellValue(row, 21),
                        "L/G Constraint": obj.GetCellValue(row, 22),
                        "Area Constraint": obj.GetCellValue(row, 23),
                        "Sigma Constraint": obj.GetCellValue(row, 24),
                        "Gamma Constraint": obj.GetCellValue(row, 25),
                        "Skew Constraint": obj.GetCellValue(row, 26)
                    } for row in range(obj.GetNumberRows())]
                }
            else:
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), decimal_places)
                              for col in range(obj.GetNumberCols())]
                             for row in range(obj.GetNumberRows())]
                }
        elif hasattr(obj, 'tolist'):
            return convert_to_serializable_and_round(obj.tolist(), decimal_places)
        else:
            return obj
    except Exception as e:
        return str(obj)

def convert_to_serializable_and_round(obj, window=None, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            # Check if array is empty
            if obj.size == 0:
                return []
            return [convert_to_serializable_and_round(item, window, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, window, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, window, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            if window and obj == window.results_grid:
                # Your existing results_grid handling code
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [{
                        # Rest of your code
                    } for row in range(obj.GetNumberRows())]
                }
            else:
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), window, decimal_places)
                              for col in range(obj.GetNumberCols())]
                             for row in range(obj.GetNumberRows())]
                }
        elif hasattr(obj, 'tolist'):
            try:
                return convert_to_serializable_and_round(obj.tolist(), window, decimal_places)
            except Exception as e:
                print(f"Error converting with tolist: {e}")
                return str(obj)
        else:
            return obj
    except Exception as e:
        print(f"Error in convert_to_serializable_and_round: {e}, type: {type(obj)}")
        return str(obj)

def convert_to_serializable(obj):
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_serializable(v) for v in obj]
    elif isinstance(obj, (np.int64, np.int32, np.float64, np.float32)):
        return obj.item()
    elif isinstance(obj, wx.grid.Grid):
        return {
            "rows": obj.GetNumberRows(),
            "cols": obj.GetNumberCols(),
            "data": [[obj.GetCellValue(row, col) for col in range(obj.GetNumberCols())]
                     for row in range(obj.GetNumberRows())]
        }
    else:
        return obj

def ensure_sliceable(data, length):
    """Ensure data is sliceable or convert it to a list of the required length."""
    if isinstance(data, (list, np.ndarray)):
        return data[:length]
    else:
        # If data is a scalar (like a float), repeat it to create a list
        return [data] * length

def save_to_excel_OLD(window, data, file_path, sheet_name):
    # Add this function near the beginning of Save.py file
    def safe_get(obj, key, default=None):
        """Safely access dictionary keys or list indices without causing errors."""
        if isinstance(obj, dict) and key in obj:
            return obj[key]
        elif isinstance(obj, (list, tuple, np.ndarray)) and isinstance(key, int) and 0 <= key < len(obj):
            return obj[key]
        return default

    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Rename column if first column is "BE"
    if existing_df.columns[0] == "BE":
        existing_df.rename(columns={"BE": "Binding Energy (eV)"}, inplace=True)

    # Determine the column containing experimental data (if present)
    exp_data_col = None
    exp_data_columns = []

    for col_idx, col_name in enumerate(existing_df.columns):
        if col_name == "Experimental Description":
            exp_data_col = col_idx
            # Collect all columns from this point to the end
            for i in range(col_idx, len(existing_df.columns)):
                exp_data_columns.append({
                    'name': existing_df.columns[i],
                    'data': existing_df.iloc[:, i]
                })
            break

    # Remove previously fitted data if it exists
    if existing_df.shape[1] > 5:
        existing_df = existing_df.iloc[:, :5]

    # Add columns if there aren't enough
    while existing_df.shape[1] < 5:
        existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

    # Ensure there's an empty column E
    if existing_df.shape[1] < 5:
        existing_df.insert(4, '', np.nan)

    if 'x_values' in data and data['x_values'] is not None:
        x_values = data['x_values'].to_numpy() if isinstance(data['x_values'], pd.Series) else data['x_values']

        # Create a new DataFrame with all data we want to add
        new_columns = {'BE': x_values}

        if data['background'] is not None and data['calculated_fit'] is not None:
            mask = np.isin(data['x_values'], x_values)
            y_values = data['y_values'][mask]

            if len(y_values) == len(data['calculated_fit']):
                new_columns['Residuals'] = y_values - data['calculated_fit']

        new_columns['Background'] = data['background'] if data['background'] is not None else np.nan
        new_columns['Calculated Fit'] = data['calculated_fit'] if data['calculated_fit'] is not None else np.nan

        if data['individual_peak_fits']:
            num_rows = len(x_values)
            num_peaks = data['peak_params_grid'].GetNumberRows() // 2
            for i in range(num_peaks):
                row = i * 2
                peak_label = data['peak_params_grid'].GetCellValue(row, 1)
                if i < len(data['individual_peak_fits']):
                    reversed_peak = np.array(data['individual_peak_fits'][i])[::-1]
                    trimmed_peak = np.roll(reversed_peak, -1)[:num_rows]
                    new_columns[peak_label] = trimmed_peak



        # Now insert all columns at position 5
        col_pos = 5
        for col_name, col_data in new_columns.items():
            # Ensure no duplicate column names
            unique_name = col_name
            counter = 1
            while unique_name in existing_df.columns:
                unique_name = f"{col_name}_{counter}"
                counter += 1

            # Insert column ensuring lengths match
            # existing_df.insert(col_pos, unique_name, col_data[:len(existing_df)])
            existing_df.insert(col_pos, unique_name, ensure_sliceable(col_data, len(existing_df)))
            col_pos += 1

        # Ensure there are at least 23 columns (A to W)
        while existing_df.shape[1] < 23:
            existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

        # Rename columns starting with "Unnamed" or "Column" to empty string
        existing_df.columns = ['' if col.startswith(('Unnamed', 'Column')) else col for col in existing_df.columns]

        # Ensure column E is empty
        if existing_df.columns[4] != '':
            existing_df.rename(columns={existing_df.columns[4]: ''}, inplace=True)

        # Create DataFrame for peak fitting parameters
        peak_params_df = pd.DataFrame()
        for col in range(window.peak_params_grid.GetNumberCols()):
            col_name = window.peak_params_grid.GetColLabelValue(col)
            col_data = [window.peak_params_grid.GetCellValue(row, col) for row in
                        range(window.peak_params_grid.GetNumberRows())]
            peak_params_df[col_name] = col_data

        # Add peak_params_df to existing_df starting from column 23 (X)
        for i, col in enumerate(peak_params_df.columns):
            existing_df.insert(23 + i, col, peak_params_df[col])

        # Handle D-parameter derivative data
        if window.selected_fitting_method == "D-parameter":
            if 'Fitting' in window.Data['Core levels'][sheet_name] and 'Peaks' in \
                    window.Data['Core levels'][sheet_name][
                        'Fitting']:
                d_param_data = window.Data['Core levels'][sheet_name]['Fitting']['Peaks'].get(
                    'D-parameter')
                if d_param_data and 'Derivative' in d_param_data:
                    filtered_data['Derivative'] = d_param_data['Derivative']
                    existing_df.insert(7, 'Derivative', d_param_data['Derivative'])

        # Restore experimental data columns if they were present
        if exp_data_columns:
            # Add three separator columns
            existing_df[''] = ''  # First separator column
            existing_df['  '] = ''  # Second separator column with two spaces
            existing_df['   '] = ''  # Third separator column with three spaces

            # Then append the experimental data columns
            for col_info in exp_data_columns:
                col_name = col_info['name']
                # Ensure the column name is unique
                suffix = 1
                original_name = col_name
                while col_name in existing_df.columns:
                    col_name = f"{original_name}_{suffix}"
                    suffix += 1

                existing_df[col_name] = col_info['data']

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Remove border from first row
            workbook = writer.book
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style=None),
                    right=openpyxl.styles.Side(style=None),
                    top=openpyxl.styles.Side(style=None),
                    bottom=openpyxl.styles.Side(style=None)
                )

                # Define styles
                thin_side = Side(style='thin')
                thick_side = Side(style='medium')
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                bold_font = Font(bold=True)

                start_row = 2  # Assuming data starts from the second row
                num_peak_rows = window.peak_params_grid.GetNumberRows()
                end_row = start_row + num_peak_rows - 1
                start_col = 24  # Column X (24th column)
                end_col = min(start_col + window.peak_params_grid.GetNumberCols(), worksheet.max_column+1)
                #print(f"start_col: {start_col}, end_col: {start_col + window.peak_params_grid.GetNumberCols() - 1}
                # OR {worksheet.max_column}")

                for row in range(start_row - 1, end_row + 1):  # Start from header row
                    for col in range(start_col, end_col):
                        if col >= worksheet.max_column+1:
                            print(f"Skipping column {col} as it exceeds max_column {worksheet.max_column}")
                            continue

                        cell = worksheet.cell(row=row, column=col)

                        # Default to thin borders
                        left = right = top = bottom = thin_side

                        # Header row
                        if row == start_row - 1:
                            cell.fill = green_fill
                            cell.font = bold_font
                            top = thick_side

                        # Add thick borders for outer edges
                        if row == start_row - 1 or row == end_row:
                            bottom = thick_side
                        if col == start_col:
                            left = thick_side
                        if col == end_col:
                            right = thick_side

                        # Add thick bottom border for every second row (constraints row)
                        if (row - start_row + 1) % 2 == 0:
                            bottom = thick_side

                        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                # Add citation row
                citation_row = end_row + 1
                citation_text = "Please cite KherveFitting software: Kerherve G. et al. Surface and Interface Analysis (2025) TBD"
                citation_cell = worksheet.cell(row=citation_row, column=start_col, value=citation_text)
                citation_cell.font = Font(italic=True)
                # Merge cells across the width of the peak params table
                worksheet.merge_cells(start_row=citation_row, start_column=start_col,
                                      end_row=citation_row, end_column=end_col)

        # After saving to Excel, update the plot with the current limits
        if hasattr(window, 'plot_config'):
            limits = window.plot_config.get_plot_limits(window, sheet_name)
            window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis
            window.ax.set_ylim(limits['Ymin'], limits['Ymax'])
            window.canvas.draw_idle()


def save_to_excel(window, data, file_path, sheet_name, update_console=None):
    # Add this function near the beginning of Save.py file
    def safe_get(obj, key, default=None):
        """Safely access dictionary keys or list indices without causing errors."""
        if isinstance(obj, dict) and key in obj:
            return obj[key]
        elif isinstance(obj, (list, tuple, np.ndarray)) and isinstance(key, int) and 0 <= key < len(obj):
            return obj[key]
        return default

    if update_console:
        update_console("Reading existing Excel data...")

    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Rename column if first column is "BE"
    if existing_df.columns[0] == "BE":
        existing_df.rename(columns={"BE": "Binding Energy (eV)"}, inplace=True)

    # Determine the column containing experimental data (if present)
    exp_data_col = None
    exp_data_columns = []

    for col_idx, col_name in enumerate(existing_df.columns):
        if col_name == "Experimental Description":
            exp_data_col = col_idx
            # Collect all columns from this point to the end
            for i in range(col_idx, len(existing_df.columns)):
                exp_data_columns.append({
                    'name': existing_df.columns[i],
                    'data': existing_df.iloc[:, i]
                })
            break

    if update_console:
        update_console("Processing fitting data...")

    # Remove previously fitted data if it exists
    if existing_df.shape[1] > 5:
        existing_df = existing_df.iloc[:, :5]

    # Add columns if there aren't enough
    while existing_df.shape[1] < 5:
        existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

    # Ensure there's an empty column E
    if existing_df.shape[1] < 5:
        existing_df.insert(4, '', np.nan)

    if 'x_values' in data and data['x_values'] is not None:
        if update_console:
            update_console("Adding fitted peaks and background data...")

        x_values = data['x_values'].to_numpy() if isinstance(data['x_values'], pd.Series) else data['x_values']

        # Create a new DataFrame with all data we want to add
        new_columns = {'BE': x_values}

        if data['background'] is not None and data['calculated_fit'] is not None:
            mask = np.isin(data['x_values'], x_values)
            y_values = data['y_values'][mask]

            if len(y_values) == len(data['calculated_fit']):
                new_columns['Residuals'] = y_values - data['calculated_fit']

        new_columns['Background'] = data['background'] if data['background'] is not None else np.nan
        new_columns['Calculated Fit'] = data['calculated_fit'] if data['calculated_fit'] is not None else np.nan

        if data['individual_peak_fits']:
            if update_console:
                update_console(f"Adding {len(data['individual_peak_fits'])} individual peak fits...")

            num_rows = len(x_values)
            num_peaks = data['peak_params_grid'].GetNumberRows() // 2
            for i in range(num_peaks):
                row = i * 2
                peak_label = data['peak_params_grid'].GetCellValue(row, 1)
                if i < len(data['individual_peak_fits']):
                    reversed_peak = np.array(data['individual_peak_fits'][i])[::-1]
                    trimmed_peak = np.roll(reversed_peak, -1)[:num_rows]
                    new_columns[peak_label] = trimmed_peak

        # Now insert all columns at position 5
        col_pos = 5
        for col_name, col_data in new_columns.items():
            # Ensure no duplicate column names
            unique_name = col_name
            counter = 1
            while unique_name in existing_df.columns:
                unique_name = f"{col_name}_{counter}"
                counter += 1

            # Insert column ensuring lengths match
            existing_df.insert(col_pos, unique_name, ensure_sliceable(col_data, len(existing_df)))
            col_pos += 1

        # Ensure there are at least 23 columns (A to W)
        while existing_df.shape[1] < 23:
            existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

        # Rename columns starting with "Unnamed" or "Column" to empty string
        existing_df.columns = ['' if col.startswith(('Unnamed', 'Column')) else col for col in existing_df.columns]

        # Ensure column E is empty
        if existing_df.columns[4] != '':
            existing_df.rename(columns={existing_df.columns[4]: ''}, inplace=True)

        if update_console:
            update_console("Adding peak parameters...")

        # Create DataFrame for peak fitting parameters
        peak_params_df = pd.DataFrame()
        for col in range(window.peak_params_grid.GetNumberCols()):
            col_name = window.peak_params_grid.GetColLabelValue(col)
            col_data = [window.peak_params_grid.GetCellValue(row, col) for row in
                        range(window.peak_params_grid.GetNumberRows())]
            peak_params_df[col_name] = col_data

        # Add peak_params_df to existing_df starting from column 23 (X)
        for i, col in enumerate(peak_params_df.columns):
            existing_df.insert(23 + i, col, peak_params_df[col])

        # Handle D-parameter derivative data
        if window.selected_fitting_method == "D-parameter":
            if 'Fitting' in window.Data['Core levels'][sheet_name] and 'Peaks' in \
                    window.Data['Core levels'][sheet_name]['Fitting']:
                d_param_data = window.Data['Core levels'][sheet_name]['Fitting']['Peaks'].get('D-parameter')
                if d_param_data and 'Derivative' in d_param_data:
                    existing_df.insert(7, 'Derivative', d_param_data['Derivative'])

        # Restore experimental data columns if they were present
        if exp_data_columns:
            # Add three separator columns
            existing_df[''] = ''
            existing_df['  '] = ''
            existing_df['   '] = ''

            # Then append the experimental data columns
            for col_info in exp_data_columns:
                col_name = col_info['name']
                suffix = 1
                original_name = col_name
                while col_name in existing_df.columns:
                    col_name = f"{original_name}_{suffix}"
                    suffix += 1
                existing_df[col_name] = col_info['data']

        if update_console:
            update_console("Writing data to Excel file...")

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            existing_df.to_excel(writer, sheet_name=sheet_name, index=False)


            # Remove border from first row
            workbook = writer.book
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style=None),
                    right=openpyxl.styles.Side(style=None),
                    top=openpyxl.styles.Side(style=None),
                    bottom=openpyxl.styles.Side(style=None)
                )

                # Define styles
                thin_side = Side(style='thin')
                thick_side = Side(style='medium')
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                bold_font = Font(bold=True)

                start_row = 2  # Assuming data starts from the second row
                num_peak_rows = window.peak_params_grid.GetNumberRows()
                end_row = start_row + num_peak_rows - 1
                start_col = 24  # Column X (24th column)
                end_col = min(start_col + window.peak_params_grid.GetNumberCols(), worksheet.max_column+1)
                #print(f"start_col: {start_col}, end_col: {start_col + window.peak_params_grid.GetNumberCols() - 1}
                # OR {worksheet.max_column}")

                for row in range(start_row - 1, end_row + 1):  # Start from header row
                    for col in range(start_col, end_col):
                        if col >= worksheet.max_column+1:
                            print(f"Skipping column {col} as it exceeds max_column {worksheet.max_column}")
                            continue

                        cell = worksheet.cell(row=row, column=col)

                        # Default to thin borders
                        left = right = top = bottom = thin_side

                        # Header row
                        if row == start_row - 1:
                            cell.fill = green_fill
                            cell.font = bold_font
                            top = thick_side

                        # Add thick borders for outer edges
                        if row == start_row - 1 or row == end_row:
                            bottom = thick_side
                        if col == start_col:
                            left = thick_side
                        if col == end_col:
                            right = thick_side

                        # Add thick bottom border for every second row (constraints row)
                        if (row - start_row + 1) % 2 == 0:
                            bottom = thick_side

                        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                # Add citation row
                citation_row = end_row + 1
                citation_text = "Please cite KherveFitting software: Kerherve G. et al. Surface and Interface Analysis (2025) TBD"
                citation_cell = worksheet.cell(row=citation_row, column=start_col, value=citation_text)
                citation_cell.font = Font(italic=True)
                # Merge cells across the width of the peak params table
                worksheet.merge_cells(start_row=citation_row, start_column=start_col,
                                      end_row=citation_row, end_column=end_col)

        # After saving to Excel, update the plot with the current limits
        if hasattr(window, 'plot_config'):
            limits = window.plot_config.get_plot_limits(window, sheet_name)
            window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis
            window.ax.set_ylim(limits['Ymin'], limits['Ymax'])
            window.canvas.draw_idle()


def refresh_sheets_OLD(window, on_sheet_selected_func):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file currently open. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    current_sheet = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        import re
        # Save BEcorrections data
        be_corrections = window.Data.get('BEcorrections', {}).copy() if 'BEcorrections' in window.Data else {}

        # Save current state to JSON
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Reopen the XLSX file
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names

        # Filter out "Results Table" and "Experimental Description" sheets
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        # Normalize sheet names
        name_changes = {}
        for old_name in sheet_names:
            new_name = old_name

            lower_name = old_name.lower()
            if 'survey' in lower_name:
                new_name = 'Survey'
            elif 'xps survey' in lower_name:
                    new_name = 'Survey'
            elif 'survey scan' in lower_name:
                    new_name = 'Survey'
            elif 'xps' in lower_name:
                    new_name = 'Survey'
            elif 'wide' in lower_name:
                new_name = 'Wide'
            elif 'wide scan' in lower_name:
                new_name = 'Wide'
            else:
                # Remove spaces between element and orbital (e.g., "C 1s" â†’ "C1s")
                match = re.search(r'([A-Z][a-z]?)\s+(\d+[spdf])', old_name)
                if match:
                    element, orbital = match.groups()
                    new_name = f"{element}{orbital}"

                # Simplify names like "C1s Scan" to just "C1s"
                match = re.search(r'([A-Z][a-z]?\d+[spdf])', new_name)
                if match and len(new_name) > len(match.group(1)):
                    new_name = match.group(1)

            # Preserve sample number suffix if it exists
            suffix_match = re.search(r'(\d+)$', old_name)
            if suffix_match and not re.search(r'\d+$', new_name):
                new_name = f"{new_name}{suffix_match.group(1)}"

            if new_name != old_name:
                name_changes[old_name] = new_name

        # Rename sheets in Excel file if needed
        if name_changes:
            wb = openpyxl.load_workbook(file_path)
            for old_name, new_name in name_changes.items():
                if old_name in wb.sheetnames and new_name not in wb.sheetnames:
                    sheet = wb[old_name]
                    sheet.title = new_name
            wb.save(file_path)

            # Reopen Excel file to get updated sheet names
            excel_file = pd.ExcelFile(file_path)
            all_sheet_names = excel_file.sheet_names
            sheet_names = [name for name in all_sheet_names if
                           name.lower() not in ["results table", "experimental description"]]

            # Update window.Data with new sheet names
            updated_core_levels = {}
            for old_name, new_name in name_changes.items():
                if old_name in window.Data['Core levels']:
                    updated_core_levels[new_name] = window.Data['Core levels'][old_name]

                    # Update plot_config data structures
                    if hasattr(window, 'plot_config'):
                        # Update plot_limits
                        if old_name in window.plot_config.plot_limits:
                            window.plot_config.plot_limits[new_name] = window.plot_config.plot_limits.pop(old_name)

                        # Update original_limits - THIS IS THE FIX
                        if hasattr(window.plot_config,
                                   'original_limits') and old_name in window.plot_config.original_limits:
                            window.plot_config.original_limits[new_name] = window.plot_config.original_limits.pop(
                                old_name)

            # Add any remaining sheets that weren't renamed
            for name in window.Data['Core levels']:
                if name not in name_changes and name not in updated_core_levels:
                    updated_core_levels[name] = window.Data['Core levels'][name]

            window.Data['Core levels'] = updated_core_levels

        # Update sheet names in the combobox
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)

        # Update window.Data with new sheet information
        for sheet_name in sheet_names:
            if sheet_name not in window.Data['Core levels']:
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        # Remove any sheets from window.Data that no longer exist in the Excel file
        sheets_to_remove = set(window.Data['Core levels'].keys()) - set(sheet_names)
        for sheet_name in sheets_to_remove:
            del window.Data['Core levels'][sheet_name]

        # Update the number of core levels
        window.Data['Number of Core levels'] = len(sheet_names)

        # Update B.E. and Raw Data with current Excel data
        import re
        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            raw_be_values = df.iloc[:, 0].tolist()

            # Get BE correction for this sheet
            sheet_correction = 0  # Default if no correction found

            # Extract row number from sheet name
            match = re.search(r'(\d+)$', sheet_name)

            if match and be_corrections:
                # Sheet name ends with a number
                sample_row = match.group(1)
                if sample_row in be_corrections:
                    sheet_correction = be_corrections[sample_row]
            elif be_corrections and "0" in be_corrections:
                # Sheet name doesn't end with a number - treat as row "0"
                sheet_correction = be_corrections["0"]

            # Apply the sheet-specific BE correction
            window.Data['Core levels'][sheet_name]['B.E.'] = [be + sheet_correction for be in raw_be_values]
            window.Data['Core levels'][sheet_name]['Raw Data'] = df.iloc[:, 1].tolist()

        # Restore BE corrections data
        if be_corrections:
            window.Data['BEcorrections'] = be_corrections

        # Handle current sheet selection after normalization
        if current_sheet in name_changes:
            current_sheet = name_changes[current_sheet]

        # Set the current sheet as selected if it still exists, otherwise select the first sheet
        if current_sheet in sheet_names:
            window.sheet_combobox.SetValue(current_sheet)
        elif sheet_names:
            window.sheet_combobox.SetValue(sheet_names[0])
            current_sheet = sheet_names[0]

        # Update the spinbox value for the current sheet
        if current_sheet:
            match = re.search(r'(\d+)$', current_sheet)
            if match and be_corrections and match.group(1) in be_corrections:
                window.be_correction = be_corrections[match.group(1)]
            elif be_corrections and "0" in be_corrections:
                # For sheets without a number suffix, use correction from row "0"
                window.be_correction = be_corrections["0"]
            else:
                window.be_correction = 0

            window.be_correction_spinbox.SetValue(window.be_correction)

        # Initialize plot limits for new sheets
        for sheet_name in sheet_names:
            if hasattr(window, 'plot_config') and sheet_name not in window.plot_config.plot_limits:
                # This will update both plot_limits and original_limits
                window.plot_config.update_plot_limits(window, sheet_name)

        # Update the plot for the current sheet
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(current_sheet)
        on_sheet_selected_func(window, event)

        # Update plot limits for the current sheet
        if hasattr(window, 'plot_config'):
            window.plot_config.update_plot_limits(window, current_sheet)

        # Refresh the plot
        window.plot_manager.plot_data(window)
        window.clear_and_replot()

        # If there were any name changes, show them in the message
        if name_changes:
            changes_msg = "\n".join([f"{old} â†’ {new}" for old, new in name_changes.items()])
            # wx.MessageBox(f"Sheets refreshed and normalized:\n{changes_msg}\n\nTotal sheets: {len(sheet_names)}",
            #               "Success", wx.OK | wx.ICON_INFORMATION)
            window.show_popup_message2(f"Sheets refreshed and normalized:\n{changes_msg}\n\nTotal sheets: {len(sheet_names)}", "Success")
        else:
            # wx.MessageBox(f"Sheets refreshed. Total sheets: {len(sheet_names)}", "Success", wx.OK | wx.ICON_INFORMATION)
            window.show_popup_message2(
                f"Sheets refreshed. Total sheets: {len(sheet_names)}", "Success")

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error refreshing sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def refresh_sheets(window, on_sheet_selected_func, update_console=None):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file currently open. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    # Create console if not provided
    console_frame = None
    if update_console is None:
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Refreshing Sheets", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

    current_sheet = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        import re
        update_console("Starting sheet refresh...")

        # Save BEcorrections data
        be_corrections = window.Data.get('BEcorrections', {}).copy() if 'BEcorrections' in window.Data else {}

        # Save current state to JSON
        update_console("Saving current state to JSON...")
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Reopen the XLSX file
        update_console("Reading Excel file structure...")
        excel_file = pd.ExcelFile(file_path)
        all_sheet_names = excel_file.sheet_names

        # Filter out "Results Table" and "Experimental Description" sheets
        sheet_names = [name for name in all_sheet_names if
                       name.lower() not in ["results table", "experimental description"]]

        # Normalize sheet names
        name_changes = {}
        for old_name in sheet_names:
            new_name = old_name

            lower_name = old_name.lower()
            if 'survey' in lower_name:
                new_name = 'Survey'
            elif 'xps survey' in lower_name:
                new_name = 'Survey'
            elif 'survey scan' in lower_name:
                new_name = 'Survey'
            elif 'xps' in lower_name:
                new_name = 'Survey'
            elif 'wide' in lower_name:
                new_name = 'Wide'
            elif 'wide scan' in lower_name:
                new_name = 'Wide'
            else:
                # Remove spaces between element and orbital (e.g., "C 1s" â†’ "C1s")
                match = re.search(r'([A-Z][a-z]?)\s+(\d+[spdf])', old_name)
                if match:
                    element, orbital = match.groups()
                    new_name = f"{element}{orbital}"

                # Simplify names like "C1s Scan" to just "C1s"
                match = re.search(r'([A-Z][a-z]?\d+[spdf])', new_name)
                if match and len(new_name) > len(match.group(1)):
                    new_name = match.group(1)

            # Preserve sample number suffix if it exists
            suffix_match = re.search(r'(\d+)$', old_name)
            if suffix_match and not re.search(r'\d+$', new_name):
                new_name = f"{new_name}{suffix_match.group(1)}"

            if new_name != old_name:
                name_changes[old_name] = new_name

        # Rename sheets in Excel file if needed
        if name_changes:
            update_console("Normalizing sheet names...")
            wb = openpyxl.load_workbook(file_path)
            for old_name, new_name in name_changes.items():
                if old_name in wb.sheetnames and new_name not in wb.sheetnames:
                    sheet = wb[old_name]
                    sheet.title = new_name
            wb.save(file_path)

            # Reopen Excel file to get updated sheet names
            excel_file = pd.ExcelFile(file_path)
            all_sheet_names = excel_file.sheet_names
            sheet_names = [name for name in all_sheet_names if
                           name.lower() not in ["results table", "experimental description"]]

            # Update window.Data with new sheet names
            updated_core_levels = {}
            for old_name, new_name in name_changes.items():
                if old_name in window.Data['Core levels']:
                    updated_core_levels[new_name] = window.Data['Core levels'][old_name]

                    # Update plot_config data structures
                    if hasattr(window, 'plot_config'):
                        # Update plot_limits
                        if old_name in window.plot_config.plot_limits:
                            window.plot_config.plot_limits[new_name] = window.plot_config.plot_limits.pop(old_name)

                        # Update original_limits - THIS IS THE FIX
                        if hasattr(window.plot_config,
                                   'original_limits') and old_name in window.plot_config.original_limits:
                            window.plot_config.original_limits[new_name] = window.plot_config.original_limits.pop(
                                old_name)

            # Add any remaining sheets that weren't renamed
            for name in window.Data['Core levels']:
                if name not in name_changes and name not in updated_core_levels:
                    updated_core_levels[name] = window.Data['Core levels'][name]

            window.Data['Core levels'] = updated_core_levels

        # Update sheet names in the combobox
        update_console("Updating interface...")
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)

        # Update window.Data with new sheet information
        for sheet_name in sheet_names:
            if sheet_name not in window.Data['Core levels']:
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        # Remove any sheets from window.Data that no longer exist in the Excel file
        sheets_to_remove = set(window.Data['Core levels'].keys()) - set(sheet_names)
        for sheet_name in sheets_to_remove:
            del window.Data['Core levels'][sheet_name]

        # Update the number of core levels
        window.Data['Number of Core levels'] = len(sheet_names)

        # Update B.E. and Raw Data with current Excel data
        update_console("Updating binding energy corrections...")
        import re
        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            raw_be_values = df.iloc[:, 0].tolist()

            # Get BE correction for this sheet
            sheet_correction = 0  # Default if no correction found

            # Extract row number from sheet name
            match = re.search(r'(\d+)$', sheet_name)

            if match and be_corrections:
                # Sheet name ends with a number
                sample_row = match.group(1)
                if sample_row in be_corrections:
                    sheet_correction = be_corrections[sample_row]
            elif be_corrections and "0" in be_corrections:
                # Sheet name doesn't end with a number - treat as row "0"
                sheet_correction = be_corrections["0"]

            # Apply the sheet-specific BE correction
            window.Data['Core levels'][sheet_name]['B.E.'] = [be + sheet_correction for be in raw_be_values]
            window.Data['Core levels'][sheet_name]['Raw Data'] = df.iloc[:, 1].tolist()

        # Restore BE corrections data
        if be_corrections:
            window.Data['BEcorrections'] = be_corrections

        # Handle current sheet selection after normalization
        if current_sheet in name_changes:
            current_sheet = name_changes[current_sheet]

        # Set the current sheet as selected if it still exists, otherwise select the first sheet
        if current_sheet in sheet_names:
            window.sheet_combobox.SetValue(current_sheet)
        elif sheet_names:
            window.sheet_combobox.SetValue(sheet_names[0])
            current_sheet = sheet_names[0]

        # Update the spinbox value for the current sheet
        if current_sheet:
            match = re.search(r'(\d+)$', current_sheet)
            if match and be_corrections and match.group(1) in be_corrections:
                window.be_correction = be_corrections[match.group(1)]
            elif be_corrections and "0" in be_corrections:
                # For sheets without a number suffix, use correction from row "0"
                window.be_correction = be_corrections["0"]
            else:
                window.be_correction = 0

            window.be_correction_spinbox.SetValue(window.be_correction)

        # Initialize plot limits for new sheets
        for sheet_name in sheet_names:
            if hasattr(window, 'plot_config') and sheet_name not in window.plot_config.plot_limits:
                # This will update both plot_limits and original_limits
                window.plot_config.update_plot_limits(window, sheet_name)

        # Update the plot for the current sheet
        update_console("Updating plots...")
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(current_sheet)
        on_sheet_selected_func(window, event)

        # Update plot limits for the current sheet
        if hasattr(window, 'plot_config'):
            window.plot_config.update_plot_limits(window, current_sheet)

        # Refresh the plot
        window.plot_manager.plot_data(window)
        window.clear_and_replot()

        # Show completion message in console
        if name_changes:
            changes_msg = ", ".join([f"{old} â†’ {new}" for old, new in name_changes.items()])
            update_console(f"Sheets refreshed and normalized: {changes_msg}")
            update_console(f"Total sheets: {len(sheet_names)}")
        else:
            update_console(f"Sheets refreshed. Total sheets: {len(sheet_names)}")

        update_console("Refresh completed successfully!")

        # Close console if we created it
        if console_frame:
            wx.CallLater(500, console_frame.Close)

    except Exception as e:
        import traceback
        traceback.print_exc()
        update_console(f"Error refreshing sheets: {str(e)}")
        if console_frame:
            wx.CallLater(1000, console_frame.Close)
        wx.MessageBox(f"Error refreshing sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)

def save_plot_only_to_excel(window):
    """Save only the plot to Excel without saving any other data"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    try:
        save_plot_to_excel(window)#, update_console)
        # window.show_popup_message2("Success", "Plot saved to Excel file")
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot to Excel: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_to_json(window, file_path):
    json_file_path = os.path.splitext(file_path)[0] + '.json'

    # Create a copy of window.Data to modify
    data_to_save = window.Data.copy()

    def convert_to_list(item):
        if isinstance(item, np.ndarray):
            return item.tolist()
        elif isinstance(item, list):
            return item
        else:
            return item  # Return as is if it's neither numpy array nor list

    # Convert numpy arrays to lists for JSON serialization
    for sheet_name, sheet_data in data_to_save['Core levels'].items():
        sheet_data['B.E.'] = convert_to_list(sheet_data['B.E.'])
        sheet_data['Raw Data'] = convert_to_list(sheet_data['Raw Data'])
        sheet_data['Background']['Bkg Y'] = convert_to_list(sheet_data['Background']['Bkg Y'])

        if 'Fitting' in sheet_data and 'Peaks' in sheet_data['Fitting']:
            for peak_label, peak_data in sheet_data['Fitting']['Peaks'].items():
                if 'Y' in peak_data:
                    peak_data['Y'] = convert_to_list(peak_data['Y'])

    with open(json_file_path, 'w') as json_file:
        json.dump(data_to_save, json_file, indent=2)


def save_plot_to_excel_OLD(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()
    is_survey = "survey" in sheet_name.lower() or "wide" in sheet_name.lower()
    is_raman = sheet_name.startswith('RA') or 'RAMAN' in sheet_name.upper() or "Ra_" in sheet_name

    try:
        # Get dimensions based on plot type
        width = window.survey_excel_width if is_survey else window.excel_width
        height = window.survey_excel_height if is_survey else window.excel_height
        dpi = window.survey_excel_dpi if is_survey else window.excel_dpi

        if is_raman:
            # Set proper axis orientation after saving
            limits = window.plot_config.get_plot_limits(window, sheet_name)
            window.ax.set_xlim(limits['Xmin'], limits['Xmax'])  # Normal direction for Raman

        # Save figure to buffer
        buf = io.BytesIO()
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(width, height)

        # Save the figure
        window.figure.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)
        buf.seek(0)

        # Save to Excel
        wb = openpyxl.load_workbook(file_path)
        ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]

        # Clear existing images
        for img in ws._images:
            ws._images.remove(img)

        # Add new image
        img = Image(buf)
        ws.add_image(img, 'D6')
        wb.save(file_path)

        print(f"Plot saved to Excel file: {file_path}, Sheet: {sheet_name}")
        window.show_popup_message2("Plot saved into Excel file", f"Under sheet: {sheet_name}")

        # Set proper axis orientation after saving
        limits = window.plot_config.get_plot_limits(window, sheet_name)
        if is_raman:
            window.ax.set_xlim(limits['Xmin'], limits['Xmax'])  # Normal direction for Raman
        else:
            window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis for XPS
        window.canvas.draw_idle()

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot to Excel: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_to_excel(window, update_console=None):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        if update_console:
            update_console("Error: No file selected")
        else:
            wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()
    is_survey = "survey" in sheet_name.lower() or "wide" in sheet_name.lower()
    is_raman = sheet_name.startswith('RA') or 'RAMAN' in sheet_name.upper() or "Ra_" in sheet_name

    try:
        if update_console:
            update_console(f"Saving plot for sheet: {sheet_name}")

        # Get dimensions based on plot type
        width = window.survey_excel_width if is_survey else window.excel_width
        height = window.survey_excel_height if is_survey else window.excel_height
        dpi = window.survey_excel_dpi if is_survey else window.excel_dpi

        if is_raman:
            # Set proper axis orientation after saving
            limits = window.plot_config.get_plot_limits(window, sheet_name)
            window.ax.set_xlim(limits['Xmin'], limits['Xmax'])  # Normal direction for Raman

        # Save figure to buffer
        buf = io.BytesIO()
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(width, height)

        # Save the figure
        window.figure.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)
        buf.seek(0)

        # Save to Excel
        wb = openpyxl.load_workbook(file_path)
        ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]

        # Clear existing images
        for img in ws._images:
            ws._images.remove(img)

        # Add new image
        img = Image(buf)
        ws.add_image(img, 'D6')
        wb.save(file_path)

        print(f"Plot saved to Excel file: {file_path}, Sheet: {sheet_name}")

        if update_console:
            update_console(f"Plot saved to Excel file under sheet: {sheet_name}")
        else:
            window.show_popup_message2("Plot saved into Excel file", f"Under sheet: {sheet_name}")

        # Set proper axis orientation after saving
        limits = window.plot_config.get_plot_limits(window, sheet_name)
        if is_raman:
            window.ax.set_xlim(limits['Xmin'], limits['Xmax'])  # Normal direction for Raman
        else:
            window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis for XPS
        window.canvas.draw_idle()

    except Exception as e:
        import traceback
        traceback.print_exc()
        error_msg = f"Error saving plot to Excel: {str(e)}"
        if update_console:
            update_console(error_msg)
        else:
            wx.MessageBox(error_msg, "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_png(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate PNG filename based on the Excel filename and sheet name
        png_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.png"
        png_filepath = os.path.join(os.path.dirname(file_path), png_filename)

        # Save the figure as PNG
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(png_filepath, format='png', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as PNG: {png_filepath}")
        window.show_popup_message2("Plot saved as PNG", f"File: {png_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as PNG: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_pdf(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate PDF filename based on the Excel filename and sheet name
        pdf_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.pdf"
        pdf_filepath = os.path.join(os.path.dirname(file_path), pdf_filename)

        # Save the figure as PDF
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(pdf_filepath, format='pdf', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as PDF: {pdf_filepath}")
        window.show_popup_message2("Plot saved as PDF", f"File: {pdf_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as PDF: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_svg(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate SVG filename based on the Excel filename and sheet name
        svg_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.svg"
        svg_filepath = os.path.join(os.path.dirname(file_path), svg_filename)

        # Save the figure as SVG
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(svg_filepath, format='svg', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as SVG: {svg_filepath}")
        window.show_popup_message2("Plot saved as SVG", f"File: {svg_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as SVG: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)




def save_results_table(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    json_file_path = os.path.splitext(file_path)[0] + '.json'

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet_name = 'Results Table'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        def get_column_letter(n):
            result = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                result = chr(65 + remainder) + result
            return result

        # Get headers from results grid
        headers = [window.results_grid.GetColLabelValue(col) for col in range(window.results_grid.GetNumberCols())]

        # Set styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # Current row in the worksheet
        current_row = 2

        # Identify result table keys in window.Data
        result_table_keys = sorted([k for k in window.Data.keys() if k.startswith('Results Table')],
                                   key=lambda x: int(x.replace('Results Table', '')) if x.replace('Results Table',
                                                                                                  '').isdigit() else 0)

        # If no Result Tables found, fallback to old format
        if not result_table_keys and 'Results' in window.Data:
            result_table_keys = ['Results']

        # Keep track of tables to remove
        tables_to_remove = []

        for table_key in result_table_keys:
            # Check if the table has any peaks
            if 'Peak' not in window.Data[table_key] or not window.Data[table_key]['Peak']:
                tables_to_remove.append(table_key)
                continue

            # Get sample number
            sample_num = table_key.replace('Results Table', '') if table_key != 'Results' else '0'

            # Write sample title
            title_cell = ws.cell(row=current_row, column=2, value=f"Sample {sample_num}")
            title_cell.font = Font(bold=True, size=12)
            current_row += 1

            # Write headers
            header_row = current_row
            for col, header in enumerate(headers, start=2):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Get peak data for this table
            peaks_data = window.Data[table_key]['Peak']

            # Sort peaks by their numeric index
            sorted_peaks = []
            for peak_key, peak_data in peaks_data.items():
                if peak_key.startswith('Peak_') and peak_key[5:].isdigit():
                    index = int(peak_key[5:])
                    sorted_peaks.append((index, peak_data))

            sorted_peaks.sort(key=lambda x: x[0])

            # Write peak data
            data_start_row = header_row + 1
            for i, (_, peak_data) in enumerate(sorted_peaks):
                row_num = data_start_row + i

                # Map fields to columns
                field_map = {
                    0: 'Name',  # Peak Label
                    1: 'Position',  # Position
                    2: 'Height',  # Height
                    3: 'FWHM',  # FWHM
                    4: 'L/G',  # L/G
                    5: 'Area',  # Area
                    6: 'at. %',  # Atomic %
                    7: 'Checkbox',  # Checkbox
                    8: 'RSF',  # RSF
                    9: 'TXFN',  # TXFN
                    10: 'ECF',  # ECF
                    11: 'Instrument',  # Instrument
                    12: 'Fitting Model',  # Fitting Model
                    13: 'Rel. Area',  # Rel Area
                    14: 'Sigma',  # Sigma
                    15: 'Gamma',  # Gamma
                    16: 'Bkg Type',  # Bkg Type
                    17: 'Bkg Low',  # Bkg Low
                    18: 'Bkg High',  # Bkg High
                    19: 'Bkg Offset Low',  # Bkg Offset Low
                    20: 'Bkg Offset High',  # Bkg Offset High
                    21: 'Sheetname',  # Sheetname
                    22: 'Pos. Constraint',  # Position Constraint
                    23: 'Height Constraint',  # Height Constraint
                    24: 'FWHM Constraint',  # FWHM Constraint
                    25: 'L/G Constraint',  # L/G Constraint
                    26: 'Area Constraint',  # Area Constraint
                    27: 'Sigma Constraint',  # Sigma Constraint
                    28: 'Gamma Constraint'  # Gamma Constraint
                }

                # Write each cell
                for col_idx in range(len(headers)):
                    field_name = field_map.get(col_idx, f"Column_{col_idx}")
                    value = peak_data.get(field_name, "")

                    # Format checkbox
                    if col_idx == 7:  # Checkbox column
                        value = 'âœ“' if value == '1' else ''

                    # Format numeric values
                    if isinstance(value, (int, float)) and col_idx in [1, 2, 3, 4, 5, 6, 13, 14, 15]:
                        value = f"{value:.2f}"

                    cell = ws.cell(row=row_num, column=col_idx + 2, value=value)
                    cell.alignment = center_align

            # Calculate last row of this table
            data_end_row = data_start_row + len(sorted_peaks) - 1 if sorted_peaks else data_start_row

            # Apply borders
            max_col = len(headers) + 1

            # Headers row - thick border
            for col in range(2, max_col + 2):
                cell = ws.cell(row=header_row, column=col)
                cell.border = Border(
                    left=Side(style='medium' if col == 2 else 'thin'),
                    right=Side(style='medium' if col == max_col + 1 else 'thin'),
                    top=Side(style='medium'),
                    bottom=Side(style='thin')
                )

            # Data cells
            for row in range(data_start_row, data_end_row + 1):
                for col in range(2, max_col + 2):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(
                        left=Side(style='medium' if col == 2 else 'thin'),
                        right=Side(style='medium' if col == max_col + 1 else 'thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium' if row == data_end_row else 'thin')
                    )

            # Move to next table position (leave gap)
            current_row = data_end_row + 4

        # Remove empty tables from window.Data
        for table_key in tables_to_remove:
            del window.Data[table_key]

        # Adjust column widths
        for column_cells in ws.columns:
            if column_cells:
                length = max(len(str(cell.value) or '') for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save Excel file
        wb.save(file_path)

        # Save JSON file with updated window.Data (after empty tables removal)
        from libraries.Save import convert_to_serializable_and_round
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving results table: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_state(window):
    current_sheet = window.sheet_combobox.GetValue()
    sheets_data = {}

    for sheet in window.Data['Core levels'].keys():
        if sheet == current_sheet:
            sheets_data[sheet] = {
                'peak_params_grid': get_grid_data(window.peak_params_grid),
                'peak_count': window.peak_count,
                'selected_peak_index': window.selected_peak_index
            }

    state = {
        'Data': deepcopy(window.Data),
        'current_sheet': current_sheet,
        'sheets': sheets_data,
        'results_grid': {
            'data': get_grid_data(window.results_grid),
            'num_rows': window.results_grid.GetNumberRows(),
            'checkbox_states': [window.results_grid.GetCellValue(row, 7)
                              for row in range(window.results_grid.GetNumberRows())]
        },
        'be_correction': window.be_correction  # Add BE correction to state
    }

    window.history.append(state)
    if len(window.history) > window.max_history:
        window.history.pop(0)
    window.redo_stack.clear()
    update_undo_redo_state(window)

def undo(window):
    if len(window.history) > 1:
        current_state = window.history.pop()
        window.redo_stack.append(current_state)
        previous_state = window.history[-1]
        restore_state(window, previous_state)
        update_undo_redo_state(window)

def redo(window):
    if window.redo_stack:
        next_state = window.redo_stack.pop()
        window.history.append(next_state)
        restore_state(window, next_state)
        update_undo_redo_state(window)


def restore_state(window, state):
    window.Data = deepcopy(state['Data'])
    window.be_correction = state.get('be_correction', 0)  # Restore BE correction with default 0
    window.be_correction_spinbox.SetValue(window.be_correction)

    # Restore sheet-specific data
    for sheet, sheet_data in state['sheets'].items():
        if sheet == state['current_sheet']:
            set_grid_data(window.peak_params_grid, sheet_data['peak_params_grid'])
            window.peak_count = sheet_data['peak_count']
            window.selected_peak_index = sheet_data['selected_peak_index']

    # Restore results grid
    results_grid_data = state['results_grid']
    current_rows = window.results_grid.GetNumberRows()
    target_rows = results_grid_data['num_rows']

    if current_rows < target_rows:
        window.results_grid.AppendRows(target_rows - current_rows)
    elif current_rows > target_rows:
        window.results_grid.DeleteRows(target_rows, current_rows - target_rows)

    set_grid_data(window.results_grid, results_grid_data['data'])

    # Restore checkbox states
    for row, checkbox_state in enumerate(results_grid_data['checkbox_states']):
        window.results_grid.SetCellRenderer(row, 7, wx.grid.GridCellBoolRenderer())
        window.results_grid.SetCellEditor(row, 7, wx.grid.GridCellBoolEditor())
        window.results_grid.SetCellValue(row, 7, checkbox_state)

    # Switch to the correct sheet
    window.sheet_combobox.SetValue(state['current_sheet'])
    on_sheet_selected(window, state['current_sheet'])

    window.clear_and_replot()

def get_grid_data(grid):
    data = []
    for row in range(grid.GetNumberRows()):
        row_data = [grid.GetCellValue(row, col) for col in range(grid.GetNumberCols())]
        data.append(row_data)
    return data

def set_grid_data(grid, data):
    grid.ClearGrid()
    if len(data) > grid.GetNumberRows():
        grid.AppendRows(len(data) - grid.GetNumberRows())
    for row, row_data in enumerate(data):
        for col, value in enumerate(row_data):
            grid.SetCellValue(row, col, value)


def update_undo_redo_state(window):
    can_undo = len(window.history) > 1
    can_redo = len(window.redo_stack) > 0

    # Update toolbar buttons if they exist
    if hasattr(window, 'undo_tool'):
        window.toolbar.EnableTool(window.undo_tool.GetId(), can_undo)
    if hasattr(window, 'redo_tool'):
        window.toolbar.EnableTool(window.redo_tool.GetId(), can_redo)

    # Update menu items if they exist
    if hasattr(window, 'edit_menu'):
        window.edit_menu.Enable(wx.ID_UNDO, can_undo)
        window.edit_menu.Enable(wx.ID_REDO, can_redo)




def save_plot_data_and_script(window, png_filepath):
    data = window.get_data_for_save()

    x_values = np.array(data['x_values'])
    y_values = np.array(data['y_values'])
    background = np.array(data['background'])
    envelope = np.array(data['calculated_fit'])
    fitted_peaks = [np.array(peak) for peak in data['individual_peak_fits']]

    # Ensure all arrays have the same length as x_values
    length = len(x_values)
    y_values = y_values[:length]
    background = background[:length]
    envelope = envelope[:length]
    fitted_peaks = [peak[:length] for peak in fitted_peaks]

    # Calculate full residuals
    full_residuals = y_values - envelope

    py_filepath = os.path.splitext(png_filepath)[0] + '.py'

    with open(py_filepath, 'w') as f:
        f.write("import numpy as np\n")
        f.write("import matplotlib.pyplot as plt\n\n")

        f.write(f"x_values = np.{repr(x_values)}\n")
        f.write(f"y_values = np.{repr(y_values)}\n")
        f.write(f"background = np.{repr(background)}\n")
        f.write(f"full_residuals = np.{repr(full_residuals)}\n")
        f.write(f"envelope = np.{repr(envelope)}\n")

        for i, peak in enumerate(fitted_peaks):
            f.write(f"fitted_peak_{i} = np.{repr(peak)}\n")

        f.write("\nplt.figure(figsize=(10, 6))\n")
        f.write("plt.scatter(x_values, y_values, color='black', s=10, label='Raw Data')\n")
        f.write("plt.plot(x_values, background, color='gray', linestyle='--', label='Background')\n")
        f.write("plt.plot(x_values, envelope, color='blue', label='Envelope')\n")
        f.write("plt.plot(x_values, full_residuals + np.max(y_values), color='green', label='Residuals')\n")

        for i in range(len(fitted_peaks)):
            f.write(
                f"plt.fill_between(x_values, background, fitted_peak_{i} + background, alpha=0.3, label='Fitted Peak {i + 1}')\n")
            f.write(f"plt.plot(x_values, fitted_peak_{i} + background, color='red', linestyle='-')\n")

        f.write("\nplt.xlabel('Binding Energy (eV)')\n")
        f.write("plt.ylabel('Intensity (CTS)')\n")
        f.write("plt.title('XPS Spectrum')\n")
        f.write("plt.legend()\n")
        f.write("plt.gca().invert_xaxis()\n")
        f.write("plt.tight_layout()\n")
        f.write("plt.show()\n")

    print(f"Plot data and script saved to {py_filepath}")


def create_plot_script_from_excel(window):
    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Extract data
    x_values = df.iloc[:, 0].values  # Assuming BE is in the first column
    y_values = df.iloc[:, 1].values  # Assuming Raw Data is in the second column
    background = df['Background'].values if 'Background' in df.columns else None
    envelope = df['Calculated Fit'].values if 'Calculated Fit' in df.columns else None
    residuals = df['Residuals'].values if 'Residuals' in df.columns else None

    # Extract fitted peaks
    fitted_peaks = []
    peak_names = []
    calculated_fit_index = df.columns.get_loc('Calculated Fit') if 'Calculated Fit' in df.columns else -1
    for col in df.columns[calculated_fit_index + 1:]:
        if df[col].notna().any():  # Check if the column is not empty
            fitted_peaks.append(df[col].values)
            peak_names.append(col)
        else:
            break  # Stop when we find an empty column

    # Create Python script
    py_filepath = os.path.splitext(file_path)[0] + f'_{sheet_name}.py'

    with open(py_filepath, 'w') as f:
        f.write("import numpy as np\n")
        f.write("import matplotlib.pyplot as plt\n\n")

        f.write(f"x_values = np.{repr(x_values)}\n")
        f.write(f"y_values = np.{repr(y_values)}\n")
        if background is not None:
            f.write(f"background = np.{repr(background)}\n")
        if envelope is not None:
            f.write(f"envelope = np.{repr(envelope)}\n")
        if residuals is not None:
            f.write(f"residuals = np.{repr(residuals)}\n")

        for i, peak in enumerate(fitted_peaks):
            f.write(f"fitted_peak_{i+1} = np.{repr(peak)}\n")

        f.write("\nplt.figure(figsize=(8, 8))\n")

        if residuals is not None:
            f.write("plt.plot(x_values, residuals + np.max(y_values), color='green', label='Residuals')\n")

        for i, peak_name in enumerate(peak_names):
            f.write(f"plt.fill_between(x_values, background, fitted_peak_{i+1}, alpha=0.6, label='{peak_name}')\n")

        if envelope is not None:
            f.write("plt.plot(x_values, envelope, color='black', label='Envelope')\n")
        if background is not None:
            f.write("plt.plot(x_values, background, color='gray', linestyle='--', label='Background')\n")
        f.write("plt.scatter(x_values, y_values, color='black', s=20, label='Raw Data')\n")

        f.write("\nplt.xlabel('Binding Energy (eV)')\n")
        f.write("plt.ylabel('Intensity (CTS)')\n")
        f.write(f"plt.title('XPS Spectrum - {sheet_name}')\n")
        f.write("plt.legend(loc='upper left')\n")
        f.write(f"plt.xlim({max(x_values)}, {min(x_values)})\n")  # Reverse x-axis
        f.write("plt.tight_layout()\n")
        f.write("plt.show()\n")

    print(f"Plot script created: {py_filepath}")

def save_peaks_library(window):
    sheet_name = window.sheet_combobox.GetValue()
    with wx.FileDialog(window, "Save peaks library", wildcard="JSON files (*.json)|*.json",
                       defaultDir="Peaks Library", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        path = fileDialog.GetPath()

        # Convert numpy arrays to lists before saving
        def convert_numpy_to_list(obj):
            if isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, dict):
                return {k: convert_numpy_to_list(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_numpy_to_list(item) for item in obj]
            return obj

        peaks_data = {
            'Core levels': {
                sheet_name: {
                    'Fitting': convert_numpy_to_list(window.Data['Core levels'][sheet_name]['Fitting'])
                }
            }
        }

        with open(path, 'w') as f:
            json.dump(peaks_data, f, indent=2)

def load_peaks_library(window):
    import wx
    import json
    from libraries.Sheet_Operations import on_sheet_selected
    from libraries.Fitting_Screen import FittingWindow
    import re

    with wx.FileDialog(window, "Load peaks library", wildcard="JSON files (*.json)|*.json",
                       defaultDir="Peaks Library", style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        path = fileDialog.GetPath()

    with open(path) as f:
        peaks_data = json.load(f)

    sheet_name = window.sheet_combobox.GetValue()
    source_sheet = list(peaks_data['Core levels'].keys())[0]
    source_data = peaks_data['Core levels'][source_sheet]

    # Check if there are existing peaks
    existing_peaks = {}
    existing_peak_count = 0
    if ('Fitting' in window.Data['Core levels'][sheet_name] and
            'Peaks' in window.Data['Core levels'][sheet_name]['Fitting']):
        existing_peaks = window.Data['Core levels'][sheet_name]['Fitting']['Peaks']
        existing_peak_count = len(existing_peaks)

    # Show choice dialog if there are existing peaks
    if existing_peak_count > 0:
        dialog = wx.MessageDialog(window,
                                  f"Found {existing_peak_count} existing peaks.\n\nChoose action:",
                                  "Load Peaks Library",
                                  wx.YES_NO | wx.CANCEL)
        dialog.SetYesNoLabels("&Overwrite existing peaks", "&Add after existing peaks")

        result = dialog.ShowModal()
        dialog.Destroy()

        if result == wx.ID_CANCEL:
            return

        overwrite = (result == wx.ID_YES)
    else:
        overwrite = True  # No existing peaks, so just load

    if overwrite:
        # Original behavior - overwrite
        window.Data['Core levels'][sheet_name]['Fitting'] = source_data['Fitting']
    else:
        # Add after existing peaks
        if 'Fitting' not in window.Data['Core levels'][sheet_name]:
            window.Data['Core levels'][sheet_name]['Fitting'] = {}
        if 'Peaks' not in window.Data['Core levels'][sheet_name]['Fitting']:
            window.Data['Core levels'][sheet_name]['Fitting']['Peaks'] = {}

        # Function to update constraint references
        def update_constraint_references(constraint_str, offset):
            if not constraint_str or constraint_str == 'Fixed':
                return constraint_str

            # Pattern to match peak references (A, B, C, etc.)
            def replace_peak_ref(match):
                peak_letter = match.group(1)
                old_index = ord(peak_letter) - ord('A')
                new_index = old_index + offset
                new_letter = chr(ord('A') + new_index)
                return match.group(0).replace(peak_letter, new_letter)

            # Update references in expressions like "A*1.5", "B+2.3", etc.
            updated = re.sub(r'\b([A-P])(?=[*+\-/])', replace_peak_ref, constraint_str)
            # Update standalone references like "A", "B"
            updated = re.sub(r'^([A-P])$', replace_peak_ref, updated)

            return updated

        # Copy library peaks with updated constraints
        library_peaks = source_data['Fitting']['Peaks']
        for peak_key, peak_data in library_peaks.items():
            new_peak_data = peak_data.copy()

            # Update constraints if they exist
            if 'Constraints' in new_peak_data:
                updated_constraints = {}
                for constraint_key, constraint_value in new_peak_data['Constraints'].items():
                    updated_constraints[constraint_key] = update_constraint_references(
                        constraint_value, existing_peak_count)
                new_peak_data['Constraints'] = updated_constraints

            window.Data['Core levels'][sheet_name]['Fitting']['Peaks'][peak_key] = new_peak_data

        # Update peak count
        window.peak_count = len(window.Data['Core levels'][sheet_name]['Fitting']['Peaks'])

    # Update display
    on_sheet_selected(window, sheet_name)

    # Ensure `fitting_window` exists and call `on_background`
    if not hasattr(window, 'fitting_window') or window.fitting_window is None:
        window.fitting_window = FittingWindow(parent=window)

    bg_low = window.peak_params_grid.GetCellValue(0, 15)
    bg_high = window.peak_params_grid.GetCellValue(0, 16)

    if bg_low and bg_high:
        window.bg_min_energy = float(bg_low)
        window.bg_max_energy = float(bg_high)
    else:
        x_values = window.Data['Core levels'][sheet_name]['B.E.']
        window.bg_min_energy = min(x_values) + 0.2
        window.bg_max_energy = max(x_values) - 0.2

def save_peaks_to_github(window, filename):
    API_URL = "https://api.github.com/repos/KherveFitting/peaks-library/contents/peaks"

    sheet_name = window.sheet_combobox.GetValue()
    peaks_data = {
        'Core levels': {
            sheet_name: {
                'Fitting': window.Data['Core levels'][sheet_name]['Fitting'],
                'Background': window.Data['Core levels'][sheet_name]['Background']
            }
        }
    }

    content = base64.b64encode(json.dumps(peaks_data).encode()).decode()

    headers = {
        "Accept": "application/vnd.github.v3+json"
    }

    # List existing files to check if file exists
    try:
        response = requests.get(f"{API_URL}/{filename}.json", headers=headers)
        response.raise_for_status()
    except requests.exceptions.RequestException:
        wx.MessageBox("Error accessing peaks library. Check your internet connection.", "Error")
        return
def load_peaks_from_github(window):
    API_URL = "https://api.github.com/repos/KherveFitting/peaks-library/contents/peaks"

    try:
        # Get list of available peak libraries
        response = requests.get(API_URL)
        response.raise_for_status()
        files = [f['name'] for f in response.json() if f['name'].endswith('.json')]

        # Show file selection dialog
        dialog = wx.SingleChoiceDialog(window, "Select a peak library to load:",
                                       "Load Peak Library", files)
        if dialog.ShowModal() == wx.ID_OK:
            filename = dialog.GetStringSelection()

            # Get file content
            response = requests.get(f"{API_URL}/{filename}")
            response.raise_for_status()
            content = base64.b64decode(response.json()['content']).decode()
            peaks_data = json.loads(content)

            # Rest of your existing load code...
            sheet_name = window.sheet_combobox.GetValue()
            source_sheet = list(peaks_data['Core levels'].keys())[0]
            source_data = peaks_data['Core levels'][source_sheet]

            window.Data['Core levels'][sheet_name]['Fitting'] = source_data['Fitting']
            if 'Background' in source_data:
                window.Data['Core levels'][sheet_name]['Background'] = source_data['Background']
                window.background = np.array(source_data['Background']['Bkg Y'])
                window.bg_min_energy = float(source_data['Background'].get('Bkg Low', 0))
                window.bg_max_energy = float(source_data['Background'].get('Bkg High', 0))
                window.background_method = str(source_data['Background'].get('Bkg Type', 'Smart'))
                window.offset_h = float(source_data['Background'].get('Bkg Offset High', 0))
                window.offset_l = float(source_data['Background'].get('Bkg Offset Low', 0))

            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, sheet_name)

    except requests.exceptions.RequestException:
        wx.MessageBox("Error accessing peaks library. Check your internet connection.", "Error")


# Add this to save.py (without importing from Open.py at the top level)

def on_save_as(window, event=None):
    """
    Save the current data to new Excel and JSON files.
    """
    import os
    import json
    import shutil
    import wx

    current_file_path = window.Data.get('FilePath')

    if not current_file_path:
        wx.MessageBox("No file is currently open.", "Error", wx.OK | wx.ICON_ERROR)
        return

    # Get default directory and filename from current file
    default_dir = os.path.dirname(current_file_path)
    default_file = os.path.basename(current_file_path)

    # Create and display the file dialog
    with wx.FileDialog(
            window,
            "Save As",
            defaultDir=default_dir,
            defaultFile=default_file,
            wildcard="Excel files (*.xlsx)|*.xlsx",
            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT
    ) as fileDialog:

        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return  # User canceled

        # Get the new file path
        new_file_path = fileDialog.GetPath()

        # Ensure it has .xlsx extension
        if not new_file_path.lower().endswith('.xlsx'):
            new_file_path += '.xlsx'

        # Get the JSON path for the current file
        current_json_path = os.path.splitext(current_file_path)[0] + '.json'

        # Calculate the new JSON path
        new_json_path = os.path.splitext(new_file_path)[0] + '.json'

        try:
            # Save the Excel file
            if os.path.exists(current_file_path):
                # If the original Excel file exists, make a copy
                shutil.copy2(current_file_path, new_file_path)
            else:
                wx.MessageBox("Original Excel file not found. Only JSON data will be saved.",
                              "Warning", wx.OK | wx.ICON_WARNING)

            # Save the JSON data
            if os.path.exists(current_json_path):
                # If JSON file exists, copy it with modifications
                with open(current_json_path, 'r') as f:
                    data = json.load(f)

                # Update the FilePath in the data
                data['FilePath'] = new_file_path

                with open(new_json_path, 'w') as f:
                    json.dump(data, f, indent=2)
            else:
                # Save current window data
                data = convert_to_serializable_and_round(window.Data)
                data['FilePath'] = new_file_path

                with open(new_json_path, 'w') as f:
                    json.dump(data, f, indent=2)

            # Update the file path in the current window
            window.Data['FilePath'] = new_file_path
            window.SetStatusText(f"Selected File: {new_file_path}", 0)

            # Update recent files list - Fixed the variable name from file_path to new_file_path
            if new_file_path in window.recent_files:
                window.recent_files.remove(new_file_path)
            window.recent_files.insert(0, new_file_path)
            window.recent_files = window.recent_files[:window.max_recent_files]

            # Update menu
            if hasattr(window, 'recent_files_menu'):
                # Clear existing items
                for item in window.recent_files_menu.GetMenuItems():
                    window.recent_files_menu.DestroyItem(item)

                # Add new items
                for i, file_path in enumerate(window.recent_files):
                    item = window.recent_files_menu.Append(wx.ID_ANY, os.path.basename(file_path))
                    # Late import for the open_xlsx_file function
                    from libraries.Open import open_xlsx_file
                    window.Bind(wx.EVT_MENU, lambda evt, fp=file_path: open_xlsx_file(window, fp), item)

            window.save_config()

            # Also save the current state to the new file
            data = window.get_data_for_save()
            save_to_excel(window, data, new_file_path, window.sheet_combobox.GetValue())
            save_plot_to_excel(window)

            window.show_popup_message2("Save As Successful", f"File saved as:\n{new_file_path}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            wx.MessageBox(f"Error saving file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_vamas_file(window, output_path=None):
    """
    Save the current data as a VAMAS file.

    This function takes the data from window.Data and creates a VAMAS file,
    reading experimental description from column AX of Excel sheets or from JSON.

    Args:
        window: The main application window object containing the data
        output_path: Optional path for output file. If None, shows save dialog.
    """
    import os
    import json
    import datetime

    try:
        if 'FilePath' not in window.Data or not window.Data['FilePath']:
            wx.MessageBox("No data loaded to save as VAMAS file.", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Show save dialog if no output path provided
        if output_path is None:
            with wx.FileDialog(window, "Save VAMAS file",
                               wildcard="VAMAS files (*.vms)|*.vms",
                               style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
                if fileDialog.ShowModal() == wx.ID_CANCEL:
                    return
                output_path = fileDialog.GetPath()

        # Ensure .vms extension
        if not output_path.lower().endswith('.vms'):
            output_path += '.vms'

        # Create console window for progress updates
        parent_pos = window.GetPosition()
        parent_size = window.GetSize()
        console_frame = wx.Frame(window, title="Creating VAMAS File", size=(300, 350))
        console_frame.SetPosition((
            parent_pos.x + (parent_size.width - 300) // 2,
            parent_pos.y + (parent_size.height - 350) // 2
        ))
        console_text = wx.TextCtrl(console_frame, style=wx.TE_MULTILINE | wx.TE_READONLY)
        console_frame.Show()

        def update_console(message):
            console_text.AppendText(message + '\n')
            console_text.Update()
            wx.SafeYield()

        update_console("Starting VAMAS file creation...")

        # Get experimental description data
        exp_data = get_experimental_description_data(window)

        # Write VAMAS file
        with open(output_path, 'w') as f:
            update_console("Writing VAMAS header...")
            write_vamas_header(f, window, exp_data)

            update_console("Writing data blocks...")
            write_vamas_blocks(f, window, exp_data, update_console)

        update_console("VAMAS file created successfully!")
        wx.CallLater(1500, console_frame.Close)

        wx.MessageBox(f"VAMAS file saved successfully:\n{output_path}",
                      "Success", wx.OK | wx.ICON_INFORMATION)

    except Exception as e:
        if 'console_frame' in locals():
            console_frame.Close()
        wx.MessageBox(f"Error creating VAMAS file: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def get_experimental_description_data(window):
    """
    Extract experimental description data from Excel sheets or JSON file.

    Returns:
        dict: Dictionary containing experimental parameters for each core level
    """
    import pandas as pd
    import json
    import os

    exp_data = {}

    # Try to get data from JSON file first (easier structure)
    json_path = os.path.splitext(window.Data['FilePath'])[0] + '.json'
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r') as f:
                json_data = json.load(f)

            # Extract experimental data from JSON if available
            for core_level in window.Data['Core levels']:
                if core_level in json_data.get('Core levels', {}):
                    exp_data[core_level] = extract_exp_data_from_json(json_data['Core levels'][core_level])
        except:
            pass

    # Fallback: read from Excel file column AX (column 50)
    if not exp_data:
        exp_data = get_exp_data_from_excel(window.Data['FilePath'], window.Data['Core levels'].keys())

    return exp_data


def extract_exp_data_from_json(core_level_data):
    """Extract experimental parameters from JSON core level data."""

    # Default experimental parameters structure
    exp_params = {
        'Sample ID': 'Unknown',
        'Year': 2024,
        'Month': 1,
        'Day': 1,
        'Hour': 0,
        'Minute': 0,
        'Second': 0,
        'Technique': 'XPS',
        'Species': 'Unknown',
        'Transition': '1s',
        'Source Label': 'Al Ka',
        'Source Energy': 1486.67,
        'Source Width X': 0.0,
        'Source Width Y': 0.0,
        'Pass Energy': 20.0,
        'Work Function': 4.5,
        'Analyzer Mode': 'CAE',
        'Take-off Polar Angle': 45.0,
        'Take-off Azimuth': 0.0,
        'Analysis Width X': 1000.0,
        'Analysis Width Y': 1000.0,
        'X Label': 'Binding Energy',
        'X Units': 'eV',
        'X Start': 0.0,
        'X Step': 0.1,
        'Num Y Values': 0,
        'Collection Time': 1.0,
        'Y Unit': 'Counts',
        'Block Comment': ''
    }

    # Try to extract actual values if they exist in the data
    if 'experimental_description' in core_level_data:
        for item in core_level_data['experimental_description']:
            if isinstance(item, list) and len(item) >= 2:
                key, value = item[0], item[1]
                if key in exp_params:
                    exp_params[key] = value

    # Calculate parameters from actual data
    if 'B.E.' in core_level_data and core_level_data['B.E.']:
        be_values = core_level_data['B.E.']
        exp_params['X Start'] = max(be_values)  # VAMAS starts with highest BE
        exp_params['X Step'] = abs(be_values[1] - be_values[0]) if len(be_values) > 1 else 0.1
        exp_params['Num Y Values'] = len(be_values)

    return exp_params


def get_exp_data_from_excel(excel_path, sheet_names):
    """Extract experimental description data from Excel column AX."""
    import pandas as pd

    exp_data = {}

    try:
        for sheet_name in sheet_names:
            exp_data[sheet_name] = {}

            # Read experimental description from column AX (column 50)
            df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[49, 50], header=None)

            # Parse experimental description
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    key = str(row.iloc[0]).strip()
                    value = row.iloc[1]
                    exp_data[sheet_name][key] = value

            # Set default values for missing parameters
            exp_data[sheet_name] = set_default_exp_values(exp_data[sheet_name])

    except Exception as e:
        print(f"Error reading experimental data from Excel: {e}")
        # Return default values for all sheets
        for sheet_name in sheet_names:
            exp_data[sheet_name] = set_default_exp_values({})

    return exp_data


def set_default_exp_values(exp_dict):
    """Set default values for missing experimental parameters."""

    defaults = {
        'Sample ID': 'Unknown',
        'Year': 2024,
        'Month': 1,
        'Day': 1,
        'Hour': 0,
        'Minute': 0,
        'Second': 0,
        'Technique': 'XPS',
        'Species': 'Unknown',
        'Transition': '1s',
        'Source Label': 'Al Ka',
        'Source Energy': 1486.67,
        'Source Width X': 0.0,
        'Source Width Y': 0.0,
        'Pass Energy': 20.0,
        'Work Function': 4.5,
        'Analyzer Mode': 'CAE',
        'Take-off Polar Angle': 45.0,
        'Take-off Azimuth': 0.0,
        'Analysis Width X': 1000.0,
        'Analysis Width Y': 1000.0,
        'X Label': 'Binding Energy',
        'X Units': 'eV',
        'X Start': 0.0,
        'X Step': 0.1,
        'Num Y Values': 0,
        'Collection Time': 1.0,
        'Y Unit': 'Counts',
        'Block Comment': ''
    }

    for key, default_value in defaults.items():
        if key not in exp_dict:
            exp_dict[key] = default_value

    return exp_dict


def write_vamas_header(f, window, exp_data):
    """Write the VAMAS file header."""

    # VAMAS format identifier
    f.write("VAMAS Surface Chemical Analysis Standard Data Transfer Format 1988 May 4\n")

    # Institution, instrument, operator, experiment identifiers
    f.write("Not Specified\n")
    f.write("Not Specified\n")
    f.write("Not Specified\n")
    f.write("Not Specified\n")

    # Comment lines
    f.write("4\n")
    f.write("Casa Info Follows CasaXPS Version 2.3.27PR1.7\n")
    f.write("0\n")
    f.write("SourceAnalyserAngle: Not Specified\n")
    f.write("CasaRowLabel:KherveFitting SampleID\n")

    # Experiment and scan mode
    f.write("MAP\n")
    f.write("REGULAR\n")

    # Number of spectral regions
    num_core_levels = len(window.Data['Core levels'])
    f.write(f"{num_core_levels}\n")

    # MAP mode parameters (FIXED: these should be 0, 0, 0)
    f.write("0\n")  # Number of analysis positions
    f.write("0\n")  # Number of discrete X coordinates
    f.write("0\n")  # Number of discrete Y coordinates

    # Experiment variables
    f.write("1\n")  # Number of experiment variables
    f.write("Exp Variable\n")  # Variable label
    f.write("d\n")  # Variable unit

    # Inclusion/exclusion list
    f.write("0\n")  # Number of entries

    # Manually entered items
    f.write("0\n")

    # Future upgrade entries
    f.write("0\n")
    f.write("0\n")

    # Number of blocks
    f.write(f"{num_core_levels}\n")


def write_vamas_blocks(f, window, exp_data, update_console):
    """Write all data blocks to the VAMAS file."""

    block_num = 0
    for core_level_name, core_level_data in window.Data['Core levels'].items():
        block_num += 1
        update_console(f"Writing block {block_num}: {core_level_name}")

        # Get experimental parameters for this core level
        exp_params = exp_data.get(core_level_name, set_default_exp_values({}))

        # Extract species and transition from core level name
        species, transition = parse_core_level_name(core_level_name)
        exp_params['Species'] = species
        exp_params['Transition'] = transition

        # Write block
        write_single_block(f, core_level_name, core_level_data, exp_params, block_num, window)

    # CRITICAL: Add the end of experiment marker
    f.write("end of experiment\n")


def parse_core_level_name(name):
    """Parse core level name to extract species and transition."""
    import re

    # Try to match patterns like "C1s", "O2p", etc.
    match = re.match(r'([A-Z][a-z]?)(\d+[spdfghi])', name)
    if match:
        return match.group(1), match.group(2)

    # Fallback for other patterns
    if name.lower() in ['survey', 'wide']:
        return 'Survey', 'none'

    return 'Unknown', '1s'


def get_rsf_for_core_level(peak_name, library_data, current_instrument):
    """
    Get RSF for a core level using the exact same method as _extract_peak_parameters.

    Args:
        peak_name: Name like "O1s", "C1s p1", etc.
        library_data: The RSF library data
        current_instrument: Current instrument from window

    Returns:
        float: RSF value
    """
    import re

    # Use the EXACT same regex as _extract_peak_parameters
    match = re.match(r'([A-Z][a-z]*)(\d+[spdf])(?:(\d+/\d+))?', peak_name)
    if match:
        element, orbital, suborbital = match.groups()
        if suborbital:
            # Use full orbital including suborbital for RSF lookup
            core_level = f"{element}{orbital}{suborbital}"
        else:
            core_level = f"{element}{orbital}"
    else:
        core_level = ''.join(filter(str.isalnum, peak_name.split()[0]))
        element, orbital, suborbital = core_level, '', None

    print(f"DEBUG: Parsed - Element: {element}, Orbital: {orbital}, Suborbital: {suborbital}")

    # Get RSF directly using complete orbital designation - EXACT same logic as _extract_peak_parameters
    key = (element, orbital + (suborbital or ''))
    print(f"DEBUG: Looking for key: {key}")
    print(f"DEBUG: Available keys: {list(library_data.keys())[:10]}...")  # Show first 10 keys

    if key in library_data and current_instrument in library_data[key]:
        rsf = library_data[key][current_instrument]['rsf']
        print(f"DEBUG: Found RSF: {rsf}")
        return rsf
    else:
        print(f"DEBUG: Key {key} not found or instrument {current_instrument} not available")
        if key in library_data:
            print(f"DEBUG: Available instruments for {key}: {list(library_data[key].keys())}")
            # Fallback to first available instrument
            available_instruments = list(library_data[key].keys())
            if available_instruments:
                fallback_instrument = available_instruments[0]
                rsf = library_data[key][fallback_instrument]['rsf']
                print(f"DEBUG: Using fallback instrument {fallback_instrument}, RSF: {rsf}")
                return rsf
        print(f"DEBUG: No RSF found, returning 1.0")
        return 1.0


def write_casa_fitting_info(f, core_level_name, core_level_data, window):
    """
    Write CASA fitting information to VAMAS file between comment count and 'XPS'.
    Uses existing RSF and atomic mass libraries from Export.py and Area_Calculation.

    Args:
        f: File object to write to
        core_level_name: Name of the core level (e.g., "O1s")
        core_level_data: Dictionary containing fitting and background data
        window: Main window object to access current_instrument
    """
    # Import existing libraries
    from libraries.Open import load_library_data
    from libraries.Area_Calculation import extract_element_symbol, ATOMIC_MASSES
    import re

    # Load RSF library data and get current instrument
    library_data = load_library_data()
    current_instrument = window.current_instrument

    print(f"DEBUG: Current instrument: {current_instrument}")
    print(f"DEBUG: Core level name: {core_level_name}")

    # Prepare all comment lines first to count them
    comment_lines = []

    # Add standard CASA header lines
    comment_lines.append("Casa Info Follows")
    comment_lines.append("0")
    comment_lines.append("0")

    # Check if fitting data exists
    fitting_data = core_level_data.get('Fitting', {})
    background_data = core_level_data.get('Background', {})
    peaks_data = fitting_data.get('Peaks', {})

    # Add background information
    num_backgrounds = 1 if background_data and background_data.get('Bkg Type') else 0
    comment_lines.append(str(num_backgrounds))

    if num_backgrounds > 0:
        # Convert BE to KE for CASA format
        photon_energy = 1486.67
        bkg_low_be = float(background_data.get('Bkg Low', 0))
        bkg_high_be = float(background_data.get('Bkg High', 0))

        # Convert BE to KE (KE = photon_energy - BE)
        low_ke = photon_energy - bkg_high_be
        high_ke = photon_energy - bkg_low_be

        # Map KherveFitting background types to CASA types
        bg_type_mapping = {
            'Smart': 'Shirley',
            'Shirley': 'Shirley',
            'Linear': 'Linear',
            'Polynomial': 'Linear',
            'Tougaard': 'Tougaard'
        }

        kf_bg_type = background_data.get('Bkg Type', 'Smart')
        casa_bg_type = bg_type_mapping.get(kf_bg_type, 'Shirley')

        # Get RSF using the EXACT same method as _extract_peak_parameters
        rsf = get_rsf_for_core_level(core_level_name, library_data, current_instrument)

        # Get atomic mass using existing library
        element_symbol = extract_element_symbol(core_level_name)
        atomic_mass = ATOMIC_MASSES.get(element_symbol, 1.0)

        print(f"DEBUG: Background RSF: {rsf}, Atomic mass: {atomic_mass}")

        # Write CASA region line
        casa_region = (f"CASA region (*{core_level_name}*) (*{casa_bg_type}*) "
                       f"{low_ke:.5f} {high_ke:.5f} {rsf} 2 0 0 0 -450 0 0 "
                       f"(*{core_level_name}*) {atomic_mass} 0 {rsf}")
        comment_lines.append(casa_region)

    # Add peak information
    num_peaks = len(peaks_data)
    comment_lines.append(str(num_peaks))

    for peak_name, peak_data in peaks_data.items():
        # Get peak parameters
        position_be = peak_data.get('Position', 0)
        area = peak_data.get('Area', 0)
        fwhm = peak_data.get('FWHM', 1.0)
        lg_ratio = peak_data.get('L/G', 30)
        sigma = peak_data.get('Sigma', 0.6)
        gamma = peak_data.get('Gamma', 0.4)

        # Convert position from BE to KE
        position_ke = photon_energy - position_be

        # Map KherveFitting model to CASA model - UPDATED FOR LA MODELS WITH GAMMA MAPPING
        fitting_model = peak_data.get('Fitting Model', 'GL (Area)')
        print(f"DEBUG: Processing peak {peak_name} with model: {fitting_model}")

        if 'LA*G' in fitting_model:
            # LA*G model: use format LA(sigma, gamma, >100)
            casa_model = f"LA({sigma:.2f}, {gamma:.2f}, 150)"
            print(f"DEBUG: LA*G model - sigma: {sigma}, gamma: {gamma}")
        elif 'LA (Area, Ïƒ/Î³, Î³)' in fitting_model:
            # LA with sigma/gamma ratio: check if sigma/gamma â‰ˆ 50% (sigma â‰ˆ gamma)
            sigma_gamma_ratio = sigma / (sigma + gamma) * 100 if (sigma + gamma) > 0 else 50

            if abs(sigma_gamma_ratio - 50) < 5:  # If approximately 50/50 split
                # Check gamma mapping from Open.py
                gamma_mapping = {
                    20: 2.7, 30: 2.4, 40: 2.2, 50: 2.0, 60: 1.8,
                    70: 1.6, 80: 1.4, 90: 1.2, 100: 1.0
                }

                # Find matching gamma value in mapping
                matching_ratio = None
                for ratio_val, mapped_gamma in gamma_mapping.items():
                    if abs(gamma - mapped_gamma) < 0.1:
                        matching_ratio = ratio_val
                        break

                if matching_ratio:
                    casa_model = f"LA({matching_ratio})"
                    print(f"DEBUG: LA Ïƒ/Î³ model matches 50% split with gamma mapping - ratio: {matching_ratio}")
                else:
                    casa_model = f"LA({sigma:.2f}, {gamma:.2f}, 5)"
                    print(f"DEBUG: LA Ïƒ/Î³ model 50% split but no gamma mapping match - using 3-param format")
            else:
                # Not 50/50 split, use 3-parameter format
                casa_model = f"LA({sigma:.2f}, {gamma:.2f}, 5)"
                print(f"DEBUG: LA Ïƒ/Î³ model not 50% split ({sigma_gamma_ratio:.1f}%) - using 3-param format")
        elif 'LA (Area, Ïƒ, Î³)' in fitting_model:
            # Standard LA model: Check if it matches the gamma mapping pattern
            # Gamma mapping from Open.py
            gamma_mapping = {
                20: 2.7, 30: 2.4, 40: 2.2, 50: 2.0, 60: 1.8,
                70: 1.6, 80: 1.4, 90: 1.2, 100: 1.0
            }

            # Check if this sigma/gamma combination matches a gamma mapping entry
            matches_mapping = False
            for ratio_val, mapped_gamma in gamma_mapping.items():
                # In Open.py: sigma_value = gamma_value and lg_value = ratio_value
                # So if sigma â‰ˆ gamma and they match a mapped value, use single parameter format
                if abs(gamma - mapped_gamma) < 0.1 and abs(sigma - gamma) < 0.1:
                    casa_model = f"LA({ratio_val})"
                    matches_mapping = True
                    print(f"DEBUG: LA Ïƒ,Î³ model matches mapping - ratio: {ratio_val}")
                    break

            if not matches_mapping:
                # Use 3-parameter format if it doesn't match the mapping
                casa_model = f"LA({sigma:.2f}, {gamma:.2f}, 0.5)"
                print(f"DEBUG: LA Ïƒ,Î³ model - sigma: {sigma}, gamma: {gamma}")
        elif 'SGL' in fitting_model:
            casa_model = f"SGL({int(lg_ratio)})"
        elif 'Voigt' in fitting_model:
            # Convert Voigt to SGL as closest CASA equivalent
            # Voigt FWHM becomes SGL MFWHM, L/G ratio becomes SGL parameter
            casa_model = f"SGL({int(lg_ratio)})"
            print(f"DEBUG: Voigt model converted to SGL - L/G: {lg_ratio} -> SGL({int(lg_ratio)})")
        else:
            # Default GL model
            casa_model = f"GL({int(lg_ratio)})"

        print(f"DEBUG: Final CASA model: {casa_model}")

        # Get background limits in KE
        if background_data:
            bg_low_ke = photon_energy - float(background_data.get('Bkg High', position_be + 5))
            bg_high_ke = photon_energy - float(background_data.get('Bkg Low', position_be - 5))
        else:
            bg_low_ke = position_ke - 5
            bg_high_ke = position_ke + 5

        # Get RSF for this specific peak using the same method as _extract_peak_parameters
        rsf = get_rsf_for_core_level(peak_name, library_data, current_instrument)

        # Get atomic mass
        element_symbol = extract_element_symbol(peak_name)
        atomic_mass = ATOMIC_MASSES.get(element_symbol, 1.0)

        print(f"DEBUG: Peak {peak_name} RSF: {rsf}, Atomic mass: {atomic_mass}")

        # Get constraints from peak data
        constraints = peak_data.get('Constraints', {})

        # Convert KherveFitting constraints back to CASA format

        # Convert position BE to KE and apply constraints
        area_constraint = convert_constraint_to_casa(constraints.get('Area', ''), area, 'Area')
        fwhm_constraint = convert_constraint_to_casa(constraints.get('FWHM', ''), fwhm, 'MFWHM')

        # For position, convert BE constraint limits to KE
        pos_constraint_str = constraints.get('Position', '')
        if ":" in pos_constraint_str:
            be_min, be_max = pos_constraint_str.split(":")
            ke_min = photon_energy - float(be_max)  # Inverted: higher BE = lower KE
            ke_max = photon_energy - float(be_min)  # Inverted: lower BE = higher KE
            pos_constraint = f"{position_ke:.2f} {ke_min:.2f} {ke_max:.2f} -1 1"
        else:
            pos_constraint = convert_constraint_to_casa(pos_constraint_str, position_ke, 'Position')

        # Write CASA comp line with constraints
        casa_comp = (f"CASA comp (*{peak_name}*) (*{casa_model}*) "
                     f"Area {area_constraint} "
                     f"MFWHM {fwhm_constraint} "
                     f"Position {pos_constraint} "
                     f"RSF {rsf} MASS {atomic_mass} INDEX -1 "
                     f"(*{core_level_name}*) CONST (**) UNCORRECTEDRSF {rsf}")
        comment_lines.append(casa_comp)

        # # Write CASA comp line
        # casa_comp = (f"CASA comp (*{peak_name}*) (*{casa_model}*) "
        #              f"Area {area} 1e-20 659597.45 -1 1 "
        #              f"MFWHM {fwhm:.2f} 0.24 6 -1 1 "
        #              f"Position {position_ke:.2f} {bg_low_ke:.5f} {bg_high_ke:.5f} -1 1 "
        #              f"RSF {rsf} MASS {atomic_mass} INDEX -1 "
        #              f"(*{core_level_name}*) CONST (**) UNCORRECTEDRSF {rsf}")
        # comment_lines.append(casa_comp)

    # Add final line
    comment_lines.append("Created by KherveFitting")

    # Write the correct number of comment lines
    f.write(f"{len(comment_lines)}\n")

    # Write all comment lines
    for line in comment_lines:
        f.write(f"{line}\n")


def convert_constraint_to_casa(constraint_str, value, param_name):
    """Convert KherveFitting constraint format to CASA format"""
    if not constraint_str or constraint_str == "":
        # Default constraints
        if param_name == "Area":
            return f"{value} 1e-20 659597.45 -1 1"
        elif param_name == "MFWHM":
            return f"{value:.2f} 0.24 6 -1 1"
        elif param_name == "Position":
            return f"{value:.2f} {value - 10:.2f} {value + 10:.2f} -1 1"

    if constraint_str == "Fixed":
        # Fixed constraint - very tight limits
        return f"{value:.2f} {value - 0.001:.3f} {value + 0.001:.3f} -1 1"

    elif ":" in constraint_str:
        # Range constraint like "1:1000" or "0.3:3.5"
        min_val, max_val = constraint_str.split(":")
        return f"{value:.2f} {min_val} {max_val} -1 1"

    elif any(letter in constraint_str for letter in "ABCDEFGHIJKLMNOP"):
        # Linked constraint like "A+1.5" or "B*2.0"
        import re
        match = re.match(r'([A-P])([+\-*/])([\d.]+)', constraint_str)
        if match:
            peak_letter, operator, const_value = match.groups()
            peak_index = ord(peak_letter) - ord('A')  # A=0, B=1, etc.
            return f"{value:.2f} {value - 10:.2f} {value + 10:.2f} {peak_index} {const_value}"

    # Default fallback
    return f"{value:.2f} {value - 10:.2f} {value + 10:.2f} -1 1"

def write_single_block(f, core_level_name, core_level_data, exp_params, block_num, window):
    """Write a single data block to the VAMAS file."""

    # Block identifier (use original sample name)
    f.write(f"KherveFitting BlockID {core_level_name}/{block_num}\n")

    # Sample identifier
    f.write("KherveFitting SampleID\n")

    # Date and time
    f.write("2025\n") # To replace by the date of acquisition
    f.write("02\n")
    f.write("31\n")
    f.write("13\n")
    f.write("13\n")
    f.write("13\n")

    # Hours advance GMT
    f.write("0\n")

    # Block comment
    # f.write("Casa Info Follows\n")
    # f.write("0\n")
    # f.write("0\n")
    # f.write("0\n")
    # f.write("0\n")
    # f.write("Created by KherveFitting\n")
    write_casa_fitting_info(f, core_level_name, core_level_data, window)

    # Technique
    f.write("XPS\n")

    # Experimental variables
    f.write("16384\n")

    # Analysis source parameters
    f.write("16384\n")
    f.write("0\n")
    f.write("X-ray Source\n") # To replace by the type of source
    f.write("1486.71\n") # To replace by the energy
    f.write("001\n") # To replace by power of the source
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("1000\n")
    f.write("1000\n")
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("FAT\n")
    f.write("20\n")
    f.write("1e+37\n")
    f.write("-4.47\n") # To replace by workfunction
    f.write("0\n")
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("1e+37\n")

    # Species and transition
    species, transition = parse_core_level_name(core_level_name)
    f.write(f"{species}\n")
    f.write(f"{transition}\n")
    f.write("-1\n")

    # Data parameters
    f.write("Binding Energy\n")
    f.write("eV\n")

    # Get actual data
    be_values = core_level_data.get('B.E.', [])
    intensity_values = (
            core_level_data.get('Corrected Data', []) or
            core_level_data.get('Raw Data', []) or
            core_level_data.get('Intensity', [])
    )

    if not be_values or not intensity_values:
        raise ValueError(f"No data found for core level {core_level_name}")

    # Ensure equal length
    min_length = min(len(be_values), len(intensity_values))
    be_values = be_values[:min_length]
    intensity_values = intensity_values[:min_length]
    intensity_values.reverse()  # Reverse to match VAMAS order

    # Use BE values - start from highest BE (be_end)
    be_end = min(be_values)  # Highest binding energy
    be_step = round(abs(be_values[1] - be_values[0]), 1) if len(be_values) > 1 else 0.1

    f.write(f"{be_end}\n")  # Changed from be_start to be_end
    f.write(f"{be_step}\n")

    # Number of corresponding variables
    f.write("2\n")

    # Corresponding variable labels
    f.write("Intensity\n")
    f.write("d\n")
    f.write("Transmission\n")
    f.write("d\n")

    # Additional parameters
    f.write("pulse counting\n")
    # f.write("0.352941\n")
    f.write("1.00\n")   # To replace by the dwell time
    f.write("1\n")          # To replace by how many time it was measured
    f.write("0\n")
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("1e+37\n")
    f.write("1\n")

    # Additional numerical parameter
    f.write("MFP Exponent\n")
    f.write("d\n")
    f.write("0\n")

    # Number of Y values - multiply by 2 for intensity + transmission pairs
    f.write(f"{len(intensity_values) * 2}\n")

    # Y mins and maxs (use actual data values)
    intensity_min = int(min(intensity_values))
    intensity_max = int(max(intensity_values))

    f.write(f"{intensity_min}\n")
    f.write(f"{intensity_max}\n")
    f.write("1.00\n")
    f.write("1.00\n")

    # Write the data pairs
    for i, intensity in enumerate(intensity_values):
        f.write(f"{int(intensity)}\n")

        # Generate transmission that matches original pattern
        transmission_start = 1.00
        transmission_end = 1.00
        transmission_step = (transmission_start - transmission_end) / (len(intensity_values) - 1)
        transmission = transmission_start - (i * transmission_step)
        f.write(f"{transmission:.6f}\n")


def save_vamas_file_dialog(window):
    """
    Show save dialog for VAMAS file and call save_vamas_file.

    Args:
        window: The main application window object.
    """
    save_vamas_file(window)




def export_sheet_to_txt(window):
    """Export current sheet data to a TXT file including all columns"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    sheet_name = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        # Read complete sheet data from Excel file including fitted data
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Create default file path
        default_file = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.txt"
        default_dir = os.path.dirname(file_path)

        # Show save dialog
        with wx.FileDialog(window, "Export to TXT file",
                           defaultDir=default_dir,
                           defaultFile=default_file,
                           wildcard="Text files (*.txt)|*.txt",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            txt_path = fileDialog.GetPath()

            # Remove any empty columns and any columns after the peak fits
            # Get index of first empty column after the last peak fit
            last_non_empty_col = df.shape[1]
            for i in range(df.shape[1] - 1, 4, -1):  # Start from the end, stopping at column E
                if df.iloc[:, i].notna().any():
                    break
                last_non_empty_col = i

            # Remove columns after the last non-empty column
            if last_non_empty_col < df.shape[1]:
                df = df.iloc[:, :last_non_empty_col]

            # Drop columns with all NaN values
            df = df.dropna(axis=1, how='all')

            # Export to text file - use utf-8 encoding to handle Unicode characters
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(f"# {sheet_name} XPS data\n")
                # Write header
                f.write("# " + "\t".join([str(col) for col in df.columns]) + "\n")
                # Write data
                for _, row in df.iterrows():
                    line = "\t".join([f"{val:.6f}" if isinstance(val, (int, float)) and not pd.isna(val)
                                      else "" if pd.isna(val) else str(val) for val in row])
                    f.write(line + "\n")

            window.show_popup_message2("Export Successful", f"Data exported to: {txt_path}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error exporting to TXT: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def export_sheet_to_csv(window):
    """Export current sheet data to a CSV file including all columns"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    sheet_name = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        # Read complete sheet data from Excel file including fitted data
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Create default file path
        default_file = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.csv"
        default_dir = os.path.dirname(file_path)

        # Show save dialog
        with wx.FileDialog(window, "Export to CSV file",
                           defaultDir=default_dir,
                           defaultFile=default_file,
                           wildcard="CSV files (*.csv)|*.csv",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            csv_path = fileDialog.GetPath()

            # Remove any empty columns and any columns after the peak fits
            # Get index of first empty column after the last peak fit
            last_non_empty_col = df.shape[1]
            for i in range(df.shape[1] - 1, 4, -1):  # Start from the end, stopping at column E
                if df.iloc[:, i].notna().any():
                    break
                last_non_empty_col = i

            # Remove columns after the last non-empty column
            if last_non_empty_col < df.shape[1]:
                df = df.iloc[:, :last_non_empty_col]

            # Drop columns with all NaN values
            df = df.dropna(axis=1, how='all')

            # Export to CSV with utf-8 encoding
            df.to_csv(csv_path, index=False, float_format='%.6f', encoding='utf-8')

            window.show_popup_message2("Export Successful", f"Data exported to: {csv_path}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error exporting to CSV: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def export_sheet_to_dat(window):
    """Export current sheet data to a DAT file including all columns"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    sheet_name = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        # Read complete sheet data from Excel file including fitted data
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Create default file path
        default_file = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.dat"
        default_dir = os.path.dirname(file_path)

        # Show save dialog
        with wx.FileDialog(window, "Export to DAT file",
                           defaultDir=default_dir,
                           defaultFile=default_file,
                           wildcard="DAT files (*.dat)|*.dat",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            dat_path = fileDialog.GetPath()

            # Remove any empty columns and any columns after the peak fits
            # Get index of first empty column after the last peak fit
            last_non_empty_col = df.shape[1]
            for i in range(df.shape[1] - 1, 4, -1):  # Start from the end, stopping at column E
                if df.iloc[:, i].notna().any():
                    break
                last_non_empty_col = i

            # Remove columns after the last non-empty column
            if last_non_empty_col < df.shape[1]:
                df = df.iloc[:, :last_non_empty_col]

            # Drop columns with all NaN values
            df = df.dropna(axis=1, how='all')

            # Export to DAT file with utf-8 encoding
            with open(dat_path, 'w', encoding='utf-8') as f:
                # Write header with column names, preceded by '#'
                f.write("# " + " ".join([str(col) for col in df.columns]) + "\n")
                # Write data
                for _, row in df.iterrows():
                    line = " ".join([f"{val:.6f}" if isinstance(val, (int, float)) and not pd.isna(val)
                                     else "" if pd.isna(val) else str(val) for val in row])
                    f.write(line + "\n")

            window.show_popup_message2("Export Successful", f"Data exported to: {dat_path}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error exporting to DAT: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def copy_core_level(window):
    import json
    import os
    import tempfile
    from copy import deepcopy

    sheet_name = window.sheet_combobox.GetValue()
    if sheet_name in window.Data['Core levels']:
        # Create a deep copy of the current core level data
        clipboard_data = {
            'core_level': deepcopy(window.Data['Core levels'][sheet_name]),
            'peak_params_grid': get_grid_data(window.peak_params_grid),
            'peak_count': window.peak_count,
            'original_sheet_name': sheet_name
        }

        # Add Excel column data preservation
        file_path = window.Data.get('FilePath', '')
        if file_path and os.path.exists(file_path):
            try:
                # Read all data including columns C and D
                import pandas as pd
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Store column names
                column_names = df.columns.tolist()
                clipboard_data['column_names'] = column_names

                # Store exact data from columns C and D (indices 2 and 3)
                if df.shape[1] > 3:
                    clipboard_data['column_C_data'] = df.iloc[:, 2].tolist()
                    clipboard_data['column_D_data'] = df.iloc[:, 3].tolist()

                # Store experimental description if available
                exp_data_col = None
                for col_idx, col_name in enumerate(df.columns):
                    if col_name == "Experimental Description":
                        exp_data_col = col_idx
                        break

                if exp_data_col is not None:
                    exp_description = []
                    for i in range(min(30, len(df))):  # Limit to reasonable number of rows
                        if i < len(df) and col_idx < df.shape[1] and col_idx + 1 < df.shape[1]:
                            key = df.iloc[i, exp_data_col] if not pd.isna(df.iloc[i, exp_data_col]) else ""
                            value = df.iloc[i, exp_data_col + 1] if not pd.isna(df.iloc[i, exp_data_col + 1]) else ""
                            if key or value:
                                exp_description.append([key, value])

                    clipboard_data['experimental_description'] = exp_description

            except Exception as e:
                print(f"Error reading Excel data for {sheet_name}: {e}")

        # Save to clipboard file
        clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')
        with open(clipboard_file, 'w') as f:
            json.dump(convert_to_serializable_and_round(clipboard_data), f)


        window.show_popup_message2("Core Level Copied", f"Core level '{sheet_name}' copied to clipboard")
    else:
        window.show_popup_message2("Copy Failed", "No core level data to copy")


def paste_core_level(window):
    import json
    import os
    import tempfile
    import pandas as pd
    import openpyxl
    from copy import deepcopy

    # Perform backup before pasting
    from libraries.Utilities import perform_auto_backup
    perform_auto_backup(window)

    clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_corelevels_clipboard.json')

    if not os.path.exists(clipboard_file):
        window.show_popup_message2("Paste Failed", "No data in clipboard")
        return

    # Check if no file is open
    is_new_file = False
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        is_new_file = True
        # Show save dialog to create a new file
        with wx.FileDialog(window, "Save As", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return  # User canceled
            file_path = fileDialog.GetPath()
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'

            # Create empty Excel file
            wb = openpyxl.Workbook()
            wb.save(file_path)

            # Set the filepath in window.Data
            window.Data['FilePath'] = file_path
            window.SetStatusText(f"Selected File: {file_path}", 0)

            # Initialize Core levels if needed
            if 'Core levels' not in window.Data:
                window.Data['Core levels'] = {}
                window.Data['Number of Core levels'] = 0

            # Update recent files
            if file_path in window.recent_files:
                window.recent_files.remove(file_path)
            window.recent_files.insert(0, file_path)
            window.recent_files = window.recent_files[:window.max_recent_files]
            window.save_config()

    try:
        with open(clipboard_file, 'r') as f:
            clipboard_data = json.load(f)

        # Save current state for undo
        save_state(window)

        # Create a new sheet name based on the original
        original_name = clipboard_data.get('original_sheet_name', 'Sheet1')  # Provide default if missing
        if not original_name or original_name.strip() == '':
            original_name = 'Sheet1'  # Fallback if original name is empty

        if is_new_file:
            # For new files, just use the original name since there are no sheets yet
            new_sheet_name = original_name
            current_sheet = ""  # No current sheet for new files
        else:
            # For existing files, create a unique name
            current_sheet = window.sheet_combobox.GetValue()
            # Get all sheets and find index of current sheet
            all_sheets = list(window.Data['Core levels'].keys())
            # Only try to get current_index if all_sheets is not empty
            if all_sheets and current_sheet:
                try:
                    current_index = all_sheets.index(current_sheet)
                except ValueError:
                    current_index = -1  # Default to end of list if current sheet not found
            else:
                current_index = -1  # Default to end of list

        # Create a unique new sheet name
        wb = openpyxl.load_workbook(window.Data['FilePath'])
        new_sheet_name = original_name
        i = 1
        while new_sheet_name in wb.sheetnames:
            new_sheet_name = f"{original_name}{i}"
            i += 1

        # Add the new core level data to window.Data
        window.Data['Core levels'][new_sheet_name] = deepcopy(clipboard_data['core_level'])
        window.Data['Number of Core levels'] += 1

        # Create a new Excel sheet with the data
        core_level_data = window.Data['Core levels'][new_sheet_name]

        # Get data length for this core level
        data_length = len(core_level_data['B.E.'])

        # Use original column names if available, otherwise use defaults
        column_names = clipboard_data.get('column_names',
                                          ['Binding Energy', 'Corrected Data', 'Raw Data', 'Transmission'])

        # Make sure we have enough column names
        while len(column_names) < 4:
            column_names.append(f"Column{len(column_names) + 1}")

        # Create DataFrame with original column names
        data_dict = {
            column_names[0]: core_level_data['B.E.'],
            column_names[1]: core_level_data['Raw Data']
        }

        # Use the exact C and D data from clipboard if available
        if 'column_C_data' in clipboard_data and len(column_names) > 2:
            col_c_data = clipboard_data['column_C_data']
            # Adjust length if needed
            if len(col_c_data) != data_length:
                if len(col_c_data) > data_length:
                    col_c_data = col_c_data[:data_length]
                else:
                    col_c_data = col_c_data + [None] * (data_length - len(col_c_data))
            data_dict[column_names[2]] = col_c_data
        elif 'Background' in core_level_data and 'Bkg Y' in core_level_data['Background']:
            data_dict[column_names[2]] = core_level_data['Background']['Bkg Y']
        else:
            data_dict[column_names[2]] = [0.0] * data_length  # Default to zeros

        if 'column_D_data' in clipboard_data and len(column_names) > 3:
            col_d_data = clipboard_data['column_D_data']
            # Adjust length if needed
            if len(col_d_data) != data_length:
                if len(col_d_data) > data_length:
                    col_d_data = col_d_data[:data_length]
                else:
                    col_d_data = col_d_data + [None] * (data_length - len(col_d_data))
            data_dict[column_names[3]] = col_d_data
        else:
            data_dict[column_names[3]] = [1.0] * data_length

        df = pd.DataFrame(data_dict)

        # Write to Excel
        with pd.ExcelWriter(window.Data['FilePath'], engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

            # Add experimental description data if available
            if 'experimental_description' in clipboard_data and clipboard_data['experimental_description']:
                # Calculate column 'AX' index (typically 49)
                exp_col = 49

                # Get the workbook and sheet
                workbook = writer.book
                worksheet = workbook[new_sheet_name]

                # Add experimental description header
                worksheet.cell(row=1, column=exp_col + 1, value="Experimental Description")
                worksheet.cell(row=1, column=exp_col + 2, value="Value")

                # Add all experimental description data
                for i, item in enumerate(clipboard_data.get('experimental_description', [])):
                    if isinstance(item, list) and len(item) >= 2:
                        worksheet.cell(row=i + 2, column=exp_col + 1, value=item[0])
                        worksheet.cell(row=i + 2, column=exp_col + 2, value=item[1])

        # Update the JSON file
        json_file_path = os.path.splitext(window.Data['FilePath'])[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Load the updated workbook
        wb = openpyxl.load_workbook(window.Data['FilePath'])

        # Remove "Sheet" if it exists and not the only sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
            wb.save(window.Data['FilePath'])

        # Update sheet list in UI
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(wb.sheetnames)

        # Make sure the new sheet name is valid
        if new_sheet_name in wb.sheetnames:
            window.sheet_combobox.SetValue(new_sheet_name)

            # Initialize plot limits for the new sheet before switching to it
            if hasattr(window, 'plot_config') and not window.plot_config.plot_limits.get(new_sheet_name):
                window.plot_config.update_plot_limits(window, new_sheet_name)

            # Switch to the new sheet
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, new_sheet_name)

            # Set the peak parameters
            if 'peak_params_grid' in clipboard_data:
                try:
                    window.peak_params_grid.ClearGrid()
                    set_grid_data(window.peak_params_grid, clipboard_data['peak_params_grid'])
                    window.peak_count = clipboard_data.get('peak_count', 0)
                except Exception as grid_err:
                    print(f"Error setting peak parameters: {grid_err}")
        else:
            # Fallback if new sheet wasn't created properly
            if wb.sheetnames:
                window.sheet_combobox.SetValue(wb.sheetnames[0])
                from libraries.Sheet_Operations import on_sheet_selected
                on_sheet_selected(window, wb.sheetnames[0])

        # Refresh sheets after pasting
        try:
            from libraries.Sheet_Operations import on_sheet_selected
            refresh_sheets(window, on_sheet_selected, update_console)
        except Exception as refresh_err:
            print(f"Error refreshing sheets: {refresh_err}")

        # Update the plot
        try:
            window.clear_and_replot()
        except Exception as plot_err:
            print(f"Error updating plot: {plot_err}")

        window.show_popup_message2("Core Level Pasted", f"Core level data pasted as new sheet '{new_sheet_name}'"
                                                        f"\nCreated backup of the original data in Backup folder"
                                                        f"\nBeware that core level retain their BE correction values")

    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Paste Failed", f"Error: {str(e)}")


def copy_all_peak_parameters(window):
    """Copy all peak parameters from the peak_params_grid and Data structure."""
    import json
    import os
    import tempfile
    from copy import deepcopy

    sheet_name = window.sheet_combobox.GetValue()
    if sheet_name not in window.Data['Core levels'] or 'Fitting' not in window.Data['Core levels'][sheet_name]:
        window.show_popup_message2("Copy Failed", "No peak parameters to copy")
        return

    clipboard_data = {
        'sheet_name': sheet_name,
        'grid_data': [],
        'peak_data': deepcopy(window.Data['Core levels'][sheet_name]['Fitting']),
        'peak_count': window.peak_count
    }

    # Get all grid data
    for row in range(window.peak_params_grid.GetNumberRows()):
        row_data = []
        for col in range(window.peak_params_grid.GetNumberCols()):
            row_data.append(window.peak_params_grid.GetCellValue(row, col))
        clipboard_data['grid_data'].append(row_data)

    # Save to temp file
    clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_peak_clipboard.json')
    with open(clipboard_file, 'w') as f:
        json.dump(convert_to_serializable_and_round(clipboard_data), f)

    window.show_popup_message2("Parameters Copied", f"All peak parameters from {sheet_name} copied")


def paste_all_peak_parameters(window):
    """Paste all peak parameters to the peak_params_grid and Data structure."""
    import json
    import os
    import tempfile
    from libraries.Sheet_Operations import on_sheet_selected

    clipboard_file = os.path.join(tempfile.gettempdir(), 'khervefitting_peak_clipboard.json')
    if not os.path.exists(clipboard_file):
        return

    save_state(window)

    try:
        with open(clipboard_file, 'r') as f:
            clipboard_data = json.load(f)

        sheet_name = window.sheet_combobox.GetValue()

        # Update window.Data structure
        if 'peak_data' in clipboard_data:
            # Ensure Fitting exists in destination
            if 'Fitting' not in window.Data['Core levels'][sheet_name]:
                window.Data['Core levels'][sheet_name]['Fitting'] = {}

            # Update Fitting data (but preserve Peaks key if it exists)
            for key, value in clipboard_data['peak_data'].items():
                if key != 'Peaks':
                    window.Data['Core levels'][sheet_name]['Fitting'][key] = value

            # Handle Peaks - preserve original peak names
            if 'Peaks' in clipboard_data['peak_data']:
                # Directly copy the Peaks dictionary without renaming
                window.Data['Core levels'][sheet_name]['Fitting']['Peaks'] = {}

                # Just copy all peaks with their original names
                for peak_key, peak_data in clipboard_data['peak_data']['Peaks'].items():
                    window.Data['Core levels'][sheet_name]['Fitting']['Peaks'][peak_key] = peak_data.copy()

        # Update peak count
        window.peak_count = len(clipboard_data['peak_data']['Peaks']) if 'peak_data' in clipboard_data and 'Peaks' in \
                                                                         clipboard_data['peak_data'] else 0

        # Use on_sheet_selected to refresh grid with proper formatting
        on_sheet_selected(window, sheet_name)

        window.update_ratios()
        window.clear_and_replot()

        window.show_popup_message2("Peak Table Pasted", f"All peak parameters pasted to {sheet_name}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        window.show_popup_message2("Paste Failed", f"Error: {str(e)}")

def on_backup_main(window):
    """Create a backup of the current Excel and JSON files from main toolbar"""
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file currently open to backup.", "Error", wx.OK | wx.ICON_ERROR)
        return

    # Get current file paths
    excel_file = window.Data['FilePath']
    json_file = os.path.splitext(excel_file)[0] + '.json'

    # Check if files exist
    if not os.path.exists(excel_file):
        wx.MessageBox(f"Excel file not found: {excel_file}", "Error", wx.OK | wx.ICON_ERROR)
        return

    # Create backup folder in the executable directory
    import sys
    executable_dir = os.path.dirname(os.path.abspath(sys.executable))
    # For development environment, fall back to current script directory
    if not "KherveFitting" in executable_dir:
        executable_dir = os.path.dirname(os.path.abspath(__file__))
        # Go up one level if in libraries folder
        if os.path.basename(executable_dir) == "libraries":
            executable_dir = os.path.dirname(executable_dir)

    backup_folder = os.path.join(executable_dir, "Backup")
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)

    # Generate timestamp in format: YYcDD_HHMMSS
    import datetime
    now = datetime.datetime.now()
    month_letter = chr(ord('a') + now.month - 1)  # a=Jan, b=Feb, c=Mar, etc.
    timestamp = f"{now.year % 100}{month_letter}{now.day:02d}_{now.hour:02d}{now.minute:02d}{now.second:02d}"

    # Create backup filenames
    excel_filename = os.path.basename(excel_file)
    excel_backup = os.path.join(backup_folder, f"{excel_filename}_{timestamp}")

    # Copy the Excel file
    try:
        shutil.copy2(excel_file, excel_backup)
        files_backed_up = [excel_backup]

        # Copy the JSON file if it exists
        if os.path.exists(json_file):
            json_filename = os.path.basename(json_file)
            json_backup = os.path.join(backup_folder, f"{json_filename}_{timestamp}")
            shutil.copy2(json_file, json_backup)
            files_backed_up.append(json_backup)

        # Show success message
        window.show_popup_message2("Backup Complete", f"Backup created successfully:\n" + "\n".join(files_backed_up))

    except Exception as e:
        window.show_popup_message2("Error", f"Error creating backup: {str(e)}")