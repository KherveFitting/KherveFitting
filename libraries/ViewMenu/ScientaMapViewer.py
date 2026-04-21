# KherveFitting - XPS Data Analysis Software
# Copyright (C) 2024-2026 Gwilherm Kerherve <g.kerherve@ic.ac.uk>
#
# KherveFitting is dual-licensed:
#   - GNU GPL v3.0 (see LICENSE-GPL.txt) for open-source use
#   - Commercial Licence (see LICENSE-COMMERCIAL.txt) for proprietary use
# SPDX-License-Identifier: GPL-3.0-only OR LicenseRef-KherveFitting-Commercial

"""
Scienta Map Viewer Window for KherveFitting

Opens when clicking on a XPS~Map sheet in Sample Manager or selecting from combobox.
Displays a single map with tools for:
- Rectangle selection to sum and crop sweeps
- Dropping sweeps via right-click or line tool
- Binning options
- Various colormaps
- Save as XPS Plot
"""

import os
import re
import numpy as np
import wx
import json
import openpyxl
import matplotlib
matplotlib.use('WXAgg')
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.widgets import RectangleSelector
import matplotlib.patches as patches


class ScientaMapViewerWindow(wx.Frame):
    """
    Window to view and manipulate a single Scienta map from window.Data.
    """

    def __init__(self, parent, map_sheet_name):
        """
        Initialize the map viewer window.

        Args:
            parent: Parent window (main KherveFitting window)
            map_sheet_name: Name of the map sheet (e.g., 'XPS~Map')
        """
        # Get base name early for title
        map_data_temp = parent.Data['Core levels'].get(map_sheet_name, {})
        base_name_temp = map_data_temp.get('_core_level', '') or map_data_temp.get('ExperimentalInfo', {}).get('Core Level', '')
        if not base_name_temp:
            region_name = map_data_temp.get('ExperimentalInfo', {}).get('Region Name', '')
            if region_name:
                base_name_temp = region_name.strip().replace(' ', '')
                for suffix in ['_fast2', '_fast', '_slow', '_quick', '_scan']:
                    if base_name_temp.endswith(suffix):
                        base_name_temp = base_name_temp[:-len(suffix)]
                        break
            else:
                base_name_temp = 'Spectrum'

        super().__init__(parent, title=f"XPS HeatMap - {base_name_temp}",
                        size=(650, 600), style=wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP)

        self.parent = parent
        self.map_sheet_name = map_sheet_name

        # Extract map data from window.Data
        if map_sheet_name not in parent.Data['Core levels']:
            wx.MessageBox(f"Map sheet '{map_sheet_name}' not found in data.",
                         "Error", wx.OK | wx.ICON_ERROR)
            self.Destroy()
            return

        map_data = parent.Data['Core levels'][map_sheet_name]

        # Get BE values
        self.be_values = np.array(map_data.get('B.E.', []))

        # Build 2D data array from Y columns
        self.num_sweeps = map_data.get('_num_sweeps', 0)
        if self.num_sweeps == 0:
            # Count Y columns
            self.num_sweeps = sum(1 for key in map_data.keys() if key.startswith('Y') and key[1:].isdigit())

        if self.num_sweeps == 0 or len(self.be_values) == 0:
            wx.MessageBox(f"Invalid map data in '{map_sheet_name}'.",
                         "Error", wx.OK | wx.ICON_ERROR)
            self.Destroy()
            return

        # Build data_2d array (num_sweeps x num_be_points)
        self.data_2d_original = np.zeros((self.num_sweeps, len(self.be_values)))
        for i in range(self.num_sweeps):
            col_name = f'Y{i + 1}'
            if col_name in map_data:
                self.data_2d_original[i, :] = np.array(map_data[col_name])

        # Working copy (may be binned)
        self.data_2d = self.data_2d_original.copy()
        self.be_values_display = self.be_values.copy()
        self.num_sweeps_display = self.num_sweeps

        self.metadata = map_data.get('ExperimentalInfo', {})
        self.dropped_sweeps = set()
        self.undo_stack = []  # For undo functionality

        # Get base name (core level) from stored data
        self.base_name = map_data.get('_core_level', '')
        if not self.base_name:
            self.base_name = self.metadata.get('Core Level', '')
        if not self.base_name:
            region_name = self.metadata.get('Region Name', '')
            if region_name:
                self.base_name = self._clean_region_name(region_name)
            else:
                self.base_name = 'Spectrum'

        # Selection state
        self.selection_mode = None  # 'zoom', 'pan', 'rectangle', 'drop'
        self.rect_selector = None
        self.selected_rect = None  # (be_min, be_max, sweep_min, sweep_max)
        self.colormap = 'Greens'
        self.bin_factor = 1

        self.normalize = True
        self._cbar = None
        self._cbar_ax = None

        # Drop line state
        self.drop_line = None
        self.drop_text = None
        self.current_mouse_sweep = None

        # Get icon path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.icon_path = os.path.join(current_dir, "..", "Icons")
        if not os.path.exists(self.icon_path):
            self.icon_path = os.path.join(current_dir, "Icons")

        self.init_ui()
        self.update_heatmap()

        self.Centre()
        self.Show()

        # Bind close event
        self.Bind(wx.EVT_CLOSE, self.on_close)

    def _clean_region_name(self, region_name):
        """Clean region name to get base core level name."""
        name = region_name.strip()
        for suffix in ['_fast2', '_fast', '_slow', '_quick', '_scan']:
            if name.endswith(suffix):
                name = name[:-len(suffix)]
                break
        name = name.replace(' ', '')
        return name

    def init_ui(self):
        """Initialize the user interface."""
        panel = wx.Panel(self, style=wx.BORDER_RAISED)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Toolbar at top
        toolbar_panel = self.create_toolbar(panel)
        main_sizer.Add(toolbar_panel, 0, wx.EXPAND | wx.ALL, 0)

        # Main content - horizontal split
        content_sizer = wx.BoxSizer(wx.HORIZONTAL)

        # Left panel - info and controls
        left_panel = wx.Panel(panel, style=wx.BORDER_SIMPLE)
        left_sizer = wx.BoxSizer(wx.VERTICAL)

        # Sweep info
        sweep_panel = wx.Panel(left_panel)
        sweep_sizer = wx.BoxSizer(wx.VERTICAL)

        self.sweep_label = wx.StaticText(sweep_panel, label=f"Sweeps: {self.num_sweeps}")
        sweep_sizer.Add(self.sweep_label, 0, wx.ALL, 5)

        self.dropped_label = wx.StaticText(sweep_panel, label="Dropped: 0")
        sweep_sizer.Add(self.dropped_label, 0, wx.ALL, 5)

        self.active_label = wx.StaticText(sweep_panel, label=f"Active: {self.num_sweeps}")
        sweep_sizer.Add(self.active_label, 0, wx.ALL, 5)

        sweep_panel.SetSizer(sweep_sizer)
        left_sizer.Add(sweep_panel, 0, wx.EXPAND | wx.ALL, 5)

        # Normalisation control
        left_sizer.Add(wx.StaticLine(left_panel), 0, wx.EXPAND | wx.ALL, 5)

        self.norm_checkbox = wx.CheckBox(left_panel, label="Norm. (0-1000)")
        self.norm_checkbox.SetValue(True)
        self.norm_checkbox.SetToolTip("Normalise each sweep to 0-1000 for display")
        self.norm_checkbox.Bind(wx.EVT_CHECKBOX, self.on_norm_toggle)
        left_sizer.Add(self.norm_checkbox, 0, wx.ALL, 5)

        # Binning control
        left_sizer.Add(wx.StaticLine(left_panel), 0, wx.EXPAND | wx.ALL, 5)

        bin_label = wx.StaticText(left_panel, label="Bin (group sweeps):")
        left_sizer.Add(bin_label, 0, wx.LEFT | wx.TOP, 5)

        self.bin_spin = wx.SpinCtrl(left_panel, min=1, max=self.num_sweeps, initial=1, size=(60, -1))
        self.bin_spin.Bind(wx.EVT_SPINCTRL, self.on_bin_change)
        left_sizer.Add(self.bin_spin, 0, wx.ALL, 5)

        self.bin_info_label = wx.StaticText(left_panel, label=f"-> {self.num_sweeps} spectra")
        left_sizer.Add(self.bin_info_label, 0, wx.LEFT | wx.BOTTOM, 5)

        # Selection info
        left_sizer.Add(wx.StaticLine(left_panel), 0, wx.EXPAND | wx.ALL, 5)

        self.selection_label = wx.StaticText(left_panel, label="Selection: None")
        left_sizer.Add(self.selection_label, 0, wx.ALL, 5)

        # Action buttons
        left_sizer.Add(wx.StaticLine(left_panel), 0, wx.EXPAND | wx.ALL, 5)

        self.sum_all_btn = wx.Button(left_panel, label="Export\nSummed Spectrum")
        self.sum_all_btn.SetMinSize((-1, 35))
        self.sum_all_btn.SetToolTip("Sum all active sweeps")
        self.sum_all_btn.Bind(wx.EVT_BUTTON, self.on_sum_all)
        left_sizer.Add(self.sum_all_btn, 0, wx.EXPAND | wx.ALL, 3)

        self.sum_selected_btn = wx.Button(left_panel, label="Export\nSummed Selected")
        self.sum_selected_btn.SetMinSize((-1, 35))
        self.sum_selected_btn.SetToolTip("Sum sweeps in selected rectangle")
        self.sum_selected_btn.Bind(wx.EVT_BUTTON, self.on_sum_selection)
        self.sum_selected_btn.Enable(False)
        left_sizer.Add(self.sum_selected_btn, 0, wx.EXPAND | wx.ALL, 3)

        self.export_binned_btn = wx.Button(left_panel, label="Export\nBinned Spectra")
        self.export_binned_btn.SetMinSize((-1, 35))
        self.export_binned_btn.SetToolTip("Export binned (grouped) spectra")
        self.export_binned_btn.Bind(wx.EVT_BUTTON, self.on_export_binned)
        left_sizer.Add(self.export_binned_btn, 0, wx.EXPAND | wx.ALL, 3)

        self.export_all_btn = wx.Button(left_panel, label="Export\nAll Individuals")
        self.export_all_btn.SetMinSize((-1, 35))
        self.export_all_btn.SetToolTip("Export each sweep as separate spectrum")
        self.export_all_btn.Bind(wx.EVT_BUTTON, self.on_export_all_spectra)
        left_sizer.Add(self.export_all_btn, 0, wx.EXPAND | wx.ALL, 3)

        self.export_selected_btn = wx.Button(left_panel, label="Export\nIndividuals Selected")
        self.export_selected_btn.SetMinSize((-1, 35))
        self.export_selected_btn.SetToolTip("Export selected spectra")
        self.export_selected_btn.Bind(wx.EVT_BUTTON, self.on_export_selected)
        self.export_selected_btn.Enable(False)
        left_sizer.Add(self.export_selected_btn, 0, wx.EXPAND | wx.ALL, 3)

        left_panel.SetSizer(left_sizer)
        left_panel.SetMinSize((130, -1))
        content_sizer.Add(left_panel, 0, wx.EXPAND | wx.ALL, 0)

        # Right panel - heatmap
        right_panel = wx.Panel(panel)
        right_sizer = wx.BoxSizer(wx.VERTICAL)

        # Create figure for heatmap
        self.figure = Figure(figsize=(5, 5), dpi=100)
        self.figure.set_tight_layout(False)
        self.canvas = FigureCanvas(right_panel, -1, self.figure)
        self.ax = self.figure.add_subplot(111)

        right_sizer.Add(self.canvas, 1, wx.EXPAND | wx.ALL, 0)

        right_panel.SetSizer(right_sizer)
        content_sizer.Add(right_panel, 1, wx.EXPAND | wx.ALL, 0)

        main_sizer.Add(content_sizer, 1, wx.EXPAND, 0)

        panel.SetSizer(main_sizer)

        # Bind canvas events
        self.canvas.mpl_connect('button_press_event', self.on_canvas_click)
        self.canvas.mpl_connect('motion_notify_event', self.on_canvas_motion)

    def create_toolbar(self, parent):
        """Create toolbar with icon buttons like EDX~Map window."""
        toolbar_panel = wx.Panel(parent, style=wx.BORDER_SIMPLE)
        toolbar_sizer = wx.BoxSizer(wx.HORIZONTAL)

        btn_size = (25, 25)

        # Navigation tools
        self.zoom_in_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.zoom_out_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        self.pan_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)

        # Load zoom icons with fallback
        zoom_in_path = os.path.join(self.icon_path, "ZoomIN-3.png")
        if os.path.exists(zoom_in_path):
            self.zoom_in_btn.SetBitmap(wx.Bitmap(zoom_in_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_in_btn.SetBitmap(self.create_icon_bitmap('zoom_in'))

        zoom_out_path = os.path.join(self.icon_path, "ZoomOUT-3.png")
        if os.path.exists(zoom_out_path):
            self.zoom_out_btn.SetBitmap(wx.Bitmap(zoom_out_path, wx.BITMAP_TYPE_PNG))
        else:
            self.zoom_out_btn.SetBitmap(self.create_icon_bitmap('zoom_out'))

        pan_path = os.path.join(self.icon_path, "Drag-25.png")
        if os.path.exists(pan_path):
            self.pan_btn.SetBitmap(wx.Bitmap(pan_path, wx.BITMAP_TYPE_PNG))
        else:
            self.pan_btn.SetBitmap(self.create_icon_bitmap('pan'))

        self.zoom_in_btn.SetToolTip("Zoom In")
        self.zoom_out_btn.SetToolTip("Zoom Out / Home")
        self.pan_btn.SetToolTip("Pan")

        toolbar_sizer.Add(self.zoom_in_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.zoom_out_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.pan_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL), 0, wx.EXPAND | wx.ALL, 3)

        # Selection tools
        self.rect_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.line_btn = wx.BitmapToggleButton(toolbar_panel, size=btn_size)
        self.clear_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        self.undo_btn = wx.BitmapButton(toolbar_panel, size=btn_size)

        self.rect_btn.SetBitmap(self.create_icon_bitmap('area'))
        self.line_btn.SetBitmap(self.create_icon_bitmap('line'))
        self.clear_btn.SetBitmap(self.create_icon_bitmap('clear'))
        self.undo_btn.SetBitmap(self.create_icon_bitmap('undo'))

        self.rect_btn.SetToolTip("Rectangle Selection")
        self.line_btn.SetToolTip("Drop Sweep Mode")
        self.clear_btn.SetToolTip("Clear Selections")
        self.undo_btn.SetToolTip("Undo Last Drop")

        toolbar_sizer.Add(self.rect_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.line_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.clear_btn, 0, wx.ALL, 2)
        toolbar_sizer.Add(self.undo_btn, 0, wx.ALL, 2)

        toolbar_sizer.Add(wx.StaticLine(toolbar_panel, style=wx.LI_VERTICAL), 0, wx.EXPAND | wx.ALL, 3)

        # Colormap button
        self.cmap_btn = wx.BitmapButton(toolbar_panel, size=btn_size)
        heatmap_path = os.path.join(self.icon_path, "heatmap-3.png")
        if os.path.exists(heatmap_path):
            self.cmap_btn.SetBitmap(wx.Bitmap(heatmap_path, wx.BITMAP_TYPE_PNG))
        else:
            self.cmap_btn.SetBitmap(self.create_icon_bitmap('intensity'))
        self.cmap_btn.SetToolTip("Change Colormap")
        toolbar_sizer.Add(self.cmap_btn, 0, wx.ALL, 2)

        toolbar_panel.SetSizer(toolbar_sizer)

        # Bind events
        self.zoom_in_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_zoom_in)
        self.zoom_out_btn.Bind(wx.EVT_BUTTON, self.on_zoom_out)
        self.pan_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_pan)
        self.rect_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_rect_mode)
        self.line_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_line_mode)
        self.clear_btn.Bind(wx.EVT_BUTTON, self.on_clear_selection)
        self.undo_btn.Bind(wx.EVT_BUTTON, self.on_undo)
        self.cmap_btn.Bind(wx.EVT_BUTTON, self.on_colormap_button)

        return toolbar_panel

    def create_icon_bitmap(self, icon_type):
        """Create simple icon bitmaps as fallback."""
        size = 20
        bmp = wx.Bitmap(size, size)
        dc = wx.MemoryDC(bmp)
        dc.SetBackground(wx.Brush(wx.Colour(240, 240, 240)))
        dc.Clear()

        dc.SetPen(wx.Pen(wx.Colour(60, 60, 60), 2))

        if icon_type == 'zoom_in':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(14, 14, 18, 18)
            dc.DrawLine(7, 10, 13, 10)
            dc.DrawLine(10, 7, 10, 13)
        elif icon_type == 'zoom_out':
            dc.DrawCircle(10, 10, 6)
            dc.DrawLine(14, 14, 18, 18)
            dc.DrawLine(7, 10, 13, 10)
        elif icon_type == 'pan':
            dc.DrawLine(10, 3, 10, 17)
            dc.DrawLine(3, 10, 17, 10)
            dc.DrawLine(10, 3, 7, 6)
            dc.DrawLine(10, 3, 13, 6)
        elif icon_type == 'area':
            dc.SetBrush(wx.TRANSPARENT_BRUSH)
            dc.DrawRectangle(3, 3, 14, 14)
        elif icon_type == 'line':
            # Horizontal line with X for drop
            dc.SetPen(wx.Pen(wx.Colour(200, 0, 0), 2))
            dc.DrawLine(3, 10, 17, 10)
            dc.DrawLine(8, 6, 12, 14)
            dc.DrawLine(8, 14, 12, 6)
        elif icon_type == 'clear':
            dc.SetPen(wx.Pen(wx.Colour(200, 0, 0), 2))
            dc.DrawLine(4, 4, 16, 16)
            dc.DrawLine(16, 4, 4, 16)
        elif icon_type == 'undo':
            dc.SetPen(wx.Pen(wx.Colour(60, 60, 60), 2))
            dc.DrawArc(5, 8, 15, 8, 10, 12)
            dc.DrawLine(5, 5, 5, 10)
            dc.DrawLine(5, 5, 10, 5)
        elif icon_type == 'intensity':
            for i in range(4):
                gray = 255 - i * 60
                dc.SetBrush(wx.Brush(wx.Colour(0, gray, 0)))
                dc.SetPen(wx.Pen(wx.Colour(0, gray, 0)))
                dc.DrawRectangle(3 + i * 4, 3, 4, 14)

        dc.SelectObject(wx.NullBitmap)
        return bmp

    def on_bin_change(self, event):
        """Handle binning change - groups sweeps together."""
        self.bin_factor = self.bin_spin.GetValue()

        # Calculate number of output spectra when binning sweeps
        active_sweeps = self.num_sweeps - len(self.dropped_sweeps)
        num_binned = active_sweeps // self.bin_factor if self.bin_factor > 0 else active_sweeps
        self.bin_info_label.SetLabel(f"-> {num_binned} spectra")

    def apply_binning(self):
        """Apply binning to the data."""
        if self.bin_factor <= 1:
            self.data_2d = self.data_2d_original.copy()
            self.be_values_display = self.be_values.copy()
            self.num_sweeps_display = self.num_sweeps
            return

        # Bin the BE axis
        n_be = len(self.be_values)
        n_be_binned = n_be // self.bin_factor

        if n_be_binned == 0:
            self.bin_spin.SetValue(1)
            self.bin_factor = 1
            self.data_2d = self.data_2d_original.copy()
            self.be_values_display = self.be_values.copy()
            return

        # Bin BE values (average)
        self.be_values_display = np.array([
            np.mean(self.be_values[i*self.bin_factor:(i+1)*self.bin_factor])
            for i in range(n_be_binned)
        ])

        # Bin intensity data (sum)
        self.data_2d = np.zeros((self.num_sweeps, n_be_binned))
        for i in range(n_be_binned):
            self.data_2d[:, i] = np.sum(
                self.data_2d_original[:, i*self.bin_factor:(i+1)*self.bin_factor],
                axis=1
            )

        self.num_sweeps_display = self.num_sweeps

    def update_heatmap(self):
        """Update the heatmap display."""
        self.ax.clear()

        # Remove old colorbar axes completely before redrawing
        if hasattr(self, '_cbar_ax') and self._cbar_ax is not None:
            try:
                self.figure.delaxes(self._cbar_ax)
            except Exception:
                pass
            self._cbar_ax = None
            self._cbar = None

        # Get data - mask dropped sweeps with NaN
        data = self.data_2d.copy().astype(float)
        for dropped_idx in self.dropped_sweeps:
            if dropped_idx < data.shape[0]:
                data[dropped_idx, :] = np.nan

        # Per-sweep normalisation to 0-1000
        if getattr(self, 'normalize', True):
            for i in range(data.shape[0]):
                row = data[i, :]
                finite = row[np.isfinite(row)]
                if len(finite) == 0:
                    continue
                r_min, r_max = finite.min(), finite.max()
                if r_max - r_min > 0:
                    data[i, :] = np.where(np.isfinite(row),
                                          (row - r_min) / (r_max - r_min) * 1000.0,
                                          np.nan)
                else:
                    data[i, :] = np.where(np.isfinite(row), 500.0, np.nan)
            vmin, vmax = 0, 1000
            cbar_label = 'Norm. Intensity (0-1000)'
        else:
            finite_all = data[np.isfinite(data)]
            vmin = float(finite_all.min()) if len(finite_all) else 0
            vmax = float(finite_all.max()) if len(finite_all) else 1
            cbar_label = 'Intensity (CPS)'

        be_min, be_max = self.be_values_display.min(), self.be_values_display.max()
        be_descending = (self.be_values_display[0] > self.be_values_display[-1]
                         if len(self.be_values_display) > 1 else False)

        # Fix main axes position before drawing
        # self.ax.set_position([0.12, 0.1, 0.68, 0.85])
        # Full-width axes (no colorbar)
        # Full-width axes (no colorbar) — set_position must be after ax.clear()
        self.ax.set_position([0.10, 0.10, 0.87, 0.87])

        if be_descending:
            self.heatmap = self.ax.imshow(
                data, aspect='auto', origin='lower',
                extent=[self.be_values_display[0], self.be_values_display[-1],
                        0, self.num_sweeps_display],
                cmap=self.colormap, vmin=vmin, vmax=vmax)
        else:
            data_flipped = np.fliplr(data)
            self.heatmap = self.ax.imshow(
                data_flipped, aspect='auto', origin='lower',
                extent=[be_max, be_min, 0, self.num_sweeps_display],
                cmap=self.colormap, vmin=vmin, vmax=vmax)

        self.ax.set_ylim(0, self.num_sweeps_display)
        self.ax.set_xlabel('Binding Energy (eV)', fontsize=10)
        self.ax.set_ylabel('Sweep Number', fontsize=10)

        # No colorbar — norm state shown via checkbox label instead

        # Draw selection rectangle if exists
        if self.selected_rect:
            be_lo, be_hi, sw_lo, sw_hi = self.selected_rect
            rect = patches.Rectangle((be_hi, sw_lo), be_lo - be_hi, sw_hi - sw_lo,
                                     linewidth=2, edgecolor='blue', facecolor='none',
                                     linestyle='-')
            self.ax.add_patch(rect)

        # Drop line (hidden initially)
        self.drop_line = self.ax.axhline(y=-10, color='red', linewidth=1.5, linestyle='--', visible=False)
        self.drop_text = self.ax.text(be_max, -10, '', color='red', fontsize=8, ha='left', visible=False)

        self.canvas.draw()
        self.update_labels()

    def update_labels(self):
        """Update info labels."""
        num_dropped = len(self.dropped_sweeps)
        num_active = self.num_sweeps - num_dropped

        self.dropped_label.SetLabel(f"Dropped: {num_dropped}")
        self.active_label.SetLabel(f"Active: {num_active}")

        if self.selected_rect:
            be_lo, be_hi, sw_lo, sw_hi = self.selected_rect
            self.selection_label.SetLabel(f"BE: {be_lo:.1f}-{be_hi:.1f}\nSweeps: {int(sw_lo)}-{int(sw_hi)}")
            self.sum_selected_btn.Enable(True)
            self.export_selected_btn.Enable(True)
        else:
            self.selection_label.SetLabel("Selection: None")
            self.sum_selected_btn.Enable(False)
            self.export_selected_btn.Enable(False)

    def on_colormap_button(self, event):
        """Show colormap selection menu."""
        menu = wx.Menu()
        cmaps = ['Greens', 'viridis', 'plasma', 'inferno', 'magma', 'hot', 'cool', 'jet', 'gray', 'bone']
        for cmap in cmaps:
            item = menu.AppendRadioItem(wx.ID_ANY, cmap)
            if cmap == self.colormap:
                item.Check(True)
            self.Bind(wx.EVT_MENU, lambda e, c=cmap: self._set_colormap(c), item)

        self.PopupMenu(menu)
        menu.Destroy()

    def _set_colormap(self, cmap):
        """Set colormap."""
        self.colormap = cmap
        self.update_heatmap()

    def on_norm_toggle(self, event):
        """Toggle per-sweep normalisation and refresh heatmap."""
        self.normalize = self.norm_checkbox.GetValue()
        self.update_heatmap()

    def on_zoom_in(self, event):
        """Toggle zoom in mode."""
        if self.zoom_in_btn.GetValue():
            self.pan_btn.SetValue(False)
            self.rect_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'zoom'
            self._hide_drop_line()
            self._setup_zoom_selector()
        else:
            self.selection_mode = None
            self._clear_selectors()

    def on_zoom_out(self, event):
        """Reset zoom to full view."""
        be_min, be_max = self.be_values_display.min(), self.be_values_display.max()
        self.ax.set_xlim(be_max, be_min)
        self.ax.set_ylim(0, self.num_sweeps_display)
        self.canvas.draw()

    def on_pan(self, event):
        """Toggle pan mode."""
        if self.pan_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.rect_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'pan'
            self._hide_drop_line()
            self._clear_selectors()
        else:
            self.selection_mode = None

    def on_rect_mode(self, event):
        """Toggle rectangle selection mode."""
        if self.rect_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.line_btn.SetValue(False)
            self.selection_mode = 'rectangle'
            self._hide_drop_line()
            self._setup_rect_selector()
        else:
            self.selection_mode = None
            self._clear_selectors()

    def on_line_mode(self, event):
        """Toggle line/drop sweep mode."""
        if self.line_btn.GetValue():
            self.zoom_in_btn.SetValue(False)
            self.pan_btn.SetValue(False)
            self.rect_btn.SetValue(False)
            self.selection_mode = 'drop'
            self._clear_selectors()
            # Show drop line
            if self.drop_line:
                self.drop_line.set_visible(True)
                self.drop_text.set_visible(True)
                self.canvas.draw_idle()
        else:
            self.selection_mode = None
            self._hide_drop_line()

    def _hide_drop_line(self):
        """Hide the drop line."""
        if self.drop_line:
            self.drop_line.set_visible(False)
        if self.drop_text:
            self.drop_text.set_visible(False)
        self.canvas.draw_idle()

    def _setup_zoom_selector(self):
        """Setup zoom rectangle selector."""
        self._clear_selectors()
        self.rect_selector = RectangleSelector(
            self.ax, self._on_zoom_select,
            useblit=True,
            button=[1],
            minspanx=5, minspany=5,
            spancoords='pixels',
            interactive=False
        )

    def _setup_rect_selector(self):
        """Setup rectangle selection for summing."""
        self._clear_selectors()
        self.rect_selector = RectangleSelector(
            self.ax, self._on_rect_select,
            useblit=True,
            button=[1],
            minspanx=5, minspany=5,
            spancoords='pixels',
            interactive=True
        )

    def _clear_selectors(self):
        """Clear all selectors."""
        if self.rect_selector:
            self.rect_selector.set_active(False)
            self.rect_selector = None

    def _on_zoom_select(self, eclick, erelease):
        """Handle zoom selection."""
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        if x1 is None or x2 is None:
            return

        self.ax.set_xlim(max(x1, x2), min(x1, x2))
        self.ax.set_ylim(min(y1, y2), max(y1, y2))
        self.canvas.draw()

    def _on_rect_select(self, eclick, erelease):
        """Handle rectangle selection for summing."""
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        if x1 is None or x2 is None:
            return

        be_min, be_max = min(x1, x2), max(x1, x2)
        sw_min, sw_max = max(0.5, min(y1, y2)), min(self.num_sweeps_display + 0.5, max(y1, y2))

        self.selected_rect = (be_min, be_max, sw_min, sw_max)
        self.update_heatmap()

    def on_clear_selection(self, event):
        """Clear the current selection and reset all modes."""
        self.selected_rect = None

        # Deselect all toggle buttons
        self.zoom_in_btn.SetValue(False)
        self.pan_btn.SetValue(False)
        self.rect_btn.SetValue(False)
        self.line_btn.SetValue(False)

        # Clear selection mode
        self.selection_mode = None

        # Clear selectors
        self._clear_selectors()

        # Hide drop line
        self._hide_drop_line()

        # Disable selection-dependent buttons
        self.sum_selected_btn.Enable(False)
        self.export_selected_btn.Enable(False)

        # Update selection label
        self.selection_label.SetLabel("Selection: None")

        # Update display
        self.update_heatmap()

    def on_undo(self, event):
        """Undo last drop."""
        if self.undo_stack:
            last_dropped = self.undo_stack.pop()
            self.dropped_sweeps.discard(last_dropped)
            self.update_heatmap()

    def on_canvas_click(self, event):
        """Handle canvas click for right-click menu and drop mode."""
        if event.button == 3:  # Right click
            self._show_context_menu(event)
        elif event.button == 1 and event.inaxes == self.ax:
            if self.selection_mode == 'drop' and event.ydata is not None:
                # Drop sweep on click
                sweep_idx = max(0, min(self.num_sweeps - 1, int(event.ydata) - 1))
                if sweep_idx not in self.dropped_sweeps:
                    self.undo_stack.append(sweep_idx)
                    self.dropped_sweeps.add(sweep_idx)
                    self.update_heatmap()

    def on_canvas_motion(self, event):
        """Handle mouse motion for drop line."""
        if self.selection_mode == 'drop' and event.inaxes == self.ax and event.ydata is not None:
            sweep_idx = max(1, min(self.num_sweeps, int(event.ydata + 0.5)))

            if self.drop_line and self.drop_text:
                self.drop_line.set_ydata([sweep_idx, sweep_idx])
                self.drop_line.set_visible(True)

                be_max = self.be_values_display.max()
                be_min = self.be_values_display.min()
                self.drop_text.set_position((be_max - (be_max - be_min) * 0.02, sweep_idx + 0.2))
                self.drop_text.set_text(f"Sweep {sweep_idx}")
                self.drop_text.set_visible(True)

                self.canvas.draw_idle()

    def _show_context_menu(self, event):
        """Show right-click context menu."""
        menu = wx.Menu()

        # Style submenu
        style_menu = wx.Menu()
        for cmap in ['Greens', 'viridis', 'plasma', 'inferno', 'hot', 'jet', 'gray']:
            item = style_menu.AppendRadioItem(wx.ID_ANY, cmap)
            if cmap == self.colormap:
                item.Check(True)
            self.Bind(wx.EVT_MENU, lambda e, c=cmap: self._set_colormap(c), item)
        menu.AppendSubMenu(style_menu, "Style")

        menu.AppendSeparator()

        # Drop sweep (if clicked on a sweep)
        if event.ydata is not None:
            sweep_idx = int(event.ydata) - 1
            if 0 <= sweep_idx < self.num_sweeps:
                if sweep_idx in self.dropped_sweeps:
                    restore_item = menu.Append(wx.ID_ANY, f"Restore Sweep {sweep_idx + 1}")
                    self.Bind(wx.EVT_MENU, lambda e, s=sweep_idx: self._restore_sweep(s), restore_item)
                else:
                    drop_item = menu.Append(wx.ID_ANY, f"Drop Sweep {sweep_idx + 1}")
                    self.Bind(wx.EVT_MENU, lambda e, s=sweep_idx: self._drop_sweep(s), drop_item)

        menu.AppendSeparator()

        # Info
        info_item = menu.Append(wx.ID_ANY, "Info...")
        self.Bind(wx.EVT_MENU, self.on_show_info, info_item)

        # Data browser
        data_browser_item = menu.Append(wx.ID_ANY, "Data Browser...")
        self.Bind(wx.EVT_MENU, self.on_data_browser, data_browser_item)

        menu.AppendSeparator()

        # Clear selections
        clear_item = menu.Append(wx.ID_ANY, "Clear Selections")
        self.Bind(wx.EVT_MENU, lambda e: self.on_clear_selection(None), clear_item)

        menu.AppendSeparator()

        # Save options
        if self.selected_rect:
            save_sel_item = menu.Append(wx.ID_ANY, "Save Selection as XPS Plot...")
            self.Bind(wx.EVT_MENU, self.on_sum_selection, save_sel_item)

        save_all_item = menu.Append(wx.ID_ANY, "Sum All as XPS Plot...")
        self.Bind(wx.EVT_MENU, self.on_sum_all, save_all_item)

        self.PopupMenu(menu)
        menu.Destroy()

    def _drop_sweep(self, sweep_idx):
        """Drop a sweep."""
        self.undo_stack.append(sweep_idx)
        self.dropped_sweeps.add(sweep_idx)
        self.update_heatmap()

    def _restore_sweep(self, sweep_idx):
        """Restore a dropped sweep."""
        self.dropped_sweeps.discard(sweep_idx)
        self.update_heatmap()

    def on_show_info(self, event):
        """Show map info dialog."""
        info = f"Map: {self.map_sheet_name}\n"
        info += f"Core Level: {self.base_name}\n"
        info += f"Total Sweeps: {self.num_sweeps}\n"
        info += f"Active Sweeps: {self.num_sweeps - len(self.dropped_sweeps)}\n"
        info += f"Dropped Sweeps: {len(self.dropped_sweeps)}\n"
        info += f"BE Points: {len(self.be_values)}\n"
        info += f"BE Range: {self.be_values.min():.2f} - {self.be_values.max():.2f} eV\n"
        info += f"Bin Factor: {self.bin_factor}\n"

        if self.metadata:
            info += "\n--- Experimental Info ---\n"
            for key, value in self.metadata.items():
                info += f"{key}: {value}\n"

        wx.MessageBox(info, "Map Info", wx.OK | wx.ICON_INFORMATION)

    def on_data_browser(self, event):
        """Show data browser window."""
        browser = wx.Frame(self, title="Data Browser", size=(400, 500))

        text = wx.TextCtrl(browser, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)

        content = f"Map: {self.map_sheet_name}\n"
        content += f"Shape: {self.num_sweeps} sweeps x {len(self.be_values)} BE points\n\n"
        content += "BE Values:\n"
        content += ", ".join([f"{be:.2f}" for be in self.be_values[:20]])
        if len(self.be_values) > 20:
            content += f"... ({len(self.be_values)} total)\n"

        content += "\n\nIntensity Statistics:\n"
        content += f"Min: {np.nanmin(self.data_2d):.2f}\n"
        content += f"Max: {np.nanmax(self.data_2d):.2f}\n"
        content += f"Mean: {np.nanmean(self.data_2d):.2f}\n"

        text.SetValue(content)
        browser.Show()

    def on_sum_selection(self, event):
        """Sum sweeps within selection and save as XPS plot."""
        if not self.selected_rect:
            wx.MessageBox("No selection. Use rectangle tool to select an area.",
                         "No Selection", wx.OK | wx.ICON_WARNING)
            return

        be_lo, be_hi, sw_min, sw_max = self.selected_rect

        sweep_start = max(0, int(sw_min) - 1)
        sweep_end = min(self.num_sweeps, int(sw_max))
        valid_sweeps = [i for i in range(sweep_start, sweep_end) if i not in self.dropped_sweeps]

        if not valid_sweeps:
            wx.MessageBox("No active sweeps in selection.", "Error", wx.OK | wx.ICON_ERROR)
            return

        # Use original data for export
        be_mask = (self.be_values >= be_lo) & (self.be_values <= be_hi)
        be_indices = np.where(be_mask)[0]

        if len(be_indices) == 0:
            wx.MessageBox("No BE points in selection.", "Error", wx.OK | wx.ICON_ERROR)
            return

        selected_data = self.data_2d_original[valid_sweeps, :][:, be_indices]
        summed = np.sum(selected_data, axis=0)
        selected_be = self.be_values[be_indices]

        self._save_as_xps_plot(selected_be, summed, len(valid_sweeps),
                              f"Selection ({be_lo:.1f}-{be_hi:.1f} eV, {len(valid_sweeps)} sweeps)")

    def on_sum_all(self, event):
        """Sum all active sweeps and save as XPS plot."""
        valid_sweeps = [i for i in range(self.num_sweeps) if i not in self.dropped_sweeps]

        if not valid_sweeps:
            wx.MessageBox("No active sweeps.", "Error", wx.OK | wx.ICON_ERROR)
            return

        summed = np.sum(self.data_2d_original[valid_sweeps, :], axis=0)

        self._save_as_xps_plot(self.be_values, summed, len(valid_sweeps),
                              f"All sweeps ({len(valid_sweeps)} summed)")

    def on_export_all_spectra(self, event):
        """Export each sweep as separate spectrum."""
        valid_sweeps = [i for i in range(self.num_sweeps) if i not in self.dropped_sweeps]

        if not valid_sweeps:
            wx.MessageBox("No active sweeps.", "Error", wx.OK | wx.ICON_ERROR)
            return

        for sweep_idx in valid_sweeps:
            intensities = self.data_2d_original[sweep_idx, :]
            self._save_as_xps_plot(self.be_values, intensities, 1,
                                  f"Sweep {sweep_idx + 1}")

        wx.MessageBox(f"Exported {len(valid_sweeps)} spectra.", "Success", wx.OK | wx.ICON_INFORMATION)

    def on_export_summed(self, event):
        """Export summed spectrum."""
        self.on_sum_all(event)

    def on_export_binned(self, event):
        """Export binned (grouped) spectra - groups sweeps by bin_factor."""
        valid_sweeps = [i for i in range(self.num_sweeps) if i not in self.dropped_sweeps]

        if not valid_sweeps:
            wx.MessageBox("No active sweeps.", "Error", wx.OK | wx.ICON_ERROR)
            return

        if self.bin_factor <= 1:
            # No binning, just export all
            self.on_export_all_spectra(event)
            return

        # Group sweeps by bin factor
        num_bins = len(valid_sweeps) // self.bin_factor
        if num_bins == 0:
            wx.MessageBox(f"Bin factor {self.bin_factor} is too large for {len(valid_sweeps)} sweeps.",
                         "Error", wx.OK | wx.ICON_ERROR)
            return

        for bin_idx in range(num_bins):
            start_idx = bin_idx * self.bin_factor
            end_idx = start_idx + self.bin_factor
            sweeps_in_bin = valid_sweeps[start_idx:end_idx]

            # Sum sweeps in this bin
            binned_data = np.sum(self.data_2d_original[sweeps_in_bin, :], axis=0)

            self._save_as_xps_plot(self.be_values, binned_data, len(sweeps_in_bin),
                                  f"Binned ({len(sweeps_in_bin)} sweeps)")

        wx.MessageBox(f"Exported {num_bins} binned spectra.", "Success", wx.OK | wx.ICON_INFORMATION)

    def on_export_selected(self, event):
        """Export selected spectra."""
        if not self.selected_rect:
            wx.MessageBox("No selection.", "Error", wx.OK | wx.ICON_WARNING)
            return

        be_lo, be_hi, sw_min, sw_max = self.selected_rect

        sweep_start = max(0, int(sw_min) - 1)
        sweep_end = min(self.num_sweeps, int(sw_max))
        valid_sweeps = [i for i in range(sweep_start, sweep_end) if i not in self.dropped_sweeps]

        if not valid_sweeps:
            wx.MessageBox("No active sweeps in selection.", "Error", wx.OK | wx.ICON_ERROR)
            return

        be_mask = (self.be_values >= be_lo) & (self.be_values <= be_hi)
        be_indices = np.where(be_mask)[0]

        for sweep_idx in valid_sweeps:
            intensities = self.data_2d_original[sweep_idx, be_indices]
            self._save_as_xps_plot(self.be_values[be_indices], intensities, 1,
                                  f"Sweep {sweep_idx + 1} (selected)")

        wx.MessageBox(f"Exported {len(valid_sweeps)} selected spectra.", "Success", wx.OK | wx.ICON_INFORMATION)

    def _save_as_xps_plot(self, be_values, intensities, num_sweeps, description):
        """Save spectrum to parent window as XPS plot sheet."""
        if self.parent is None:
            return

        # Check if FileManager is open and store its position
        file_manager_was_open = False
        file_manager_position = None
        file_manager_size = None
        if hasattr(self.parent, 'file_manager') and self.parent.file_manager:
            try:
                file_manager_was_open = True
                file_manager_position = self.parent.file_manager.GetPosition()
                file_manager_size = self.parent.file_manager.GetSize()
                self.parent.file_manager.Close()
                self.parent.file_manager = None
            except:
                pass

        # Divide by number of sweeps to get averaged intensity
        if num_sweeps > 1:
            intensities = intensities / num_sweeps

        # Generate unique sheet name: base_name, base_name1, base_name2, etc.
        sheet_name = self.base_name
        existing = list(self.parent.Data['Core levels'].keys())
        if sheet_name in existing:
            counter = 1
            while f"{self.base_name}{counter}" in existing:
                counter += 1
            sheet_name = f"{self.base_name}{counter}"

        # Get experimental info from the map data
        map_data = self.parent.Data['Core levels'].get(self.map_sheet_name, {})
        source_exp_info = map_data.get('ExperimentalInfo', {})

        # Build comprehensive ExperimentalInfo
        exp_info = {
            'Sample ID': source_exp_info.get('Sample ID', ''),
            'Spectrum Name': source_exp_info.get('Spectrum Name', ''),
            'Region Name': source_exp_info.get('Region Name', ''),
            'Core Level': self.base_name,
            'Source Map': self.map_sheet_name,
            'Instrument': source_exp_info.get('Instrument', ''),
            'Location': source_exp_info.get('Location', ''),
            'User': source_exp_info.get('User', ''),
            'Date': source_exp_info.get('Date', ''),
            'Time': source_exp_info.get('Time', ''),
            'Technique': 'XPS',
            'Excitation Energy': source_exp_info.get('Excitation Energy', ''),
            'Pass Energy': source_exp_info.get('Pass Energy', ''),
            'Energy Scale': source_exp_info.get('Energy Scale', ''),
            'Lens Mode': source_exp_info.get('Lens Mode', ''),
            'Step Time': source_exp_info.get('Step Time', ''),
            'Number of Points': str(len(be_values)),
            'BE Start': f"{be_values[0]:.2f}",
            'BE End': f"{be_values[-1]:.2f}",
            'Total Sweeps in Map': source_exp_info.get('Number of Sweeps', str(self.num_sweeps)),
            'Sweeps Summed': str(num_sweeps),
            'Description': description,
        }

        core_level_data = {
            'Name': sheet_name,
            'B.E.': [round(float(be), 2) for be in be_values],
            'Raw Data': [round(float(val), 2) for val in intensities],
            'Corrected Data': [round(float(val), 2) for val in intensities],
            'Background': {
                'Bkg Type': '',
                'Bkg Low': round(float(min(be_values)), 2),
                'Bkg High': round(float(max(be_values)), 2),
                'Bkg Offset Low': 0,
                'Bkg Offset High': 0,
                'Bkg Y': [round(float(val), 2) for val in intensities]
            },
            'ExperimentalInfo': exp_info
        }

        self.parent.Data['Core levels'][sheet_name] = core_level_data
        self.parent.Data['Number of Core levels'] = len(self.parent.Data['Core levels'])

        self._add_sheet_to_excel(sheet_name, be_values, intensities, exp_info)

        if sheet_name not in [self.parent.sheet_combobox.GetString(i)
                              for i in range(self.parent.sheet_combobox.GetCount())]:
            self.parent.sheet_combobox.Append(sheet_name)

        self.parent.sheet_combobox.SetStringSelection(sheet_name)
        self.parent.current_sheet = sheet_name

        from libraries.Sheet_Operations import on_sheet_selected
        on_sheet_selected(self.parent, None)

        from libraries.FileMenu.Save import save_state
        save_state(self.parent)

        # Refresh all to update intensity display
        if hasattr(self.parent, 'refresh_all'):
            self.parent.refresh_all()

        # Restore FileManager if it was open
        if file_manager_was_open:
            try:
                from libraries.ViewMenu.FileManager import FileManagerWindow
                self.parent.file_manager = FileManagerWindow(self.parent)
                if file_manager_position:
                    self.parent.file_manager.SetPosition(file_manager_position)
                if file_manager_size:
                    self.parent.file_manager.SetSize(file_manager_size)
                self.parent.file_manager.Show()
            except Exception as e:
                print(f"Error reopening FileManager: {e}")

    def _add_sheet_to_excel(self, sheet_name, be_values, intensities, exp_info=None):
        """Add new sheet to Excel file with experimental info."""
        excel_path = self.parent.Data.get('FilePath', '')
        if not excel_path or not os.path.exists(excel_path):
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.create_sheet(sheet_name)

            ws.cell(row=1, column=1, value='BE')
            ws.cell(row=1, column=2, value='Raw Data')

            for i, (be, intensity) in enumerate(zip(be_values, intensities), start=2):
                ws.cell(row=i, column=1, value=round(float(be), 2))
                ws.cell(row=i, column=2, value=round(float(intensity), 2))

            # Add experimental info
            if exp_info:
                exp_col = 50
                ws.cell(row=1, column=exp_col, value="Experimental Description")

                row = 2
                for key, value in exp_info.items():
                    ws.cell(row=row, column=exp_col, value=key)
                    ws.cell(row=row, column=exp_col + 1, value=str(value))
                    row += 1

                ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
                ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40

            wb.save(excel_path)
            wb.close()
        except Exception as e:
            print(f"Error adding sheet to Excel: {e}")

    def on_show_info(self, event=None):
        """Show experimental info window using parent's method."""
        if hasattr(self.parent, 'show_info_window'):
            # Get map data's experimental info
            map_data = self.parent.Data['Core levels'].get(self.map_sheet_name, {})
            exp_info = map_data.get('ExperimentalInfo', {})

            if exp_info:
                self.parent.show_info_window(exp_info, f"Map Info - {self.map_sheet_name}")
            else:
                wx.MessageBox("No experimental information available for this map.",
                              "Info", wx.OK | wx.ICON_INFORMATION)

    def on_close(self, event):
        """Handle window close."""
        if hasattr(self.parent, 'scienta_map_window'):
            self.parent.scienta_map_window = None
        self.Destroy()


def open_scienta_map_viewer(parent, map_sheet_name):
    """
    Open the Scienta Map Viewer window for an XPS~Map sheet.
    """
    try:
        viewer = ScientaMapViewerWindow(parent, map_sheet_name)
        return viewer
    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error opening map viewer:\n{str(e)}",
                     "Error", wx.OK | wx.ICON_ERROR)
        return None